"""
Mewsie's ItemParam Toolbox — v12
  Tool 1  · Box XML Generator  (live validation, last-ID memory, compound/exchange output)
  Tool 1b · ItemParam Generator (dropdown selectors, tooltips, compound/exchange, PresentItemParam)
  Tool 2  · Rate / Count Adjuster
  Tool 3  · NCash Updater (simple CSV)
  Tool 4  · NCash Updater (parent-box CSV + sub-box)
  Tool 5  · NCash ↔ Ticket Calculator

Run: python box_tool_suite.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv, io, re, os, copy, json as _json
try:
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None
try:
    import openpyxl
    _HAVE_OPENPYXL = True
except ImportError:
    _HAVE_OPENPYXL = False

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — character / options data  (used by Tool 1)
# ══════════════════════════════════════════════════════════════════════════════
CHR_NAMES = ["Bunny","Buffalo","Sheep","Dragon","Fox","Lion","Cat","Raccoon","Paula"]
CHR_JOBS  = ["1st","2nd","3rd"]
CHR_FLAG_MAP = {
    "Bunny 1st":1,"Buffalo 1st":2,"Sheep 1st":4,"Dragon 1st":8,"Fox 1st":16,
    "Lion 1st":32,"Cat 1st":64,"Raccoon 1st":124,"Paula 1st":256,
    "Bunny 2nd":512,"Buffalo 2nd":1024,"Sheep 2nd":2048,"Dragon 2nd":4096,
    "Fox 2nd":8192,"Lion 2nd":16384,"Cat 2nd":32768,"Raccoon 2nd":65536,
    "Paula 2nd":131072,"Bunny 3rd":262144,"Buffalo 3rd":524288,"Sheep 3rd":1048576,
    "Dragon 3rd":2097152,"Fox 3rd":4194304,"Lion 3rd":8388608,"Cat 3rd":16777216,
    "Raccoon 3rd":33554432,"Paula 3rd":67108864,
}
CHR_FLAG_REVERSE = {v:k for k,v in CHR_FLAG_MAP.items()}

# ── Job-name to (race, job_tier) lookup for CSV auto-populate ───────────────
# Each entry: normalised_keyword -> (CHR_NAMES entry, "1st"/"2nd"/"3rd")
_JOB_NAME_MAP = {}
def _reg(*keywords, race, tier):
    for kw in keywords:
        _JOB_NAME_MAP[re.sub(r"[^a-z0-9]","", kw.lower())] = (race, tier)
# Bunny
_reg("schoolgirl","bunnyschoolgirl", race="Bunny", tier="1st")
_reg("boxer","bunnyboxer", race="Bunny", tier="2nd")
_reg("champion","duelist","bunnychampion","bunnyduelist", race="Bunny", tier="3rd")
# Buffalo
_reg("fighter","buffalofighter", race="Buffalo", tier="1st")
_reg("warrior","buffalowarrior", race="Buffalo", tier="2nd")
_reg("mercenary","gladiator","buffalomercenary","buffalogladiator", race="Buffalo", tier="3rd")
# Sheep
_reg("librarian","sheeplibrarian", race="Sheep", tier="1st")
_reg("bard","sheepbard", race="Sheep", tier="2nd")
_reg("witch","soulmaster","sheepwitch","sheepsoulmaster", race="Sheep", tier="3rd")
# Dragon
_reg("shaman","dragonshaman", race="Dragon", tier="1st")
_reg("magician","dragonmagician", race="Dragon", tier="2nd")
_reg("priest","darklord","wizard","dragonpriest","dragondarklord","dragonwizard", race="Dragon", tier="3rd")
# Fox
_reg("foxfirst","fox1st", race="Fox", tier="1st")
_reg("foxsecond","fox2nd", race="Fox", tier="2nd")
_reg("foxthird","fox3rd", race="Fox", tier="3rd")
# Lion
_reg("lionfirst","lion1st", race="Lion", tier="1st")
_reg("lionsecond","lion2nd", race="Lion", tier="2nd")
_reg("lionthird","lion3rd", race="Lion", tier="3rd")
# Cat
_reg("model","catmodel", race="Cat", tier="1st")
_reg("entertainer","catentertainer", race="Cat", tier="2nd")
_reg("primadonna","diva","catprimadonna","catdiva", race="Cat", tier="3rd")
# Raccoon
_reg("teacher","raccoonteacher", race="Raccoon", tier="1st")
_reg("cardmaster","raccooncardmaster", race="Raccoon", tier="2nd")
_reg("gambler","duke","raccoongambler","raccounduke", race="Raccoon", tier="3rd")
# Paula / Polar Bear / Bear
_reg("animallover","paulaanimallover","polarbearanimallover", race="Paula", tier="1st")
_reg("trainer","paulatrainer","polarbeartrainer", race="Paula", tier="2nd")
_reg("zoologist","paulazoologist","polarbearzoologist", race="Paula", tier="3rd")

# Race aliases: bear / polarbear / paula all -> "Paula"
_RACE_ALIAS = {}
for aliases, canon in [
    (["bunny"], "Bunny"), (["buffalo"], "Buffalo"),
    (["sheep"], "Sheep"), (["dragon"], "Dragon"),
    (["fox"], "Fox"), (["lion"], "Lion"), (["cat"], "Cat"),
    (["raccoon"], "Raccoon"),
    (["paula","bear","polarbear","polar bear","polar"], "Paula"),
]:
    for a in aliases: _RACE_ALIAS[re.sub(r"[^a-z]","", a.lower())] = canon

# Tier aliases
_TIER_ALIAS = {}
for aliases, canon in [
    (["1st","first","1","job1","firstjob","1stjob","job1st"], "1st"),
    (["2nd","second","2","job2","secondjob","2ndjob","job2nd"], "2nd"),
    (["3rd","third","3","job3","thirdjob","3rdjob","job3rd"], "3rd"),
]:
    for a in aliases: _TIER_ALIAS[re.sub(r"[^a-z0-9]","", a.lower())] = canon

def resolve_chr_flag(text):
    """Try to resolve free-text like '2nd job Sheep' or 'Witch' to a CHR_FLAG_MAP value.
    Returns int flag or None."""
    raw = text.strip()
    norm = re.sub(r"[^a-z0-9]", "", raw.lower())
    # Direct job name lookup
    if norm in _JOB_NAME_MAP:
        race, tier = _JOB_NAME_MAP[norm]
        return CHR_FLAG_MAP.get(f"{race} {tier}")
    # Try to parse "tier race" or "race tier" patterns
    # Strip common joiners
    parts = re.split(r"[-–—/\s]+", raw.lower())
    parts = [re.sub(r"[^a-z0-9]","", p) for p in parts if p.strip()]
    found_race = found_tier = None
    for p in parts:
        if p in _RACE_ALIAS: found_race = _RACE_ALIAS[p]
        if p in _TIER_ALIAS: found_tier = _TIER_ALIAS[p]
    # Also try each part as a job name (e.g. "witch")
    if not found_race or not found_tier:
        for p in parts:
            if p in _JOB_NAME_MAP:
                r, t = _JOB_NAME_MAP[p]
                if not found_race: found_race = r
                if not found_tier: found_tier = t
    if found_race and found_tier:
        return CHR_FLAG_MAP.get(f"{found_race} {found_tier}")
    return None
OPTIONS_CHECKS = [
    ("Not Buyable",256),("Not Sellable",512),("Not Exchangeable",1024),
    ("Not Pickable",2048),("Not Droppable",4096),("Not Vanishable",8192),
    ("No Angelina Bank",65536),("No Lisa Bank",131072),
]

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — XML regex  (used by Tools 1/2/3/4)
# ══════════════════════════════════════════════════════════════════════════════
ROW_RE   = re.compile(r'<ROW>.*?</ROW>', re.DOTALL)
CDATA_RE = re.compile(r'<!\[CDATA\[(.*?)\]\]>', re.DOTALL)

def _get_tag(block, tag):
    m = re.search(rf'<{re.escape(tag)}>(.*?)</{re.escape(tag)}>', block, re.DOTALL)
    if not m: return ""
    cd = CDATA_RE.search(m.group(1))
    return cd.group(1).strip() if cd else m.group(1).strip()

def _set_tag(block, tag, val):
    return re.sub(rf'<{re.escape(tag)}>.*?</{re.escape(tag)}>',
                  f'<{tag}>{val}</{tag}>', block, flags=re.DOTALL)

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — ItemParam / NCash helpers  (Tools 1/3/4)
# ══════════════════════════════════════════════════════════════════════════════
TARGET_FILES = {"itemparam2.xml","itemparamcm2.xml","itemparamex2.xml","itemparamex.xml"}
PRESENT_FILE = "presentitemparam2.xml"

def build_item_lib(files):
    lib = {}
    for _, text in files:
        for row in ROW_RE.findall(text):
            rid  = _get_tag(row, "ID")
            name = _get_tag(row, "Name")
            if rid.isdigit() and name:
                lib[rid] = name
    return lib

def bulk_update_ncash(xml_text, updates):
    found = {k: False for k in updates}
    def replace_row(m):
        block = m.group(0)
        rid   = _get_tag(block, "ID")
        if rid not in updates: return block
        found[rid] = True
        return re.sub(r'<Ncash>\d+</Ncash>', f'<Ncash>{updates[rid]}</Ncash>', block)
    return ROW_RE.sub(replace_row, xml_text), found

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 1 CSV + XML builders
# ══════════════════════════════════════════════════════════════════════════════
# ── Column headers that are NEVER box-name columns ──────────────────────────
# ── Column-header sets for CSV field recognition ─────────────────────────────
_SKIP_HEADERS = {
    "id","level","rate","lv","luck","lvl","chance","prob","itemcnt","count",
    "droprate","dropcnt","dropid","drop","qty","quantity","amount","weight",
}

def _norm_hdr(h):
    return re.sub(r"[^a-z0-9]", "", h.strip().lower())

# Headers that identify box-name columns (the column header IS the box name)
# A column is a box-name column when its header is non-empty, not a known skip
# word, not purely numeric, and NOT a recognised field/metadata header.
_FIELD_HDRS = {
    "name","comment","use","filename","fn","bundlenum","bn",
    "cmtfilename","cmtfn","cmtbundlenum","cmtbn",
    "weight","value","minlevel","money","ncash","tickets","ticket",
    "options","chrtypeflags","chrtypeflag","recycle","recyclable",
    "boxid","contents","content","id",
}

def _is_box_name_header(h):
    hn = _norm_hdr(h)
    if not hn: return False
    if hn.isdigit(): return False
    if hn in _SKIP_HEADERS: return False
    if hn in _FIELD_HDRS: return False
    return True

def _parse_options_cell(cell):
    """Parse Options cell → (opt_checks list, recycle_int).
    Accepts:
      • Text labels  e.g. 'Not Buyable, Not Sellable'
      • Raw bitmask  e.g. '2' or '2/256/512' (slash/comma separated ints)
        The bitmask is matched against each OPTIONS_CHECKS flag value.
    """
    opt_checks = [False] * 8
    recycle = 0
    cell = cell.strip()
    # Try numeric / slash-separated-numeric first
    raw_parts = re.split(r"[,;|/\s]+", cell)
    all_numeric = all(p.strip().isdigit() for p in raw_parts if p.strip())
    if all_numeric and any(p.strip() for p in raw_parts):
        combined = 0
        for p in raw_parts:
            p = p.strip()
            if p.isdigit(): combined |= int(p)
        # Map each OPTIONS_CHECKS value
        for i, (_, v) in enumerate(OPTIONS_CHECKS):
            if combined & v: opt_checks[i] = True
        # Recycle flags
        if combined & 262144:  recycle = 262144
        if combined & 8388608: recycle = 8388608
        return opt_checks, recycle
    # Text label parsing
    tokens = [re.sub(r"[^a-z0-9]", "", t.lower()) for t in re.split(r"[,;|]+", cell) if t.strip()]
    for tok in tokens:
        for i, (lbl, _) in enumerate(OPTIONS_CHECKS):
            if re.sub(r"[^a-z0-9]", "", lbl.lower()) in tok or tok in re.sub(r"[^a-z0-9]", "", lbl.lower()):
                opt_checks[i] = True
        if tok in ("recyclable", "recycle", "recyc"):                         recycle = 262144
        if tok in ("nonrecyclable", "norecycle", "nonrecycle", "notrecyclable"): recycle = 8388608
    return opt_checks, recycle

def _parse_chrtypeflags_cell(cell):
    """Parse ChrTypeFlags cell → list of flag ints.
    Accepts:
      • Text  e.g. 'Sheep 2nd, Dragon 3rd' or 'Witch, Priest'
      • Raw bitmask integer e.g. '2050'  (decoded to individual flag values)
      • Slash/comma separated raw ints e.g. '2048/2'
    """
    cell = cell.strip()
    # Check if the whole cell is a pure bitmask (one int or slash-sep ints)
    raw_parts = re.split(r"[,;|/\s]+", cell)
    all_numeric = all(p.strip().isdigit() for p in raw_parts if p.strip())
    if all_numeric and any(p.strip() for p in raw_parts):
        combined = 0
        for p in raw_parts:
            if p.strip().isdigit(): combined |= int(p.strip())
        # Expand combined bitmask to individual known flags
        flags = []
        for v in sorted(CHR_FLAG_MAP.values()):
            if combined & v: flags.append(v)
        return flags
    # Text parsing
    flags = []
    for part in re.split(r"[,;|/]+", cell):
        part = part.strip()
        if not part: continue
        if part.isdigit(): flags.append(int(part)); continue
        f = resolve_chr_flag(part)
        if f is not None: flags.append(f)
    return flags

# Map normalised header → cfg_override key (special keys start with _)
_HDR_TO_CFGKEY = {
    "name":         "name",
    "comment":      "comment",
    "use":          "use",
    "filename":     "file_name",   "fn": "file_name",
    "bundlenum":    "bundle_num",  "bn": "bundle_num",
    "cmtfilename":  "cmt_file_name", "cmtfn": "cmt_file_name",
    "cmtbundlenum": "cmt_bundle_num", "cmtbn": "cmt_bundle_num",
    "weight":       "weight",
    "value":        "value",
    "minlevel":     "min_level",
    "money":        "money",
    "ncash":        "_ncash",      # handled specially
    "tickets":      "_tickets",    # handled specially
    "ticket":       "_tickets",
    "options":      "_options",
    "chrtypeflags": "_chr",
    "chrtypeflag":  "_chr",
    "recycle":      "_recycle",
    "recyclable":   "_recycle",
    "boxid":        "_boxid",      # box's own ID
    "id":           "_id",         # ambiguous – could be box ID or item ID
    "contents":     "_contents",
    "content":      "_contents",
}

# Filename fields — keep only safe path characters (letters, digits, . _ - \ / : space)
_FILENAME_KEYS = {"file_name", "cmt_file_name"}

def _sanitise_filename(val):
    """Strip special characters from a filename/path, keeping backslash for Windows paths."""
    return re.sub(r"[^\w .\/:_\-]", "", val, flags=re.ASCII)

def _apply_field_col(cfg_override, key, val):
    """Apply a parsed field value to cfg_override dict."""
    if key in _FILENAME_KEYS:
        cfg_override[key] = _sanitise_filename(val)
    elif key == "_options":
        oc, rv = _parse_options_cell(val)
        cfg_override["opt_checks"] = oc
        if rv: cfg_override["opt_recycle"] = rv
    elif key == "_chr":
        flags = _parse_chrtypeflags_cell(val)
        if flags: cfg_override["chr_type_flags"] = flags
    elif key == "_recycle":
        nv = _norm_hdr(val)
        if nv in ("recyclable","recycle","recyc"):                       cfg_override["opt_recycle"] = 262144
        elif nv in ("nonrecyclable","norecycle","nonrecycle","notrecyclable"): cfg_override["opt_recycle"] = 8388608
    elif key == "_ncash":
        try: cfg_override["ncash_direct"] = int(round(float(val)))
        except: pass
    elif key == "_tickets":
        try: cfg_override["ticket"] = val
        except: pass
    elif not key.startswith("_"):
        cfg_override[key] = val

def parse_grouped_csv(text):
    """Parse a box-definition CSV.

    Layout rules (same as original, with additions):
    - Header row: columns whose headers are not recognised field/skip words are
      "box-name columns".  Each defines a GROUP.
    - A group's span is: columns from (previous group's box-name col + 1) up to
      and including its own box-name col.
    - Within each span:
        * The box-name col: each data row's cell is an item NAME (or box name if
          a "Box ID" / "Box …" column is present).
        * Any recognised field header (Name, Comment, Options, ChrTypeFlags,
          BoxID, Contents, …) is parsed and stored in cfg_override.
        * Numeric middle columns are item DropRate.
        * "ItemCnt" / "Count" column sets item_cnt.
        * "Contents" column overrides item name source.
        * "Box ID" or "BoxID" column sets the box's own ID in cfg_override.
    - Returns list of {box_name, items, cfg_override} dicts.
    """
    reader = csv.reader(io.StringIO(text))
    rows   = list(reader)
    if not rows: return []
    headers = [h.strip() for h in rows[0]]
    hnorms  = [_norm_hdr(h) for h in headers]
    data_rows = rows[1:]

    # Identify box-name columns
    box_col_indices = [i for i, h in enumerate(headers) if _is_box_name_header(h)]
    if not box_col_indices: return []

    results = []
    prev_end = -1
    # Compute next box-col for each group (to know where trailing fields end)
    next_box_cols = box_col_indices[1:] + [len(headers)]

    for bi, bc in enumerate(box_col_indices):
        # Core span: prev_end+1 .. bc  (the original item columns)
        # Trailing span: bc+1 .. next_box_col-1 BUT only recognised field headers
        next_bc = next_box_cols[bi]
        _TRAILING_OK = {k for k,v in _HDR_TO_CFGKEY.items()
                        if v not in ("_id",) and k not in _SKIP_HEADERS}
        trailing = [ci for ci in range(bc + 1, next_bc)
                    if _norm_hdr(headers[ci]) in _TRAILING_OK]
        span = list(range(prev_end + 1, bc + 1)) + trailing
        prev_end = bc

        box_name_hdr = headers[bc]   # the column header = box name for the group

        # Classify every column in the span
        id_col       = None   # item ID column index (per-item)
        box_id_col   = None   # box's own ID column index
        contents_col = None   # alternate item-name column
        rate_cols    = []     # numeric rate columns
        itemcnt_col  = None
        field_cols   = {}     # col_index -> cfg_key (field overrides)

        for ci in span:
            if ci == bc: continue   # box-name col itself
            hn = hnorms[ci]
            if hn in _HDR_TO_CFGKEY:
                cfgkey = _HDR_TO_CFGKEY[hn]
                if cfgkey == "_boxid":
                    box_id_col = ci
                elif cfgkey == "_id":
                    # bare "ID" column — treat as item ID
                    id_col = ci
                elif cfgkey == "_contents":
                    contents_col = ci
                else:
                    field_cols[ci] = cfgkey
            elif hn in ("itemcnt", "count"):
                itemcnt_col = ci
            elif hn in _SKIP_HEADERS:
                rate_cols.append(ci)   # treat rate/level/etc as rate source
            else:
                rate_cols.append(ci)   # unknown middle col → try as rate

        # Build cfg_override from field columns (take first non-empty value per field)
        group_cfg = {}
        for ci, cfgkey in field_cols.items():
            for row in data_rows:
                val = row[ci].strip() if ci < len(row) else ""
                if val:
                    _apply_field_col(group_cfg, cfgkey, val)
                    break

        # Box's own ID from box_id_col (first non-empty)
        if box_id_col is not None:
            for row in data_rows:
                val = row[box_id_col].strip() if box_id_col < len(row) else ""
                if val.isdigit():
                    group_cfg["id"] = val
                    break

        # Build items list
        items = []
        for row in data_rows:
            # Item name comes from box-name col (or contents_col if present)
            raw_name = row[bc].strip() if bc < len(row) else ""
            if not raw_name: continue

            item_name = raw_name
            if contents_col is not None:
                item_name = row[contents_col].strip() if contents_col < len(row) else raw_name

            # Item ID
            item_id = ""
            if id_col is not None:
                v = row[id_col].strip() if id_col < len(row) else ""
                if v.isdigit(): item_id = v

            # Rate: first parseable integer from rate_cols
            rate = None
            for ci in rate_cols:
                v = row[ci].strip() if ci < len(row) else ""
                try: rate = int(v.replace("%","").strip()); break
                except: pass

            # ItemCnt
            item_cnt = 1
            if itemcnt_col is not None:
                v = row[itemcnt_col].strip() if itemcnt_col < len(row) else ""
                try: item_cnt = int(v)
                except: pass

            items.append({"id": item_id, "name": item_name,
                          "extra": [], "rate": rate, "item_cnt": item_cnt})

        if items:
            results.append({"box_name": box_name_hdr,
                            "items": items,
                            "cfg_override": group_cfg})

    return results

def substitute_box_name(template, old, new):
    if not old or not template: return template
    return re.sub(re.escape(old), new, template, flags=re.IGNORECASE)

def apply_name_template(template, prev, new):
    if not template or not prev: return new
    return substitute_box_name(template, prev, new)

def deduplicate_name(proposed, existing):
    if proposed not in existing: return proposed
    i = 2
    while f"{proposed} ({i})" in existing: i += 1
    return f"{proposed} ({i})"

def build_options_str(check_states, recycle_val):
    base    = [2,32]
    checked = [v for (_,v),on in zip(OPTIONS_CHECKS,check_states) if on]
    rec     = [recycle_val] if recycle_val > 0 else []
    return "/".join(str(x) for x in sorted(set(base+checked+rec)))

def build_itemparam_row(cfg):
    opts      = build_options_str(cfg["opt_checks"], cfg["opt_recycle"])
    chr_flags = sum(cfg["chr_type_flags"])
    fn        = cfg["file_name"]
    bn        = cfg["bundle_num"]
    return f"""<ROW>
<ID>{cfg['id']}</ID>
<Class>1</Class>
<Type>15</Type>
<SubType>0</SubType>
<ItemFType>0</ItemFType>
<n><![CDATA[{cfg['name']}]]></n>
<Comment><![CDATA[{cfg['comment']}]]></Comment>
<Use><![CDATA[{cfg['use']}]]></Use>
<Name_Eng><![CDATA[ ]]></Name_Eng>
<Comment_Eng><![CDATA[ ]]></Comment_Eng>
<FileName><![CDATA[{fn}]]></FileName>
<BundleNum>{bn}</BundleNum>
<InvFileName><![CDATA[{fn}]]></InvFileName>
<InvBundleNum>{bn}</InvBundleNum>
<CmtFileName><![CDATA[{cfg['cmt_file_name']}]]></CmtFileName>
<CmtBundleNum>{cfg['cmt_bundle_num']}</CmtBundleNum>
<EquipFileName><![CDATA[ ]]></EquipFileName>
<PivotID>0</PivotID>
<PaletteId>0</PaletteId>
<Options>{opts}</Options>
<HideHat>0</HideHat>
<ChrTypeFlags>{chr_flags}</ChrTypeFlags>
<GroundFlags>0</GroundFlags>
<SystemFlags>0</SystemFlags>
<OptionsEx>0</OptionsEx>
<Weight>{cfg['weight']}</Weight>
<Value>{cfg['value']}</Value>
<MinLevel>{cfg['min_level']}</MinLevel>
<Effect>22</Effect>
<EffectFlags2>0</EffectFlags2>
<SelRange>0</SelRange>
<Life>0</Life>
<Depth>0</Depth>
<Delay>0.000000</Delay>
<AP>0</AP>
<HP>0</HP>
<HPCon>0</HPCon>
<MP>0</MP>
<MPCon>0</MPCon>
<Money>{cfg['money']}</Money>
<APPlus>0</APPlus>
<ACPlus>0</ACPlus>
<DXPlus>0</DXPlus>
<MaxMPPlus>0</MaxMPPlus>
<MAPlus>0</MAPlus>
<MDPlus>0</MDPlus>
<MaxWTPlus>0</MaxWTPlus>
<DAPlus>0</DAPlus>
<LKPlus>0</LKPlus>
<MaxHPPlus>0</MaxHPPlus>
<DPPlus>0</DPPlus>
<HVPlus>0</HVPlus>
<HPRecoveryRate>0.000000</HPRecoveryRate>
<MPRecoveryRate>0.000000</MPRecoveryRate>
<CardNum>0</CardNum>
<CardGenGrade>0</CardGenGrade>
<CardGenParam>0.000000</CardGenParam>
<DailyGenCnt>0</DailyGenCnt>
<PartFileName><![CDATA[ ]]></PartFileName>
<ChrFTypeFlag>0</ChrFTypeFlag>
<ChrGender>0</ChrGender>
<ExistType>0</ExistType>
<Ncash>{cfg['ncash']}</Ncash>
<NewCM>0</NewCM>
<FamCM>0</FamCM>
<Summary><![CDATA[ ]]></Summary>
<ShopFileName><![CDATA[ ]]></ShopFileName>
<ShopBundleNum>0</ShopBundleNum>
<MinStatType>0</MinStatType>
<MinStatLv>0</MinStatLv>
<RefineIndex>0</RefineIndex>
<RefineType>0</RefineType>
<CompoundSlot>0</CompoundSlot>
<SetItemID>0</SetItemID>
<ReformCount>0</ReformCount>
<GroupId>0</GroupId>
</ROW>"""

def build_presentparam_row(box_id, items, ptype, drop_cnt, default_rate,
                           item_cnts=None, box_name=None):
    """box_name appended as <!-- name --> after <Id>.
    items[i]["name"] appended as <!-- name --> after each <DropId_#>."""
    is_distrib      = (ptype == 2)
    actual_drop_cnt = len(items) if is_distrib else drop_cnt
    id_line = f"<Id>{box_id}</Id>"
    if box_name:
        id_line += f" <!-- {box_name} -->"
    lines = ["<ROW>", id_line,
             f"<Type>{ptype}</Type>", f"<DropCnt>{actual_drop_cnt}</DropCnt>"]
    for i in range(20):
        if i < len(items):
            did   = items[i]["id"]
            drate = 100 if is_distrib else (items[i].get("rate") or default_rate)
            icnt  = (item_cnts[i] if item_cnts and i < len(item_cnts) else 1) or 1
            item_name = items[i].get("name", "").strip()
            drop_id_line = f"<DropId_{i}>{did}</DropId_{i}>"
            if item_name:
                drop_id_line += f" <!-- {item_name} -->"
        else:
            did, drate, icnt = 0, 0, 0
            drop_id_line = f"<DropId_{i}>{did}</DropId_{i}>"
        lines += [drop_id_line,
                  f"<DropRate_{i}>{drate}</DropRate_{i}>",
                  f"<ItemCnt_{i}>{icnt}</ItemCnt_{i}>"]
    lines.append("</ROW>")
    return "\n".join(lines)

def build_recycle_except_row(box_id, name):
    return f"<ROW>\n<ItemID>{box_id}</ItemID>\n<Comment><![CDATA[{name}]]></Comment>\n</ROW>"

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 2 XML helpers
# ══════════════════════════════════════════════════════════════════════════════
def real_drop_slots(block):
    pairs = re.findall(r'<DropId_(\d+)>(\d+)</DropId_\d+>', block)
    return [(int(i),v) for i,v in sorted(pairs,key=lambda x:int(x[0])) if v!="0"]

def apply_cfg_to_row(block, cfg):
    block = _set_tag(block,"Type",str(cfg["type"]))
    block = _set_tag(block,"DropCnt",str(cfg["drop_cnt"]))
    for pos,(idx,_) in enumerate(real_drop_slots(block)):
        sc = cfg["slots"][pos] if pos<len(cfg["slots"]) else {"rate":100,"count":1}
        block = _set_tag(block,f"DropRate_{idx}",str(sc["rate"]))
        block = _set_tag(block,f"ItemCnt_{idx}",str(sc["count"]))
    return block

def load_itemparam_folder(folder):
    lib = {}
    for fname in os.listdir(folder):
        if not fname.lower().endswith(".xml"): continue
        try:
            with open(os.path.join(folder,fname),encoding="utf-8-sig",errors="replace") as f:
                text = f.read()
            for row in ROW_RE.findall(text):
                rid  = _get_tag(row,"ID")
                name = _get_tag(row,"n")
                if rid.isdigit() and name: lib[rid] = name
        except: pass
    return lib

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 2 CSV parser
# ══════════════════════════════════════════════════════════════════════════════
_T2_SKIP = {"id","level","rate","lv","luck","lvl","chance","prob"}

def parse_box_id_csv(text):
    reader  = csv.reader(io.StringIO(text))
    rows    = list(reader)
    if not rows: return {}
    headers = [h.strip() for h in rows[0]]
    id_positions = [i for i,h in enumerate(headers) if h.strip().lower()=="id"]
    if not id_positions: id_positions = [0]
    box_map = {}
    for g,id_pos in enumerate(id_positions):
        next_id = id_positions[g+1] if g+1<len(id_positions) else len(headers)
        gcols   = list(range(id_pos,next_id))
        ghdrs   = [headers[c] for c in gcols]
        name_local = next((i for i,h in enumerate(ghdrs)
                           if bool(h) and h.strip().lower() not in _T2_SKIP and not h.strip().isdigit()),None)
        for row in rows[1:]:
            id_val = row[id_pos].strip() if id_pos<len(row) else ""
            if not id_val or not id_val.isdigit(): continue
            if name_local is not None:
                nc = gcols[name_local]
                name_val = row[nc].strip() if nc<len(row) else ""
            else:
                name_val = ""
            box_map[id_val] = name_val
    return box_map

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 3 CSV parser
# ══════════════════════════════════════════════════════════════════════════════
def parse_csv_text_t3(text):
    stripped = text.strip()
    if not stripped: return []
    all_rows = list(csv.reader(io.StringIO(stripped)))
    if not all_rows: return []
    raw_headers = [h.strip() for h in all_rows[0]]
    data_rows   = all_rows[1:]
    items, seen = [], set()
    def add(id_str, cost=None):
        id_str = id_str.strip()
        if id_str and id_str.isdigit() and id_str not in seen:
            seen.add(id_str); items.append({"id":id_str,"ticket_cost":cost})
    item_col_positions = [i for i,h in enumerate(raw_headers) if re.match(r'Item\d+_ID',h,re.I)]
    if item_col_positions:
        for row in data_rows:
            for pos in item_col_positions:
                if pos<len(row): add(row[pos])
        return items
    id_positions = [i for i,h in enumerate(raw_headers) if h.lower()=="id"]
    if id_positions:
        for row in data_rows:
            for pos in id_positions:
                if pos<len(row): add(row[pos])
        return items
    if len(raw_headers)>=2:
        for row in data_rows:
            if not row: continue
            raw_cost = row[1].strip() if len(row)>1 else ""
            try:    cost = float(raw_cost)
            except: cost = None
            add(row[0], cost)
        return items
    for row in data_rows:
        if row: add(row[0])
    return items

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 4 CSV parser
# ══════════════════════════════════════════════════════════════════════════════
_NON_ID_HEADERS = {
    "acplus","ap","applus","bundlenum","cardgengrade","cardgenparam","cardnum",
    "chrftypeflag","chrgender","chrtypeflags","class","cmtbundlenum","cmtfilename",
    "comment","comment_eng","compoundslot","daplus","dpplus","dxplus","dailygencnt",
    "delay","depth","effect","effectflags2","equipfilename","existtype","famcm",
    "filename","groundflags","groupid","hp","hpcon","hprecoveryrate","hvplus",
    "hidehat","invbundlenum","invfilename","itemftype","lkplus","life","maplus",
    "mdplus","mp","mpcon","mprecoveryrate","maxhpplus","maxmpplus","maxwtplus",
    "minlevel","minstatLv","minstattype","money","name","name_eng","newcm",
    "options","optionsex","paletteid","partfilename","pivotid","refineindex",
    "refinetype","reformcount","selrange","setitemid","shopbundlenum","shopfilename",
    "subtype","summary","systemflags","type","use","value","weight",
    "level","rate","lv","luck","lvl","chance","prob","drop","qty",
    "quantity","count","amount","row",
    "droprate","dropcnt","dropid","itemcnt","dropcnt_0","dropcnt_1",
    "droprate_0","droprate_1","droprate_2","droprate_3","droprate_4",
    "dropid_0","dropid_1","dropid_2","dropid_3","dropid_4",
    "itemcnt_0","itemcnt_1","itemcnt_2","itemcnt_3","itemcnt_4",
}
_TICKET_NAMES     = {"tickets","ticket"}
_NCASH_NAMES      = {"ncash","ncash_val","ncashval"}
_BOX_TICKET_NAMES = {
    "tickets of box contents","box contents tickets",
    "box content tickets","tickets of box content",
    "sub-box tickets","subbox tickets",
}

def _find_value_col(raw_headers, id_pos):
    next_id = next((i for i in range(id_pos+1,len(raw_headers))
                    if raw_headers[i].lower()=="id"), len(raw_headers))
    for i in range(id_pos+1, next_id):
        h = raw_headers[i].lower()
        if h in _TICKET_NAMES: return i,"tickets"
        if h in _NCASH_NAMES:  return i,"ncash"
    return None, None

def _find_box_ticket_col(raw_headers, id_pos):
    next_id = next((i for i in range(id_pos+1,len(raw_headers))
                    if raw_headers[i].lower()=="id"), len(raw_headers))
    for i in range(id_pos+1, next_id):
        if raw_headers[i].lower().strip() in _BOX_TICKET_NAMES: return i
    return None

def parse_parentbox_csv(text):
    stripped = text.strip()
    if not stripped: return []
    all_rows = list(csv.reader(io.StringIO(stripped)))
    if not all_rows: return []
    raw_headers = [h.strip() for h in all_rows[0]]
    data_rows   = all_rows[1:]
    id_positions = [i for i,h in enumerate(raw_headers) if h.lower()=="id"]
    val_map, box_tick_map = {}, {}
    for id_pos in id_positions:
        vcol,vtype = _find_value_col(raw_headers, id_pos)
        if vcol is not None: val_map[id_pos] = (vcol,vtype)
        btcol = _find_box_ticket_col(raw_headers, id_pos)
        if btcol is not None: box_tick_map[id_pos] = btcol
    items, seen = [], set()
    def add(id_str, ticket_cost, ncash_direct, box_ticket_cost, group_idx=0):
        id_str = id_str.strip()
        if id_str and id_str.isdigit() and id_str not in seen:
            seen.add(id_str)
            items.append({"id":id_str,"ticket_cost":ticket_cost,"ncash_direct":ncash_direct,
                          "box_ticket_cost":box_ticket_cost,"group_idx":group_idx,"name":""})
    def _parse_num(row, col):
        if col is None or col>=len(row): return None
        try:    return float(row[col].strip())
        except: return None
    if id_positions:
        for row in data_rows:
            for gi,id_pos in enumerate(id_positions):
                if id_pos>=len(row): continue
                id_val = row[id_pos].strip()
                if not (id_val and id_val.isdigit()): continue
                ticket_cost = ncash_direct = None
                if id_pos in val_map:
                    vcol,vtype = val_map[id_pos]
                    num = _parse_num(row, vcol)
                    if num is not None:
                        if vtype=="tickets": ticket_cost  = num
                        else:               ncash_direct = int(round(num))
                btcol = box_tick_map.get(id_pos)
                box_ticket_cost = _parse_num(row,btcol) if btcol is not None else None
                add(id_val, ticket_cost, ncash_direct, box_ticket_cost, group_idx=gi)
        return items
    for row in data_rows:
        for i,cell in enumerate(row):
            hdr = raw_headers[i].lower() if i<len(raw_headers) else ""
            if hdr not in _NON_ID_HEADERS:
                add(cell, None, None, None)
    return items

def extract_drop_ids_from_present(present_text, box_ids):
    result = {}
    for row in ROW_RE.findall(present_text):
        bid = _get_tag(row,"Id")
        if bid not in box_ids: continue
        drops = []
        for i in range(20):
            did = _get_tag(row,f"DropId_{i}")
            if did and did.isdigit() and did!="0": drops.append(did)
        result[bid] = drops
    return result


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STORE  — shared state between tools
# ══════════════════════════════════════════════════════════════════════════════
class AppSession:
    """Central store. Tool 1 writes here; Tools 2-4 can import from it."""
    def __init__(self):
        self.box_id_list_csv  = None   # Tool1 output: box IDs CSV text
        self.box_id_map       = {}     # Tool1 output: {box_id: box_name}
        self.box_contents_csv = None   # Tool2 output (future)
        self.present_xml_path = None

# ══════════════════════════════════════════════════════════════════════════════
# UI HELPERS — shared across all tool frames
# ══════════════════════════════════════════════════════════════════════════════
BG      = "#1e1e2e"
BG2     = "#181825"
BG3     = "#313244"
BG4     = "#45475a"
FG      = "#cdd6f4"
FG_DIM  = "#a6adc8"
FG_GREY = "#6c7086"
ACC1    = "#cba6f7"   # purple  — tool 1
ACC2    = "#89dceb"   # cyan    — tool 2
ACC3    = "#f38ba8"   # red     — tool 3
ACC4    = "#fab387"   # peach   — tool 4
ACC5    = "#f9e2af"   # yellow  — calculator
ACC6    = "#94e2d5"   # teal    — itemparam gen
GREEN   = "#a6e3a1"
BLUE    = "#89b4fa"

def mk_section(parent, title):
    f = tk.LabelFrame(parent, text=title, bg=BG, fg=BLUE,
                      font=("Consolas",10,"bold"), bd=1, relief="groove")
    f.pack(fill="x", padx=12, pady=5)
    return f

def mk_btn(parent, text, command, color=BG3, fg=FG, **kw):
    defaults = dict(font=("Consolas",10), relief="flat", padx=12, pady=6)
    defaults.update(kw)
    return tk.Button(parent, text=text, command=command, bg=color, fg=fg, **defaults)

def mk_scroll_canvas(parent, init_width=900):
    """Returns (canvas, container_frame). Container scrolls inside canvas.
    init_width: initial width hint so content renders before the first Configure event."""
    sb     = ttk.Scrollbar(parent, orient="vertical")
    canvas = tk.Canvas(parent, bg=BG, highlightthickness=0,
                       yscrollcommand=sb.set, width=init_width)
    sb.configure(command=canvas.yview)
    sb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    cont = tk.Frame(canvas, bg=BG)
    wid  = canvas.create_window((0,0), window=cont, anchor="nw", width=init_width)

    def _on_inner_resize(e):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _on_canvas_resize(e):
        canvas.itemconfig(wid, width=e.width)

    cont.bind("<Configure>", _on_inner_resize)
    canvas.bind("<Configure>", _on_canvas_resize)

    def _on_enter(e): canvas.bind_all("<MouseWheel>", lambda ev: canvas.yview_scroll(-1*(ev.delta//120),"units"))
    def _on_leave(e): canvas.unbind_all("<MouseWheel>")
    canvas.bind("<Enter>", _on_enter)
    canvas.bind("<Leave>", _on_leave)

    return canvas, cont

class ScrolledFrame(tk.Frame):
    """A frame with a vertical scrollbar. Place children in .inner.
    Uses a canvas with an explicit initial width so content renders immediately."""
    def __init__(self, parent, **kw):
        super().__init__(parent, **kw)
        self.configure(bg=BG)
        self._canvas = tk.Canvas(self, bg=BG, highlightthickness=0,
                                 borderwidth=0, width=900)
        self._sb = ttk.Scrollbar(self, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._sb.set)
        self._sb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(self._canvas, bg=BG)
        self._win = self._canvas.create_window((0, 0), window=self.inner,
                                               anchor="nw", width=900)
        self.inner.bind("<Configure>", self._on_inner)
        self._canvas.bind("<Configure>", self._on_canvas)
        self._canvas.bind("<Enter>",
            lambda e: self._canvas.bind_all("<MouseWheel>", self._scroll))
        self._canvas.bind("<Leave>",
            lambda e: self._canvas.unbind_all("<MouseWheel>"))

    def _on_inner(self, e):
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas(self, e):
        self._canvas.itemconfig(self._win, width=e.width)

    def _scroll(self, e):
        self._canvas.yview_scroll(-1 * (e.delta // 120), "units")

def make_output_tab(nb, title, content, fname, root):
    frm = tk.Frame(nb, bg=BG); nb.add(frm, text=title)
    br  = tk.Frame(frm, bg=BG); br.pack(side="bottom", fill="x")
    mk_btn(br, "📋 Copy All",
           lambda c=content: (root.clipboard_clear(), root.clipboard_append(c),
                               messagebox.showinfo("Copied","Copied to clipboard.")),
           padx=10, pady=4).pack(side="left", padx=6, pady=4)
    def _save(c=content, f=fname):
        p = filedialog.asksaveasfilename(initialfile=f, defaultextension=".xml",
                filetypes=[("XML","*.xml"),("CSV","*.csv"),("Text","*.txt"),("All","*.*")])
        if p:
            with open(p,"w",encoding="utf-8") as fh: fh.write(c)
            messagebox.showinfo("Saved",f"Saved to {p}")
    mk_btn(br,"💾 Save As…",_save,color=GREEN,fg=BG2,padx=10,pady=4).pack(side="left",padx=6,pady=4)
    txt = scrolledtext.ScrolledText(frm, font=("Consolas",9), bg=BG2, fg=FG)
    txt.pack(fill="both", expand=True, padx=4, pady=4)
    txt.insert("1.0", content); txt.config(state="disabled")


def _sheet_to_csv(sheet):
    """Convert an openpyxl worksheet to CSV text, preserving cell values."""
    out = io.StringIO()
    writer = csv.writer(out)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(["" if v is None else str(v) for v in row])
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 1 — Box XML Generator
# ══════════════════════════════════════════════════════════════════════════════
class Tool1(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self.saved_settings = None
        self.continue_mode  = None
        self.groups         = []
        self.current_group_idx = 0
        self.box_configs    = []
        self.no_csv_mode    = False
        self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    # ── Load screen ──────────────────────────────────────────────────────────
    def _build_load_screen(self):
        self._clear()
        frm = tk.Frame(self, bg=BG); frm.pack(expand=True)
        tk.Label(frm, text="BOX XML GENERATOR", font=("Consolas",20,"bold"),
                 bg=BG, fg=ACC1).pack(pady=(30,5))
        tk.Label(frm, text="Load a CSV with groups:  ID  |  Level/Rate  |  Parent Box Name",
                 bg=BG, fg=FG_DIM, font=("Consolas",10)).pack(pady=5)
        bf = tk.Frame(frm, bg=BG); bf.pack(pady=15)
        mk_btn(bf,"📂  Load File (CSV / Excel)",self._load_csv_file).pack(side="left",padx=8)
        mk_btn(bf,"📋  Paste CSV Text",  self._paste_csv    ).pack(side="left",padx=8)
        mk_btn(bf,"✏️  No CSV — Manual Entry", self._start_no_csv,
               color=BG4).pack(side="left",padx=8)

    def _load_csv_file(self):
        types = [("CSV / Excel","*.csv *.xlsx *.xls"),("CSV","*.csv"),("All","*.*")]
        if _HAVE_OPENPYXL:
            types.insert(1,("Excel","*.xlsx *.xls"))
        p = filedialog.askopenfilename(filetypes=types)
        if not p: return
        ext = os.path.splitext(p)[1].lower()
        if ext in (".xlsx",".xls"):
            self._load_excel_file(p)
        else:
            with open(p, encoding="utf-8-sig") as f: self._process_csv(f.read())

    def _load_excel_file(self, path):
        if not _HAVE_OPENPYXL:
            messagebox.showerror("Missing library",
                "openpyxl is required to open Excel files.\n"
                "Install it with:  pip install openpyxl"); return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Excel error", str(e)); return
        sheet_names = wb.sheetnames
        if not sheet_names:
            messagebox.showerror("Excel error","Workbook has no sheets."); return
        # If only one sheet, load it directly
        if len(sheet_names) == 1:
            self._process_csv(_sheet_to_csv(wb[sheet_names[0]])); return
        # Multiple sheets — ask user which ones to load
        win = tk.Toplevel(self.root); win.title("Select Sheets")
        win.configure(bg=BG); win.grab_set()
        tk.Label(win, text="Select sheets to import:", bg=BG, fg=FG,
                 font=("Consolas",11,"bold")).pack(pady=(14,6),padx=16,anchor="w")
        tk.Label(win, text="Each sheet becomes a separate group.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(padx=16,anchor="w")
        check_vars = []
        for name in sheet_names:
            v = tk.BooleanVar(value=True)
            tk.Checkbutton(win, text=name, variable=v, bg=BG, fg=FG,
                           selectcolor=BG3, activebackground=BG,
                           font=("Consolas",10)).pack(anchor="w",padx=20,pady=2)
            check_vars.append((name, v))
        def do_import():
            selected = [name for name,v in check_vars if v.get()]
            if not selected:
                messagebox.showwarning("Nothing selected","Select at least one sheet."); return
            # Combine selected sheets: convert each to CSV text and concatenate
            # Each sheet is treated as an independent CSV parse, results merged
            all_groups = []
            for name in selected:
                csv_text = _sheet_to_csv(wb[name])
                groups   = parse_grouped_csv(csv_text)
                all_groups.extend(groups)
            win.destroy()
            if not all_groups:
                messagebox.showerror("Error","No valid box groups found in selected sheets."); return
            self.groups = all_groups; self.current_group_idx = 0; self.box_configs = []
            self.continue_mode = None; self.saved_settings = None; self.no_csv_mode = False
            self._build_config_screen()
        bf2 = tk.Frame(win, bg=BG); bf2.pack(pady=12,padx=16,fill="x")
        mk_btn(bf2,"Import",do_import,color=GREEN,fg=BG2).pack(side="left",padx=4)
        mk_btn(bf2,"Cancel",win.destroy).pack(side="left",padx=4)
        win.wait_window()

    def _paste_csv(self):
        win = tk.Toplevel(self.root); win.title("Paste CSV")
        win.geometry("600x400"); win.configure(bg=BG)
        tk.Label(win, text="Paste CSV below:", bg=BG, fg=FG,
                 font=("Consolas",10)).pack(anchor="w",padx=10,pady=5)
        txt = scrolledtext.ScrolledText(win, font=("Consolas",9))
        txt.pack(fill="both",expand=True,padx=10,pady=5)
        def confirm():
            self._process_csv(txt.get("1.0","end")); win.destroy()
        mk_btn(win,"Confirm",confirm,color=GREEN,fg=BG2).pack(pady=8)

    def _process_csv(self, text):
        groups = parse_grouped_csv(text)
        if not groups:
            messagebox.showerror("Error","No valid box groups found in CSV."); return
        self.groups = groups; self.current_group_idx=0; self.box_configs=[]
        self.continue_mode=None; self.saved_settings=None; self.no_csv_mode=False
        self._build_config_screen()

    def _start_no_csv(self):
        self.no_csv_mode=True; self.current_group_idx=0; self.box_configs=[]
        self.continue_mode=None; self.saved_settings=None
        self.groups=[{"box_name":"Manual Entry","items":[
            {"id":"","name":"","extra":[],"rate":None,"item_cnt":1}]}]
        self._build_config_screen()

    # ── Config screen ─────────────────────────────────────────────────────────
    def _build_config_screen(self):
        self._clear()
        idx      = self.current_group_idx
        grp      = self.groups[idx]
        box_name = grp["box_name"]
        total    = len(self.groups)
        s        = self.saved_settings or {}
        prev_box_name = s.get("box_name","")

        # Wrapper uses grid so rows have explicit weights
        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0, weight=0)
        wrap.grid_rowconfigure(1, weight=1)
        wrap.grid_rowconfigure(2, weight=0)
        wrap.grid_columnconfigure(0, weight=1)

        # Header — row 0, fixed height
        hdr = tk.Frame(wrap, bg=BG2)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text=f"Box {idx+1} / {total}:  {box_name}",
                 font=("Consolas",14,"bold"), bg=BG2, fg=ACC1, pady=8
                 ).pack(side="left", padx=15)
        if self.continue_mode:
            tk.Label(hdr, text="🤖 AUTO" if self.continue_mode=="automate" else "👁 MONITOR",
                     font=("Consolas",10), bg=BG2, fg=ACC4).pack(side="right", padx=15)

        # Nav bar — row 2, always visible
        nav = tk.Frame(wrap, bg=BG2)
        nav.grid(row=2, column=0, sticky="ew")

        # Scrollable content — row 1, expands to fill
        scroll_host = tk.Frame(wrap, bg=BG)
        scroll_host.grid(row=1, column=0, sticky="nsew")
        canvas, container = mk_scroll_canvas(scroll_host)

        def section(title): return mk_section(container, title)
        def row_entry(p,lbl,var,w=38):
            r=tk.Frame(p,bg=BG); r.pack(fill="x",padx=8,pady=2)
            tk.Label(r,text=lbl,width=26,anchor="w",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
            tk.Entry(r,textvariable=var,width=w,bg=BG3,fg=FG,insertbackground=FG,
                     font=("Consolas",9),relief="flat").pack(side="left",padx=4)
        def row_num(p,lbl,var,w=10):
            r=tk.Frame(p,bg=BG); r.pack(fill="x",padx=8,pady=2)
            tk.Label(r,text=lbl,width=26,anchor="w",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
            tk.Entry(r,textvariable=var,width=w,bg=BG3,fg=FG,insertbackground=FG,
                     font=("Consolas",9),relief="flat").pack(side="left",padx=4)
        def note(p,txt):
            tk.Label(p,text=txt,bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(anchor="w",padx=10,pady=(3,0))

        # cfg_override from CSV — must be resolved FIRST before any StringVar uses it
        cfg_ov = grp.get("cfg_override", {})

        try:    next_id = str(int(s.get("id",""))+1)
        except: next_id = s.get("id","")
        if not next_id:
            # Fall back to last used box ID from session
            last = _get_last_id("t1_box", 0)
            next_id = str(last + 1) if last else ""
        if cfg_ov.get("id"): next_id = cfg_ov["id"]
        saved_name_tmpl = s.get("name_template","")
        if saved_name_tmpl and prev_box_name:
            proposed_name = apply_name_template(saved_name_tmpl, prev_box_name, box_name)
        else:
            proposed_name = box_name
        used_names = [c["name"] for c in self.box_configs]
        proposed_name = deduplicate_name(proposed_name, used_names)
        if cfg_ov.get("name"): proposed_name = deduplicate_name(cfg_ov["name"], used_names)
        saved_cmt_tmpl = s.get("comment_template","")
        initial_comment = apply_name_template(saved_cmt_tmpl,prev_box_name,box_name) if saved_cmt_tmpl and prev_box_name else s.get("comment","Special dice that contains amazing items.")
        if cfg_ov.get("comment"): initial_comment = cfg_ov["comment"]
        saved_use_tmpl = s.get("use_template","")
        initial_use = apply_name_template(saved_use_tmpl,prev_box_name,box_name) if saved_use_tmpl and prev_box_name else s.get("use","Event Box.")
        if cfg_ov.get("use"): initial_use = cfg_ov["use"]

        v_id=tk.StringVar(value=next_id); v_name=tk.StringVar(value=proposed_name)
        v_comment=tk.StringVar(value=initial_comment); v_use=tk.StringVar(value=initial_use)
        v_file_name=tk.StringVar(value=cfg_ov.get("file_name") or s.get("file_name",r"data\item\itm_pre_107.nri"))
        v_bundle_num=tk.StringVar(value=cfg_ov.get("bundle_num") or s.get("bundle_num","0"))
        v_cmt_file_name=tk.StringVar(value=cfg_ov.get("cmt_file_name") or s.get("cmt_file_name",r"data\item\itm_pre_illu_107.nri"))
        v_cmt_bundle=tk.StringVar(value=cfg_ov.get("cmt_bundle_num") or s.get("cmt_bundle_num","0"))
        ov_checks = cfg_ov.get("opt_checks", s.get("opt_checks",[False]*8))
        ov_recycle = cfg_ov.get("opt_recycle", s.get("opt_recycle",0))
        ov_chr     = cfg_ov.get("chr_type_flags", s.get("chr_type_flags",[]))
        opt_check_vars=[tk.BooleanVar(value=ov_checks[i] if i<len(ov_checks) else False) for i in range(8)]
        opt_recycle_var=tk.IntVar(value=ov_recycle)
        chr_selected=list(ov_chr)
        present_type_var=tk.IntVar(value=s.get("present_type",0))
        drop_cnt_var=tk.StringVar(value=s.get("drop_cnt","1"))
        default_rate_var=tk.StringVar(value=s.get("default_rate","50"))
        remember_present=tk.BooleanVar(value=s.get("remember_present",False))
        v_weight=tk.StringVar(value=cfg_ov.get("weight") or s.get("weight","1"))
        v_value=tk.StringVar(value=cfg_ov.get("value") or s.get("value","0"))
        v_min_level=tk.StringVar(value=cfg_ov.get("min_level") or s.get("min_level","1"))
        v_money=tk.StringVar(value=cfg_ov.get("money") or s.get("money","0"))
        v_ticket=tk.StringVar(value=cfg_ov.get("ticket") or s.get("ticket","0"))
        dc_mode_var=tk.StringVar(value=s.get("dc_mode","flexible"))
        rate_mode_var=tk.StringVar(value=s.get("rate_mode","manual"))
        item_cnt_mode_var=tk.StringVar(value=s.get("item_cnt_mode","flexible"))
        item_cnt_univ_var=tk.StringVar(value=s.get("item_cnt_univ","1"))

        live_items=list(grp["items"])
        item_rate_vars=[]; item_cnt_vars=[]
        for it in live_items:
            if present_type_var.get()==2: initial_r="100"
            elif it.get("rate") is not None: initial_r=str(it["rate"])
            else:
                try: initial_r=str(int(s.get("default_rate","50")))
                except: initial_r="50"
            item_rate_vars.append(tk.StringVar(value=initial_r))
            item_cnt_vars.append(tk.StringVar(value=str(it.get("item_cnt",1))))

        # Basic Info
        sb = section("  ItemParam – Basic Info  ")
        row_entry(sb,"Box ID (itemparam):",v_id,20)
        note(sb,"  Name auto-filled from CSV header — editable.")
        row_entry(sb,"Name (CDATA):",v_name)
        row_entry(sb,"Comment (CDATA):",v_comment)
        row_entry(sb,"Use (CDATA):",v_use)

        # Filepaths
        sf = section("  Filepaths & Bundle Numbers  ")
        note(sf,"  FileName — also written to InvFileName.")
        row_entry(sf,"FileName (CDATA):",v_file_name)
        row_num(sf,"BundleNum:",v_bundle_num,8)
        row_entry(sf,"CmtFileName (CDATA):",v_cmt_file_name)
        row_num(sf,"CmtBundleNum:",v_cmt_bundle,8)

        # Options
        so = section("  Options  ")
        chk_frm=tk.Frame(so,bg=BG); chk_frm.pack(anchor="w",padx=8,pady=4)
        for i,(lbl,_) in enumerate(OPTIONS_CHECKS):
            tk.Checkbutton(chk_frm,text=lbl,variable=opt_check_vars[i],
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                           font=("Consolas",9)).grid(row=i//4,column=i%4,sticky="w",padx=6,pady=2)
        rec_frm=tk.Frame(so,bg=BG); rec_frm.pack(anchor="w",padx=8,pady=4)
        tk.Label(rec_frm,text="Recycle:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left",padx=(0,8))
        for lbl,val in [("None",0),("Recyclable",262144),("Non-Recyclable",8388608)]:
            tk.Radiobutton(rec_frm,text=lbl,variable=opt_recycle_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                           font=("Consolas",9)).pack(side="left",padx=6)
        nc_container=tk.Frame(so,bg=BG); nc_container.pack(anchor="w",padx=8,pady=2)
        nc_inner=tk.Frame(nc_container,bg=BG)
        tk.Label(nc_inner,text="Ticket value (NCash = tickets × 133, rounded):",
                 bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Entry(nc_inner,textvariable=v_ticket,width=10,bg=BG3,fg=FG,
                 insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=8)
        ncash_lbl=tk.Label(nc_inner,text="→ NCash: 0",bg=BG,fg=GREEN,font=("Consolas",9))
        ncash_lbl.pack(side="left")
        def update_ncash(*_):
            try: ncash_lbl.config(text=f"→ NCash: {round(float(v_ticket.get())*133)}")
            except: ncash_lbl.config(text="→ NCash: ?")
        v_ticket.trace_add("write",update_ncash); update_ncash()
        def toggle_ncash(*_):
            if opt_recycle_var.get()==262144: nc_inner.pack(anchor="w")
            else: nc_inner.pack_forget()
        opt_recycle_var.trace_add("write",toggle_ncash); toggle_ncash()

        # ChrTypeFlags
        sc_chr=section("  ChrTypeFlags  ")
        pf=tk.Frame(sc_chr,bg=BG); pf.pack(fill="x",padx=8,pady=4)
        tk.Label(pf,text="Character Type:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        chr_name_combo=ttk.Combobox(pf,values=CHR_NAMES,state="readonly",width=14,font=("Consolas",9))
        chr_name_combo.pack(side="left",padx=(6,12))
        tk.Label(pf,text="Job:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        chr_job_combo=ttk.Combobox(pf,values=CHR_JOBS,state="readonly",width=6,font=("Consolas",9))
        chr_job_combo.pack(side="left",padx=(6,12))
        lb_frm=tk.Frame(sc_chr,bg=BG); lb_frm.pack(fill="x",padx=8,pady=(0,6))
        tk.Label(lb_frm,text="Added:",bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(anchor="w")
        chr_lb=tk.Listbox(lb_frm,height=4,width=36,bg=BG3,fg=FG,
                          font=("Consolas",9),selectbackground=BG4,activestyle="none")
        chr_lb.pack(anchor="w")
        def refresh_chr_lb():
            chr_lb.delete(0,"end")
            for val in chr_selected: chr_lb.insert("end",CHR_FLAG_REVERSE.get(val,str(val)))
        def add_chr_flag():
            name=chr_name_combo.get(); job=chr_job_combo.get()
            if not name or not job: return
            key=f"{name} {job}"; val=CHR_FLAG_MAP.get(key)
            if val and val not in chr_selected and len(chr_selected)<24:
                chr_selected.append(val); refresh_chr_lb()
        def rem_chr_flag():
            sel=chr_lb.curselection()
            if sel: chr_selected.pop(sel[0]); refresh_chr_lb()
        tk.Button(pf,text="+",command=add_chr_flag,bg=GREEN,fg=BG2,
                  font=("Consolas",11,"bold"),relief="flat",width=3).pack(side="left",padx=2)
        tk.Button(pf,text="−",command=rem_chr_flag,bg=ACC3,fg=BG2,
                  font=("Consolas",11,"bold"),relief="flat",width=3).pack(side="left",padx=2)
        refresh_chr_lb()

        # Numeric
        sn=section("  Numeric Fields  ")
        rf=tk.Frame(sn,bg=BG); rf.pack(fill="x",padx=8,pady=4)
        for ci,(lbl,var) in enumerate([("Weight:",v_weight),("Value:",v_value),
                                        ("MinLevel:",v_min_level),("Money:",v_money)]):
            tk.Label(rf,text=lbl,bg=BG,fg=FG,font=("Consolas",9),width=10,anchor="w").grid(row=0,column=ci*2,padx=4)
            tk.Entry(rf,textvariable=var,width=10,bg=BG3,fg=FG,insertbackground=FG,
                     font=("Consolas",9),relief="flat").grid(row=0,column=ci*2+1,padx=4)

        # PresentItemParam2 settings
        sp=section("  PresentItemParam2 Settings  ")
        id_disp=tk.Label(sp,text=f"Box ID: {next_id}  |  {box_name}",bg=BG,fg=FG_GREY,font=("Consolas",9))
        id_disp.pack(anchor="w",padx=8,pady=(4,0))
        v_id.trace_add("write",lambda *_: id_disp.config(text=f"Box ID: {v_id.get() or '—'}  |  {box_name}"))
        type_frm=tk.Frame(sp,bg=BG); type_frm.pack(anchor="w",padx=8,pady=4)
        tk.Label(type_frm,text="Drop Type:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        for lbl,val in [("Random",0),("Distributive",2)]:
            tk.Radiobutton(type_frm,text=lbl,variable=present_type_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                           font=("Consolas",9)).pack(side="left",padx=8)
        tk.Checkbutton(type_frm,text="Remember this setting",variable=remember_present,
                       bg=BG,fg=ACC4,selectcolor=BG3,activebackground=BG,
                       font=("Consolas",9)).pack(side="left",padx=14)

        dc_mode_frm=tk.Frame(sp,bg=BG); dc_mode_frm.pack(anchor="w",padx=8,pady=(2,0))
        tk.Label(dc_mode_frm,text="DropCnt:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Radiobutton(dc_mode_frm,text="Flexible  (= item count)",variable=dc_mode_var,value="flexible",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_dc_mode()).pack(side="left",padx=6)
        tk.Radiobutton(dc_mode_frm,text="Manual",variable=dc_mode_var,value="manual",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_dc_mode()).pack(side="left",padx=4)
        dc_frm=tk.Frame(sp,bg=BG); dc_frm.pack(anchor="w",padx=8,pady=2)
        dc_lbl=tk.Label(dc_frm,text="DropCnt value:",bg=BG,fg=FG,font=("Consolas",9))
        dc_ent=tk.Entry(dc_frm,textvariable=drop_cnt_var,width=6,bg=BG3,fg=FG,
                        insertbackground=FG,font=("Consolas",9),relief="flat")
        rate_note=tk.Label(sp,text="",bg=BG,fg=FG_GREY,font=("Consolas",8))
        rate_note.pack(anchor="w",padx=8)
        def _toggle_dc_mode():
            if dc_mode_var.get()=="manual": dc_lbl.pack(side="left"); dc_ent.pack(side="left",padx=6)
            else: dc_lbl.pack_forget(); dc_ent.pack_forget()

        dr_mode_frm=tk.Frame(sp,bg=BG); dr_mode_frm.pack(anchor="w",padx=8,pady=(4,0))
        tk.Label(dr_mode_frm,text="DropRate:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Radiobutton(dr_mode_frm,text="Manual",variable=rate_mode_var,value="manual",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_rate_mode()).pack(side="left",padx=6)
        tk.Radiobutton(dr_mode_frm,text="Distributive  (set all entries equally)",
                       variable=rate_mode_var,value="distributive",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_rate_mode()).pack(side="left",padx=4)
        dr_dist_frm=tk.Frame(sp,bg=BG)
        tk.Label(dr_dist_frm,text="Default DropRate:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        dr_dist_ent=tk.Entry(dr_dist_frm,textvariable=default_rate_var,width=6,bg=BG3,fg=FG,
                             insertbackground=FG,font=("Consolas",9),relief="flat")
        dr_dist_ent.pack(side="left",padx=6)
        dr_dist_preview=tk.Label(dr_dist_frm,text="",bg=BG,fg=FG_GREY,font=("Consolas",8))
        dr_dist_preview.pack(side="left")
        def _apply_dist_rate(*_):
            if rate_mode_var.get()=="distributive":
                try:   val=str(int(default_rate_var.get()))
                except: val="50"
                for var in item_rate_vars: var.set(val)
                dr_dist_preview.config(text=f"→ all items set to {val}")
        default_rate_var.trace_add("write",_apply_dist_rate)

        # Box Contents table — defined BEFORE the ItemCnt radios so the command lambda works
        sec_items=tk.LabelFrame(container,text="  Box Contents  ",
                                bg=BG,fg=BLUE,font=("Consolas",10,"bold"),bd=1,relief="groove")
        sec_items.pack(fill="x",padx=12,pady=5)
        items_outer=tk.Frame(sec_items,bg=BG); items_outer.pack(fill="x",padx=8,pady=4)
        rate_entry_widgets=[]; icnt_entry_widgets=[]
        icnt_univ_frm=tk.Frame(sp,bg=BG)
        tk.Label(icnt_univ_frm,text="ItemCnt (all):",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Entry(icnt_univ_frm,textvariable=item_cnt_univ_var,width=6,bg=BG3,fg=FG,
                 insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=6)

        def _rebuild_items_table():
            nonlocal rate_entry_widgets, icnt_entry_widgets
            while len(item_rate_vars)<len(live_items): item_rate_vars.append(tk.StringVar(value=default_rate_var.get() or "50"))
            while len(item_rate_vars)>len(live_items): item_rate_vars.pop()
            while len(item_cnt_vars)<len(live_items): item_cnt_vars.append(tk.StringVar(value="1"))
            while len(item_cnt_vars)>len(live_items): item_cnt_vars.pop()
            for w in items_outer.winfo_children(): w.destroy()
            rate_entry_widgets=[]; icnt_entry_widgets=[]
            sec_items.config(text=f"  Box Contents ({len(live_items)} items)  ")
            icnt_mode=item_cnt_mode_var.get(); is_no_csv=self.no_csv_mode
            if icnt_mode=="universal": icnt_univ_frm.pack(anchor="w",padx=8,pady=(0,4))
            else: icnt_univ_frm.pack_forget()
            cols=[("#",3),("ID",9 if is_no_csv else 10),("Name",40 if is_no_csv else 44),("DropRate",8)]
            if icnt_mode=="manual": cols.append(("ItemCnt",7))
            if is_no_csv: cols.append(("",5))
            for ci,(txt,w) in enumerate(cols):
                tk.Label(items_outer,text=txt,width=w,bg=BG2,fg=BLUE,
                         font=("Consolas",9,"bold"),anchor="w").grid(row=0,column=ci,padx=2,pady=2,sticky="w")
            for i,item in enumerate(live_items):
                bg = BG if i%2==0 else BG2; col=0
                tk.Label(items_outer,text=str(i),width=3,bg=bg,fg=FG_GREY,
                         font=("Consolas",9)).grid(row=i+1,column=col,padx=2,pady=1); col+=1
                if is_no_csv:
                    id_var=tk.StringVar(value=item.get("id",""))
                    tk.Entry(items_outer,textvariable=id_var,width=9,bg=BG3,fg=FG,
                             insertbackground=FG,font=("Consolas",9),relief="flat").grid(row=i+1,column=col,padx=2)
                    id_var.trace_add("write",lambda *_,v=id_var,it=item: it.update({"id":v.get()}))
                    name_var=tk.StringVar(value=item.get("name",""))
                    tk.Entry(items_outer,textvariable=name_var,width=40,bg=BG3,fg=FG,
                             insertbackground=FG,font=("Consolas",9),relief="flat").grid(row=i+1,column=col+1,padx=2,sticky="w")
                    name_var.trace_add("write",lambda *_,v=name_var,it=item: it.update({"name":v.get()}))
                else:
                    tk.Label(items_outer,text=item.get("id",""),width=9,bg=bg,fg=FG,
                             font=("Consolas",9)).grid(row=i+1,column=col,padx=2)
                    tk.Label(items_outer,text=item.get("name","")[:52],width=44,bg=bg,fg=FG_DIM,
                             font=("Consolas",9),anchor="w").grid(row=i+1,column=col+1,padx=2,sticky="w")
                col+=2
                rate_ent=tk.Entry(items_outer,textvariable=item_rate_vars[i],width=8,bg=BG3,fg=FG,
                                  insertbackground=FG,font=("Consolas",9),relief="flat")
                rate_ent.grid(row=i+1,column=col,padx=2); rate_entry_widgets.append(rate_ent); col+=1
                if icnt_mode=="manual":
                    icnt_ent=tk.Entry(items_outer,textvariable=item_cnt_vars[i],width=7,bg=BG3,fg=FG,
                                      insertbackground=FG,font=("Consolas",9),relief="flat")
                    icnt_ent.grid(row=i+1,column=col,padx=2); icnt_entry_widgets.append(icnt_ent); col+=1
                if is_no_csv:
                    def make_remove(idx):
                        def _rem():
                            if len(live_items)<=1: return
                            live_items.pop(idx)
                            if idx<len(item_rate_vars): item_rate_vars.pop(idx)
                            if idx<len(item_cnt_vars): item_cnt_vars.pop(idx)
                            _rebuild_items_table()
                        return _rem
                    tk.Button(items_outer,text="−",width=2,command=make_remove(i),
                              bg=ACC3,fg=BG2,font=("Consolas",9,"bold"),relief="flat"
                              ).grid(row=i+1,column=col,padx=2)
            if is_no_csv and len(live_items)<20:
                def _add_item():
                    live_items.append({"id":"","name":"","extra":[],"rate":None,"item_cnt":1})
                    _rebuild_items_table()
                tk.Button(items_outer,text="＋ Add row",command=_add_item,
                          bg=GREEN,fg=BG2,font=("Consolas",9),relief="flat",padx=6,pady=2
                          ).grid(row=len(live_items)+1,column=0,columnspan=6,sticky="w",padx=4,pady=4)
            _toggle_rate_mode_inner()

        # ItemCnt mode radios — placed here so _rebuild_items_table is already defined
        icnt_mode_frm=tk.Frame(sp,bg=BG); icnt_mode_frm.pack(anchor="w",padx=8,pady=(4,2))
        tk.Label(icnt_mode_frm,text="ItemCnt:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        for _lbl_ic,_val_ic in [("Flexible (=1)","flexible"),("Universal","universal"),("Manual (per row)","manual")]:
            tk.Radiobutton(icnt_mode_frm,text=_lbl_ic,variable=item_cnt_mode_var,value=_val_ic,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                           command=_rebuild_items_table).pack(side="left",padx=4)

        def _toggle_rate_mode_inner():
            is_dist=rate_mode_var.get()=="distributive"
            for ent in rate_entry_widgets:
                ent.config(state="disabled" if is_dist else "normal",disabledforeground=FG_GREY)
            if is_dist: _apply_dist_rate()

        def _toggle_rate_mode():
            if rate_mode_var.get()=="distributive":
                dr_dist_frm.pack(anchor="w",padx=8,pady=2); _toggle_rate_mode_inner()
            else:
                dr_dist_frm.pack_forget()
                for ent in rate_entry_widgets: ent.config(state="normal")
                dr_dist_preview.config(text="")

        def toggle_drop_fields(*_):
            _toggle_dc_mode(); rate_note.config(text="")

        _rebuild_items_table()
        present_type_var.trace_add("write",toggle_drop_fields)
        _toggle_dc_mode(); toggle_drop_fields(); _toggle_rate_mode()

        # Gather / nav
        def gather_config():
            try: ncash=round(float(v_ticket.get() or "0")*133)
            except: ncash=0
            item_rates=[]
            for iv in item_rate_vars:
                try: item_rates.append(int(iv.get()))
                except: item_rates.append(100 if present_type_var.get()==2 else 50)
            icnt_mode=item_cnt_mode_var.get()
            if icnt_mode=="flexible": item_cnts=[1]*len(live_items)
            elif icnt_mode=="universal":
                try: uval=int(item_cnt_univ_var.get()) or 1
                except: uval=1
                item_cnts=[uval]*len(live_items)
            else:
                item_cnts=[]
                for iv in item_cnt_vars:
                    try: item_cnts.append(int(iv.get()) or 1)
                    except: item_cnts.append(1)
            return {
                "id":v_id.get().strip(),"name":v_name.get(),
                "name_template":v_name.get(),"comment":v_comment.get(),
                "comment_template":v_comment.get(),"use":v_use.get(),
                "use_template":v_use.get(),"box_name":box_name,
                "items":list(live_items),"item_rates":item_rates,"item_cnts":item_cnts,
                "item_cnt_mode":icnt_mode,"item_cnt_univ":item_cnt_univ_var.get() or "1",
                "file_name":v_file_name.get(),"bundle_num":v_bundle_num.get() or "0",
                "cmt_file_name":v_cmt_file_name.get(),"cmt_bundle_num":v_cmt_bundle.get() or "0",
                "weight":v_weight.get() or "1","value":v_value.get() or "0",
                "min_level":v_min_level.get() or "1","money":v_money.get() or "0",
                "ncash":ncash,"ticket":v_ticket.get() or "0",
                "opt_checks":[bv.get() for bv in opt_check_vars],
                "opt_recycle":opt_recycle_var.get(),"chr_type_flags":list(chr_selected),
                "present_type":present_type_var.get(),
                "drop_cnt":drop_cnt_var.get() or "1" if dc_mode_var.get()=="manual" else str(len(live_items)),
                "default_rate":default_rate_var.get() or "50",
                "remember_present":remember_present.get(),
                "dc_mode":dc_mode_var.get(),"rate_mode":rate_mode_var.get(),
            }

        def save_settings(cfg):
            self.saved_settings={k:v for k,v in cfg.items() if k!="items"}

        # nav is already created above, pinned to bottom of outer

        def go_next():
            cfg=gather_config()
            if not cfg["id"]: messagebox.showwarning("Missing ID","Please enter a Box ID."); return
            blanks=[]
            if not cfg["name"].strip(): blanks.append("Name")
            if not cfg["file_name"].strip(): blanks.append("FileName")
            if not cfg["cmt_file_name"].strip(): blanks.append("CmtFileName")
            if blanks:
                if not messagebox.askyesno("Missed a spot","These fields are empty:\n\n  "+"\n  ".join(blanks)+"\n\nContinue anyway?"): return
            self.box_configs.append(cfg); save_settings(cfg); self.current_group_idx+=1
            if self.no_csv_mode or self.current_group_idx>=len(self.groups): self._build_output_screen()
            elif self.continue_mode=="automate": self._automate_remaining(cfg)
            elif self.continue_mode=="monitor": self._build_config_screen()
            else: self._ask_automate_or_monitor(cfg)

        mk_btn(nav,"◀  Back to CSV",self._build_load_screen).pack(side="left",padx=10,pady=8)
        if idx>0:
            def go_prev():
                self.current_group_idx-=1
                if self.box_configs: self.box_configs.pop()
                self._build_config_screen()
            mk_btn(nav,"◀  Previous Box",go_prev).pack(side="left",padx=4,pady=8)
        if self.continue_mode:
            mk_btn(nav,"⚙ Change Mode",lambda: (setattr(self,"continue_mode",None),self._build_config_screen()),
                   color=BG4).pack(side="left",padx=4,pady=8)

        # ── Live-validating Generate/Next button ──────────────────────────
        next_lbl = "⚡  Generate XML ✓" if idx==total-1 else "Next Box ▶"
        next_btn = mk_btn(nav, next_lbl, go_next, color=GREEN, fg=BG2,
                          font=("Consolas",10,"bold"))
        next_btn.pack(side="right", padx=10, pady=8)

        # Validation label (shows what's missing)
        val_lbl = tk.Label(nav, text="", bg=BG2, fg=ACC3, font=("Consolas",8))
        val_lbl.pack(side="right", padx=4)

        def _check_ready(*_):
            missing = []
            if not v_id.get().strip():           missing.append("ID")
            if not v_name.get().strip():          missing.append("Name")
            if not v_file_name.get().strip():     missing.append("FileName")
            if not v_cmt_file_name.get().strip(): missing.append("CmtFileName")
            # Check that at least one item has an ID
            has_item = any(it.get("id","").strip() for it in live_items)
            if not has_item:                      missing.append("Box Contents (need ≥1 item ID)")
            if missing:
                next_btn.config(state="disabled", bg=BG4, fg=FG_GREY)
                val_lbl.config(text="Missing: " + ", ".join(missing))
            else:
                next_btn.config(state="normal", bg=GREEN, fg=BG2)
                val_lbl.config(text="✓ Ready")

        # Trace required vars
        for _var in (v_id, v_name, v_file_name, v_cmt_file_name):
            _var.trace_add("write", _check_ready)
        # Re-check after item table rebuilds (hook into _rebuild_items_table)
        _orig_rebuild = _rebuild_items_table
        def _rebuild_and_check():
            _orig_rebuild()
            self.after(50, _check_ready)
        # Patch live_items check — just schedule it
        self.after(100, _check_ready)

    def _ask_automate_or_monitor(self, last_cfg):
        remaining=len(self.groups)-self.current_group_idx
        win=tk.Toplevel(self.root); win.title("Continue?"); win.geometry("530x260")
        win.configure(bg=BG); win.grab_set()
        tk.Label(win,text=f"{remaining} box(es) remaining.",bg=BG,fg=FG,font=("Consolas",13,"bold")).pack(pady=14)
        tk.Label(win,text="How would you like to continue?",bg=BG,fg=FG_DIM,font=("Consolas",10)).pack()
        remember_var=tk.BooleanVar(value=False)
        bf=tk.Frame(win,bg=BG); bf.pack(pady=10)
        def choose(mode):
            if remember_var.get(): self.continue_mode=mode
            win.destroy()
            if mode=="automate": self._automate_remaining(last_cfg)
            else: self._build_config_screen()
        mk_btn(bf,"🤖  Automate  —  use saved settings for all remaining boxes",
               lambda:choose("automate"),color=ACC1,fg=BG2).pack(pady=5)
        mk_btn(bf,"👁  Monitor  —  review each box individually",
               lambda:choose("monitor"),color=BLUE,fg=BG2).pack(pady=5)
        tk.Checkbutton(win,text="Remember my choice for the rest of this session",
                       variable=remember_var,bg=BG,fg=ACC4,selectcolor=BG3,
                       activebackground=BG,font=("Consolas",9)).pack(pady=4)

    def _automate_remaining(self, last_cfg):
        try: base_id=int(last_cfg["id"])
        except:
            messagebox.showerror("Error","Cannot automate: last Box ID was not a plain integer.")
            self.continue_mode="monitor"; self._build_config_screen(); return
        try: default_rate=int(last_cfg.get("default_rate",50))
        except: default_rate=50
        id_counter=base_id+1; prev_name=last_cfg["box_name"]
        used_names=[c["name"] for c in self.box_configs]
        for i in range(self.current_group_idx, len(self.groups)):
            grp=self.groups[i]; cfg=copy.deepcopy(last_cfg)
            cfg["id"]=str(id_counter)
            proposed=apply_name_template(cfg.get("name_template",""),prev_name,grp["box_name"])
            proposed=deduplicate_name(proposed,used_names)
            cfg["name"]=proposed; cfg["name_template"]=proposed
            cfg["comment"]=apply_name_template(cfg.get("comment_template",""),prev_name,grp["box_name"])
            cfg["use"]=apply_name_template(cfg.get("use_template",""),prev_name,grp["box_name"])
            cfg["comment_template"]=cfg["comment"]; cfg["use_template"]=cfg["use"]
            cfg["box_name"]=grp["box_name"]; cfg["items"]=grp["items"]
            is_distrib=cfg["present_type"]==2
            cfg["item_rates"]=[100 if is_distrib else (it.get("rate") if it.get("rate") is not None else default_rate) for it in grp["items"]]
            icnt_mode=cfg.get("item_cnt_mode","flexible")
            if icnt_mode=="flexible": cfg["item_cnts"]=[1]*len(grp["items"])
            elif icnt_mode=="universal":
                try: uval=int(cfg.get("item_cnt_univ","1")) or 1
                except: uval=1
                cfg["item_cnts"]=[uval]*len(grp["items"])
            else: cfg["item_cnts"]=[it.get("item_cnt",1) for it in grp["items"]]
            used_names.append(proposed); prev_name=grp["box_name"]
            self.box_configs.append(cfg); id_counter+=1
        self.current_group_idx=len(self.groups); self._build_output_screen()

    def _build_output_screen(self, _compound_rows=None, _exchange_rows=None):
        self._clear()
        if _compound_rows is None: _compound_rows = []
        if _exchange_rows is None: _exchange_rows = []

        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0,weight=0); wrap.grid_rowconfigure(1,weight=1)
        wrap.grid_rowconfigure(2,weight=0); wrap.grid_columnconfigure(0,weight=1)
        hdr = tk.Frame(wrap, bg=BG2); hdr.grid(row=0,column=0,sticky="ew")
        tk.Label(hdr,text="Generated XML Output",font=("Consolas",14,"bold"),
                 bg=BG2,fg=ACC1,pady=8).pack(side="left",padx=15)

        nb_host = tk.Frame(wrap, bg=BG); nb_host.grid(row=1,column=0,sticky="nsew")
        nb=ttk.Notebook(nb_host); nb.pack(fill="both",expand=True,padx=8,pady=4)

        itemparam_rows=[]; presentparam_rows=[]; recycle_except_rows=[]
        include_present = {}   # cfg["id"] -> bool

        for cfg in self.box_configs:
            try: default_rate=int(cfg.get("default_rate",50))
            except: default_rate=50
            try: drop_cnt=int(cfg.get("drop_cnt",1))
            except: drop_cnt=1
            is_distrib=(cfg["present_type"]==2)
            items_with_rates=[]
            for j,it in enumerate(cfg["items"]):
                rate=100 if is_distrib else (cfg["item_rates"][j] if j<len(cfg["item_rates"]) else default_rate)
                items_with_rates.append({**it,"rate":rate})
            item_cnts=cfg.get("item_cnts") or [1]*len(cfg["items"])
            itemparam_rows.append(build_itemparam_row(cfg))
            presentparam_rows.append(build_presentparam_row(cfg["id"],items_with_rates,cfg["present_type"],drop_cnt,default_rate,item_cnts=item_cnts,box_name=cfg.get("name","")))
            include_present[cfg["id"]] = True
            if cfg.get("opt_recycle",0) in (0,8388608):
                recycle_except_rows.append(build_recycle_except_row(cfg["id"],cfg["name"]))

        csv_lines=["ID,BoxName"]+[f"{c['id']},{c['box_name']}" for c in self.box_configs]
        csv_text="\n".join(csv_lines)
        self.session.box_id_list_csv = csv_text
        self.session.box_id_map = {c["id"]: c["box_name"] for c in self.box_configs}

        # Save last box ID used
        if self.box_configs:
            try: _set_last_id("t1_box", int(self.box_configs[-1]["id"]))
            except: pass

        _exports=[]
        _exports.append(("itemparam_rows.xml","\n".join(itemparam_rows)))
        make_output_tab(nb,"itemparam.xml rows","\n".join(itemparam_rows),"itemparam_rows.xml",self.root)

        _exports.append(("presentparam_rows.xml","\n".join(presentparam_rows)))
        make_output_tab(nb,"PresentItemParam2.xml rows","\n".join(presentparam_rows),"presentparam_rows.xml",self.root)

        _exports.append(("box_id_list.csv",csv_text))
        make_output_tab(nb,"Box ID List (→ Tool 2)","\n".join(csv_lines),"box_id_list.csv",self.root)

        if recycle_except_rows:
            _exports.append(("RecycleExceptItem_rows.xml","\n".join(recycle_except_rows)))
            make_output_tab(nb,"RecycleExceptItem.xml rows","\n".join(recycle_except_rows),"RecycleExceptItem_rows.xml",self.root)

        # Compound/exchange tabs
        if _compound_rows:
            cp_xml = "\n".join(r[0] for r in _compound_rows)
            cl_xml = "\n".join(r[1] for r in _compound_rows)
            _exports += [("Compound_Potion_rows.xml",cp_xml),("Compounder_Location_rows.xml",cl_xml)]
            make_output_tab(nb,"Compound_Potion rows",cp_xml,"Compound_Potion_rows.xml",self.root)
            make_output_tab(nb,"Compounder_Location rows",cl_xml,"Compounder_Location_rows.xml",self.root)
        if _exchange_rows:
            es_xml = "\n".join(r[0] for r in _exchange_rows)
            el_xml = "\n".join(r[1] for r in _exchange_rows)
            _exports += [("ExchangeShopContents_rows.xml",es_xml),("Exchange_Location_rows.xml",el_xml)]
            make_output_tab(nb,"ExchangeShopContents rows",es_xml,"ExchangeShopContents_rows.xml",self.root)
            make_output_tab(nb,"Exchange_Location rows",el_xml,"Exchange_Location_rows.xml",self.root)

        nb.select(0)

        # ── Nav footer ────────────────────────────────────────────────────
        bot = tk.Frame(wrap, bg=BG2); bot.grid(row=2,column=0,sticky="ew")

        # Compound/exchange post-generate button
        _ce_remember = tk.StringVar(value="ask")   # "ask"|"compound"|"exchange"|"none"
        def _open_ce_for_box(cfg_box):
            """Open compound/exchange dialog for a single box config."""
            def _on_compound(ce_cfg):
                cpr = build_compound_row(ce_cfg)
                clr = build_compound_location_row(ce_cfg["compound_id"])
                _compound_rows.append((cpr, clr))
                messagebox.showinfo("Added", f"Compound row added.\nRegenerate to see updated tabs.")
            def _on_exchange(ce_cfg):
                esr = build_exchange_row(ce_cfg)
                elr = build_exchange_location_row(ce_cfg["exchange_id"])
                _exchange_rows.append((esr, elr))
                messagebox.showinfo("Added", f"Exchange row added.")
            _show_compound_exchange_dialog(
                self.root, cfg_box.get("name",""), cfg_box.get("comment",""),
                cfg_box.get("id",""), _on_compound, _on_exchange, lambda: None)

        def _add_ce():
            if len(self.box_configs) == 1:
                _open_ce_for_box(self.box_configs[0])
                self._build_output_screen(_compound_rows, _exchange_rows)
            else:
                # Pick which box
                win2 = tk.Toplevel(self.root); win2.title("Select Box"); win2.configure(bg=BG)
                tk.Label(win2,text="Which box to add Compound/Exchange for?",
                         bg=BG,fg=FG,font=("Consolas",10,"bold")).pack(pady=10,padx=12)
                for cfg_box in self.box_configs:
                    lbl = f"{cfg_box['id']} — {cfg_box.get('name','')}"
                    mk_btn(win2, lbl, lambda c=cfg_box: (win2.destroy(), _open_ce_for_box(c),
                           self._build_output_screen(_compound_rows, _exchange_rows)),
                           color=BG3).pack(fill="x",padx=12,pady=2)
                mk_btn(win2,"Cancel",win2.destroy).pack(pady=8)

        def export_all():
            folder=filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            saved=[]
            for fname,content in _exports:
                with open(os.path.join(folder,fname),"w",encoding="utf-8") as f: f.write(content)
                saved.append(fname)
            messagebox.showinfo("Export Complete",f"Saved to:\n{folder}\n\n"+"\n".join(saved))

        mk_btn(bot,"💾  Export All Files",export_all,color=ACC1,fg=BG2,
               font=("Consolas",11,"bold")).pack(side="left",padx=14,pady=6)
        mk_btn(bot,"⚗  Add Compound/Exchange",_add_ce,color=BG4).pack(side="left",padx=4,pady=6)
        def _import_ce_t1():
            def _on_compound_cfgs(cfgs):
                for ce_cfg in cfgs:
                    cpr = build_compound_row(ce_cfg)
                    clr = build_compound_location_row(ce_cfg["compound_id"])
                    _compound_rows.append((cpr, clr))
                messagebox.showinfo("Imported",
                    f"{len(cfgs)} compound row(s) added.\nRegenerate to see updated tabs.")
            def _on_exchange_cfgs(cfgs):
                for ce_cfg in cfgs:
                    esr = build_exchange_row(ce_cfg)
                    elr = build_exchange_location_row(ce_cfg["exchange_id"])
                    _exchange_rows.append((esr, elr))
                messagebox.showinfo("Imported",
                    f"{len(cfgs)} exchange row(s) added.\nRegenerate to see updated tabs.")
            _ask_import_mode_then_file(self.root, _on_compound_cfgs, _on_exchange_cfgs)
        mk_btn(bot,"📥  Import CSV/Excel",_import_ce_t1,color=BG4).pack(side="left",padx=4,pady=6)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4,pady=6)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 2 — PresentItemParam2 Rate Adjuster
# ══════════════════════════════════════════════════════════════════════════════
class Tool2(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session; self.csv_text=""; self.xml_text=""
        self.item_lib={}; self.mode_var=tk.StringVar(value="automatic")
        self._rate_var=tk.StringVar(value="100"); self._count_var=tk.StringVar(value="1")
        self._lib_status=tk.StringVar(value="No library loaded  (item names won't appear)")
        self._mode_panel_frame=None; self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _build_load_screen(self):
        self._clear()
        tk.Label(self,text="PRESENTITEMPARAM2 RATE ADJUSTER",font=("Consolas",16,"bold"),
                 bg=BG,fg=ACC2).pack(pady=(18,2))
        tk.Label(self,text="CSV must contain the BOX IDs  (<Id> in PresentItemParam2).\nUse the 'Box ID List' exported by Tool 1, or any 2-col ID / Name CSV.",
                 bg=BG,fg=FG_GREY,font=("Consolas",8),justify="center").pack(pady=(0,6))

        csv_frm=mk_section(self,"  Step 1 — Box ID CSV  ")
        csv_status=tk.StringVar(value="No file loaded")
        tk.Label(csv_frm,textvariable=csv_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_csv():
            p=filedialog.askopenfilename(filetypes=[("CSV","*.csv"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig") as f: self.csv_text=f.read()
            bm=parse_box_id_csv(self.csv_text)
            csv_status.set(f"✓  {os.path.basename(p)}  ({len(bm)} box IDs found)")
        mk_btn(csv_frm,"📂 Load CSV",load_csv,padx=10,pady=4).pack(side="right",padx=8,pady=6)
        def import_session_t2():
            if not self.session.box_id_list_csv:
                messagebox.showinfo("No Session Data","Run Tool 1 first to generate box IDs.")
                return
            self.csv_text = self.session.box_id_list_csv
            bm = parse_box_id_csv(self.csv_text)
            csv_status.set(f"✓  Imported from Tool 1  ({len(bm)} box IDs)")
        mk_btn(csv_frm,"⬇  Import from Tool 1",import_session_t2,color=ACC2,fg=BG2,padx=8,pady=4).pack(side="right",padx=4,pady=6)

        xml_frm=mk_section(self,"  Step 2 — PresentItemParam2.xml  ")
        xml_status=tk.StringVar(value="No file loaded")
        tk.Label(xml_frm,textvariable=xml_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_xml():
            p=filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig") as f: self.xml_text=f.read()
            xml_status.set(f"✓  {os.path.basename(p)}")
        mk_btn(xml_frm,"📂 Load XML",load_xml,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        mode_frm=mk_section(self,"  Step 3 — Mode  ")
        mf=tk.Frame(mode_frm,bg=BG); mf.pack(anchor="w",padx=10,pady=6)
        for lbl,val in [("Manual     — review and configure each box individually","manual"),
                        ("Automatic  — apply the same values to every matched box","automatic")]:
            tk.Radiobutton(mf,text=lbl,variable=self.mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10),
                           command=self._refresh_mode_panel).pack(anchor="w",pady=2)

        self._mode_panel_frame=tk.Frame(self,bg=BG)
        self._mode_panel_frame.pack(fill="x",padx=30,pady=2)
        self._refresh_mode_panel()
        mk_btn(self,"▶  Continue →",self._on_continue,color=GREEN,fg=BG2,
               font=("Consolas",12,"bold")).pack(pady=14)

    def _refresh_mode_panel(self):
        if not self._mode_panel_frame: return
        for w in self._mode_panel_frame.winfo_children(): w.destroy()
        if self.mode_var.get()=="automatic":
            frm=tk.LabelFrame(self._mode_panel_frame,text="  Adjustment Values  ",
                              bg=BG,fg=BLUE,font=("Consolas",10,"bold"),bd=1,relief="groove")
            frm.pack(fill="x")
            def num_row(lbl,var,note=""):
                r=tk.Frame(frm,bg=BG); r.pack(fill="x",padx=10,pady=4)
                tk.Label(r,text=lbl,width=18,anchor="w",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
                tk.Entry(r,textvariable=var,width=8,bg=BG3,fg=FG,insertbackground=FG,
                         font=("Consolas",9),relief="flat").pack(side="left",padx=6)
                tk.Label(r,text=note,bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(side="left")
            num_row("Adjust Rate:",self._rate_var,"(1–32766)  applied to every used DropRate_# slot")
            num_row("Adjust Count:",self._count_var,"(1–32766)  applied to every used ItemCnt_# slot")
            tk.Label(frm,text="  Type will be set to 2.  DropCnt will be set to the number of real items.",
                     bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(anchor="w",padx=10,pady=(0,6))
        else:
            frm=tk.LabelFrame(self._mode_panel_frame,text="  ItemParam Library (optional)  ",
                              bg=BG,fg=BLUE,font=("Consolas",10,"bold"),bd=1,relief="groove")
            frm.pack(fill="x")
            tk.Label(frm,textvariable=self._lib_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10,pady=6)
            def load_lib():
                folder=filedialog.askdirectory(title="Select folder containing ItemParam XML files")
                if not folder: return
                self.item_lib=load_itemparam_folder(folder)
                self._lib_status.set(f"✓  {len(self.item_lib)} items from {os.path.basename(folder)}")
            mk_btn(frm,"📂 Load Folder",load_lib,padx=10,pady=4).pack(side="right",padx=8,pady=6)

    def _on_continue(self):
        if not self.csv_text: messagebox.showwarning("Missing","Please load a CSV first."); return
        if not self.xml_text: messagebox.showwarning("Missing","Please load PresentItemParam2.xml first."); return
        box_map=parse_box_id_csv(self.csv_text)
        if not box_map: messagebox.showerror("Error","No box IDs found in CSV."); return
        matched=[]; seen=set()
        for row in ROW_RE.findall(self.xml_text):
            rid=_get_tag(row,"Id")
            if rid in box_map and rid not in seen:
                matched.append((rid,box_map[rid],row)); seen.add(rid)
        if not matched:
            messagebox.showwarning("No Matches","None of the CSV box IDs matched any <Id> in the XML.\n\nMake sure you're using the Box ID CSV from Tool 1."); return
        if self.mode_var.get()=="automatic":
            try:
                rate=int(self._rate_var.get()); count=int(self._count_var.get())
                if not (1<=rate<=32766) or not (1<=count<=32766): raise ValueError
            except: messagebox.showerror("Invalid","Rate and Count must be integers 1–32766."); return
            self._run_automatic(matched,rate,count)
        else: self._run_manual(matched)

    def _run_automatic(self, matched, rate, count):
        matched_ids={rid:row for rid,_,row in matched}; csv_rows=[]
        def replace_row(m):
            row=m.group(0); rid=_get_tag(row,"Id")
            if rid not in matched_ids: return row
            slots=real_drop_slots(row)
            cfg={"type":2,"drop_cnt":len(slots),"slots":[{"rate":rate,"count":count} for _ in slots]}
            new_row=apply_cfg_to_row(row,cfg)
            drop_ids=[v for _,v in real_drop_slots(new_row)]
            name=next((n for r,n,_ in matched if r==rid),"")
            csv_rows.append([rid,name,*drop_ids]); return new_row
        full_out=ROW_RE.sub(replace_row,self.xml_text)
        self._build_output_screen(full_out,csv_rows,len(matched))

    def _run_manual(self, matched):
        self.manual_matched=matched; self.manual_idx=0
        self.manual_configs={}; self.manual_saved=None; self.manual_continue_mode=None
        self._build_manual_screen()

    def _build_manual_screen(self):
        self._clear()
        idx=self.manual_idx; total=len(self.manual_matched)
        rid,csv_name,row_block=self.manual_matched[idx]
        slots=real_drop_slots(row_block)
        s=self.manual_saved or {}
        last_type=s.get("type",2); last_dc=s.get("drop_cnt",len(slots)); last_slots=s.get("slots",[])

        wrap=tk.Frame(self,bg=BG); wrap.pack(fill="both",expand=True)
        wrap.grid_rowconfigure(0,weight=0); wrap.grid_rowconfigure(1,weight=1); wrap.grid_rowconfigure(2,weight=0)
        wrap.grid_columnconfigure(0,weight=1)
        hdr=tk.Frame(wrap,bg=BG2); hdr.grid(row=0,column=0,sticky="ew")
        hdr_txt=f"  Box {idx+1} / {total}   ID: {rid}"
        if csv_name: hdr_txt+=f"   —   {csv_name}"
        tk.Label(hdr,text=hdr_txt,font=("Consolas",12,"bold"),bg=BG2,fg=ACC2,pady=8).pack(side="left",padx=10)
        if self.manual_continue_mode:
            tk.Label(hdr,text="🤖 AUTO" if self.manual_continue_mode=="automate" else "👁 MONITOR",
                     font=("Consolas",10),bg=BG2,fg=ACC4).pack(side="right",padx=15)
        nav2=tk.Frame(wrap,bg=BG2); nav2.grid(row=2,column=0,sticky="ew")
        scroll_host2=tk.Frame(wrap,bg=BG); scroll_host2.grid(row=1,column=0,sticky="nsew")
        canvas,cont=mk_scroll_canvas(scroll_host2)

        sec_type=mk_section(cont,"  Drop Type  ")
        type_var=tk.IntVar(value=last_type)
        tf=tk.Frame(sec_type,bg=BG); tf.pack(anchor="w",padx=8,pady=6)
        for lbl,val in [("Random (Type=0)",0),("Egalitarian (Type=2)",2)]:
            tk.Radiobutton(tf,text=lbl,variable=type_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(side="left",padx=8)
        dc_frame=tk.Frame(sec_type,bg=BG); dc_frame.pack(anchor="w",padx=8,pady=(0,6))
        dc_var=tk.StringVar(value=str(last_dc))
        dc_lbl=tk.Label(dc_frame,text="Types of Items (DropCnt):",bg=BG,fg=FG,font=("Consolas",9))
        dc_ent=tk.Entry(dc_frame,textvariable=dc_var,width=6,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat")
        dc_auto=tk.Label(dc_frame,text=f"DropCnt auto-set to {len(slots)}  (all items in this box)",bg=BG,fg=FG_GREY,font=("Consolas",9))
        def toggle_dc(*_):
            if type_var.get()==0: dc_auto.pack_forget(); dc_lbl.pack(side="left"); dc_ent.pack(side="left",padx=6)
            else: dc_lbl.pack_forget(); dc_ent.pack_forget(); dc_auto.pack(anchor="w")
        type_var.trace_add("write",toggle_dc); toggle_dc()

        sec_slots=mk_section(cont,f"  Drop Slots  ({len(slots)} items)  ")
        hrow=tk.Frame(sec_slots,bg=BG2); hrow.pack(fill="x",padx=8,pady=2)
        for txt,w in [("Slot #",6),("Drop ID",12),("Item Name",34),("Rate %",9),("Item Count",10)]:
            tk.Label(hrow,text=txt,width=w,bg=BG2,fg=BLUE,font=("Consolas",9,"bold"),anchor="w").pack(side="left",padx=3)
        slot_rate_vars=[]; slot_count_vars=[]
        for pos,(sidx,drop_id) in enumerate(slots):
            bg=BG if pos%2==0 else BG2
            srow=tk.Frame(sec_slots,bg=bg); srow.pack(fill="x",padx=8,pady=1)
            prev_r=last_slots[pos]["rate"] if pos<len(last_slots) else 100
            prev_c=last_slots[pos]["count"] if pos<len(last_slots) else 1
            tk.Label(srow,text=str(sidx),width=6,bg=bg,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=3)
            tk.Label(srow,text=drop_id,width=12,bg=bg,fg=BG4,font=("Consolas",9)).pack(side="left",padx=3)
            name=self.item_lib.get(drop_id,"—")
            tk.Label(srow,text=name[:36],width=34,bg=bg,fg=FG_DIM,font=("Consolas",9),anchor="w").pack(side="left",padx=3)
            rv=tk.StringVar(value=str(prev_r))
            tk.Entry(srow,textvariable=rv,width=7,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=3)
            slot_rate_vars.append(rv)
            cv=tk.StringVar(value=str(prev_c))
            tk.Entry(srow,textvariable=cv,width=7,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=3)
            slot_count_vars.append(cv)

        def gather():
            t=type_var.get()
            dc=len(slots) if t==2 else max(1,int(dc_var.get() or 1))
            sl=[]
            for rv,cv in zip(slot_rate_vars,slot_count_vars):
                try: r=max(1,min(32766,int(rv.get())))
                except: r=100
                try: c=max(1,min(32766,int(cv.get())))
                except: c=1
                sl.append({"rate":r,"count":c})
            return {"type":t,"drop_cnt":dc,"slots":sl}

        def save_and_advance():
            if type_var.get()==0:
                dc_raw=dc_var.get().strip()
                if not dc_raw or not dc_raw.isdigit() or int(dc_raw)<1:
                    if not messagebox.askyesno("Missed a spot","DropCnt is blank/invalid. Will default to 1. Continue?"): return
            cfg=gather(); self.manual_configs[rid]=cfg; self.manual_saved=cfg
            self.manual_idx+=1
            if self.manual_idx>=total: self._finish_manual()
            elif self.manual_continue_mode=="automate": self._automate_manual_remaining(cfg)
            elif self.manual_continue_mode=="monitor": self._build_manual_screen()
            else: self._ask_manual_mode(cfg)

        # Use nav2 which is pinned outside the scroll canvas
        mk_btn(nav2,"◀  Start Over",self._build_load_screen).pack(side="left",padx=8,pady=6)
        if idx>0:
            mk_btn(nav2,"◀  Prev",lambda:(setattr(self,"manual_idx",self.manual_idx-1),self._build_manual_screen())).pack(side="left",padx=4,pady=6)
        if self.manual_continue_mode:
            mk_btn(nav2,"⚙ Change Mode",lambda:(setattr(self,"manual_continue_mode",None),self._build_manual_screen()),color=BG4).pack(side="left",padx=4,pady=6)
        mk_btn(nav2,"Finish ✓" if idx==total-1 else "Next ▶",save_and_advance,color=GREEN,fg=BG2,font=("Consolas",10,"bold")).pack(side="right",padx=8,pady=6)

    def _ask_manual_mode(self, last_cfg):
        remaining=len(self.manual_matched)-self.manual_idx
        win=tk.Toplevel(self.root); win.title("Continue?"); win.geometry("520x240")
        win.configure(bg=BG); win.grab_set()
        tk.Label(win,text=f"{remaining} box(es) remaining.",bg=BG,fg=FG,font=("Consolas",13,"bold")).pack(pady=12)
        remember=tk.BooleanVar(value=False)
        bf=tk.Frame(win,bg=BG); bf.pack(pady=8)
        def choose(mode):
            if remember.get(): self.manual_continue_mode=mode
            win.destroy()
            if mode=="automate": self._automate_manual_remaining(last_cfg)
            else: self._build_manual_screen()
        mk_btn(bf,"🤖  Automate  —  copy settings to all remaining boxes",lambda:choose("automate"),color=ACC1,fg=BG2).pack(pady=4)
        mk_btn(bf,"👁  Monitor  —  review each box",lambda:choose("monitor"),color=BLUE,fg=BG2).pack(pady=4)
        tk.Checkbutton(win,text="Remember for rest of session",variable=remember,
                       bg=BG,fg=ACC4,selectcolor=BG3,activebackground=BG,font=("Consolas",9)).pack(pady=4)

    def _automate_manual_remaining(self, last_cfg):
        while self.manual_idx<len(self.manual_matched):
            rid,_,row_block=self.manual_matched[self.manual_idx]
            slots=real_drop_slots(row_block); cfg=copy.deepcopy(last_cfg)
            while len(cfg["slots"])<len(slots): cfg["slots"].append(cfg["slots"][-1] if cfg["slots"] else {"rate":100,"count":1})
            cfg["slots"]=cfg["slots"][:len(slots)]
            if cfg["type"]==2: cfg["drop_cnt"]=len(slots)
            self.manual_configs[rid]=cfg; self.manual_idx+=1
        self._finish_manual()

    def _finish_manual(self):
        csv_rows=[]
        def replace_row(m):
            row=m.group(0); rid=_get_tag(row,"Id")
            if rid not in self.manual_configs: return row
            new_row=apply_cfg_to_row(row,self.manual_configs[rid])
            drop_ids=[v for _,v in real_drop_slots(new_row)]
            name=next((n for r,n,_ in self.manual_matched if r==rid),"")
            csv_rows.append([rid,name,*drop_ids]); return new_row
        full_out=ROW_RE.sub(replace_row,self.xml_text)
        self._build_output_screen(full_out,csv_rows,len(self.manual_configs))

    def _build_output_screen(self, full_xml, csv_rows, count):
        self._clear()
        tk.Label(self,text=f"Done — {count} box(es) modified",font=("Consolas",13,"bold"),
                 bg=BG,fg=GREEN).pack(pady=12)
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=12,pady=4)
        max_items=max((len(r)-2 for r in csv_rows),default=0)
        header=["BoxID","BoxName"]+[f"Item{i+1}_ID" for i in range(max_items)]
        csv_cont="\n".join([",".join(header)]+[",".join(str(x) for x in r) for r in csv_rows])
        make_output_tab(nb,"Full PresentItemParam2.xml (modified)",full_xml,"PresentItemParam2_modified.xml",self.root)
        make_output_tab(nb,"Box Contents CSV (→ Tool 3)",csv_cont,"box_contents_for_tool3.csv",self.root)
        nb.select(0)
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        def export_all():
            folder=filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            for fname,content in [("PresentItemParam2_modified.xml",full_xml),("box_contents_for_tool3.csv",csv_cont)]:
                with open(os.path.join(folder,fname),"w",encoding="utf-8") as f: f.write(content)
            messagebox.showinfo("Export Complete",f"Saved to:\n{folder}")
        mk_btn(bot,"💾  Export All Files",export_all,color=ACC2,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=14)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 3 — NCash / Ticket Updater  (simple CSV)
# ══════════════════════════════════════════════════════════════════════════════
class Tool3(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session; self.csv_items=[]; self.xml_files=[]; self.item_lib={}
        self.mode_var=tk.StringVar(value="uniform"); self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _build_load_screen(self):
        self._clear()
        tk.Label(self,text="NCASH / TICKET UPDATER",font=("Consolas",18,"bold"),bg=BG,fg=ACC3).pack(pady=(24,4))
        tk.Label(self,text="Formula: NCash = round(tickets × 133)",bg=BG,fg=FG_DIM,font=("Consolas",10)).pack(pady=(0,8))
        csv_status=tk.StringVar(value="No file loaded")
        xml_status=tk.StringVar(value="No file loaded")

        csv_frm=mk_section(self,"Step 1 — Box Contents CSV (from Tool 2, or ID list)")
        tk.Label(csv_frm,textvariable=csv_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_csv():
            p=filedialog.askopenfilename(filetypes=[("CSV","*.csv"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig") as f: text=f.read()
            items=parse_csv_text_t3(text)
            if not items: messagebox.showerror("Error","No item IDs found in CSV."); return
            self.csv_items=items
            if self.item_lib:
                for it in self.csv_items: it["name"]=self.item_lib.get(it["id"],"")
            csv_status.set(f"✓  {os.path.basename(p)}  —  {len(items)} items")
        mk_btn(csv_frm,"📂 Load",load_csv,padx=10,pady=4).pack(side="right",padx=8,pady=6)
        def import_session_t3():
            csv = self.session.box_id_list_csv or self.session.box_contents_csv
            if not csv:
                messagebox.showinfo("No Session Data","Run Tool 1 or 2 first to generate data.")
                return
            items = parse_csv_text_t3(csv)
            if not items:
                messagebox.showerror("Error","No item IDs found in session data."); return
            self.csv_items = items
            if self.item_lib:
                for it in self.csv_items: it["name"] = self.item_lib.get(it["id"],"")
            src_label = "Tool 1" if self.session.box_id_list_csv else "Tool 2"
            csv_status.set(f"✓  Imported from {src_label}  —  {len(items)} items")
        mk_btn(csv_frm,"⬇  Import from Tool 1/2",import_session_t3,color=ACC3,fg=BG2,padx=8,pady=4).pack(side="right",padx=4,pady=6)

        xml_frm=mk_section(self,"Step 2 — ItemParam XML (pick any one of the 4 files)")
        tk.Label(xml_frm,textvariable=xml_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_xml():
            p=filedialog.askopenfilename(title="Select any one of the 4 ItemParam XML files",
                                         filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            folder=os.path.dirname(p); loaded=[]
            for fname in os.listdir(folder):
                if fname.lower() in TARGET_FILES:
                    try:
                        with open(os.path.join(folder,fname),encoding="utf-8-sig",errors="replace") as f:
                            loaded.append((fname,f.read()))
                    except: pass
            if not loaded: messagebox.showerror("Error","None of the 4 ItemParam files found."); return
            self.xml_files=loaded; self.item_lib=build_item_lib(loaded)
            for it in self.csv_items: it["name"]=self.item_lib.get(it["id"],"")
            xml_status.set(f"✓  {len(loaded)}/4 files  |  {len(self.item_lib)} items indexed")
        mk_btn(xml_frm,"📂 Load",load_xml,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        mode_frm=mk_section(self,"Step 3 — Mode")
        mf=tk.Frame(mode_frm,bg=BG); mf.pack(anchor="w",padx=10,pady=6)
        for lbl,val in [("Uniform  —  one ticket cost applied to every item","uniform"),
                        ("Manual   —  set ticket cost per item individually","manual")]:
            tk.Radiobutton(mf,text=lbl,variable=self.mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(anchor="w",pady=2)

        def proceed():
            if not self.csv_items: messagebox.showwarning("Missing","Load a CSV first."); return
            if not self.xml_files: messagebox.showwarning("Missing","Load ItemParam XML first."); return
            if self.mode_var.get()=="uniform": self._build_uniform_screen()
            else: self._build_manual_screen()
        mk_btn(self,"▶  Continue →",proceed,color=GREEN,fg=BG2,font=("Consolas",12,"bold")).pack(pady=18)

    def _build_uniform_screen(self):
        self._clear()
        tk.Label(self,text="Uniform Ticket Cost",font=("Consolas",14,"bold"),bg=BG,fg=ACC3).pack(pady=(20,4))
        tk.Label(self,text=f"Applied to all {len(self.csv_items)} items.",bg=BG,fg=FG_DIM,font=("Consolas",10)).pack(pady=(0,12))
        frm=tk.Frame(self,bg=BG); frm.pack()
        tk.Label(frm,text="Ticket Cost:",bg=BG,fg=FG,font=("Consolas",12)).pack(side="left",padx=8)
        tv=tk.StringVar()
        ent=tk.Entry(frm,textvariable=tv,width=12,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",12),relief="flat")
        ent.pack(side="left",padx=8); ent.focus()
        ncash_var=tk.StringVar(value="NCash: —")
        tk.Label(self,textvariable=ncash_var,bg=BG,fg=GREEN,font=("Consolas",12,"bold")).pack(pady=8)
        def on_change(*_):
            try: ncash_var.set(f"NCash: {round(float(tv.get())*133)}")
            except: ncash_var.set("NCash: —")
        tv.trace_add("write",on_change)
        def apply_uniform():
            try: cost=float(tv.get())
            except: messagebox.showwarning("Invalid","Enter a valid ticket cost."); return
            for it in self.csv_items: it["ticket_cost"]=cost
            self._process_and_show()
        bot=tk.Frame(self,bg=BG); bot.pack(pady=16)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=8)
        mk_btn(bot,"✓  Apply to All & Update XML",apply_uniform,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=8)

    def _build_manual_screen(self):
        self._clear()
        tk.Label(self,text="Manual Ticket Costs",font=("Consolas",14,"bold"),bg=BG,fg=ACC3).pack(pady=(12,2))
        tk.Label(self,text="Leave blank to skip an item.",bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(pady=(0,4))
        outer=tk.Frame(self,bg=BG); outer.pack(fill="both",expand=True,padx=20,pady=4)
        canvas,cont=mk_scroll_canvas(outer)
        hdr=tk.Frame(cont,bg=BG2); hdr.pack(fill="x",pady=2)
        for txt,w in [("Item ID",12),("Item Name",36),("Ticket Cost",14),("NCash (calc)",14)]:
            tk.Label(hdr,text=txt,width=w,bg=BG2,fg=BLUE,font=("Consolas",9,"bold"),anchor="w").pack(side="left",padx=6,pady=4)
        ticket_vars=[]
        for i,item in enumerate(self.csv_items):
            bg=BG if i%2==0 else BG2
            row=tk.Frame(cont,bg=bg); row.pack(fill="x")
            tk.Label(row,text=item["id"],width=12,bg=bg,fg=BG4,font=("Consolas",9),anchor="w").pack(side="left",padx=6,pady=2)
            name=item.get("name") or self.item_lib.get(item["id"],"—")
            tk.Label(row,text=name[:38],width=36,bg=bg,fg=FG_DIM,font=("Consolas",9),anchor="w").pack(side="left",padx=6,pady=2)
            tv=tk.StringVar()
            if item.get("ticket_cost") is not None: tv.set(str(item["ticket_cost"]))
            ticket_vars.append(tv)
            tk.Entry(row,textvariable=tv,width=12,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=6,pady=2)
            ncash_lbl=tk.Label(row,text="—",width=14,bg=bg,fg=GREEN,font=("Consolas",9),anchor="w")
            ncash_lbl.pack(side="left",padx=6)
            def make_trace(var,lbl):
                def cb(*_):
                    try: lbl.config(text=str(round(float(var.get())*133)))
                    except: lbl.config(text="—")
                var.trace_add("write",cb); cb()
            make_trace(tv,ncash_lbl)
        def confirm():
            blanks=[]
            for i,item in enumerate(self.csv_items):
                raw=ticket_vars[i].get().strip()
                try: item["ticket_cost"]=float(raw)
                except: item["ticket_cost"]=None; blanks.append(item["id"])
            if blanks:
                if not messagebox.askyesno("Missed a spot",f"{len(blanks)} item(s) will be SKIPPED:\n\n"+", ".join(blanks[:20])+("\n\nContinue anyway?")): return
            self._process_and_show()
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=14)
        mk_btn(bot,"✓  Apply & Update XML",confirm,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="right",padx=14)

    def _process_and_show(self):
        updates={it["id"]:round(it["ticket_cost"]*133) for it in self.csv_items if it["ticket_cost"] is not None}
        name_map={it["id"]:it.get("name","") for it in self.csv_items}
        file_results=[]
        for fname,text in self.xml_files:
            modified,found_map=bulk_update_ncash(text,updates)
            file_results.append((fname,modified,found_map))
        found_in={}
        for fname,_,found_map in file_results:
            for iid,hit in found_map.items():
                if hit and iid not in found_in: found_in[iid]=fname
        results=[]
        for item in self.csv_items:
            iid=item["id"]; name=name_map.get(iid,"")
            if item["ticket_cost"] is None: results.append((iid,name,None,None))
            else: results.append((iid,name,updates[iid],found_in.get(iid)))
        self._build_output_screen(file_results,results,updates)

    def _build_output_screen(self, file_results, results, updates):
        self._clear()
        updated_count=sum(1 for _,_,n,f in results if n is not None and f is not None)
        skipped_count=sum(1 for _,_,n,_ in results if n is None)
        missing_count=sum(1 for _,_,n,f in results if n is not None and f is None)
        tk.Label(self,text=f"✓ Updated: {updated_count}    ⚠ Not found: {missing_count}    — Skipped: {skipped_count}",
                 font=("Consolas",10,"bold"),bg=BG,fg=GREEN).pack(pady=8)
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=12,pady=4)
        exports=[]
        for fname,modified_text,found_map in file_results:
            if not any(hit for hit in found_map.values()): continue
            exports.append((fname,modified_text))
            make_output_tab(nb,os.path.splitext(fname)[0],modified_text,fname,self.root)
        col_hdr=f"{'ID':<15}{'Name':<34}{'NCash':<13}Status"; col_sep="─"*74
        log_parts=[]
        for fname,_,found_map in file_results:
            file_rows=[(iid,name,ncash,ff) for iid,name,ncash,ff in results if ff==fname]
            if not file_rows: log_parts.append(f"{fname}  →  No matching IDs — Skipped!\n"); continue
            log_parts.append(f"{fname}  →  {len(file_rows)} match(es)\n  {col_hdr}\n  {col_sep}")
            for iid,name,ncash,_ in file_rows: log_parts.append(f"  {iid:<15}{(name or '—')[:32]:<34}{ncash:<13}✓ Updated")
            log_parts.append("")
        unassigned=[(iid,name,ncash,ff) for iid,name,ncash,ff in results if ff is None]
        missing_rows=[(iid,name,ncash) for iid,name,ncash,_ in unassigned if ncash is not None]
        skipped_rows=[(iid,name) for iid,name,ncash,_ in unassigned if ncash is None]
        log_parts.append("── Unassigned / Skipped ──────────────────────────────────")
        for iid,name,ncash in missing_rows: log_parts.append(f"  {iid:<15}{(name or '—')[:32]:<34}{ncash:<13}⚠ Not found")
        for iid,name in skipped_rows: log_parts.append(f"  {iid:<15}{(name or '—')[:32]:<34}{'—':<13}SKIPPED")
        if not missing_rows and not skipped_rows: log_parts.append("  (none)")
        log_content="\n".join(log_parts)
        exports.append(("ncash_update_log.txt",log_content))
        make_output_tab(nb,"Update Log",log_content,"ncash_update_log.txt",self.root)
        nb.select(0)
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        def export_all():
            folder=filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            saved=[]
            for efname,content in exports:
                with open(os.path.join(folder,efname),"w",encoding="utf-8") as f: f.write(content)
                saved.append(efname)
            messagebox.showinfo("Export Complete",f"Saved to:\n{folder}\n\n"+"\n".join(saved))
        mk_btn(bot,"💾  Export All Files",export_all,color=ACC1,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=14)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 4 — NCash Updater  (parent-box CSV + sub-box via PresentItemParam2)
# ══════════════════════════════════════════════════════════════════════════════
class Tool4(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session
        self.parent_items=[]; self.xml_files=[]; self.item_lib={}
        self.present_text=None; self.present_enabled=tk.BooleanVar(value=False)
        self.parent_mode_var=tk.StringVar(value="uniform")
        self.sub_mode_var=tk.StringVar(value="uniform")
        self.sub_items=[]; self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _build_load_screen(self):
        self._clear()
        tk.Label(self,text="NCASH UPDATER — PARENT BOX",font=("Consolas",16,"bold"),bg=BG,fg=ACC4).pack(pady=(18,2))
        tk.Label(self,text="Formula: NCash = round(tickets × 133)",bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(pady=(0,6))
        outer=tk.Frame(self,bg=BG); outer.pack(fill="both",expand=True)
        scroll_host4=tk.Frame(outer,bg=BG); scroll_host4.pack(fill="both",expand=True)
        canvas,cont=mk_scroll_canvas(scroll_host4)
        # Session import panel
        if self.session.box_id_list_csv:
            sess_frm=mk_section(cont,"  ⬇  Session Import  ")
            def import_t4():
                items=parse_parentbox_csv(self.session.box_id_list_csv)
                if not items:
                    items=[{"id":i,"tickets":None,"ncash_direct":None,"box_ticket_cost":None,"group_idx":0,"name":n}
                           for i,n in self.session.box_id_map.items()]
                if not items: messagebox.showinfo("Empty","Session has no usable box IDs."); return
                self.parent_items=items
                messagebox.showinfo("Imported",f"Imported {len(items)} box IDs from Tool 1 session.")
            mk_btn(sess_frm,"Import box IDs from Tool 1",import_t4,color=ACC4,fg=BG2).pack(side="left",padx=10,pady=6)
            tk.Label(sess_frm,text=f"{len(parse_box_id_csv(self.session.box_id_list_csv))} IDs available",
                     bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=8)

        csv_status=tk.StringVar(value="No file loaded")
        xml_status=tk.StringVar(value="No file loaded")
        pres_status=tk.StringVar(value="Not loaded")

        s1=mk_section(cont,"  Step 1 — Parent-Box CSV  (ID, Tickets/NCash, Tickets of Box Contents)  ")
        tk.Label(s1,textvariable=csv_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_csv():
            p=filedialog.askopenfilename(filetypes=[("CSV","*.csv"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig") as f: text=f.read()
            items=parse_parentbox_csv(text)
            if not items: messagebox.showerror("Error","No valid item IDs found in CSV."); return
            self.parent_items=items
            csv_status.set(f"✓  {os.path.basename(p)}  —  {len(items)} IDs")
        mk_btn(s1,"📂 Load CSV",load_csv,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        s2=mk_section(cont,"  Step 2 — ItemParam XML  (pick any one of the 4 files)  ")
        tk.Label(s2,textvariable=xml_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_xml():
            p=filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            folder=os.path.dirname(p); loaded=[]
            for fname in os.listdir(folder):
                if fname.lower() in TARGET_FILES:
                    try:
                        with open(os.path.join(folder,fname),encoding="utf-8-sig",errors="replace") as f:
                            loaded.append((fname,f.read()))
                    except: pass
            if not loaded: messagebox.showerror("Error","None of the 4 ItemParam files found."); return
            self.xml_files=loaded; self.item_lib=build_item_lib(loaded)
            for it in self.parent_items: it["name"]=self.item_lib.get(it["id"],"")
            # Also silently scan for PresentItemParam2.xml
            ppath=os.path.join(folder,PRESENT_FILE)
            if os.path.exists(ppath):
                try:
                    with open(ppath,encoding="utf-8-sig",errors="replace") as f:
                        self.present_text=f.read()
                    pres_status.set(f"✓  {PRESENT_FILE} auto-loaded")
                except: pass
            xml_status.set(f"✓  {len(loaded)}/4 files  |  {len(self.item_lib)} items indexed")
        mk_btn(s2,"📂 Load",load_xml,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        s3=mk_section(cont,"  Step 3 — Mode  (parent-box IDs)  ")
        mf=tk.Frame(s3,bg=BG); mf.pack(anchor="w",padx=10,pady=6)
        for lbl,val in [("Uniform  —  one value per group","uniform"),("Manual  —  set per item","manual")]:
            tk.Radiobutton(mf,text=lbl,variable=self.parent_mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(anchor="w",pady=2)

        s4=mk_section(cont,"  Step 4 (Optional) — Sub-box NCash via PresentItemParam2  ")
        tk.Label(s4,textvariable=pres_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(anchor="w",padx=10,pady=(4,0))
        tk.Checkbutton(s4,text="Enable sub-box NCash update via PresentItemParam2",
                       variable=self.present_enabled,bg=BG,fg=FG,selectcolor=BG3,
                       activebackground=BG,font=("Consolas",10)).pack(anchor="w",padx=10,pady=4)
        def load_present_manual():
            p=filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig",errors="replace") as f: self.present_text=f.read()
            pres_status.set(f"✓  {os.path.basename(p)}")
        mk_btn(s4,"📂 Load PresentItemParam2.xml manually",load_present_manual,padx=10,pady=4).pack(anchor="w",padx=10,pady=(0,6))

        s5=mk_section(cont,"  Step 5 (Optional) — Sub-box Mode  ")
        sf=tk.Frame(s5,bg=BG); sf.pack(anchor="w",padx=10,pady=6)
        for lbl,val in [("Uniform  —  one value for all sub items","uniform"),("Manual  —  set per sub item","manual")]:
            tk.Radiobutton(sf,text=lbl,variable=self.sub_mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(anchor="w",pady=2)

        bot_frm=tk.Frame(cont,bg=BG); bot_frm.pack(fill="x",pady=10)
        def proceed():
            if not self.parent_items: messagebox.showwarning("Missing","Load a CSV first."); return
            if not self.xml_files: messagebox.showwarning("Missing","Load ItemParam XML first."); return
            if self.parent_mode_var.get()=="uniform": self._build_uniform_screen()
            else: self._build_manual_screen(self.parent_items,"parent",self._after_parent_configured)
        mk_btn(bot_frm,"▶  Continue →",proceed,color=GREEN,fg=BG2,font=("Consolas",12,"bold")).pack(pady=10)

    # ── Uniform screen (group-aware) ──────────────────────────────────────────
    def _build_uniform_screen(self, _saved_group_vals=None):
        self._clear()
        groups={}
        for it in self.parent_items:
            gi=it.get("group_idx",0)
            groups.setdefault(gi,[]).append(it)
        group_keys=sorted(groups.keys())
        saved_vals=_saved_group_vals or {}
        confirmed_vals={}  # gi -> {ticket_cost or ncash_direct}
        current_group_pos=[0]

        def show_group(pos):
            self._clear()
            gi=group_keys[pos]
            group_items=groups[gi]
            sample_names=[self.item_lib.get(it["id"],"") for it in group_items if self.item_lib.get(it["id"],"")][:3]
            tk.Label(self,text=f"Uniform Settings — Group {pos+1} of {len(group_keys)}",
                     font=("Consolas",14,"bold"),bg=BG,fg=ACC4).pack(pady=(18,4))
            if sample_names: tk.Label(self,text="e.g. "+", ".join(sample_names),bg=BG,fg=FG_DIM,font=("Consolas",9)).pack()

            if confirmed_vals:
                prev_frm=tk.LabelFrame(self,text="  Previously confirmed  ",bg=BG,fg=FG_GREY,
                                       font=("Consolas",9),bd=1,relief="groove")
                prev_frm.pack(fill="x",padx=24,pady=4)
                for prev_gi,pval in confirmed_vals.items():
                    pg_items=groups[prev_gi]
                    pg_sample=[self.item_lib.get(it["id"],"") for it in pg_items if self.item_lib.get(it["id"],"")][:2]
                    desc=", ".join(pg_sample) or f"Group {prev_gi+1}"
                    if pval.get("ticket_cost") is not None: disp=f"Tickets={pval['ticket_cost']} → NCash={round(pval['ticket_cost']*133)}"
                    else: disp=f"NCash={pval.get('ncash_direct','?')}"
                    tk.Label(prev_frm,text=f"  Group {list(group_keys).index(prev_gi)+1}: {desc[:40]}  →  {disp}",
                             bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(anchor="w",padx=6,pady=1)

            type_var=tk.StringVar(value="tickets")
            sample_it=group_items[0]
            if sample_it.get("ticket_cost") is not None: type_var.set("tickets"); init_val=str(sample_it["ticket_cost"])
            elif sample_it.get("ncash_direct") is not None: type_var.set("ncash"); init_val=str(sample_it["ncash_direct"])
            else: init_val=saved_vals.get(gi,{}).get("init_val","")
            val_var=tk.StringVar(value=init_val)

            inp_frm=tk.Frame(self,bg=BG); inp_frm.pack(pady=10)
            tk.Radiobutton(inp_frm,text="Tickets ×133",variable=type_var,value="tickets",bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(side="left",padx=8)
            tk.Radiobutton(inp_frm,text="NCash exact",variable=type_var,value="ncash",bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(side="left",padx=8)
            ent_frm=tk.Frame(self,bg=BG); ent_frm.pack(pady=6)
            tk.Label(ent_frm,text="Value:",bg=BG,fg=FG,font=("Consolas",11)).pack(side="left",padx=8)
            ent=tk.Entry(ent_frm,textvariable=val_var,width=14,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",12),relief="flat")
            ent.pack(side="left",padx=8); ent.focus()
            result_lbl=tk.Label(self,text="",bg=BG,fg=GREEN,font=("Consolas",11,"bold"))
            result_lbl.pack(pady=4)
            def update_result(*_):
                try:
                    if type_var.get()=="tickets": result_lbl.config(text=f"→ NCash: {round(float(val_var.get())*133)}")
                    else: result_lbl.config(text=f"→ Approx tickets: {round(float(val_var.get())/133,4)}")
                except: result_lbl.config(text="")
            val_var.trace_add("write",update_result); type_var.trace_add("write",update_result); update_result()

            def confirm_group():
                try: v=float(val_var.get())
                except: messagebox.showwarning("Invalid","Enter a valid number."); return
                if type_var.get()=="tickets": confirmed_vals[gi]={"ticket_cost":v,"ncash_direct":None}
                else: confirmed_vals[gi]={"ticket_cost":None,"ncash_direct":int(round(v))}
                if pos+1<len(group_keys): show_group(pos+1)
                else:
                    for it in self.parent_items:
                        gval=confirmed_vals.get(it.get("group_idx",0),{})
                        it["ticket_cost"]=gval.get("ticket_cost"); it["ncash_direct"]=gval.get("ncash_direct")
                    self._after_parent_configured()

            def go_back():
                if pos>0:
                    prev_gi=group_keys[pos-1]
                    confirmed_vals.pop(prev_gi,None)
                    show_group(pos-1)
                else: self._build_load_screen()

            nav=tk.Frame(self,bg=BG); nav.pack(pady=14)
            mk_btn(nav,"◀  Back",go_back).pack(side="left",padx=10)
            mk_btn(nav,"Confirm ▶" if pos<len(group_keys)-1 else "✓  Apply All",confirm_group,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=10)

        show_group(0)

    # ── Manual screen (parent or sub) ─────────────────────────────────────────
    def _build_manual_screen(self, items, label, on_done):
        self._clear()
        tk.Label(self,text=f"Manual — {label.title()} Item Costs",font=("Consolas",14,"bold"),bg=BG,fg=ACC4).pack(pady=(12,2))
        tk.Label(self,text="Leave blank to skip.",bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(pady=(0,4))
        outer=tk.Frame(self,bg=BG); outer.pack(fill="both",expand=True,padx=20,pady=4)
        canvas,cont=mk_scroll_canvas(outer)
        hdr=tk.Frame(cont,bg=BG2); hdr.pack(fill="x",pady=2)
        for txt,w in [("Item ID",12),("Item Name",34),("Type",12),("Value",14),("NCash →",14)]:
            tk.Label(hdr,text=txt,width=w,bg=BG2,fg=BLUE,font=("Consolas",9,"bold"),anchor="w").pack(side="left",padx=4,pady=4)
        type_vars=[]; val_vars=[]
        for i,item in enumerate(items):
            bg=BG if i%2==0 else BG2
            row=tk.Frame(cont,bg=bg); row.pack(fill="x")
            tk.Label(row,text=item["id"],width=12,bg=bg,fg=BG4,font=("Consolas",9),anchor="w").pack(side="left",padx=4,pady=2)
            name=item.get("name") or self.item_lib.get(item["id"],"—")
            tk.Label(row,text=name[:36],width=34,bg=bg,fg=FG_DIM,font=("Consolas",9),anchor="w").pack(side="left",padx=4,pady=2)
            if item.get("ticket_cost") is not None: init_type,init_val="tickets",str(item["ticket_cost"])
            elif item.get("ncash_direct") is not None: init_type,init_val="ncash",str(item["ncash_direct"])
            else: init_type,init_val="tickets",""
            tv=tk.StringVar(value=init_type); type_vars.append(tv)
            type_combo=ttk.Combobox(row,textvariable=tv,values=["tickets","ncash"],state="readonly",width=8,font=("Consolas",9))
            type_combo.pack(side="left",padx=4,pady=2)
            vv=tk.StringVar(value=init_val); val_vars.append(vv)
            tk.Entry(row,textvariable=vv,width=12,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=4,pady=2)
            ncash_lbl=tk.Label(row,text="—",width=14,bg=bg,fg=GREEN,font=("Consolas",9),anchor="w")
            ncash_lbl.pack(side="left",padx=4)
            def make_trace(tv2,vv2,lbl):
                def cb(*_):
                    try:
                        v=float(vv2.get())
                        if tv2.get()=="tickets": lbl.config(text=str(round(v*133)))
                        else: lbl.config(text=str(int(round(v))))
                    except: lbl.config(text="—")
                tv2.trace_add("write",cb); vv2.trace_add("write",cb); cb()
            make_trace(tv,vv,ncash_lbl)

        def confirm():
            blanks=[]
            for i,item in enumerate(items):
                raw=val_vars[i].get().strip()
                try:
                    v=float(raw)
                    if type_vars[i].get()=="tickets": item["ticket_cost"]=v; item["ncash_direct"]=None
                    else: item["ncash_direct"]=int(round(v)); item["ticket_cost"]=None
                except: item["ticket_cost"]=None; item["ncash_direct"]=None; blanks.append(item["id"])
            if blanks:
                if not messagebox.askyesno("Missed a spot",f"{len(blanks)} item(s) will be SKIPPED. Continue anyway?"): return
            on_done()

        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=14)
        mk_btn(bot,"✓  Apply",confirm,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="right",padx=14)

    def _after_parent_configured(self):
        if self.present_enabled.get() and self.present_text:
            self._build_sub_configure_screen()
        else:
            self._process_and_show()

    def _build_sub_configure_screen(self):
        box_ids={it["id"] for it in self.parent_items}
        drop_map=extract_drop_ids_from_present(self.present_text, box_ids)
        self.sub_items=[]
        seen=set()
        for it in self.parent_items:
            for drop_id in drop_map.get(it["id"],[]):
                if drop_id not in seen:
                    seen.add(drop_id)
                    tc=it.get("box_ticket_cost")
                    self.sub_items.append({"id":drop_id,"name":self.item_lib.get(drop_id,""),
                                           "ticket_cost":tc,"ncash_direct":None,"box_ticket_cost":tc,"group_idx":0})
        if not self.sub_items:
            messagebox.showinfo("No sub-items","No DropId entries found for the parent box IDs in PresentItemParam2.")
            self._process_and_show(); return
        all_prefilled=all(it.get("ticket_cost") is not None or it.get("ncash_direct") is not None for it in self.sub_items)
        if all_prefilled:
            self._process_and_show(); return
        if self.sub_mode_var.get()=="uniform":
            self._build_sub_uniform_screen()
        else:
            self._build_manual_screen(self.sub_items,"sub-box",self._process_and_show)

    def _build_sub_uniform_screen(self):
        self._clear()
        tk.Label(self,text="Uniform Sub-Box Tickets",font=("Consolas",14,"bold"),bg=BG,fg=ACC4).pack(pady=(20,4))
        tk.Label(self,text=f"Applied to all {len(self.sub_items)} sub-box drop IDs.",bg=BG,fg=FG_DIM,font=("Consolas",10)).pack(pady=(0,12))
        sample_it=next((it for it in self.sub_items if it.get("ticket_cost") is not None),None)
        init_val=str(sample_it["ticket_cost"]) if sample_it else ""
        frm=tk.Frame(self,bg=BG); frm.pack()
        tv=tk.StringVar(value=init_val)
        tk.Label(frm,text="Ticket Cost:",bg=BG,fg=FG,font=("Consolas",12)).pack(side="left",padx=8)
        ent=tk.Entry(frm,textvariable=tv,width=14,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",12),relief="flat")
        ent.pack(side="left",padx=8); ent.focus()
        ncash_lbl=tk.Label(self,text="",bg=BG,fg=GREEN,font=("Consolas",12,"bold")); ncash_lbl.pack(pady=6)
        def on_change(*_):
            try: ncash_lbl.config(text=f"→ NCash: {round(float(tv.get())*133)}")
            except: ncash_lbl.config(text="")
        tv.trace_add("write",on_change)
        def apply():
            try: cost=float(tv.get())
            except: messagebox.showwarning("Invalid","Enter a valid ticket cost."); return
            for it in self.sub_items: it["ticket_cost"]=cost; it["ncash_direct"]=None
            self._process_and_show()
        bot=tk.Frame(self,bg=BG); bot.pack(pady=16)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=8)
        mk_btn(bot,"✓  Apply & Update XML",apply,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=8)

    def _resolve_ncash(self, item):
        if item.get("ticket_cost") is not None: return round(item["ticket_cost"]*133)
        if item.get("ncash_direct") is not None: return item["ncash_direct"]
        return None

    def _process_and_show(self):
        all_items=list(self.parent_items)+list(self.sub_items)
        updates={}
        for it in all_items:
            n=self._resolve_ncash(it)
            if n is not None: updates[it["id"]]=n
        file_results=[]
        for fname,text in self.xml_files:
            modified,found_map=bulk_update_ncash(text,updates)
            file_results.append((fname,modified,found_map))
        found_in={}
        for fname,_,found_map in file_results:
            for iid,hit in found_map.items():
                if hit and iid not in found_in: found_in[iid]=fname
        parent_ids={it["id"] for it in self.parent_items}
        sub_ids={it["id"] for it in self.sub_items}
        updated_p=sum(1 for iid in parent_ids if found_in.get(iid))
        updated_s=sum(1 for iid in sub_ids if found_in.get(iid))
        missing=sum(1 for iid,ncash in updates.items() if not found_in.get(iid))
        skipped=sum(1 for it in all_items if self._resolve_ncash(it) is None)
        self._build_output_screen(file_results,updates,found_in,parent_ids,sub_ids,updated_p,updated_s,missing,skipped)

    def _build_output_screen(self, file_results, updates, found_in, parent_ids, sub_ids,
                              updated_p, updated_s, missing, skipped):
        self._clear()
        total_updated=updated_p+updated_s
        tk.Label(self,text=f"✓ Updated: {total_updated}  (parent: {updated_p}, sub-box: {updated_s})   ⚠ Not found: {missing}   — Skipped: {skipped}",
                 font=("Consolas",10,"bold"),bg=BG,fg=GREEN).pack(pady=8)
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=12,pady=4)
        exports=[]
        for fname,modified_text,found_map in file_results:
            if not any(hit for hit in found_map.values()): continue
            exports.append((fname,modified_text))
            make_output_tab(nb,os.path.splitext(fname)[0],modified_text,fname,self.root)
        log_parts=[]
        for fname,_,found_map in file_results:
            hits=[(iid,updates[iid],found_in.get(iid)=="parent" or iid in parent_ids) for iid,hit in found_map.items() if hit]
            if not hits: log_parts.append(f"{fname}  →  No matches"); continue
            log_parts.append(f"{fname}  →  {len(hits)} update(s)")
            for iid,ncash,is_parent in hits:
                label="parent" if iid in parent_ids else "sub-drop"
                name=self.item_lib.get(iid,"—")
                log_parts.append(f"  [{label}]  {iid:<12}  {name[:30]:<30}  NCash={ncash}")
            log_parts.append("")
        log_content="\n".join(log_parts)
        exports.append(("ncash_update_log.txt",log_content))
        make_output_tab(nb,"Update Log",log_content,"ncash_update_log.txt",self.root)
        nb.select(0)
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        def export_all():
            folder=filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            for efname,content in exports:
                with open(os.path.join(folder,efname),"w",encoding="utf-8") as f: f.write(content)
            messagebox.showinfo("Export Complete",f"Saved to:\n{folder}")
        mk_btn(bot,"💾  Export All Files",export_all,color=ACC4,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=14)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 5 — NCash ↔ Ticket Calculator
# ══════════════════════════════════════════════════════════════════════════════
class Tool5(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session; self._build()

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        tk.Label(self,text="NCash  ↔  Ticket  Calculator",font=("Consolas",16,"bold"),
                 bg=BG,fg=ACC5).pack(pady=(30,4))
        tk.Label(self,text="Formula:  NCash = round( Tickets × 133 )",
                 font=("Consolas",9),bg=BG,fg=FG_GREY).pack(pady=(0,18))

        box_a=tk.LabelFrame(self,text="  Tickets  →  NCash  ",bg=BG,fg=BLUE,
                             font=("Consolas",10,"bold"),bd=1,relief="groove")
        box_a.pack(fill="x",padx=40,pady=8)
        ra=tk.Frame(box_a,bg=BG); ra.pack(padx=14,pady=12)
        tk.Label(ra,text="Tickets:",width=10,anchor="w",font=("Consolas",11),bg=BG,fg=FG).pack(side="left")
        v_tickets=tk.StringVar()
        tk.Entry(ra,textvariable=v_tickets,width=14,bg=BG3,fg=FG,insertbackground=FG,
                 font=("Consolas",13),relief="flat").pack(side="left",padx=8)
        tk.Label(ra,text="=",font=("Consolas",13),bg=BG,fg=FG_GREY).pack(side="left",padx=4)
        lbl_ncash=tk.Label(ra,text="—",width=14,anchor="w",font=("Consolas",13,"bold"),bg=BG,fg=GREEN)
        lbl_ncash.pack(side="left",padx=4)
        tk.Label(ra,text="NCash",font=("Consolas",10),bg=BG,fg=FG_GREY).pack(side="left")

        box_b=tk.LabelFrame(self,text="  NCash  →  Tickets  ",bg=BG,fg=BLUE,
                             font=("Consolas",10,"bold"),bd=1,relief="groove")
        box_b.pack(fill="x",padx=40,pady=8)
        rb=tk.Frame(box_b,bg=BG); rb.pack(padx=14,pady=12)
        tk.Label(rb,text="NCash:",width=10,anchor="w",font=("Consolas",11),bg=BG,fg=FG).pack(side="left")
        v_ncash=tk.StringVar()
        tk.Entry(rb,textvariable=v_ncash,width=14,bg=BG3,fg=FG,insertbackground=FG,
                 font=("Consolas",13),relief="flat").pack(side="left",padx=8)
        tk.Label(rb,text="=",font=("Consolas",13),bg=BG,fg=FG_GREY).pack(side="left",padx=4)
        lbl_tickets=tk.Label(rb,text="—",width=14,anchor="w",font=("Consolas",13,"bold"),bg=BG,fg=ACC5)
        lbl_tickets.pack(side="left",padx=4)
        tk.Label(rb,text="Tickets",font=("Consolas",10),bg=BG,fg=FG_GREY).pack(side="left")

        def calc_ncash(*_):
            try: lbl_ncash.config(text=f"{round(float(v_tickets.get())*133):,}")
            except: lbl_ncash.config(text="—")
        def calc_tickets(*_):
            try:
                raw=float(v_ncash.get())/133
                lbl_tickets.config(text=f"{int(raw):,}" if raw==int(raw) else f"{round(raw,4):,}")
            except: lbl_tickets.config(text="—")
        v_tickets.trace_add("write",calc_ncash)
        v_ncash.trace_add("write",calc_tickets)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 6 — ItemParam Generator
# ══════════════════════════════════════════════════════════════════════════════

# ── Reference tables ──────────────────────────────────────────────────────────
_CLASS_MAP  = [(0,"Unselected"),(1,"Item"),(2,"Equipment"),(3,"Drill")]

_TYPE_MAP = [
    (0,"Unselected"),(1,"Galder Coupons"),(2,"Sword Equip"),(3,"Hat Equip"),
    (4,"Inner Equip"),(5,"Fashion (Shoes)"),(6,"Drill"),(7,"Accessory / Sprint"),
    (8,"Shield Equip"),(9,"Unused"),(10,"Potions"),(11,"Cards"),
    (12,"Animated Jewels"),(13,"Pets"),(14,"Compound Etc."),(15,"Useables (Boxes/Ports)"),
    (16,"Throwing Items"),(17,"Ears"),(18,"Tails"),(19,"Fashion (Top/Jacket)"),
    (20,"Fashion (Bottom/Pants)"),(21,"Fashion (Torso/Shoulder)"),(22,"Fashion (Belt)"),
    (23,"Fashion (Accessory)"),(24,"Head Accessory"),(25,"Face Accessory"),
    (26,"Fashion (Hand/Glove)"),(27,"Fashion (Socks)"),(28,"Skill Card"),
    (29,"Ammo"),(30,"Galder (4 stacks)"),(34,"Cape"),(35,"Fortune Card"),
    (999,"Internal (unused)"),
]

_SUBTYPE_MAP = [
    (0,"N/A"),(2,"Sword"),(3,"Hat"),(4,"Innerwear"),(5,"Fashion (Shoe)"),
    (6,"Drill"),(7,"Accessory"),(8,"Shield"),(10,"HP Recovery"),
    (11,"MP Recovery"),(12,"HP&MP Recovery"),(14,"Fashion Set"),
    (18,"Buffalo/Bunny Card"),(19,"Raccoon/Cat Card"),(20,"Lion/Fox Card"),
    (21,"Dragon/Sheep Card"),(22,"NPC Card"),(23,"Power Card"),(24,"Charm Card"),
    (25,"Sense Card"),(26,"Magic Card"),(27,"Old Monster2 Card"),(33,"Neutral Card"),
    (36,"Protein/Lock Candy"),(39,"Throwing Item"),(41,"Ears"),(42,"Tail"),
    (43,"Fashion (Top/Jacket)"),(44,"Fashion (Bottom/Pants)"),
    (45,"Fashion (Torso/Shoulder)"),(46,"Fashion (Belt)"),(47,"Fashion (Accessory)"),
    (48,"Head Accessory"),(49,"Face Accessory"),(50,"Fashion (Hand/Glove)"),
    (51,"Fashion (Socks)"),(52,"GM Amulet"),(53,"Sprint"),(54,"Gun"),
    (55,"Skill Card"),(56,"Ammo"),(61,"MyShop Drill"),(62,"Bracer Accessory"),
    (63,"Repair Powder"),(64,"Galders (Tiny Stack)"),(65,"Galders (Small Stack)"),
    (66,"Galders (Medium Stack)"),(67,"Galders (Large Stack)"),(69,"Scrolls"),
    (70,"Fortune Card"),(71,"Disguise Kit"),(72,"Poseidon/Spinel Seed"),
    (73,"Secret Card"),(74,"Soul Gauge"),(75,"Soul Ticket"),(76,"Soul Feather Pen"),
    (77,"Guardian Runes"),(78,"Custom Character Pets"),(81,"Pet Item Hunt"),
]

_ITEMFTYPE_MAP = [
    (0,"N/A"),(1,"Power"),(2,"Magic"),(3,"Sense"),(4,"Charm"),(5,"Neutral"),
]

_EFFECT_MAP = [
    (0,"N/A"),(11,"Recovery Item"),(12,"Beta Heal (unused)"),(13,"Beta Heal (unused)"),
    (14,"Portable Port"),(15,"Hair Dye"),(16,"Equip Ammo"),(19,"Emergency Port"),
    (20,"Portable Port AD"),(21,"Befana's Unstable Elixir"),(22,"Open Box"),
    (23,"Master's Authority"),(24,"Point Back 40"),(25,"Build Graph Reset"),
    (26,"Point Back All"),(27,"Store More Permit"),(28,"Teleport/Transfer Map"),
    (29,"Resurrect/Revive"),(30,"Recharge Coupon"),(31,"Reidentifier 2000"),
    (33,"Mic"),(34,"Memory Port"),(35,"Read Book/Letter"),(36,"Artisan's Flame"),
    (37,"Star Tear"),(38,"Allen Bottle"),(39,"Friend Finder"),
    (40,"Erin's Secret Book"),(41,"Voice/Auto Loot Pet"),(42,"Poseidon Seed"),
    (43,"Worn Treasure Map"),(44,"Weird Treasure Map"),(45,"Soul Charge"),
    (46,"Fiesta Ticket"),(47,"Character Boxes"),(48,"Earring"),(49,"Locked Boxes"),
    (50,"Key"),(51,"EXP/TM Booster"),(52,"Seals"),
]

_OPTIONS_FULL = [
    (1,    "Rare"),
    (2,    "Usable"),
    (4,    "Select Target Object"),
    (8,    "Select Target Position"),
    (16,   "Equipable"),
    (32,   "Usable To Self"),
    (64,   "Usable To Player"),
    (128,  "Usable To Monster"),
    (256,  "Not Buyable"),
    (512,  "Not Sellable"),
    (1024, "Not Exchangeable"),
    (2048, "Not Pickable"),
    (4096, "Not Droppable"),
    (8192, "Not Vanishable"),
    (16384,"Equipable To ChrType"),
    (32768,"Custom Data"),
    (65536,"Not Addable To Warehouse"),
    (131072,"Not Addable To NeoWarehouse"),
    (262144,"CM (MyShop)"),
    (524288,"Auto Drill"),
    (1048576,"Absolute Unique"),
    (2097152,"Legend (Boss EQ)"),
    (4194304,"Quest"),
    (8388608,"Not Composable"),
    (16777216,"Not Reform"),
    (33554432,"Skin"),
    (67108864,"Event"),
    (134217728,"Dummy (Exchange)"),
    (268435456,"Use Safe Zone"),
]

_OPTIONSEX_MAP = [
    (0,"N/A"),(1,"Apply random stat range"),(2,"Apply elemental property"),
    (16,"Boss equipment set"),(32,"Unknown"),(64,"Non-consumed teleport"),
    (512,"Hair color reset"),(1024,"GM Drill"),(2048,"Unknown (Soul/Tartarus)"),
    (8192,"Unknown (Disabled Swords)"),(16384,"Soul Weapon"),
]

_REFINEINDEX_MAP = [
    (0,"N/A"),(1,"Form/Skin Equipment"),(2,"Fashion"),(21,"Unknown"),
    (22,"Event/Exchange Equipment"),(23,"In-game Obtainable Pets"),
    (37,"Character-specific EQ"),(38,"Various Cash/Gacha Equipment"),
    (39,"Egg Shop Equipment"),(47,"Cash Item Equipment"),
    (48,"Sylvia/Kooh Equipment"),(56,"Cash Pet"),
    (64,"Character/Boss/Voice Pet"),
]

_REFINETYPE_MAP = [
    (0,"N/A"),(1,"AP Sword/Hammer/Lance"),(2,"DP Hat/Shield"),(3,"MA Staff/Cane"),
    (4,"MD Hat/Shield"),(5,"Gun AP"),(6,"DA (Goddess Circlet)"),(7,"HV (Goddess Shield)"),
]

_MINSTATTYPE_MAP = [(0,"Required AP"),(4,"Required MA"),(7,"Required DA")]

# ── Tooltip text for every field ──────────────────────────────────────────────
_TOOLTIPS = {
    # ── Identity ──────────────────────────────────────────────────────────
    "ID":           "Item ID — must be unique in the XML table. REQUIRED.",
    "Class":        "Item class.  1=Item  2=Equipment  3=Drill\n"
                    "⚠ REQUIRED — cannot be 0 or Unselected.",
    "Type":         "Item type category.\n"
                    "1=Galder Coupons  2=Sword  3=Hat  4=Inner  5=Shoes  6=Drill\n"
                    "7=Accessory/Sprint  8=Shield  10=Potions  11=Cards  12=Jewels\n"
                    "13=Pets  14=Compound  15=Useables/Boxes  16=Throwing  17=Ears\n"
                    "18=Tails  19=Top  20=Bottom  21=Torso  22=Belt  23=Fashion Acc\n"
                    "24=Head Acc  25=Face Acc  26=Glove  27=Socks  28=Skill Card\n"
                    "29=Ammo  30=Galder(4 stacks)  34=Cape  35=Fortune Card\n"
                    "⚠ REQUIRED — cannot be 0 or Unselected.",
    "SubType":      "Sub-category of the item type.  0=N/A (default).\n"
                    "Hover for full table or see 📖 Field Reference.",
    "ItemFType":    "Stat affinity.  0=N/A  1=Power  2=Magic  3=Sense  4=Charm  5=Neutral.",
    # ── Names / Text ──────────────────────────────────────────────────────
    "Name":         "The name displayed in-game. CDATA wrapper added automatically.",
    "Comment":      "The item description shown in-game. CDATA wrapper added automatically.",
    "Use":          "The purpose of the item shown in the in-game tooltip.",
    "Name_Eng":     "Dev/localization notes. Rarely visible in-game. Usually a single space.",
    "Comment_Eng":  "May contain dev references or 'description coming soon'. Usually a single space.",
    # ── Files ─────────────────────────────────────────────────────────────
    "FileName":     "Path to the item sprite NRI file.\n"
                    "Example: data\\item\\itm000.nri\n"
                    "This is copied automatically to InvFileName.",
    "BundleNum":    "Sprite index within the NRI file. Starts at 0.\n"
                    "To find your sprite: open NRI in viewer → Animations tab → slot = BundleNum+1.\n"
                    "This is copied automatically to InvBundleNum.",
    "InvFileName":  "Identical to FileName — auto-copied, no manual entry needed.",
    "InvBundleNum": "Identical to BundleNum — auto-copied, no manual entry needed.",
    "CmtFileName":  "Path to the item illustration NRI file (shown in item tooltip/description window).\n"
                    "Example: data\\item\\itm_illu000.nri  (separate from FileName).",
    "CmtBundleNum": "Sprite index within CmtFileName NRI. Same +1 offset rule as BundleNum.",
    "EquipFileName":"Path to the equipment or drill model. Leave as a single space if not equipment/drill.",
    # ── PivotID / Palette ─────────────────────────────────────────────────
    "PivotID":      "Source item ID reference. Used for equipment with multiple levels or option variants.\n"
                    "Suggested 0 for most items.",
    "PaletteId":    "Palette ID. Leave at 0 for most items. Some exceptions exist but can be ignored.",
    # ── Options ───────────────────────────────────────────────────────────
    "Options":      "Item option flags (eItemOption).  Values OR together.\n"
                    "Key flags: 2=Usable  16=Equipable  32=UsableToSelf  256=NotBuyable\n"
                    "512=NotSellable  262144=CM(MyShop)  268435456=UseSafeZone(Skins)\n"
                    "See 📖 Field Reference for full table.",
    # ── HideHat ───────────────────────────────────────────────────────────
    "HideHat":      "Per-character flag that hides the character's ear model when this item is equipped.\n"
                    "Uses the same flag values as ChrTypeFlags (1st/2nd/3rd job per race).\n"
                    "Suggested 0 unless specifically needed.",
    # ── ChrTypeFlags ──────────────────────────────────────────────────────
    "ChrTypeFlags": "Character-type access flags.  Sum flags for each character/job that should be allowed.\n"
                    "Bunny 1st=1  Buffalo 1st=2  Sheep 1st=4  Dragon 1st=8  Fox 1st=16\n"
                    "Lion 1st=32  Cat 1st=64  Raccoon 1st=124  Paula 1st=256\n"
                    "2nd job: multiply base by 512.  3rd job: see reference table.\n"
                    "0 = no restriction (all characters allowed).  Suggested 0.",
    # ── Ground / System ───────────────────────────────────────────────────
    "GroundFlags":  "Always 0 in standard items. Non-zero may cause unintended behaviour. SUGGESTED 0.",
    "SystemFlags":  "Always 0 in standard items. Non-zero may cause unintended behaviour. SUGGESTED 0.",
    # ── OptionsEx ─────────────────────────────────────────────────────────
    "OptionsEx":    "Extended item options. Values OR together.\n"
                    "1=Random stat range  2=Elemental property  16=Boss EQ set\n"
                    "64=Non-consumed teleport  512=Hair color reset  1024=GM Drill\n"
                    "2048=Unknown(Soul/Tartarus)  8192=Unknown(Disabled Swords)  16384=Soul Weapon\n"
                    "Mainly used on equipment in ItemParamCM2. Suggested 0 for standard items.",
    # ── Numeric stats ─────────────────────────────────────────────────────
    "Weight":       "Item weight value (WT stat). Affects carry capacity. Default 1.",
    "Value":        "Galder (in-game gold) retail price. 0 = cannot be sold to NPC.",
    "MinLevel":     "Minimum character level required to use or equip the item. Default 1.",
    # ── Effect ────────────────────────────────────────────────────────────
    "Effect":       "Action triggered when the item is used.\n"
                    "22=Open Box (REQUIRED for boxes/Type 15)  14=Portable Port  15=Hair Dye\n"
                    "19=Emergency Port  22=Open Box  29=Resurrect  47=Character Boxes\n"
                    "49=Locked Boxes  50=Key  51=EXP/TM Booster\n"
                    "Multiple effects can be OR'd: enter slash-separated e.g. 22/47\n"
                    "Suggested 0 for non-use items.",
    "EffectFlags2": "Always 0. SUGGESTED 0.",
    "SelRange":     "Leave at 0 except for Beta Magic Cards (no longer used). SUGGESTED 0.",
    "Life":         "Duration for timed items (EXP/TM Boosters) or drill life span.\n"
                    "0 = no time limit.",
    "Depth":        "Used only for the three test drills in the table. SUGGESTED 0.",
    "Delay":        "Tied to Beta Magic Cards (no longer used). Leave at 0.000000. SUGGESTED 0.",
    # ── Stat bonuses ─────────────────────────────────────────────────────
    "AP":           "Gun / Throwing Attack Power. NOT the regular AP stat (that is APPlus). Default 0.",
    "HP":           "Amount of HP recovered on use (potions, recovery items). Default 0.",
    "HPCon":        "Amount of HP consumed on use. Rarely used. Default 0.",
    "MP":           "Amount of MP recovered on use. Default 0.",
    "MPCon":        "Amount of MP consumed on use. Also used for Card MP values. Default 0.",
    "Money":        "Amount of Galder obtained from Galder Coupons. Default 0.",
    "APPlus":       "Attack Power stat bonus (equipment). Not to be confused with AP field above.",
    "ACPlus":       "Accuracy stat bonus.",
    "DXPlus":       "Dexterity stat bonus.",
    "MaxMPPlus":    "Maximum Magic Point capacity bonus.",
    "MAPlus":       "Magic Attack stat bonus.",
    "MDPlus":       "Magic Defense stat bonus.",
    "MaxWTPlus":    "Maximum Weight Capacity bonus.",
    "DAPlus":       "Detect Ability stat bonus.",
    "LKPlus":       "Luck stat bonus.",
    "MaxHPPlus":    "Maximum Health Point capacity bonus.",
    "DPPlus":       "Defense Power stat bonus.",
    "HVPlus":       "Evasion stat bonus.",
    "HPRecoveryRate":"HP regen rate (used on pets). Format: 0.000000  Default: 0.000000",
    "MPRecoveryRate":"MP regen rate (used on pets). Format: 0.000000  Default: 0.000000",
    # ── Card params ───────────────────────────────────────────────────────
    "CardNum":      "Card rank. Controls life span and grade range.\n"
                    "0=None/dummy  1=Boss(life 5-6)  2-3=High tier(4-5)  4-6=Mid(4)\n"
                    "7-9=Low/NPC(3)  |  Skill cards: CardNum=7, CardGenGrade=0, CardGenParam=0\n"
                    "SUGGESTED 0 for non-card items.",
    "CardGenGrade": "Determines which grade ranges are available for this card.\n"
                    "SUGGESTED 0 for non-card items.",
    "CardGenParam": "Unsure. Skill cards always use 0.000000. Format: 0.000000\n"
                    "SUGGESTED 0.000000 for non-card items.",
    "DailyGenCnt":  "Possibly a daily card-battle generation limit. SUGGESTED 0.",
    # ── Other ─────────────────────────────────────────────────────────────
    "PartFileName": "ItemParamCM2 only — path to the fashion item model. Usually a single space.",
    "ChrFTypeFlag": "Always 0. SUGGESTED 0.",
    "ChrGender":    "Always 0. SUGGESTED 0.",
    "ExistType":    "Items that cannot stack simultaneously on a character (sprints, boosters, etc).\n"
                    "0=disabled  1=timer / cannot stack.  Actual time values are WIP.",
    "Ncash":        "Cash shop (MyShop) price tag. 0 = not sold in cash shop.",
    "NewCM":        "Always 0. SUGGESTED 0.",
    "FamCM":        "Always 0. SUGGESTED 0.",
    "Summary":      "Always blank (single space). SUGGESTED blank.",
    "ShopFileName": "Path to promotional Cash Mall item image NRI. Usually a single space.",
    "ShopBundleNum":"Animation sprite index for ShopFileName. Same +1 offset rule as BundleNum. Suggested 0.",
    "MinStatType":  "Stat type required to equip.  0=Required AP  4=Required MA  7=Required DA.\n"
                    "Set MinStatLv=0 to show no requirement.",
    "MinStatLv":    "Minimum stat value required to equip. 0 = no stat requirement shown.\n"
                    "Note: this is the culprit behind no equip-swapping during Rust debuff.",
    "RefineIndex":  "References RefineLevelTable. See 📖 Field Reference for values.\n"
                    "Common: 1=Form/Skin  2=Fashion  22=Event EQ  23=Obtainable Pet\n"
                    "38=Cash/Gacha EQ  47=Cash Item  56=Cash Pet  64=Character/Boss/Voice Pet",
    "RefineType":   "Equipment type for refining.  0=N/A  1=AP Sword  2=DP Hat/Shield\n"
                    "3=MA Staff  4=MD Hat/Shield  5=Gun AP  6=DA Circlet  7=HV Shield",
    "CompoundSlot": "Number of compound slots (0-5). Intended range: 0-5.\n"
                    "Higher values will cause UI errors (EE) when the item is inspected.",
    "SetItemID":    "Equipment set ID reference. Used for set bonuses. Suggested 0.",
    "ReformCount":  "Reform count — possibly tied to skins. Not fully researched. Suggested 0.",
    "GroupId":      "Group ID. Always 0 in standard items. SUGGESTED 0.",
    "Money":        "Amount of Galder from coupons. Default 0.",
}

# ── Tooltip widget helper ─────────────────────────────────────────────────────
def _attach_tooltip(widget, text):
    tip = None
    def show(e):
        nonlocal tip
        if tip: return
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + widget.winfo_height() + 4
        tip = tk.Toplevel(widget)
        tip.wm_overrideredirect(True)
        tip.wm_geometry(f"+{x}+{y}")
        tk.Label(tip, text=text, justify="left", bg="#313244", fg="#cdd6f4",
                 font=("Consolas", 8), relief="flat", bd=1, wraplength=380,
                 padx=6, pady=4).pack()
    def hide(e):
        nonlocal tip
        if tip: tip.destroy(); tip = None
    widget.bind("<Enter>", show, add="+")
    widget.bind("<Leave>", hide, add="+")

# ── Persistent settings helper ───────────────────────────────────────────────

# ── Per-tool persistent settings (separate files per tool) ──────────────────
_SETTINGS_DIR = os.path.expanduser("~")

def _settings_path(tool_key):
    return os.path.join(_SETTINGS_DIR, f".box_tool_suite_{tool_key}.json")

def _load_settings(tool_key):
    try:
        with open(_settings_path(tool_key), encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}

def _save_settings(tool_key, data):
    try:
        with open(_settings_path(tool_key), "w", encoding="utf-8") as f:
            _json.dump(data, f, indent=2)
    except Exception:
        pass

# Backwards-compat aliases for Tool6
def _load_t6_settings():    return _load_settings("t6")
def _save_t6_settings(d):   _save_settings("t6", d)

def _get_last_id(tool_key, default=1):
    """Return the last-used ID for a given tool."""
    return _load_settings(tool_key).get("last_id", default)

def _set_last_id(tool_key, id_val):
    d = _load_settings(tool_key)
    d["last_id"] = id_val
    _save_settings(tool_key, d)

# ── XML builder for generic ItemParam ────────────────────────────────────────
def _parse_effect_val(raw):
    """Parse Effect field — may be int, '0', '22', '22/47', etc."""
    if isinstance(raw, int): return raw
    raw = str(raw).strip()
    parts = [p.strip() for p in raw.replace("/",",").split(",") if p.strip()]
    vals = []
    for p in parts:
        try: vals.append(int(p))
        except: pass
    if not vals: return 0
    # Single value -> return int; multiple -> slash-separated string
    if len(vals) == 1: return vals[0]
    return "/".join(str(v) for v in vals)

def build_generic_itemparam_row(cfg):
    """Build a full ItemParam <ROW> from a cfg dict."""
    def _cd(v):  return f"<![CDATA[{v}]]>"
    def _fmt6(v):
        try:    return f"{float(v):.6f}"
        except: return "0.000000"
    def _int(v, default=0):
        try:    return int(v)
        except: return default

    # Options bitmask from list of checked flags
    opts_val = 0
    for flag, checked in zip([f for f,_ in _OPTIONS_FULL], cfg.get("options_flags", [])):
        if checked: opts_val |= flag
    # OptionsEx — sum of selected values
    optex_val = cfg.get("options_ex", 0)

    lines = [
        "<ROW>",
        f"<ID>{cfg['id']}</ID>",
        f"<Class>{cfg.get('class_val','1')}</Class>",
        f"<Type>{cfg.get('type_val','15')}</Type>",
        f"<SubType>{cfg.get('subtype_val','0')}</SubType>",
        f"<ItemFType>{cfg.get('itemftype_val','0')}</ItemFType>",
        f"<n>{_cd(cfg.get('name',''))}</n>",
        f"<Comment>{_cd(cfg.get('comment',''))}</Comment>",
        f"<Use>{_cd(cfg.get('use',''))}</Use>",
        f"<Name_Eng>{_cd(cfg.get('name_eng',' '))}</Name_Eng>",
        f"<Comment_Eng>{_cd(cfg.get('comment_eng',' '))}</Comment_Eng>",
        f"<FileName>{_cd(cfg.get('file_name',''))}</FileName>",
        f"<BundleNum>{_int(cfg.get('bundle_num',0))}</BundleNum>",
        f"<InvFileName>{_cd(cfg.get('file_name',''))}</InvFileName>",
        f"<InvBundleNum>{_int(cfg.get('bundle_num',0))}</InvBundleNum>",
        f"<CmtFileName>{_cd(cfg.get('cmt_file_name',''))}</CmtFileName>",
        f"<CmtBundleNum>{_int(cfg.get('cmt_bundle_num',0))}</CmtBundleNum>",
        f"<EquipFileName>{_cd(cfg.get('equip_file_name',' '))}</EquipFileName>",
        f"<PivotID>{_int(cfg.get('pivot_id',0))}</PivotID>",
        f"<PaletteId>{_int(cfg.get('palette_id',0))}</PaletteId>",
        f"<Options>{opts_val}</Options>",
        f"<HideHat>{_int(cfg.get('hide_hat',0))}</HideHat>",
        f"<ChrTypeFlags>{_int(cfg.get('chr_type_flags',0))}</ChrTypeFlags>",
        f"<GroundFlags>{_int(cfg.get('ground_flags',0))}</GroundFlags>",
        f"<SystemFlags>{_int(cfg.get('system_flags',0))}</SystemFlags>",
        f"<OptionsEx>{optex_val}</OptionsEx>",
        f"<Weight>{_int(cfg.get('weight',1))}</Weight>",
        f"<Value>{_int(cfg.get('value',0))}</Value>",
        f"<MinLevel>{_int(cfg.get('min_level',1))}</MinLevel>",
        f"<Effect>{_parse_effect_val(cfg.get('effect',0))}</Effect>",
        f"<EffectFlags2>{_int(cfg.get('effect_flags2',0))}</EffectFlags2>",
        f"<SelRange>{_int(cfg.get('sel_range',0))}</SelRange>",
        f"<Life>{_int(cfg.get('life',0))}</Life>",
        f"<Depth>{_int(cfg.get('depth',0))}</Depth>",
        f"<Delay>{_fmt6(cfg.get('delay',0))}</Delay>",
        f"<AP>{_int(cfg.get('ap',0))}</AP>",
        f"<HP>{_int(cfg.get('hp',0))}</HP>",
        f"<HPCon>{_int(cfg.get('hpcon',0))}</HPCon>",
        f"<MP>{_int(cfg.get('mp',0))}</MP>",
        f"<MPCon>{_int(cfg.get('mpcon',0))}</MPCon>",
        f"<Money>{_int(cfg.get('money',0))}</Money>",
        f"<APPlus>{_int(cfg.get('applus',0))}</APPlus>",
        f"<ACPlus>{_int(cfg.get('acplus',0))}</ACPlus>",
        f"<DXPlus>{_int(cfg.get('dxplus',0))}</DXPlus>",
        f"<MaxMPPlus>{_int(cfg.get('maxmpplus',0))}</MaxMPPlus>",
        f"<MAPlus>{_int(cfg.get('maplus',0))}</MAPlus>",
        f"<MDPlus>{_int(cfg.get('mdplus',0))}</MDPlus>",
        f"<MaxWTPlus>{_int(cfg.get('maxwtplus',0))}</MaxWTPlus>",
        f"<DAPlus>{_int(cfg.get('daplus',0))}</DAPlus>",
        f"<LKPlus>{_int(cfg.get('lkplus',0))}</LKPlus>",
        f"<MaxHPPlus>{_int(cfg.get('maxhpplus',0))}</MaxHPPlus>",
        f"<DPPlus>{_int(cfg.get('dpplus',0))}</DPPlus>",
        f"<HVPlus>{_int(cfg.get('hvplus',0))}</HVPlus>",
        f"<HPRecoveryRate>{_fmt6(cfg.get('hprecoveryrate',0))}</HPRecoveryRate>",
        f"<MPRecoveryRate>{_fmt6(cfg.get('mprecoveryrate',0))}</MPRecoveryRate>",
        f"<CardNum>{_int(cfg.get('cardnum',0))}</CardNum>",
        f"<CardGenGrade>{_int(cfg.get('cardgengrade',0))}</CardGenGrade>",
        f"<CardGenParam>{_fmt6(cfg.get('cardgenparam',0))}</CardGenParam>",
        f"<DailyGenCnt>{_int(cfg.get('dailygencnt',0))}</DailyGenCnt>",
        f"<PartFileName>{_cd(cfg.get('part_file_name',' '))}</PartFileName>",
        f"<ChrFTypeFlag>{_int(cfg.get('chr_ftype_flag',0))}</ChrFTypeFlag>",
        f"<ChrGender>{_int(cfg.get('chr_gender',0))}</ChrGender>",
        f"<ExistType>{_int(cfg.get('exist_type',0))}</ExistType>",
        f"<Ncash>{_int(cfg.get('ncash',0))}</Ncash>",
        f"<NewCM>{_int(cfg.get('new_cm',0))}</NewCM>",
        f"<FamCM>{_int(cfg.get('fam_cm',0))}</FamCM>",
        f"<Summary>{_cd(cfg.get('summary',' '))}</Summary>",
        f"<ShopFileName>{_cd(cfg.get('shop_file_name',' '))}</ShopFileName>",
        f"<ShopBundleNum>{_int(cfg.get('shop_bundle_num',0))}</ShopBundleNum>",
        f"<MinStatType>{_int(cfg.get('min_stat_type',0))}</MinStatType>",
        f"<MinStatLv>{_int(cfg.get('min_stat_lv',0))}</MinStatLv>",
        f"<RefineIndex>{_int(cfg.get('refine_index',0))}</RefineIndex>",
        f"<RefineType>{_int(cfg.get('refine_type',0))}</RefineType>",
        f"<CompoundSlot>{_int(cfg.get('compound_slot',0))}</CompoundSlot>",
        f"<SetItemID>{_int(cfg.get('set_item_id',0))}</SetItemID>",
        f"<ReformCount>{_int(cfg.get('reform_count',0))}</ReformCount>",
        f"<GroupId>{_int(cfg.get('group_id',0))}</GroupId>",
        "</ROW>",
    ]
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# COMPOUND / EXCHANGE XML BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_compound_row(cfg):
    """Build Compound_Potion.xml <ROW>."""
    def _cd(v): return f"<![CDATA[{v}]]>"
    def _i(v, d=0):
        try: return int(v)
        except: return d
    lines = ["<ROW>",
        f"<CompoundID>{_i(cfg.get('compound_id',0))}</CompoundID>",
        f"<Name>{_cd(cfg.get('name',''))}</Name>",
        f"<Comment>{_cd(cfg.get('comment',''))}</Comment>",
        f"<ResLv>{_i(cfg.get('res_lv',1))}</ResLv>",
    ]
    for n in range(1,4):
        lines.append(f"<ResID{n}>{_i(cfg.get(f'res_id{n}',0))}</ResID{n}>")
    for n in range(1,6):
        lines.append(f"<ReqID{n}>{_i(cfg.get(f'req_id{n}',0))}</ReqID{n}>")
        lines.append(f"<ReqNum{n}>{_i(cfg.get(f'req_num{n}',0))}</ReqNum{n}>")
    lines += [
        f"<Probability>{_i(cfg.get('probability',50))}</Probability>",
        f"<Fee>{_i(cfg.get('fee',1))}</Fee>",
        f"<WasteItem>{_i(cfg.get('waste_item',12000))}</WasteItem>",
        "</ROW>",
    ]
    return "\n".join(lines)

def build_compound_location_row(compound_id):
    """Build Compounder_Location.xml <ROW> (Probability and Hidden always 0)."""
    return f"<ROW>\n<CompoundID>{compound_id}</CompoundID>\n<Probability>0</Probability>\n<Hidden>0</Hidden>\n</ROW>"

def build_exchange_row(cfg):
    """Build ExchangeShopContents.xml <ROW>."""
    def _cd(v): return f"<![CDATA[{v}]]>"
    def _i(v, d=0):
        try: return int(v)
        except: return d
    lines = ["<ROW>",
        f"<ExchangeID>{_i(cfg.get('exchange_id',0))}</ExchangeID>",
        f"<Name>{_cd(cfg.get('name',''))}</Name>",
        f"<Comment>{_cd(cfg.get('comment',''))}</Comment>",
        f"<ResLv>{_i(cfg.get('res_lv',1))}</ResLv>",
    ]
    for n in range(1,4):
        lines.append(f"<ResID{n}>{_i(cfg.get(f'res_id{n}',0))}</ResID{n}>")
    for n in range(1,6):
        lines.append(f"<ReqID{n}>{_i(cfg.get(f'req_id{n}',0))}</ReqID{n}>")
        lines.append(f"<ReqNum{n}>{_i(cfg.get(f'req_num{n}',0))}</ReqNum{n}>")
    lines += [f"<Fee>{_i(cfg.get('fee',0))}</Fee>", "</ROW>"]
    return "\n".join(lines)

def build_exchange_location_row(exchange_id):
    """Build Exchange_Location.xml <ROW>."""
    return f"<ROW>\n<ExchangeID>{exchange_id}</ExchangeID>\n</ROW>"

# Session-level CE preference store (runtime only, not persisted to file)
_ce_session_pref = {"choice": None}   # None = always ask, else "compound"/"exchange"/"none"

def _show_compound_exchange_dialog(root, item_name, item_comment, item_id,
                                   on_compound, on_exchange, on_skip,
                                   remember_var=None, box_effect=None):
    """Popup asking whether to also generate compound/exchange data.
    Respects session-level remembered choice to skip re-asking."""
    # If session pref is set, skip dialog and act immediately
    if _ce_session_pref["choice"] is not None:
        ch = _ce_session_pref["choice"]
        if ch == "compound":
            _show_compound_form(root, item_name, item_comment, item_id, on_compound)
        elif ch == "exchange":
            _show_exchange_form(root, item_name, item_comment, item_id, on_exchange)
        else:
            on_skip()
        return

    win = tk.Toplevel(root)
    win.title("Generate Compound / Exchange?")
    win.configure(bg=BG)
    win.grab_set()
    win.resizable(False, False)

    tk.Label(win, text="Add Compound or Exchange entry?",
             bg=BG, fg=FG, font=("Consolas",12,"bold")).pack(pady=(16,4), padx=20)
    tk.Label(win, text=f"Item: {item_name or item_id}",
             bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(padx=20)

    choice_var = tk.StringVar(value="none")
    bf = tk.Frame(win, bg=BG); bf.pack(pady=10)
    for lbl, val in [("Compound","compound"),("Exchange","exchange"),("Neither — skip","none")]:
        tk.Radiobutton(bf, text=lbl, variable=choice_var, value=val,
                       bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                       font=("Consolas",10)).pack(anchor="w", padx=10, pady=2)

    rem_var = tk.BooleanVar(value=False)
    tk.Checkbutton(win, text="Remember this choice for the rest of this session",
                   variable=rem_var, bg=BG, fg=ACC4,
                   selectcolor=BG3, activebackground=BG,
                   font=("Consolas",9)).pack(padx=20, pady=(0,8))

    def _proceed():
        ch = choice_var.get()
        if rem_var.get():
            _ce_session_pref["choice"] = ch
        win.destroy()
        if ch == "compound":
            _show_compound_form(root, item_name, item_comment, item_id, on_compound)
        elif ch == "exchange":
            _show_exchange_form(root, item_name, item_comment, item_id, on_exchange)
        else:
            on_skip()

    # Reset session pref button
    def _reset_pref():
        _ce_session_pref["choice"] = None
        messagebox.showinfo("Reset", "Session preference cleared — you will be asked each time.")

    bf2 = tk.Frame(win, bg=BG); bf2.pack()
    mk_btn(bf2, "Continue ▶", _proceed, color=GREEN, fg=BG2,
           font=("Consolas",10,"bold")).pack(side="left", padx=8, pady=(0,14))
    mk_btn(bf2, "Reset Remembered Choice", _reset_pref, color=BG4,
           font=("Consolas",8)).pack(side="left", padx=4, pady=(0,14))
    win.wait_window()

def _make_id_row(parent, label, var, tip=""):
    r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
    lw = tk.Label(r, text=label, width=20, anchor="w", bg=BG, fg=FG, font=("Consolas",9))
    lw.pack(side="left")
    ent = tk.Entry(r, textvariable=var, width=14, bg=BG3, fg=FG,
                   insertbackground=FG, font=("Consolas",9), relief="flat")
    ent.pack(side="left", padx=4)
    if tip: _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
    return ent

# ══════════════════════════════════════════════════════════════════════════════
# CSV / Excel batch importer for Compound and Exchange
# ══════════════════════════════════════════════════════════════════════════════

# ── Column aliases ─────────────────────────────────────────────────────────
_COMPOUND_COL_ALIASES = {
    "compoundid":    "compound_id",
    "compound_id":   "compound_id",
    "id":            "compound_id",
    "name":          "name",
    "n":             "name",
    "comment":       "comment",
    "reslv":         "res_lv",
    "res_lv":        "res_lv",
    "requiredlevel": "res_lv",
    "resid1":        "res_id1",  "res_id1": "res_id1",
    "resid2":        "res_id2",  "res_id2": "res_id2",
    "resid3":        "res_id3",  "res_id3": "res_id3",
    "reqid1":        "req_id1",  "req_id1": "req_id1",
    "reqid2":        "req_id2",  "req_id2": "req_id2",
    "reqid3":        "req_id3",  "req_id3": "req_id3",
    "reqid4":        "req_id4",  "req_id4": "req_id4",
    "reqid5":        "req_id5",  "req_id5": "req_id5",
    "reqnum1":       "req_num1", "req_num1": "req_num1",
    "reqnum2":       "req_num2", "req_num2": "req_num2",
    "reqnum3":       "req_num3", "req_num3": "req_num3",
    "reqnum4":       "req_num4", "req_num4": "req_num4",
    "reqnum5":       "req_num5", "req_num5": "req_num5",
    "probability":   "probability",
    "fee":           "fee",
    "wasteitem":     "waste_item", "waste_item": "waste_item",
}

_EXCHANGE_COL_ALIASES = {
    "exchangeid":    "exchange_id",
    "exchange_id":   "exchange_id",
    "id":            "exchange_id",
    "name":          "name",
    "n":             "name",
    "comment":       "comment",
    "reslv":         "res_lv",
    "res_lv":        "res_lv",
    "requiredlevel": "res_lv",
    "resid1":        "res_id1",  "res_id1": "res_id1",
    "resid2":        "res_id2",  "res_id2": "res_id2",
    "resid3":        "res_id3",  "res_id3": "res_id3",
    "reqid1":        "req_id1",  "req_id1": "req_id1",
    "reqid2":        "req_id2",  "req_id2": "req_id2",
    "reqid3":        "req_id3",  "req_id3": "req_id3",
    "reqid4":        "req_id4",  "req_id4": "req_id4",
    "reqid5":        "req_id5",  "req_id5": "req_id5",
    "reqnum1":       "req_num1", "req_num1": "req_num1",
    "reqnum2":       "req_num2", "req_num2": "req_num2",
    "reqnum3":       "req_num3", "req_num3": "req_num3",
    "reqnum4":       "req_num4", "req_num4": "req_num4",
    "reqnum5":       "req_num5", "req_num5": "req_num5",
    "fee":           "fee",
}

# ── Compound defaults ──────────────────────────────────────────────────────
_COMPOUND_DEFAULTS = {
    "compound_id": "0", "name": "", "comment": "", "res_lv": "1",
    "res_id1": "0", "res_id2": "0", "res_id3": "0",
    "req_id1": "0", "req_id2": "0", "req_id3": "0", "req_id4": "0", "req_id5": "0",
    "req_num1": "0", "req_num2": "0", "req_num3": "0", "req_num4": "0", "req_num5": "0",
    "probability": "50", "fee": "1", "waste_item": "12000",
}

# ── Exchange defaults ──────────────────────────────────────────────────────
_EXCHANGE_DEFAULTS = {
    "exchange_id": "0", "name": "", "comment": "", "res_lv": "1",
    "res_id1": "0", "res_id2": "0", "res_id3": "0",
    "req_id1": "0", "req_id2": "0", "req_id3": "0", "req_id4": "0", "req_id5": "0",
    "req_num1": "0", "req_num2": "0", "req_num3": "0", "req_num4": "0", "req_num5": "0",
    "fee": "0",
}

def _normalise_col(name):
    """Strip, lowercase, remove spaces/underscores for fuzzy matching."""
    return re.sub(r'[\s_]+', '', str(name).lower().strip())


def _read_csv_rows(filepath):
    """Return (headers, list-of-dicts) from a CSV file."""
    with open(filepath, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    return headers, rows


def _read_xlsx_rows(filepath, sheet_index=0):
    """Return (headers, list-of-dicts) from an xlsx file (first sheet by default)."""
    if _openpyxl is None:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")
    wb = _openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows_raw = list(ws.values)
    if not rows_raw:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows_raw[0]]
    rows = []
    for row in rows_raw[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip fully blank rows
        rows.append({headers[i]: (str(row[i]) if row[i] is not None else "")
                     for i in range(len(headers))})
    return headers, rows


def _map_row_to_cfg(raw_row, alias_map, defaults):
    """Map a raw CSV/xlsx row dict -> cfg dict using alias_map, filling defaults."""
    cfg = dict(defaults)
    for col, val in raw_row.items():
        key = alias_map.get(_normalise_col(col))
        if key and str(val).strip() != "":
            cfg[key] = str(val).strip()
    return cfg


def _import_preview_window(root, mode, filepath, on_confirm):
    """
    Show a scrollable preview of parsed rows before committing.
    mode = "compound" or "exchange"
    on_confirm(cfgs) called with list of cfg dicts.
    """
    alias_map = _COMPOUND_COL_ALIASES if mode == "compound" else _EXCHANGE_COL_ALIASES
    defaults  = _COMPOUND_DEFAULTS    if mode == "compound" else _EXCHANGE_DEFAULTS
    id_key    = "compound_id"         if mode == "compound" else "exchange_id"
    color     = ACC1                  if mode == "compound" else ACC2
    title_lbl = "Compound_Potion.xml" if mode == "compound" else "ExchangeShopContents.xml"

    # ── Parse ──────────────────────────────────────────────────────────────
    try:
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            headers, raw_rows = _read_xlsx_rows(filepath)
        else:
            headers, raw_rows = _read_csv_rows(filepath)
    except Exception as e:
        messagebox.showerror("Import Error", f"Could not read file:\n{e}", parent=root)
        return

    if not raw_rows:
        messagebox.showwarning("Empty File", "No data rows found in the file.", parent=root)
        return

    cfgs = [_map_row_to_cfg(r, alias_map, defaults) for r in raw_rows]

    # ── Detect unmapped columns ────────────────────────────────────────────
    unmapped = []
    for col in headers:
        if _normalise_col(col) not in alias_map:
            unmapped.append(col)

    # ── Preview window ─────────────────────────────────────────────────────
    win = tk.Toplevel(root)
    win.title(f"Import Preview — {mode.capitalize()} ({len(cfgs)} rows)")
    win.configure(bg=BG)
    win.geometry("960x620")
    win.grab_set()

    # Header
    hdr = tk.Frame(win, bg=BG2); hdr.pack(fill="x")
    tk.Label(hdr, text=f"📥  Import {mode.capitalize()} — {title_lbl}",
             bg=BG2, fg=color, font=("Consolas", 12, "bold"),
             pady=8).pack(side="left", padx=14)
    tk.Label(hdr, text=f"{len(cfgs)} rows from {os.path.basename(filepath)}",
             bg=BG2, fg=FG_DIM, font=("Consolas", 9)).pack(side="left", padx=6)

    # Unmapped warning
    if unmapped:
        wf = tk.Frame(win, bg="#45475a"); wf.pack(fill="x", padx=8, pady=(4, 0))
        tk.Label(wf, text=f"⚠  Columns not recognised (will be ignored): {', '.join(unmapped)}",
                 bg="#45475a", fg=ACC4, font=("Consolas", 8),
                 pady=4).pack(anchor="w", padx=10)

    # Scrollable table
    tbl_host = tk.Frame(win, bg=BG); tbl_host.pack(fill="both", expand=True, padx=8, pady=6)

    x_scroll = tk.Scrollbar(tbl_host, orient="horizontal")
    y_scroll = tk.Scrollbar(tbl_host, orient="vertical")
    tbl_canvas = tk.Canvas(tbl_host, bg=BG, bd=0, highlightthickness=0,
                           xscrollcommand=x_scroll.set,
                           yscrollcommand=y_scroll.set)
    x_scroll.config(command=tbl_canvas.xview)
    y_scroll.config(command=tbl_canvas.yview)
    x_scroll.pack(side="bottom", fill="x")
    y_scroll.pack(side="right",  fill="y")
    tbl_canvas.pack(side="left", fill="both", expand=True)

    tbl_inner = tk.Frame(tbl_canvas, bg=BG)
    tbl_canvas.create_window((0, 0), window=tbl_inner, anchor="nw")
    tbl_inner.bind("<Configure>",
        lambda e: tbl_canvas.configure(scrollregion=tbl_canvas.bbox("all")))

    # Columns to show
    if mode == "compound":
        show_cols = [id_key, "name", "comment", "res_lv",
                     "res_id1", "res_id2", "res_id3",
                     "req_id1", "req_num1", "req_id2", "req_num2",
                     "req_id3", "req_num3", "req_id4", "req_num4", "req_id5", "req_num5",
                     "probability", "fee", "waste_item"]
    else:
        show_cols = [id_key, "name", "comment", "res_lv",
                     "res_id1", "res_id2", "res_id3",
                     "req_id1", "req_num1", "req_id2", "req_num2",
                     "req_id3", "req_num3", "req_id4", "req_num4", "req_id5", "req_num5",
                     "fee"]

    col_w = 10
    # Header row
    for ci, col in enumerate(show_cols):
        tk.Label(tbl_inner, text=col, bg=BG2, fg=BLUE,
                 font=("Consolas", 8, "bold"), width=col_w, anchor="w",
                 relief="flat", padx=3).grid(row=0, column=ci, sticky="w", padx=1, pady=1)

    # Data rows — editable
    row_vars = []  # list of dicts: col -> StringVar
    for ri, cfg in enumerate(cfgs):
        rv = {}
        bg_row = BG if ri % 2 == 0 else BG2
        for ci, col in enumerate(show_cols):
            v = tk.StringVar(value=cfg.get(col, ""))
            rv[col] = v
            ent = tk.Entry(tbl_inner, textvariable=v, width=col_w,
                           bg=bg_row, fg=FG, insertbackground=FG,
                           font=("Consolas", 8), relief="flat", bd=0)
            ent.grid(row=ri + 1, column=ci, sticky="w", padx=1, pady=1)
        row_vars.append(rv)

    # Footer
    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")
    tk.Label(nav, text="You can edit cells above before confirming.",
             bg=BG2, fg=FG_DIM, font=("Consolas", 8)).pack(side="left", padx=14, pady=6)

    def _confirm():
        # Read back edited values
        final_cfgs = []
        for rv in row_vars:
            cfg_out = dict(defaults)  # start from defaults
            for col, v in rv.items():
                cfg_out[col] = v.get()
            final_cfgs.append(cfg_out)
        win.destroy()
        on_confirm(final_cfgs)

    mk_btn(nav, f"✓  Import {len(cfgs)} rows", _confirm,
           color=GREEN, fg=BG2, font=("Consolas", 10, "bold")).pack(side="right", padx=14, pady=6)
    mk_btn(nav, "Cancel", win.destroy, color=BG4).pack(side="right", padx=4, pady=6)
    win.wait_window()


def _ask_import_mode_then_file(root, on_compound_cfgs, on_exchange_cfgs):
    """
    Step 1: ask Compound or Exchange.
    Step 2: file picker.
    Step 3: preview + confirm.
    on_compound_cfgs(cfgs) / on_exchange_cfgs(cfgs) called on confirm.
    """
    win = tk.Toplevel(root)
    win.title("Import — Choose Mode")
    win.configure(bg=BG)
    win.resizable(False, False)
    win.grab_set()

    tk.Label(win, text="📥  Import from CSV / Excel",
             bg=BG, fg=BLUE, font=("Consolas", 13, "bold"),
             pady=12).pack(padx=28)
    tk.Label(win, text="What type of data are you importing?",
             bg=BG, fg=FG, font=("Consolas", 10)).pack(padx=28, pady=(0, 16))

    btn_frm = tk.Frame(win, bg=BG); btn_frm.pack(padx=28, pady=4)

    def _pick_file(mode):
        win.destroy()
        path = filedialog.askopenfilename(
            title=f"Import {mode.capitalize()} data",
            filetypes=[
                ("Spreadsheet / CSV", "*.csv *.xlsx *.xlsm *.xls"),
                ("CSV files",  "*.csv"),
                ("Excel files","*.xlsx *.xlsm *.xls"),
                ("All files",  "*.*"),
            ],
            parent=root,
        )
        if not path:
            return
        if mode == "compound":
            _import_preview_window(root, "compound", path, on_compound_cfgs)
        else:
            _import_preview_window(root, "exchange", path, on_exchange_cfgs)

    mk_btn(btn_frm, "⚗  Compound_Potion.xml", lambda: _pick_file("compound"),
           color=ACC1, fg=BG2, font=("Consolas", 11, "bold"),
           padx=18, pady=10).grid(row=0, column=0, padx=10, pady=6)
    mk_btn(btn_frm, "🔄  ExchangeShopContents.xml", lambda: _pick_file("exchange"),
           color=ACC2, fg=BG2, font=("Consolas", 11, "bold"),
           padx=18, pady=10).grid(row=0, column=1, padx=10, pady=6)

    tk.Label(win,
             text="CSV column names are matched automatically.\n"
                  "Click 📖 Template below to download a blank template.",
             bg=BG, fg=FG_GREY, font=("Consolas", 8),
             justify="center").pack(pady=(10, 4))

    def _save_template(mode):
        if mode == "compound":
            cols = list(_COMPOUND_DEFAULTS.keys())
        else:
            cols = list(_EXCHANGE_DEFAULTS.keys())
        path = filedialog.asksaveasfilename(
            title=f"Save {mode} template",
            defaultextension=".csv",
            filetypes=[("CSV","*.csv")],
            initialfile=f"{mode}_template.csv",
            parent=root,
        )
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(cols)
            writer.writerow(["" for _ in cols])  # blank example row
        messagebox.showinfo("Template saved", f"Saved to:\n{path}", parent=root)

    tpl_frm = tk.Frame(win, bg=BG); tpl_frm.pack(pady=(0, 14))
    mk_btn(tpl_frm, "📄  Compound template", lambda: _save_template("compound"),
           color=BG4, font=("Consolas", 9)).pack(side="left", padx=8)
    mk_btn(tpl_frm, "📄  Exchange template", lambda: _save_template("exchange"),
           color=BG4, font=("Consolas", 9)).pack(side="left", padx=8)
    mk_btn(win, "Cancel", win.destroy, color=BG4).pack(pady=(0, 12))
    win.wait_window()


def _show_compound_form(root, item_name, item_comment, item_id, on_done):
    """Full compound entry dialog."""
    last_id = _get_last_id("compound", 100)
    win = tk.Toplevel(root)
    win.title("Compound Entry")
    win.configure(bg=BG)
    win.grab_set()

    sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True)
    canv, C = mk_scroll_canvas(sh, init_width=540)
    win.geometry("580x620")

    tk.Label(C, text="Compound_Potion.xml Entry",
             bg=BG, fg=ACC1, font=("Consolas",12,"bold")).pack(pady=(12,2), padx=12, anchor="w")

    def sec(t):
        f = tk.LabelFrame(C, text=t, bg=BG, fg=BLUE,
                          font=("Consolas",9,"bold"), bd=1, relief="groove")
        f.pack(fill="x", padx=10, pady=4)
        return f

    s1 = sec("  Identity  ")
    v_cid  = tk.StringVar(value=str(last_id + 1))
    v_name = tk.StringVar(value=item_name or "")
    v_cmt  = tk.StringVar(value=item_comment or "")
    v_reslv= tk.StringVar(value="1")
    _make_id_row(s1, "CompoundID:", v_cid, "Unique ID for this compound recipe.")
    _make_id_row(s1, "Name:", v_name)
    _make_id_row(s1, "Comment:", v_cmt)
    _make_id_row(s1, "Required Level:", v_reslv, "ResLv — minimum level to use this compound.")

    s2 = sec("  Receiving IDs (ResID 1-3)  ")
    tk.Label(s2, text="Item IDs produced by this recipe (0 = unused).",
             bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", padx=8)
    res_vars = []
    for n in range(1,4):
        v = tk.StringVar(value=item_id if n==1 else "0")
        _make_id_row(s2, f"ResID{n}:", v, f"Receiving item ID #{n}")
        res_vars.append(v)

    s3 = sec("  Required Items (ReqID/ReqNum 1-5)  ")
    tk.Label(s3, text="Items required to perform this compound (0 = unused).",
             bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", padx=8)
    req_id_vars = []; req_num_vars = []
    for n in range(1,6):
        r = tk.Frame(s3, bg=BG); r.pack(fill="x", padx=8, pady=1)
        tk.Label(r, text=f"ReqID{n}:", width=8, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left")
        vid = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vid, width=12, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        tk.Label(r, text=f"ReqNum{n}:", width=10, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left", padx=(8,0))
        vnum = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vnum, width=6, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        req_id_vars.append(vid); req_num_vars.append(vnum)

    s4 = sec("  Recipe Settings  ")
    v_prob  = tk.StringVar(value="50")
    v_fee   = tk.StringVar(value="1")
    v_waste = tk.StringVar(value="12000")
    _make_id_row(s4, "Probability:", v_prob, "Success chance (0-100).")
    _make_id_row(s4, "Fee:", v_fee, "Galder fee to attempt compound.")
    _make_id_row(s4, "WasteItem:", v_waste, "Item consumed on failure (default 12000).")

    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")

    def _gen():
        cfg = {
            "compound_id": v_cid.get(), "name": v_name.get(), "comment": v_cmt.get(),
            "res_lv": v_reslv.get(),
            "res_id1": res_vars[0].get(), "res_id2": res_vars[1].get(), "res_id3": res_vars[2].get(),
            "probability": v_prob.get(), "fee": v_fee.get(), "waste_item": v_waste.get(),
        }
        for n in range(1,6):
            cfg[f"req_id{n}"]  = req_id_vars[n-1].get()
            cfg[f"req_num{n}"] = req_num_vars[n-1].get()
        try: _set_last_id("compound", int(v_cid.get()))
        except: pass
        win.destroy()
        on_done(cfg)

    mk_btn(nav, "✓  Generate", _gen, color=GREEN, fg=BG2,
           font=("Consolas",10,"bold")).pack(side="left", padx=12, pady=6)
    mk_btn(nav, "Cancel", win.destroy).pack(side="left", padx=4, pady=6)
    win.wait_window()

def _show_exchange_form(root, item_name, item_comment, item_id, on_done):
    """Full exchange entry dialog."""
    last_id = _get_last_id("exchange", 0)
    win = tk.Toplevel(root)
    win.title("Exchange Entry")
    win.configure(bg=BG)
    win.grab_set()

    sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True)
    canv, C = mk_scroll_canvas(sh, init_width=540)
    win.geometry("580x580")

    tk.Label(C, text="ExchangeShopContents.xml Entry",
             bg=BG, fg=ACC2, font=("Consolas",12,"bold")).pack(pady=(12,2), padx=12, anchor="w")

    def sec(t):
        f = tk.LabelFrame(C, text=t, bg=BG, fg=BLUE,
                          font=("Consolas",9,"bold"), bd=1, relief="groove")
        f.pack(fill="x", padx=10, pady=4)
        return f

    s1 = sec("  Identity  ")
    v_eid  = tk.StringVar(value=str(last_id + 1))
    v_name = tk.StringVar(value=item_name or "")
    v_cmt  = tk.StringVar(value=item_comment or "")
    v_reslv= tk.StringVar(value="1")
    _make_id_row(s1, "ExchangeID:", v_eid, "Unique ID for this exchange entry.")
    _make_id_row(s1, "Name:", v_name)
    _make_id_row(s1, "Comment:", v_cmt)
    _make_id_row(s1, "Required Level:", v_reslv, "ResLv — minimum level.")

    s2 = sec("  Receiving IDs (ResID 1-3)  ")
    tk.Label(s2, text="Items received from exchange (0 = unused).",
             bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", padx=8)
    res_vars = []
    for n in range(1,4):
        v = tk.StringVar(value=item_id if n==1 else "0")
        _make_id_row(s2, f"ResID{n}:", v)
        res_vars.append(v)

    s3 = sec("  Required Items (ReqID/ReqNum 1-5)  ")
    req_id_vars = []; req_num_vars = []
    for n in range(1,6):
        r = tk.Frame(s3, bg=BG); r.pack(fill="x", padx=8, pady=1)
        tk.Label(r, text=f"ReqID{n}:", width=8, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left")
        vid = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vid, width=12, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        tk.Label(r, text=f"ReqNum{n}:", width=10, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left", padx=(8,0))
        vnum = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vnum, width=6, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        req_id_vars.append(vid); req_num_vars.append(vnum)

    s4 = sec("  Settings  ")
    v_fee = tk.StringVar(value="0")
    _make_id_row(s4, "Fee:", v_fee, "Galder fee for exchange.")

    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")

    def _gen():
        cfg = {
            "exchange_id": v_eid.get(), "name": v_name.get(), "comment": v_cmt.get(),
            "res_lv": v_reslv.get(),
            "res_id1": res_vars[0].get(), "res_id2": res_vars[1].get(), "res_id3": res_vars[2].get(),
            "fee": v_fee.get(),
        }
        for n in range(1,6):
            cfg[f"req_id{n}"]  = req_id_vars[n-1].get()
            cfg[f"req_num{n}"] = req_num_vars[n-1].get()
        try: _set_last_id("exchange", int(v_eid.get()))
        except: pass
        win.destroy()
        on_done(cfg)

    mk_btn(nav, "✓  Generate", _gen, color=GREEN, fg=BG2,
           font=("Consolas",10,"bold")).pack(side="left", padx=12, pady=6)
    mk_btn(nav, "Cancel", win.destroy).pack(side="left", padx=4, pady=6)
    win.wait_window()

class Tool6(tk.Frame):
    """ItemParam Generator — builds full <ROW> entries for any item type."""

    ACC = "#94e2d5"   # teal — tool 6

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root    = root
        self.session = session
        self._settings = _load_t6_settings()
        self._first_run = not bool(self._settings.get("_first_run_done", False))
        self._rows = []           # list of generated row strings
        self._build_editor()

    # ─────────────────────────────────────────────────────────────────────────
    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    # ─────────────────────────────────────────────────────────────────────────
    def _build_editor(self):
        self._clear()
        s = self._settings
        ACC = self.ACC

        # ── Outer wrapper with header / scroll / footer layout ────────────
        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0, weight=0)
        wrap.grid_rowconfigure(1, weight=1)
        wrap.grid_rowconfigure(2, weight=0)
        wrap.grid_columnconfigure(0, weight=1)

        # ── Header ────────────────────────────────────────────────────────
        hdr = tk.Frame(wrap, bg=BG2)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="ItemParam Generator",
                 font=("Consolas", 14, "bold"), bg=BG2, fg=ACC, pady=8
                 ).pack(side="left", padx=15)
        tk.Label(hdr, text="Generates full <ROW> entries for any item type.",
                 font=("Consolas", 9), bg=BG2, fg=FG_DIM
                 ).pack(side="left", padx=4)
        mk_btn(hdr, "📖  Field Reference", self._show_reference,
               color=BG4, font=("Consolas", 9)).pack(side="right", padx=10, pady=4)

        # ── Scrollable body ───────────────────────────────────────────────
        scroll_host = tk.Frame(wrap, bg=BG)
        scroll_host.grid(row=1, column=0, sticky="nsew")
        canvas, C = mk_scroll_canvas(scroll_host)

        def sec(title):
            return mk_section(C, title)

        def lbl_entry(parent, label, var, width=30, tip=None):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            lw = tk.Label(r, text=label, width=22, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9)); lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=width, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip: _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent

        def lbl_note(parent, text, fg=FG_GREY):
            tk.Label(parent, text=text, bg=BG, fg=fg,
                     font=("Consolas", 8)).pack(anchor="w", padx=10, pady=(2, 0))

        # ── First-run welcome banner ──────────────────────────────────────
        if self._first_run:
            wb = tk.Frame(C, bg=BG4, padx=12, pady=8)
            wb.pack(fill="x", padx=12, pady=(10, 4))
            tk.Label(wb, text="👋  Welcome to the ItemParam Generator",
                     bg=BG4, fg=ACC, font=("Consolas", 11, "bold")).pack(anchor="w")
            tk.Label(wb, text=(
                "All fields pre-filled with safe defaults.  Hover any label for its description.\n"
                "Class and Type default to Unselected — set them before generating.\n"
                "Your settings are saved between sessions.  Click 📖 Field Reference for lookup tables."),
                bg=BG4, fg=FG, font=("Consolas", 9), justify="left").pack(anchor="w", pady=(4, 0))
            mk_btn(wb, "Got it — dismiss", lambda: wb.destroy(),
                   color=BG3, font=("Consolas", 9)).pack(anchor="w", pady=(6, 0))

        # ── Mode toggle ───────────────────────────────────────────────────
        mode_var = tk.StringVar(value=s.get("input_mode", "dropdown"))
        mode_frm = tk.Frame(C, bg=BG); mode_frm.pack(fill="x", padx=14, pady=(8, 2))
        tk.Label(mode_frm, text="Input mode:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 9)).pack(side="left")
        for lbl_m, val_m in [("Dropdown", "dropdown"), ("Manual", "manual")]:
            tk.Radiobutton(mode_frm, text=lbl_m, variable=mode_var, value=val_m,
                           bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                           font=("Consolas", 9)).pack(side="left", padx=8)
        lbl_note(mode_frm.master,
                 "  Manual mode: Class/Type/SubType/ItemFType become free-text; "
                 "Options & OptionsEx become raw numeric fields.")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 1 — Identity
        # ═════════════════════════════════════════════════════════════════
        s1 = sec("  Identity  ⚠ Class and Type MUST be set (not 0 / Unselected)")

        _default_id = s.get("id", "") or str(_get_last_id("t6_item", 0) + 1 or "")
        if _default_id == "1": _default_id = s.get("id", "")
        v_id    = tk.StringVar(value=_default_id)
        v_class = tk.StringVar(value=s.get("class_val", "0"))
        v_type  = tk.StringVar(value=s.get("type_val", "0"))
        v_sub   = tk.StringVar(value=s.get("subtype_val", "0"))
        v_ift   = tk.StringVar(value=s.get("itemftype_val", "0"))

        lbl_entry(s1, "ID:", v_id, width=14,
                  tip="Item ID — must be unique in the XML table.  Required.")

        # ── Dropdown builder (shared by class/type/sub/ift) ───────────────
        def _build_dd_row(parent, label, mapping, var, tip):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            lw = tk.Label(r, text=label, width=22, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9)); lw.pack(side="left")
            vals_dd = [f"{v} — {d}" for v, d in mapping]
            combo = ttk.Combobox(r, values=vals_dd, state="readonly",
                                 width=38, font=("Consolas", 9))
            combo.pack(side="left", padx=4)
            raw_ent = tk.Entry(r, textvariable=var, width=7, bg=BG3, fg=FG,
                               insertbackground=FG, font=("Consolas", 9), relief="flat")
            raw_ent.pack(side="left", padx=(0, 4))
            def _dd_sel(e):
                sel = combo.current()
                if sel >= 0: var.set(str(mapping[sel][0]))
            combo.bind("<<ComboboxSelected>>", _dd_sel)
            def _raw_ch(*_):
                val = var.get().strip()
                for i, (v, _) in enumerate(mapping):
                    if str(v) == val: combo.set(vals_dd[i]); return
                combo.set("")
            var.trace_add("write", _raw_ch); _raw_ch()
            if tip:
                _attach_tooltip(lw, tip); _attach_tooltip(combo, tip)
                _attach_tooltip(raw_ent, tip)
            return combo, raw_ent

        def _build_man_row(parent, label, var, tip):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            lw = tk.Label(r, text=label, width=22, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9)); lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=14, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip: _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent

        dd_class_frm = tk.Frame(s1, bg=BG)
        dd_type_frm  = tk.Frame(s1, bg=BG)
        dd_sub_frm   = tk.Frame(s1, bg=BG)
        dd_ift_frm   = tk.Frame(s1, bg=BG)
        man_class_frm = tk.Frame(s1, bg=BG)
        man_type_frm  = tk.Frame(s1, bg=BG)
        man_sub_frm   = tk.Frame(s1, bg=BG)
        man_ift_frm   = tk.Frame(s1, bg=BG)

        _build_dd_row(dd_class_frm, "Class: *",   _CLASS_MAP,    v_class,
                      _TOOLTIPS["Class"] + "\n⚠ REQUIRED — must not be 0 or Unselected.")
        _build_dd_row(dd_type_frm,  "Type: *",    _TYPE_MAP,     v_type,
                      _TOOLTIPS["Type"] + "\n⚠ REQUIRED — must not be 0 or Unselected.")
        _build_dd_row(dd_sub_frm,   "SubType:",   _SUBTYPE_MAP,  v_sub,   _TOOLTIPS["SubType"])
        _build_dd_row(dd_ift_frm,   "ItemFType:", _ITEMFTYPE_MAP, v_ift,  _TOOLTIPS["ItemFType"])
        _build_man_row(man_class_frm, "Class: *",   v_class,
                       _TOOLTIPS["Class"] + "\n⚠ REQUIRED.")
        _build_man_row(man_type_frm,  "Type: *",    v_type,
                       _TOOLTIPS["Type"] + "\n⚠ REQUIRED.")
        _build_man_row(man_sub_frm,   "SubType:",   v_sub,   _TOOLTIPS["SubType"])
        _build_man_row(man_ift_frm,   "ItemFType:", v_ift,   _TOOLTIPS["ItemFType"])

        def _refresh_mode(*_):
            dd = mode_var.get() == "dropdown"
            for f in (dd_class_frm, dd_type_frm, dd_sub_frm, dd_ift_frm):
                if dd:  f.pack(fill="x")
                else:   f.pack_forget()
            for f in (man_class_frm, man_type_frm, man_sub_frm, man_ift_frm):
                if not dd: f.pack(fill="x")
                else:      f.pack_forget()

        mode_var.trace_add("write", _refresh_mode)
        _refresh_mode()

        lbl_note(s1, "  SubType and ItemFType default to 0 (N/A).  Hover any label for description.")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 2 — Names & Text
        # ═════════════════════════════════════════════════════════════════
        s2 = sec("  Names & Text  ")
        v_name     = tk.StringVar(value=s.get("name", ""))
        v_comment  = tk.StringVar(value=s.get("comment", ""))
        v_use      = tk.StringVar(value=s.get("use", ""))
        v_name_eng = tk.StringVar(value=s.get("name_eng", " "))
        v_cmt_eng  = tk.StringVar(value=s.get("comment_eng", " "))
        lbl_entry(s2, "Name:",        v_name,    40, _TOOLTIPS["Name"])
        lbl_entry(s2, "Comment:",     v_comment, 60, _TOOLTIPS["Comment"])
        lbl_entry(s2, "Use:",         v_use,     60, _TOOLTIPS["Use"])
        lbl_entry(s2, "Name_Eng:",    v_name_eng, 40, _TOOLTIPS["Name_Eng"])
        lbl_entry(s2, "Comment_Eng:", v_cmt_eng,  40, _TOOLTIPS["Comment_Eng"])
        lbl_note(s2, "  CDATA wrappers are added automatically. Enter plain text.")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 3 — Files & Bundle Numbers
        # ═════════════════════════════════════════════════════════════════
        s3 = sec("  Filepaths & Bundle Numbers  ")
        v_fn      = tk.StringVar(value=s.get("file_name",     r"data\item\itm000.nri"))
        v_bn      = tk.StringVar(value=s.get("bundle_num",    "0"))
        v_cmtfn   = tk.StringVar(value=s.get("cmt_file_name", r"data\item\itm_illu000.nri"))
        v_cmtbn   = tk.StringVar(value=s.get("cmt_bundle_num","0"))
        v_equipfn = tk.StringVar(value=s.get("equip_file_name"," "))
        lbl_entry(s3, "FileName:",     v_fn,     50, _TOOLTIPS["FileName"])
        lbl_entry(s3, "BundleNum:",    v_bn,      8, _TOOLTIPS["BundleNum"])
        lbl_note(s3, "  InvFileName / InvBundleNum are auto-copied from FileName / BundleNum above.")
        lbl_entry(s3, "CmtFileName:",  v_cmtfn,  50, _TOOLTIPS["CmtFileName"])
        lbl_entry(s3, "CmtBundleNum:", v_cmtbn,   8, _TOOLTIPS["CmtBundleNum"])
        lbl_entry(s3, "EquipFileName:", v_equipfn, 50, _TOOLTIPS["EquipFileName"])
        lbl_note(s3, "  EquipFileName: leave as a single space if not equipment or drill.")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 4 — PivotID & PaletteId (after files, before Options)
        # ═════════════════════════════════════════════════════════════════
        s_pivot = sec("  PivotID & PaletteId  (Suggested 0) ")
        v_pivot   = tk.StringVar(value=str(s.get("pivot_id",  0)))
        v_palette = tk.StringVar(value=str(s.get("palette_id", 0)))
        lbl_entry(s_pivot, "PivotID:",   v_pivot,   10, _TOOLTIPS["PivotID"])
        lbl_entry(s_pivot, "PaletteId:", v_palette, 10, _TOOLTIPS["PaletteId"])
        lbl_note(s_pivot, "  PivotID: source item ID for equipment variants. "
                          "PaletteId: leave at 0 for most items.")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 5 — Options  (checkboxes in dropdown mode, raw text in manual)
        # ═════════════════════════════════════════════════════════════════
        s4 = sec("  Options  (eItemOption bitmask — flags OR together)")

        # Saved options_flags list OR decode from saved bitmask int
        saved_flags_list = s.get("options_flags", [])
        saved_opts_raw   = s.get("options_raw_manual", "0")
        # If list shorter than full table, pad with False
        while len(saved_flags_list) < len(_OPTIONS_FULL):
            saved_flags_list.append(False)

        opts_vars = [tk.BooleanVar(value=bool(saved_flags_list[i]))
                     for i in range(len(_OPTIONS_FULL))]

        v_opts_manual = tk.StringVar(value=saved_opts_raw)

        # Dropdown mode: checkboxes + bitmask preview
        dd_opts_frm = tk.Frame(s4, bg=BG)
        opt_chk_frm = tk.Frame(dd_opts_frm, bg=BG); opt_chk_frm.pack(fill="x", padx=8, pady=4)
        for i, (fval, flbl) in enumerate(_OPTIONS_FULL):
            cb = tk.Checkbutton(opt_chk_frm, text=f"{flbl}  ({fval})",
                                variable=opts_vars[i],
                                bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                                font=("Consolas", 8))
            cb.grid(row=i // 3, column=i % 3, sticky="w", padx=6, pady=1)
            _attach_tooltip(cb, f"Flag value: {fval}  —  {flbl}")
        opt_preview = tk.Label(dd_opts_frm, text="Bitmask: 0", bg=BG, fg=GREEN,
                               font=("Consolas", 9))
        opt_preview.pack(anchor="w", padx=10, pady=(2, 4))

        def _update_opts_preview(*_):
            total = sum(f for (f, _), v in zip(_OPTIONS_FULL, opts_vars) if v.get())
            opt_preview.config(text=f"Bitmask: {total}")
            v_opts_manual.set(str(total))   # keep manual in sync

        for v in opts_vars: v.trace_add("write", _update_opts_preview)
        _update_opts_preview()

        # Manual mode: raw text entry
        man_opts_frm = tk.Frame(s4, bg=BG)
        r_mo = tk.Frame(man_opts_frm, bg=BG); r_mo.pack(fill="x", padx=8, pady=4)
        tk.Label(r_mo, text="Options (raw int):", width=22, anchor="w",
                 bg=BG, fg=FG, font=("Consolas", 9)).pack(side="left")
        man_opts_ent = tk.Entry(r_mo, textvariable=v_opts_manual, width=14,
                                bg=BG3, fg=FG, insertbackground=FG,
                                font=("Consolas", 9), relief="flat")
        man_opts_ent.pack(side="left", padx=4)
        _attach_tooltip(man_opts_ent, _TOOLTIPS["Options"])
        # Sync manual -> checkboxes
        def _manual_opts_to_chk(*_):
            try: total = int(v_opts_manual.get())
            except: return
            for (fval, _), v in zip(_OPTIONS_FULL, opts_vars):
                v.set(bool(total & fval))
        v_opts_manual.trace_add("write", _manual_opts_to_chk)

        # ═════════════════════════════════════════════════════════════════
        # SECTION 6 — OptionsEx
        # ═════════════════════════════════════════════════════════════════
        s5 = sec("  OptionsEx  (flags OR together — unknown values noted)")
        saved_optex = s.get("options_ex", 0)
        v_optex_raw = tk.StringVar(value=str(saved_optex))
        v_optex_manual = tk.StringVar(value=str(saved_optex))

        optex_vars = {}
        dd_optex_frm = tk.Frame(s5, bg=BG)
        optex_chk_frm = tk.Frame(dd_optex_frm, bg=BG); optex_chk_frm.pack(fill="x", padx=8, pady=4)
        for i, (fval, fdesc) in enumerate(_OPTIONSEX_MAP):
            if fval == 0: continue
            v = tk.BooleanVar()
            cb = tk.Checkbutton(optex_chk_frm, text=f"{fval} — {fdesc}", variable=v,
                                bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                                font=("Consolas", 8))
            cb.grid(row=(i - 1) // 2, column=(i - 1) % 2, sticky="w", padx=6, pady=1)
            _attach_tooltip(cb, f"Flag: {fval}  —  {fdesc}")
            optex_vars[fval] = v
        r_optex = tk.Frame(dd_optex_frm, bg=BG); r_optex.pack(fill="x", padx=8, pady=(2, 4))
        tk.Label(r_optex, text="Bitmask preview:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 9)).pack(side="left")
        optex_preview = tk.Label(r_optex, text="0", bg=BG, fg=GREEN, font=("Consolas", 9))
        optex_preview.pack(side="left", padx=6)

        def _optex_chk_to_raw(*_):
            total = sum(f for f, v in optex_vars.items() if v.get())
            v_optex_raw.set(str(total))
            v_optex_manual.set(str(total))
            optex_preview.config(text=str(total))
        def _optex_raw_to_chk(*_):
            try: total = int(v_optex_raw.get())
            except: total = 0
            for f, v in optex_vars.items(): v.set(bool(total & f))
            optex_preview.config(text=str(total))

        for v in optex_vars.values(): v.trace_add("write", _optex_chk_to_raw)
        v_optex_raw.trace_add("write", _optex_raw_to_chk)
        _optex_raw_to_chk()

        # Manual mode: raw entry
        man_optex_frm = tk.Frame(s5, bg=BG)
        r_moe = tk.Frame(man_optex_frm, bg=BG); r_moe.pack(fill="x", padx=8, pady=4)
        tk.Label(r_moe, text="OptionsEx (raw int):", width=22, anchor="w",
                 bg=BG, fg=FG, font=("Consolas", 9)).pack(side="left")
        man_optex_ent = tk.Entry(r_moe, textvariable=v_optex_manual, width=14,
                                 bg=BG3, fg=FG, insertbackground=FG,
                                 font=("Consolas", 9), relief="flat")
        man_optex_ent.pack(side="left", padx=4)
        _attach_tooltip(man_optex_ent, _TOOLTIPS["OptionsEx"])
        def _man_optex_to_raw(*_):
            v_optex_raw.set(v_optex_manual.get())
        v_optex_manual.trace_add("write", _man_optex_to_raw)

        # ═════════════════════════════════════════════════════════════════
        # SECTION 7 — HideHat  (same picker as ChrTypeFlags)
        # ═════════════════════════════════════════════════════════════════
        s_hh = sec("  HideHat  (per-character ear hide flag — Suggested 0)")
        v_hide_hat = tk.StringVar(value=str(s.get("hide_hat", 0)))
        lbl_note(s_hh,
                 "  Defines which character model hides ears when this item is equipped.\n"
                 "  Uses the same flag values as ChrTypeFlags.  Usually left at 0.")

        hh_sel_frm = tk.Frame(s_hh, bg=BG); hh_sel_frm.pack(fill="x", padx=8, pady=4)
        tk.Label(hh_sel_frm, text="Character:", bg=BG, fg=FG,
                 font=("Consolas", 9)).pack(side="left")
        hh_name_dd = ttk.Combobox(hh_sel_frm, values=CHR_NAMES, state="readonly",
                                  width=12, font=("Consolas", 9))
        hh_name_dd.pack(side="left", padx=6)
        tk.Label(hh_sel_frm, text="Job:", bg=BG, fg=FG,
                 font=("Consolas", 9)).pack(side="left")
        hh_job_dd = ttk.Combobox(hh_sel_frm, values=["1st", "2nd", "3rd"],
                                 state="readonly", width=6, font=("Consolas", 9))
        hh_job_dd.pack(side="left", padx=6)
        hh_add_btn = mk_btn(hh_sel_frm, "+", None, color=GREEN, fg=BG2,
                            font=("Consolas", 11, "bold"), width=3)
        hh_add_btn.pack(side="left", padx=2)
        hh_rem_btn = mk_btn(hh_sel_frm, "−", None, color=ACC3, fg=BG2,
                            font=("Consolas", 11, "bold"), width=3)
        hh_rem_btn.pack(side="left", padx=2)
        hh_lb_frm = tk.Frame(s_hh, bg=BG); hh_lb_frm.pack(fill="x", padx=8, pady=(0, 4))
        tk.Label(hh_lb_frm, text="Added:", bg=BG, fg=FG_GREY,
                 font=("Consolas", 8)).pack(anchor="w")
        hh_lb = tk.Listbox(hh_lb_frm, height=3, width=36, bg=BG3, fg=FG,
                           font=("Consolas", 9), selectbackground=BG4, activestyle="none")
        hh_lb.pack(anchor="w")
        r_hh = tk.Frame(s_hh, bg=BG); r_hh.pack(fill="x", padx=8, pady=(0, 4))
        tk.Label(r_hh, text="Raw bitmask:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 9)).pack(side="left")
        hh_raw_ent = tk.Entry(r_hh, textvariable=v_hide_hat, width=12,
                              bg=BG3, fg=FG, insertbackground=FG,
                              font=("Consolas", 9), relief="flat")
        hh_raw_ent.pack(side="left", padx=6)
        _attach_tooltip(hh_raw_ent, _TOOLTIPS["HideHat"])
        hh_selected = []

        def _refresh_hh():
            hh_lb.delete(0, "end")
            for val in hh_selected:
                hh_lb.insert("end", CHR_FLAG_REVERSE.get(val, str(val)))
            v_hide_hat.set(str(sum(hh_selected)))

        def _add_hh():
            name = hh_name_dd.get(); job = hh_job_dd.get()
            if not name or not job: return
            val = CHR_FLAG_MAP.get(f"{name} {job}")
            if val and val not in hh_selected:
                hh_selected.append(val); _refresh_hh()

        def _rem_hh():
            sel = hh_lb.curselection()
            if sel: hh_selected.pop(sel[0]); _refresh_hh()

        hh_add_btn.config(command=_add_hh)
        hh_rem_btn.config(command=_rem_hh)
        # Init from saved value
        saved_hh = s.get("hide_hat", 0)
        try: saved_hh_int = int(saved_hh)
        except: saved_hh_int = 0
        for v_hh in sorted(CHR_FLAG_MAP.values()):
            if saved_hh_int & v_hh: hh_selected.append(v_hh)
        _refresh_hh()

        # ═════════════════════════════════════════════════════════════════
        # SECTION 8 — ChrTypeFlags
        # ═════════════════════════════════════════════════════════════════
        s6 = sec("  ChrTypeFlags  (per-character access / ruleset flags)")
        v_chr_raw = tk.StringVar(value=str(s.get("chr_type_flags", 0)))
        lbl_note(s6,
                 "  Sum of per-character per-job flags. "
                 "0 = all characters allowed (no restriction).  Suggested 0 unless restricting.")

        chr_sel_frm = tk.Frame(s6, bg=BG); chr_sel_frm.pack(fill="x", padx=8, pady=4)
        tk.Label(chr_sel_frm, text="Character:", bg=BG, fg=FG,
                 font=("Consolas", 9)).pack(side="left")
        chr_name_dd = ttk.Combobox(chr_sel_frm, values=CHR_NAMES, state="readonly",
                                   width=12, font=("Consolas", 9))
        chr_name_dd.pack(side="left", padx=6)
        tk.Label(chr_sel_frm, text="Job:", bg=BG, fg=FG,
                 font=("Consolas", 9)).pack(side="left")
        chr_job_dd = ttk.Combobox(chr_sel_frm, values=["1st", "2nd", "3rd"],
                                  state="readonly", width=6, font=("Consolas", 9))
        chr_job_dd.pack(side="left", padx=6)
        chr_add_btn = mk_btn(chr_sel_frm, "+", None, color=GREEN, fg=BG2,
                             font=("Consolas", 11, "bold"), width=3)
        chr_add_btn.pack(side="left", padx=2)
        chr_rem_btn = mk_btn(chr_sel_frm, "−", None, color=ACC3, fg=BG2,
                             font=("Consolas", 11, "bold"), width=3)
        chr_rem_btn.pack(side="left", padx=2)
        chr_lb_frm = tk.Frame(s6, bg=BG); chr_lb_frm.pack(fill="x", padx=8, pady=(0, 4))
        tk.Label(chr_lb_frm, text="Added:", bg=BG, fg=FG_GREY,
                 font=("Consolas", 8)).pack(anchor="w")
        chr_lb = tk.Listbox(chr_lb_frm, height=4, width=36, bg=BG3, fg=FG,
                            font=("Consolas", 9), selectbackground=BG4, activestyle="none")
        chr_lb.pack(anchor="w")
        r_chr = tk.Frame(s6, bg=BG); r_chr.pack(fill="x", padx=8, pady=(0, 4))
        tk.Label(r_chr, text="Raw bitmask:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 9)).pack(side="left")
        chr_raw_ent = tk.Entry(r_chr, textvariable=v_chr_raw, width=12,
                               bg=BG3, fg=FG, insertbackground=FG,
                               font=("Consolas", 9), relief="flat")
        chr_raw_ent.pack(side="left", padx=6)
        _attach_tooltip(chr_raw_ent, _TOOLTIPS["ChrTypeFlags"])
        chr_selected6 = []

        def _refresh_chr6():
            chr_lb.delete(0, "end")
            for val in chr_selected6:
                chr_lb.insert("end", CHR_FLAG_REVERSE.get(val, str(val)))
            v_chr_raw.set(str(sum(chr_selected6)))

        def _add_chr6():
            name = chr_name_dd.get(); job = chr_job_dd.get()
            if not name or not job: return
            val = CHR_FLAG_MAP.get(f"{name} {job}")
            if val and val not in chr_selected6:
                chr_selected6.append(val); _refresh_chr6()

        def _rem_chr6():
            sel = chr_lb.curselection()
            if sel: chr_selected6.pop(sel[0]); _refresh_chr6()

        chr_add_btn.config(command=_add_chr6)
        chr_rem_btn.config(command=_rem_chr6)
        saved_chr = s.get("chr_type_flags", 0)
        try: saved_chr_int = int(saved_chr)
        except: saved_chr_int = 0
        for v_c in sorted(CHR_FLAG_MAP.values()):
            if saved_chr_int & v_c: chr_selected6.append(v_c)
        _refresh_chr6()

        # ═════════════════════════════════════════════════════════════════
        # SECTION 9 — GroundFlags / SystemFlags  (always 0)
        # ═════════════════════════════════════════════════════════════════
        s7 = sec("  GroundFlags & SystemFlags  (Suggested 0 — warn if changed)")
        v_ground = tk.StringVar(value=str(s.get("ground_flags", 0)))
        v_system = tk.StringVar(value=str(s.get("system_flags", 0)))

        def _attach_warn_nonzero(entry_widget, fname):
            def _cb(e):
                val = entry_widget.get().strip()
                if val not in ("", "0"):
                    messagebox.showwarning("Warning",
                        f"{fname} is not 0.\nThis could have unintended consequences.\n"
                        "Continue at your own risk.")
            entry_widget.bind("<FocusOut>", _cb)

        gnd_ent = lbl_entry(s7, "GroundFlags:  (S0)", v_ground, 10, _TOOLTIPS["GroundFlags"])
        sys_ent = lbl_entry(s7, "SystemFlags:  (S0)", v_system, 10, _TOOLTIPS["SystemFlags"])
        lbl_note(s7, "  These are always 0 in standard items. A warning will appear if you change them.")
        _attach_warn_nonzero(gnd_ent, "GroundFlags")
        _attach_warn_nonzero(sys_ent, "SystemFlags")

        v_exist_type = tk.StringVar(value=str(s.get("exist_type", 0)))
        lbl_entry(s7, "ExistType:", v_exist_type, 10, _TOOLTIPS["ExistType"])
        lbl_note(s7, "  ExistType: 0=disabled, 1=timer/cannot stack (sprints, boosters).")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 10 — Numeric Stats
        # ═════════════════════════════════════════════════════════════════
        s8 = sec("  Numeric Stats  ")
        _num_fields = [
            ("Weight:",   "weight",    "1",  _TOOLTIPS["Weight"]),
            ("Value:",    "value",     "0",  _TOOLTIPS["Value"]),
            ("MinLevel:", "min_level", "1",  _TOOLTIPS["MinLevel"]),
            ("Money:",    "money",     "0",  _TOOLTIPS["Money"]),
            ("Ncash:",    "ncash",     "0",  _TOOLTIPS["Ncash"]),
        ]
        _stat_fields = [
            ("AP:",      "ap",      _TOOLTIPS["AP"]),
            ("HP:",      "hp",      _TOOLTIPS["HP"]),
            ("HPCon:",   "hpcon",   _TOOLTIPS["HPCon"]),
            ("MP:",      "mp",      _TOOLTIPS["MP"]),
            ("MPCon:",   "mpcon",   _TOOLTIPS["MPCon"]),
            ("APPlus:",  "applus",  _TOOLTIPS["APPlus"]),
            ("ACPlus:",  "acplus",  _TOOLTIPS["ACPlus"]),
            ("DXPlus:",  "dxplus",  _TOOLTIPS["DXPlus"]),
            ("MaxMPPlus:","maxmpplus",_TOOLTIPS["MaxMPPlus"]),
            ("MAPlus:",  "maplus",  _TOOLTIPS["MAPlus"]),
            ("MDPlus:",  "mdplus",  _TOOLTIPS["MDPlus"]),
            ("MaxWTPlus:","maxwtplus",_TOOLTIPS["MaxWTPlus"]),
            ("DAPlus:",  "daplus",  _TOOLTIPS["DAPlus"]),
            ("LKPlus:",  "lkplus",  _TOOLTIPS["LKPlus"]),
            ("MaxHPPlus:","maxhpplus",_TOOLTIPS["MaxHPPlus"]),
            ("DPPlus:",  "dpplus",  _TOOLTIPS["DPPlus"]),
            ("HVPlus:",  "hvplus",  _TOOLTIPS["HVPlus"]),
        ]
        num_vars = {}
        nr = tk.Frame(s8, bg=BG); nr.pack(fill="x", padx=8, pady=4)
        for ci, (lbl_t, key, dflt, tip) in enumerate(_num_fields):
            v = tk.StringVar(value=str(s.get(key, dflt)))
            num_vars[key] = v
            lw2 = tk.Label(nr, text=lbl_t, bg=BG, fg=FG, font=("Consolas", 9),
                           width=10, anchor="w")
            lw2.grid(row=0, column=ci * 2, padx=3)
            ent2 = tk.Entry(nr, textvariable=v, width=10, bg=BG3, fg=FG,
                            insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent2.grid(row=0, column=ci * 2 + 1, padx=3)
            if tip: _attach_tooltip(lw2, tip); _attach_tooltip(ent2, tip)

        stat_frm = tk.Frame(s8, bg=BG); stat_frm.pack(fill="x", padx=8, pady=4)
        for i, (lbl_t, key, tip) in enumerate(_stat_fields):
            v = tk.StringVar(value=str(s.get(key, "0")))
            num_vars[key] = v
            lw2 = tk.Label(stat_frm, text=lbl_t, bg=BG, fg=FG, font=("Consolas", 9),
                           width=10, anchor="w")
            lw2.grid(row=i // 4, column=(i % 4) * 2, padx=3, pady=1)
            ent2 = tk.Entry(stat_frm, textvariable=v, width=8, bg=BG3, fg=FG,
                            insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent2.grid(row=i // 4, column=(i % 4) * 2 + 1, padx=3, pady=1)
            if tip: _attach_tooltip(lw2, tip); _attach_tooltip(ent2, tip)

        f_frm = tk.Frame(s8, bg=BG); f_frm.pack(fill="x", padx=8, pady=4)
        for ci, (lbl_t, key, tip) in enumerate([
            ("HPRecovery:", "hprecoveryrate", _TOOLTIPS["HPRecoveryRate"]),
            ("MPRecovery:", "mprecoveryrate", _TOOLTIPS["MPRecoveryRate"]),
            ("Delay:",      "delay",          _TOOLTIPS["Delay"]),
            ("CardGenParam:","cardgenparam",  _TOOLTIPS["CardGenParam"]),
        ]):
            v = tk.StringVar(value=str(s.get(key, "0")))
            num_vars[key] = v
            lw2 = tk.Label(f_frm, text=lbl_t, bg=BG, fg=FG_DIM, font=("Consolas", 9),
                           width=12, anchor="w")
            lw2.grid(row=0, column=ci * 2, padx=3)
            ent2 = tk.Entry(f_frm, textvariable=v, width=12, bg=BG3, fg=FG,
                            insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent2.grid(row=0, column=ci * 2 + 1, padx=3)
            if tip: _attach_tooltip(lw2, tip); _attach_tooltip(ent2, tip)
        lbl_note(s8,
                 "  Float fields (HPRecovery, MPRecovery, Delay, CardGenParam) "
                 "auto-formatted to 6 decimal places.")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 11 — Effect  (multi-select checkboxes + raw)
        # ═════════════════════════════════════════════════════════════════
        s_eff = sec("  Effect  (item use result — multiple values OR together)")
        lbl_note(s_eff,
                 "  For boxes: MUST include 22 (Open Box).  "
                 "Multiple effects can be combined.  Suggested 0 for non-use items.")

        # Save as comma-separated raw string internally; parse on load
        saved_effect_raw = str(s.get("effect", 0))
        # Parse saved value into a set of ints
        _eff_active = set()
        for _ep in str(saved_effect_raw).replace("/", ",").split(","):
            try: _eff_active.add(int(_ep.strip()))
            except: pass

        eff_vars = {}
        eff_chk_frm = tk.Frame(s_eff, bg=BG); eff_chk_frm.pack(fill="x", padx=8, pady=4)
        for i, (fval, fdesc) in enumerate(_EFFECT_MAP):
            v = tk.BooleanVar(value=(fval in _eff_active))
            cb = tk.Checkbutton(eff_chk_frm,
                                text=f"{fval} — {fdesc}",
                                variable=v, bg=BG, fg=FG, selectcolor=BG3,
                                activebackground=BG, font=("Consolas", 8))
            cb.grid(row=i // 3, column=i % 3, sticky="w", padx=6, pady=1)
            _attach_tooltip(cb, f"Effect {fval}: {fdesc}")
            eff_vars[fval] = v

        eff_preview = tk.Label(s_eff, text="Value: 0", bg=BG, fg=GREEN,
                               font=("Consolas", 9))
        eff_preview.pack(anchor="w", padx=10)

        v_effect_raw = tk.StringVar(value=saved_effect_raw)

        r_eff = tk.Frame(s_eff, bg=BG); r_eff.pack(fill="x", padx=8, pady=(0, 4))
        tk.Label(r_eff, text="Raw value:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 9)).pack(side="left")
        eff_raw_ent = tk.Entry(r_eff, textvariable=v_effect_raw, width=14,
                               bg=BG3, fg=FG, insertbackground=FG,
                               font=("Consolas", 9), relief="flat")
        eff_raw_ent.pack(side="left", padx=6)
        _attach_tooltip(eff_raw_ent, _TOOLTIPS["Effect"])

        def _eff_chk_to_raw(*_):
            vals = sorted(f for f, v in eff_vars.items() if v.get())
            if not vals or vals == [0]:
                raw = "0"
            else:
                raw = "/".join(str(x) for x in vals if x != 0) or "0"
            v_effect_raw.set(raw)
            eff_preview.config(text="Value: " + raw)

        def _eff_raw_to_chk(*_):
            raw = v_effect_raw.get()
            active = set()
            for p in raw.replace("/", ",").split(","):
                try: active.add(int(p.strip()))
                except: pass
            for fval, v in eff_vars.items():
                v.set(fval in active)
            eff_preview.config(text="Value: " + raw)

        for v in eff_vars.values(): v.trace_add("write", _eff_chk_to_raw)
        v_effect_raw.trace_add("write", _eff_raw_to_chk)
        _eff_raw_to_chk()

        # Manual mode effect: raw entry only (no checkboxes)
        man_eff_frm = tk.Frame(s_eff, bg=BG)
        r_me = tk.Frame(man_eff_frm, bg=BG); r_me.pack(fill="x", padx=8, pady=4)
        tk.Label(r_me, text="Effect (raw):", width=22, anchor="w",
                 bg=BG, fg=FG, font=("Consolas", 9)).pack(side="left")
        tk.Entry(r_me, textvariable=v_effect_raw, width=14, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)

        # ═════════════════════════════════════════════════════════════════
        # SECTION 12 — Misc single-value fields (suggested 0 noted)
        # ═════════════════════════════════════════════════════════════════
        s9 = sec("  Misc Fields  (EffectFlags2, SelRange, Life, Depth, Delay — Suggested 0 unless needed)")
        v_eff2     = tk.StringVar(value=str(s.get("effect_flags2", 0)))
        v_sel_range= tk.StringVar(value=str(s.get("sel_range", 0)))
        v_life     = tk.StringVar(value=str(s.get("life", 0)))
        v_depth    = tk.StringVar(value=str(s.get("depth", 0)))
        lbl_entry(s9, "EffectFlags2: (S0)", v_eff2,      10, _TOOLTIPS["EffectFlags2"])
        lbl_entry(s9, "SelRange:     (S0)", v_sel_range, 10, _TOOLTIPS["SelRange"])
        lbl_entry(s9, "Life:",             v_life,       10, _TOOLTIPS["Life"])
        lbl_entry(s9, "Depth:        (S0)", v_depth,     10, _TOOLTIPS["Depth"])

        # ═════════════════════════════════════════════════════════════════
        # SECTION 13 — Card params
        # ═════════════════════════════════════════════════════════════════
        s10 = sec("  Card Parameters  (Suggested 0 for non-card items)")
        v_cardnum  = tk.StringVar(value=str(s.get("cardnum",     0)))
        v_cardgrade= tk.StringVar(value=str(s.get("cardgengrade",0)))
        v_daily    = tk.StringVar(value=str(s.get("dailygencnt", 0)))
        lbl_entry(s10, "CardNum:",       v_cardnum,   8, _TOOLTIPS["CardNum"])
        lbl_entry(s10, "CardGenGrade:",  v_cardgrade, 8, _TOOLTIPS["CardGenGrade"])
        lbl_entry(s10, "DailyGenCnt: (S0)", v_daily,  8, _TOOLTIPS["DailyGenCnt"])
        lbl_note(s10, "  Skill cards: CardNum=7, CardGenGrade=0, CardGenParam=0.000000")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 14 — Refine / Equipment / Gear params
        # ═════════════════════════════════════════════════════════════════
        s11 = sec("  Refine & Equipment Params  ")
        v_refine_idx  = tk.StringVar(value=str(s.get("refine_index",  0)))
        v_refine_type = tk.StringVar(value=str(s.get("refine_type",   0)))
        v_minstattype = tk.StringVar(value=str(s.get("min_stat_type", 0)))
        v_minstatLv   = tk.StringVar(value=str(s.get("min_stat_lv",   0)))
        v_compound    = tk.StringVar(value=str(s.get("compound_slot", 0)))
        v_setitem     = tk.StringVar(value=str(s.get("set_item_id",   0)))
        v_reform      = tk.StringVar(value=str(s.get("reform_count",  0)))

        def _dd_row_ref(parent, label, mapping, raw_var, tip_key=""):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            tip = _TOOLTIPS.get(tip_key, "")
            lw2 = tk.Label(r, text=label, width=22, anchor="w", bg=BG, fg=FG,
                           font=("Consolas", 9)); lw2.pack(side="left")
            vals_dd = [f"{v} — {d}" for v, d in mapping]
            combo2 = ttk.Combobox(r, values=vals_dd, state="readonly",
                                  width=34, font=("Consolas", 9))
            combo2.pack(side="left", padx=4)
            raw_ent2 = tk.Entry(r, textvariable=raw_var, width=7, bg=BG3, fg=FG,
                                insertbackground=FG, font=("Consolas", 9), relief="flat")
            raw_ent2.pack(side="left", padx=(0, 4))
            def _sel2(e):
                sel2 = combo2.current()
                if sel2 >= 0: raw_var.set(str(mapping[sel2][0]))
            combo2.bind("<<ComboboxSelected>>", _sel2)
            def _raw2(*_):
                val2 = raw_var.get().strip()
                for i2, (v2, _) in enumerate(mapping):
                    if str(v2) == val2: combo2.set(vals_dd[i2]); return
                combo2.set("")
            raw_var.trace_add("write", _raw2); _raw2()
            if tip: _attach_tooltip(lw2, tip); _attach_tooltip(combo2, tip); _attach_tooltip(raw_ent2, tip)

        _dd_row_ref(s11, "RefineIndex:",   _REFINEINDEX_MAP, v_refine_idx,  "RefineIndex")
        _dd_row_ref(s11, "RefineType:",    _REFINETYPE_MAP,  v_refine_type, "RefineType")
        _dd_row_ref(s11, "MinStatType:",   _MINSTATTYPE_MAP, v_minstattype, "MinStatType")
        lbl_entry(s11, "MinStatLv:",       v_minstatLv, 8, _TOOLTIPS["MinStatLv"])
        lbl_entry(s11, "CompoundSlot:",    v_compound,  8, _TOOLTIPS["CompoundSlot"])
        lbl_entry(s11, "SetItemID:   (S0)",v_setitem,   8, _TOOLTIPS["SetItemID"])
        lbl_entry(s11, "ReformCount: (S0)",v_reform,    8, _TOOLTIPS["ReformCount"])
        lbl_note(s11, "  CompoundSlot: intended range 0-5. Higher values cause UI errors on item inspect.")

        # ═════════════════════════════════════════════════════════════════
        # SECTION 15 — CM / Shop / Part / Summary  (suggested 0/blank)
        # ═════════════════════════════════════════════════════════════════
        s12 = sec("  CM / Shop / PartFileName  (Suggested 0 or blank)")
        v_shopfn  = tk.StringVar(value=s.get("shop_file_name",  " "))
        v_shopbn  = tk.StringVar(value=str(s.get("shop_bundle_num", 0)))
        v_partfn  = tk.StringVar(value=s.get("part_file_name",  " "))
        v_group   = tk.StringVar(value=str(s.get("group_id",    0)))
        lbl_entry(s12, "ShopFileName:  (S0)", v_shopfn, 50, _TOOLTIPS["ShopFileName"])
        lbl_entry(s12, "ShopBundleNum: (S0)", v_shopbn,  8, _TOOLTIPS["ShopBundleNum"])
        lbl_entry(s12, "PartFileName:  (S0)", v_partfn, 50, _TOOLTIPS["PartFileName"])
        lbl_entry(s12, "GroupId:       (S0)", v_group,  10, _TOOLTIPS["GroupId"])
        lbl_note(s12,
                 "  ShopFileName: promotional Cash Mall image.\n"
                 "  PartFileName: ItemParamCM2 only — fashion item model path.\n"
                 "  ChrFTypeFlag, ChrGender, NewCM, FamCM, Summary always 0/blank — not shown.")

        # ─── Show/hide dropdown vs manual sections ────────────────────────
        def _apply_mode(*_):
            dd = mode_var.get() == "dropdown"
            # Options
            if dd:
                dd_opts_frm.pack(fill="x"); man_opts_frm.pack_forget()
            else:
                man_opts_frm.pack(fill="x"); dd_opts_frm.pack_forget()
            # OptionsEx
            if dd:
                dd_optex_frm.pack(fill="x"); man_optex_frm.pack_forget()
            else:
                man_optex_frm.pack(fill="x"); dd_optex_frm.pack_forget()
            # Effect
            if dd:
                eff_chk_frm.pack(fill="x")
                eff_preview.pack(anchor="w", padx=10)
                r_eff.pack(fill="x", padx=8, pady=(0,4))
                man_eff_frm.pack_forget()
            else:
                eff_chk_frm.pack_forget()
                eff_preview.pack_forget()
                r_eff.pack_forget()
                man_eff_frm.pack(fill="x")

        mode_var.trace_add("write", _apply_mode)
        _apply_mode()

        # ═════════════════════════════════════════════════════════════════
        # GATHER
        # ═════════════════════════════════════════════════════════════════
        def _gather():
            return {
                "id":            v_id.get().strip(),
                "class_val":     v_class.get().strip(),
                "type_val":      v_type.get().strip(),
                "subtype_val":   v_sub.get().strip(),
                "itemftype_val": v_ift.get().strip(),
                "name":          v_name.get(),
                "comment":       v_comment.get(),
                "use":           v_use.get(),
                "name_eng":      v_name_eng.get() or " ",
                "comment_eng":   v_cmt_eng.get() or " ",
                "file_name":     v_fn.get(),
                "bundle_num":    v_bn.get(),
                "cmt_file_name": v_cmtfn.get(),
                "cmt_bundle_num":v_cmtbn.get(),
                "equip_file_name": v_equipfn.get() or " ",
                "pivot_id":      v_pivot.get(),
                "palette_id":    v_palette.get(),
                "options_flags": [v.get() for v in opts_vars],
                "options_raw_manual": v_opts_manual.get(),
                "options_ex":    int(v_optex_raw.get() or 0),
                "hide_hat":      v_hide_hat.get(),
                "chr_type_flags":int(v_chr_raw.get() or 0),
                "ground_flags":  v_ground.get(),
                "system_flags":  v_system.get(),
                "exist_type":    v_exist_type.get(),
                "weight":        num_vars["weight"].get(),
                "value":         num_vars["value"].get(),
                "min_level":     num_vars["min_level"].get(),
                "money":         num_vars["money"].get(),
                "ncash":         num_vars["ncash"].get(),
                "ap":            num_vars["ap"].get(),
                "hp":            num_vars["hp"].get(),
                "hpcon":         num_vars["hpcon"].get(),
                "mp":            num_vars["mp"].get(),
                "mpcon":         num_vars["mpcon"].get(),
                "applus":        num_vars["applus"].get(),
                "acplus":        num_vars["acplus"].get(),
                "dxplus":        num_vars["dxplus"].get(),
                "maxmpplus":     num_vars["maxmpplus"].get(),
                "maplus":        num_vars["maplus"].get(),
                "mdplus":        num_vars["mdplus"].get(),
                "maxwtplus":     num_vars["maxwtplus"].get(),
                "daplus":        num_vars["daplus"].get(),
                "lkplus":        num_vars["lkplus"].get(),
                "maxhpplus":     num_vars["maxhpplus"].get(),
                "dpplus":        num_vars["dpplus"].get(),
                "hvplus":        num_vars["hvplus"].get(),
                "hprecoveryrate": num_vars["hprecoveryrate"].get(),
                "mprecoveryrate": num_vars["mprecoveryrate"].get(),
                "delay":          num_vars["delay"].get(),
                "cardgenparam":   num_vars["cardgenparam"].get(),
                "effect":         v_effect_raw.get(),
                "effect_flags2":  v_eff2.get(),
                "sel_range":      v_sel_range.get(),
                "life":           v_life.get(),
                "depth":          v_depth.get(),
                "cardnum":        v_cardnum.get(),
                "cardgengrade":   v_cardgrade.get(),
                "dailygencnt":    v_daily.get(),
                "refine_index":   v_refine_idx.get(),
                "refine_type":    v_refine_type.get(),
                "min_stat_type":  v_minstattype.get(),
                "min_stat_lv":    v_minstatLv.get(),
                "compound_slot":  v_compound.get(),
                "set_item_id":    v_setitem.get(),
                "reform_count":   v_reform.get(),
                "shop_file_name": v_shopfn.get() or " ",
                "shop_bundle_num":v_shopbn.get(),
                "part_file_name": v_partfn.get() or " ",
                "group_id":       v_group.get(),
                "input_mode":     mode_var.get(),
            }

        # ═════════════════════════════════════════════════════════════════
        # GENERATE ACTIONS
        # ═════════════════════════════════════════════════════════════════
        def _generate():
            cfg = _gather()
            if not cfg["id"].strip():
                messagebox.showerror("Missing ID", "Please enter an Item ID."); return
            if cfg["class_val"] in ("", "0", "Unselected"):
                if not messagebox.askyesno("Confirm", "Class is Unselected / 0. Continue?"):
                    return
            if cfg["type_val"] in ("", "0", "Unselected"):
                if not messagebox.askyesno("Confirm", "Type is Unselected / 0. Continue?"):
                    return
            xml = build_generic_itemparam_row(cfg)
            cfg["_first_run_done"] = True
            _save_t6_settings(cfg)
            try: _set_last_id("t6_item", int(cfg["id"]))
            except: pass
            self._settings = cfg
            self._first_run = False

            # Box warning / PresentItemParam prompt
            try:  type_val_i = int(cfg.get("type_val", 0))
            except: type_val_i = 0
            try:
                eff_str = str(cfg.get("effect", "0"))
                eff_set = {int(x.strip()) for x in eff_str.replace("/",",").split(",") if x.strip()}
            except: eff_set = set()
            present_xml = None
            if type_val_i == 15:
                if 22 not in eff_set:
                    messagebox.showwarning("Box Effect Warning",
                        "Type 15 (Useables/Boxes) detected but Effect 22 (Open Box) is NOT set.\n"
                        "The box cannot be opened without Effect=22.\n\n"
                        "Go back and add Effect 22, or continue if intentional.")
                want_present = messagebox.askyesno("PresentItemParam2",
                    "This is a box (Type 15).\n"
                    "Generate a PresentItemParam2 row (drop contents) too?")
                if want_present:
                    present_xml = self._ask_present_contents(cfg)
            self._show_output(xml, cfg, present_xml=present_xml)

        def _generate_and_next():
            cfg = _gather()
            if not cfg["id"].strip():
                messagebox.showerror("Missing ID", "Please enter an Item ID."); return
            xml = build_generic_itemparam_row(cfg)
            try:    cfg["id"] = str(int(cfg["id"]) + 1)
            except: pass
            cfg["_first_run_done"] = True
            _save_t6_settings(cfg)
            try: _set_last_id("t6_item", int(cfg["id"]) - 1)
            except: pass
            self._settings = cfg
            self._first_run = False
            self._show_output(xml, cfg, auto_next=True, present_xml=None)

        # ── Footer nav  (Back | Clear | Generate) ────────────────────────
        nav = tk.Frame(wrap, bg=BG2)
        nav.grid(row=2, column=0, sticky="ew")

        mk_btn(nav, "◀  Back", self._build_editor,
               color=BG4).pack(side="left", padx=14, pady=6)
        mk_btn(nav, "🗑  Clear / Reset", self._reset,
               color=BG4).pack(side="left", padx=4, pady=6)
        mk_btn(nav, "➕  Generate & Next ID", _generate_and_next,
               color=GREEN, fg=BG2).pack(side="right", padx=4, pady=6)
        mk_btn(nav, "⚡  Generate XML Row", _generate,
               color=self.ACC, fg=BG2,
               font=("Consolas", 11, "bold")).pack(side="right", padx=14, pady=6)

    # ─────────────────────────────────────────────────────────────────────────
    def _ask_present_contents(self, cfg):
        """Simple dialog to collect item IDs for a PresentItemParam2 row."""
        win = tk.Toplevel(self.root)
        win.title("PresentItemParam2 — Box Contents")
        win.geometry("500x480")
        win.configure(bg=BG)
        win.grab_set()

        tk.Label(win, text="Box Drop Contents",
                 bg=BG, fg=ACC6, font=("Consolas",12,"bold")).pack(pady=(12,2), padx=12, anchor="w")
        tk.Label(win, text="Enter the item IDs that drop from this box (up to 20).",
                 bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(padx=12, anchor="w")

        sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True, padx=8, pady=4)
        canv, C = mk_scroll_canvas(sh, init_width=460)

        rows_frame = tk.Frame(C, bg=BG); rows_frame.pack(fill="x", padx=4, pady=4)
        item_id_vars = []; item_rate_vars = []; item_cnt_vars = []
        for n in range(1, 21):
            r = tk.Frame(rows_frame, bg=BG); r.pack(fill="x", pady=1)
            tk.Label(r, text=f"Drop {n:2d}:", bg=BG, fg=FG,
                     font=("Consolas",9), width=8).pack(side="left")
            vid = tk.StringVar(value="0")
            tk.Entry(r, textvariable=vid, width=10, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=2)
            tk.Label(r, text="Rate:", bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(side="left", padx=(6,0))
            vrate = tk.StringVar(value="50")
            tk.Entry(r, textvariable=vrate, width=6, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=2)
            tk.Label(r, text="Cnt:", bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(side="left", padx=(6,0))
            vcnt = tk.StringVar(value="1")
            tk.Entry(r, textvariable=vcnt, width=4, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=2)
            item_id_vars.append(vid); item_rate_vars.append(vrate); item_cnt_vars.append(vcnt)

        ptype_var = tk.IntVar(value=1)
        pt_frm = tk.Frame(C, bg=BG); pt_frm.pack(fill="x", padx=8, pady=(4,0))
        tk.Label(pt_frm, text="Present Type:", bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        for lbl, val in [("Normal (1)","1"),("Distributive (2)","2")]:
            tk.Radiobutton(pt_frm, text=lbl, variable=ptype_var, value=int(val),
                           bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                           font=("Consolas",9)).pack(side="left", padx=8)

        drop_cnt_var = tk.StringVar(value="1")
        dc_frm = tk.Frame(C, bg=BG); dc_frm.pack(fill="x", padx=8, pady=2)
        tk.Label(dc_frm, text="Drop Cnt:", bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        tk.Entry(dc_frm, textvariable=drop_cnt_var, width=6, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=6)

        result = [None]
        nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")

        def _build():
            items = []
            for vid, vr, vc in zip(item_id_vars, item_rate_vars, item_cnt_vars):
                iid = vid.get().strip()
                if iid and iid != "0":
                    try: rate = int(vr.get())
                    except: rate = 50
                    try: cnt = int(vc.get())
                    except: cnt = 1
                    items.append({"id": iid, "rate": rate, "item_cnt": cnt})
            if not items:
                messagebox.showwarning("No Items", "Add at least one drop item."); return
            try: dc = int(drop_cnt_var.get())
            except: dc = 1
            result[0] = build_presentparam_row(
                cfg["id"], items, ptype_var.get(), dc, 50,
                item_cnts=[it["item_cnt"] for it in items],
                box_name=cfg.get("name",""))
            win.destroy()

        mk_btn(nav, "✓  Build Present Row", _build, color=GREEN, fg=BG2,
               font=("Consolas",10,"bold")).pack(side="left", padx=12, pady=6)
        mk_btn(nav, "Skip", win.destroy).pack(side="left", padx=4, pady=6)
        win.wait_window()
        return result[0]

    def _show_output(self, xml, cfg, auto_next=False,
                     _compound_rows=None, _exchange_rows=None, present_xml=None):
        self._clear()
        if _compound_rows is None: _compound_rows = []
        if _exchange_rows is None: _exchange_rows = []
        ACC = self.ACC

        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0, weight=0)
        wrap.grid_rowconfigure(1, weight=1)
        wrap.grid_rowconfigure(2, weight=0)
        wrap.grid_columnconfigure(0, weight=1)

        hdr = tk.Frame(wrap, bg=BG2)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text=f"✓  Generated: ID {cfg['id']}  —  {cfg.get('name','')}",
                 font=("Consolas", 13, "bold"), bg=BG2, fg=ACC, pady=8
                 ).pack(side="left", padx=15)

        # Warn if Effect != 22 but PresentItemParam is expected (box use)
        try:
            effect_val = int(cfg.get("effect", 0))
        except: effect_val = 0
        if effect_val != 0 and effect_val != 22:
            pass  # non-box item, no warning needed
        # If type is 15 (Useables/Boxes) and effect != 22 warn
        try: type_val = int(cfg.get("type_val", 0))
        except: type_val = 0
        if type_val == 15 and effect_val != 22:
            tk.Label(hdr, text="⚠ Effect should be 22 (Open Box) for Type 15 items!",
                     bg=BG2, fg=ACC3, font=("Consolas",9,"bold")).pack(side="right", padx=10)

        nb = ttk.Notebook(wrap)
        nb.grid(row=1, column=0, sticky="nsew", padx=8, pady=4)
        make_output_tab(nb, "itemparam row", xml, "itemparam_row.xml", self.root)
        if present_xml:
            make_output_tab(nb, "PresentItemParam2 row", present_xml,
                            "presentparam_row.xml", self.root)

        # Compound/exchange extra tabs
        if _compound_rows:
            cp_xml = "\n".join(r[0] for r in _compound_rows)
            cl_xml = "\n".join(r[1] for r in _compound_rows)
            make_output_tab(nb,"Compound_Potion rows",cp_xml,"Compound_Potion_rows.xml",self.root)
            make_output_tab(nb,"Compounder_Location rows",cl_xml,"Compounder_Location_rows.xml",self.root)
        if _exchange_rows:
            es_xml = "\n".join(r[0] for r in _exchange_rows)
            el_xml = "\n".join(r[1] for r in _exchange_rows)
            make_output_tab(nb,"ExchangeShopContents rows",es_xml,"ExchangeShopContents_rows.xml",self.root)
            make_output_tab(nb,"Exchange_Location rows",el_xml,"Exchange_Location_rows.xml",self.root)

        nav = tk.Frame(wrap, bg=BG2)
        nav.grid(row=2, column=0, sticky="ew")

        def _add_ce():
            def _on_compound(ce_cfg):
                cpr = build_compound_row(ce_cfg)
                clr = build_compound_location_row(ce_cfg["compound_id"])
                _compound_rows.append((cpr, clr))
                self._show_output(xml, cfg, _compound_rows=_compound_rows,
                                  _exchange_rows=_exchange_rows)
            def _on_exchange(ce_cfg):
                esr = build_exchange_row(ce_cfg)
                elr = build_exchange_location_row(ce_cfg["exchange_id"])
                _exchange_rows.append((esr, elr))
                self._show_output(xml, cfg, _compound_rows=_compound_rows,
                                  _exchange_rows=_exchange_rows)
            _show_compound_exchange_dialog(
                self.root, cfg.get("name",""), cfg.get("comment",""),
                cfg.get("id",""), _on_compound, _on_exchange, lambda: None)

        def _export_all():
            folder = filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            exports = [("itemparam_row.xml", xml)]
            if present_xml:
                exports.append(("presentparam_row.xml", present_xml))
            if _compound_rows:
                exports += [("Compound_Potion_rows.xml", "\n".join(r[0] for r in _compound_rows)),
                            ("Compounder_Location_rows.xml", "\n".join(r[1] for r in _compound_rows))]
            if _exchange_rows:
                exports += [("ExchangeShopContents_rows.xml", "\n".join(r[0] for r in _exchange_rows)),
                            ("Exchange_Location_rows.xml", "\n".join(r[1] for r in _exchange_rows))]
            saved = []
            for fname, content in exports:
                with open(os.path.join(folder, fname), "w", encoding="utf-8") as f:
                    f.write(content)
                saved.append(fname)
            messagebox.showinfo("Export Complete", f"Saved to:\n{folder}\n\n" + "\n".join(saved))

        mk_btn(nav, "➕  New Item (next ID)", self._next_from_output(cfg),
               color=self.ACC, fg=BG2, font=("Consolas",10,"bold")).pack(
               side="left", padx=14, pady=6)
        mk_btn(nav, "⚗  Add Compound/Exchange", _add_ce, color=BG4).pack(
               side="left", padx=4, pady=6)
        def _import_ce_t6():
            def _on_compound_cfgs(cfgs):
                for ce_cfg in cfgs:
                    cpr = build_compound_row(ce_cfg)
                    clr = build_compound_location_row(ce_cfg["compound_id"])
                    _compound_rows.append((cpr, clr))
                self._show_output(xml, cfg,
                                  _compound_rows=_compound_rows,
                                  _exchange_rows=_exchange_rows)
            def _on_exchange_cfgs(cfgs):
                for ce_cfg in cfgs:
                    esr = build_exchange_row(ce_cfg)
                    elr = build_exchange_location_row(ce_cfg["exchange_id"])
                    _exchange_rows.append((esr, elr))
                self._show_output(xml, cfg,
                                  _compound_rows=_compound_rows,
                                  _exchange_rows=_exchange_rows)
            _ask_import_mode_then_file(self.root, _on_compound_cfgs, _on_exchange_cfgs)
        mk_btn(nav, "📥  Import CSV/Excel", _import_ce_t6, color=BG4).pack(
               side="left", padx=4, pady=6)
        mk_btn(nav, "💾  Export All", _export_all, color=GREEN, fg=BG2).pack(
               side="left", padx=4, pady=6)
        mk_btn(nav, "◀  Back to Edit", self._build_editor).pack(
               side="left", padx=4, pady=6)
        mk_btn(nav, "🗑  Reset All", self._reset).pack(side="right", padx=14, pady=6)
        if auto_next:
            self.after(10, self._build_editor)

    def _next_from_output(self, cfg):
        def _go():
            try: cfg["id"] = str(int(cfg["id"]) + 1)
            except: pass
            try: _set_last_id("t6_item", int(cfg["id"]) - 1)
            except: pass
            _save_t6_settings(cfg)
            self._settings = cfg
            self._build_editor()
        return _go

    def _show_reference(self):
        """Scrollable reference window — all lookup tables in XML field order."""
        win = tk.Toplevel(self.root)
        win.title("ItemParam Field Reference")
        win.geometry("920x700")
        win.configure(bg=BG)
        sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True)
        canv, cont = mk_scroll_canvas(sh, init_width=880)

        def rsec(title, required=False):
            f = tk.Frame(cont, bg=BG2, pady=4)
            f.pack(fill="x", padx=8, pady=(8, 2))
            color = ACC3 if required else BLUE
            tk.Label(f, text=title, bg=BG2, fg=color,
                     font=("Consolas", 10, "bold")).pack(anchor="w", padx=8)
            body = tk.Frame(cont, bg=BG); body.pack(fill="x", padx=12, pady=2)
            return body

        def rtable(parent, rows, cols=None):
            frm = tk.Frame(parent, bg=BG); frm.pack(anchor="w", pady=2, padx=4)
            if cols:
                for ci, c in enumerate(cols):
                    w = max(12, len(str(c)) + 2)
                    tk.Label(frm, text=c, bg=BG2, fg=BLUE,
                             font=("Consolas", 8, "bold"), width=w,
                             anchor="w").grid(row=0, column=ci, padx=2, pady=1, sticky="w")
            for ri, row in enumerate(rows):
                bg = BG if ri % 2 == 0 else BG2
                for ci, cell in enumerate(row):
                    w = max(12, len(str(cell)) + 2)
                    tk.Label(frm, text=str(cell), bg=bg, fg=FG,
                             font=("Consolas", 8), width=w,
                             anchor="w").grid(row=ri + (1 if cols else 0),
                                              column=ci, padx=2, pady=1, sticky="w")

        def rnote(parent, text):
            tk.Label(parent, text=text, bg=BG, fg=FG_GREY,
                     font=("Consolas", 8), justify="left", wraplength=840).pack(
                     anchor="w", padx=4, pady=(1, 4))

        # ── In XML order ─────────────────────────────────────────────────
        s = rsec("ID", required=False)
        rnote(s, "Unique item ID number. Must not duplicate any existing ID in the table.")

        s = rsec("Class  ⚠ REQUIRED — must not be 0", required=True)
        rtable(s, _CLASS_MAP, ["Value","Description"])

        s = rsec("Type  ⚠ REQUIRED — must not be 0", required=True)
        rtable(s, _TYPE_MAP, ["Value","Description"])

        s = rsec("SubType  (default 0)")
        rtable(s, _SUBTYPE_MAP, ["Value","Description"])
        rnote(s, "Values marked > in original docs are from ItemParamCM2.")

        s = rsec("ItemFType  (default 0)")
        rtable(s, _ITEMFTYPE_MAP, ["Value","Description"])

        s = rsec("Name / Comment / Use / Name_Eng / Comment_Eng")
        rnote(s, "All are CDATA text fields.  Name_Eng and Comment_Eng are usually a single space.")

        s = rsec("FileName / BundleNum / InvFileName / InvBundleNum")
        rnote(s,
            "FileName: path to sprite NRI (e.g. data\\item\\itm000.nri).\n"
            "BundleNum: sprite index starting at 0.  InvFileName and InvBundleNum are "
            "always identical to FileName and BundleNum — auto-copied by this tool.\n"
            "To find a sprite: open NRI in viewer → Animations tab → slot number = BundleNum + 1.")

        s = rsec("CmtFileName / CmtBundleNum")
        rnote(s, "Points to the item illustration shown in the tooltip window. "
                 "Different file from FileName. Same BundleNum +1 offset rule.")

        s = rsec("EquipFileName")
        rnote(s, "Path to equipment or drill model. Leave as a single space if not equipment/drill.")

        s = rsec("PivotID  (Suggested 0)")
        rnote(s, "Source item ID reference — mainly used for equipment with multiple level/option variants.")

        s = rsec("PaletteId  (Suggested 0)")
        rnote(s, "Palette ID. 0 for almost all items. Rare exceptions can be ignored.")

        s = rsec("Options  (eItemOption — flags OR together)")
        rtable(s, _OPTIONS_FULL, ["Flag Value", "Description"])
        rnote(s, "Common defaults to always include: 2 (Usable) and 32 (UsableToSelf) for use-items.")

        s = rsec("HideHat  (Suggested 0)")
        rnote(s,
            "Defines per character-type which model hides its ear when this item is equipped.\n"
            "Uses the SAME flag values as ChrTypeFlags (see below).  Usually 0.")
        rtable(s, [
            ("Bunny",   1, 512,    262144),
            ("Buffalo", 2, 1024,   524288),
            ("Sheep",   4, 2048,   1048576),
            ("Dragon",  8, 4096,   2097152),
            ("Fox",    16, 8192,   4194304),
            ("Lion",   32, 16384,  8388608),
            ("Cat",    64, 32768,  16777216),
            ("Raccoon",124,65536,  33554432),
            ("Paula",  256,131072, 67108864),
        ], ["Character", "1st Job Flag", "2nd Job Flag", "3rd Job Flag"])

        s = rsec("ChrTypeFlags  (Suggested 0)")
        rnote(s,
            "Character-type restrictions/events. Sum the flags for every character+job "
            "that should be allowed.  0 = no restriction (all characters can use).")
        rtable(s, [
            ("Bunny",   1, 512,    262144),
            ("Buffalo", 2, 1024,   524288),
            ("Sheep",   4, 2048,   1048576),
            ("Dragon",  8, 4096,   2097152),
            ("Fox",    16, 8192,   4194304),
            ("Lion",   32, 16384,  8388608),
            ("Cat",    64, 32768,  16777216),
            ("Raccoon",124,65536,  33554432),
            ("Paula",  256,131072, 67108864),
        ], ["Character", "1st Job Flag", "2nd Job Flag", "3rd Job Flag"])

        s = rsec("GroundFlags / SystemFlags  (ALWAYS 0 — Suggested 0)")
        rnote(s, "Never changed. Non-zero values may cause unintended behaviour.")

        s = rsec("OptionsEx  (flags OR together — Suggested 0)")
        rtable(s, _OPTIONSEX_MAP, ["Flag Value", "Description"])
        rnote(s, "Values can be combined. Example: 1/2 = random stat range + elemental property.")

        s = rsec("Weight / Value / MinLevel  (defaults: 1, 0, 1)")
        rnote(s, "Weight = WT stat cost.  Value = Galder NPC sell price.  MinLevel = level gate.")

        s = rsec("Effect  (action on use — Suggested 0 for non-use items)")
        rtable(s, _EFFECT_MAP, ["Value", "Description"])
        rnote(s, "⚠ Type 15 (Useables/Boxes) MUST have Effect 22 (Open Box) to be usable.")

        s = rsec("EffectFlags2 / SelRange / Depth / Delay  (ALWAYS 0 — Suggested 0)")
        rnote(s, "All left at 0 in standard items. SelRange has a non-zero use only for Beta Magic Cards.")

        s = rsec("Life")
        rnote(s, "Duration for timed items (EXP/TM Boosters) or drill life span.  0 = no limit.")

        s = rsec("AP / HP / HPCon / MP / MPCon / Money  (Suggested 0 unless item grants stats)")
        rnote(s,
            "AP = Gun/Throwing ATK (not the APPlus equip stat).\n"
            "HP/MP = recovery amount.  HPCon/MPCon = consumed amount.  "
            "Money = Galder from coupon.")

        s = rsec("APPlus / ACPlus / DXPlus / MAPlus / MDPlus / DPPlus / HVPlus / DAPlus / LKPlus / MaxHP/MP/WTPlus")
        rnote(s, "Static stat bonuses for equipment. All default 0.")

        s = rsec("HPRecoveryRate / MPRecoveryRate  (format: 0.000000)")
        rnote(s, "HP/MP regen rate — used on pets. Format: 0.000000.")

        s = rsec("CardNum / CardGenGrade / CardGenParam / DailyGenCnt  (Suggested 0 for non-cards)")
        rtable(s, [
            (0,"None/dummy/invalid","—","Lowest"),
            (1,"Boss Cards","5-6","—"),
            (2,"High Tier","4-5","—"),
            (3,"Low-mid","4-5","—"),
            (4,"Mid","4","—"),
            (5,"High-low Monster","4","—"),
            (6,"Mid-Low Monster","4","—"),
            (7,"Characters / Skill","3","—"),
            (8,"Low Monster","3","—"),
            (9,"NPCs","3","Highest"),
        ], ["CardNum","Target","Life","Fortune Rank"])
        rnote(s, "Skill cards: CardNum=7, CardGenGrade=0, CardGenParam=0.000000\n"
                 "Star cards: CardNum=1, CardGenGrade=0, span full range.")

        s = rsec("PartFileName  (Suggested blank/space)")
        rnote(s, "ItemParamCM2 only — path to fashion item model. Usually a single space.")

        s = rsec("ChrFTypeFlag / ChrGender / NewCM / FamCM / Summary  (ALWAYS 0 / blank)")
        rnote(s, "All left at 0 or blank. Not shown in editor.")

        s = rsec("ExistType")
        rnote(s, "0 = disabled.  1 = timer enabled / cannot stack (sprints, boosters, etc).")

        s = rsec("Ncash")
        rnote(s, "Cash shop (MyShop) price. 0 = not sold in cash shop.")

        s = rsec("ShopFileName / ShopBundleNum  (Suggested 0/blank)")
        rnote(s, "Points to promotional Cash Mall image. Same NRI +1 offset rule for BundleNum.")

        s = rsec("MinStatType / MinStatLv")
        rtable(s, _MINSTATTYPE_MAP, ["Value", "Description"])
        rnote(s, "MinStatLv = 0 shows no requirement.\n"
                 "⚠ Non-zero MinStatLv blocks equip-swapping while under the Rust debuff.")

        s = rsec("RefineIndex  (Suggested 0)")
        rtable(s, _REFINEINDEX_MAP, ["Value", "Description"])

        s = rsec("RefineType  (Suggested 0)")
        rtable(s, _REFINETYPE_MAP, ["Value", "Description"])

        s = rsec("CompoundSlot  (Suggested 0)")
        rnote(s, "Number of compound slots.  Intended range: 0-5.  "
                 "Higher values cause UI errors (EE) when item is inspected.")

        s = rsec("SetItemID  (Suggested 0)")
        rnote(s, "Equipment set ID reference for set bonuses.")

        s = rsec("ReformCount  (Suggested 0)")
        rnote(s, "Reform count — possibly tied to Skins. Not fully researched.")

        s = rsec("GroupId  (ALWAYS 0)")
        rnote(s, "Always 0 in all standard items.")

        mk_btn(win, "Close", win.destroy, color=BG4).pack(pady=8)


    def _reset(self):
        if messagebox.askyesno("Reset", "Clear all fields and reset to defaults?"):
            _save_t6_settings({})
            self._settings = {}
            self._first_run = True
            self._build_editor()


# ══════════════════════════════════════════════════════════════════════════════
# COMBINED SHELL
# ══════════════════════════════════════════════════════════════════════════════
TOOLS = [
    ("1", "Box XML\nGenerator",   ACC1, Tool1),
    ("1b","ItemParam\nGenerator", ACC6, Tool6),
    ("2", "Rate / Count\nAdjuster", ACC2, Tool2),
    ("3", "NCash Updater\n(Simple)", ACC3, Tool3),
    ("4", "NCash Updater\n(Parent Box)", ACC4, Tool4),
    ("5", "NCash ↔ Ticket\nCalculator", ACC5, Tool5),
]

class CombinedApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Mewsie's ItemParam Toolbox")
        self.geometry("1100x820")
        self.configure(bg=BG2)
        self._current_tool = None
        self._tool_instances = {}
        self._nav_buttons = {}
        self.session = AppSession()
        self._build_layout()
        self._switch_tool(0)

    def _build_layout(self):
        # ── Left sidebar ──────────────────────────────────────────────────
        sidebar = tk.Frame(self, bg=BG2, width=148)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        tk.Label(sidebar, text="BOX\nTOOL\nSUITE", font=("Consolas",13,"bold"),
                 bg=BG2, fg=FG, justify="center").pack(pady=(20,16))

        sep = tk.Frame(sidebar, bg=BG4, height=1); sep.pack(fill="x", padx=10, pady=4)

        for i,(num,label,color,_) in enumerate(TOOLS):
            frm = tk.Frame(sidebar, bg=BG2, cursor="hand2")
            frm.pack(fill="x", padx=8, pady=3)
            dot = tk.Label(frm, text="●", font=("Consolas",9), bg=BG2, fg=color, width=2)
            dot.pack(side="left")
            btn = tk.Button(frm, text=f"  {label}", font=("Consolas",9),
                            bg=BG2, fg=FG_DIM, relief="flat", anchor="w",
                            justify="left", padx=4, pady=6,
                            activebackground=BG3, activeforeground=FG,
                            command=lambda idx=i: self._switch_tool(idx))
            btn.pack(side="left", fill="x", expand=True)
            self._nav_buttons[i] = (frm, btn, dot, color)
            frm.bind("<Button-1>", lambda e, idx=i: self._switch_tool(idx))

        sep2 = tk.Frame(sidebar, bg=BG4, height=1); sep2.pack(fill="x", padx=10, pady=(14,4))
        tk.Label(sidebar, text="Session data\nflows between\ntools. Use\n'Import Session'\nin each tool.",
                 font=("Consolas",7), bg=BG2, fg=FG_GREY, justify="left").pack(padx=12, pady=6, anchor="w")

        # ── Content area ──────────────────────────────────────────────────
        self._content = tk.Frame(self, bg=BG)
        self._content.pack(side="left", fill="both", expand=True)

    def _switch_tool(self, idx):
        if self._current_tool == idx: return

        # Update nav highlight
        for i,(frm,btn,dot,color) in self._nav_buttons.items():
            if i == idx:
                frm.config(bg=BG3); btn.config(bg=BG3, fg=FG); dot.config(bg=BG3, fg=color)
            else:
                frm.config(bg=BG2); btn.config(bg=BG2, fg=FG_DIM); dot.config(bg=BG2, fg=self._nav_buttons[i][3])

        # Hide current
        if self._current_tool is not None:
            if self._current_tool in self._tool_instances:
                self._tool_instances[self._current_tool].pack_forget()

        # Instantiate or show
        if idx not in self._tool_instances:
            _,_,_,ToolClass = TOOLS[idx]
            instance = ToolClass(self._content, self, self.session)
            self._tool_instances[idx] = instance

        self._tool_instances[idx].pack(fill="both", expand=True)
        self._current_tool = idx
        num,label,_,_ = TOOLS[idx]
        self.title(f"Mewsie's ItemParam Toolbox  —  Tool {num}: {label.replace(chr(10),' ')}")


if __name__ == "__main__":
    CombinedApp().mainloop()
