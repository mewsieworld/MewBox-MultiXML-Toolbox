"""
Mewsie's ItemParam Toolbox — v15
  Tool 1  · Box XML Generator  (live validation, last-ID memory, compound/exchange output)
  Tool 1b · ItemParam Generator (dropdown selectors, tooltips, compound/exchange, PresentItemParam)
  Tool 2  · Rate / Count Adjuster
  Tool 3  · NCash Updater (simple CSV)
  Tool 4  · NCash Updater (parent-box CSV + sub-box)
  Tool 5  · NCash ↔ Ticket Calculator

Run: python box_tool_suite.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import csv
import json, io, re, os, copy, json as _json
import datetime
try:
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None
try:
    import openpyxl
    _HAVE_OPENPYXL = True
except ImportError:
    _HAVE_OPENPYXL = False

try:
    from lxml import etree as _lxml_ET
    _HAVE_LXML = True
except ImportError:
    try:
        import xml.etree.ElementTree as _lxml_ET
        _HAVE_LXML = False
    except Exception:
        _lxml_ET = None
        _HAVE_LXML = False

import threading
import time as _time_module

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — character / options data  (used by Tool 1)
# ══════════════════════════════════════════════════════════════════════════════
CHR_NAMES = ["Bunny","Buffalo","Sheep","Dragon","Fox","Lion","Cat","Raccoon","Paula"]
CHR_JOBS  = ["1st","2nd","3rd"]
CHR_FLAG_MAP = {
    "Bunny 1st":1,"Buffalo 1st":2,"Sheep 1st":4,"Dragon 1st":8,"Fox 1st":16,
    "Lion 1st":32,"Cat 1st":64,"Raccoon 1st":124,"Paula 1st":256,
    "Bunny 2nd":512,"Buffalo 2nd":1024,"Sheep 2nd":2048,"Dragon 2nd":4096,
    "Fox 2nd":8192,"Lion 2nd":16384,"Cat 2nd":32768,"Raccoon 2nd":65536,
    "Paula 2nd":131072,"Bunny 3rd":262144,"Buffalo 3rd":524288,"Sheep 3rd":1048576,
    "Dragon 3rd":2097152,"Fox 3rd":4194304,"Lion 3rd":8388608,"Cat 3rd":16777216,
    "Raccoon 3rd":33554432,"Paula 3rd":67108864,
}
CHR_FLAG_REVERSE = {v:k for k,v in CHR_FLAG_MAP.items()}

# ── Job-name to (race, job_tier) lookup for CSV auto-populate ───────────────
# Each entry: normalised_keyword -> (CHR_NAMES entry, "1st"/"2nd"/"3rd")
_JOB_NAME_MAP = {}
def _reg(*keywords, race, tier):
    for kw in keywords:
        _JOB_NAME_MAP[re.sub(r"[^a-z0-9]","", kw.lower())] = (race, tier)
# Bunny
_reg("schoolgirl","bunnyschoolgirl", race="Bunny", tier="1st")
_reg("boxer","bunnyboxer", race="Bunny", tier="2nd")
_reg("champion","duelist","bunnychampion","bunnyduelist", race="Bunny", tier="3rd")
# Buffalo
_reg("fighter","buffalofighter", race="Buffalo", tier="1st")
_reg("warrior","buffalowarrior", race="Buffalo", tier="2nd")
_reg("mercenary","gladiator","buffalomercenary","buffalogladiator", race="Buffalo", tier="3rd")
# Sheep
_reg("librarian","sheeplibrarian", race="Sheep", tier="1st")
_reg("bard","sheepbard", race="Sheep", tier="2nd")
_reg("witch","soulmaster","sheepwitch","sheepsoulmaster", race="Sheep", tier="3rd")
# Dragon
_reg("shaman","dragonshaman", race="Dragon", tier="1st")
_reg("magician","dragonmagician", race="Dragon", tier="2nd")
_reg("priest","darklord","wizard","dragonpriest","dragondarklord","dragonwizard", race="Dragon", tier="3rd")
# Fox
_reg("foxfirst","fox1st", race="Fox", tier="1st")
_reg("foxsecond","fox2nd", race="Fox", tier="2nd")
_reg("foxthird","fox3rd", race="Fox", tier="3rd")
# Lion
_reg("lionfirst","lion1st", race="Lion", tier="1st")
_reg("lionsecond","lion2nd", race="Lion", tier="2nd")
_reg("lionthird","lion3rd", race="Lion", tier="3rd")
# Cat
_reg("model","catmodel", race="Cat", tier="1st")
_reg("entertainer","catentertainer", race="Cat", tier="2nd")
_reg("primadonna","diva","catprimadonna","catdiva", race="Cat", tier="3rd")
# Raccoon
_reg("teacher","raccoonteacher", race="Raccoon", tier="1st")
_reg("cardmaster","raccooncardmaster", race="Raccoon", tier="2nd")
_reg("gambler","duke","raccoongambler","raccounduke", race="Raccoon", tier="3rd")
# Paula / Polar Bear / Bear
_reg("animallover","paulaanimallover","polarbearanimallover", race="Paula", tier="1st")
_reg("trainer","paulatrainer","polarbeartrainer", race="Paula", tier="2nd")
_reg("zoologist","paulazoologist","polarbearzoologist", race="Paula", tier="3rd")

# Race aliases: bear / polarbear / paula all -> "Paula"
_RACE_ALIAS = {}
for aliases, canon in [
    (["bunny"], "Bunny"), (["buffalo"], "Buffalo"),
    (["sheep"], "Sheep"), (["dragon"], "Dragon"),
    (["fox"], "Fox"), (["lion"], "Lion"), (["cat"], "Cat"),
    (["raccoon"], "Raccoon"),
    (["paula","bear","polarbear","polar bear","polar"], "Paula"),
]:
    for a in aliases: _RACE_ALIAS[re.sub(r"[^a-z]","", a.lower())] = canon

# Tier aliases
_TIER_ALIAS = {}
for aliases, canon in [
    (["1st","first","1","job1","firstjob","1stjob","job1st"], "1st"),
    (["2nd","second","2","job2","secondjob","2ndjob","job2nd"], "2nd"),
    (["3rd","third","3","job3","thirdjob","3rdjob","job3rd"], "3rd"),
]:
    for a in aliases: _TIER_ALIAS[re.sub(r"[^a-z0-9]","", a.lower())] = canon

# Sorted longest-first so "1st" matches before "1" when scanning fused strings
_TIER_TOKENS = sorted(_TIER_ALIAS.keys(), key=len, reverse=True)
_RACE_TOKENS = sorted(_RACE_ALIAS.keys(), key=len, reverse=True)

def _split_fused(norm):
    """Extract (race_canon, tier_canon) from a fused normalised string like 'bunny1st'.
    Tries peeling a known race token from start/end, then a tier token from what remains,
    and vice-versa. Returns (None, None) if nothing matched.
    """
    for first_map, first_alias, second_map, second_alias in [
        (_RACE_TOKENS, _RACE_ALIAS, _TIER_TOKENS, _TIER_ALIAS),
        (_TIER_TOKENS, _TIER_ALIAS, _RACE_TOKENS, _RACE_ALIAS),
    ]:
        for tok in first_map:
            if norm.startswith(tok) or norm.endswith(tok):
                rest = norm[len(tok):] if norm.startswith(tok) else norm[:-len(tok)]
                for tok2 in second_map:
                    if rest == tok2 or rest.startswith(tok2) or rest.endswith(tok2):
                        a = first_alias[tok]; b = second_alias[tok2]
                        # Figure out which is race and which is tier
                        if tok in _RACE_ALIAS:  return a, b
                        else:                   return b, a
    return None, None

def resolve_chr_flag(text):
    """Try to resolve free-text like '2nd job Sheep' or 'Witch' to a CHR_FLAG_MAP value.
    Returns int flag or None."""
    raw = text.strip()
    norm = re.sub(r"[^a-z0-9]", "", raw.lower())
    # Direct job name lookup
    if norm in _JOB_NAME_MAP:
        race, tier = _JOB_NAME_MAP[norm]
        return CHR_FLAG_MAP.get(f"{race} {tier}")
    # Try to parse "tier race" or "race tier" patterns
    # Strip common joiners
    parts = re.split(r"[-–—/\s]+", raw.lower())
    parts = [re.sub(r"[^a-z0-9]","", p) for p in parts if p.strip()]
    found_race = found_tier = None
    for p in parts:
        if p in _RACE_ALIAS: found_race = _RACE_ALIAS[p]
        if p in _TIER_ALIAS: found_tier = _TIER_ALIAS[p]
    # Also try each part as a job name (e.g. "witch")
    if not found_race or not found_tier:
        for p in parts:
            if p in _JOB_NAME_MAP:
                r, t = _JOB_NAME_MAP[p]
                if not found_race: found_race = r
                if not found_tier: found_tier = t
    if found_race and found_tier:
        return CHR_FLAG_MAP.get(f"{found_race} {found_tier}")
    # Last pass: try splitting the fully fused normalised string (e.g. "bunny1", "2ndbunny")
    fused_race, fused_tier = _split_fused(norm)
    if fused_race and fused_tier:
        return CHR_FLAG_MAP.get(f"{fused_race} {fused_tier}")
    return None
OPTIONS_CHECKS = [
    ("Not Buyable",256),("Not Sellable",512),("Not Exchangeable",1024),
    ("Not Pickable",2048),("Not Droppable",4096),("Not Vanishable",8192),
    ("No Angelina Bank",65536),("No Lisa Bank",131072),
]

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — XML regex  (used by Tools 1/2/3/4)
# ══════════════════════════════════════════════════════════════════════════════
ROW_RE   = re.compile(r'<ROW>.*?</ROW>', re.DOTALL)
CDATA_RE = re.compile(r'<!\[CDATA\[(.*?)\]\]>', re.DOTALL)

def _get_tag(block, tag):
    m = re.search(rf'<{re.escape(tag)}>(.*?)</{re.escape(tag)}>', block, re.DOTALL)
    if not m: return ""
    cd = CDATA_RE.search(m.group(1))
    return cd.group(1).strip() if cd else m.group(1).strip()

def _set_tag(block, tag, val):
    return re.sub(rf'<{re.escape(tag)}>.*?</{re.escape(tag)}>',
                  f'<{tag}>{val}</{tag}>', block, flags=re.DOTALL)

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — ItemParam / NCash helpers  (Tools 1/3/4)
# ══════════════════════════════════════════════════════════════════════════════
TARGET_FILES = {"itemparam2.xml","itemparamcm2.xml","itemparamex2.xml","itemparamex.xml"}
PRESENT_FILE = "presentitemparam2.xml"

def build_item_lib(files):
    lib = {}
    for _, text in files:
        for row in ROW_RE.findall(text):
            rid  = _get_tag(row, "ID")
            name = _get_tag(row, "Name")
            if rid.isdigit() and name:
                lib[rid] = name
    return lib

def _add_recycle_flag(block):
    """Add 262144 (recyclable) to <Options> if not already present."""
    def _patch_opts(m):
        opts_str = m.group(1)
        parts = [x.strip() for x in opts_str.split('/') if x.strip()]
        int_parts = []
        for p in parts:
            try: int_parts.append(int(p))
            except: int_parts.append(p)
        if 262144 not in int_parts:
            int_parts.append(262144)
        return f'<Options>{"/".join(str(x) for x in int_parts)}</Options>'
    return re.sub(r'<Options>(.*?)</Options>', _patch_opts, block)

def bulk_update_ncash(xml_text, updates):
    found = {k: False for k in updates}
    def replace_row(m):
        block = m.group(0)
        rid   = _get_tag(block, "ID")
        if rid not in updates: return block
        found[rid] = True
        block = re.sub(r'<Ncash>\d+</Ncash>', f'<Ncash>{updates[rid]}</Ncash>', block)
        block = _add_recycle_flag(block)
        return block
    return ROW_RE.sub(replace_row, xml_text), found

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 1 CSV + XML builders
# ══════════════════════════════════════════════════════════════════════════════
# ── Column headers that are NEVER box-name columns ──────────────────────────
# ── Column-header sets for CSV field recognition ─────────────────────────────
_SKIP_HEADERS = {
    "id","level","rate","lv","luck","lvl","chance","prob","itemcnt","count",
    "droprate","dropcnt","dropid","drop","qty","quantity","amount","weight",
}

def _norm_hdr(h):
    return re.sub(r"[^a-z0-9]", "", h.strip().lower())

# Headers that identify box-name columns (the column header IS the box name)
# A column is a box-name column when its header is non-empty, not a known skip
# word, not purely numeric, and NOT a recognised field/metadata header.
_FIELD_HDRS = {
    "name","comment","use","filename","fn","bundlenum","bn",
    "cmtfilename","cmtfn","cmtbundlenum","cmtbn",
    "weight","value","minlevel","level","lvl","money","ncash","tickets","ticket",
    "options","chrtypeflags","chrtypeflag","recycle","recyclable",
    "boxid","contents","content","id",
}

def _is_box_name_header(h):
    hn = _norm_hdr(h)
    if not hn: return False
    if hn.isdigit(): return False
    if hn in _SKIP_HEADERS: return False
    if hn in _FIELD_HDRS: return False
    return True

def _parse_options_cell(cell):
    """Parse Options cell → (opt_checks list, recycle_int).
    Accepts:
      • Text labels  e.g. 'Not Buyable, Not Sellable'
      • Raw bitmask  e.g. '2' or '2/256/512' (slash/comma separated ints)
        The bitmask is matched against each OPTIONS_CHECKS flag value.
    """
    opt_checks = [False] * 8
    recycle = 0
    cell = cell.strip()
    # Try numeric / slash-separated-numeric first
    raw_parts = re.split(r"[,;|/\s]+", cell)
    all_numeric = all(p.strip().isdigit() for p in raw_parts if p.strip())
    if all_numeric and any(p.strip() for p in raw_parts):
        combined = 0
        for p in raw_parts:
            p = p.strip()
            if p.isdigit(): combined |= int(p)
        # Map each OPTIONS_CHECKS value
        for i, (_, v) in enumerate(OPTIONS_CHECKS):
            if combined & v: opt_checks[i] = True
        # Recycle flags
        if combined & 262144:  recycle = 262144
        if combined & 8388608: recycle = 8388608
        return opt_checks, recycle
    # Text label parsing
    tokens = [re.sub(r"[^a-z0-9]", "", t.lower()) for t in re.split(r"[,;|]+", cell) if t.strip()]
    for tok in tokens:
        for i, (lbl, _) in enumerate(OPTIONS_CHECKS):
            if re.sub(r"[^a-z0-9]", "", lbl.lower()) in tok or tok in re.sub(r"[^a-z0-9]", "", lbl.lower()):
                opt_checks[i] = True
        if tok in ("recyclable", "recycle", "recyc"):                         recycle = 262144
        if tok in ("nonrecyclable", "norecycle", "nonrecycle", "notrecyclable"): recycle = 8388608
    return opt_checks, recycle

def _parse_chrtypeflags_cell(cell):
    """Parse ChrTypeFlags cell → list of flag ints.
    Accepts:
      • Text  e.g. 'Sheep 2nd, Dragon 3rd' or 'Witch, Priest'
      • Raw bitmask integer e.g. '2050'  (decoded to individual flag values)
      • Slash/comma separated raw ints e.g. '2048/2'
    """
    cell = cell.strip()
    # Check if the whole cell is a pure bitmask (one int or slash-sep ints)
    raw_parts = re.split(r"[,;|/\s]+", cell)
    all_numeric = all(p.strip().isdigit() for p in raw_parts if p.strip())
    if all_numeric and any(p.strip() for p in raw_parts):
        combined = 0
        for p in raw_parts:
            if p.strip().isdigit(): combined |= int(p.strip())
        # Expand combined bitmask to individual known flags
        flags = []
        for v in sorted(CHR_FLAG_MAP.values()):
            if combined & v: flags.append(v)
        return flags
    # Text parsing
    flags = []
    for part in re.split(r"[,;|/]+", cell):
        part = part.strip()
        if not part: continue
        if part.isdigit(): flags.append(int(part)); continue
        f = resolve_chr_flag(part)
        if f is not None: flags.append(f)
    return flags

# Map normalised header → cfg_override key (special keys start with _)
_HDR_TO_CFGKEY = {
    "name":         "name",
    "comment":      "comment",
    "use":          "use",
    "filename":     "file_name",   "fn": "file_name",
    "bundlenum":    "bundle_num",  "bn": "bundle_num",
    "cmtfilename":  "cmt_file_name", "cmtfn": "cmt_file_name",
    "cmtbundlenum": "cmt_bundle_num", "cmtbn": "cmt_bundle_num",
    "weight":       "weight",
    "value":        "value",
    "minlevel":     "min_level",
    "level":        "min_level",
    "lvl":          "min_level",
    "money":        "money",
    "ncash":        "_ncash",      # handled specially
    "tickets":      "_tickets",    # handled specially
    "ticket":       "_tickets",
    "options":      "_options",
    "chrtypeflags": "_chr",
    "chrtypeflag":  "_chr",
    "recycle":      "_recycle",
    "recyclable":   "_recycle",
    "boxid":        "_boxid",      # box's own ID
    "id":           "_id",         # ambiguous – could be box ID or item ID
    "contents":     "_contents",
    "content":      "_contents",
}

# Filename fields — keep only safe path characters (letters, digits, . _ - \ / : space)
_FILENAME_KEYS = {"file_name", "cmt_file_name"}

def _sanitise_filename(val):
    """Keep only safe path characters including Windows backslash."""
    # Explicitly keep backslash so Windows paths like data\\item\\foo.nri survive
    return re.sub(r"[^\w .\/:\\\-]", "", val, flags=re.ASCII)

def _apply_field_col(cfg_override, key, val):
    """Apply a parsed field value to cfg_override dict."""
    if key in _FILENAME_KEYS:
        cfg_override[key] = _sanitise_filename(val)
    elif key == "_options":
        oc, rv = _parse_options_cell(val)
        cfg_override["opt_checks"] = oc
        if rv: cfg_override["opt_recycle"] = rv
    elif key == "_chr":
        flags = _parse_chrtypeflags_cell(val)
        if flags: cfg_override["chr_type_flags"] = flags
    elif key == "_recycle":
        nv = _norm_hdr(val)
        if nv in ("recyclable","recycle","recyc"):                       cfg_override["opt_recycle"] = 262144
        elif nv in ("nonrecyclable","norecycle","nonrecycle","notrecyclable"): cfg_override["opt_recycle"] = 8388608
    elif key == "_ncash":
        try: cfg_override["ncash_direct"] = int(round(float(val)))
        except: pass
    elif key == "_tickets":
        try:
            cfg_override["ticket"] = val
            # A Ticket/Tickets column means Recyclable — auto-set recycle flag
            cfg_override["opt_recycle"] = 262144
        except: pass
    elif not key.startswith("_"):
        cfg_override[key] = val

# ══════════════════════════════════════════════════════════════════════════════
# CMSetItemParam.xml — Set Generator helpers
# ══════════════════════════════════════════════════════════════════════════════

def _norm_set_hdr(h):
    """Normalise a set-CSV header: strip, lowercase, remove non-alnum."""
    return re.sub(r"[^a-z0-9]", "", h.strip().lower())

# Recognised column names for set CSVs
_SET_ID_KEYS   = {"setid", "setitemid", "setid1", "sid"}   # → the set's own ID (<ID>)
_SET_NAME_KEYS = {"setname", "nameofset", "setitemname",
                  "name"}                                    # → the set's <Name>
_ITEM_ID_KEYS  = {"id", "itemid", "item", "itemids",
                  "items"}                                   # → items inside the set
_ITEM_NAME_KEYS = {"itemname", "nameofitem", "boxname",
                   "nameofbox", "itemnames", "boxnames"}     # → inline <!-- comment -->
# "item#" / "item #" columns  → treated as item ID columns when # is digit-only
# "#" alone or leading with digit → item ID (e.g. column "0", "1" … "7")

def _is_item_col(hn):
    """Return True if a normalised header looks like an item-slot column."""
    if hn.isdigit(): return True
    if re.match(r"^item\d+$", hn): return True   # item0, item1 …
    if re.match(r"^item#?$", hn): return True      # item# or item
    return False

def _classify_set_col(hn):
    """Return ('set_id'|'set_name'|'item_id'|'item_name'|'item_slot'|None, slot_index)."""
    if hn in _SET_ID_KEYS:   return "set_id",   None
    if hn in _SET_NAME_KEYS: return "set_name", None
    if hn in _ITEM_ID_KEYS:  return "item_id",  None
    if hn in _ITEM_NAME_KEYS: return "item_name", None
    if hn.isdigit():          return "item_slot", int(hn)
    m = re.match(r"^item(\d+)$", hn)
    if m: return "item_slot", int(m.group(1))
    return None, None


def parse_set_csv(text):
    """Parse a CMSetItemParam CSV.

    Two supported layouts:

    Layout A — one row = one set  (SetID column present)
    ────────────────────────────────────────────────────
    Header:   SetID | Name | ID | Item Name | ID | Item Name | …
      or      SetID | Set Name | Item0 | Item1 | … Item7
    Each data row produces one set ROW.
    ID / Item# columns contain item IDs; Name columns contain inline comments.

    Layout B — multiple rows = one set  (no SetID column; group by set name header)
    ────────────────────────────────────────────────────────────────────────────────
    Header:   <Set Name Header> | ID | # (item slot) | Item Name  (any order)
    The set name column header IS the set name; data rows are items of that set.

    Returns list of dicts:
        {set_id, set_name, items: [{item_id, item_name}]}
    """
    reader   = csv.reader(io.StringIO(text))
    raw_rows = list(reader)
    if not raw_rows: return []
    headers   = [h.strip() for h in raw_rows[0]]
    hn_list   = [_norm_set_hdr(h) for h in headers]
    data_rows = raw_rows[1:]

    # Detect layout
    has_set_id_col  = any(hn in _SET_ID_KEYS  for hn in hn_list)
    has_item_id_col = any(hn in _ITEM_ID_KEYS for hn in hn_list)
    has_item_slots  = any(_classify_set_col(hn)[0] == "item_slot" for hn in hn_list)

    # ── Layout A: SetID column present ──────────────────────────────────────
    if has_set_id_col:
        # Map column indices
        set_id_ci    = next((i for i,hn in enumerate(hn_list) if hn in _SET_ID_KEYS),   None)
        set_name_ci  = next((i for i,hn in enumerate(hn_list) if hn in _SET_NAME_KEYS), None)

        # Item columns: either "Item0".."Item7" slots, or repeated "ID" + optional "Name of Item" pairs
        slot_cols  = {}  # slot_index -> (id_col_idx, name_col_idx)
        if has_item_slots:
            # Item# / # style: each slot column is an item ID; look for adjacent name col
            slot_ci_list = [(i, _classify_set_col(hn)[1])
                            for i,hn in enumerate(hn_list)
                            if _classify_set_col(hn)[0] == "item_slot"]
            slot_ci_list.sort(key=lambda x: x[1])
            for ci, slot in slot_ci_list:
                # Check for an adjacent "item name" / "name of item" column
                name_ci = None
                for offset in (1, -1):
                    ni = ci + offset
                    if 0 <= ni < len(hn_list) and hn_list[ni] in _ITEM_NAME_KEYS:
                        name_ci = ni; break
                slot_cols[slot] = (ci, name_ci)
        elif has_item_id_col:
            # Plain repeated "ID" columns — collect all of them
            id_indices = [i for i,hn in enumerate(hn_list) if hn in _ITEM_ID_KEYS]
            name_indices = [i for i,hn in enumerate(hn_list) if hn in _ITEM_NAME_KEYS]
            for slot, ci in enumerate(id_indices):
                ni = name_indices[slot] if slot < len(name_indices) else None
                slot_cols[slot] = (ci, ni)

        results = []
        for row in data_rows:
            def _cell(ci): return row[ci].strip() if ci is not None and ci < len(row) else ""
            set_id   = _cell(set_id_ci)
            set_name = _cell(set_name_ci) if set_name_ci is not None else ""
            if not set_id: continue
            items = []
            for slot in sorted(slot_cols.keys()):
                id_ci, nm_ci = slot_cols[slot]
                item_id   = _cell(id_ci)
                item_name = _cell(nm_ci) if nm_ci is not None else ""
                items.append({"item_id": item_id, "item_name": item_name})
            # Pad to 8 slots
            while len(items) < 8:
                items.append({"item_id": "0", "item_name": ""})
            results.append({"set_id": set_id, "set_name": set_name,
                            "items": items[:8]})
        return results

    # ── Layout B: group-by-column (set name is the column header) ───────────
    # Find group columns (not a known field header)
    known_hn = (_SET_ID_KEYS | _SET_NAME_KEYS | _ITEM_ID_KEYS |
                _ITEM_NAME_KEYS | {"id","name","comment"})
    group_col_indices = [i for i,hn in enumerate(hn_list)
                         if hn_list[i] and not hn_list[i].isdigit()
                         and hn_list[i] not in known_hn
                         and not _is_item_col(hn_list[i])]

    if not group_col_indices:
        return []

    # For each group column, collect items from all data rows under it
    results = []
    for gc in group_col_indices:
        set_name = headers[gc]   # column header = set name

        # Find id_ci and name_ci within the span
        id_ci   = next((i for i,hn in enumerate(hn_list)
                        if hn in _ITEM_ID_KEYS and i != gc), None)
        name_ci = next((i for i,hn in enumerate(hn_list)
                        if hn in _ITEM_NAME_KEYS), None)
        # slot column (# or item#) if present
        slot_ci = next((i for i,hn in enumerate(hn_list)
                        if _is_item_col(hn)), None)

        # Set's own ID: look for first non-empty cell in a set_id-like column,
        # or fall back to the set_name if it looks like a number
        set_id_ci_b = next((i for i,hn in enumerate(hn_list)
                            if hn in _SET_ID_KEYS), None)

        items = []
        set_id = ""
        for row in data_rows:
            def _c(ci): return row[ci].strip() if ci is not None and ci < len(row) else ""
            if not set_id and set_id_ci_b is not None:
                v = _c(set_id_ci_b)
                if v: set_id = v
            item_id   = _c(id_ci) if id_ci is not None else _c(gc)
            item_name = _c(name_ci) if name_ci is not None else ""
            if item_id: items.append({"item_id": item_id, "item_name": item_name})

        # If set_id not found, try using set_name if numeric
        if not set_id:
            set_id = set_name if set_name.isdigit() else ""

        while len(items) < 8:
            items.append({"item_id": "0", "item_name": ""})
        results.append({"set_id": set_id, "set_name": set_name,
                        "items": items[:8]})
    return results


def build_set_row(set_cfg):
    """Build a <ROW> for CMSetItemParam.xml.
    set_cfg: {set_id, set_name, items:[{item_id, item_name}]}
    item_name is appended as <!-- name --> when non-empty.
    """
    lines = ["<ROW>",
             f"<ID>{set_cfg['set_id']}</ID>",
             f"<Name><![CDATA[{set_cfg['set_name']}]]></Name>"]
    for i, item in enumerate(set_cfg["items"][:8]):
        item_id   = str(item.get("item_id", "0") or "0")
        item_name = str(item.get("item_name", "")).strip()
        line = f"<Item{i}>{item_id}</Item{i}>"
        if item_name:
            line += f" <!-- {item_name} -->"
        lines.append(line)
    # Pad to 8 if fewer provided
    for i in range(len(set_cfg["items"]), 8):
        lines.append(f"<Item{i}>0</Item{i}>")
    lines.append("</ROW>")
    return "\n".join(lines)


def parse_grouped_csv(text):
    """Parse a box-definition CSV.

    Layout rules (same as original, with additions):
    - Header row: columns whose headers are not recognised field/skip words are
      "box-name columns".  Each defines a GROUP.
    - A group's span is: columns from (previous group's box-name col + 1) up to
      and including its own box-name col.
    - Within each span:
        * The box-name col: each data row's cell is an item NAME (or box name if
          a "Box ID" / "Box …" column is present).
        * Any recognised field header (Name, Comment, Options, ChrTypeFlags,
          BoxID, Contents, …) is parsed and stored in cfg_override.
        * Numeric middle columns are item DropRate.
        * "ItemCnt" / "Count" column sets item_cnt.
        * "Contents" column overrides item name source.
        * "Box ID" or "BoxID" column sets the box's own ID in cfg_override.
    - Returns list of {box_name, items, cfg_override} dicts.
    """
    reader = csv.reader(io.StringIO(text))
    rows   = list(reader)
    if not rows: return []
    headers = [h.strip() for h in rows[0]]
    hnorms  = [_norm_hdr(h) for h in headers]
    data_rows = rows[1:]

    # Identify box-name columns
    box_col_indices = [i for i, h in enumerate(headers) if _is_box_name_header(h)]
    if not box_col_indices: return []

    results = []
    prev_end = -1
    # Compute next box-col for each group (to know where trailing fields end)
    next_box_cols = box_col_indices[1:] + [len(headers)]

    for bi, bc in enumerate(box_col_indices):
        # Core span: prev_end+1 .. bc  (the original item columns)
        # Trailing span: bc+1 .. next_box_col-1 BUT only recognised field headers
        next_bc = next_box_cols[bi]
        _TRAILING_OK = {k for k,v in _HDR_TO_CFGKEY.items()
                        if v not in ("_id",) and k not in _SKIP_HEADERS}
        trailing = [ci for ci in range(bc + 1, next_bc)
                    if _norm_hdr(headers[ci]) in _TRAILING_OK]
        span = list(range(prev_end + 1, bc + 1)) + trailing
        prev_end = bc

        box_name_hdr = headers[bc]   # the column header = box name for the group

        # Classify every column in the span
        id_col       = None   # item ID column index (per-item)
        box_id_col   = None   # box's own ID column index
        contents_col = None   # alternate item-name column
        rate_cols    = []     # numeric rate columns
        itemcnt_col  = None
        field_cols   = {}     # col_index -> cfg_key (field overrides)

        for ci in span:
            if ci == bc: continue   # box-name col itself
            hn = hnorms[ci]
            if hn == "id":
                # Bare "ID" column always means item ID, regardless of skip lists
                if id_col is None: id_col = ci
            elif hn in _HDR_TO_CFGKEY:
                cfgkey = _HDR_TO_CFGKEY[hn]
                if cfgkey == "_boxid":
                    box_id_col = ci
                elif cfgkey == "_id":
                    if id_col is None: id_col = ci
                elif cfgkey == "_contents":
                    contents_col = ci
                else:
                    field_cols[ci] = cfgkey
            elif hn in ("itemcnt", "count"):
                itemcnt_col = ci
            elif hn in _SKIP_HEADERS:
                rate_cols.append(ci)   # treat rate/level/etc as rate source
            else:
                rate_cols.append(ci)   # unknown middle col → try as rate

        # Build cfg_override from field columns (take first non-empty value per field)
        group_cfg = {}
        for ci, cfgkey in field_cols.items():
            for row in data_rows:
                val = row[ci].strip() if ci < len(row) else ""
                if val:
                    _apply_field_col(group_cfg, cfgkey, val)
                    break

        # Box's own ID from box_id_col (first non-empty)
        if box_id_col is not None:
            for row in data_rows:
                val = row[box_id_col].strip() if box_id_col < len(row) else ""
                if val.isdigit():
                    group_cfg["id"] = val
                    break

        # Build items list
        # Check for explicit DropID column (items inside box are the DropID values)
        dropid_col  = next((ci for ci, cfgkey in field_cols.items() if cfgkey == "_dropid"),  None)
        boxname_col = next((ci for ci, cfgkey in field_cols.items() if cfgkey == "_boxname"), None)

        items = []
        for row in data_rows:
            if dropid_col is not None:
                # DropID mode: items come from DropID column; box name from bc col or boxname_col
                raw_name = row[dropid_col].strip() if dropid_col < len(row) else ""
                if not raw_name: continue
                item_id   = raw_name if raw_name.isdigit() else ""
                item_name = ""
                if boxname_col is not None:
                    item_name = row[boxname_col].strip() if boxname_col < len(row) else ""
            else:
                # Normal mode
                # Item ID — prefer id_col; fall back to first digit in box-name col
                item_id = ""
                if id_col is not None:
                    v = row[id_col].strip() if id_col < len(row) else ""
                    if v.isdigit(): item_id = v
                # Item name — from contents_col, or box-name col
                # If id_col gave us an ID, the box-name col text is descriptive (skip it)
                raw_name = row[bc].strip() if bc < len(row) else ""
                if contents_col is not None:
                    item_name = row[contents_col].strip() if contents_col < len(row) else raw_name
                elif id_col is not None:
                    # ID-based format: name col not present, use raw_name only if non-numeric
                    item_name = "" if raw_name.isdigit() else raw_name
                else:
                    item_name = raw_name
                if not item_id and not item_name: continue

            # Rate: first parseable integer from rate_cols
            rate = None
            for ci in rate_cols:
                v = row[ci].strip() if ci < len(row) else ""
                try: rate = int(v.replace("%","").strip()); break
                except: pass

            # ItemCnt
            item_cnt = 1
            if itemcnt_col is not None:
                v = row[itemcnt_col].strip() if itemcnt_col < len(row) else ""
                try: item_cnt = int(v)
                except: pass

            items.append({"id": item_id, "name": item_name,
                          "extra": [], "rate": rate, "item_cnt": item_cnt})

        if items:
            results.append({"box_name": box_name_hdr,
                            "items": items,
                            "cfg_override": group_cfg})

    return results

def substitute_box_name(template, old, new):
    if not old or not template: return template
    return re.sub(re.escape(old), new, template, flags=re.IGNORECASE)

def apply_name_template(template, prev, new):
    if not template or not prev: return new
    return substitute_box_name(template, prev, new)

def deduplicate_name(proposed, existing):
    if proposed not in existing: return proposed
    i = 2
    while f"{proposed} ({i})" in existing: i += 1
    return f"{proposed} ({i})"

def build_options_str(check_states, recycle_val):
    base    = [2,32]
    checked = [v for (_,v),on in zip(OPTIONS_CHECKS,check_states) if on]
    rec     = [recycle_val] if recycle_val > 0 else []
    return "/".join(str(x) for x in sorted(set(base+checked+rec)))

def build_itemparam_row(cfg):
    opts      = build_options_str(cfg["opt_checks"], cfg["opt_recycle"])
    chr_flags = sum(cfg["chr_type_flags"])
    fn        = cfg["file_name"]
    bn        = cfg["bundle_num"]
    return f"""<ROW>
<ID>{cfg['id']}</ID>
<Class>1</Class>
<Type>15</Type>
<SubType>0</SubType>
<ItemFType>0</ItemFType>
<n><![CDATA[{cfg['name']}]]></n>
<Comment><![CDATA[{cfg['comment']}]]></Comment>
<Use><![CDATA[{cfg['use']}]]></Use>
<Name_Eng><![CDATA[ ]]></Name_Eng>
<Comment_Eng><![CDATA[ ]]></Comment_Eng>
<FileName><![CDATA[{fn}]]></FileName>
<BundleNum>{bn}</BundleNum>
<InvFileName><![CDATA[{fn}]]></InvFileName>
<InvBundleNum>{bn}</InvBundleNum>
<CmtFileName><![CDATA[{cfg['cmt_file_name']}]]></CmtFileName>
<CmtBundleNum>{cfg['cmt_bundle_num']}</CmtBundleNum>
<EquipFileName><![CDATA[ ]]></EquipFileName>
<PivotID>0</PivotID>
<PaletteId>0</PaletteId>
<Options>{opts}</Options>
<HideHat>0</HideHat>
<ChrTypeFlags>{chr_flags}</ChrTypeFlags>
<GroundFlags>0</GroundFlags>
<SystemFlags>0</SystemFlags>
<OptionsEx>0</OptionsEx>
<Weight>{cfg['weight']}</Weight>
<Value>{cfg['value']}</Value>
<MinLevel>{cfg['min_level']}</MinLevel>
<Effect>22</Effect>
<EffectFlags2>0</EffectFlags2>
<SelRange>0</SelRange>
<Life>0</Life>
<Depth>0</Depth>
<Delay>0.000000</Delay>
<AP>0</AP>
<HP>0</HP>
<HPCon>0</HPCon>
<MP>0</MP>
<MPCon>0</MPCon>
<Money>{cfg['money']}</Money>
<APPlus>0</APPlus>
<ACPlus>0</ACPlus>
<DXPlus>0</DXPlus>
<MaxMPPlus>0</MaxMPPlus>
<MAPlus>0</MAPlus>
<MDPlus>0</MDPlus>
<MaxWTPlus>0</MaxWTPlus>
<DAPlus>0</DAPlus>
<LKPlus>0</LKPlus>
<MaxHPPlus>0</MaxHPPlus>
<DPPlus>0</DPPlus>
<HVPlus>0</HVPlus>
<HPRecoveryRate>0.000000</HPRecoveryRate>
<MPRecoveryRate>0.000000</MPRecoveryRate>
<CardNum>0</CardNum>
<CardGenGrade>0</CardGenGrade>
<CardGenParam>0.000000</CardGenParam>
<DailyGenCnt>0</DailyGenCnt>
<PartFileName><![CDATA[ ]]></PartFileName>
<ChrFTypeFlag>0</ChrFTypeFlag>
<ChrGender>0</ChrGender>
<ExistType>0</ExistType>
<Ncash>{cfg['ncash']}</Ncash>
<NewCM>0</NewCM>
<FamCM>0</FamCM>
<Summary><![CDATA[ ]]></Summary>
<ShopFileName><![CDATA[ ]]></ShopFileName>
<ShopBundleNum>0</ShopBundleNum>
<MinStatType>0</MinStatType>
<MinStatLv>0</MinStatLv>
<RefineIndex>0</RefineIndex>
<RefineType>0</RefineType>
<CompoundSlot>0</CompoundSlot>
<SetItemID>0</SetItemID>
<ReformCount>0</ReformCount>
<GroupId>0</GroupId>
</ROW>"""

def build_presentparam_row(box_id, items, ptype, drop_cnt, default_rate,
                           item_cnts=None, box_name=None):
    """box_name appended as <!-- name --> after <Id>.
    items[i]["name"] appended as <!-- name --> after each <DropId_#>."""
    is_distrib      = (ptype == 2)
    actual_drop_cnt = len(items) if is_distrib else drop_cnt
    id_line = f"<Id>{box_id}</Id>"
    if box_name:
        id_line += f" <!-- {box_name} -->"
    lines = ["<ROW>", id_line,
             f"<Type>{ptype}</Type>", f"<DropCnt>{actual_drop_cnt}</DropCnt>"]
    for i in range(20):
        if i < len(items):
            did   = items[i]["id"]
            drate = 100 if is_distrib else (items[i].get("rate") or default_rate)
            icnt  = (item_cnts[i] if item_cnts and i < len(item_cnts) else 1) or 1
            item_name = items[i].get("name", "").strip()
            drop_id_line = f"<DropId_{i}>{did}</DropId_{i}>"
            if item_name:
                drop_id_line += f" <!-- {item_name} -->"
        else:
            did, drate, icnt = 0, 0, 0
            drop_id_line = f"<DropId_{i}>{did}</DropId_{i}>"
        lines += [drop_id_line,
                  f"<DropRate_{i}>{drate}</DropRate_{i}>",
                  f"<ItemCnt_{i}>{icnt}</ItemCnt_{i}>"]
    lines.append("</ROW>")
    return "\n".join(lines)

def build_recycle_except_row(box_id, name):
    return f"<ROW>\n<ItemID>{box_id}</ItemID>\n<Comment><![CDATA[{name}]]></Comment>\n</ROW>"

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 2 XML helpers
# ══════════════════════════════════════════════════════════════════════════════

def _sanitise_xml_path(path):
    """Ensure a file path uses single backslashes in XML CDATA (never \\\\).
    Input may use / or \\ or \\\\; output always uses single \\.
    """
    if not path or path.strip() in ('', ' '):
        return path
    # Normalise: replace \\\\ -> \\, then / -> \\
    p = path.replace('\\\\', '\\').replace('/', '\\')
    return p


def real_drop_slots(block):
    pairs = re.findall(r'<DropId_(\d+)>(\d+)</DropId_\d+>', block)
    return [(int(i),v) for i,v in sorted(pairs,key=lambda x:int(x[0])) if v!="0"]

def apply_cfg_to_row(block, cfg):
    block = _set_tag(block,"Type",str(cfg["type"]))
    block = _set_tag(block,"DropCnt",str(cfg["drop_cnt"]))
    for pos,(idx,_) in enumerate(real_drop_slots(block)):
        sc = cfg["slots"][pos] if pos<len(cfg["slots"]) else {"rate":100,"count":1}
        block = _set_tag(block,f"DropRate_{idx}",str(sc["rate"]))
        block = _set_tag(block,f"ItemCnt_{idx}",str(sc["count"]))
    return block

def load_itemparam_folder(folder):
    lib = {}
    for fname in os.listdir(folder):
        if not fname.lower().endswith(".xml"): continue
        try:
            with open(os.path.join(folder,fname),encoding="utf-8-sig",errors="replace") as f:
                text = f.read()
            for row in ROW_RE.findall(text):
                rid  = _get_tag(row,"ID")
                name = _get_tag(row,"n")
                if rid.isdigit() and name: lib[rid] = name
        except: pass
    return lib

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 2 CSV parser
# ══════════════════════════════════════════════════════════════════════════════
_T2_SKIP = {"id","level","rate","lv","luck","lvl","chance","prob"}

def parse_box_id_csv(text):
    """Parse a box-ID CSV. Returns {id: name} dict.
    Accepts tall or wide format, with columns named:
      ID / BoxID / box_id  → item ID (required)
      Name / BoxName / <any non-field header>  → display name (optional)
      Level / MinLevel / lvl  → ignored for box_map but preserved in parse_box_csv_groups
    Multiple ID columns = multiple box groups (wide format).
    """
    result = {}
    for gid, gname, _items in parse_box_csv_groups(text):
        for iid, iname, _meta in _items:
            if iid and iid not in result:
                result[iid] = iname
        # Also register the outer box name itself
        if gid and gid not in result:
            result[gid] = gname
    return result


def parse_box_csv_groups(text):
    """Parse a box CSV into groups. Returns list of (outer_box_id, outer_box_name, items).
    items = list of (id, name, meta_dict) where meta_dict may contain:
      'level', 'droprate', 'droprate_N', 'itemcnt', 'itemcnt_N', 'rate'

    Accepts:
    - Wide format: repeating groups of cols [ID, optional-fields..., BoxName-header]
    - Tall format: single group with ID col + Name col
    - DropRate / DropRate_3 / Rate / ItemCnt / ItemCnt_2 / Level / MinLevel columns
    """
    import re as _re
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows: return []
    headers = [h.strip() for h in rows[0]]
    data_rows = rows[1:]

    def _norm(h): return _re.sub(r"[^a-z0-9]","",h.lower())

    # Field header norms that mean "this col carries data, not a group name"
    FIELD_NORMS = {"id","boxid","box_id","itemid","name","boxname","itemname",
                   "level","minlevel","lvl","rate","droprate","itemcnt","count"}
    def _is_field(h):
        n = _norm(h)
        if not n: return True
        if n in FIELD_NORMS: return True
        if _re.match(r"droprate\d*$", n): return True
        if _re.match(r"itemcnt\d*$", n): return True
        if _re.match(r"rate\d*$", n): return True
        if _re.match(r"itemcount\d*$", n): return True
        return False

    # Find box-name columns (non-field headers → group name)
    box_name_cols = [i for i,h in enumerate(headers) if h.strip() and not _is_field(h)]

    if not box_name_cols:
        # Tall format — single group, no outer box name
        # Find the ID col
        id_col = next((i for i,h in enumerate(headers) if _norm(h) in ("id","boxid","itemid")), 0)
        name_col = next((i for i,h in enumerate(headers)
                         if i != id_col and _norm(h) in ("name","boxname","itemname")), None)
        items = []
        for row in data_rows:
            iid = row[id_col].strip() if id_col < len(row) else ""
            if not iid or not iid.isdigit(): continue
            nm = row[name_col].strip() if name_col is not None and name_col < len(row) else ""
            items.append((iid, nm, {}))
        return [("", "", items)]

    # Wide format — one group per box_name_col
    groups = []
    prev = -1
    for bi, bc in enumerate(box_name_cols):
        outer_name = headers[bc]
        outer_id   = ""  # outer box has no explicit ID column in this format
        # Span: prev+1 .. bc
        span = list(range(prev+1, bc+1))
        prev = bc

        # Classify span cols
        id_col_local   = None
        name_col_local = None
        field_cols     = {}  # col_idx → field_key

        for ci in span:
            if ci == bc: continue  # box-name col itself = the outer box name
            n = _norm(headers[ci])
            if n in ("id","boxid","itemid") and id_col_local is None:
                id_col_local = ci
            elif n in ("name","boxname","itemname") and name_col_local is None:
                name_col_local = ci
            elif _re.match(r"droprate(\d*)$", n):
                m_ = _re.match(r"droprate(\d*)$", n)
                field_cols[ci] = f"droprate_{m_.group(1) or ''}".rstrip("_")
            elif _re.match(r"rate(\d*)$", n):
                field_cols[ci] = "rate"
            elif _re.match(r"itemcnt(\d*)$|itemcount(\d*)$", n):
                m_ = _re.match(r"(?:itemcnt|itemcount)(\d*)$", n)
                field_cols[ci] = f"itemcnt_{m_.group(1) or ''}".rstrip("_")
            elif n in ("level","minlevel","lvl"):
                field_cols[ci] = "level"

        items = []
        for row in data_rows:
            iid = ""
            if id_col_local is not None:
                iid = row[id_col_local].strip() if id_col_local < len(row) else ""
            if not iid or not iid.isdigit(): continue
            # Name: from name col, or from box-name col cell, or from last field col
            nm = ""
            if name_col_local is not None:
                nm = row[name_col_local].strip() if name_col_local < len(row) else ""
            if not nm:
                nm = row[bc].strip() if bc < len(row) else ""
            meta = {}
            for ci, fkey in field_cols.items():
                v = row[ci].strip() if ci < len(row) else ""
                if v: meta[fkey] = v
            items.append((iid, nm, meta))

        groups.append((outer_id, outer_name, items))

    return groups


# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 3 CSV parser
# ══════════════════════════════════════════════════════════════════════════════
def parse_csv_text_t3(text):
    stripped = text.strip()
    if not stripped: return []
    all_rows = list(csv.reader(io.StringIO(stripped)))
    if not all_rows: return []
    raw_headers = [h.strip() for h in all_rows[0]]
    data_rows   = all_rows[1:]
    items = []
    item_col_positions = [i for i,h in enumerate(raw_headers) if re.match(r'Item\d+_ID',h,re.I)]
    if item_col_positions:
        seen = set()
        for row in data_rows:
            for pos in item_col_positions:
                if pos<len(row):
                    v = row[pos].strip()
                    if v and v.isdigit() and v not in seen:
                        seen.add(v); items.append({"id":v,"ticket_cost":None})
        return items
    id_positions = [i for i,h in enumerate(raw_headers) if h.lower()=="id"]
    if id_positions:
        # Per-group dedup — same ID can legitimately appear in multiple groups
        group_seen = {gi: set() for gi in range(len(id_positions))}
        for row in data_rows:
            for gi,pos in enumerate(id_positions):
                if pos<len(row):
                    v = row[pos].strip()
                    if v and v.isdigit() and v not in group_seen[gi]:
                        group_seen[gi].add(v)
                        items.append({"id":v,"ticket_cost":None})
        return items
    if len(raw_headers)>=2:
        for row in data_rows:
            if not row: continue
            raw_cost = row[1].strip() if len(row)>1 else ""
            try:    cost = float(raw_cost)
            except: cost = None
            add(row[0], cost)
        return items
    for row in data_rows:
        if row: add(row[0])
    return items

# ══════════════════════════════════════════════════════════════════════════════
# SHARED — Tool 4 CSV parser
# ══════════════════════════════════════════════════════════════════════════════
_NON_ID_HEADERS = {
    "acplus","ap","applus","bundlenum","cardgengrade","cardgenparam","cardnum",
    "chrftypeflag","chrgender","chrtypeflags","class","cmtbundlenum","cmtfilename",
    "comment","comment_eng","compoundslot","daplus","dpplus","dxplus","dailygencnt",
    "delay","depth","effect","effectflags2","equipfilename","existtype","famcm",
    "filename","groundflags","groupid","hp","hpcon","hprecoveryrate","hvplus",
    "hidehat","invbundlenum","invfilename","itemftype","lkplus","life","maplus",
    "mdplus","mp","mpcon","mprecoveryrate","maxhpplus","maxmpplus","maxwtplus",
    "minlevel","minstatLv","minstattype","money","name","name_eng","newcm",
    "options","optionsex","paletteid","partfilename","pivotid","refineindex",
    "refinetype","reformcount","selrange","setitemid","shopbundlenum","shopfilename",
    "subtype","summary","systemflags","type","use","value","weight",
    "level","rate","lv","luck","lvl","chance","prob","drop","qty",
    "quantity","count","amount","row",
    "droprate","dropcnt","dropid","itemcnt","dropcnt_0","dropcnt_1",
    "droprate_0","droprate_1","droprate_2","droprate_3","droprate_4",
    "dropid_0","dropid_1","dropid_2","dropid_3","dropid_4",
    "itemcnt_0","itemcnt_1","itemcnt_2","itemcnt_3","itemcnt_4",
}
_TICKET_NAMES     = {"tickets","ticket"}
_NCASH_NAMES      = {"ncash","ncash_val","ncashval"}
_BOX_TICKET_NAMES = {
    "tickets of box contents","box contents tickets",
    "box content tickets","tickets of box content",
    "sub-box tickets","subbox tickets",
}

def _find_value_col(raw_headers, id_pos):
    next_id = next((i for i in range(id_pos+1,len(raw_headers))
                    if raw_headers[i].lower()=="id"), len(raw_headers))
    for i in range(id_pos+1, next_id):
        h = raw_headers[i].lower()
        if h in _TICKET_NAMES: return i,"tickets"
        if h in _NCASH_NAMES:  return i,"ncash"
    return None, None

def _find_box_ticket_col(raw_headers, id_pos):
    next_id = next((i for i in range(id_pos+1,len(raw_headers))
                    if raw_headers[i].lower()=="id"), len(raw_headers))
    for i in range(id_pos+1, next_id):
        if raw_headers[i].lower().strip() in _BOX_TICKET_NAMES: return i
    return None

def parse_parentbox_csv(text):
    stripped = text.strip()
    if not stripped: return []
    all_rows = list(csv.reader(io.StringIO(stripped)))
    if not all_rows: return []
    raw_headers = [h.strip() for h in all_rows[0]]
    data_rows   = all_rows[1:]
    id_positions = [i for i,h in enumerate(raw_headers) if h.lower()=="id"]
    val_map, box_tick_map = {}, {}
    for id_pos in id_positions:
        vcol,vtype = _find_value_col(raw_headers, id_pos)
        if vcol is not None: val_map[id_pos] = (vcol,vtype)
        btcol = _find_box_ticket_col(raw_headers, id_pos)
        if btcol is not None: box_tick_map[id_pos] = btcol
    items = []
    def _parse_num(row, col):
        if col is None or col>=len(row): return None
        try:    return float(row[col].strip())
        except: return None
    if id_positions:
        # Deduplicate per-group only — same ID can appear in multiple groups (boxes)
        group_seen = {gi: set() for gi in range(len(id_positions))}
        for row in data_rows:
            for gi,id_pos in enumerate(id_positions):
                if id_pos>=len(row): continue
                id_val = row[id_pos].strip()
                if not (id_val and id_val.isdigit()): continue
                if id_val in group_seen[gi]: continue
                group_seen[gi].add(id_val)
                ticket_cost = ncash_direct = None
                if id_pos in val_map:
                    vcol,vtype = val_map[id_pos]
                    num = _parse_num(row, vcol)
                    if num is not None:
                        if vtype=="tickets": ticket_cost  = num
                        else:               ncash_direct = int(round(num))
                btcol = box_tick_map.get(id_pos)
                box_ticket_cost = _parse_num(row,btcol) if btcol is not None else None
                items.append({"id":id_val,"ticket_cost":ticket_cost,"ncash_direct":ncash_direct,
                              "box_ticket_cost":box_ticket_cost,"group_idx":gi,"name":""})
        return items
    for row in data_rows:
        for i,cell in enumerate(row):
            hdr = raw_headers[i].lower() if i<len(raw_headers) else ""
            if hdr not in _NON_ID_HEADERS:
                add(cell, None, None, None)
    return items

def extract_drop_ids_from_present(present_text, box_ids):
    """Single-level drop extraction — returns {box_id: [drop_ids]}."""
    result = {}
    for row in ROW_RE.findall(present_text):
        bid = _get_tag(row,"Id")
        if bid not in box_ids: continue
        drops = []
        for i in range(20):
            did = _get_tag(row,f"DropId_{i}")
            if did and did.isdigit() and did!="0": drops.append(did)
        result[bid] = drops
    return result


def extract_drop_ids_recursive(present_text, start_box_ids):
    """
    Recursively extract ALL leaf-level item IDs reachable from start_box_ids
    through PresentItemParam2.
    Any drop ID that itself has a PresentItemParam2 row is a box — recurse into it.
    Returns (leaf_ids, all_box_ids_traversed) both as sets.
    """
    # Build full map of every present row: id -> [drop_ids]
    all_present = {}
    for row in ROW_RE.findall(present_text):
        bid = _get_tag(row, "Id")
        if not bid: continue
        drops = []
        for i in range(20):
            did = _get_tag(row, f"DropId_{i}")
            if did and did.isdigit() and did != "0": drops.append(did)
        if drops:
            all_present[bid] = drops

    leaf_ids = set()
    box_ids_traversed = set()
    queue = list(start_box_ids)
    visited = set(start_box_ids)

    while queue:
        bid = queue.pop()
        if bid not in all_present:
            # Not a box in PresentItemParam2 — it's a leaf item
            if bid not in start_box_ids:
                leaf_ids.add(bid)
            continue
        box_ids_traversed.add(bid)
        for did in all_present[bid]:
            if did in visited: continue
            visited.add(did)
            if did in all_present:
                # It's a sub-box — recurse
                queue.append(did)
            else:
                # It's a leaf item
                leaf_ids.add(did)

    return leaf_ids, box_ids_traversed


# ══════════════════════════════════════════════════════════════════════════════
# SESSION STORE  — shared state between tools
# ══════════════════════════════════════════════════════════════════════════════
class AppSession:
    """Central store. Tool 1 writes here; Tools 2-4 can import from it."""
    def __init__(self):
        self.box_id_list_csv  = None   # Tool1 output: box IDs CSV text
        self.box_id_map       = {}     # Tool1 output: {box_id: box_name}
        self.box_contents_csv = None   # Tool2 output (future)
        self.present_xml_path = None
        self.compound_rows     = []   # Tool7 compound row tuples
        self.exchange_rows     = []   # Tool7 exchange row tuples

# ══════════════════════════════════════════════════════════════════════════════
# GLOBAL APP SETTINGS  (gear menu, libconfig dir, filename overrides, timestamp)
# ══════════════════════════════════════════════════════════════════════════════

_APP_SETTINGS_PATH = os.path.expanduser("~/.box_tool_suite_app_settings.json")

_DEFAULT_APP_SETTINGS = {
    "libconfig_dir": os.path.join(os.getcwd(), "libconfig"),
    "timestamp_files": False,
    "filenames": {
        "itemparam":            "itemparam2.xml",
        "presentparam":         "presentitemparam2.xml",
        "compound_potion":      "Compound_Potion.xml",
        "compounder_location":  "Compounder_Location.xml",
        "exchange_contents":    "ExchangeShopContents.xml",
        "exchange_location":    "Exchange_Location.xml",
        "recycle_except":       "RecycleExceptItem.xml",
        "box_id_csv":           "box_id_list.csv",
        "set_item_param":       "CMSetItemParam.xml",
    }
}

def _load_app_settings():
    try:
        with open(_APP_SETTINGS_PATH, encoding="utf-8") as f:
            d = json.load(f)
        # Merge missing keys from defaults
        out = dict(_DEFAULT_APP_SETTINGS)
        out.update({k: v for k, v in d.items() if k != "filenames"})
        out["filenames"] = dict(_DEFAULT_APP_SETTINGS["filenames"])
        out["filenames"].update(d.get("filenames", {}))
        return out
    except:
        return dict(_DEFAULT_APP_SETTINGS)

def _save_app_settings(d):
    try:
        with open(_APP_SETTINGS_PATH, "w", encoding="utf-8") as f:
            json.dump(d, f, indent=2)
    except Exception as e:
        print(f"Could not save app settings: {e}")

# Runtime cache — loaded once at startup
def _ensure_settings_file():
    """Create the app settings JSON with defaults if it doesn't exist."""
    if not os.path.exists(_APP_SETTINGS_PATH):
        try:
            with open(_APP_SETTINGS_PATH, "w", encoding="utf-8") as _f:
                json.dump(_DEFAULT_APP_SETTINGS, _f, indent=2)
        except Exception:
            pass

_ensure_settings_file()
_APP_SETTINGS = _load_app_settings()

def _get_output_path(file_key):
    """Return full output path for a named file key, applying libconfig dir + optional timestamp."""
    s = _APP_SETTINGS
    base = s["filenames"].get(file_key, file_key + ".xml")
    if s.get("timestamp_files", False):
        import time
        ts = time.strftime("%d%m%y-%S%M%H")
        name, ext = os.path.splitext(base)
        base = f"{name}_{ts}{ext}"
    lib_dir = s.get("libconfig_dir") or os.path.join(os.getcwd(), "libconfig")
    os.makedirs(lib_dir, exist_ok=True)
    return os.path.join(lib_dir, base)

def _open_settings_window(root):
    """Gear ⚙ settings dialog."""
    win = tk.Toplevel(root)
    win.title("⚙  Toolbox Settings")
    win.configure(bg=BG)
    win.geometry("700x580")
    win.grab_set()

    s = _APP_SETTINGS  # reference to live dict

    tk.Label(win, text="⚙  Toolbox Settings", bg=BG, fg=BLUE,
             font=("Consolas", 14, "bold"), pady=10).pack()

    sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True, padx=12)
    canv, C = mk_scroll_canvas(sh)

    # ── Output directory ──────────────────────────────────────────────
    s_dir = mk_section(container, "  Output Directory  (libconfig folder)")
    dir_var = tk.StringVar(value=s.get("libconfig_dir", ""))
    r_dir = tk.Frame(s_dir, bg=BG); r_dir.pack(fill="x", padx=8, pady=6)
    dir_ent = tk.Entry(r_dir, textvariable=dir_var, width=52, bg=BG3, fg=FG,
                       insertbackground=FG, font=("Consolas", 9), relief="flat")
    dir_ent.pack(side="left", padx=(0, 6))
    def _pick_dir():
        d = filedialog.askdirectory(title="Select libconfig folder", parent=win)
        if d: dir_var.set(d)
    mk_btn(r_dir, "📂  Browse", _pick_dir, color=BG4, font=("Consolas", 9)).pack(side="left")
    tk.Label(s_dir, text="  XML output files will be written here. Created if it doesn't exist.",
             bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(anchor="w", padx=10, pady=(0, 4))

    # ── Timestamp option ──────────────────────────────────────────────
    s_ts = mk_section(container, "  File Naming")
    ts_var = tk.BooleanVar(value=s.get("timestamp_files", False))
    tk.Checkbutton(s_ts, text="Generate Unique File Every Time  (append DDMMYY-SSMMHH timestamp)",
                   variable=ts_var, bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                   font=("Consolas", 9)).pack(anchor="w", padx=10, pady=6)
    tk.Label(s_ts, text="  Example:  Compound_Potion_150325-421014.xml\n  Disabled = always overwrite the same file (default).",
             bg=BG, fg=FG_GREY, font=("Consolas", 8), justify="left").pack(anchor="w", padx=10, pady=(0,4))

    # ── Filename overrides ────────────────────────────────────────────
    s_fn = mk_section(container, "  XML Filename Overrides")
    fn_vars = {}
    fn_labels = {
        "itemparam":            "itemparam2.xml  (default)",
        "presentparam":         "presentitemparam2.xml  (default)",
        "compound_potion":      "Compound_Potion.xml  (default)",
        "compounder_location":  "Compounder_Spot.xml  (default)",
        "exchange_contents":    "ExchangeShopContents.xml  (default)",
        "exchange_location":    "Exchange_Location.xml  (default)",
        "recycle_except":       "RecycleExceptItem.xml  (default)",
        "box_id_csv":           "box_id_list.csv  (default)",
        "set_item_param":       "CMSetItemParam.xml  (default)",
        "shop_item":            "R_ShopItem.xml  (default)",
    }
    for key, placeholder in fn_labels.items():
        rr = tk.Frame(s_fn, bg=BG); rr.pack(fill="x", padx=8, pady=2)
        tk.Label(rr, text=f"{placeholder}:", width=38, anchor="w",
                 bg=BG, fg=FG_DIM, font=("Consolas", 8)).pack(side="left")
        v = tk.StringVar(value=s["filenames"].get(key, ""))
        tk.Entry(rr, textvariable=v, width=32, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
        fn_vars[key] = v

    # ── Reports directory ─────────────────────────────────────────────────
    s_rpt = mk_section(container, "  Reports Directory  (audit/compare outputs)")
    rpt_var = tk.StringVar(value=s.get("reports_dir", os.path.join(os.getcwd(), "reports")))
    r_rpt = tk.Frame(s_rpt, bg=BG); r_rpt.pack(fill="x", padx=8, pady=6)
    tk.Entry(r_rpt, textvariable=rpt_var, width=52, bg=BG3, fg=FG,
             insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=(0,6))
    def _pick_rpt():
        d = filedialog.askdirectory(title="Select reports folder", parent=win)
        if d: rpt_var.set(d)
    mk_btn(r_rpt, "📂  Browse", _pick_rpt, color=BG4, font=("Consolas",9)).pack(side="left")

    # ── MyShop output directory ────────────────────────────────────────────
    s_ms = mk_section(container, "  MyShop Directory  (SQL + libcmgds_e outputs)")
    ms_var = tk.StringVar(value=s.get("myshop_dir", os.path.join(os.getcwd(), "MyShop")))
    r_ms = tk.Frame(s_ms, bg=BG); r_ms.pack(fill="x", padx=8, pady=6)
    tk.Entry(r_ms, textvariable=ms_var, width=52, bg=BG3, fg=FG,
             insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=(0,6))
    def _pick_ms():
        d = filedialog.askdirectory(title="Select MyShop folder", parent=win)
        if d: ms_var.set(d)
    mk_btn(r_ms, "📂  Browse", _pick_ms, color=BG4, font=("Consolas",9)).pack(side="left")

    # ── UI / Performance ───────────────────────────────────────────────────
    s_perf = mk_section(container, "  UI & Performance")
    tt_var = tk.BooleanVar(value=s.get("tooltips_enabled", True))
    tk.Checkbutton(s_perf, text="Show field tooltips  (disable to speed up UI on slow machines)",
                   variable=tt_var, bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                   font=("Consolas",9)).pack(anchor="w", padx=10, pady=4)
    tk.Label(s_perf, text="  Tooltip changes apply immediately.",
             bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", padx=10, pady=(0,4))

    # ── Advanced Manual Renaming ───────────────────────────────────────────
    s_adv = mk_section(container, "  Advanced Manual Renaming  ⚠")
    adv_var = tk.BooleanVar(value=s.get("advanced_renaming_enabled", False))

    def _adv_toggle():
        if adv_var.get():
            messagebox.showwarning(
                "⚠  Advanced Renaming — Danger",
                "Advanced Manual Renaming uses regex and pattern replacement directly on XML field values.\n\n"
                "IMPORTANT: This can corrupt your XML structure, break CDATA sections, or introduce "
                "invalid characters if used incorrectly.\n\n"
                "Please be careful and always keep a backup before applying changes.\n\n"
                "A regex reference guide is available within each tool that supports this feature.",
                parent=win)
    tk.Checkbutton(s_adv,
                   text="Enable Advanced Manual Renaming  (regex / pattern replacement in manual input fields)",
                   variable=adv_var, bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                   font=("Consolas", 9), command=_adv_toggle).pack(anchor="w", padx=10, pady=4)
    tk.Label(s_adv,
             text="  When enabled: adds a regex/pattern replace option to manual input fields in supported tools.\n"
                  "  A 📖 Reference Guide button will appear. CDATA and XML structure are preserved where possible.\n"
                  "  ⚠  Use with caution — incorrect patterns can break XML output.",
             bg=BG, fg=FG_GREY, font=("Consolas", 8), justify="left").pack(anchor="w", padx=10, pady=(0, 4))


    # ── Variable Editor ──────────────────────────────────────────────────────
    def _open_var_editor(parent_win):
        """Pop-out editor for all persisted tool settings files."""
        ved = tk.Toplevel(parent_win)
        ved.title("🗂  Variable Editor — All Saved Tool Data")
        ved.configure(bg=BG)
        ved.geometry("860x620")

        tk.Label(ved, text="🗂  Variable Editor",
                 bg=BG, fg=BLUE, font=("Consolas", 13, "bold"), pady=6).pack()
        tk.Label(ved, text="  Edit, clear, or delete any persisted value. "
                           "Changes apply immediately when you press Save.",
                 bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(anchor="w", padx=14)

        sh2 = tk.Frame(ved, bg=BG); sh2.pack(fill="both", expand=True, padx=8, pady=4)
        canv2, C2 = mk_scroll_canvas(sh2)

        # Tool ordering for display
        TOOL_ORDER = [
            ("t6",             "ItemParam Generator (Tool 1)"),
            ("t18_fashion",    "Fashion Creation (Tool 3)"),
            ("t1_box",         "Box XML Generator (Tool 2)"),
            ("t1_myshop_ver",  "Box Generator — MyShop version_code"),
            ("t8_set",         "Set Item Generator (Tool 4)"),
            ("compound",       "Compound/Exchange/Shop (Tool 5) — Compound"),
            ("exchange",       "Compound/Exchange/Shop (Tool 5) — Exchange"),
            ("t6_item",        "ItemParam Generator (Tool 1) — last item ID"),
            ("t18_fashion_item","Fashion Creation — last fashion item ID"),
        ]

        # Load all settings files
        all_data = {}
        for key, label in TOOL_ORDER:
            all_data[key] = _load_settings(key)

        # Per-field editor widgets: {tool_key: {field_key: StringVar}}
        edit_vars = {}
        dirty = {}  # track which fields changed

        for tool_key, tool_label in TOOL_ORDER:
            data = all_data[tool_key]
            if not data:
                continue

            sec_f = tk.LabelFrame(C2, text=f"  {tool_label}  ",
                                   bg=BG, fg=ACC6, font=("Consolas", 8, "bold"),
                                   bd=1, relief="groove")
            sec_f.pack(fill="x", padx=8, pady=4)
            edit_vars[tool_key] = {}

            # Sort fields alphabetically
            for field_key in sorted(data.keys()):
                val = data[field_key]
                if isinstance(val, dict):
                    # Nested dict (e.g. filenames, items list) — show as JSON string
                    display = str(val)[:200]
                elif isinstance(val, list):
                    display = f"[{len(val)} items]"
                else:
                    display = str(val)

                row_f = tk.Frame(sec_f, bg=BG); row_f.pack(fill="x", padx=8, pady=1)
                tk.Label(row_f, text=f"{tool_key} → {field_key}:",
                         bg=BG, fg=FG, font=("Consolas", 8), width=36,
                         anchor="w").pack(side="left")
                sv = tk.StringVar(value=display)
                ent = tk.Entry(row_f, textvariable=sv, width=42, bg=BG3, fg=FG,
                               insertbackground=FG, font=("Consolas", 8),
                               relief="flat")
                ent.pack(side="left", padx=4)

                def _mark_dirty(e, tk_=tool_key, fk_=field_key):
                    dirty[(tk_, fk_)] = True

                ent.bind("<Key>", _mark_dirty)
                edit_vars[tool_key][field_key] = (sv, val)

                def _clear_field(tk_=tool_key, fk_=field_key, sv_=sv):
                    sv_.set("")
                    dirty[(tk_, fk_)] = True

                def _del_field(tk_=tool_key, fk_=field_key, r_=row_f):
                    if messagebox.askyesno("Delete field",
                            f"Delete {tk_} → {fk_}?", parent=ved):
                        d = _load_settings(tk_)
                        d.pop(fk_, None)
                        _save_settings(tk_, d)
                        r_.destroy()

                btn_f = tk.Frame(row_f, bg=BG); btn_f.pack(side="left")
                tk.Button(btn_f, text="✕ Clear", command=_clear_field,
                          bg=BG4, fg=FG_DIM, font=("Consolas", 7),
                          relief="flat").pack(side="left", padx=2)
                tk.Button(btn_f, text="🗑 Del", command=_del_field,
                          bg=BG4, fg=ACC3, font=("Consolas", 7),
                          relief="flat").pack(side="left", padx=2)

        # Nav
        nav2 = tk.Frame(ved, bg=BG2); nav2.pack(fill="x", side="bottom")

        def _save_all():
            count = 0
            for tool_key, fields in edit_vars.items():
                d = _load_settings(tool_key)
                changed = False
                for field_key, (sv, orig_val) in fields.items():
                    if (tool_key, field_key) in dirty:
                        new_str = sv.get()
                        # Try to coerce back to original type
                        if isinstance(orig_val, bool):
                            d[field_key] = new_str.lower() in ("true", "1", "yes")
                        elif isinstance(orig_val, int):
                            try: d[field_key] = int(new_str)
                            except: d[field_key] = new_str
                        elif isinstance(orig_val, float):
                            try: d[field_key] = float(new_str)
                            except: d[field_key] = new_str
                        elif isinstance(orig_val, (dict, list)):
                            pass  # don't write back complex structures from text
                        else:
                            d[field_key] = new_str
                        changed = True
                        count += 1
                if changed:
                    _save_settings(tool_key, d)
            messagebox.showinfo("Saved", f"{count} field(s) saved.", parent=ved)
            dirty.clear()
            ved.destroy()

        def _clear_tool():
            # Clear ALL data for a selected tool
            choices = [label for _, label in TOOL_ORDER]
            keys    = [key   for key, _ in TOOL_ORDER]
            # Simple dialog
            win3 = tk.Toplevel(ved); win3.title("Clear tool data"); win3.configure(bg=BG)
            tk.Label(win3, text="Select tool to clear ALL saved data:",
                     bg=BG, fg=FG, font=("Consolas", 9)).pack(pady=8, padx=12)
            for key, lbl in TOOL_ORDER:
                def _do_clear(k=key, l=lbl):
                    if messagebox.askyesno("Confirm", f"Clear ALL saved data for:\n{l}?", parent=win3):
                        _save_settings(k, {})
                        win3.destroy()
                        ved.destroy()
                        _open_var_editor(parent_win)
                tk.Button(win3, text=f"Clear  {lbl}", command=_do_clear,
                          bg=BG4, fg=ACC3, font=("Consolas", 8),
                          relief="flat").pack(fill="x", padx=12, pady=2)
            mk_btn(win3, "Cancel", win3.destroy, color=BG3).pack(pady=6)

        mk_btn(nav2, "💾  Save Changes", _save_all,
               color=GREEN, fg=BG2, font=("Consolas", 10, "bold")).pack(side="right", padx=14, pady=6)
        mk_btn(nav2, "🗑  Clear Tool Data…", _clear_tool,
               color=BG4).pack(side="left", padx=8, pady=6)
        mk_btn(nav2, "Cancel", ved.destroy, color=BG4).pack(side="right", padx=4, pady=6)

    mk_btn(C, "🗂  Open Variable Editor", lambda: _open_var_editor(win),
           color=BG3, font=("Consolas", 9)).pack(anchor="w", padx=10, pady=6)

    # ── Buttons ───────────────────────────────────────────────────────────
    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")

    def _save():
        s["libconfig_dir"]   = dir_var.get().strip() or os.path.join(os.getcwd(), "libconfig")
        s["reports_dir"]     = rpt_var.get().strip() or os.path.join(os.getcwd(), "reports")
        s["myshop_dir"]      = ms_var.get().strip() or os.path.join(os.getcwd(), "MyShop")
        s["tooltips_enabled"]= tt_var.get()
        s["timestamp_files"] = ts_var.get()
        s["advanced_renaming_enabled"] = adv_var.get()
        for key, v in fn_vars.items():
            val = v.get().strip()
            if val:
                s["filenames"][key] = val
            else:
                s["filenames"][key] = _DEFAULT_APP_SETTINGS["filenames"][key]
        _save_app_settings(s)
        messagebox.showinfo("Saved", f"Settings saved.\nOutput directory: {s['libconfig_dir']}", parent=win)
        win.destroy()

    def _reset_defaults():
        if messagebox.askyesno("Reset", "Reset ALL settings to defaults?", parent=win):
            _APP_SETTINGS.update(_DEFAULT_APP_SETTINGS)
            _save_app_settings(_APP_SETTINGS)
            win.destroy()
            _open_settings_window(root)

    mk_btn(nav, "💾  Save", _save, color=GREEN, fg=BG2,
           font=("Consolas", 10, "bold")).pack(side="right", padx=14, pady=6)
    mk_btn(nav, "Cancel", win.destroy, color=BG4).pack(side="right", padx=4, pady=6)
    mk_btn(nav, "↺  Reset to Defaults", _reset_defaults, color=BG4).pack(side="left", padx=14, pady=6)
    win.wait_window()


# ══════════════════════════════════════════════════════════════════════════════
# UI HELPERS — shared across all tool frames
# ══════════════════════════════════════════════════════════════════════════════
BG      = "#1e1e2e"
BG2     = "#181825"
BG3     = "#313244"
BG4     = "#45475a"
FG      = "#cdd6f4"
FG_DIM  = "#a6adc8"
FG_GREY = "#6c7086"
ACC1    = "#cba6f7"   # purple  — tool 1
ACC2    = "#89dceb"   # cyan    — tool 2
ACC3    = "#f38ba8"   # red     — tool 3
ACC4    = "#fab387"   # peach   — tool 4
ACC5    = "#f9e2af"   # yellow  — calculator
ACC6    = "#94e2d5"   # teal    — itemparam gen
GREEN   = "#a6e3a1"
BLUE    = "#89b4fa"

def mk_section(parent, title):
    f = tk.LabelFrame(parent, text=title, bg=BG, fg=BLUE,
                      font=("Consolas",10,"bold"), bd=1, relief="groove")
    f.pack(fill="x", padx=12, pady=5)
    return f

def mk_btn(parent, text, command, color=BG3, fg=FG, **kw):
    defaults = dict(font=("Consolas",10), relief="flat", padx=12, pady=6)
    defaults.update(kw)
    return tk.Button(parent, text=text, command=command, bg=color, fg=fg, **defaults)

def mk_scroll_canvas(parent, init_width=900):
    """Returns (canvas, container_frame). Container scrolls inside canvas.
    init_width: initial width hint so content renders before the first Configure event."""
    sb     = ttk.Scrollbar(parent, orient="vertical")
    canvas = tk.Canvas(parent, bg=BG, highlightthickness=0,
                       yscrollcommand=sb.set, width=init_width)
    sb.configure(command=canvas.yview)
    sb.pack(side="right", fill="y")
    canvas.pack(side="left", fill="both", expand=True)
    cont = tk.Frame(canvas, bg=BG)
    wid  = canvas.create_window((0,0), window=cont, anchor="nw", width=init_width)

    def _on_inner_resize(e):
        canvas.configure(scrollregion=canvas.bbox("all"))

    def _on_canvas_resize(e):
        canvas.itemconfig(wid, width=e.width)

    cont.bind("<Configure>", _on_inner_resize)
    canvas.bind("<Configure>", _on_canvas_resize)

    def _on_enter(e): canvas.bind_all("<MouseWheel>", lambda ev: canvas.yview_scroll(-1*(ev.delta//120),"units"))
    def _on_leave(e): canvas.unbind_all("<MouseWheel>")
    canvas.bind("<Enter>", _on_enter)
    canvas.bind("<Leave>", _on_leave)

    return canvas, cont

class ScrolledFrame(tk.Frame):
    """A frame with a vertical scrollbar. Place children in .inner.
    Uses a canvas with an explicit initial width so content renders immediately."""
    def __init__(self, parent, **kw):
        super().__init__(parent, **kw)
        self.configure(bg=BG)
        self._canvas = tk.Canvas(self, bg=BG, highlightthickness=0,
                                 borderwidth=0, width=900)
        self._sb = ttk.Scrollbar(self, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=self._sb.set)
        self._sb.pack(side="right", fill="y")
        self._canvas.pack(side="left", fill="both", expand=True)
        self.inner = tk.Frame(self._canvas, bg=BG)
        self._win = self._canvas.create_window((0, 0), window=self.inner,
                                               anchor="nw", width=900)
        self.inner.bind("<Configure>", self._on_inner)
        self._canvas.bind("<Configure>", self._on_canvas)
        self._canvas.bind("<Enter>",
            lambda e: self._canvas.bind_all("<MouseWheel>", self._scroll))
        self._canvas.bind("<Leave>",
            lambda e: self._canvas.unbind_all("<MouseWheel>"))

    def _on_inner(self, e):
        self._canvas.configure(scrollregion=self._canvas.bbox("all"))

    def _on_canvas(self, e):
        self._canvas.itemconfig(self._win, width=e.width)

    def _scroll(self, e):
        self._canvas.yview_scroll(-1 * (e.delta // 120), "units")

def make_output_tab(nb, title, content, fname, root):
    frm = tk.Frame(nb, bg=BG); nb.add(frm, text=title)
    br  = tk.Frame(frm, bg=BG); br.pack(side="bottom", fill="x")
    mk_btn(br, "📋 Copy All",
           lambda c=content: (root.clipboard_clear(), root.clipboard_append(c),
                               messagebox.showinfo("Copied","Copied to clipboard.")),
           padx=10, pady=4).pack(side="left", padx=6, pady=4)
    def _save(c=content, f=fname):
        p = filedialog.asksaveasfilename(initialfile=f, defaultextension=".xml",
                filetypes=[("XML","*.xml"),("CSV","*.csv"),("Text","*.txt"),("All","*.*")])
        if p:
            with open(p,"w",encoding="utf-8") as fh: fh.write(c)
            messagebox.showinfo("Saved",f"Saved to {p}")
    mk_btn(br,"💾 Save As…",_save,color=GREEN,fg=BG2,padx=10,pady=4).pack(side="left",padx=6,pady=4)
    txt = scrolledtext.ScrolledText(frm, font=("Consolas",9), bg=BG2, fg=FG)
    txt.pack(fill="both", expand=True, padx=4, pady=4)
    txt.insert("1.0", content); txt.config(state="disabled")


def _sheet_to_csv(sheet):
    """Convert an openpyxl worksheet to CSV text, preserving cell values."""
    out = io.StringIO()
    writer = csv.writer(out)
    for row in sheet.iter_rows(values_only=True):
        writer.writerow(["" if v is None else str(v) for v in row])
    return out.getvalue()

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 1 — Box XML Generator
# ══════════════════════════════════════════════════════════════════════════════
class Tool1(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self.saved_settings = None
        self.continue_mode  = None
        self.groups         = []
        self.current_group_idx = 0
        self.box_configs    = []
        self.no_csv_mode    = False
        self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    # ── Load screen ──────────────────────────────────────────────────────────
    def _build_load_screen(self):
        self._clear()
        frm = tk.Frame(self, bg=BG); frm.pack(expand=True)
        tk.Label(frm, text="BOX XML GENERATOR", font=("Consolas",20,"bold"),
                 bg=BG, fg=ACC1).pack(pady=(30,5))
        tk.Label(frm, text="Load a spreadsheet (CSV or Excel) to batch-generate box XML rows.",
                 bg=BG, fg=FG_DIM, font=("Consolas",10)).pack(pady=5)
        info_frm = tk.Frame(frm, bg=BG2); info_frm.pack(pady=6, padx=20, fill="x")
        tk.Label(info_frm,
            text=(
                "Column guide:\n"
                "  • Any non-keyword column header becomes a Box Group name.\n"
                "  • Recognised columns (any order, with or without _# suffix):\n"
                "      ID / BoxID                  — item ID (required)\n"
                "      Name / BoxName              — display name\n"
                "      Level / MinLevel / Lvl      — level shown in config screen\n"
                "      DropRate / DropRate_3       — drop rate (all slots, or slot 3)\n"
                "      Rate                        — alias for DropRate (all slots)\n"
                "      ItemCnt / ItemCnt_2 / Count — item count (all slots, or slot 2)\n"
                "      FileName / BundleNum        — NRI paths for this box\n"
                "      CmtFileName / CmtBundleNum  — portrait NRI\n"
                "      NCash / Tickets             — pricing (Tickets × 133 = NCash)\n"
                "      Options / ChrTypeFlags / Recycle / BoxID\n"
                "      Contents / Comment / Use\n"
                "  • \"Contents\" column: item names/IDs that go inside the box.\n"
                "  • Wide format: each non-keyword header = one box group.\n"
                "    e.g.  ID | Level | Dragon Box  |  ID | Level | Sheep Box  | …\n"
                "  • Tall format: one ID column + optional field columns.\n"
                "  • Excel: each sheet becomes a separate group."
            ),
            bg=BG2, fg=FG, font=("Consolas",8), justify="left",
            padx=10, pady=8).pack(anchor="w")
        bf = tk.Frame(frm, bg=BG); bf.pack(pady=15)
        mk_btn(bf,"📂  Load File (CSV / Excel)",self._load_csv_file).pack(side="left",padx=8)
        mk_btn(bf,"📋  Paste CSV Text",  self._paste_csv    ).pack(side="left",padx=8)
        mk_btn(bf,"✏️  No CSV — Manual Entry", self._start_no_csv,
               color=BG4).pack(side="left",padx=8)

    def _load_csv_file(self):
        types = [("Spreadsheet","*.csv *.xlsx *.xlsm *.xls"),("CSV","*.csv"),("Excel","*.xlsx *.xlsm *.xls"),("All","*.*")]
        p = filedialog.askopenfilename(filetypes=types)
        if not p: return
        ext = os.path.splitext(p)[1].lower()
        if ext in (".xlsx",".xlsm",".xls"):
            self._load_excel_file(p)
        else:
            with open(p, encoding="utf-8-sig") as f: self._process_csv(f.read())

    def _load_excel_file(self, path):
        if not _HAVE_OPENPYXL:
            messagebox.showerror("Missing library",
                "openpyxl is required to open Excel files.\n"
                "Install it with:  pip install openpyxl"); return
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            messagebox.showerror("Excel error", str(e)); return
        sheet_names = wb.sheetnames
        if not sheet_names:
            messagebox.showerror("Excel error","Workbook has no sheets."); return
        # If only one sheet, load it directly
        if len(sheet_names) == 1:
            self._process_csv(_sheet_to_csv(wb[sheet_names[0]])); return
        # Multiple sheets — ask user which ones to load
        win = tk.Toplevel(self.root); win.title("Select Sheets")
        win.configure(bg=BG); win.grab_set()
        tk.Label(win, text="Select sheets to import:", bg=BG, fg=FG,
                 font=("Consolas",11,"bold")).pack(pady=(14,6),padx=16,anchor="w")
        tk.Label(win, text="Each sheet becomes a separate group.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(padx=16,anchor="w")
        check_vars = []
        for name in sheet_names:
            v = tk.BooleanVar(value=True)
            tk.Checkbutton(win, text=name, variable=v, bg=BG, fg=FG,
                           selectcolor=BG3, activebackground=BG,
                           font=("Consolas",10)).pack(anchor="w",padx=20,pady=2)
            check_vars.append((name, v))
        def do_import():
            selected = [name for name,v in check_vars if v.get()]
            if not selected:
                messagebox.showwarning("Nothing selected","Select at least one sheet."); return
            # Combine selected sheets: convert each to CSV text and concatenate
            # Each sheet is treated as an independent CSV parse, results merged
            all_groups = []
            for name in selected:
                csv_text = _sheet_to_csv(wb[name])
                groups   = parse_grouped_csv(csv_text)
                all_groups.extend(groups)
            win.destroy()
            if not all_groups:
                messagebox.showerror("Error","No valid box groups found in selected sheets."); return
            self.groups = all_groups; self.current_group_idx = 0; self.box_configs = []
            self.continue_mode = None; self.saved_settings = None; self.no_csv_mode = False
            self._build_config_screen()
        bf2 = tk.Frame(win, bg=BG); bf2.pack(pady=12,padx=16,fill="x")
        mk_btn(bf2,"Import",do_import,color=GREEN,fg=BG2).pack(side="left",padx=4)
        mk_btn(bf2,"Cancel",win.destroy).pack(side="left",padx=4)
        win.wait_window()

    def _paste_csv(self):
        win = tk.Toplevel(self.root); win.title("Paste CSV")
        win.geometry("600x400"); win.configure(bg=BG)
        tk.Label(win, text="Paste CSV below:", bg=BG, fg=FG,
                 font=("Consolas",10)).pack(anchor="w",padx=10,pady=5)
        txt = scrolledtext.ScrolledText(win, font=("Consolas",9))
        txt.pack(fill="both",expand=True,padx=10,pady=5)
        def confirm():
            self._process_csv(txt.get("1.0","end")); win.destroy()
        mk_btn(win,"Confirm",confirm,color=GREEN,fg=BG2).pack(pady=8)

    def _process_csv(self, text):
        groups = parse_grouped_csv(text)
        if not groups:
            messagebox.showerror("Error","No valid box groups found in CSV."); return
        self.groups = groups; self.current_group_idx=0; self.box_configs=[]
        self.continue_mode=None; self.saved_settings=None; self.no_csv_mode=False
        self._extra_shop_rows = []
        self._build_config_screen()

    def _start_no_csv(self):
        self.no_csv_mode=True; self.current_group_idx=0; self.box_configs=[]
        self.continue_mode=None; self.saved_settings=None
        self.groups=[{"box_name":"Manual Entry","items":[
            {"id":"","name":"","extra":[],"rate":None,"item_cnt":1}]}]
        self._build_config_screen()

    # ── Config screen ─────────────────────────────────────────────────────────
    def _build_config_screen(self):
        self._clear()
        idx      = self.current_group_idx
        grp      = self.groups[idx]
        box_name = grp["box_name"]
        total    = len(self.groups)
        s        = self.saved_settings or {}
        prev_box_name = s.get("box_name","")

        # Wrapper uses grid so rows have explicit weights
        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0, weight=0)
        wrap.grid_rowconfigure(1, weight=1)
        wrap.grid_rowconfigure(2, weight=0)
        wrap.grid_columnconfigure(0, weight=1)

        # Header — row 0, fixed height
        hdr = tk.Frame(wrap, bg=BG2)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text=f"Box {idx+1} / {total}:  {box_name}",
                 font=("Consolas",14,"bold"), bg=BG2, fg=ACC1, pady=8
                 ).pack(side="left", padx=15)
        if self.continue_mode:
            tk.Label(hdr, text="🤖 AUTO" if self.continue_mode=="automate" else "👁 MONITOR",
                     font=("Consolas",10), bg=BG2, fg=ACC4).pack(side="right", padx=15)

        # Nav bar — row 2, always visible
        nav = tk.Frame(wrap, bg=BG2)
        nav.grid(row=2, column=0, sticky="ew")

        # Scrollable content — row 1, expands to fill
        scroll_host = tk.Frame(wrap, bg=BG)
        scroll_host.grid(row=1, column=0, sticky="nsew")
        canvas, container = mk_scroll_canvas(scroll_host)

        def section(title): return mk_section(container, title)
        def row_entry(p,lbl,var,w=38):
            r=tk.Frame(p,bg=BG); r.pack(fill="x",padx=8,pady=2)
            tk.Label(r,text=lbl,width=26,anchor="w",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
            tk.Entry(r,textvariable=var,width=w,bg=BG3,fg=FG,insertbackground=FG,
                     font=("Consolas",9),relief="flat").pack(side="left",padx=4)
        def row_num(p,lbl,var,w=10):
            r=tk.Frame(p,bg=BG); r.pack(fill="x",padx=8,pady=2)
            tk.Label(r,text=lbl,width=26,anchor="w",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
            tk.Entry(r,textvariable=var,width=w,bg=BG3,fg=FG,insertbackground=FG,
                     font=("Consolas",9),relief="flat").pack(side="left",padx=4)
        def note(p,txt):
            tk.Label(p,text=txt,bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(anchor="w",padx=10,pady=(3,0))

        # cfg_override from CSV — must be resolved FIRST before any StringVar uses it
        cfg_ov = grp.get("cfg_override", {})

        if s.get("_restore_id"):
            # Going backwards: use the exact ID the box had, no +1
            next_id = s["_restore_id"]
        else:
            try:    next_id = str(int(s.get("id",""))+1)
            except: next_id = s.get("id","")
            if not next_id:
                last = _get_last_id("t1_box", 0)
                next_id = str(last + 1) if last else ""
        if cfg_ov.get("id"): next_id = cfg_ov["id"]
        saved_name_tmpl = s.get("name_template","")
        if saved_name_tmpl and prev_box_name:
            proposed_name = apply_name_template(saved_name_tmpl, prev_box_name, box_name)
        else:
            proposed_name = box_name
        used_names = [c["name"] for c in self.box_configs]
        proposed_name = deduplicate_name(proposed_name, used_names)
        if cfg_ov.get("name"): proposed_name = deduplicate_name(cfg_ov["name"], used_names)
        saved_cmt_tmpl = s.get("comment_template","")
        initial_comment = apply_name_template(saved_cmt_tmpl,prev_box_name,box_name) if saved_cmt_tmpl and prev_box_name else s.get("comment","Special dice that contains amazing items.")
        if cfg_ov.get("comment"): initial_comment = cfg_ov["comment"]
        saved_use_tmpl = s.get("use_template","")
        initial_use = apply_name_template(saved_use_tmpl,prev_box_name,box_name) if saved_use_tmpl and prev_box_name else s.get("use","Event Box.")
        if cfg_ov.get("use"): initial_use = cfg_ov["use"]

        v_id=tk.StringVar(value=next_id); v_name=tk.StringVar(value=proposed_name)
        v_comment=tk.StringVar(value=initial_comment); v_use=tk.StringVar(value=initial_use)
        v_file_name=tk.StringVar(value=cfg_ov.get("file_name") or s.get("file_name",r"data\item\itm_pre_107.nri"))
        v_bundle_num=tk.StringVar(value=cfg_ov.get("bundle_num") or s.get("bundle_num","0"))
        v_cmt_file_name=tk.StringVar(value=cfg_ov.get("cmt_file_name") or s.get("cmt_file_name",r"data\item\itm_pre_illu_107.nri"))
        v_cmt_bundle=tk.StringVar(value=cfg_ov.get("cmt_bundle_num") or s.get("cmt_bundle_num","0"))
        ov_checks = cfg_ov.get("opt_checks", s.get("opt_checks",[False]*8))
        ov_recycle = cfg_ov.get("opt_recycle", s.get("opt_recycle",0))
        ov_chr     = cfg_ov.get("chr_type_flags", s.get("chr_type_flags",[]))
        opt_check_vars=[tk.BooleanVar(value=ov_checks[i] if i<len(ov_checks) else False) for i in range(8)]
        opt_recycle_var=tk.IntVar(value=ov_recycle)
        chr_selected=list(ov_chr)
        present_type_var=tk.IntVar(value=s.get("present_type",0))
        drop_cnt_var=tk.StringVar(value=s.get("drop_cnt","1"))
        default_rate_var=tk.StringVar(value=s.get("default_rate","50"))
        v_weight=tk.StringVar(value=cfg_ov.get("weight") or s.get("weight","1"))
        v_value=tk.StringVar(value=cfg_ov.get("value") or s.get("value","0"))
        v_min_level=tk.StringVar(value=cfg_ov.get("min_level") or s.get("min_level","1"))
        v_money=tk.StringVar(value=cfg_ov.get("money") or s.get("money","0"))
        v_ticket=tk.StringVar(value=cfg_ov.get("ticket") or s.get("ticket","0"))
        dc_mode_var=tk.StringVar(value=s.get("dc_mode","manual"))
        rate_mode_var=tk.StringVar(value=s.get("rate_mode","manual"))
        item_cnt_mode_var=tk.StringVar(value=s.get("item_cnt_mode","manual"))
        item_cnt_univ_var=tk.StringVar(value=s.get("item_cnt_univ","1"))

        live_items=list(grp["items"])
        item_rate_vars=[]; item_cnt_vars=[]
        for it in live_items:
            if present_type_var.get()==2: initial_r="100"
            elif it.get("rate") is not None: initial_r=str(it["rate"])
            else:
                try: initial_r=str(int(s.get("default_rate","50")))
                except: initial_r="50"
            item_rate_vars.append(tk.StringVar(value=initial_r))
            item_cnt_vars.append(tk.StringVar(value=str(it.get("item_cnt",1))))

        # Basic Info
        sb = section("  ItemParam – Basic Info  ")
        row_entry(sb,"Box ID (itemparam):",v_id,20)
        note(sb,"  Name auto-filled from CSV header — editable.")
        row_entry(sb,"Name (CDATA):",v_name)
        row_entry(sb,"Comment (CDATA):",v_comment)
        row_entry(sb,"Use (CDATA):",v_use)

        # Filepaths
        sf = section("  Filepaths & Bundle Numbers  ")
        note(sf,"  FileName — also written to InvFileName.")
        row_entry(sf,"FileName (CDATA):",v_file_name)
        row_num(sf,"BundleNum:",v_bundle_num,8)
        row_entry(sf,"CmtFileName (CDATA):",v_cmt_file_name)
        row_num(sf,"CmtBundleNum:",v_cmt_bundle,8)

        # Options
        so = section("  Options  ")
        chk_frm=tk.Frame(so,bg=BG); chk_frm.pack(anchor="w",padx=8,pady=4)
        for i,(lbl,_) in enumerate(OPTIONS_CHECKS):
            tk.Checkbutton(chk_frm,text=lbl,variable=opt_check_vars[i],
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                           font=("Consolas",9)).grid(row=i//4,column=i%4,sticky="w",padx=6,pady=2)
        rec_frm=tk.Frame(so,bg=BG); rec_frm.pack(anchor="w",padx=8,pady=4)
        tk.Label(rec_frm,text="Recycle:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left",padx=(0,8))
        for lbl,val in [("None",0),("Recyclable",262144),("Non-Recyclable",8388608)]:
            tk.Radiobutton(rec_frm,text=lbl,variable=opt_recycle_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                           font=("Consolas",9)).pack(side="left",padx=6)
        nc_container=tk.Frame(so,bg=BG); nc_container.pack(anchor="w",padx=8,pady=2)
        nc_inner=tk.Frame(nc_container,bg=BG)
        tk.Label(nc_inner,text="Ticket value (NCash = tickets × 133, rounded):",
                 bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Entry(nc_inner,textvariable=v_ticket,width=10,bg=BG3,fg=FG,
                 insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=8)
        ncash_lbl=tk.Label(nc_inner,text="→ NCash: 0",bg=BG,fg=GREEN,font=("Consolas",9))
        ncash_lbl.pack(side="left")
        def update_ncash(*_):
            try: ncash_lbl.config(text=f"→ NCash: {round(float(v_ticket.get())*133)}")
            except: ncash_lbl.config(text="→ NCash: ?")
        v_ticket.trace_add("write",update_ncash); update_ncash()
        def toggle_ncash(*_):
            if opt_recycle_var.get()==262144: nc_inner.pack(anchor="w")
            else: nc_inner.pack_forget()
        opt_recycle_var.trace_add("write",toggle_ncash); toggle_ncash()

        # ChrTypeFlags
        sc_chr=section("  ChrTypeFlags  ")
        pf=tk.Frame(sc_chr,bg=BG); pf.pack(fill="x",padx=8,pady=4)
        tk.Label(pf,text="Character Type:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        chr_name_combo=ttk.Combobox(pf,values=CHR_NAMES,state="readonly",width=14,font=("Consolas",9))
        chr_name_combo.pack(side="left",padx=(6,12))
        tk.Label(pf,text="Job:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        chr_job_combo=ttk.Combobox(pf,values=CHR_JOBS,state="readonly",width=6,font=("Consolas",9))
        chr_job_combo.pack(side="left",padx=(6,12))
        lb_frm=tk.Frame(sc_chr,bg=BG); lb_frm.pack(fill="x",padx=8,pady=(0,6))
        tk.Label(lb_frm,text="Added:",bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(anchor="w")
        chr_lb=tk.Listbox(lb_frm,height=4,width=36,bg=BG3,fg=FG,
                          font=("Consolas",9),selectbackground=BG4,activestyle="none")
        chr_lb.pack(anchor="w")
        def refresh_chr_lb():
            chr_lb.delete(0,"end")
            for val in chr_selected: chr_lb.insert("end",CHR_FLAG_REVERSE.get(val,str(val)))
        def add_chr_flag():
            name=chr_name_combo.get(); job=chr_job_combo.get()
            if not name or not job: return
            key=f"{name} {job}"; val=CHR_FLAG_MAP.get(key)
            if val and val not in chr_selected and len(chr_selected)<24:
                chr_selected.append(val); refresh_chr_lb()
        def rem_chr_flag():
            sel=chr_lb.curselection()
            if sel: chr_selected.pop(sel[0]); refresh_chr_lb()
        tk.Button(pf,text="+",command=add_chr_flag,bg=GREEN,fg=BG2,
                  font=("Consolas",11,"bold"),relief="flat",width=3).pack(side="left",padx=2)
        tk.Button(pf,text="−",command=rem_chr_flag,bg=ACC3,fg=BG2,
                  font=("Consolas",11,"bold"),relief="flat",width=3).pack(side="left",padx=2)
        refresh_chr_lb()

        # Numeric
        sn=section("  Numeric Fields  ")
        rf=tk.Frame(sn,bg=BG); rf.pack(fill="x",padx=8,pady=4)
        for ci,(lbl,var) in enumerate([("Weight:",v_weight),("Value:",v_value),
                                        ("MinLevel:",v_min_level),("Money:",v_money)]):
            tk.Label(rf,text=lbl,bg=BG,fg=FG,font=("Consolas",9),width=10,anchor="w").grid(row=0,column=ci*2,padx=4)
            tk.Entry(rf,textvariable=var,width=10,bg=BG3,fg=FG,insertbackground=FG,
                     font=("Consolas",9),relief="flat").grid(row=0,column=ci*2+1,padx=4)

        # PresentItemParam2 settings
        sp=section("  PresentItemParam2 Settings  ")
        id_disp=tk.Label(sp,text=f"Box ID: {next_id}  |  {box_name}",bg=BG,fg=FG_GREY,font=("Consolas",9))
        id_disp.pack(anchor="w",padx=8,pady=(4,0))
        v_id.trace_add("write",lambda *_: id_disp.config(text=f"Box ID: {v_id.get() or '—'}  |  {box_name}"))
        type_frm=tk.Frame(sp,bg=BG); type_frm.pack(anchor="w",padx=8,pady=4)
        tk.Label(type_frm,text="Drop Type:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        for lbl,val in [("Random",0),("Distributive",2)]:
            tk.Radiobutton(type_frm,text=lbl,variable=present_type_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                           font=("Consolas",9)).pack(side="left",padx=8)

        dc_mode_frm=tk.Frame(sp,bg=BG); dc_mode_frm.pack(anchor="w",padx=8,pady=(2,0))
        tk.Label(dc_mode_frm,text="DropCnt:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Radiobutton(dc_mode_frm,text="Flexible  (= item count)",variable=dc_mode_var,value="flexible",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_dc_mode()).pack(side="left",padx=6)
        tk.Radiobutton(dc_mode_frm,text="Manual",variable=dc_mode_var,value="manual",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_dc_mode()).pack(side="left",padx=4)
        dc_frm=tk.Frame(sp,bg=BG); dc_frm.pack(anchor="w",padx=8,pady=2)
        dc_lbl=tk.Label(dc_frm,text="DropCnt value:",bg=BG,fg=FG,font=("Consolas",9))
        dc_ent=tk.Entry(dc_frm,textvariable=drop_cnt_var,width=6,bg=BG3,fg=FG,
                        insertbackground=FG,font=("Consolas",9),relief="flat")
        rate_note=tk.Label(sp,text="",bg=BG,fg=FG_GREY,font=("Consolas",8))
        rate_note.pack(anchor="w",padx=8)
        def _toggle_dc_mode():
            if dc_mode_var.get()=="manual": dc_lbl.pack(side="left"); dc_ent.pack(side="left",padx=6)
            else: dc_lbl.pack_forget(); dc_ent.pack_forget()

        dr_mode_frm=tk.Frame(sp,bg=BG); dr_mode_frm.pack(anchor="w",padx=8,pady=(4,0))
        tk.Label(dr_mode_frm,text="DropRate:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Radiobutton(dr_mode_frm,text="Manual",variable=rate_mode_var,value="manual",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_rate_mode()).pack(side="left",padx=6)
        tk.Radiobutton(dr_mode_frm,text="Distributive  (set all entries equally)",
                       variable=rate_mode_var,value="distributive",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                       command=lambda:_toggle_rate_mode()).pack(side="left",padx=4)
        dr_dist_frm=tk.Frame(sp,bg=BG)
        tk.Label(dr_dist_frm,text="Default DropRate:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        dr_dist_ent=tk.Entry(dr_dist_frm,textvariable=default_rate_var,width=6,bg=BG3,fg=FG,
                             insertbackground=FG,font=("Consolas",9),relief="flat")
        dr_dist_ent.pack(side="left",padx=6)
        dr_dist_preview=tk.Label(dr_dist_frm,text="",bg=BG,fg=FG_GREY,font=("Consolas",8))
        dr_dist_preview.pack(side="left")
        def _apply_dist_rate(*_):
            if rate_mode_var.get()=="distributive":
                try:   val=str(int(default_rate_var.get()))
                except: val="50"
                for var in item_rate_vars: var.set(val)
                dr_dist_preview.config(text=f"→ all items set to {val}")
        default_rate_var.trace_add("write",_apply_dist_rate)

        # Box Contents table — defined BEFORE the ItemCnt radios so the command lambda works
        sec_items=tk.LabelFrame(container,text="  Box Contents  ",
                                bg=BG,fg=BLUE,font=("Consolas",10,"bold"),bd=1,relief="groove")
        sec_items.pack(fill="x",padx=12,pady=5)
        items_outer=tk.Frame(sec_items,bg=BG); items_outer.pack(fill="x",padx=8,pady=4)
        rate_entry_widgets=[]; icnt_entry_widgets=[]
        icnt_univ_frm=tk.Frame(sp,bg=BG)
        tk.Label(icnt_univ_frm,text="ItemCnt (all):",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        tk.Entry(icnt_univ_frm,textvariable=item_cnt_univ_var,width=6,bg=BG3,fg=FG,
                 insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=6)

        def _rebuild_items_table():
            nonlocal rate_entry_widgets, icnt_entry_widgets
            while len(item_rate_vars)<len(live_items): item_rate_vars.append(tk.StringVar(value=default_rate_var.get() or "50"))
            while len(item_rate_vars)>len(live_items): item_rate_vars.pop()
            while len(item_cnt_vars)<len(live_items): item_cnt_vars.append(tk.StringVar(value="1"))
            while len(item_cnt_vars)>len(live_items): item_cnt_vars.pop()
            for w in items_outer.winfo_children(): w.destroy()
            rate_entry_widgets=[]; icnt_entry_widgets=[]
            sec_items.config(text=f"  Box Contents ({len(live_items)} items)  ")
            icnt_mode=item_cnt_mode_var.get(); is_no_csv=self.no_csv_mode
            if icnt_mode=="universal": icnt_univ_frm.pack(anchor="w",padx=8,pady=(0,4))
            else: icnt_univ_frm.pack_forget()
            cols=[("#",3),("ID",9 if is_no_csv else 10),("Name",40 if is_no_csv else 44),("DropRate",8)]
            if icnt_mode=="manual": cols.append(("ItemCnt",7))
            if is_no_csv: cols.append(("",5))
            for ci,(txt,w) in enumerate(cols):
                tk.Label(items_outer,text=txt,width=w,bg=BG2,fg=BLUE,
                         font=("Consolas",9,"bold"),anchor="w").grid(row=0,column=ci,padx=2,pady=2,sticky="w")
            for i,item in enumerate(live_items):
                bg = BG if i%2==0 else BG2; col=0
                tk.Label(items_outer,text=str(i),width=3,bg=bg,fg=FG_GREY,
                         font=("Consolas",9)).grid(row=i+1,column=col,padx=2,pady=1); col+=1
                if is_no_csv:
                    id_var=tk.StringVar(value=item.get("id",""))
                    tk.Entry(items_outer,textvariable=id_var,width=9,bg=BG3,fg=FG,
                             insertbackground=FG,font=("Consolas",9),relief="flat").grid(row=i+1,column=col,padx=2)
                    id_var.trace_add("write",lambda *_,v=id_var,it=item: it.update({"id":v.get()}))
                    name_var=tk.StringVar(value=item.get("name",""))
                    tk.Entry(items_outer,textvariable=name_var,width=40,bg=BG3,fg=FG,
                             insertbackground=FG,font=("Consolas",9),relief="flat").grid(row=i+1,column=col+1,padx=2,sticky="w")
                    name_var.trace_add("write",lambda *_,v=name_var,it=item: it.update({"name":v.get()}))
                else:
                    tk.Label(items_outer,text=item.get("id",""),width=9,bg=bg,fg=FG,
                             font=("Consolas",9)).grid(row=i+1,column=col,padx=2)
                    tk.Label(items_outer,text=item.get("name","")[:52],width=44,bg=bg,fg=FG_DIM,
                             font=("Consolas",9),anchor="w").grid(row=i+1,column=col+1,padx=2,sticky="w")
                col+=2
                rate_ent=tk.Entry(items_outer,textvariable=item_rate_vars[i],width=8,bg=BG3,fg=FG,
                                  insertbackground=FG,font=("Consolas",9),relief="flat")
                rate_ent.grid(row=i+1,column=col,padx=2); rate_entry_widgets.append(rate_ent); col+=1
                if icnt_mode=="manual":
                    icnt_ent=tk.Entry(items_outer,textvariable=item_cnt_vars[i],width=7,bg=BG3,fg=FG,
                                      insertbackground=FG,font=("Consolas",9),relief="flat")
                    icnt_ent.grid(row=i+1,column=col,padx=2); icnt_entry_widgets.append(icnt_ent); col+=1
                if is_no_csv:
                    def make_remove(idx):
                        def _rem():
                            if len(live_items)<=1: return
                            live_items.pop(idx)
                            if idx<len(item_rate_vars): item_rate_vars.pop(idx)
                            if idx<len(item_cnt_vars): item_cnt_vars.pop(idx)
                            _rebuild_items_table()
                        return _rem
                    tk.Button(items_outer,text="−",width=2,command=make_remove(i),
                              bg=ACC3,fg=BG2,font=("Consolas",9,"bold"),relief="flat"
                              ).grid(row=i+1,column=col,padx=2)
            if is_no_csv and len(live_items)<20:
                def _add_item():
                    live_items.append({"id":"","name":"","extra":[],"rate":None,"item_cnt":1})
                    _rebuild_items_table()
                tk.Button(items_outer,text="＋ Add row",command=_add_item,
                          bg=GREEN,fg=BG2,font=("Consolas",9),relief="flat",padx=6,pady=2
                          ).grid(row=len(live_items)+1,column=0,columnspan=6,sticky="w",padx=4,pady=4)
            _toggle_rate_mode_inner()

        # ItemCnt mode radios — placed here so _rebuild_items_table is already defined
        icnt_mode_frm=tk.Frame(sp,bg=BG); icnt_mode_frm.pack(anchor="w",padx=8,pady=(4,2))
        tk.Label(icnt_mode_frm,text="ItemCnt:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        for _lbl_ic,_val_ic in [("Flexible (=1)","flexible"),("Universal","universal"),("Manual (per row)","manual")]:
            tk.Radiobutton(icnt_mode_frm,text=_lbl_ic,variable=item_cnt_mode_var,value=_val_ic,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9),
                           command=_rebuild_items_table).pack(side="left",padx=4)

        def _toggle_rate_mode_inner():
            is_dist=rate_mode_var.get()=="distributive"
            for ent in rate_entry_widgets:
                ent.config(state="disabled" if is_dist else "normal",disabledforeground=FG_GREY)
            if is_dist: _apply_dist_rate()

        def _toggle_rate_mode():
            if rate_mode_var.get()=="distributive":
                dr_dist_frm.pack(anchor="w",padx=8,pady=2); _toggle_rate_mode_inner()
            else:
                dr_dist_frm.pack_forget()
                for ent in rate_entry_widgets: ent.config(state="normal")
                dr_dist_preview.config(text="")

        def toggle_drop_fields(*_):
            _toggle_dc_mode(); rate_note.config(text="")

        _rebuild_items_table()
        present_type_var.trace_add("write",toggle_drop_fields)
        _toggle_dc_mode(); toggle_drop_fields(); _toggle_rate_mode()

        # ── MyShop Generation ────────────────────────────────────────────────
        _s_myshop = self.saved_settings or {}
        _last_ver_raw = _get_last_id("t1_myshop_ver", 0)

        myshop_sec = mk_section(container, "  MyShop Generation  (libcmgds_e + SQL)  ")
        ms_hdr = tk.Frame(myshop_sec, bg=BG); ms_hdr.pack(fill="x", padx=8, pady=4)
        myshop_var = tk.BooleanVar(value=_s_myshop.get("myshop_enabled", False))
        def _toggle_ms_body(*_):
            if myshop_var.get(): ms_body.pack(fill="x", padx=8, pady=2)
            else: ms_body.pack_forget()
        tk.Checkbutton(ms_hdr, text="Generate MyShop files for this box group",
                       variable=myshop_var, bg=BG, fg=FG, selectcolor=BG3,
                       activebackground=BG, font=("Consolas",9),
                       command=_toggle_ms_body).pack(side="left")
        tk.Label(ms_hdr, text="  (exports to MyShop folder)", bg=BG, fg=FG_GREY,
                 font=("Consolas",8)).pack(side="left")

        ms_body = tk.Frame(myshop_sec, bg=BG)
        # Only show body when checkbox is enabled
        if myshop_var.get(): ms_body.pack(fill="x", padx=8, pady=2)

        def _row(lbl, var, width=14, tip=""):
            r = tk.Frame(ms_body, bg=BG); r.pack(fill="x", pady=2)
            lw = tk.Label(r, text=lbl, width=22, anchor="w", bg=BG, fg=FG, font=("Consolas",9))
            lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=width, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas",9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip and _APP_SETTINGS.get("tooltips_enabled", True):
                _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent

        myshop_price_var = tk.StringVar(value=_s_myshop.get("myshop_price", "0"))
        myshop_level_var = tk.StringVar(value=_s_myshop.get("myshop_level", "0"))
        myshop_ver_var   = tk.StringVar(value=str(_s_myshop.get("myshop_ver", _last_ver_raw + 1)))

        _row("goods_cash_price:", myshop_price_var, tip="MyShop NCash price for this box.")
        _row("goods_char_level:", myshop_level_var, tip="Level requirement (0 = none). ItemParam MinLevel will be 1 unless this is set higher.")
        _row("version_code:", myshop_ver_var, tip="Auto-incremented from last used. Edit freely.")
        myshop_item_count_var = tk.StringVar(value=str(_s_myshop.get("myshop_item_count", "1")))
        _row("item_count (tbl_list):", myshop_item_count_var, tip="Number of this item per MyShop listing (tbl_goods_list item_count).")
        _last_glc_t1 = _get_last_id("t1_goods_list_code", 21040)
        # Pre-fill: prev box's glc + 1, same pattern as ID field.
        # If saved_settings has a glc from the previous box, use that +1.
        # Falls back to last persisted session value +1 if no prior box.
        # Pre-fill logic:
        # - If previous box had a glc saved, next box = that value + 1
        # - If no previous box (box 1), use last persisted + 1 as the starting value
        if "myshop_glc_start" in _s_myshop:
            try:    _next_glc = str(int(_s_myshop["myshop_glc_start"]) + 1)
            except: _next_glc = str(_last_glc_t1 + 1)
        else:
            _next_glc = str(_last_glc_t1 + 1)
        myshop_glc_var = tk.StringVar(value=_next_glc)
        _row("goods_list_code:", myshop_glc_var, tip="goods_list_code/parents_list_code for this box in libcmgds_e. Auto-increments +1 from previous box. Override freely.")

        # Category + stamp row
        cat_row = tk.Frame(ms_body, bg=BG); cat_row.pack(fill="x", pady=4)
        tk.Label(cat_row, text="Box category:", width=22, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left")
        myshop_cat_var = tk.StringVar(value=_s_myshop.get("myshop_cat", "special"))
        tk.Radiobutton(cat_row, text="Special Box  (goods_category2=2)",
                       variable=myshop_cat_var, value="special",
                       bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                       font=("Consolas",9)).pack(side="left", padx=6)
        tk.Radiobutton(cat_row, text="New Item  (goods_category2=0)",
                       variable=myshop_cat_var, value="new",
                       bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                       font=("Consolas",9)).pack(side="left", padx=6)

        stamp_row = tk.Frame(ms_body, bg=BG); stamp_row.pack(fill="x", pady=4)
        tk.Label(stamp_row, text="Stamps:", width=22, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left")
        myshop_new_var = tk.BooleanVar(value=bool(_s_myshop.get("myshop_new", False)))
        myshop_pop_var = tk.BooleanVar(value=bool(_s_myshop.get("myshop_pop", False)))
        tk.Checkbutton(stamp_row, text="New Stamp  (goods_shop_new=1)",
                       variable=myshop_new_var, bg=BG, fg=FG, selectcolor=BG3,
                       activebackground=BG, font=("Consolas",9)).pack(side="left", padx=6)
        tk.Checkbutton(stamp_row, text="Popular Stamp  (goods_shop_popular=1)",
                       variable=myshop_pop_var, bg=BG, fg=FG, selectcolor=BG3,
                       activebackground=BG, font=("Consolas",9)).pack(side="left", padx=6)

        tk.Label(myshop_sec,
                 text=(
                     "  goods_category=0, goods_category0=6, goods_category1=3, goods_limit_use=2,\n"
                     "  goods_char_sex=0, goods_char_type=15, goods_limit_desc='All Characters'"
                 ),
                 bg=BG, fg=FG_GREY, font=("Consolas",7)).pack(anchor="w", padx=10, pady=(0,4))

        # Gather / nav
        def gather_config():
            try: ncash=round(float(v_ticket.get() or "0")*133)
            except: ncash=0
            item_rates=[]
            for iv in item_rate_vars:
                try: item_rates.append(int(iv.get()))
                except: item_rates.append(100 if present_type_var.get()==2 else 50)
            icnt_mode=item_cnt_mode_var.get()
            if icnt_mode=="flexible": item_cnts=[1]*len(live_items)
            elif icnt_mode=="universal":
                try: uval=int(item_cnt_univ_var.get()) or 1
                except: uval=1
                item_cnts=[uval]*len(live_items)
            else:
                item_cnts=[]
                for iv in item_cnt_vars:
                    try: item_cnts.append(int(iv.get()) or 1)
                    except: item_cnts.append(1)
            return {
                "id":v_id.get().strip(),"name":v_name.get(),
                "name_template":v_name.get(),"comment":v_comment.get(),
                "comment_template":v_comment.get(),"use":v_use.get(),
                "use_template":v_use.get(),"box_name":box_name,
                "items":list(live_items),"item_rates":item_rates,"item_cnts":item_cnts,
                "item_cnt_mode":icnt_mode,"item_cnt_univ":item_cnt_univ_var.get() or "1",
                "file_name":v_file_name.get(),"bundle_num":v_bundle_num.get() or "0",
                "cmt_file_name":v_cmt_file_name.get(),"cmt_bundle_num":v_cmt_bundle.get() or "0",
                "weight":v_weight.get() or "1","value":v_value.get() or "0",
                "min_level":v_min_level.get() or "1","money":v_money.get() or "0",
                "ncash":ncash,"ticket":v_ticket.get() or "0",
                "opt_checks":[bv.get() for bv in opt_check_vars],
                "opt_recycle":opt_recycle_var.get(),"chr_type_flags":list(chr_selected),
                "present_type":present_type_var.get(),
                "drop_cnt":drop_cnt_var.get() or "1" if dc_mode_var.get()=="manual" else str(len(live_items)),
                "default_rate":default_rate_var.get() or "50",
                "dc_mode":dc_mode_var.get(),"rate_mode":rate_mode_var.get(),
                "myshop_enabled":     myshop_var.get(),
                "myshop_price":       myshop_price_var.get().strip() or "0",
                "myshop_cat":         myshop_cat_var.get(),
                "myshop_new":         int(myshop_new_var.get()),
                "myshop_pop":         int(myshop_pop_var.get()),
                "myshop_level":       myshop_level_var.get().strip() or "0",
                "myshop_ver":         myshop_ver_var.get().strip() or "1",
                "myshop_item_count":  myshop_item_count_var.get().strip() or "1",
                "myshop_glc_start":   myshop_glc_var.get().strip() or "21000",
            }

        def save_settings(cfg):
            self.saved_settings={k:v for k,v in cfg.items() if k!="items"}

        # nav is already created above, pinned to bottom of outer

        def go_next():
            cfg=gather_config()
            if not cfg["id"]: messagebox.showwarning("Missing ID","Please enter a Box ID."); return
            blanks=[]
            if not cfg["name"].strip(): blanks.append("Name")
            if not cfg["file_name"].strip(): blanks.append("FileName")
            if not cfg["cmt_file_name"].strip(): blanks.append("CmtFileName")
            if blanks:
                if not messagebox.askyesno("Missed a spot","These fields are empty:\n\n  "+"\n  ".join(blanks)+"\n\nContinue anyway?"): return
            self.box_configs.append(cfg); save_settings(cfg); self.current_group_idx+=1
            if self.no_csv_mode or self.current_group_idx>=len(self.groups): self._build_output_screen()
            elif self.continue_mode=="automate": self._automate_remaining(cfg)
            elif self.continue_mode=="monitor": self._build_config_screen()
            else: self._ask_automate_or_monitor(cfg)

        mk_btn(nav,"◀  Back to CSV",self._build_load_screen).pack(side="left",padx=10,pady=8)
        if idx>0:
            def go_prev():
                self.current_group_idx -= 1
                # Pop the box we're returning to so it can be re-confirmed.
                # Store its exact cfg as saved_settings, with _restore_id set so
                # _build_config_screen uses the id directly instead of doing +1.
                if self.box_configs:
                    popped = self.box_configs.pop()
                    restored = {k: v for k, v in popped.items() if k != "items"}
                    restored["_restore_id"] = popped.get("id", "")
                    self.saved_settings = restored
                else:
                    self.saved_settings = None
                self._build_config_screen()
            mk_btn(nav,"◀  Previous Box",go_prev).pack(side="left",padx=4,pady=8)
        if self.continue_mode:
            mk_btn(nav,"⚙ Change Mode",lambda: (setattr(self,"continue_mode",None),self._build_config_screen()),
                   color=BG4).pack(side="left",padx=4,pady=8)

        # ── Live-validating Generate/Next button ──────────────────────────
        next_lbl = "⚡  Generate XML ✓" if idx==total-1 else "Next Box ▶"
        next_btn = mk_btn(nav, next_lbl, go_next, color=GREEN, fg=BG2,
                          font=("Consolas",10,"bold"))
        next_btn.pack(side="right", padx=10, pady=8)

        # Validation label (shows what's missing)
        val_lbl = tk.Label(nav, text="", bg=BG2, fg=ACC3, font=("Consolas",8))
        val_lbl.pack(side="right", padx=4)

        def _check_ready(*_):
            missing = []
            if not v_id.get().strip():           missing.append("ID")
            if not v_name.get().strip():          missing.append("Name")
            if not v_file_name.get().strip():     missing.append("FileName")
            if not v_cmt_file_name.get().strip(): missing.append("CmtFileName")
            # Check that at least one item has an ID
            has_item = any(it.get("id","").strip() for it in live_items)
            if not has_item:                      missing.append("Box Contents (need ≥1 item ID)")
            if missing:
                next_btn.config(state="disabled", bg=BG4, fg=FG_GREY)
                val_lbl.config(text="Missing: " + ", ".join(missing))
            else:
                next_btn.config(state="normal", bg=GREEN, fg=BG2)
                val_lbl.config(text="✓ Ready")

        # Trace required vars
        for _var in (v_id, v_name, v_file_name, v_cmt_file_name):
            _var.trace_add("write", _check_ready)
        # Re-check after item table rebuilds (hook into _rebuild_items_table)
        _orig_rebuild = _rebuild_items_table
        def _rebuild_and_check():
            _orig_rebuild()
            self.after(50, _check_ready)
        # Patch live_items check — just schedule it
        self.after(100, _check_ready)

    def _ask_automate_or_monitor(self, last_cfg):
        remaining=len(self.groups)-self.current_group_idx
        win=tk.Toplevel(self.root); win.title("Continue?"); win.geometry("530x260")
        win.configure(bg=BG); win.grab_set()
        tk.Label(win,text=f"{remaining} box(es) remaining.",bg=BG,fg=FG,font=("Consolas",13,"bold")).pack(pady=14)
        tk.Label(win,text="How would you like to continue?",bg=BG,fg=FG_DIM,font=("Consolas",10)).pack()
        remember_var=tk.BooleanVar(value=False)
        bf=tk.Frame(win,bg=BG); bf.pack(pady=10)
        def choose(mode):
            if remember_var.get(): self.continue_mode=mode
            win.destroy()
            if mode=="automate": self._automate_remaining(last_cfg)
            else: self._build_config_screen()
        mk_btn(bf,"🤖  Automate  —  use saved settings for all remaining boxes",
               lambda:choose("automate"),color=ACC1,fg=BG2).pack(pady=5)
        mk_btn(bf,"👁  Monitor  —  review each box individually",
               lambda:choose("monitor"),color=BLUE,fg=BG2).pack(pady=5)
        tk.Checkbutton(win,text="Remember my choice for the rest of this session",
                       variable=remember_var,bg=BG,fg=ACC4,selectcolor=BG3,
                       activebackground=BG,font=("Consolas",9)).pack(pady=4)

    def _automate_remaining(self, last_cfg):
        try: base_id=int(last_cfg["id"])
        except:
            messagebox.showerror("Error","Cannot automate: last Box ID was not a plain integer.")
            self.continue_mode="monitor"; self._build_config_screen(); return
        try: default_rate=int(last_cfg.get("default_rate",50))
        except: default_rate=50
        id_counter=base_id+1; prev_name=last_cfg["box_name"]
        used_names=[c["name"] for c in self.box_configs]
        for i in range(self.current_group_idx, len(self.groups)):
            grp=self.groups[i]; cfg=copy.deepcopy(last_cfg)
            cfg_ov=grp.get("cfg_override",{})

            # ── ID ────────────────────────────────────────────────────────
            # CSV-provided ID wins; otherwise auto-increment from last
            cfg["id"] = cfg_ov.get("id") or str(id_counter)
            try: id_counter = int(cfg["id"]) + 1
            except: id_counter += 1

            # ── Name ──────────────────────────────────────────────────────
            # CSV name col > box_name header directly > template substitution
            # The column header IS the canonical name for this box.
            if cfg_ov.get("name"):
                proposed = deduplicate_name(cfg_ov["name"], used_names)
            else:
                raw_box = grp["box_name"]
                tmpl = cfg.get("name_template", "")
                sub  = apply_name_template(tmpl, prev_name, raw_box)
                # Use template result only if it actually transformed the name
                # (i.e. prev_name appeared in the template and was replaced).
                # Otherwise fall back to the raw box header name.
                if sub and sub != tmpl and prev_name and prev_name.lower() in tmpl.lower():
                    proposed = deduplicate_name(sub, used_names)
                else:
                    proposed = deduplicate_name(raw_box, used_names)
            cfg["name"]=proposed; cfg["name_template"]=proposed

            # ── Comment / Use ─────────────────────────────────────────────
            # CSV field wins; otherwise apply template substitution
            if cfg_ov.get("comment"):
                cfg["comment"]=cfg_ov["comment"]; cfg["comment_template"]=cfg_ov["comment"]
            else:
                cfg["comment"]=apply_name_template(cfg.get("comment_template",""),prev_name,grp["box_name"])
                cfg["comment_template"]=cfg["comment"]
            if cfg_ov.get("use"):
                cfg["use"]=cfg_ov["use"]; cfg["use_template"]=cfg_ov["use"]
            else:
                cfg["use"]=apply_name_template(cfg.get("use_template",""),prev_name,grp["box_name"])
                cfg["use_template"]=cfg["use"]

            # ── Per-box CSV field overrides ───────────────────────────────
            # Any field the CSV provides for this specific group wins over the
            # persisted last-config value. This is the critical fix: every group
            # in the wide-format sheet has its own FileName/CmtFileName/BundleNum
            # etc, and those must be applied rather than carrying the first box's
            # values across all 250 boxes.
            for field_key in ("file_name", "bundle_num",
                              "cmt_file_name", "cmt_bundle_num",
                              "weight", "value", "min_level", "money",
                              "ticket", "opt_checks", "opt_recycle",
                              "chr_type_flags", "ncash_direct"):
                if field_key in cfg_ov:
                    cfg[field_key] = cfg_ov[field_key]

            # ── Items / rates / counts ─────────────────────────────────────
            cfg["box_name"]=grp["box_name"]; cfg["items"]=grp["items"]
            is_distrib=cfg["present_type"]==2
            cfg["item_rates"]=[100 if is_distrib else (it.get("rate") if it.get("rate") is not None else default_rate) for it in grp["items"]]
            icnt_mode=cfg.get("item_cnt_mode","manual")
            if icnt_mode=="flexible": cfg["item_cnts"]=[1]*len(grp["items"])
            elif icnt_mode=="universal":
                try: uval=int(cfg.get("item_cnt_univ","1")) or 1
                except: uval=1
                cfg["item_cnts"]=[uval]*len(grp["items"])
            else: cfg["item_cnts"]=[it.get("item_cnt",1) for it in grp["items"]]
            used_names.append(proposed); prev_name=grp["box_name"]
            self.box_configs.append(cfg)
        self.current_group_idx=len(self.groups); self._build_output_screen()

    def _build_output_screen(self, _compound_rows=None, _exchange_rows=None):
        self._clear()
        if _compound_rows is None: _compound_rows = []
        if _exchange_rows is None: _exchange_rows = []

        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0,weight=0); wrap.grid_rowconfigure(1,weight=1)
        wrap.grid_rowconfigure(2,weight=0); wrap.grid_columnconfigure(0,weight=1)
        hdr = tk.Frame(wrap, bg=BG2); hdr.grid(row=0,column=0,sticky="ew")
        tk.Label(hdr,text="Generated XML Output",font=("Consolas",14,"bold"),
                 bg=BG2,fg=ACC1,pady=8).pack(side="left",padx=15)

        nb_host = tk.Frame(wrap, bg=BG); nb_host.grid(row=1,column=0,sticky="nsew")
        nb=ttk.Notebook(nb_host); nb.pack(fill="both",expand=True,padx=8,pady=4)

        itemparam_rows=[]; presentparam_rows=[]; recycle_except_rows=[]
        include_present = {}   # cfg["id"] -> bool

        for cfg in self.box_configs:
            try: default_rate=int(cfg.get("default_rate",50))
            except: default_rate=50
            try: drop_cnt=int(cfg.get("drop_cnt",1))
            except: drop_cnt=1
            is_distrib=(cfg["present_type"]==2)
            items_with_rates=[]
            for j,it in enumerate(cfg["items"]):
                rate=100 if is_distrib else (cfg["item_rates"][j] if j<len(cfg["item_rates"]) else default_rate)
                items_with_rates.append({**it,"rate":rate})
            item_cnts=cfg.get("item_cnts") or [1]*len(cfg["items"])
            itemparam_rows.append(build_itemparam_row(cfg))
            presentparam_rows.append(build_presentparam_row(cfg["id"],items_with_rates,cfg["present_type"],drop_cnt,default_rate,item_cnts=item_cnts,box_name=cfg.get("name","")))
            include_present[cfg["id"]] = True
            if cfg.get("opt_recycle",0) in (0,8388608):
                recycle_except_rows.append(build_recycle_except_row(cfg["id"],cfg["name"]))

        csv_lines=["ID,BoxName"]+[f"{c['id']},{c['box_name']}" for c in self.box_configs]
        csv_text="\n".join(csv_lines)
        self.session.box_id_list_csv = csv_text
        self.session.box_id_map = {c["id"]: c["box_name"] for c in self.box_configs}

        # Save last box ID used
        if self.box_configs:
            try: _set_last_id("t1_box", int(self.box_configs[-1]["id"]))
            except: pass

        _exports=[]
        _exports.append(("itemparam_rows.xml","\n".join(itemparam_rows)))
        make_output_tab(nb,"itemparam.xml rows","\n".join(itemparam_rows),"itemparam_rows.xml",self.root)

        _exports.append(("presentparam_rows.xml","\n".join(presentparam_rows)))
        make_output_tab(nb,"PresentItemParam2.xml rows","\n".join(presentparam_rows),"presentparam_rows.xml",self.root)

        _exports.append(("box_id_list.csv",csv_text))
        make_output_tab(nb,"Box ID List (→ Tool 2)","\n".join(csv_lines),"box_id_list.csv",self.root)

        if recycle_except_rows:
            _exports.append(("RecycleExceptItem_rows.xml","\n".join(recycle_except_rows)))
            make_output_tab(nb,"RecycleExceptItem.xml rows","\n".join(recycle_except_rows),"RecycleExceptItem_rows.xml",self.root)

        # ── MyShop outputs (libcmgds_e + SQL) ───────────────────────────────────
        _myshop_xml_blocks, _myshop_sql_rows, _myshop_final_glc = _build_box_myshop_outputs(self.box_configs)
        _myshop_exports = []   # list of (fname, text) for MyShop dir
        if _myshop_xml_blocks:
            libcmgds_full = "\n".join(_myshop_xml_blocks)
            sql_full      = "\n".join(_myshop_sql_rows)
            make_output_tab(nb, "libcmgds_e (GOODS)", libcmgds_full,
                            "box_libcmgds_e.xml", self.root)
            make_output_tab(nb, "SQL (goods/list/limit)", sql_full,
                            "box_goods_sql.sql", self.root)
            _myshop_exports = [("box_libcmgds_e.xml", libcmgds_full),
                               ("box_goods_sql.sql",  sql_full)]
            # Persist last version_code used
            for cfg in self.box_configs:
                if cfg.get("myshop_enabled"):
                    try: _set_last_id("t1_myshop_ver", int(cfg.get("myshop_ver", 1)))
                    except: pass
            # Persist final goods_list_code so next session starts where this one ended
            try: _set_last_id("t1_goods_list_code", _myshop_final_glc - 1)
            except: pass

        # Compound/exchange tabs
        if _compound_rows:
            cp_xml = "\n".join(r[0] for r in _compound_rows)
            cl_xml = "\n".join(r[1] for r in _compound_rows)
            _exports += [("Compound_Potion_rows.xml",cp_xml),("Compounder_Spot_rows.xml",cl_xml)]
            make_output_tab(nb,"Compound_Potion rows",cp_xml,"Compound_Potion_rows.xml",self.root)
            make_output_tab(nb,"Compounder_Location rows",cl_xml,"Compounder_Spot_rows.xml",self.root)
        if _exchange_rows:
            es_xml = "\n".join(r[0] for r in _exchange_rows)
            el_xml = "\n".join(r[1] for r in _exchange_rows)
            _exports += [("ExchangeShopContents_rows.xml",es_xml),("Exchange_Location_rows.xml",el_xml)]
            make_output_tab(nb,"ExchangeShopContents rows",es_xml,"ExchangeShopContents_rows.xml",self.root)
            make_output_tab(nb,"Exchange_Location rows",el_xml,"Exchange_Location_rows.xml",self.root)
        _shop_extra = getattr(self, "_extra_shop_rows", [])
        if _shop_extra:
            sh_xml = "\n".join(_shop_extra)
            _exports.append(("R_ShopItem_rows.xml", sh_xml))
            make_output_tab(nb,"R_ShopItem rows",sh_xml,"R_ShopItem_rows.xml",self.root)

        nb.select(0)

        # ── Nav footer ────────────────────────────────────────────────────
        bot = tk.Frame(wrap, bg=BG2); bot.grid(row=2,column=0,sticky="ew")

        # Compound/exchange post-generate button
        _ce_remember = tk.StringVar(value="ask")   # "ask"|"compound"|"exchange"|"none"
        def _open_ce_for_box(cfg_box):
            """Open compound/exchange dialog for a single box config."""
            def _on_compound(ce_cfg):
                cpr = build_compound_row(ce_cfg)
                clr = build_compound_location_row(ce_cfg["compound_id"])
                _compound_rows.append((cpr, clr))
                self.session.compound_rows = list(_compound_rows)
                messagebox.showinfo("Added", f"Compound row added.\nRegenerate to see updated tabs.")
            def _on_exchange(ce_cfg):
                esr = build_exchange_row(ce_cfg)
                elr = build_exchange_location_row(ce_cfg["exchange_id"])
                _exchange_rows.append((esr, elr))
                self.session.exchange_rows = list(_exchange_rows)
                messagebox.showinfo("Added", f"Exchange row added.")
            _show_compound_exchange_dialog(
                self.root, cfg_box.get("name",""), cfg_box.get("comment",""),
                cfg_box.get("id",""), _on_compound, _on_exchange, lambda: None)

        def _add_ce():
            _show_box_ce_dialog(
                self.root, self.box_configs,
                _compound_rows, _exchange_rows,
                lambda: self._build_output_screen(_compound_rows, _exchange_rows)
            )

        def export_all():
            saved = []
            # Use libconfig dir from settings; let user override
            default_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(), "libconfig"))
            folder = filedialog.askdirectory(title="Choose export folder (default: libconfig)",
                                             initialdir=default_dir)
            if not folder: folder = default_dir
            os.makedirs(folder, exist_ok=True)
            for fname, content in _exports:
                # Apply timestamp if enabled
                if _APP_SETTINGS.get("timestamp_files", False):
                    import time as _time
                    ts = _time.strftime("%d%m%y-%S%M%H")
                    name, ext = os.path.splitext(fname)
                    fname_out = f"{name}_{ts}{ext}"
                else:
                    fname_out = fname
                with open(os.path.join(folder, fname_out), "w", encoding="utf-8") as f:
                    f.write(content)
                saved.append(fname_out)
            messagebox.showinfo("Export Complete", f"Saved to:\n{folder}\n\n" + "\n".join(saved))

        def export_myshop():
            if not _myshop_exports:
                messagebox.showinfo("No MyShop data",
                    "No boxes have MyShop enabled.\nEnable it on the config screen."); return
            ms_dir = _APP_SETTINGS.get("myshop_dir", os.path.join(os.getcwd(), "MyShop"))
            folder = filedialog.askdirectory(
                title="Export MyShop files to…", initialdir=ms_dir, parent=self.root)
            if not folder: folder = ms_dir
            os.makedirs(folder, exist_ok=True)
            saved2 = []
            for fname, content in _myshop_exports:
                if _APP_SETTINGS.get("timestamp_files", False):
                    import time as _time
                    ts = _time.strftime("%d%m%y-%S%M%H")
                    n2, e2 = os.path.splitext(fname)
                    fname = f"{n2}_{ts}{e2}"
                with open(os.path.join(folder, fname), "w", encoding="utf-8") as f:
                    f.write(content)
                saved2.append(fname)
            messagebox.showinfo("MyShop Export",
                f"Saved to:\n{folder}\n\n" + "\n".join(saved2))

        mk_btn(bot,"💾  Export All Files",export_all,color=ACC1,fg=BG2,
               font=("Consolas",11,"bold")).pack(side="left",padx=14,pady=6)
        # Update session so other tools can import
        def _publish_session():
            csv_lines=["ID,BoxName"]+[f"{c['id']},{c['box_name']}" for c in self.box_configs]
            self.session.box_id_list_csv = "\n".join(csv_lines)
            self.session.box_id_map = {c["id"]: c["box_name"] for c in self.box_configs}
            messagebox.showinfo("Session Updated", f"Session updated with {len(self.box_configs)} box ID(s).\nTools 2-5 can now Import Session.")
        mk_btn(bot,"🔗  Publish to Session",_publish_session,color=BG4).pack(side="left",padx=4,pady=6)
        if _myshop_exports:
            mk_btn(bot,"🏪  Export MyShop",export_myshop,color=ACC5,fg=BG2,
                   font=("Consolas",9,"bold")).pack(side="left",padx=4,pady=6)
        mk_btn(bot,"⚗  Add Compound/Exchange",_add_ce,color=BG4).pack(side="left",padx=4,pady=6)
        def _import_ce_t1():
            def _on_compound_cfgs(cfgs):
                for ce_cfg in cfgs:
                    cpr = build_compound_row(ce_cfg)
                    clr = build_compound_location_row(ce_cfg["compound_id"])
                    _compound_rows.append((cpr, clr))
                messagebox.showinfo("Imported",
                    f"{len(cfgs)} compound row(s) added.\nRegenerate to see updated tabs.")
            def _on_exchange_cfgs(cfgs):
                for ce_cfg in cfgs:
                    esr = build_exchange_row(ce_cfg)
                    elr = build_exchange_location_row(ce_cfg["exchange_id"])
                    _exchange_rows.append((esr, elr))
                messagebox.showinfo("Imported",
                    f"{len(cfgs)} exchange row(s) added.\nRegenerate to see updated tabs.")
            _ask_import_mode_then_file(self.root, _on_compound_cfgs, _on_exchange_cfgs)
        mk_btn(bot,"📥  Import CSV/Excel",_import_ce_t1,color=BG4).pack(side="left",padx=4,pady=6)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4,pady=6)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 2 — PresentItemParam2 Rate Adjuster
# ══════════════════════════════════════════════════════════════════════════════
class Tool2(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session; self.csv_text=""; self.xml_text=""
        self.item_lib={}; self.mode_var=tk.StringVar(value="automatic")
        self._rate_var=tk.StringVar(value="100"); self._count_var=tk.StringVar(value="1")
        self._lib_status=tk.StringVar(value="No library loaded  (item names won't appear)")
        # Nested mode: "direct" = CSV IDs are the PresentItemParam2 <Id> rows to adjust
        #              "nested" = CSV IDs are contents of outer boxes; find THOSE IDs in XML
        self._nested_var=tk.StringVar(value="direct")
        self._mode_panel_frame=None
        # Optional box-param rewrite vars
        self._rewrite_params = tk.BooleanVar(value=False)
        self._rewrite_type   = tk.IntVar(value=2)       # 0=Random, 2=Egalitarian
        self._rewrite_dc_mode= tk.StringVar(value="flexible")  # flexible|manual
        self._rewrite_dc_val = tk.StringVar(value="1")
        self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _build_load_screen(self):
        self._clear()
        # Scrollable wrapper — use _c (content frame) not self to avoid shadowing instance
        _outer = tk.Frame(self, bg=BG); _outer.pack(fill="both", expand=True)
        _canv  = tk.Canvas(_outer, bg=BG, bd=0, highlightthickness=0)
        _vsb   = tk.Scrollbar(_outer, orient="vertical", command=_canv.yview)
        _canv.configure(yscrollcommand=_vsb.set)
        _vsb.pack(side="right", fill="y"); _canv.pack(side="left", fill="both", expand=True)
        _c = tk.Frame(_canv, bg=BG)
        _cw = _canv.create_window((0,0), window=_c, anchor="nw")
        _c.bind("<Configure>", lambda e: _canv.configure(scrollregion=_canv.bbox("all")))
        _canv.bind("<Configure>", lambda e: _canv.itemconfig(_cw, width=e.width))
        _canv.bind_all("<MouseWheel>", lambda e: _canv.yview_scroll(int(-1*(e.delta/120)), "units"))
        tk.Label(_c,text="PRESENTITEMPARAM2 RATE ADJUSTER",font=("Consolas",16,"bold"),
                 bg=BG,fg=ACC2).pack(pady=(18,2))
        tk.Label(self,
                 text="Adjusts DropRate_# and ItemCnt_# slots in PresentItemParam2.xml rows.\n"
                      "Direct mode: CSV IDs match the <Id> rows to adjust directly.\n"
                      "Nested mode: CSV IDs are items inside outer boxes — finds those\n"
                      "             IDs as <Id> rows in XML and adjusts their drop rates.",
                 bg=BG,fg=FG_GREY,font=("Consolas",8),justify="center").pack(pady=(0,4))

        # ── Column info ───────────────────────────────────────────────────
        _col_info2 = tk.Frame(_c, bg=BG2); _col_info2.pack(pady=(0,6), padx=20, fill="x")
        tk.Label(_col_info2,
            text=(
                "  SUPPORTED CSV FORMATS — tall or wide, any column order:\n"
                "  Required:  ID / BoxID / ItemID               — the box/item ID\n"
                "  Optional:  Name / BoxName                    — display name (shown in UI)\n"
                "  Optional:  Level / MinLevel / Lvl            — level (shown in header)\n"
                "  Optional:  DropRate / DropRate_3 / Rate      — pre-fill drop rate for that slot\n"
                "               DropRate with no number = apply to all slots\n"
                "               DropRate_3 = apply only to slot 3\n"
                "  Optional:  ItemCnt / ItemCnt_2 / Count       — pre-fill item count per slot\n"
                "\n"
                "  WIDE FORMAT (your CSV):  repeating [ID, Level, BoxName] groups\n"
                "    Dragon Low Gear Box  |  Sheep Low Gear Box  | …\n"
                "    ID | Level | Box     |  ID | Level | Box    |\n"
                "    727011 | 60 | …      |  727011 | 60 | …     |\n"
                "  → Each column-group = one outer box; IDs below = its contents\n"
                "\n"
                "  TALL FORMAT:  ID column + optional Name / DropRate / Level columns\n"
                "  Excel (.xlsx/.xlsm/.xls) also supported."
            ),
            bg=BG2, fg=FG, font=("Consolas",8), justify="left", padx=8, pady=6
        ).pack(anchor="w")

        # ── Step 1: CSV ───────────────────────────────────────────────────
        csv_frm=mk_section(_c,"  Step 1 — Box ID CSV  ")
        csv_status=tk.StringVar(value="No file loaded")
        tk.Label(csv_frm,textvariable=csv_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_csv():
            p=filedialog.askopenfilename(
                filetypes=[("Spreadsheet","*.csv *.xlsx *.xlsm *.xls"),
                           ("CSV","*.csv"),("Excel","*.xlsx *.xlsm *.xls"),("All","*.*")])
            if not p: return
            ext=os.path.splitext(p)[1].lower()
            if ext in (".xlsx",".xlsm",".xls"):
                try:
                    import openpyxl as _opx
                    wb=_opx.load_workbook(p,data_only=True)
                    import csv as _csv,io as _io
                    out=_io.StringIO()
                    for row in wb.worksheets[0].iter_rows(values_only=True):
                        _csv.writer(out).writerow(["" if v is None else str(v) for v in row])
                    self.csv_text=out.getvalue()
                except Exception as e:
                    messagebox.showerror("Excel error",str(e)); return
            else:
                with open(p,encoding="utf-8-sig") as f: self.csv_text=f.read()
            groups=parse_box_csv_groups(self.csv_text)
            total_ids=sum(len(items) for _,_,items in groups)
            csv_status.set(f"✓  {os.path.basename(p)}  ({len(groups)} group(s), {total_ids} IDs)")
        mk_btn(csv_frm,"📂 Load CSV/Excel",load_csv,padx=10,pady=4).pack(side="right",padx=8,pady=6)
        def import_session_t2():
            if not self.session.box_id_list_csv:
                messagebox.showinfo("No Session Data","Run Tool 1 first to generate box IDs.")
                return
            self.csv_text=self.session.box_id_list_csv
            bm=parse_box_id_csv(self.csv_text)
            csv_status.set(f"✓  Imported from Tool 1  ({len(bm)} box IDs)")
        mk_btn(csv_frm,"⬇  Import from Tool 1",import_session_t2,color=ACC2,fg=BG2,
               padx=8,pady=4).pack(side="right",padx=4,pady=6)

        # ── Step 2: XML ───────────────────────────────────────────────────
        xml_frm=mk_section(_c,"  Step 2 — PresentItemParam2.xml  ")
        xml_status=tk.StringVar(value="No file loaded")
        tk.Label(xml_frm,textvariable=xml_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_xml():
            p=filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig") as f: self.xml_text=f.read()
            n_rows=len(ROW_RE.findall(self.xml_text))
            xml_status.set(f"✓  {os.path.basename(p)}  ({n_rows} ROW entries)")
        mk_btn(xml_frm,"📂 Load XML",load_xml,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        # ── Step 3: Nested mode ───────────────────────────────────────────
        nest_sec=mk_section(_c,"  Step 3 — Lookup Mode  ")
        tk.Label(nest_sec,
                 text="  Choose how to match CSV IDs to PresentItemParam2 rows:",
                 bg=BG,fg=FG,font=("Consolas",9)).pack(anchor="w",padx=10,pady=(4,2))

        nf=tk.Frame(nest_sec,bg=BG); nf.pack(anchor="w",padx=10,pady=4)
        tk.Radiobutton(nf,text="Direct  — CSV IDs are the <Id> rows to adjust (default)",
                       variable=self._nested_var,value="direct",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                       font=("Consolas",10)).pack(anchor="w",pady=2)
        tk.Radiobutton(nf,text="Nested  — CSV IDs are contents of outer boxes; find those IDs in XML and adjust their DropRates",
                       variable=self._nested_var,value="nested",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,
                       font=("Consolas",10)).pack(anchor="w",pady=2)

        tk.Label(nest_sec,
                 text="  Direct:  CSV IDs = the <Id> in PresentItemParam2 to adjust directly.\n"
                      "           Use when your CSV IS the box list (e.g. Tool 1 export).\n"
                      "\n"
                      "  Nested:  CSV IDs = contents (drop items) inside outer boxes.\n"
                      "           Tool looks up each ID as an <Id> row in the XML and adjusts\n"
                      "           its DropRate_# slots. Use when your CSV has IDs of the\n"
                      "           inner boxes sitting inside your outer boxes.\n"
                      "\n"
                      "  Example (Nested): Dragon Low Gear Box → [727011, 432047, 727014]\n"
                      "  Finds 727011, 432047, 727014 as <Id> rows and adjusts their rates.",
                 bg=BG,fg=FG_GREY,font=("Consolas",8),justify="left").pack(anchor="w",padx=14,pady=(0,6))

        # ── Step 4: Mode ──────────────────────────────────────────────────
        mode_frm=mk_section(_c,"  Step 4 — Adjustment Mode  ")
        mf=tk.Frame(mode_frm,bg=BG); mf.pack(anchor="w",padx=10,pady=6)
        for lbl,val in [("Manual     — review and configure each box individually","manual"),
                        ("Automatic  — apply the same values to every matched box","automatic")]:
            tk.Radiobutton(mf,text=lbl,variable=self.mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10),
                           command=self._refresh_mode_panel).pack(anchor="w",pady=2)

        self._mode_panel_frame=tk.Frame(_c,bg=BG)
        self._mode_panel_frame.pack(fill="x",padx=30,pady=2)
        self._refresh_mode_panel()

        # ── Step 5: Optional box parameter rewrite ────────────────────────
        rw_sec = mk_section(_c, "  Step 5 — Rewrite Box Parameters  (optional)  ")
        rw_hdr = tk.Frame(rw_sec, bg=BG); rw_hdr.pack(fill="x", padx=8, pady=4)
        rw_body = tk.Frame(rw_sec, bg=BG)
        def _toggle_rw_body(*_):
            if self._rewrite_params.get(): rw_body.pack(fill="x", padx=14, pady=(0,6))
            else: rw_body.pack_forget()
        tk.Checkbutton(rw_hdr,
                       text="Also rewrite Type / DropCnt on matched rows",
                       variable=self._rewrite_params, command=_toggle_rw_body,
                       bg=BG, fg=ACC2, selectcolor=BG3, activebackground=BG,
                       font=("Consolas",9)).pack(side="left")
        tk.Label(rw_hdr, text="  (leave unchecked to only adjust rates/counts)",
                 bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(side="left")
        # Type
        rw_type_row = tk.Frame(rw_body, bg=BG); rw_type_row.pack(anchor="w", pady=(4,2))
        tk.Label(rw_type_row, text="Drop Type:", width=14, anchor="w",
                 bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        for lbl, val in [("Random  (Type=0)", 0), ("Egalitarian  (Type=2)", 2)]:
            tk.Radiobutton(rw_type_row, text=lbl, variable=self._rewrite_type, value=val,
                           bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                           font=("Consolas",9)).pack(side="left", padx=10)
        # DropCnt mode
        rw_dc_row = tk.Frame(rw_body, bg=BG); rw_dc_row.pack(anchor="w", pady=2)
        tk.Label(rw_dc_row, text="DropCnt:", width=14, anchor="w",
                 bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        rw_dc_ent = tk.Entry(rw_body, textvariable=self._rewrite_dc_val, width=6,
                             bg=BG3, fg=FG, insertbackground=FG,
                             font=("Consolas",9), relief="flat")
        def _toggle_dc_rw(*_):
            if self._rewrite_dc_mode.get() == "manual":
                rw_dc_ent.pack(anchor="w", padx=14, pady=(0,4))
            else:
                rw_dc_ent.pack_forget()
        for lbl, val in [("Flexible  (= number of drop slots)", "flexible"),
                         ("Manual", "manual")]:
            tk.Radiobutton(rw_dc_row, text=lbl, variable=self._rewrite_dc_mode, value=val,
                           bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                           font=("Consolas",9), command=_toggle_dc_rw).pack(side="left", padx=8)
        _toggle_rw_body()

        mk_btn(_c,"▶  Continue →",self._on_continue,color=GREEN,fg=BG2,
               font=("Consolas",12,"bold")).pack(pady=14)

    def _refresh_mode_panel(self):
        if not self._mode_panel_frame: return
        for w in self._mode_panel_frame.winfo_children(): w.destroy()
        if self.mode_var.get()=="automatic":
            frm=tk.LabelFrame(self._mode_panel_frame,text="  Adjustment Values  ",
                              bg=BG,fg=BLUE,font=("Consolas",10,"bold"),bd=1,relief="groove")
            frm.pack(fill="x")
            def num_row(lbl,var,note=""):
                r=tk.Frame(frm,bg=BG); r.pack(fill="x",padx=10,pady=4)
                tk.Label(r,text=lbl,width=18,anchor="w",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
                tk.Entry(r,textvariable=var,width=8,bg=BG3,fg=FG,insertbackground=FG,
                         font=("Consolas",9),relief="flat").pack(side="left",padx=6)
                tk.Label(r,text=note,bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(side="left")
            num_row("Adjust Rate:",self._rate_var,"(1–32766)  applied to every used DropRate_# slot")
            num_row("Adjust Count:",self._count_var,"(1–32766)  applied to every used ItemCnt_# slot")
            tk.Label(frm,text="  Type will be set to 2.  DropCnt will be set to the number of real items.",
                     bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(anchor="w",padx=10,pady=(0,6))
        else:
            frm=tk.LabelFrame(self._mode_panel_frame,text="  ItemParam Library (optional)  ",
                              bg=BG,fg=BLUE,font=("Consolas",10,"bold"),bd=1,relief="groove")
            frm.pack(fill="x")
            tk.Label(frm,textvariable=self._lib_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10,pady=6)
            def load_lib():
                folder=filedialog.askdirectory(title="Select folder containing ItemParam XML files")
                if not folder: return
                self.item_lib=load_itemparam_folder(folder)
                self._lib_status.set(f"✓  {len(self.item_lib)} items from {os.path.basename(folder)}")
            mk_btn(frm,"📂 Load Folder",load_lib,padx=10,pady=4).pack(side="right",padx=8,pady=6)

    def _on_continue(self):
        if not self.csv_text: messagebox.showwarning("Missing","Please load a CSV first."); return
        if not self.xml_text: messagebox.showwarning("Missing","Please load PresentItemParam2.xml first."); return

        nested = self._nested_var.get() == "nested"

        if nested:
            # Nested mode: CSV contains inner-box IDs (contents of outer boxes).
            # We find those inner IDs as <Id> rows in the XML and adjust them.
            groups = parse_box_csv_groups(self.csv_text)
            # Collect all inner item IDs across all groups, with their names
            inner_map = {}  # id -> (name, meta, group_name)
            for outer_id, outer_name, items in groups:
                for iid, nm, meta in items:
                    if iid and iid not in inner_map:
                        inner_map[iid] = (nm or iid, meta, outer_name)
            if not inner_map:
                messagebox.showerror("Error","No item IDs found in CSV."); return
            # Match inner IDs against XML rows
            matched=[]; seen=set()
            for row in ROW_RE.findall(self.xml_text):
                rid=_get_tag(row,"Id")
                if rid in inner_map and rid not in seen:
                    nm2, meta2, grp = inner_map[rid]
                    matched.append((rid, nm2, row, meta2, grp)); seen.add(rid)
            if not matched:
                # Build helpful message showing which IDs were in CSV vs XML
                csv_sample = list(inner_map.keys())[:5]
                xml_ids = [_get_tag(r,"Id") for r in ROW_RE.findall(self.xml_text)[:5]]
                csv_groups_shown = [(grp, list(inner_map.keys())[:3])
                                    for _, grp, _ in list(inner_map.values())[:2]]
                messagebox.showwarning("No Matches",
                    f"Nested mode: none of the {len(inner_map)} CSV item IDs matched any <Id> in XML.\n\n"
                    f"CSV sample IDs: {csv_sample}\n"
                    f"XML sample IDs: {xml_ids}\n\n"
                    "Make sure the inner-box IDs in your CSV match <Id> values in PresentItemParam2.xml.\n"
                    "If the CSV lists the outer box IDs directly, use Direct mode instead."); return
        else:
            # Direct mode: CSV IDs are the PresentItemParam2 <Id> rows themselves
            box_map=parse_box_id_csv(self.csv_text)
            if not box_map: messagebox.showerror("Error","No box IDs found in CSV."); return
            # Also capture meta from groups for per-row droprate overrides
            _direct_meta = {}
            for _, grp_name, items in parse_box_csv_groups(self.csv_text):
                for iid, nm, meta in items:
                    if iid and iid not in _direct_meta:
                        _direct_meta[iid] = (meta, grp_name)
            matched=[]; seen=set()
            for row in ROW_RE.findall(self.xml_text):
                rid=_get_tag(row,"Id")
                if rid in box_map and rid not in seen:
                    meta2, grp2 = _direct_meta.get(rid, ({}, ""))
                    matched.append((rid, box_map[rid], row, meta2, grp2)); seen.add(rid)
            if not matched:
                messagebox.showwarning("No Matches",
                    "None of the CSV box IDs matched any <Id> in the XML.\n\n"
                    "Make sure the IDs in your CSV match the <Id> values in PresentItemParam2.xml.\n\n"
                    "Tip: Use Nested mode if your CSV lists the contents inside the outer boxes."); return

        if self.mode_var.get()=="automatic":
            try:
                rate=int(self._rate_var.get()); count=int(self._count_var.get())
                if not (1<=rate<=32766) or not (1<=count<=32766): raise ValueError
            except: messagebox.showerror("Invalid","Rate and Count must be integers 1–32766."); return
            self._run_automatic(matched,rate,count)
        else: self._run_manual(matched)

    def _run_automatic(self, matched, rate, count):
        # matched is list of (rid, name, row_block, meta, group_name)
        matched_map={rid: (row, meta) for rid,_,row,meta,_ in matched}; csv_rows=[]
        def replace_row(m):
            row=m.group(0); rid=_get_tag(row,"Id")
            if rid not in matched_map: return row
            orig_row, meta = matched_map[rid]
            slots=real_drop_slots(row)
            slot_cfgs=[]
            for pos,(sidx,_) in enumerate(slots):
                # Check if CSV provides specific rate/count for this slot
                slot_rate  = meta.get(f"droprate_{sidx}", meta.get("droprate", meta.get("rate", rate)))
                slot_count = meta.get(f"itemcnt_{sidx}",  meta.get("itemcnt",  count))
                try: slot_rate  = int(slot_rate)
                except: slot_rate  = rate
                try: slot_count = int(slot_count)
                except: slot_count = count
                slot_cfgs.append({"rate": slot_rate, "count": slot_count})
            if self._rewrite_params.get():
                rw_type = self._rewrite_type.get()
                if self._rewrite_dc_mode.get() == "manual":
                    try: rw_dc = int(self._rewrite_dc_val.get())
                    except: rw_dc = len(slots)
                else:
                    rw_dc = len(slots)
            else:
                rw_type = 2; rw_dc = len(slots)
            cfg={"type":rw_type,"drop_cnt":rw_dc,"slots":slot_cfgs}
            new_row=apply_cfg_to_row(row,cfg)
            drop_ids=[v for _,v in real_drop_slots(new_row)]
            name=next((n for r,n,_,_,_ in matched if r==rid),"")
            csv_rows.append([rid,name,*drop_ids]); return new_row
        full_out=ROW_RE.sub(replace_row,self.xml_text)
        self._build_output_screen(full_out,csv_rows,len(matched))

    def _run_manual(self, matched):
        # matched is list of (rid, name, row_block, meta, group_name)
        # Normalise to ensure 5-tuple even if called from older code paths
        normalised = []
        for item in matched:
            if len(item) == 3:
                normalised.append(item + ({}, ""))
            else:
                normalised.append(item)
        self.manual_matched=normalised; self.manual_idx=0
        self.manual_configs={}; self.manual_saved=None; self.manual_continue_mode=None
        self._build_manual_screen()

    def _build_manual_screen(self):
        self._clear()
        idx=self.manual_idx; total=len(self.manual_matched)
        _item = self.manual_matched[idx]
        rid, csv_name, row_block = _item[0], _item[1], _item[2]
        meta = _item[3] if len(_item) > 3 else {}
        group_name = _item[4] if len(_item) > 4 else ""
        slots=real_drop_slots(row_block)
        s=self.manual_saved or {}
        # If optional param rewrite is enabled and no prior manual save, seed from load-screen settings
        if self._rewrite_params.get() and not self.manual_saved:
            last_type = self._rewrite_type.get()
            try: last_dc = int(self._rewrite_dc_val.get()) if self._rewrite_dc_mode.get()=="manual" else len(slots)
            except: last_dc = len(slots)
        else:
            last_type=s.get("type",2); last_dc=s.get("drop_cnt",len(slots))
        last_slots=s.get("slots",[])

        # Pre-fill slots from CSV meta if present
        csv_slot_rates  = {}  # slot_idx -> rate from CSV
        csv_slot_counts = {}  # slot_idx -> count from CSV
        if meta:
            for k, v in meta.items():
                import re as _re2
                m_dr = _re2.match(r"droprate_?(\d+)$", k)
                m_ic = _re2.match(r"itemcnt_?(\d+)$", k)
                m_r  = _re2.match(r"rate$", k)
                if m_dr:
                    try: csv_slot_rates[int(m_dr.group(1))] = int(v)
                    except: pass
                elif m_ic:
                    try: csv_slot_counts[int(m_ic.group(1))] = int(v)
                    except: pass
                elif m_r:
                    try:
                        r_val = int(v)
                        for sidx, _ in slots: csv_slot_rates[sidx] = r_val
                    except: pass
            # Bare "droprate" or "itemcnt" without index → all slots
            if "droprate" in meta:
                try:
                    dr = int(meta["droprate"])
                    for sidx, _ in slots: csv_slot_rates.setdefault(sidx, dr)
                except: pass
            if "itemcnt" in meta:
                try:
                    ic = int(meta["itemcnt"])
                    for sidx, _ in slots: csv_slot_counts.setdefault(sidx, ic)
                except: pass
        level_str = meta.get("level", "")

        wrap=tk.Frame(self,bg=BG); wrap.pack(fill="both",expand=True)
        wrap.grid_rowconfigure(0,weight=0); wrap.grid_rowconfigure(1,weight=1); wrap.grid_rowconfigure(2,weight=0)
        wrap.grid_columnconfigure(0,weight=1)
        hdr=tk.Frame(wrap,bg=BG2); hdr.grid(row=0,column=0,sticky="ew")
        hdr_txt=f"  Box {idx+1} / {total}   ID: {rid}"
        if csv_name: hdr_txt+=f"   —   {csv_name}"
        if level_str: hdr_txt+=f"   (Lv {level_str})"
        if group_name: hdr_txt+=f"   [{group_name}]"
        tk.Label(hdr,text=hdr_txt,font=("Consolas",12,"bold"),bg=BG2,fg=ACC2,pady=8).pack(side="left",padx=10)
        if self.manual_continue_mode:
            tk.Label(hdr,text="🤖 AUTO" if self.manual_continue_mode=="automate" else "👁 MONITOR",
                     font=("Consolas",10),bg=BG2,fg=ACC4).pack(side="right",padx=15)
        nav2=tk.Frame(wrap,bg=BG2); nav2.grid(row=2,column=0,sticky="ew")
        scroll_host2=tk.Frame(wrap,bg=BG); scroll_host2.grid(row=1,column=0,sticky="nsew")
        canvas,cont=mk_scroll_canvas(scroll_host2)

        sec_type=mk_section(cont,"  Drop Type  ")
        type_var=tk.IntVar(value=last_type)
        tf=tk.Frame(sec_type,bg=BG); tf.pack(anchor="w",padx=8,pady=6)
        for lbl,val in [("Random (Type=0)",0),("Egalitarian (Type=2)",2)]:
            tk.Radiobutton(tf,text=lbl,variable=type_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(side="left",padx=8)
        dc_frame=tk.Frame(sec_type,bg=BG); dc_frame.pack(anchor="w",padx=8,pady=(0,6))
        dc_var=tk.StringVar(value=str(last_dc))
        dc_lbl=tk.Label(dc_frame,text="Types of Items (DropCnt):",bg=BG,fg=FG,font=("Consolas",9))
        dc_ent=tk.Entry(dc_frame,textvariable=dc_var,width=6,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat")
        dc_auto=tk.Label(dc_frame,text=f"DropCnt auto-set to {len(slots)}  (all items in this box)",bg=BG,fg=FG_GREY,font=("Consolas",9))
        def toggle_dc(*_):
            if type_var.get()==0: dc_auto.pack_forget(); dc_lbl.pack(side="left"); dc_ent.pack(side="left",padx=6)
            else: dc_lbl.pack_forget(); dc_ent.pack_forget(); dc_auto.pack(anchor="w")
        type_var.trace_add("write",toggle_dc); toggle_dc()

        sec_slots=mk_section(cont,f"  Drop Slots  ({len(slots)} items)  ")
        hrow=tk.Frame(sec_slots,bg=BG2); hrow.pack(fill="x",padx=8,pady=2)
        for txt,w in [("Slot #",6),("Drop ID",12),("Item Name",34),("Rate %",9),("Item Count",10)]:
            tk.Label(hrow,text=txt,width=w,bg=BG2,fg=BLUE,font=("Consolas",9,"bold"),anchor="w").pack(side="left",padx=3)
        slot_rate_vars=[]; slot_count_vars=[]
        for pos,(sidx,drop_id) in enumerate(slots):
            bg=BG if pos%2==0 else BG2
            srow=tk.Frame(sec_slots,bg=bg); srow.pack(fill="x",padx=8,pady=1)
            # Priority: last_slots (user edited) > CSV meta for this slot > fallback 100/1
            if pos < len(last_slots):
                prev_r = last_slots[pos]["rate"]
                prev_c = last_slots[pos]["count"]
            else:
                prev_r = csv_slot_rates.get(sidx, 100)
                prev_c = csv_slot_counts.get(sidx, 1)
            tk.Label(srow,text=str(sidx),width=6,bg=bg,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=3)
            tk.Label(srow,text=drop_id,width=12,bg=bg,fg=BG4,font=("Consolas",9)).pack(side="left",padx=3)
            name=self.item_lib.get(drop_id,"—")
            tk.Label(srow,text=name[:36],width=34,bg=bg,fg=FG_DIM,font=("Consolas",9),anchor="w").pack(side="left",padx=3)
            rv=tk.StringVar(value=str(prev_r))
            tk.Entry(srow,textvariable=rv,width=7,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=3)
            slot_rate_vars.append(rv)
            cv=tk.StringVar(value=str(prev_c))
            tk.Entry(srow,textvariable=cv,width=7,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=3)
            slot_count_vars.append(cv)

        def gather():
            t=type_var.get()
            dc=len(slots) if t==2 else max(1,int(dc_var.get() or 1))
            sl=[]
            for rv,cv in zip(slot_rate_vars,slot_count_vars):
                try: r=max(1,min(32766,int(rv.get())))
                except: r=100
                try: c=max(1,min(32766,int(cv.get())))
                except: c=1
                sl.append({"rate":r,"count":c})
            return {"type":t,"drop_cnt":dc,"slots":sl}

        def save_and_advance():
            if type_var.get()==0:
                dc_raw=dc_var.get().strip()
                if not dc_raw or not dc_raw.isdigit() or int(dc_raw)<1:
                    if not messagebox.askyesno("Missed a spot","DropCnt is blank/invalid. Will default to 1. Continue?"): return
            cfg=gather(); self.manual_configs[rid]=cfg; self.manual_saved=cfg
            self.manual_idx+=1
            if self.manual_idx>=total: self._finish_manual()
            elif self.manual_continue_mode=="automate": self._automate_manual_remaining(cfg)
            elif self.manual_continue_mode=="monitor": self._build_manual_screen()
            else: self._ask_manual_mode(cfg)

        # Use nav2 which is pinned outside the scroll canvas
        mk_btn(nav2,"◀  Start Over",self._build_load_screen).pack(side="left",padx=8,pady=6)
        if idx>0:
            mk_btn(nav2,"◀  Prev",lambda:(setattr(self,"manual_idx",self.manual_idx-1),self._build_manual_screen())).pack(side="left",padx=4,pady=6)
        if self.manual_continue_mode:
            mk_btn(nav2,"⚙ Change Mode",lambda:(setattr(self,"manual_continue_mode",None),self._build_manual_screen()),color=BG4).pack(side="left",padx=4,pady=6)
        mk_btn(nav2,"Finish ✓" if idx==total-1 else "Next ▶",save_and_advance,color=GREEN,fg=BG2,font=("Consolas",10,"bold")).pack(side="right",padx=8,pady=6)

    def _ask_manual_mode(self, last_cfg):
        remaining=len(self.manual_matched)-self.manual_idx
        win=tk.Toplevel(self.root); win.title("Continue?"); win.geometry("520x240")
        win.configure(bg=BG); win.grab_set()
        tk.Label(win,text=f"{remaining} box(es) remaining.",bg=BG,fg=FG,font=("Consolas",13,"bold")).pack(pady=12)
        remember=tk.BooleanVar(value=False)
        bf=tk.Frame(win,bg=BG); bf.pack(pady=8)
        def choose(mode):
            if remember.get(): self.manual_continue_mode=mode
            win.destroy()
            if mode=="automate": self._automate_manual_remaining(last_cfg)
            else: self._build_manual_screen()
        mk_btn(bf,"🤖  Automate  —  copy settings to all remaining boxes",lambda:choose("automate"),color=ACC1,fg=BG2).pack(pady=4)
        mk_btn(bf,"👁  Monitor  —  review each box",lambda:choose("monitor"),color=BLUE,fg=BG2).pack(pady=4)
        tk.Checkbutton(win,text="Remember for rest of session",variable=remember,
                       bg=BG,fg=ACC4,selectcolor=BG3,activebackground=BG,font=("Consolas",9)).pack(pady=4)

    def _automate_manual_remaining(self, last_cfg):
        while self.manual_idx<len(self.manual_matched):
            _item=self.manual_matched[self.manual_idx]
            rid=_item[0]; row_block=_item[2]
            meta=_item[3] if len(_item)>3 else {}
            slots=real_drop_slots(row_block); cfg=copy.deepcopy(last_cfg)
            # Apply any per-slot CSV overrides from meta
            for pos,(sidx,_) in enumerate(slots):
                sr = meta.get(f"droprate_{sidx}", meta.get("droprate", meta.get("rate")))
                sc = meta.get(f"itemcnt_{sidx}", meta.get("itemcnt"))
                if pos < len(cfg["slots"]):
                    if sr is not None:
                        try: cfg["slots"][pos]["rate"] = int(sr)
                        except: pass
                    if sc is not None:
                        try: cfg["slots"][pos]["count"] = int(sc)
                        except: pass
            while len(cfg["slots"])<len(slots): cfg["slots"].append(cfg["slots"][-1] if cfg["slots"] else {"rate":100,"count":1})
            cfg["slots"]=cfg["slots"][:len(slots)]
            if self._rewrite_params.get():
                cfg["type"] = self._rewrite_type.get()
                if self._rewrite_dc_mode.get() == "manual":
                    try: cfg["drop_cnt"] = int(self._rewrite_dc_val.get())
                    except: pass
                elif cfg["type"] == 2:
                    cfg["drop_cnt"] = len(slots)
            elif cfg["type"]==2:
                cfg["drop_cnt"]=len(slots)
            self.manual_configs[rid]=cfg; self.manual_idx+=1
        self._finish_manual()

    def _finish_manual(self):
        csv_rows=[]
        # Build name lookup from 5-tuple matched list
        _name_map={item[0]:item[1] for item in self.manual_matched}
        def replace_row(m):
            row=m.group(0); rid=_get_tag(row,"Id")
            if rid not in self.manual_configs: return row
            new_row=apply_cfg_to_row(row,self.manual_configs[rid])
            drop_ids=[v for _,v in real_drop_slots(new_row)]
            name=_name_map.get(rid,"")
            csv_rows.append([rid,name,*drop_ids]); return new_row
        full_out=ROW_RE.sub(replace_row,self.xml_text)
        self._build_output_screen(full_out,csv_rows,len(self.manual_configs))

    def _build_output_screen(self, full_xml, csv_rows, count):
        self._clear()
        tk.Label(self,text=f"Done — {count} box(es) modified",font=("Consolas",13,"bold"),
                 bg=BG,fg=GREEN).pack(pady=12)
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=12,pady=4)
        max_items=max((len(r)-2 for r in csv_rows),default=0)
        header=["BoxID","BoxName"]+[f"Item{i+1}_ID" for i in range(max_items)]
        csv_cont="\n".join([",".join(header)]+[",".join(str(x) for x in r) for r in csv_rows])
        make_output_tab(nb,"Full PresentItemParam2.xml (modified)",full_xml,"PresentItemParam2_modified.xml",self.root)
        make_output_tab(nb,"Box Contents CSV (→ Tool 3)",csv_cont,"box_contents_for_tool3.csv",self.root)
        nb.select(0)
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        def export_all():
            folder=filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            for fname,content in [("PresentItemParam2_modified.xml",full_xml),("box_contents_for_tool3.csv",csv_cont)]:
                with open(os.path.join(folder,fname),"w",encoding="utf-8") as f: f.write(content)
            messagebox.showinfo("Export Complete",f"Saved to:\n{folder}")
        mk_btn(bot,"💾  Export All Files",export_all,color=ACC2,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=14)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 3 — NCash / Ticket Updater  (simple CSV)
# ══════════════════════════════════════════════════════════════════════════════
class Tool3(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session; self.csv_items=[]; self.xml_files=[]; self.item_lib={}
        self.mode_var=tk.StringVar(value="uniform"); self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _build_load_screen(self):
        self._clear()
        tk.Label(self,text="NCASH / TICKET UPDATER",font=("Consolas",18,"bold"),bg=BG,fg=ACC3).pack(pady=(24,4))
        tk.Label(self,text="Formula: NCash = round(tickets × 133)",bg=BG,fg=FG_DIM,font=("Consolas",10)).pack(pady=(0,4))
        _col_info3 = tk.Frame(self, bg=BG2); _col_info3.pack(pady=(0,6), padx=20, fill="x")
        tk.Label(_col_info3,
            text=("  CSV / Excel column guide:\n"
                  "    ID or ItemID — item ID to look up in the XML (required)\n"
                  "    Tickets      — ticket price (used to compute NCash = tickets×133)\n"
                  "    NCash        — set NCash directly (overrides Tickets)\n"
                  "    Name         — display name (optional, cosmetic only)\n"
                  "  Use Tool 1 or 2 session import to skip the CSV entirely."),
            bg=BG2, fg=FG, font=("Consolas",8), justify="left", padx=8, pady=6
        ).pack(anchor="w")
        csv_status=tk.StringVar(value="No file loaded")
        xml_status=tk.StringVar(value="No file loaded")

        csv_frm=mk_section(self,"Step 1 — Box Contents CSV (from Tool 2, or ID list)")
        tk.Label(csv_frm,textvariable=csv_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_csv():
            p=filedialog.askopenfilename(filetypes=[("Spreadsheet","*.csv *.xlsx *.xlsm *.xls"),("CSV","*.csv"),("Excel","*.xlsx *.xlsm *.xls"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig") as f: text=f.read()
            items=parse_csv_text_t3(text)
            if not items: messagebox.showerror("Error","No item IDs found in CSV."); return
            self.csv_items=items
            if self.item_lib:
                for it in self.csv_items: it["name"]=self.item_lib.get(it["id"],"")
            csv_status.set(f"✓  {os.path.basename(p)}  —  {len(items)} items")
        mk_btn(csv_frm,"📂 Load",load_csv,padx=10,pady=4).pack(side="right",padx=8,pady=6)
        def import_session_t3():
            csv = self.session.box_id_list_csv or self.session.box_contents_csv
            if not csv:
                messagebox.showinfo("No Session Data","Run Tool 1 or 2 first to generate data.")
                return
            items = parse_csv_text_t3(csv)
            if not items:
                messagebox.showerror("Error","No item IDs found in session data."); return
            self.csv_items = items
            if self.item_lib:
                for it in self.csv_items: it["name"] = self.item_lib.get(it["id"],"")
            src_label = "Tool 1" if self.session.box_id_list_csv else "Tool 2"
            csv_status.set(f"✓  Imported from {src_label}  —  {len(items)} items")
        mk_btn(csv_frm,"⬇  Import from Tool 1/2",import_session_t3,color=ACC3,fg=BG2,padx=8,pady=4).pack(side="right",padx=4,pady=6)

        xml_frm=mk_section(self,"Step 2 — ItemParam XML (pick any one of the 4 files)")
        tk.Label(xml_frm,textvariable=xml_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_xml():
            p=filedialog.askopenfilename(title="Select any one of the 4 ItemParam XML files",
                                         filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            folder=os.path.dirname(p); loaded=[]
            for fname in os.listdir(folder):
                if fname.lower() in TARGET_FILES:
                    try:
                        with open(os.path.join(folder,fname),encoding="utf-8-sig",errors="replace") as f:
                            loaded.append((fname,f.read()))
                    except: pass
            if not loaded: messagebox.showerror("Error","None of the 4 ItemParam files found."); return
            self.xml_files=loaded; self.item_lib=build_item_lib(loaded)
            for it in self.csv_items: it["name"]=self.item_lib.get(it["id"],"")
            xml_status.set(f"✓  {len(loaded)}/4 files  |  {len(self.item_lib)} items indexed")
        mk_btn(xml_frm,"📂 Load",load_xml,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        mode_frm=mk_section(self,"Step 3 — Mode")
        mf=tk.Frame(mode_frm,bg=BG); mf.pack(anchor="w",padx=10,pady=6)
        for lbl,val in [("Uniform  —  one ticket cost applied to every item","uniform"),
                        ("Manual   —  set ticket cost per item individually","manual")]:
            tk.Radiobutton(mf,text=lbl,variable=self.mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(anchor="w",pady=2)

        def proceed():
            if not self.csv_items: messagebox.showwarning("Missing","Load a CSV first."); return
            if not self.xml_files: messagebox.showwarning("Missing","Load ItemParam XML first."); return
            if self.mode_var.get()=="uniform": self._build_uniform_screen()
            else: self._build_manual_screen()
        mk_btn(self,"▶  Continue →",proceed,color=GREEN,fg=BG2,font=("Consolas",12,"bold")).pack(pady=18)

    def _build_uniform_screen(self):
        self._clear()
        tk.Label(self,text="Uniform Ticket Cost",font=("Consolas",14,"bold"),bg=BG,fg=ACC3).pack(pady=(20,4))
        tk.Label(self,text=f"Applied to all {len(self.csv_items)} items.",bg=BG,fg=FG_DIM,font=("Consolas",10)).pack(pady=(0,12))
        frm=tk.Frame(self,bg=BG); frm.pack()
        tk.Label(frm,text="Ticket Cost:",bg=BG,fg=FG,font=("Consolas",12)).pack(side="left",padx=8)
        tv=tk.StringVar()
        ent=tk.Entry(frm,textvariable=tv,width=12,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",12),relief="flat")
        ent.pack(side="left",padx=8); ent.focus()
        ncash_var=tk.StringVar(value="NCash: —")
        tk.Label(self,textvariable=ncash_var,bg=BG,fg=GREEN,font=("Consolas",12,"bold")).pack(pady=8)
        def on_change(*_):
            try: ncash_var.set(f"NCash: {round(float(tv.get())*133)}")
            except: ncash_var.set("NCash: —")
        tv.trace_add("write",on_change)
        def apply_uniform():
            try: cost=float(tv.get())
            except: messagebox.showwarning("Invalid","Enter a valid ticket cost."); return
            for it in self.csv_items: it["ticket_cost"]=cost
            self._process_and_show()
        bot=tk.Frame(self,bg=BG); bot.pack(pady=16)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=8)
        mk_btn(bot,"✓  Apply to All & Update XML",apply_uniform,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=8)

    def _build_manual_screen(self):
        self._clear()
        tk.Label(self,text="Manual Ticket Costs",font=("Consolas",14,"bold"),bg=BG,fg=ACC3).pack(pady=(12,2))
        tk.Label(self,text="Leave blank to skip an item.",bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(pady=(0,4))
        outer=tk.Frame(self,bg=BG); outer.pack(fill="both",expand=True,padx=20,pady=4)
        canvas,cont=mk_scroll_canvas(outer)
        hdr=tk.Frame(cont,bg=BG2); hdr.pack(fill="x",pady=2)
        for txt,w in [("Item ID",12),("Item Name",36),("Ticket Cost",14),("NCash (calc)",14)]:
            tk.Label(hdr,text=txt,width=w,bg=BG2,fg=BLUE,font=("Consolas",9,"bold"),anchor="w").pack(side="left",padx=6,pady=4)
        ticket_vars=[]
        for i,item in enumerate(self.csv_items):
            bg=BG if i%2==0 else BG2
            row=tk.Frame(cont,bg=bg); row.pack(fill="x")
            tk.Label(row,text=item["id"],width=12,bg=bg,fg=BG4,font=("Consolas",9),anchor="w").pack(side="left",padx=6,pady=2)
            name=item.get("name") or self.item_lib.get(item["id"],"—")
            tk.Label(row,text=name[:38],width=36,bg=bg,fg=FG_DIM,font=("Consolas",9),anchor="w").pack(side="left",padx=6,pady=2)
            tv=tk.StringVar()
            if item.get("ticket_cost") is not None: tv.set(str(item["ticket_cost"]))
            ticket_vars.append(tv)
            tk.Entry(row,textvariable=tv,width=12,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=6,pady=2)
            ncash_lbl=tk.Label(row,text="—",width=14,bg=bg,fg=GREEN,font=("Consolas",9),anchor="w")
            ncash_lbl.pack(side="left",padx=6)
            def make_trace(var,lbl):
                def cb(*_):
                    try: lbl.config(text=str(round(float(var.get())*133)))
                    except: lbl.config(text="—")
                var.trace_add("write",cb); cb()
            make_trace(tv,ncash_lbl)
        def confirm():
            blanks=[]
            for i,item in enumerate(self.csv_items):
                raw=ticket_vars[i].get().strip()
                try: item["ticket_cost"]=float(raw)
                except: item["ticket_cost"]=None; blanks.append(item["id"])
            if blanks:
                if not messagebox.askyesno("Missed a spot",f"{len(blanks)} item(s) will be SKIPPED:\n\n"+", ".join(blanks[:20])+("\n\nContinue anyway?")): return
            self._process_and_show()
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=14)
        mk_btn(bot,"✓  Apply & Update XML",confirm,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="right",padx=14)

    def _process_and_show(self):
        updates={it["id"]:round(it["ticket_cost"]*133) for it in self.csv_items if it["ticket_cost"] is not None}
        name_map={it["id"]:it.get("name","") for it in self.csv_items}
        file_results=[]
        for fname,text in self.xml_files:
            modified,found_map=bulk_update_ncash(text,updates)
            file_results.append((fname,modified,found_map))
        found_in={}
        for fname,_,found_map in file_results:
            for iid,hit in found_map.items():
                if hit and iid not in found_in: found_in[iid]=fname
        results=[]
        for item in self.csv_items:
            iid=item["id"]; name=name_map.get(iid,"")
            if item["ticket_cost"] is None: results.append((iid,name,None,None))
            else: results.append((iid,name,updates[iid],found_in.get(iid)))
        self._build_output_screen(file_results,results,updates)

    def _build_output_screen(self, file_results, results, updates):
        self._clear()
        updated_count=sum(1 for _,_,n,f in results if n is not None and f is not None)
        skipped_count=sum(1 for _,_,n,_ in results if n is None)
        missing_count=sum(1 for _,_,n,f in results if n is not None and f is None)
        tk.Label(self,text=f"✓ Updated: {updated_count}    ⚠ Not found: {missing_count}    — Skipped: {skipped_count}",
                 font=("Consolas",10,"bold"),bg=BG,fg=GREEN).pack(pady=8)
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=12,pady=4)
        exports=[]
        for fname,modified_text,found_map in file_results:
            if not any(hit for hit in found_map.values()): continue
            exports.append((fname,modified_text))
            make_output_tab(nb,os.path.splitext(fname)[0],modified_text,fname,self.root)
        col_hdr=f"{'ID':<15}{'Name':<34}{'NCash':<13}Status"; col_sep="─"*74
        log_parts=[]
        for fname,_,found_map in file_results:
            file_rows=[(iid,name,ncash,ff) for iid,name,ncash,ff in results if ff==fname]
            if not file_rows: log_parts.append(f"{fname}  →  No matching IDs — Skipped!\n"); continue
            log_parts.append(f"{fname}  →  {len(file_rows)} match(es)\n  {col_hdr}\n  {col_sep}")
            for iid,name,ncash,_ in file_rows: log_parts.append(f"  {iid:<15}{(name or '—')[:32]:<34}{ncash:<13}✓ Updated")
            log_parts.append("")
        unassigned=[(iid,name,ncash,ff) for iid,name,ncash,ff in results if ff is None]
        missing_rows=[(iid,name,ncash) for iid,name,ncash,_ in unassigned if ncash is not None]
        skipped_rows=[(iid,name) for iid,name,ncash,_ in unassigned if ncash is None]
        log_parts.append("── Unassigned / Skipped ──────────────────────────────────")
        for iid,name,ncash in missing_rows: log_parts.append(f"  {iid:<15}{(name or '—')[:32]:<34}{ncash:<13}⚠ Not found")
        for iid,name in skipped_rows: log_parts.append(f"  {iid:<15}{(name or '—')[:32]:<34}{'—':<13}SKIPPED")
        if not missing_rows and not skipped_rows: log_parts.append("  (none)")
        log_content="\n".join(log_parts)
        exports.append(("ncash_update_log.txt",log_content))
        make_output_tab(nb,"Update Log",log_content,"ncash_update_log.txt",self.root)
        nb.select(0)
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        def export_all():
            folder=filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            saved=[]
            for efname,content in exports:
                with open(os.path.join(folder,efname),"w",encoding="utf-8") as f: f.write(content)
                saved.append(efname)
            messagebox.showinfo("Export Complete",f"Saved to:\n{folder}\n\n"+"\n".join(saved))
        mk_btn(bot,"💾  Export All Files",export_all,color=ACC1,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=14)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 4 — NCash Updater  (parent-box CSV + sub-box via PresentItemParam2)
# ══════════════════════════════════════════════════════════════════════════════
class Tool4(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session
        self.parent_items=[]; self.xml_files=[]; self.item_lib={}
        self.present_text=None; self.present_enabled=tk.BooleanVar(value=False)
        self._traversed_box_ids=set()
        self.parent_mode_var=tk.StringVar(value="uniform")
        self.sub_mode_var=tk.StringVar(value="uniform")
        self.sub_items=[]; self._build_load_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _build_load_screen(self):
        self._clear()
        tk.Label(self,text="NCASH UPDATER — PARENT BOX",font=("Consolas",16,"bold"),bg=BG,fg=ACC4).pack(pady=(18,2))
        tk.Label(self,text="Formula: NCash = round(tickets × 133)",bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(pady=(0,4))
        _col_info4 = tk.Frame(self, bg=BG2); _col_info4.pack(pady=(0,4), padx=12, fill="x")
        tk.Label(_col_info4,
            text=("  Parent-Box CSV column guide:\n"
                  "    ID           — parent box item ID (required)\n"
                  "    Tickets      — ticket cost of the parent box\n"
                  "    NCash        — set NCash directly for parent box\n"
                  "    BoxTickets   — ticket cost of sub-boxes inside this parent\n"
                  "    Name         — display name (optional)\n"
                  "  Excel (.xlsx/.xlsm/.xls) is also supported."),
            bg=BG2, fg=FG, font=("Consolas",8), justify="left", padx=8, pady=6
        ).pack(anchor="w")
        outer=tk.Frame(self,bg=BG); outer.pack(fill="both",expand=True)
        scroll_host4=tk.Frame(outer,bg=BG); scroll_host4.pack(fill="both",expand=True)
        canvas,cont=mk_scroll_canvas(scroll_host4)
        # Session import panel
        if self.session.box_id_list_csv:
            sess_frm=mk_section(cont,"  ⬇  Session Import  ")
            def import_t4():
                items=parse_parentbox_csv(self.session.box_id_list_csv)
                if not items:
                    items=[{"id":i,"tickets":None,"ncash_direct":None,"box_ticket_cost":None,"group_idx":0,"name":n}
                           for i,n in self.session.box_id_map.items()]
                if not items: messagebox.showinfo("Empty","Session has no usable box IDs."); return
                self.parent_items=items
                messagebox.showinfo("Imported",f"Imported {len(items)} box IDs from Tool 1 session.")
            mk_btn(sess_frm,"Import box IDs from Tool 1",import_t4,color=ACC4,fg=BG2).pack(side="left",padx=10,pady=6)
            tk.Label(sess_frm,text=f"{len(parse_box_id_csv(self.session.box_id_list_csv))} IDs available",
                     bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=8)

        csv_status=tk.StringVar(value="No file loaded")
        xml_status=tk.StringVar(value="No file loaded")
        pres_status=tk.StringVar(value="Not loaded")

        s1=mk_section(cont,"  Step 1 — Parent-Box CSV  (ID, Tickets/NCash, Tickets of Box Contents)  ")
        tk.Label(s1,textvariable=csv_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_csv():
            p=filedialog.askopenfilename(filetypes=[("Spreadsheet","*.csv *.xlsx *.xlsm *.xls"),("CSV","*.csv"),("Excel","*.xlsx *.xlsm *.xls"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig") as f: text=f.read()
            items=parse_parentbox_csv(text)
            if not items: messagebox.showerror("Error","No valid item IDs found in CSV."); return
            self.parent_items=items
            csv_status.set(f"✓  {os.path.basename(p)}  —  {len(items)} IDs")
        mk_btn(s1,"📂 Load CSV",load_csv,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        s2=mk_section(cont,"  Step 2 — ItemParam XML  (pick any one of the 4 files)  ")
        tk.Label(s2,textvariable=xml_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(side="left",padx=10)
        def load_xml():
            p=filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            folder=os.path.dirname(p); loaded=[]
            for fname in os.listdir(folder):
                if fname.lower() in TARGET_FILES:
                    try:
                        with open(os.path.join(folder,fname),encoding="utf-8-sig",errors="replace") as f:
                            loaded.append((fname,f.read()))
                    except: pass
            if not loaded: messagebox.showerror("Error","None of the 4 ItemParam files found."); return
            self.xml_files=loaded; self.item_lib=build_item_lib(loaded)
            for it in self.parent_items: it["name"]=self.item_lib.get(it["id"],"")
            # Also silently scan for PresentItemParam2.xml
            ppath=os.path.join(folder,PRESENT_FILE)
            if os.path.exists(ppath):
                try:
                    with open(ppath,encoding="utf-8-sig",errors="replace") as f:
                        self.present_text=f.read()
                    pres_status.set(f"✓  {PRESENT_FILE} auto-loaded")
                except: pass
            xml_status.set(f"✓  {len(loaded)}/4 files  |  {len(self.item_lib)} items indexed")
        mk_btn(s2,"📂 Load",load_xml,padx=10,pady=4).pack(side="right",padx=8,pady=6)

        s3=mk_section(cont,"  Step 3 — Mode  ")
        tk.Label(s3,
                 text="  Uniform: one ticket/NCash value applied to every item in the group.\n"
                      "  Manual: set the value per item individually.",
                 bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(anchor="w",padx=10,pady=(4,0))
        mf=tk.Frame(s3,bg=BG); mf.pack(anchor="w",padx=10,pady=6)

        # Step 5 (sub-box mode) — only shown when Manual is selected in Step 3
        s5=mk_section(cont,"  Step 4 (Optional) — Update NCash on the boxes INSIDE your parent box  ")
        tk.Label(s5,
                 text="  Your parent box contains sub-boxes (e.g. Dragon Low Gear Box contains [JP] Hanyu Box, etc.).\n"
                      "  Those sub-boxes each have their own NCash value in ItemParam.\n"
                      "  Load PresentItemParam2.xml here to ALSO update those inner box NCash values.\n"
                      "  Leave unchecked if you only want to update the parent box NCash.",
                 bg=BG,fg=FG_GREY,font=("Consolas",8),justify="left").pack(anchor="w",padx=10,pady=(4,2))
        tk.Label(s5,textvariable=pres_status,bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(anchor="w",padx=10,pady=(0,0))
        tk.Checkbutton(s5,text="Yes — also update the sub-boxes (items INSIDE the parent box) via PresentItemParam2",
                       variable=self.present_enabled,bg=BG,fg=FG,selectcolor=BG3,
                       activebackground=BG,font=("Consolas",10)).pack(anchor="w",padx=10,pady=4)
        def load_present_manual():
            p=filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")])
            if not p: return
            with open(p,encoding="utf-8-sig",errors="replace") as f: self.present_text=f.read()
            pres_status.set(f"✓  {os.path.basename(p)}")
        mk_btn(s5,"📂 Load PresentItemParam2.xml",load_present_manual,padx=10,pady=4).pack(anchor="w",padx=10,pady=(0,6))

        s6=mk_section(cont,"  Step 5 — How to set NCash on the sub-boxes inside your parent box  (only if Step 4 enabled)  ")
        tk.Label(s6,
                 text="  Each sub-box sitting INSIDE your parent box (e.g. [JP] Hanyu Box, 2012 Goodie Bag, etc.)\n"
                      "  can be updated to its own NCash/ticket value.\n"
                      "  Uniform: apply one ticket price to ALL of those inner boxes at once.\n"
                      "  Manual:  set a different ticket price for each inner box individually.",
                 bg=BG,fg=FG_GREY,font=("Consolas",8),justify="left").pack(anchor="w",padx=10,pady=(4,2))
        sf=tk.Frame(s6,bg=BG); sf.pack(anchor="w",padx=10,pady=6)
        for lbl,val in [("Uniform  —  same ticket/NCash value for all sub-boxes inside the parent","uniform"),
                        ("Manual   —  set a different value per sub-box","manual")]:
            tk.Radiobutton(sf,text=lbl,variable=self.sub_mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(anchor="w",pady=2)

        def _toggle_s6(*_):
            if self.parent_mode_var.get()=="manual": s6.pack(fill="x",padx=4,pady=2)
            else: s6.pack_forget()
        # Initial state and wire toggle
        _toggle_s6()

        for lbl,val in [("Uniform  —  one value per group","uniform"),("Manual  —  set per item","manual")]:
            tk.Radiobutton(mf,text=lbl,variable=self.parent_mode_var,value=val,
                           bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10),
                           command=_toggle_s6).pack(anchor="w",pady=2)

        bot_frm=tk.Frame(cont,bg=BG); bot_frm.pack(fill="x",pady=10)
        def proceed():
            if not self.parent_items: messagebox.showwarning("Missing","Load a CSV first."); return
            if not self.xml_files: messagebox.showwarning("Missing","Load ItemParam XML first."); return
            if self.parent_mode_var.get()=="uniform":
                self.sub_mode_var.set("uniform")  # force uniform sub-mode when parent is uniform
                self._build_uniform_screen()
            else: self._build_manual_screen(self.parent_items,"parent",self._after_parent_configured)
        mk_btn(bot_frm,"▶  Continue →",proceed,color=GREEN,fg=BG2,font=("Consolas",12,"bold")).pack(pady=10)

    # ── Uniform screen (group-aware) ──────────────────────────────────────────
    def _build_uniform_screen(self, _saved_group_vals=None):
        self._clear()
        groups={}
        for it in self.parent_items:
            gi=it.get("group_idx",0)
            groups.setdefault(gi,[]).append(it)
        group_keys=sorted(groups.keys())
        saved_vals=_saved_group_vals or {}
        confirmed_vals={}  # gi -> {ticket_cost or ncash_direct}
        current_group_pos=[0]
        last_confirmed_val={"type":"tickets","val":""}  # persists across groups

        def show_group(pos):
            self._clear()
            gi=group_keys[pos]
            group_items=groups[gi]
            # Get the box name from the group header (first item's group name or CSV box name)
            box_name = next((it.get("name","") for it in group_items if it.get("name","")), f"Group {pos+1}")
            sample_names=[self.item_lib.get(it["id"],"") for it in group_items if self.item_lib.get(it["id"],"")][:4]

            tk.Label(self,text=f"Parent Box {pos+1} of {len(group_keys)}",
                     font=("Consolas",14,"bold"),bg=BG,fg=ACC4).pack(pady=(18,2))
            tk.Label(self,text=f"Set the ticket or NCash price for this parent box.",
                     bg=BG,fg=FG_DIM,font=("Consolas",10)).pack()
            tk.Label(self,text=f"This parent box contains {len(group_items)} sub-box(es) inside it.",
                     bg=BG,fg=FG_GREY,font=("Consolas",9)).pack(pady=(2,0))
            if sample_names:
                tk.Label(self,text="Sub-boxes inside: "+", ".join(sample_names),
                         bg=BG,fg=FG_GREY,font=("Consolas",8)).pack()

            if confirmed_vals:
                prev_frm=tk.LabelFrame(self,text="  Already confirmed  ",bg=BG,fg=FG_GREY,
                                       font=("Consolas",9),bd=1,relief="groove")
                prev_frm.pack(fill="x",padx=24,pady=6)
                for prev_gi,pval in confirmed_vals.items():
                    pg_items=groups[prev_gi]
                    pg_sample=[self.item_lib.get(it["id"],"") for it in pg_items if self.item_lib.get(it["id"],"")][:2]
                    desc=", ".join(pg_sample) or f"Group {prev_gi+1}"
                    if pval.get("ticket_cost") is not None:
                        tc = int(round(pval["ticket_cost"]))
                        disp=f"{tc} tickets → NCash {tc*133}"
                    else: disp=f"NCash {pval.get('ncash_direct','?')} (direct)"
                    tk.Label(prev_frm,text=f"  Box {list(group_keys).index(prev_gi)+1}: {desc[:40]}  →  {disp}",
                             bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(anchor="w",padx=6,pady=1)

            type_var=tk.StringVar(value="tickets")
            sample_it=group_items[0]
            if sample_it.get("ticket_cost") is not None:
                type_var.set("tickets")
                init_val=str(int(round(sample_it["ticket_cost"])))
            elif sample_it.get("ncash_direct") is not None:
                type_var.set("ncash")
                init_val=str(int(round(sample_it["ncash_direct"])))
            elif last_confirmed_val["val"]:
                # Carry forward the last confirmed value and type
                type_var.set(last_confirmed_val["type"])
                init_val=last_confirmed_val["val"]
            else:
                init_val=saved_vals.get(gi,{}).get("init_val","")
            val_var=tk.StringVar(value=init_val)

            inp_frm=tk.Frame(self,bg=BG); inp_frm.pack(pady=10)
            tk.Label(inp_frm,text="How is the price set?",bg=BG,fg=FG,font=("Consolas",10)).pack(side="left",padx=8)
            tk.Radiobutton(inp_frm,text="Ticket cost  (×133 = NCash)",variable=type_var,value="tickets",bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(side="left",padx=8)
            tk.Radiobutton(inp_frm,text="NCash directly",variable=type_var,value="ncash",bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",10)).pack(side="left",padx=8)
            ent_frm=tk.Frame(self,bg=BG); ent_frm.pack(pady=6)
            def _type_label(*_):
                lbl_txt = "Ticket cost for this parent box:" if type_var.get()=="tickets" else "NCash value for this parent box:"
                type_lbl.config(text=lbl_txt)
            type_lbl=tk.Label(ent_frm,text="Ticket cost for this parent box:",bg=BG,fg=FG,font=("Consolas",11))
            type_lbl.pack(side="left",padx=8)
            type_var.trace_add("write",_type_label)
            ent=tk.Entry(ent_frm,textvariable=val_var,width=10,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",12),relief="flat")
            ent.pack(side="left",padx=8); ent.focus()
            result_lbl=tk.Label(self,text="",bg=BG,fg=GREEN,font=("Consolas",11,"bold"))
            result_lbl.pack(pady=4)
            def update_result(*_):
                try:
                    v = int(val_var.get()) if val_var.get().strip().isdigit() else float(val_var.get())
                    if type_var.get()=="tickets": result_lbl.config(text=f"→ NCash will be set to: {int(round(v*133))}")
                    else: result_lbl.config(text=f"→ Approx ticket cost: {int(round(v/133))}")
                except: result_lbl.config(text="")
            val_var.trace_add("write",update_result); type_var.trace_add("write",update_result); update_result()

            def confirm_group():
                raw = val_var.get().strip()
                try: v=int(raw) if raw.isdigit() else float(raw)
                except: messagebox.showwarning("Invalid","Enter a whole number (no decimals)."); return
                # Persist so next group pre-fills with this value
                last_confirmed_val["type"]=type_var.get()
                last_confirmed_val["val"]=raw
                if type_var.get()=="tickets": confirmed_vals[gi]={"ticket_cost":v,"ncash_direct":None}
                else: confirmed_vals[gi]={"ticket_cost":None,"ncash_direct":int(round(v))}
                if pos+1<len(group_keys): show_group(pos+1)
                else:
                    for it in self.parent_items:
                        gval=confirmed_vals.get(it.get("group_idx",0),{})
                        it["ticket_cost"]=gval.get("ticket_cost"); it["ncash_direct"]=gval.get("ncash_direct")
                    self._after_parent_configured()

            def go_back():
                if pos>0:
                    prev_gi=group_keys[pos-1]
                    confirmed_vals.pop(prev_gi,None)
                    show_group(pos-1)
                else: self._build_load_screen()

            nav=tk.Frame(self,bg=BG); nav.pack(pady=14)
            mk_btn(nav,"◀  Back",go_back).pack(side="left",padx=10)
            mk_btn(nav,"Confirm ▶" if pos<len(group_keys)-1 else "✓  Apply All",confirm_group,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=10)

        show_group(0)

    # ── Manual screen (parent or sub) ─────────────────────────────────────────
    def _build_manual_screen(self, items, label, on_done):
        self._clear()
        tk.Label(self,text=f"Manual — {label.title()} Item Costs",font=("Consolas",14,"bold"),bg=BG,fg=ACC4).pack(pady=(12,2))
        tk.Label(self,text="Leave blank to skip.",bg=BG,fg=FG_DIM,font=("Consolas",9)).pack(pady=(0,4))
        outer=tk.Frame(self,bg=BG); outer.pack(fill="both",expand=True,padx=20,pady=4)
        canvas,cont=mk_scroll_canvas(outer)
        hdr=tk.Frame(cont,bg=BG2); hdr.pack(fill="x",pady=2)
        for txt,w in [("Item ID",12),("Item Name",34),("Type",12),("Value",14),("NCash →",14)]:
            tk.Label(hdr,text=txt,width=w,bg=BG2,fg=BLUE,font=("Consolas",9,"bold"),anchor="w").pack(side="left",padx=4,pady=4)
        type_vars=[]; val_vars=[]
        for i,item in enumerate(items):
            bg=BG if i%2==0 else BG2
            row=tk.Frame(cont,bg=bg); row.pack(fill="x")
            tk.Label(row,text=item["id"],width=12,bg=bg,fg=BG4,font=("Consolas",9),anchor="w").pack(side="left",padx=4,pady=2)
            name=item.get("name") or self.item_lib.get(item["id"],"—")
            tk.Label(row,text=name[:36],width=34,bg=bg,fg=FG_DIM,font=("Consolas",9),anchor="w").pack(side="left",padx=4,pady=2)
            if item.get("ticket_cost") is not None: init_type,init_val="tickets",str(item["ticket_cost"])
            elif item.get("ncash_direct") is not None: init_type,init_val="ncash",str(item["ncash_direct"])
            else: init_type,init_val="tickets",""
            tv=tk.StringVar(value=init_type); type_vars.append(tv)
            type_combo=ttk.Combobox(row,textvariable=tv,values=["tickets","ncash"],state="readonly",width=8,font=("Consolas",9))
            type_combo.pack(side="left",padx=4,pady=2)
            vv=tk.StringVar(value=init_val); val_vars.append(vv)
            tk.Entry(row,textvariable=vv,width=12,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=4,pady=2)
            ncash_lbl=tk.Label(row,text="—",width=14,bg=bg,fg=GREEN,font=("Consolas",9),anchor="w")
            ncash_lbl.pack(side="left",padx=4)
            def make_trace(tv2,vv2,lbl):
                def cb(*_):
                    try:
                        v=float(vv2.get())
                        if tv2.get()=="tickets": lbl.config(text=str(round(v*133)))
                        else: lbl.config(text=str(int(round(v))))
                    except: lbl.config(text="—")
                tv2.trace_add("write",cb); vv2.trace_add("write",cb); cb()
            make_trace(tv,vv,ncash_lbl)

        def confirm():
            blanks=[]
            for i,item in enumerate(items):
                raw=val_vars[i].get().strip()
                try:
                    v=float(raw)
                    if type_vars[i].get()=="tickets": item["ticket_cost"]=v; item["ncash_direct"]=None
                    else: item["ncash_direct"]=int(round(v)); item["ticket_cost"]=None
                except: item["ticket_cost"]=None; item["ncash_direct"]=None; blanks.append(item["id"])
            if blanks:
                if not messagebox.askyesno("Missed a spot",f"{len(blanks)} item(s) will be SKIPPED. Continue anyway?"): return
            on_done()

        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=14)
        mk_btn(bot,"✓  Apply",confirm,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="right",padx=14)

    def _after_parent_configured(self):
        if self.present_enabled.get() and self.present_text:
            self._build_sub_configure_screen()
        else:
            self._process_and_show()

    def _build_sub_configure_screen(self):
        box_ids={it["id"] for it in self.parent_items}
        # Recursively walk PresentItemParam2: parent boxes → sub-boxes → sub-sub-boxes → leaf items
        leaf_ids, traversed_box_ids = extract_drop_ids_recursive(self.present_text, box_ids)
        self._traversed_box_ids = traversed_box_ids  # store for output display
        self.sub_items=[]
        seen=set()
        # Use box_ticket_cost from the parent item as the default ticket cost for leaves
        sample_tc = next((it.get("box_ticket_cost") for it in self.parent_items
                          if it.get("box_ticket_cost") is not None), None)
        for drop_id in sorted(leaf_ids, key=lambda x: int(x) if x.isdigit() else x):
            if drop_id not in seen:
                seen.add(drop_id)
                self.sub_items.append({"id":drop_id,"name":self.item_lib.get(drop_id,""),
                                       "ticket_cost":sample_tc,"ncash_direct":None,
                                       "box_ticket_cost":sample_tc,"group_idx":0})
        if not self.sub_items:
            messagebox.showinfo("No sub-items",
                f"Recursive search found no leaf items under the {len(box_ids)} parent box IDs.\n"
                f"Traversed {len(traversed_box_ids)} intermediate box level(s) in PresentItemParam2.")
            self._process_and_show(); return
        all_prefilled=all(it.get("ticket_cost") is not None or it.get("ncash_direct") is not None for it in self.sub_items)
        if all_prefilled:
            self._process_and_show(); return
        if self.sub_mode_var.get()=="uniform":
            self._build_sub_uniform_screen()
        else:
            self._build_manual_screen(self.sub_items,"sub-box",self._process_and_show)

    def _build_sub_uniform_screen(self):
        self._clear()
        tk.Label(self,text="Set Ticket Price for Sub-Boxes (Items Inside the Parent Boxes)",
                 font=("Consolas",13,"bold"),bg=BG,fg=ACC4).pack(pady=(20,4))
        tk.Label(self,
                 text="You already set the price for the PARENT boxes in the previous step.\n"
                      "Now you are setting the price for the boxes that sit INSIDE those parent boxes —\n"
                      f"the {len(self.sub_items)} individual sub-box items (e.g. [JP] Hanyu Box, 2012 Goodie Bag, etc.)\n"
                      "that the player actually receives when they open the parent box.\n"
                      "This ONE ticket value will be applied to ALL of those inner items at once.",
                 bg=BG,fg=FG_DIM,font=("Consolas",9),justify="center").pack(pady=(0,10))
        sample_it=next((it for it in self.sub_items if it.get("ticket_cost") is not None),None)
        init_val=str(int(round(sample_it["ticket_cost"]))) if sample_it else ""
        frm=tk.Frame(self,bg=BG); frm.pack()
        tv=tk.StringVar(value=init_val)
        tk.Label(frm,text="Ticket cost for each sub-box item:",bg=BG,fg=FG,font=("Consolas",11)).pack(side="left",padx=8)
        ent=tk.Entry(frm,textvariable=tv,width=10,bg=BG3,fg=FG,insertbackground=FG,font=("Consolas",12),relief="flat")
        ent.pack(side="left",padx=8); ent.focus()
        ncash_lbl=tk.Label(self,text="",bg=BG,fg=GREEN,font=("Consolas",12,"bold")); ncash_lbl.pack(pady=6)
        def on_change(*_):
            raw = tv.get().strip()
            try:
                v = int(raw) if raw.isdigit() else float(raw)
                ncash_lbl.config(text=f"→ NCash will be set to: {int(round(v*133))}  on all {len(self.sub_items)} sub-box items")
            except: ncash_lbl.config(text="")
        tv.trace_add("write",on_change); on_change()
        # Option to skip updating parent boxes and only update sub-box contents
        skip_parents_var = tk.BooleanVar(value=False)
        skip_frm = tk.Frame(self,bg=BG); skip_frm.pack(pady=(4,0))
        tk.Checkbutton(skip_frm,
                       text="Only update the sub-box items inside the parent boxes\n"
                            "(skip updating the parent boxes themselves — their NCash stays unchanged)",
                       variable=skip_parents_var,bg=BG,fg=ACC4,selectcolor=BG3,
                       activebackground=BG,font=("Consolas",9),justify="left").pack(anchor="w",padx=10)

        def apply():
            raw = tv.get().strip()
            try: cost = int(raw) if raw.isdigit() else float(raw)
            except: messagebox.showwarning("Invalid","Enter a whole number (no decimals)."); return
            for it in self.sub_items: it["ticket_cost"]=cost; it["ncash_direct"]=None
            if skip_parents_var.get():
                # Temporarily zero out parent items so they are skipped in _process_and_show
                _saved_parents = [(it, it.get("ticket_cost"), it.get("ncash_direct"))
                                  for it in self.parent_items]
                for it in self.parent_items:
                    it["ticket_cost"] = None; it["ncash_direct"] = None
                self._process_and_show()
                # Restore parent items
                for it, tc, nd in _saved_parents:
                    it["ticket_cost"] = tc; it["ncash_direct"] = nd
            else:
                self._process_and_show()
        bot=tk.Frame(self,bg=BG); bot.pack(pady=16)
        mk_btn(bot,"◀  Back",self._build_load_screen).pack(side="left",padx=8)
        mk_btn(bot,"✓  Apply to All Sub-Box Items & Update XML",apply,color=GREEN,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=8)

    def _resolve_ncash(self, item):
        if item.get("ticket_cost") is not None: return round(item["ticket_cost"]*133)
        if item.get("ncash_direct") is not None: return item["ncash_direct"]
        return None

    def _process_and_show(self):
        # ONLY leaf items (sub_items) get NCash+recycle updated.
        # Parent boxes (parent_items) are already correctly configured — never touch them.
        # Intermediate boxes (traversed during recursion) are also boxes — never touch them.
        updates={}
        for it in self.sub_items:
            n=self._resolve_ncash(it)
            if n is not None: updates[it["id"]]=n
        file_results=[]
        for fname,text in self.xml_files:
            modified,found_map=bulk_update_ncash(text,updates)
            file_results.append((fname,modified,found_map))
        found_in={}
        for fname,_,found_map in file_results:
            for iid,hit in found_map.items():
                if hit and iid not in found_in: found_in[iid]=fname
        parent_ids={it["id"] for it in self.parent_items}
        sub_ids={it["id"] for it in self.sub_items}
        updated_p=0  # parent boxes are never updated
        updated_s=sum(1 for iid in sub_ids if found_in.get(iid))
        missing=sum(1 for iid in updates if not found_in.get(iid))
        skipped=sum(1 for it in self.sub_items if self._resolve_ncash(it) is None)
        self._build_output_screen(file_results,updates,found_in,parent_ids,sub_ids,updated_p,updated_s,missing,skipped)

    def _build_output_screen(self, file_results, updates, found_in, parent_ids, sub_ids,
                              updated_p, updated_s, missing, skipped):
        self._clear()
        traversed = getattr(self, "_traversed_box_ids", set())
        tk.Label(self,
                 text=f"✓ {updated_s} leaf items updated (NCash + recycle)   "
                      f"⚠ Not found in XML: {missing}   — Skipped (no price set): {skipped}\n"
                      f"Boxes traversed (not updated): {len(parent_ids)} parent + {len(traversed)} intermediate",
                 font=("Consolas",10,"bold"),bg=BG,fg=GREEN,justify="center").pack(pady=8)
        nb=ttk.Notebook(self); nb.pack(fill="both",expand=True,padx=12,pady=4)
        exports=[]
        for fname,modified_text,found_map in file_results:
            if not any(hit for hit in found_map.values()): continue
            exports.append((fname,modified_text))
            make_output_tab(nb,os.path.splitext(fname)[0],modified_text,fname,self.root)
        log_parts=[]
        log_parts.append(f"Parent boxes (NOT updated — already configured):")
        for iid in sorted(parent_ids):
            name=self.item_lib.get(iid,"—")
            log_parts.append(f"  [parent-box]  {iid:<12}  {name[:40]}")
        log_parts.append("")
        log_parts.append(f"Intermediate boxes traversed (NOT updated — they are boxes, not items):")
        for iid in sorted(traversed - parent_ids):
            name=self.item_lib.get(iid,"—")
            log_parts.append(f"  [inter-box]   {iid:<12}  {name[:40]}")
        log_parts.append("")
        for fname,_,found_map in file_results:
            hits=[(iid,updates[iid]) for iid,hit in found_map.items() if hit]
            if not hits: log_parts.append(f"{fname}  →  No matches"); continue
            log_parts.append(f"{fname}  →  {len(hits)} leaf item(s) updated")
            for iid,ncash in hits:
                name=self.item_lib.get(iid,"—")
                log_parts.append(f"  [leaf-item]   {iid:<12}  {name[:40]:<40}  NCash={ncash}")
            log_parts.append("")
        log_content="\n".join(log_parts)
        exports.append(("ncash_update_log.txt",log_content))
        make_output_tab(nb,"Update Log",log_content,"ncash_update_log.txt",self.root)
        nb.select(0)
        bot=tk.Frame(self,bg=BG); bot.pack(fill="x",pady=6)
        def export_all():
            folder=filedialog.askdirectory(title="Choose export folder")
            if not folder: return
            for efname,content in exports:
                with open(os.path.join(folder,efname),"w",encoding="utf-8") as f: f.write(content)
            messagebox.showinfo("Export Complete",f"Saved to:\n{folder}")
        mk_btn(bot,"💾  Export All Files",export_all,color=ACC4,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=14)
        mk_btn(bot,"◀  Start Over",self._build_load_screen).pack(side="left",padx=4)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 5 — NCash ↔ Ticket Calculator
# ══════════════════════════════════════════════════════════════════════════════
class Tool5(tk.Frame):
    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root=root; self.session=session; self._build()

    def _build(self):
        for w in self.winfo_children(): w.destroy()
        tk.Label(self,text="NCash  ↔  Ticket  Calculator",font=("Consolas",16,"bold"),
                 bg=BG,fg=ACC5).pack(pady=(30,4))
        tk.Label(self,text="Formula:  NCash = round( Tickets × 133 )",
                 font=("Consolas",9),bg=BG,fg=FG_GREY).pack(pady=(0,18))

        box_a=tk.LabelFrame(self,text="  Tickets  →  NCash  ",bg=BG,fg=BLUE,
                             font=("Consolas",10,"bold"),bd=1,relief="groove")
        box_a.pack(fill="x",padx=40,pady=8)
        ra=tk.Frame(box_a,bg=BG); ra.pack(padx=14,pady=12)
        tk.Label(ra,text="Tickets:",width=10,anchor="w",font=("Consolas",11),bg=BG,fg=FG).pack(side="left")
        v_tickets=tk.StringVar()
        tk.Entry(ra,textvariable=v_tickets,width=14,bg=BG3,fg=FG,insertbackground=FG,
                 font=("Consolas",13),relief="flat").pack(side="left",padx=8)
        tk.Label(ra,text="=",font=("Consolas",13),bg=BG,fg=FG_GREY).pack(side="left",padx=4)
        lbl_ncash=tk.Label(ra,text="—",width=14,anchor="w",font=("Consolas",13,"bold"),bg=BG,fg=GREEN)
        lbl_ncash.pack(side="left",padx=4)
        tk.Label(ra,text="NCash",font=("Consolas",10),bg=BG,fg=FG_GREY).pack(side="left")

        box_b=tk.LabelFrame(self,text="  NCash  →  Tickets  ",bg=BG,fg=BLUE,
                             font=("Consolas",10,"bold"),bd=1,relief="groove")
        box_b.pack(fill="x",padx=40,pady=8)
        rb=tk.Frame(box_b,bg=BG); rb.pack(padx=14,pady=12)
        tk.Label(rb,text="NCash:",width=10,anchor="w",font=("Consolas",11),bg=BG,fg=FG).pack(side="left")
        v_ncash=tk.StringVar()
        tk.Entry(rb,textvariable=v_ncash,width=14,bg=BG3,fg=FG,insertbackground=FG,
                 font=("Consolas",13),relief="flat").pack(side="left",padx=8)
        tk.Label(rb,text="=",font=("Consolas",13),bg=BG,fg=FG_GREY).pack(side="left",padx=4)
        lbl_tickets=tk.Label(rb,text="—",width=14,anchor="w",font=("Consolas",13,"bold"),bg=BG,fg=ACC5)
        lbl_tickets.pack(side="left",padx=4)
        tk.Label(rb,text="Tickets",font=("Consolas",10),bg=BG,fg=FG_GREY).pack(side="left")

        def calc_ncash(*_):
            try: lbl_ncash.config(text=f"{round(float(v_tickets.get())*133):,}")
            except: lbl_ncash.config(text="—")
        def calc_tickets(*_):
            try:
                raw=float(v_ncash.get())/133
                lbl_tickets.config(text=f"{int(raw):,}" if raw==int(raw) else f"{round(raw,4):,}")
            except: lbl_tickets.config(text="—")
        v_tickets.trace_add("write",calc_ncash)
        v_ncash.trace_add("write",calc_tickets)

# ══════════════════════════════════════════════════════════════════════════════
# TOOL 6 — ItemParam Generator
# ══════════════════════════════════════════════════════════════════════════════

# ── Reference tables ──────────────────────────────────────────────────────────
_CLASS_MAP  = [(0,"Unselected"),(1,"Item"),(2,"Equipment"),(3,"Drill")]

_TYPE_MAP = [
    (0,"Unselected"),(1,"Galder Coupons"),(2,"Sword Equip"),(3,"Hat Equip"),
    (4,"Inner Equip"),(5,"Fashion (Shoes)"),(6,"Drill"),(7,"Accessory / Sprint"),
    (8,"Shield Equip"),(9,"Unused"),(10,"Potions"),(11,"Cards"),
    (12,"Animated Jewels"),(13,"Pets"),(14,"Compound Etc."),(15,"Useables (Boxes/Ports)"),
    (16,"Throwing Items"),(17,"Ears"),(18,"Tails"),(19,"Fashion (Top/Jacket)"),
    (20,"Fashion (Bottom/Pants)"),(21,"Fashion (Torso/Shoulder)"),(22,"Fashion (Belt)"),
    (23,"Fashion (Accessory)"),(24,"Head Accessory"),(25,"Face Accessory"),
    (26,"Fashion (Hand/Glove)"),(27,"Fashion (Socks)"),(28,"Skill Card"),
    (29,"Ammo"),(30,"Galder (4 stacks)"),(34,"Cape"),(35,"Fortune Card"),
    (999,"Internal (unused)"),
]

_SUBTYPE_MAP = [
    (0,"N/A"),(2,"Sword"),(3,"Hat"),(4,"Innerwear"),(5,"Fashion (Shoe)"),
    (6,"Drill"),(7,"Accessory"),(8,"Shield"),(10,"HP Recovery"),
    (11,"MP Recovery"),(12,"HP&MP Recovery"),(14,"Fashion Set"),
    (18,"Buffalo/Bunny Card"),(19,"Raccoon/Cat Card"),(20,"Lion/Fox Card"),
    (21,"Dragon/Sheep Card"),(22,"NPC Card"),(23,"Power Card"),(24,"Charm Card"),
    (25,"Sense Card"),(26,"Magic Card"),(27,"Old Monster2 Card"),(33,"Neutral Card"),
    (36,"Protein/Lock Candy"),(39,"Throwing Item"),(41,"Ears"),(42,"Tail"),
    (43,"Fashion (Top/Jacket)"),(44,"Fashion (Bottom/Pants)"),
    (45,"Fashion (Torso/Shoulder)"),(46,"Fashion (Belt)"),(47,"Fashion (Accessory)"),
    (48,"Head Accessory"),(49,"Face Accessory"),(50,"Fashion (Hand/Glove)"),
    (51,"Fashion (Socks)"),(52,"GM Amulet"),(53,"Sprint"),(54,"Gun"),
    (55,"Skill Card"),(56,"Ammo"),(61,"MyShop Drill"),(62,"Bracer Accessory"),
    (63,"Repair Powder"),(64,"Galders (Tiny Stack)"),(65,"Galders (Small Stack)"),
    (66,"Galders (Medium Stack)"),(67,"Galders (Large Stack)"),(69,"Scrolls"),
    (70,"Fortune Card"),(71,"Disguise Kit"),(72,"Poseidon/Spinel Seed"),
    (73,"Secret Card"),(74,"Soul Gauge"),(75,"Soul Ticket"),(76,"Soul Feather Pen"),
    (77,"Guardian Runes"),(78,"Custom Character Pets"),(81,"Pet Item Hunt"),
]

_ITEMFTYPE_MAP = [
    (0,"N/A"),(1,"Power"),(2,"Magic"),(3,"Sense"),(4,"Charm"),(5,"Neutral"),
]

_EFFECT_MAP = [
    (0,"N/A"),(11,"Recovery Item"),(12,"Beta Heal (unused)"),(13,"Beta Heal (unused)"),
    (14,"Portable Port"),(15,"Hair Dye"),(16,"Equip Ammo"),(19,"Emergency Port"),
    (20,"Portable Port AD"),(21,"Befana's Unstable Elixir"),(22,"Open Box"),
    (23,"Master's Authority"),(24,"Point Back 40"),(25,"Build Graph Reset"),
    (26,"Point Back All"),(27,"Store More Permit"),(28,"Teleport/Transfer Map"),
    (29,"Resurrect/Revive"),(30,"Recharge Coupon"),(31,"Reidentifier 2000"),
    (33,"Mic"),(34,"Memory Port"),(35,"Read Book/Letter"),(36,"Artisan's Flame"),
    (37,"Star Tear"),(38,"Allen Bottle"),(39,"Friend Finder"),
    (40,"Erin's Secret Book"),(41,"Voice/Auto Loot Pet"),(42,"Poseidon Seed"),
    (43,"Worn Treasure Map"),(44,"Weird Treasure Map"),(45,"Soul Charge"),
    (46,"Fiesta Ticket"),(47,"Character Boxes"),(48,"Earring"),(49,"Locked Boxes"),
    (50,"Key"),(51,"EXP/TM Booster"),(52,"Seals"),
]

_OPTIONS_FULL = [
    (1,    "Rare"),
    (2,    "Usable"),
    (4,    "Select Target Object"),
    (8,    "Select Target Position"),
    (16,   "Equipable"),
    (32,   "Usable To Self"),
    (64,   "Usable To Player"),
    (128,  "Usable To Monster"),
    (256,  "Not Buyable"),
    (512,  "Not Sellable"),
    (1024, "Not Exchangeable"),
    (2048, "Not Pickable"),
    (4096, "Not Droppable"),
    (8192, "Not Vanishable"),
    (16384,"Equipable To ChrType"),
    (32768,"Custom Data"),
    (65536,"Not Addable To Warehouse"),
    (131072,"Not Addable To NeoWarehouse"),
    (262144,"CM (MyShop)"),
    (524288,"Auto Drill"),
    (1048576,"Absolute Unique"),
    (2097152,"Legend (Boss EQ)"),
    (4194304,"Quest"),
    (8388608,"Not Composable"),
    (16777216,"Not Reform"),
    (33554432,"Skin"),
    (67108864,"Event"),
    (134217728,"Dummy (Exchange)"),
    (268435456,"Use Safe Zone"),
]

_OPTIONSEX_MAP = [
    (0,"N/A"),(1,"Apply random stat range"),(2,"Apply elemental property"),
    (16,"Boss equipment set"),(32,"Unknown"),(64,"Non-consumed teleport"),
    (512,"Hair color reset"),(1024,"GM Drill"),(2048,"Unknown (Soul/Tartarus)"),
    (8192,"Unknown (Disabled Swords)"),(16384,"Soul Weapon"),
]

_REFINEINDEX_MAP = [
    (0,"N/A"),(1,"Form/Skin Equipment"),(2,"Fashion"),(21,"Unknown"),
    (22,"Event/Exchange Equipment"),(23,"In-game Obtainable Pets"),
    (37,"Character-specific EQ"),(38,"Various Cash/Gacha Equipment"),
    (39,"Egg Shop Equipment"),(47,"Cash Item Equipment"),
    (48,"Sylvia/Kooh Equipment"),(56,"Cash Pet"),
    (64,"Character/Boss/Voice Pet"),
]

_REFINETYPE_MAP = [
    (0,"N/A"),(1,"AP Sword/Hammer/Lance"),(2,"DP Hat/Shield"),(3,"MA Staff/Cane"),
    (4,"MD Hat/Shield"),(5,"Gun AP"),(6,"DA (Goddess Circlet)"),(7,"HV (Goddess Shield)"),
]

_MINSTATTYPE_MAP = [(0,"Required AP"),(4,"Required MA"),(7,"Required DA")]

# ── Tooltip text for every field ──────────────────────────────────────────────
_TOOLTIPS = {
    # ── Identity ──────────────────────────────────────────────────────────
    "ID":           "Item ID — must be unique in the XML table. REQUIRED.",
    "Class":        "Item class.  1=Item  2=Equipment  3=Drill\n"
                    "⚠ REQUIRED — cannot be 0 or Unselected.",
    "Type":         "Item type category.\n"
                    "1=Galder Coupons  2=Sword  3=Hat  4=Inner  5=Shoes  6=Drill\n"
                    "7=Accessory/Sprint  8=Shield  10=Potions  11=Cards  12=Jewels\n"
                    "13=Pets  14=Compound  15=Useables/Boxes  16=Throwing  17=Ears\n"
                    "18=Tails  19=Top  20=Bottom  21=Torso  22=Belt  23=Fashion Acc\n"
                    "24=Head Acc  25=Face Acc  26=Glove  27=Socks  28=Skill Card\n"
                    "29=Ammo  30=Galder(4 stacks)  34=Cape  35=Fortune Card\n"
                    "⚠ REQUIRED — cannot be 0 or Unselected.",
    "SubType":      "Sub-category of the item type.  0=N/A (default).\n"
                    "Hover for full table or see 📖 Field Reference.",
    "ItemFType":    "Stat affinity.  0=N/A  1=Power  2=Magic  3=Sense  4=Charm  5=Neutral.",
    # ── Names / Text ──────────────────────────────────────────────────────
    "Name":         "The name displayed in-game. CDATA wrapper added automatically.",
    "Comment":      "The item description shown in-game. CDATA wrapper added automatically.",
    "Use":          "The purpose of the item shown in the in-game tooltip.",
    "Name_Eng":     "Dev/localization notes. Rarely visible in-game. Usually a single space.",
    "Comment_Eng":  "May contain dev references or 'description coming soon'. Usually a single space.",
    # ── Files ─────────────────────────────────────────────────────────────
    "FileName":     "Path to the item sprite NRI file.\n"
                    "Example: data\\item\\itm000.nri\n"
                    "This is copied automatically to InvFileName.",
    "BundleNum":    "Sprite index within the NRI file. Starts at 0.\n"
                    "To find your sprite: open NRI in viewer → Animations tab → slot = BundleNum+1.\n"
                    "This is copied automatically to InvBundleNum.",
    "InvFileName":  "Identical to FileName — auto-copied, no manual entry needed.",
    "InvBundleNum": "Identical to BundleNum — auto-copied, no manual entry needed.",
    "CmtFileName":  "Path to the item illustration NRI file (shown in item tooltip/description window).\n"
                    "Example: data\\item\\itm_illu000.nri  (separate from FileName).",
    "CmtBundleNum": "Sprite index within CmtFileName NRI. Same +1 offset rule as BundleNum.",
    "EquipFileName":"Path to the equipment or drill model. Leave blank if not equipment/drill.",
    # ── PivotID / Palette ─────────────────────────────────────────────────
    "PivotID":      "Source item ID reference. Used for equipment with multiple levels or option variants.\n"
                    "Suggested 0 for most items.",
    "PaletteId":    "Palette ID. Leave at 0 for most items. Some exceptions exist but can be ignored.",
    # ── Options ───────────────────────────────────────────────────────────
    "Options":      "Item option flags (eItemOption).  Values OR together.\n"
                    "Key flags: 2=Usable  16=Equipable  32=UsableToSelf  256=NotBuyable\n"
                    "512=NotSellable  262144=CM(MyShop)  268435456=UseSafeZone(Skins)\n"
                    "See 📖 Field Reference for full table.",
    # ── HideHat ───────────────────────────────────────────────────────────
    "HideHat":      "Per-character flag that hides the character's ear model when this item is equipped.\n"
                    "Uses the same flag values as ChrTypeFlags (1st/2nd/3rd job per race).\n"
                    "Suggested 0 unless specifically needed.",
    # ── ChrTypeFlags ──────────────────────────────────────────────────────
    "ChrTypeFlags": "Character-type access flags.  Sum flags for each character/job that should be allowed.\n"
                    "Bunny 1st=1  Buffalo 1st=2  Sheep 1st=4  Dragon 1st=8  Fox 1st=16\n"
                    "Lion 1st=32  Cat 1st=64  Raccoon 1st=124  Paula 1st=256\n"
                    "2nd job: multiply base by 512.  3rd job: see reference table.\n"
                    "0 = no restriction (all characters allowed).  Suggested 0.",
    # ── Ground / System ───────────────────────────────────────────────────
    "GroundFlags":  "Always 0 in standard items. Non-zero may cause unintended behaviour. Default: 0.",
    "SystemFlags":  "Always 0 in standard items. Non-zero may cause unintended behaviour. Default: 0.",
    # ── OptionsEx ─────────────────────────────────────────────────────────
    "OptionsEx":    "Extended item options. Values OR together.\n"
                    "1=Random stat range  2=Elemental property  16=Boss EQ set\n"
                    "64=Non-consumed teleport  512=Hair color reset  1024=GM Drill\n"
                    "2048=Unknown(Soul/Tartarus)  8192=Unknown(Disabled Swords)  16384=Soul Weapon\n"
                    "Mainly used on equipment in ItemParamCM2. Suggested 0 for standard items.",
    # ── Numeric stats ─────────────────────────────────────────────────────
    "Weight":       "Item weight value (WT stat). Affects carry capacity. Default 1.",
    "Value":        "Galder (in-game gold) retail price. 0 = cannot be sold to NPC.",
    "MinLevel":     "Minimum character level required to use or equip the item. Default 1.",
    # ── Effect ────────────────────────────────────────────────────────────
    "Effect":       "Action triggered when the item is used.\n"
                    "22=Open Box (REQUIRED for boxes/Type 15)  14=Portable Port  15=Hair Dye\n"
                    "19=Emergency Port  22=Open Box  29=Resurrect  47=Character Boxes\n"
                    "49=Locked Boxes  50=Key  51=EXP/TM Booster\n"
                    "Multiple effects can be OR'd: enter slash-separated e.g. 22/47\n"
                    "Suggested 0 for non-use items.",
    "EffectFlags2": "Always 0. Default: 0.",
    "SelRange":     "Leave at 0 except for Beta Magic Cards (no longer used). Default: 0.",
    "Life":         "Duration for timed items (EXP/TM Boosters) or drill life span.\n"
                    "0 = no time limit.",
    "Depth":        "Used only for the three test drills in the table. Default: 0.",
    "Delay":        "Tied to Beta Magic Cards (no longer used). Leave at 0.000000. Default: 0.",
    # ── Stat bonuses ─────────────────────────────────────────────────────
    "AP":           "Gun / Throwing Attack Power. NOT the regular AP stat (that is APPlus). Default 0.",
    "HP":           "Amount of HP recovered on use (potions, recovery items). Default 0.",
    "HPCon":        "Amount of HP consumed on use. Rarely used. Default 0.",
    "MP":           "Amount of MP recovered on use. Default 0.",
    "MPCon":        "Amount of MP consumed on use. Also used for Card MP values. Default 0.",
    "Money":        "Amount of Galder obtained from Galder Coupons. Default 0.",
    "APPlus":       "Attack Power stat bonus (equipment). Not to be confused with AP field above.",
    "ACPlus":       "Accuracy stat bonus.",
    "DXPlus":       "Dexterity stat bonus.",
    "MaxMPPlus":    "Maximum Magic Point capacity bonus.",
    "MAPlus":       "Magic Attack stat bonus.",
    "MDPlus":       "Magic Defense stat bonus.",
    "MaxWTPlus":    "Maximum Weight Capacity bonus.",
    "DAPlus":       "Detect Ability stat bonus.",
    "LKPlus":       "Luck stat bonus.",
    "MaxHPPlus":    "Maximum Health Point capacity bonus.",
    "DPPlus":       "Defense Power stat bonus.",
    "HVPlus":       "Evasion stat bonus.",
    "HPRecoveryRate":"HP regen rate (used on pets). Format: 0.000000  Default: 0.000000",
    "MPRecoveryRate":"MP regen rate (used on pets). Format: 0.000000  Default: 0.000000",
    # ── Card params ───────────────────────────────────────────────────────
    "CardNum":      "Card rank. Controls life span and grade range.\n"
                    "0=None/dummy  1=Boss(life 5-6)  2-3=High tier(4-5)  4-6=Mid(4)\n"
                    "7-9=Low/NPC(3)  |  Skill cards: CardNum=7, CardGenGrade=0, CardGenParam=0\n"
                    "Default: 0 for non-card items.",
    "CardGenGrade": "Determines which grade ranges are available for this card.\n"
                    "Default: 0 for non-card items.",
    "CardGenParam": "Unsure. Skill cards always use 0.000000. Format: 0.000000\n"
                    "Default: 0.000000 for non-card items.",
    "DailyGenCnt":  "Possibly a daily card-battle generation limit. Default: 0.",
    # ── Other ─────────────────────────────────────────────────────────────
    "PartFileName": "ItemParamCM2 only — path to the fashion item model. Usually a single space.",
    "ChrFTypeFlag": "Always 0. Default: 0.",
    "ChrGender":    "Always 0. Default: 0.",
    "ExistType":    "Items that cannot stack simultaneously on a character (sprints, boosters, etc).\n"
                    "0=disabled  1=timer / cannot stack.  Actual time values are WIP.",
    "Ncash":        "Cash shop (MyShop) price tag. 0 = not sold in cash shop.",
    "NewCM":        "Always 0. Default: 0.",
    "FamCM":        "Always 0. Default: 0.",
    "Summary":      "Always blank (single space). SUGGESTED blank.",
    "ShopFileName": "Path to promotional Cash Mall item image NRI. Usually a single space.",
    "ShopBundleNum":"Animation sprite index for ShopFileName. Same +1 offset rule as BundleNum. Suggested 0.",
    "MinStatType":  "Stat type required to equip.  0=Required AP  4=Required MA  7=Required DA.\n"
                    "Set MinStatLv=0 to show no requirement.",
    "MinStatLv":    "Minimum stat value required to equip. 0 = no stat requirement shown.\n"
                    "Note: this is the culprit behind no equip-swapping during Rust debuff.",
    "RefineIndex":  "References RefineLevelTable. See 📖 Field Reference for values.\n"
                    "Common: 1=Form/Skin  2=Fashion  22=Event EQ  23=Obtainable Pet\n"
                    "38=Cash/Gacha EQ  47=Cash Item  56=Cash Pet  64=Character/Boss/Voice Pet",
    "RefineType":   "Equipment type for refining.  0=N/A  1=AP Sword  2=DP Hat/Shield\n"
                    "3=MA Staff  4=MD Hat/Shield  5=Gun AP  6=DA Circlet  7=HV Shield",
    "CompoundSlot": "Number of compound slots (0-5). Intended range: 0-5.\n"
                    "Higher values will cause UI errors (EE) when the item is inspected.",
    "SetItemID":    "Equipment set ID reference. Used for set bonuses. Suggested 0.",
    "ReformCount":  "Reform count — possibly tied to skins. Not fully researched. Suggested 0.",
    "GroupId":      "Group ID. Always 0 in standard items. Default: 0.",
    "Money":        "Amount of Galder from coupons. Default 0.",
}

# ── Tooltip widget helper ─────────────────────────────────────────────────────
def _attach_tooltip(widget, text):
    tip = None
    def show(e):
        nonlocal tip
        if tip: return
        x = widget.winfo_rootx() + 20
        y = widget.winfo_rooty() + widget.winfo_height() + 4
        tip = tk.Toplevel(widget)
        tip.wm_overrideredirect(True)
        tip.wm_geometry(f"+{x}+{y}")
        tk.Label(tip, text=text, justify="left", bg="#313244", fg="#cdd6f4",
                 font=("Consolas", 8), relief="flat", bd=1, wraplength=380,
                 padx=6, pady=4).pack()
    def hide(e):
        nonlocal tip
        if tip: tip.destroy(); tip = None
    widget.bind("<Enter>", show, add="+")
    widget.bind("<Leave>", hide, add="+")

# ── Persistent settings helper ───────────────────────────────────────────────

# ── Per-tool persistent settings (separate files per tool) ──────────────────
_SETTINGS_DIR = os.path.expanduser("~")

def _settings_path(tool_key):
    return os.path.join(_SETTINGS_DIR, f".box_tool_suite_{tool_key}.json")

def _load_settings(tool_key):
    try:
        with open(_settings_path(tool_key), encoding="utf-8") as f:
            return _json.load(f)
    except Exception:
        return {}

def _save_settings(tool_key, data):
    try:
        with open(_settings_path(tool_key), "w", encoding="utf-8") as f:
            _json.dump(data, f, indent=2)
    except Exception:
        pass

# Backwards-compat aliases for Tool6
def _load_t6_settings():    return _load_settings("t6")
def _save_t6_settings(d):   _save_settings("t6", d)

def _get_last_id(tool_key, default=1):
    """Return the last-used ID for a given tool."""
    return _load_settings(tool_key).get("last_id", default)

def _set_last_id(tool_key, id_val):
    d = _load_settings(tool_key)
    d["last_id"] = id_val
    _save_settings(tool_key, d)

# ── XML builder for generic ItemParam ────────────────────────────────────────
def _parse_effect_val(raw):
    """Parse Effect field — may be int, '0', '22', '22/47', etc."""
    if isinstance(raw, int): return raw
    raw = str(raw).strip()
    parts = [p.strip() for p in raw.replace("/",",").split(",") if p.strip()]
    vals = []
    for p in parts:
        try: vals.append(int(p))
        except: pass
    if not vals: return 0
    # Single value -> return int; multiple -> slash-separated string
    if len(vals) == 1: return vals[0]
    return "/".join(str(v) for v in vals)

def build_generic_itemparam_row(cfg):
    """Build a full ItemParam <ROW> from a cfg dict."""
    def _cd(v):  return f"<![CDATA[{v}]]>"
    def _fmt6(v):
        try:    return f"{float(v):.6f}"
        except: return "0.000000"
    def _int(v, default=0):
        try:    return int(v)
        except: return default

    # Options bitmask from list of checked flags
    opts_val = 0
    for flag, checked in zip([f for f,_ in _OPTIONS_FULL], cfg.get("options_flags", [])):
        if checked: opts_val |= flag
    # OptionsEx — slash-separated (stored as "16/32" or "0")
    _optex_raw = str(cfg.get("options_ex", 0)).strip()
    # If it's a plain integer (legacy), decompose it into flags
    if _optex_raw.lstrip("-").isdigit():
        _optex_int = int(_optex_raw)
        _optex_flags = [str(f) for f,_ in _OPTIONSEX_MAP if f > 0 and (_optex_int & f)]
        optex_val = "/".join(_optex_flags) if _optex_flags else "0"
    else:
        # Already slash-separated — pass through as-is
        optex_val = _optex_raw if _optex_raw else "0"

    lines = [
        "<ROW>",
        f"<ID>{cfg['id']}</ID>",
        f"<Class>{cfg.get('class_val','1')}</Class>",
        f"<Type>{cfg.get('type_val','15')}</Type>",
        f"<SubType>{cfg.get('subtype_val','0')}</SubType>",
        f"<ItemFType>{cfg.get('itemftype_val','0')}</ItemFType>",
        f"<n>{_cd(cfg.get('name',''))}</n>",
        f"<Comment>{_cd(cfg.get('comment',''))}</Comment>",
        f"<Use>{_cd(cfg.get('use',''))}</Use>",
        f"<Name_Eng>{_cd(cfg.get('name_eng',' '))}</Name_Eng>",
        f"<Comment_Eng>{_cd(cfg.get('comment_eng',' '))}</Comment_Eng>",
        f"<FileName>{_cd(cfg.get('file_name',''))}</FileName>",
        f"<BundleNum>{_int(cfg.get('bundle_num',0))}</BundleNum>",
        f"<InvFileName>{_cd(cfg.get('file_name',''))}</InvFileName>",
        f"<InvBundleNum>{_int(cfg.get('bundle_num',0))}</InvBundleNum>",
        f"<CmtFileName>{_cd(cfg.get('cmt_file_name',''))}</CmtFileName>",
        f"<CmtBundleNum>{_int(cfg.get('cmt_bundle_num',0))}</CmtBundleNum>",
        f"<EquipFileName>{_cd(cfg.get('equip_file_name',''))}</EquipFileName>",
        f"<PivotID>{_int(cfg.get('pivot_id',0))}</PivotID>",
        f"<PaletteId>{_int(cfg.get('palette_id',0))}</PaletteId>",
        f"<Options>{opts_val}</Options>",
        f"<HideHat>{_int(cfg.get('hide_hat',0))}</HideHat>",
        f"<ChrTypeFlags>{_int(cfg.get('chr_type_flags',0))}</ChrTypeFlags>",
        f"<GroundFlags>{_int(cfg.get('ground_flags',0))}</GroundFlags>",
        f"<SystemFlags>{_int(cfg.get('system_flags',0))}</SystemFlags>",
        f"<OptionsEx>{optex_val}</OptionsEx>",
        f"<Weight>{_int(cfg.get('weight',1))}</Weight>",
        f"<Value>{_int(cfg.get('value',0))}</Value>",
        f"<MinLevel>{_int(cfg.get('min_level',1))}</MinLevel>",
        f"<Effect>{_parse_effect_val(cfg.get('effect',0))}</Effect>",
        f"<EffectFlags2>{_int(cfg.get('effect_flags2',0))}</EffectFlags2>",
        f"<SelRange>{_int(cfg.get('sel_range',0))}</SelRange>",
        f"<Life>{_int(cfg.get('life',0))}</Life>",
        f"<Depth>{_int(cfg.get('depth',0))}</Depth>",
        f"<Delay>{_fmt6(cfg.get('delay',0))}</Delay>",
        f"<AP>{_int(cfg.get('ap',0))}</AP>",
        f"<HP>{_int(cfg.get('hp',0))}</HP>",
        f"<HPCon>{_int(cfg.get('hpcon',0))}</HPCon>",
        f"<MP>{_int(cfg.get('mp',0))}</MP>",
        f"<MPCon>{_int(cfg.get('mpcon',0))}</MPCon>",
        f"<Money>{_int(cfg.get('money',0))}</Money>",
        f"<APPlus>{_int(cfg.get('applus',0))}</APPlus>",
        f"<ACPlus>{_int(cfg.get('acplus',0))}</ACPlus>",
        f"<DXPlus>{_int(cfg.get('dxplus',0))}</DXPlus>",
        f"<MaxMPPlus>{_int(cfg.get('maxmpplus',0))}</MaxMPPlus>",
        f"<MAPlus>{_int(cfg.get('maplus',0))}</MAPlus>",
        f"<MDPlus>{_int(cfg.get('mdplus',0))}</MDPlus>",
        f"<MaxWTPlus>{_int(cfg.get('maxwtplus',0))}</MaxWTPlus>",
        f"<DAPlus>{_int(cfg.get('daplus',0))}</DAPlus>",
        f"<LKPlus>{_int(cfg.get('lkplus',0))}</LKPlus>",
        f"<MaxHPPlus>{_int(cfg.get('maxhpplus',0))}</MaxHPPlus>",
        f"<DPPlus>{_int(cfg.get('dpplus',0))}</DPPlus>",
        f"<HVPlus>{_int(cfg.get('hvplus',0))}</HVPlus>",
        f"<HPRecoveryRate>{_fmt6(cfg.get('hprecoveryrate',0))}</HPRecoveryRate>",
        f"<MPRecoveryRate>{_fmt6(cfg.get('mprecoveryrate',0))}</MPRecoveryRate>",
        f"<CardNum>{_int(cfg.get('cardnum',0))}</CardNum>",
        f"<CardGenGrade>{_int(cfg.get('cardgengrade',0))}</CardGenGrade>",
        f"<CardGenParam>{_fmt6(cfg.get('cardgenparam',0))}</CardGenParam>",
        f"<DailyGenCnt>{_int(cfg.get('dailygencnt',0))}</DailyGenCnt>",
        f"<PartFileName>{_cd(cfg.get('part_file_name',' '))}</PartFileName>",
        f"<ChrFTypeFlag>{_int(cfg.get('chr_ftype_flag',0))}</ChrFTypeFlag>",
        f"<ChrGender>{_int(cfg.get('chr_gender',0))}</ChrGender>",
        f"<ExistType>{_int(cfg.get('exist_type',0))}</ExistType>",
        f"<Ncash>{_int(cfg.get('ncash',0))}</Ncash>",
        f"<NewCM>{_int(cfg.get('new_cm',0))}</NewCM>",
        f"<FamCM>{_int(cfg.get('fam_cm',0))}</FamCM>",
        f"<Summary>{_cd(cfg.get('summary',' '))}</Summary>",
        f"<ShopFileName>{_cd(cfg.get('shop_file_name',' '))}</ShopFileName>",
        f"<ShopBundleNum>{_int(cfg.get('shop_bundle_num',0))}</ShopBundleNum>",
        f"<MinStatType>{_int(cfg.get('min_stat_type',0))}</MinStatType>",
        f"<MinStatLv>{_int(cfg.get('min_stat_lv',0))}</MinStatLv>",
        f"<RefineIndex>{_int(cfg.get('refine_index',0))}</RefineIndex>",
        f"<RefineType>{_int(cfg.get('refine_type',0))}</RefineType>",
        f"<CompoundSlot>{_int(cfg.get('compound_slot',0))}</CompoundSlot>",
        f"<SetItemID>{_int(cfg.get('set_item_id',0))}</SetItemID>",
        f"<ReformCount>{_int(cfg.get('reform_count',0))}</ReformCount>",
        f"<GroupId>{_int(cfg.get('group_id',0))}</GroupId>",
        "</ROW>",
    ]
    return "\n".join(lines)


# ══════════════════════════════════════════════════════════════════════════════
# COMPOUND / EXCHANGE XML BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def build_compound_row(cfg):
    """Build Compound_Potion.xml <ROW>.
    CompoundID / Name / Comment / ResLv / ResID1-3 /
    ReqID1+ReqNum1 ... ReqID5+ReqNum5 / Probability / Fee / WasteItem.
    ResID  = item IDs received on success (up to 3).
    ReqID  = ingredient item IDs (up to 5); ReqNum = quantity each.
    Probability = success % (0-100).
    Fee        = galder cost per attempt.
    WasteItem  = item consumed on failure (default 12000).
    """
    def _cd(v): return f"<![CDATA[{v}]]>"
    def _i(v, d=0):
        try: return int(v)
        except: return d
    lines = ["<ROW>",
        f"<CompoundID>{_i(cfg.get('compound_id',0))}</CompoundID>",
        f"<Name>{_cd(cfg.get('name',''))}</Name>",
        f"<Comment>{_cd(cfg.get('comment',''))}</Comment>",
        f"<ResLv>{_i(cfg.get('res_lv',1))}</ResLv>",
    ]
    for n in range(1,4):
        lines.append(f"<ResID{n}>{_i(cfg.get(f'res_id{n}',0))}</ResID{n}>")
    for n in range(1,6):
        lines.append(f"<ReqID{n}>{_i(cfg.get(f'req_id{n}',0))}</ReqID{n}>")
        lines.append(f"<ReqNum{n}>{_i(cfg.get(f'req_num{n}',0))}</ReqNum{n}>")
    lines += [
        f"<Probability>{_i(cfg.get('probability',50))}</Probability>",
        f"<Fee>{_i(cfg.get('fee',1))}</Fee>",
        f"<WasteItem>{_i(cfg.get('waste_item',12000))}</WasteItem>",
        "</ROW>",
    ]
    return "\n".join(lines)

def build_compound_location_row(compound_id):
    """Build Compounder_Location.xml <ROW>.
    CompoundID matches the Compound_Potion.xml entry.
    Probability and Hidden are always 0 in this file.
    """
    return f"<ROW>\n<CompoundID>{compound_id}</CompoundID>\n<Probability>0</Probability>\n<Hidden>0</Hidden>\n</ROW>"

def build_exchange_row(cfg):
    """Build ExchangeShopContents.xml <ROW>."""
    def _cd(v): return f"<![CDATA[{v}]]>"
    def _i(v, d=0):
        try: return int(v)
        except: return d
    lines = ["<ROW>",
        f"<ExchangeID>{_i(cfg.get('exchange_id',0))}</ExchangeID>",
        f"<Name>{_cd(cfg.get('name',''))}</Name>",
        f"<Comment>{_cd(cfg.get('comment',''))}</Comment>",
        f"<ResLv>{_i(cfg.get('res_lv',1))}</ResLv>",
    ]
    for n in range(1,4):
        lines.append(f"<ResID{n}>{_i(cfg.get(f'res_id{n}',0))}</ResID{n}>")
    for n in range(1,6):
        lines.append(f"<ReqID{n}>{_i(cfg.get(f'req_id{n}',0))}</ReqID{n}>")
        lines.append(f"<ReqNum{n}>{_i(cfg.get(f'req_num{n}',0))}</ReqNum{n}>")
    lines += [f"<Fee>{_i(cfg.get('fee',0))}</Fee>", "</ROW>"]
    return "\n".join(lines)

def build_exchange_location_row(exchange_id):
    """Build Exchange_Location.xml <ROW> — all fields always present."""
    return (
        f"<ROW>\n"
        f"<ExchangeID>{exchange_id}</ExchangeID>\n"
        f"<Map><![CDATA[ ]]></Map>\n"
        f"<NPCTypeID>0</NPCTypeID>\n"
        f"<Hidden>0</Hidden>\n"
        f"</ROW>"
    )

# Session-level CE preference store (runtime only, not persisted to file)


def _show_box_ce_dialog(root, box_configs, compound_rows, exchange_rows, on_done_cb):
    """
    Per-box Exchange / Compound dialog for Tool1.
    Each box gets its own row: type selector, unique ID, ResID1 (what box gives),
    ResLv, ReqID1, ReqNum1, Fee.
    No group-selection nonsense — one box = one row = one XML row.
    """
    win = tk.Toplevel(root)
    win.title("Add Exchange / Compound — Per Box")
    win.configure(bg=BG)
    win.geometry("1100x640")
    win.minsize(900, 500)

    last_e = _get_last_id("exchange", 0)
    last_c = _get_last_id("compound", 100)

    tk.Label(win, text="Exchange / Compound — Per Box",
             font=("Consolas", 13, "bold"), bg=BG, fg=ACC1).pack(pady=(12,2))
    tk.Label(win,
             text="Each box gets its own row. Set type, IDs, and values. Leave type as None to skip that box.",
             bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(pady=(0,8))

    # ── Scrollable table ──────────────────────────────────────────────────────
    outer = tk.Frame(win, bg=BG); outer.pack(fill="both", expand=True, padx=10)
    canv = tk.Canvas(outer, bg=BG, bd=0, highlightthickness=0)
    vsb  = tk.Scrollbar(outer, orient="vertical", command=canv.yview)
    canv.configure(yscrollcommand=vsb.set)
    vsb.pack(side="right", fill="y"); canv.pack(side="left", fill="both", expand=True)
    tbl = tk.Frame(canv, bg=BG)
    cw = canv.create_window((0,0), window=tbl, anchor="nw")
    tbl.bind("<Configure>", lambda e: canv.configure(scrollregion=canv.bbox("all")))
    canv.bind("<Configure>", lambda e: canv.itemconfig(cw, width=e.width))

    # Header
    COL_W = [6, 28, 8, 8, 8, 8, 10, 10, 6]
    HDRS  = ["Type","Box Name","CE/Cmp ID","ResID1","ResLv","ReqID1","ReqNum1","Fee","ResID2"]
    hf = tk.Frame(tbl, bg=BG2); hf.pack(fill="x")
    for i,(h,w) in enumerate(zip(HDRS, COL_W)):
        tk.Label(hf, text=h, width=w, bg=BG2, fg=BLUE,
                 font=("Consolas",8,"bold"), anchor="w").grid(row=0, column=i, padx=2, pady=2)

    # One row per box
    row_vars = []
    for bi, cfg in enumerate(box_configs):
        box_id   = cfg.get("id", "")
        box_name = cfg.get("name", "")

        # Auto-increment IDs
        last_e += 1; last_c += 1
        e_id_default = str(last_e)
        c_id_default = str(last_c)

        bg = BG if bi % 2 == 0 else BG2
        rf = tk.Frame(tbl, bg=bg); rf.pack(fill="x")

        type_var   = tk.StringVar(value="Exchange")
        id_var     = tk.StringVar(value=e_id_default)
        res1_var   = tk.StringVar(value=box_id)   # box gives itself by default
        reslv_var  = tk.StringVar(value="1")
        req1_var   = tk.StringVar(value="")
        reqn1_var  = tk.StringVar(value="1")
        fee_var    = tk.StringVar(value="0")
        res2_var   = tk.StringVar(value="0")

        # When type changes, swap the default ID
        def _on_type_change(tv=type_var, iv=id_var, eid=e_id_default, cid=c_id_default):
            iv.set(eid if tv.get() == "Exchange" else cid if tv.get() == "Compound" else "")

        type_cb = ttk.Combobox(rf, textvariable=type_var,
                               values=["Exchange","Compound","None"],
                               width=9, state="readonly")
        type_cb.grid(row=0, column=0, padx=2, pady=2, sticky="w")
        type_cb.bind("<<ComboboxSelected>>", lambda e, tv=type_var, iv=id_var,
                     eid=e_id_default, cid=c_id_default:
                     iv.set(eid if tv.get()=="Exchange" else cid if tv.get()=="Compound" else ""))

        tk.Label(rf, text=f"{box_id}  {box_name}", width=28, anchor="w",
                 bg=bg, fg=FG, font=("Consolas",8)).grid(row=0, column=1, padx=2)

        for ci, (var, w) in enumerate(zip(
                [id_var, res1_var, reslv_var, req1_var, reqn1_var, fee_var, res2_var],
                [8, 8, 6, 10, 8, 8, 8]), start=2):
            tk.Entry(rf, textvariable=var, width=w, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",8), relief="flat"
                     ).grid(row=0, column=ci, padx=2, pady=2)

        row_vars.append((type_var, id_var, res1_var, reslv_var,
                         req1_var, reqn1_var, fee_var, res2_var, cfg))

    # ── Buttons ───────────────────────────────────────────────────────────────
    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom", pady=0)

    def _confirm():
        new_e = 0; new_c = 0
        for (type_var, id_var, res1_var, reslv_var,
             req1_var, reqn1_var, fee_var, res2_var, cfg) in row_vars:
            t = type_var.get()
            if t == "None" or not t: continue
            cid  = id_var.get().strip()
            name = cfg.get("name","")
            r1   = res1_var.get().strip() or cfg.get("id","0")
            r2   = res2_var.get().strip() or "0"
            rlv  = reslv_var.get().strip() or "1"
            q1   = req1_var.get().strip() or "0"
            qn1  = reqn1_var.get().strip() or "1"
            fee  = fee_var.get().strip() or "0"

            if t == "Exchange":
                ecfg = {"exchange_id": cid, "name": name, "comment": "",
                        "res_lv": rlv, "res_id1": r1, "res_id2": r2, "res_id3": "0",
                        "req_id1": q1, "req_num1": qn1,
                        "req_id2":"0","req_num2":"0","req_id3":"0","req_num3":"0",
                        "req_id4":"0","req_num4":"0","req_id5":"0","req_num5":"0",
                        "fee": fee}
                exchange_rows.append((build_exchange_row(ecfg),
                                      build_exchange_location_row(cid)))
                try: new_e = max(new_e, int(cid))
                except: pass

            elif t == "Compound":
                ccfg = {"compound_id": cid, "name": name, "comment": "",
                        "res_lv": rlv, "res_id1": r1, "res_id2": r2, "res_id3": "0",
                        "req_id1": q1, "req_num1": qn1,
                        "req_id2":"0","req_num2":"0","req_id3":"0","req_num3":"0",
                        "req_id4":"0","req_num4":"0","req_id5":"0","req_num5":"0",
                        "probability":"100", "fee": fee, "waste_item":"12000"}
                compound_rows.append((build_compound_row(ccfg),
                                      build_compound_location_row(cid)))
                try: new_c = max(new_c, int(cid))
                except: pass

        if new_e: _set_last_id("exchange", new_e)
        if new_c: _set_last_id("compound", new_c)
        win.destroy()
        on_done_cb()

    mk_btn(nav, "✓  Confirm", _confirm,
           color=GREEN, fg=BG2, font=("Consolas",11,"bold")).pack(side="right", padx=14, pady=8)
    mk_btn(nav, "✗  Cancel", win.destroy, color=BG4).pack(side="right", padx=4, pady=8)


def _show_multi_ce_picker(root, all_items, on_done, mode_hint="compound"):
    """Multi-select compound/exchange/shop picker.

    all_items: list of dicts with keys: id, name, comment
    on_done(compound_cfgs, exchange_cfgs, shop_cfgs):  called when user confirms.
      Each cfg list is a list of dicts ready for build_compound_row / build_exchange_row /
      build_shop_row respectively.

    Layout: left pane = item selector checklist, right pane = mode config + grouping.
    User can select a subset, configure their CE/shop entry, confirm, then come back
    to select a new round.
    """
    win = tk.Toplevel(root)
    win.title("Add Compound / Exchange / Shop  —  Multi-Select")
    win.configure(bg=BG)
    win.geometry("1080x640")
    win.minsize(960, 560)
    win.grab_set()

    compound_out = []
    exchange_out = []
    shop_out     = []

    # ── Header ─────────────────────────────────────────────────────────────
    hdr = tk.Frame(win, bg=BG2); hdr.pack(fill="x")
    tk.Label(hdr, text="Add Compound / Exchange / Shop",
             bg=BG2, fg=ACC7, font=("Consolas",12,"bold"), pady=8).pack(side="left", padx=14)
    tk.Label(hdr,
             text="Select items on the left → configure on the right → Add Group.\n"
                  "Repeat for new groupings. Confirm when all groups are done.",
             bg=BG2, fg=FG_DIM, font=("Consolas",8), justify="left").pack(side="left", padx=10)

    # ── Split pane ─────────────────────────────────────────────────────────
    body = tk.Frame(win, bg=BG); body.pack(fill="both", expand=True, padx=8, pady=4)
    body.grid_columnconfigure(0, weight=1)
    body.grid_columnconfigure(2, weight=2)
    body.grid_rowconfigure(0, weight=1)

    # ── LEFT: item checklist ───────────────────────────────────────────────
    left_outer = tk.Frame(body, bg=BG2, bd=1, relief="groove")
    left_outer.grid(row=0, column=0, sticky="nsew", padx=(0,4))
    tk.Label(left_outer, text="① Select Items", bg=BG2, fg=BLUE,
             font=("Consolas",10,"bold"), pady=6).pack(anchor="w", padx=10)

    sel_all_var = tk.BooleanVar(value=False)
    sel_bar = tk.Frame(left_outer, bg=BG2); sel_bar.pack(fill="x", padx=8, pady=(0,4))
    def _toggle_all(*_):
        for v in check_vars:
            v.set(sel_all_var.get())
    tk.Checkbutton(sel_bar, text="Select All", variable=sel_all_var,
                   command=_toggle_all, bg=BG2, fg=FG, selectcolor=BG3,
                   activebackground=BG2, font=("Consolas",8)).pack(side="left")

    list_sh = tk.Frame(left_outer, bg=BG2); list_sh.pack(fill="both", expand=True, padx=4)
    canv_l, C_l = mk_scroll_canvas(list_sh)

    check_vars = []
    item_rows  = []   # (var, id, name)

    def _on_check_change(*_):
        """When any checkbox changes, auto-populate config fields with first selected item."""
        first = next(((iid, nm, cmt) for v, iid, nm, cmt in item_rows if v.get()), None)
        if first is None:
            return
        iid, nm, cmt = first
        m = mode_var.get()
        if m == "compound":
            if not comp_vars.get("res_id1") or not comp_vars["res_id1"].get():
                comp_vars.setdefault("res_id1", tk.StringVar()).set(iid)
            else:
                comp_vars["res_id1"].set(iid)
            if "name" in comp_vars and not comp_vars["name"].get():
                comp_vars["name"].set(nm)
        elif m == "exchange":
            if "res_id1" in exch_vars:
                exch_vars["res_id1"].set(iid)
            if "name" in exch_vars and not exch_vars["name"].get():
                exch_vars["name"].set(nm)
    for it in all_items:
        v = tk.BooleanVar(value=False)
        row = tk.Frame(C_l, bg=BG2); row.pack(fill="x", pady=1)
        cb = tk.Checkbutton(row, variable=v, bg=BG2, fg=FG, selectcolor=BG3,
                             activebackground=BG2, font=("Consolas",8))
        cb.pack(side="left")
        lbl_txt = f"{it.get('id','')}  {it.get('name','')[:36]}"
        tk.Label(row, text=lbl_txt, bg=BG2, fg=FG,
                 font=("Consolas",8), anchor="w").pack(side="left")
        check_vars.append(v)
        item_rows.append((v, it.get("id",""), it.get("name",""), it.get("comment","")))
        v.trace_add("write", _on_check_change)

    # ── Divider ────────────────────────────────────────────────────────────
    tk.Frame(body, bg=BG4, width=2).grid(row=0, column=1, sticky="ns", padx=2)

    # ── RIGHT: mode + config ───────────────────────────────────────────────
    right_outer = tk.Frame(body, bg=BG); right_outer.grid(row=0, column=2, sticky="nsew")

    right_sh = tk.Frame(right_outer, bg=BG); right_sh.pack(fill="both", expand=True)
    canv_r, C_r = mk_scroll_canvas(right_sh)

    # Mode selector
    mode_sec = mk_section(C_r, "  ② Choose Type  ")
    mode_var = tk.StringVar(value=mode_hint)
    mode_detail = tk.Frame(C_r, bg=BG)
    mode_detail.pack(fill="x", padx=8)

    def _rebuild_mode_detail(*_):
        for w in mode_detail.winfo_children(): w.destroy()
        m = mode_var.get()
        if m == "compound":
            _build_compound_fields(mode_detail)
        elif m == "exchange":
            _build_exchange_fields(mode_detail)
        elif m == "shop":
            _build_shop_fields(mode_detail)
        elif m == "none":
            tk.Label(mode_detail, text="  Nothing will be generated for selected items.",
                     bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", pady=8)

    mode_row = tk.Frame(mode_sec, bg=BG); mode_row.pack(fill="x", padx=8, pady=4)
    for lbl2, val2 in [("Compound", "compound"), ("Exchange", "exchange"),
                        ("R_ShopItem", "shop"), ("None / Skip", "none")]:
        tk.Radiobutton(mode_row, text=lbl2, variable=mode_var, value=val2,
                       bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                       font=("Consolas",9),
                       command=_rebuild_mode_detail).pack(side="left", padx=8)

    # ── Compound fields ────────────────────────────────────────────────────
    comp_vars = {}
    def _build_compound_fields(parent):
        sec = mk_section(parent, "  Compound_Potion.xml config  ")
        last_c = _get_last_id("compound", 100)
        for key, lbl, default in [
            ("compound_id", "CompoundID:",  str(last_c + 1)),
            ("name",        "Name:",        ""),
            ("comment",     "Comment:",     ""),
            ("res_lv",      "ResLv:",       "1"),
            ("res_id1",     "ResID1:",      ""),
            ("res_id2",     "ResID2:",      "0"),
            ("res_id3",     "ResID3:",      "0"),
            ("req_id1",     "ReqID1:",      ""),
            ("req_num1",    "ReqNum1:",     "1"),
            ("req_id2",     "ReqID2:",      "0"),
            ("req_num2",    "ReqNum2:",     "0"),
            ("req_id3",     "ReqID3:",      "0"),
            ("req_num3",    "ReqNum3:",     "0"),
            ("req_id4",     "ReqID4:",      "0"),
            ("req_num4",    "ReqNum4:",     "0"),
            ("req_id5",     "ReqID5:",      "0"),
            ("req_num5",    "ReqNum5:",     "0"),
            ("probability", "Probability:", "50"),
            ("fee",         "Fee:",         "1"),
            ("waste_item",  "WasteItem:",   "12000"),
        ]:
            v = comp_vars.get(key, tk.StringVar(value=default))
            comp_vars[key] = v
            r = tk.Frame(sec, bg=BG); r.pack(fill="x", padx=8, pady=1)
            tk.Label(r, text=lbl, width=14, anchor="w", bg=BG, fg=FG,
                     font=("Consolas",8)).pack(side="left")
            tk.Entry(r, textvariable=v, width=20, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",8), relief="flat").pack(side="left", padx=4)

    # ── Exchange fields ────────────────────────────────────────────────────
    exch_vars = {}
    def _build_exchange_fields(parent):
        sec = mk_section(parent, "  ExchangeShopContents.xml config  ")
        last_e = _get_last_id("exchange", 0)
        for key, lbl, default in [
            ("exchange_id", "ExchangeID:", str(last_e + 1)),
            ("name",        "Name:",       ""),
            ("comment",     "Comment:",    ""),
            ("res_lv",      "ResLv:",      "1"),
            ("res_id1",     "ResID1:",     ""),
            ("res_id2",     "ResID2:",     "0"),
            ("res_id3",     "ResID3:",     "0"),
            ("req_id1",     "ReqID1:",     ""),
            ("req_num1",    "ReqNum1:",    "1"),
            ("req_id2",     "ReqID2:",     "0"),
            ("req_num2",    "ReqNum2:",    "0"),
            ("req_id3",     "ReqID3:",     "0"),
            ("req_num3",    "ReqNum3:",    "0"),
            ("req_id4",     "ReqID4:",     "0"),
            ("req_num4",    "ReqNum4:",    "0"),
            ("req_id5",     "ReqID5:",     "0"),
            ("req_num5",    "ReqNum5:",    "0"),
            ("fee",         "Fee:",        "0"),
        ]:
            v = exch_vars.get(key, tk.StringVar(value=default))
            exch_vars[key] = v
            r = tk.Frame(sec, bg=BG); r.pack(fill="x", padx=8, pady=1)
            tk.Label(r, text=lbl, width=14, anchor="w", bg=BG, fg=FG,
                     font=("Consolas",8)).pack(side="left")
            tk.Entry(r, textvariable=v, width=20, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",8), relief="flat").pack(side="left", padx=4)

    # ── Shop fields ────────────────────────────────────────────────────────
    shop_vars = {}
    def _build_shop_fields(parent):
        sec = mk_section(parent, "  R_ShopItem.xml config  ")
        for key, lbl, default in [
            ("count", "Count:",         "100"),
            ("price", "Price (NCash):", "0"),
        ]:
            v = shop_vars.get(key, tk.StringVar(value=default))
            shop_vars[key] = v
            r = tk.Frame(sec, bg=BG); r.pack(fill="x", padx=8, pady=2)
            tk.Label(r, text=lbl, width=14, anchor="w", bg=BG, fg=FG,
                     font=("Consolas",9)).pack(side="left")
            tk.Entry(r, textvariable=v, width=12, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)

    # Status label showing what has been queued
    status_sec = mk_section(C_r, "  ③ Queued Groups  ")
    status_lbl = tk.Label(status_sec, text="(nothing queued yet)",
                          bg=BG, fg=FG_GREY, font=("Consolas",8), justify="left")
    status_lbl.pack(anchor="w", padx=10, pady=4)

    def _update_status():
        parts = []
        if compound_out: parts.append(f"{len(compound_out)} compound")
        if exchange_out:  parts.append(f"{len(exchange_out)} exchange")
        if shop_out:       parts.append(f"{len(shop_out)} shop")
        status_lbl.config(text=", ".join(parts) if parts else "(nothing queued yet)",
                          fg=GREEN if parts else FG_GREY)

    def _add_group():
        """Add CE/Shop entries for all selected items using current config."""
        selected = [(iid, nm, cmt) for v, iid, nm, cmt in item_rows if v.get()]
        if not selected:
            messagebox.showwarning("No Items", "Select at least one item on the left.", parent=win)
            return
        m = mode_var.get()
        if m == "compound":
            cfg = {k: v.get() for k, v in comp_vars.items()}
            # ResID1 defaults to each selected item's ID if blank
            for iid, nm, cmt in selected:
                c = dict(cfg)
                if not c.get("res_id1"): c["res_id1"] = iid
                if not c.get("name"):    c["name"]    = nm
                if not c.get("comment"): c["comment"] = cmt
                compound_out.append(c)
                # Auto-increment compound_id
                try:
                    comp_vars["compound_id"].set(str(int(comp_vars["compound_id"].get()) + 1))
                except: pass
        elif m == "exchange":
            cfg = {k: v.get() for k, v in exch_vars.items()}
            for iid, nm, cmt in selected:
                c = dict(cfg)
                if not c.get("res_id1"): c["res_id1"] = iid
                if not c.get("name"):    c["name"]    = nm
                if not c.get("comment"): c["comment"] = cmt
                exchange_out.append(c)
                try:
                    exch_vars["exchange_id"].set(str(int(exch_vars["exchange_id"].get()) + 1))
                except: pass
        elif m == "shop":
            count = shop_vars.get("count", tk.StringVar(value="100")).get()
            price = shop_vars.get("price", tk.StringVar(value="0")).get()
            for iid, nm, cmt in selected:
                shop_out.append({"id": iid, "count": count, "price": price, "name": nm})

        # Uncheck all items so user can pick a new round
        for v in check_vars: v.set(False)
        sel_all_var.set(False)
        _update_status()
        messagebox.showinfo("Group Added",
            f"Added {len(selected)} item(s) to {m} queue.\n"
            "Unselected all items — pick a new group or click Confirm.",
            parent=win)

    _rebuild_mode_detail()

    # ── Footer ─────────────────────────────────────────────────────────────
    footer = tk.Frame(win, bg=BG2); footer.pack(fill="x", side="bottom")

    mk_btn(footer, "➕  Add Group for Selected", _add_group,
           color=ACC7, fg=BG2, font=("Consolas",10,"bold")).pack(side="left", padx=12, pady=8)
    tk.Label(footer,
             text="Add group, then pick a new round of items.\nRepeat until done, then Confirm.",
             bg=BG2, fg=FG_DIM, font=("Consolas",7)).pack(side="left", padx=8)

    def _confirm():
        win.destroy()
        on_done(compound_out, exchange_out, shop_out)

    mk_btn(footer, "✓  Confirm All Groups", _confirm,
           color=GREEN, fg=BG2, font=("Consolas",10,"bold")).pack(side="right", padx=12, pady=8)
    mk_btn(footer, "Cancel", win.destroy, color=BG4).pack(side="right", padx=4, pady=8)

    win.wait_window()


_ce_session_pref = {"choice": None}   # None = always ask, else "compound"/"exchange"/"none"

def _show_compound_exchange_dialog(root, item_name, item_comment, item_id,
                                   on_compound, on_exchange, on_skip,
                                   remember_var=None, box_effect=None):
    """Popup asking whether to also generate compound/exchange data.
    Respects session-level remembered choice to skip re-asking."""
    # If session pref is set, skip dialog and act immediately
    if _ce_session_pref["choice"] is not None:
        ch = _ce_session_pref["choice"]
        if ch == "compound":
            _show_compound_form(root, item_name, item_comment, item_id, on_compound)
        elif ch == "exchange":
            _show_exchange_form(root, item_name, item_comment, item_id, on_exchange)
        else:
            on_skip()
        return

    win = tk.Toplevel(root)
    win.title("Generate Compound / Exchange?")
    win.configure(bg=BG)
    win.grab_set()
    win.resizable(False, False)

    tk.Label(win, text="Add Compound or Exchange entry?",
             bg=BG, fg=FG, font=("Consolas",12,"bold")).pack(pady=(16,4), padx=20)
    tk.Label(win, text=f"Item: {item_name or item_id}",
             bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(padx=20)

    choice_var = tk.StringVar(value="none")
    bf = tk.Frame(win, bg=BG); bf.pack(pady=10)
    for lbl, val in [("Compound","compound"),("Exchange","exchange"),("Neither — skip","none")]:
        tk.Radiobutton(bf, text=lbl, variable=choice_var, value=val,
                       bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                       font=("Consolas",10)).pack(anchor="w", padx=10, pady=2)

    rem_var = tk.BooleanVar(value=False)
    tk.Checkbutton(win, text="Remember this choice for the rest of this session",
                   variable=rem_var, bg=BG, fg=ACC4,
                   selectcolor=BG3, activebackground=BG,
                   font=("Consolas",9)).pack(padx=20, pady=(0,8))

    def _proceed():
        ch = choice_var.get()
        if rem_var.get():
            _ce_session_pref["choice"] = ch
        win.destroy()
        if ch == "compound":
            _show_compound_form(root, item_name, item_comment, item_id, on_compound)
        elif ch == "exchange":
            _show_exchange_form(root, item_name, item_comment, item_id, on_exchange)
        else:
            on_skip()

    # Reset session pref button
    def _reset_pref():
        _ce_session_pref["choice"] = None
        messagebox.showinfo("Reset", "Session preference cleared — you will be asked each time.")

    bf2 = tk.Frame(win, bg=BG); bf2.pack()
    mk_btn(bf2, "Continue ▶", _proceed, color=GREEN, fg=BG2,
           font=("Consolas",10,"bold")).pack(side="left", padx=8, pady=(0,14))
    mk_btn(bf2, "Reset Remembered Choice", _reset_pref, color=BG4,
           font=("Consolas",8)).pack(side="left", padx=4, pady=(0,14))
    win.wait_window()

def _make_id_row(parent, label, var, tip=""):
    r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
    lw = tk.Label(r, text=label, width=20, anchor="w", bg=BG, fg=FG, font=("Consolas",9))
    lw.pack(side="left")
    ent = tk.Entry(r, textvariable=var, width=14, bg=BG3, fg=FG,
                   insertbackground=FG, font=("Consolas",9), relief="flat")
    ent.pack(side="left", padx=4)
    if tip: _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
    return ent

# ══════════════════════════════════════════════════════════════════════════════
# CSV / Excel batch importer for Compound and Exchange
# ══════════════════════════════════════════════════════════════════════════════

# ── Column aliases ─────────────────────────────────────────────────────────
_COMPOUND_COL_ALIASES = {
    "compoundid":    "compound_id",
    "compound_id":   "compound_id",
    "id":            "compound_id",
    "name":          "name",
    "n":             "name",
    "comment":       "comment",
    "reslv":         "res_lv",
    "res_lv":        "res_lv",
    "requiredlevel": "res_lv",
    "resid1":        "res_id1",  "res_id1": "res_id1",
    "resid2":        "res_id2",  "res_id2": "res_id2",
    "resid3":        "res_id3",  "res_id3": "res_id3",
    "reqid1":        "req_id1",  "req_id1": "req_id1",
    "reqid2":        "req_id2",  "req_id2": "req_id2",
    "reqid3":        "req_id3",  "req_id3": "req_id3",
    "reqid4":        "req_id4",  "req_id4": "req_id4",
    "reqid5":        "req_id5",  "req_id5": "req_id5",
    "reqnum1":       "req_num1", "req_num1": "req_num1",
    "reqnum2":       "req_num2", "req_num2": "req_num2",
    "reqnum3":       "req_num3", "req_num3": "req_num3",
    "reqnum4":       "req_num4", "req_num4": "req_num4",
    "reqnum5":       "req_num5", "req_num5": "req_num5",
    "probability":   "probability",
    "fee":           "fee",
    "wasteitem":     "waste_item", "waste_item": "waste_item",
}

_EXCHANGE_COL_ALIASES = {
    "exchangeid":    "exchange_id",
    "exchange_id":   "exchange_id",
    "id":            "exchange_id",
    "name":          "name",
    "n":             "name",
    "comment":       "comment",
    "reslv":         "res_lv",
    "res_lv":        "res_lv",
    "requiredlevel": "res_lv",
    "resid1":        "res_id1",  "res_id1": "res_id1",
    "resid2":        "res_id2",  "res_id2": "res_id2",
    "resid3":        "res_id3",  "res_id3": "res_id3",
    "reqid1":        "req_id1",  "req_id1": "req_id1",
    "reqid2":        "req_id2",  "req_id2": "req_id2",
    "reqid3":        "req_id3",  "req_id3": "req_id3",
    "reqid4":        "req_id4",  "req_id4": "req_id4",
    "reqid5":        "req_id5",  "req_id5": "req_id5",
    "reqnum1":       "req_num1", "req_num1": "req_num1",
    "reqnum2":       "req_num2", "req_num2": "req_num2",
    "reqnum3":       "req_num3", "req_num3": "req_num3",
    "reqnum4":       "req_num4", "req_num4": "req_num4",
    "reqnum5":       "req_num5", "req_num5": "req_num5",
    "fee":           "fee",
}

# ── Compound defaults ──────────────────────────────────────────────────────
_COMPOUND_DEFAULTS = {
    "compound_id": "0", "name": "", "comment": "", "res_lv": "1",
    "res_id1": "0", "res_id2": "0", "res_id3": "0",
    "req_id1": "0", "req_id2": "0", "req_id3": "0", "req_id4": "0", "req_id5": "0",
    "req_num1": "0", "req_num2": "0", "req_num3": "0", "req_num4": "0", "req_num5": "0",
    "probability": "50", "fee": "1", "waste_item": "12000",
}

# ── Exchange defaults ──────────────────────────────────────────────────────
_EXCHANGE_DEFAULTS = {
    "exchange_id": "0", "name": "", "comment": "", "res_lv": "1",
    "res_id1": "0", "res_id2": "0", "res_id3": "0",
    "req_id1": "0", "req_id2": "0", "req_id3": "0", "req_id4": "0", "req_id5": "0",
    "req_num1": "0", "req_num2": "0", "req_num3": "0", "req_num4": "0", "req_num5": "0",
    "fee": "0",
}

def _normalise_col(name):
    """Strip, lowercase, remove spaces/underscores for fuzzy matching."""
    return re.sub(r'[\s_]+', '', str(name).lower().strip())


def _read_csv_rows(filepath):
    """Return (headers, list-of-dicts) from a CSV file."""
    with open(filepath, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = list(reader.fieldnames or [])
    return headers, rows


def _read_xlsx_rows(filepath, sheet_index=0):
    """Return (headers, list-of-dicts) from an xlsx file (first sheet by default)."""
    if _openpyxl is None:
        raise ImportError("openpyxl is not installed. Run: pip install openpyxl")
    wb = _openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.worksheets[sheet_index]
    rows_raw = list(ws.values)
    if not rows_raw:
        return [], []
    headers = [str(h) if h is not None else "" for h in rows_raw[0]]
    rows = []
    for row in rows_raw[1:]:
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # skip fully blank rows
        rows.append({headers[i]: (str(row[i]) if row[i] is not None else "")
                     for i in range(len(headers))})
    return headers, rows


def _map_row_to_cfg(raw_row, alias_map, defaults):
    """Map a raw CSV/xlsx row dict -> cfg dict using alias_map, filling defaults."""
    cfg = dict(defaults)
    for col, val in raw_row.items():
        key = alias_map.get(_normalise_col(col))  # normalise HEADER only
        if key and str(val).strip() != "":
            cfg[key] = str(val).strip()  # value kept verbatim — paths/dots/slashes preserved
    return cfg


def _import_preview_window(root, mode, filepath, on_confirm):
    """
    Show a scrollable preview of parsed rows before committing.
    mode = "compound" or "exchange"
    on_confirm(cfgs) called with list of cfg dicts.
    """
    alias_map = _COMPOUND_COL_ALIASES if mode == "compound" else _EXCHANGE_COL_ALIASES
    defaults  = _COMPOUND_DEFAULTS    if mode == "compound" else _EXCHANGE_DEFAULTS
    id_key    = "compound_id"         if mode == "compound" else "exchange_id"
    color     = ACC1                  if mode == "compound" else ACC2
    title_lbl = "Compound_Potion.xml" if mode == "compound" else "ExchangeShopContents.xml"

    # ── Parse ──────────────────────────────────────────────────────────────
    try:
        ext = os.path.splitext(filepath)[1].lower()
        if ext in (".xlsx", ".xlsm", ".xls"):
            headers, raw_rows = _read_xlsx_rows(filepath)
        else:
            headers, raw_rows = _read_csv_rows(filepath)
    except Exception as e:
        messagebox.showerror("Import Error", f"Could not read file:\n{e}", parent=root)
        return

    if not raw_rows:
        messagebox.showwarning("Empty File", "No data rows found in the file.", parent=root)
        return

    cfgs = [_map_row_to_cfg(r, alias_map, defaults) for r in raw_rows]

    # ── Detect unmapped columns ────────────────────────────────────────────
    unmapped = []
    for col in headers:
        if _normalise_col(col) not in alias_map:
            unmapped.append(col)

    # ── Preview window ─────────────────────────────────────────────────────
    win = tk.Toplevel(root)
    win.title(f"Import Preview — {mode.capitalize()} ({len(cfgs)} rows)")
    win.configure(bg=BG)
    win.geometry("960x620")
    win.grab_set()

    # Header
    hdr = tk.Frame(win, bg=BG2); hdr.pack(fill="x")
    tk.Label(hdr, text=f"📥  Import {mode.capitalize()} — {title_lbl}",
             bg=BG2, fg=color, font=("Consolas", 12, "bold"),
             pady=8).pack(side="left", padx=14)
    tk.Label(hdr, text=f"{len(cfgs)} rows from {os.path.basename(filepath)}",
             bg=BG2, fg=FG_DIM, font=("Consolas", 9)).pack(side="left", padx=6)

    # Unmapped warning
    if unmapped:
        wf = tk.Frame(win, bg="#45475a"); wf.pack(fill="x", padx=8, pady=(4, 0))
        tk.Label(wf, text=f"⚠  Columns not recognised (will be ignored): {', '.join(unmapped)}",
                 bg="#45475a", fg=ACC4, font=("Consolas", 8),
                 pady=4).pack(anchor="w", padx=10)

    # Scrollable table
    tbl_host = tk.Frame(win, bg=BG); tbl_host.pack(fill="both", expand=True, padx=8, pady=6)

    x_scroll = tk.Scrollbar(tbl_host, orient="horizontal")
    y_scroll = tk.Scrollbar(tbl_host, orient="vertical")
    tbl_canvas = tk.Canvas(tbl_host, bg=BG, bd=0, highlightthickness=0,
                           xscrollcommand=x_scroll.set,
                           yscrollcommand=y_scroll.set)
    x_scroll.config(command=tbl_canvas.xview)
    y_scroll.config(command=tbl_canvas.yview)
    x_scroll.pack(side="bottom", fill="x")
    y_scroll.pack(side="right",  fill="y")
    tbl_canvas.pack(side="left", fill="both", expand=True)

    tbl_inner = tk.Frame(tbl_canvas, bg=BG)
    tbl_canvas.create_window((0, 0), window=tbl_inner, anchor="nw")
    tbl_inner.bind("<Configure>",
        lambda e: tbl_canvas.configure(scrollregion=tbl_canvas.bbox("all")))

    # Columns to show
    if mode == "compound":
        show_cols = [id_key, "name", "comment", "res_lv",
                     "res_id1", "res_id2", "res_id3",
                     "req_id1", "req_num1", "req_id2", "req_num2",
                     "req_id3", "req_num3", "req_id4", "req_num4", "req_id5", "req_num5",
                     "probability", "fee", "waste_item"]
    else:
        show_cols = [id_key, "name", "comment", "res_lv",
                     "res_id1", "res_id2", "res_id3",
                     "req_id1", "req_num1", "req_id2", "req_num2",
                     "req_id3", "req_num3", "req_id4", "req_num4", "req_id5", "req_num5",
                     "fee"]

    col_w = 10
    # Header row
    for ci, col in enumerate(show_cols):
        tk.Label(tbl_inner, text=col, bg=BG2, fg=BLUE,
                 font=("Consolas", 8, "bold"), width=col_w, anchor="w",
                 relief="flat", padx=3).grid(row=0, column=ci, sticky="w", padx=1, pady=1)

    # Data rows — editable
    row_vars = []  # list of dicts: col -> StringVar
    for ri, cfg in enumerate(cfgs):
        rv = {}
        bg_row = BG if ri % 2 == 0 else BG2
        for ci, col in enumerate(show_cols):
            v = tk.StringVar(value=cfg.get(col, ""))
            rv[col] = v
            ent = tk.Entry(tbl_inner, textvariable=v, width=col_w,
                           bg=bg_row, fg=FG, insertbackground=FG,
                           font=("Consolas", 8), relief="flat", bd=0)
            ent.grid(row=ri + 1, column=ci, sticky="w", padx=1, pady=1)
        row_vars.append(rv)

    # Footer
    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")
    tk.Label(nav, text="You can edit cells above before confirming.",
             bg=BG2, fg=FG_DIM, font=("Consolas", 8)).pack(side="left", padx=14, pady=6)

    def _confirm():
        # Read back edited values
        final_cfgs = []
        for rv in row_vars:
            cfg_out = dict(defaults)  # start from defaults
            for col, v in rv.items():
                cfg_out[col] = v.get()
            final_cfgs.append(cfg_out)
        win.destroy()
        on_confirm(final_cfgs)

    mk_btn(nav, f"✓  Import {len(cfgs)} rows", _confirm,
           color=GREEN, fg=BG2, font=("Consolas", 10, "bold")).pack(side="right", padx=14, pady=6)
    mk_btn(nav, "Cancel", win.destroy, color=BG4).pack(side="right", padx=4, pady=6)
    win.wait_window()


def _ask_import_mode_then_file(root, on_compound_cfgs, on_exchange_cfgs):
    """
    Step 1: ask Compound or Exchange.
    Step 2: file picker.
    Step 3: preview + confirm.
    on_compound_cfgs(cfgs) / on_exchange_cfgs(cfgs) called on confirm.
    """
    win = tk.Toplevel(root)
    win.title("Import — Choose Mode")
    win.configure(bg=BG)
    win.resizable(False, False)
    win.grab_set()

    tk.Label(win, text="📥  Import from CSV / Excel",
             bg=BG, fg=BLUE, font=("Consolas", 13, "bold"),
             pady=12).pack(padx=28)
    tk.Label(win, text="What type of data are you importing?",
             bg=BG, fg=FG, font=("Consolas", 10)).pack(padx=28, pady=(0, 16))

    btn_frm = tk.Frame(win, bg=BG); btn_frm.pack(padx=28, pady=4)

    def _pick_file(mode):
        win.destroy()
        path = filedialog.askopenfilename(
            title=f"Import {mode.capitalize()} data",
            filetypes=[
                ("Spreadsheet / CSV", "*.csv *.xlsx *.xlsm *.xls"),
                ("CSV files",  "*.csv"),
                ("Excel files","*.xlsx *.xlsm *.xls"),
                ("All files",  "*.*"),
            ],
            parent=root,
        )
        if not path:
            return
        if mode == "compound":
            _import_preview_window(root, "compound", path, on_compound_cfgs)
        else:
            _import_preview_window(root, "exchange", path, on_exchange_cfgs)

    mk_btn(btn_frm, "⚗  Compound_Potion.xml", lambda: _pick_file("compound"),
           color=ACC1, fg=BG2, font=("Consolas", 11, "bold"),
           padx=18, pady=10).grid(row=0, column=0, padx=10, pady=6)
    mk_btn(btn_frm, "🔄  ExchangeShopContents.xml", lambda: _pick_file("exchange"),
           color=ACC2, fg=BG2, font=("Consolas", 11, "bold"),
           padx=18, pady=10).grid(row=0, column=1, padx=10, pady=6)

    tk.Label(win,
             text="CSV column names are matched automatically.\n"
                  "Click 📖 Template below to download a blank template.",
             bg=BG, fg=FG_GREY, font=("Consolas", 8),
             justify="center").pack(pady=(10, 4))

    def _save_template(mode):
        if mode == "compound":
            cols = list(_COMPOUND_DEFAULTS.keys())
        else:
            cols = list(_EXCHANGE_DEFAULTS.keys())
        path = filedialog.asksaveasfilename(
            title=f"Save {mode} template",
            defaultextension=".csv",
            filetypes=[("CSV","*.csv"),("Excel","*.xlsx")],
            initialfile=f"{mode}_template.csv",
            parent=root,
        )
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(cols)
            writer.writerow(["" for _ in cols])  # blank example row
        messagebox.showinfo("Template saved", f"Saved to:\n{path}", parent=root)

    tpl_frm = tk.Frame(win, bg=BG); tpl_frm.pack(pady=(0, 14))
    mk_btn(tpl_frm, "📄  Compound template", lambda: _save_template("compound"),
           color=BG4, font=("Consolas", 9)).pack(side="left", padx=8)
    mk_btn(tpl_frm, "📄  Exchange template", lambda: _save_template("exchange"),
           color=BG4, font=("Consolas", 9)).pack(side="left", padx=8)
    mk_btn(win, "Cancel", win.destroy, color=BG4).pack(pady=(0, 12))
    win.wait_window()


def _show_compound_form(root, item_name, item_comment, item_id, on_done):
    """Full compound entry dialog."""
    last_id = _get_last_id("compound", 100)
    win = tk.Toplevel(root)
    win.title("Compound Entry")
    win.configure(bg=BG)
    win.grab_set()

    sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True)
    canv, C = mk_scroll_canvas(sh, init_width=540)
    win.geometry("580x620")

    tk.Label(C, text="Compound_Potion.xml Entry",
             bg=BG, fg=ACC1, font=("Consolas",12,"bold")).pack(pady=(12,2), padx=12, anchor="w")

    def sec(t):
        f = tk.LabelFrame(C, text=t, bg=BG, fg=BLUE,
                          font=("Consolas",9,"bold"), bd=1, relief="groove")
        f.pack(fill="x", padx=10, pady=4)
        return f

    s1 = sec("  Identity  ")
    v_cid  = tk.StringVar(value=str(last_id + 1))
    v_name = tk.StringVar(value=item_name or "")
    v_cmt  = tk.StringVar(value=item_comment or "")
    v_reslv= tk.StringVar(value="1")
    _make_id_row(s1, "CompoundID:", v_cid, "Unique ID for this compound recipe.")
    _make_id_row(s1, "Name:", v_name)
    _make_id_row(s1, "Comment:", v_cmt)
    _make_id_row(s1, "Required Level:", v_reslv, "ResLv — minimum level to use this compound.")

    s2 = sec("  Receiving IDs (ResID 1-3)  ")
    tk.Label(s2, text="Item IDs produced by this recipe (0 = unused).",
             bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", padx=8)
    res_vars = []
    for n in range(1,4):
        v = tk.StringVar(value=item_id if n==1 else "0")
        _make_id_row(s2, f"ResID{n}:", v, f"Receiving item ID #{n}")
        res_vars.append(v)

    s3 = sec("  Required Items (ReqID/ReqNum 1-5)  ")
    tk.Label(s3, text="Items required to perform this compound (0 = unused).",
             bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", padx=8)
    req_id_vars = []; req_num_vars = []
    for n in range(1,6):
        r = tk.Frame(s3, bg=BG); r.pack(fill="x", padx=8, pady=1)
        tk.Label(r, text=f"ReqID{n}:", width=8, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left")
        vid = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vid, width=12, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        tk.Label(r, text=f"ReqNum{n}:", width=10, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left", padx=(8,0))
        vnum = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vnum, width=6, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        req_id_vars.append(vid); req_num_vars.append(vnum)

    s4 = sec("  Recipe Settings  ")
    v_prob  = tk.StringVar(value="50")
    v_fee   = tk.StringVar(value="1")
    v_waste = tk.StringVar(value="12000")
    _make_id_row(s4, "Probability:", v_prob, "Success chance (0-100).")
    _make_id_row(s4, "Fee:", v_fee, "Galder fee to attempt compound.")
    _make_id_row(s4, "WasteItem:", v_waste, "Item consumed on failure (default 12000).")

    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")

    def _gen():
        cfg = {
            "compound_id": v_cid.get(), "name": v_name.get(), "comment": v_cmt.get(),
            "res_lv": v_reslv.get(),
            "res_id1": res_vars[0].get(), "res_id2": res_vars[1].get(), "res_id3": res_vars[2].get(),
            "probability": v_prob.get(), "fee": v_fee.get(), "waste_item": v_waste.get(),
        }
        for n in range(1,6):
            cfg[f"req_id{n}"]  = req_id_vars[n-1].get()
            cfg[f"req_num{n}"] = req_num_vars[n-1].get()
        try: _set_last_id("compound", int(v_cid.get()))
        except: pass
        win.destroy()
        on_done(cfg)

    mk_btn(nav, "✓  Generate", _gen, color=GREEN, fg=BG2,
           font=("Consolas",10,"bold")).pack(side="left", padx=12, pady=6)
    mk_btn(nav, "Cancel", win.destroy).pack(side="left", padx=4, pady=6)
    win.wait_window()

def _show_exchange_form(root, item_name, item_comment, item_id, on_done):
    """Full exchange entry dialog."""
    last_id = _get_last_id("exchange", 0)
    win = tk.Toplevel(root)
    win.title("Exchange Entry")
    win.configure(bg=BG)
    win.grab_set()

    sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True)
    canv, C = mk_scroll_canvas(sh, init_width=540)
    win.geometry("580x580")

    tk.Label(C, text="ExchangeShopContents.xml Entry",
             bg=BG, fg=ACC2, font=("Consolas",12,"bold")).pack(pady=(12,2), padx=12, anchor="w")

    def sec(t):
        f = tk.LabelFrame(C, text=t, bg=BG, fg=BLUE,
                          font=("Consolas",9,"bold"), bd=1, relief="groove")
        f.pack(fill="x", padx=10, pady=4)
        return f

    s1 = sec("  Identity  ")
    v_eid  = tk.StringVar(value=str(last_id + 1))
    v_name = tk.StringVar(value=item_name or "")
    v_cmt  = tk.StringVar(value=item_comment or "")
    v_reslv= tk.StringVar(value="1")
    _make_id_row(s1, "ExchangeID:", v_eid, "Unique ID for this exchange entry.")
    _make_id_row(s1, "Name:", v_name)
    _make_id_row(s1, "Comment:", v_cmt)
    _make_id_row(s1, "Required Level:", v_reslv, "ResLv — minimum level.")

    s2 = sec("  Receiving IDs (ResID 1-3)  ")
    tk.Label(s2, text="Items received from exchange (0 = unused).",
             bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w", padx=8)
    res_vars = []
    for n in range(1,4):
        v = tk.StringVar(value=item_id if n==1 else "0")
        _make_id_row(s2, f"ResID{n}:", v)
        res_vars.append(v)

    s3 = sec("  Required Items (ReqID/ReqNum 1-5)  ")
    req_id_vars = []; req_num_vars = []
    for n in range(1,6):
        r = tk.Frame(s3, bg=BG); r.pack(fill="x", padx=8, pady=1)
        tk.Label(r, text=f"ReqID{n}:", width=8, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left")
        vid = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vid, width=12, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        tk.Label(r, text=f"ReqNum{n}:", width=10, anchor="w", bg=BG, fg=FG,
                 font=("Consolas",9)).pack(side="left", padx=(8,0))
        vnum = tk.StringVar(value="0")
        tk.Entry(r, textvariable=vnum, width=6, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=4)
        req_id_vars.append(vid); req_num_vars.append(vnum)

    s4 = sec("  Settings  ")
    v_fee = tk.StringVar(value="0")
    _make_id_row(s4, "Fee:", v_fee, "Galder fee for exchange.")

    nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")

    def _gen():
        cfg = {
            "exchange_id": v_eid.get(), "name": v_name.get(), "comment": v_cmt.get(),
            "res_lv": v_reslv.get(),
            "res_id1": res_vars[0].get(), "res_id2": res_vars[1].get(), "res_id3": res_vars[2].get(),
            "fee": v_fee.get(),
        }
        for n in range(1,6):
            cfg[f"req_id{n}"]  = req_id_vars[n-1].get()
            cfg[f"req_num{n}"] = req_num_vars[n-1].get()
        try: _set_last_id("exchange", int(v_eid.get()))
        except: pass
        win.destroy()
        on_done(cfg)

    mk_btn(nav, "✓  Generate", _gen, color=GREEN, fg=BG2,
           font=("Consolas",10,"bold")).pack(side="left", padx=12, pady=6)
    mk_btn(nav, "Cancel", win.destroy).pack(side="left", padx=4, pady=6)
    win.wait_window()


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 7 — Compound / Exchange Generator (standalone)
# ══════════════════════════════════════════════════════════════════════════════

ACC7 = "#f5c2e7"   # pink — compound/exchange tool

# ── Column guide tooltips ──────────────────────────────────────────────────
_CE_COMPOUND_COL_GUIDE = (
    "Compound spreadsheet — recognised columns (any order):\n"
    "  CompoundID / ID   — unique recipe ID (required)\n"
    "  Name              — recipe name\n"
    "  Comment           — description\n"
    "  ResLv / RequiredLevel — minimum level to use\n"
    "  ResID1 / ResID2 / ResID3  — receiving item IDs\n"
    "  ReqID1..ReqID5    — required ingredient item IDs\n"
    "  ReqNum1..ReqNum5  — quantity of each ingredient\n"
    "  Probability       — success chance (default 50)\n"
    "  Fee               — Galder fee (default 1)\n"
    "  WasteItem         — item consumed on fail (default 12000)\n"
    "Group column: any column whose header is NOT one of the above\n"
    "  becomes a Group name — rows under it form one batch."
)

_CE_EXCHANGE_COL_GUIDE = (
    "Exchange spreadsheet — recognised columns (any order):\n"
    "  ExchangeID / ID   — unique exchange ID (required)\n"
    "  Name              — entry name\n"
    "  Comment           — description\n"
    "  ResLv / RequiredLevel — minimum level\n"
    "  ResID1 / ResID2 / ResID3  — receiving item IDs\n"
    "  ReqID1..ReqID5    — required item IDs\n"
    "  ReqNum1..ReqNum5  — quantity of each required item\n"
    "  Fee               — exchange fee (default 0)\n"
    "Group column: any column whose header is NOT one of the above\n"
    "  becomes a Group name — rows under it form one batch."
)


def _parse_ce_grouped_csv(text, mode):
    """Parse a grouped compound/exchange CSV — identical logic to parse_grouped_csv.
    mode = 'compound' or 'exchange'.
    Returns list of {group_name, rows:[cfg_dict], mode}.
    """
    alias_map = _COMPOUND_COL_ALIASES if mode == "compound" else _EXCHANGE_COL_ALIASES
    defaults  = _COMPOUND_DEFAULTS    if mode == "compound" else _EXCHANGE_DEFAULTS
    id_key    = "compound_id"         if mode == "compound" else "exchange_id"

    reader = csv.reader(io.StringIO(text))
    rows   = list(reader)
    if not rows: return []
    headers  = [h.strip() for h in rows[0]]
    data_rows = rows[1:]

    def _norm(h): return re.sub(r'[\s_]+', '', str(h).lower().strip())

    known = set(alias_map.keys())
    # Group columns: not a known field column
    group_col_indices = [i for i, h in enumerate(headers)
                         if h.strip() and _norm(h) not in known]
    if not group_col_indices:
        # No group column — treat whole sheet as one group named "Default"
        cfgs = []
        for raw_row in data_rows:
            d = {h: (raw_row[i].strip() if i < len(raw_row) else "")
                 for i, h in enumerate(headers)}
            cfg = _map_row_to_cfg(d, alias_map, defaults)
            if cfg.get(id_key, "0").strip() not in ("", "0"):
                cfgs.append(cfg)
        return [{"group_name": "Default", "rows": cfgs, "mode": mode}] if cfgs else []

    results = []
    next_group_cols = group_col_indices[1:] + [len(headers)]
    prev_end = -1

    for bi, gc in enumerate(group_col_indices):
        group_name = headers[gc]
        span_indices = list(range(prev_end + 1, gc + 1)) + \
                       [ci for ci in range(gc + 1, next_group_cols[bi])
                        if _norm(headers[ci]) in known]
        prev_end = gc

        cfgs = []
        for raw_row in data_rows:
            # Name cell = group col
            name_val = raw_row[gc].strip() if gc < len(raw_row) else ""
            if not name_val: continue
            d = {}
            for ci in span_indices:
                h = headers[ci]
                d[h] = raw_row[ci].strip() if ci < len(raw_row) else ""
            cfg = _map_row_to_cfg(d, alias_map, defaults)
            # If no name in cfg but we have a group col value, use it as name
            if not cfg.get("name") and name_val:
                cfg["name"] = name_val
            cfgs.append(cfg)

        if cfgs:
            results.append({"group_name": group_name, "rows": cfgs, "mode": mode})

    return results


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 7 — Compound / Exchange / Shop Listing Generator
# Manual entry now runs inside the main window (no popup for the form itself)
# ══════════════════════════════════════════════════════════════════════════════

def build_shop_row(item_id, count="100", price="0.000000", unk="0"):
    """Generate a single R_ShopItem.xml data line: id# count price unk
    Ensures 6 decimal places on count and price regardless of input format."""
    def _fmt(v):
        s = str(v).strip()
        # Remove any existing .000000 to avoid double decimal
        if "." in s:
            try: return f"{float(s):.6f}"
            except: return s
        return f"{s}.000000"
    return f"{item_id}# {_fmt(count)} {_fmt(price)} {unk}"


def _build_box_myshop_outputs(box_configs):
    """
    Given a list of box_configs (from Tool1), generate:
      - libcmgds_e XML GOODS blocks  (list of strings)
      - SQL rows: tbl_goods + tbl_goods_list + tbl_goods_limit (list of strings)
    Only processes configs where myshop_enabled == True.
    Each box gets its own unique goods_list_code that does not repeat across boxes.
    Returns (xml_blocks, sql_rows) — both may be empty lists.
    """
    today = datetime.date.today().strftime("%Y-%m-%d")
    xml_blocks = []
    sql_goods   = []
    sql_list    = []
    sql_limit   = []

    _last_glc_seen = None   # tracks the highest glc used for persistence

    for cfg in box_configs:
        if not cfg.get("myshop_enabled", False):
            continue

        box_id     = cfg.get("id", "0")
        box_name   = cfg.get("name", "")
        box_desc   = cfg.get("comment", "")
        price      = cfg.get("myshop_price", "0")
        cat        = cfg.get("myshop_cat", "special")  # "special" | "new"
        new_stamp  = int(cfg.get("myshop_new", 0))
        pop_stamp  = int(cfg.get("myshop_pop", 0))
        level      = cfg.get("myshop_level", "0")
        version    = cfg.get("myshop_ver",   "1")
        item_count = cfg.get("myshop_item_count", "1")
        limit_use  = "2"   # goods_limit_use
        created    = today

        # Each box carries its own confirmed glc value — use it directly.
        try: glc = int(cfg.get("myshop_glc_start", 21000))
        except: glc = 21000
        _last_glc_seen = glc

        # goods_category1: 113 = Special Box, 101 = New Item
        goods_cat1_xml = "113" if cat == "special" else "101"
        sql_cat2        = "2"  if cat == "special" else "0"

        # ── libcmgds_e XML block ──────────────────────────────────────────
        # goods_list_code / parents_list_code = version (unique per box, user-set)
        xml = (
            f'\t\t<GOODS goods_code="{box_id}" goods_name="{box_name}" '
            f'goods_desc="{box_desc}" goods_set_count="1" goods_limit_use="{limit_use}" '
            f'goods_limit_time="0" goods_cash_price="{price}" '
            f'goods_shop_new="{new_stamp}" goods_shop_popular="{pop_stamp}" '
            f'goods_category="1" goods_category0="11" goods_category1="{goods_cat1_xml}" '
            f'goods_category2="0" goods_limit_desc="All Characters" '
            f'goods_char_level="{level}" goods_char_sex="0" goods_char_type="15" '
            f'goods_issell="0" goods_created="{created.replace("-","")}" '
            f'goods_filtercode1="0" goods_filtercode2="0" goods_filtercode3="0" '
            f'goods_filtercode4="0" goods_filterlevel="0" '
            f'discount_start_date="1900-01-01 00:00:00" '
            f'discount_end_date="1900-01-01 00:00:00" discount_display_date="">\n'
            f'\t\t\t<GOODS_LIST item_index="{box_id}" goods_name="{box_name}" '
            f'item_count="{item_count}" item_class="1" preview_x="" preview_y="" preview_z="" '
            f'preview_d="" goods_list_code="{glc}" '
            f'parents_list_code="{glc}" />\n'
            f'\t\t</GOODS>'
        )
        xml_blocks.append(xml)

        # ── tbl_goods INSERT ──────────────────────────────────────────────
        sql_goods.append(
            f"INSERT INTO gmg_account.dbo.tbl_goods ("
            f"goods_code, goods_name, goods_desc, goods_capacity, goods_category, "
            f"goods_set_count, goods_item_index, goods_item_count, "
            f"goods_limit_use, goods_limit_time, goods_cash_price, goods_created, "
            f"goods_shop_new, goods_shop_popular, goods_sellcount, "
            f"goods_category0, goods_category1, goods_category2, "
            f"goods_limit_desc, goods_char_level, goods_char_sex, goods_char_type, "
            f"version_code, goods_issell, goods_image"
            f") VALUES ("
            f"{box_id}, N'{box_name}', N'{box_desc}', NULL, 0, "
            f"1, NULL, NULL, "
            f"{limit_use}, NULL, {price}, '{created} 00:00:00', "
            f"{new_stamp}, {pop_stamp}, 0, "
            f"11, 3, {sql_cat2}, "
            f"N'All Characters', {level}, 0, 15, "
            f"{version}, 0, ''"
            f");"
        )

        # ── tbl_goods_list INSERT ─────────────────────────────────────────
        # goods_list_code and parents_list_code = version (unique per box listing)
        # item_class = 1 for boxes
        sql_list.append(
            f"INSERT INTO gmg_account.dbo.tbl_goods_list ("
            f"goods_code, item_index, item_count, goods_scode, item_class, "
            f"preview_x, preview_y, preview_z, preview_d, "
            f"goods_list_code, parents_list_code, goods_list_limit"
            f") VALUES ("
            f"{box_id}, {box_id}, {item_count}, {box_id}, 1, "
            f"NULL, NULL, NULL, NULL, "
            f"{glc}, {glc}, 0"
            f");"
        )

        # ── tbl_goods_limit INSERT ────────────────────────────────────────
        # limit_code = goods_limit_use value
        # goods_limit_price = cash price
        # default_display = True (no quotes)
        sql_limit.append(
            f"INSERT INTO gmg_account.dbo.tbl_goods_limit ("
            f"goods_code, limit_code, goods_limit_price, default_display"
            f") VALUES ("
            f"{box_id}, {limit_use}, {price}, True"
            f");"
        )

    # Combine all SQL sections with headers
    if not sql_goods:
        return xml_blocks, [], (_last_glc_seen or 0)

    all_sql = (
        ["-- tbl_goods"]           + sql_goods  + [""] +
        ["-- tbl_goods_list"]      + sql_list   + [""] +
        ["-- tbl_goods_limit"]     + sql_limit
    )
    return xml_blocks, all_sql, (_last_glc_seen or 0)


class Tool7(tk.Frame):
    """Compound / Exchange / Shop Listing Generator."""

    ACC = ACC7

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root    = root
        self.session = session
        self._compound_rows = []   # list of (potion_xml, location_xml)
        self._exchange_rows = []   # list of (contents_xml, location_xml)
        self._shop_rows     = []   # list of "id# 100.000000 0.000000 0" strings
        self._build_start_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    # ─────────────────────────────────────────────────────────────────────────
    # START SCREEN
    # ─────────────────────────────────────────────────────────────────────────
    def _build_start_screen(self):
        self._clear()
        sh = tk.Frame(self, bg=BG); sh.pack(fill="both", expand=True)
        canv, C = mk_scroll_canvas(sh)

        tk.Label(C, text="COMPOUND / EXCHANGE / SHOP GENERATOR",
                 font=("Consolas", 16, "bold"), bg=BG, fg=self.ACC).pack(pady=(28, 4))
        tk.Label(C,
                 text="Generate Compound_Potion.xml · Compounder_Spot.xml\n"
                      "ExchangeShopContents.xml · Exchange_Location.xml · R_ShopItem.xml",
                 bg=BG, fg=FG_DIM, font=("Consolas", 9), justify="center").pack(pady=(0, 6))

        # Column guide
        info = tk.Frame(C, bg=BG2); info.pack(pady=4, padx=20, fill="x")
        tk.Label(info,
            text=(
                "  CSV / Excel columns:\n"
                "  Compound: CompoundID, Name, Comment, ResLv, ResID1-3, ReqID1-5, ReqNum1-5,\n"
                "            Probability, Fee, WasteItem\n"
                "  Exchange: ExchangeID, Name, Comment, ResLv, ResID1-3, ReqID1-5, ReqNum1-5, Fee\n"
                "  Shop:     ItemID (or ID), Count, Price, Unk  — or just ItemID for defaults\n"
                "  Group headers: any non-field column becomes a batch group name.\n"
                "  Excel: multiple sheets are each treated as a separate group."
            ),
            bg=BG2, fg=FG, font=("Consolas", 8), justify="left", padx=10, pady=8).pack(anchor="w")

        # Status badge
        n_c, n_e, n_s = len(self._compound_rows), len(self._exchange_rows), len(self._shop_rows)
        if n_c or n_e or n_s:
            tk.Label(C,
                     text=f"Pending: {n_c} compound  ·  {n_e} exchange  ·  {n_s} shop rows",
                     bg=BG, fg=GREEN, font=("Consolas", 9, "italic")).pack(pady=4)

        # Main action buttons
        bf = tk.Frame(C, bg=BG); bf.pack(pady=12)

        mk_btn(bf, "⚗  New Compound Entry",
               lambda: self._build_entry_form("compound"),
               color=ACC1, fg=BG2, font=("Consolas", 10, "bold")).pack(side="left", padx=6)
        mk_btn(bf, "🔄  New Exchange Entry",
               lambda: self._build_entry_form("exchange"),
               color=ACC2, fg=BG2, font=("Consolas", 10, "bold")).pack(side="left", padx=6)
        mk_btn(bf, "🛒  New Shop Entry",
               lambda: self._build_entry_form("shop"),
               color=ACC5, fg=BG2, font=("Consolas", 10, "bold")).pack(side="left", padx=6)

        bf2 = tk.Frame(C, bg=BG); bf2.pack(pady=4)
        mk_btn(bf2, "📂  Import Spreadsheet", self._import_from_file,
               color=BG3).pack(side="left", padx=6)
        if self.session.compound_rows or self.session.exchange_rows:
            mk_btn(bf2, "⬇  From Session", self._import_session,
                   color=BG4).pack(side="left", padx=6)

        # Template buttons
        tf = tk.Frame(C, bg=BG); tf.pack(pady=4)
        for label, mode in [("📄 Compound CSV", "compound"), ("📄 Exchange CSV", "exchange"),
                             ("📄 Shop CSV", "shop")]:
            mk_btn(tf, label, lambda m=mode: self._save_template(m),
                   color=BG4, font=("Consolas", 8)).pack(side="left", padx=4)

        if n_c or n_e or n_s:
            mk_btn(C, "📋  View / Export pending rows", self._build_output_screen,
                   color=BG3).pack(pady=8)

    # ─────────────────────────────────────────────────────────────────────────
    # IN-WINDOW ENTRY FORMS
    # ─────────────────────────────────────────────────────────────────────────
    def _build_entry_form(self, mode):
        """Build compound, exchange, or shop entry form directly in the tool window."""
        self._clear()
        self.grid_rowconfigure(0, weight=1)

        sh = tk.Frame(self, bg=BG); sh.pack(fill="both", expand=True)
        canv, C = mk_scroll_canvas(sh)

        mode_labels = {"compound": "⚗  Compound Entry", "exchange": "🔄  Exchange Entry",
                       "shop": "🛒  Shop Listing Entry"}
        mode_colors = {"compound": ACC1, "exchange": ACC2, "shop": ACC5}
        color = mode_colors.get(mode, self.ACC)

        tk.Label(C, text=mode_labels.get(mode, mode.title()),
                 font=("Consolas", 14, "bold"), bg=BG, fg=color).pack(pady=(16, 6), padx=14, anchor="w")

        def sec(title):
            f = tk.LabelFrame(C, text=title, bg=BG, fg=BLUE,
                              font=("Consolas", 9, "bold"), bd=1, relief="groove")
            f.pack(fill="x", padx=12, pady=5)
            return f

        def row(parent, label, var, tip="", width=18):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            lw = tk.Label(r, text=label, width=20, anchor="w", bg=BG, fg=FG, font=("Consolas", 9))
            lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=width, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip and _APP_SETTINGS.get("tooltips_enabled", True):
                _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent

        # ── COMPOUND FORM ──────────────────────────────────────────────────
        if mode == "compound":
            last_id = _get_last_id("compound", 100)
            v_cid   = tk.StringVar(value=str(last_id + 1))
            v_name  = tk.StringVar()
            v_cmt   = tk.StringVar()
            v_reslv = tk.StringVar(value="1")

            s1 = sec("  Identity  ")
            row(s1, "CompoundID:", v_cid, "Unique ID for this compound recipe.")
            row(s1, "Name:", v_name)
            row(s1, "Comment:", v_cmt)
            row(s1, "Required Level (ResLv):", v_reslv, "Minimum level to use this compound.")

            s2 = sec("  Result Items (ResID 1-3)  ")
            tk.Label(s2, text="  Item IDs produced. 0 = unused.", bg=BG, fg=FG_GREY,
                     font=("Consolas", 8)).pack(anchor="w", padx=8)
            res_vars = []
            for n in range(1, 4):
                v = tk.StringVar(value="0")
                row(s2, f"ResID{n}:", v, f"Result item ID #{n}")
                res_vars.append(v)

            s3 = sec("  Required Items (ReqID / ReqNum 1-5)  ")
            tk.Label(s3, text="  Items required to compound. 0 = unused.", bg=BG, fg=FG_GREY,
                     font=("Consolas", 8)).pack(anchor="w", padx=8)
            req_id_vars = []; req_num_vars = []
            for n in range(1, 6):
                r2 = tk.Frame(s3, bg=BG); r2.pack(fill="x", padx=8, pady=2)
                tk.Label(r2, text=f"ReqID{n}:", width=10, anchor="w", bg=BG, fg=FG,
                         font=("Consolas", 9)).pack(side="left")
                vid = tk.StringVar(value="0")
                tk.Entry(r2, textvariable=vid, width=12, bg=BG3, fg=FG,
                         insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
                tk.Label(r2, text=f"ReqNum{n}:", width=11, anchor="w", bg=BG, fg=FG_DIM,
                         font=("Consolas", 9)).pack(side="left", padx=(8, 0))
                vnum = tk.StringVar(value="0")
                tk.Entry(r2, textvariable=vnum, width=7, bg=BG3, fg=FG,
                         insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
                req_id_vars.append(vid); req_num_vars.append(vnum)

            s4 = sec("  Recipe Settings  ")
            v_prob  = tk.StringVar(value="50")
            v_fee   = tk.StringVar(value="1")
            v_waste = tk.StringVar(value="12000")
            row(s4, "Probability:", v_prob, "Success chance (0-100).")
            row(s4, "Fee:", v_fee, "Galder fee to attempt compound.")
            row(s4, "WasteItem:", v_waste, "Item consumed on failure.")

            def _add():
                cfg = {
                    "compound_id": v_cid.get(), "name": v_name.get(), "comment": v_cmt.get(),
                    "res_lv": v_reslv.get(),
                    "res_id1": res_vars[0].get(), "res_id2": res_vars[1].get(),
                    "res_id3": res_vars[2].get(),
                    "probability": v_prob.get(), "fee": v_fee.get(), "waste_item": v_waste.get(),
                }
                for n in range(1, 6):
                    cfg[f"req_id{n}"]  = req_id_vars[n-1].get()
                    cfg[f"req_num{n}"] = req_num_vars[n-1].get()
                try: _set_last_id("compound", int(v_cid.get()))
                except: pass
                self._compound_rows.append((build_compound_row(cfg),
                                            build_compound_location_row(cfg["compound_id"])))
                self.session.compound_rows = list(self._compound_rows)
                if messagebox.askyesno("Added",
                        f"Compound ID {v_cid.get()} added.\nAdd another?"):
                    try: v_cid.set(str(int(v_cid.get()) + 1))
                    except: pass
                else:
                    self._build_output_screen()

        # ── EXCHANGE FORM ──────────────────────────────────────────────────
        elif mode == "exchange":
            last_id = _get_last_id("exchange", 0)
            v_eid   = tk.StringVar(value=str(last_id + 1))
            v_name  = tk.StringVar()
            v_cmt   = tk.StringVar()
            v_reslv = tk.StringVar(value="1")

            s1 = sec("  Identity  ")
            row(s1, "ExchangeID:", v_eid, "Unique ID for this exchange entry.")
            row(s1, "Name:", v_name)
            row(s1, "Comment:", v_cmt)
            row(s1, "Required Level (ResLv):", v_reslv)

            s2 = sec("  Result Items (ResID 1-3)  ")
            tk.Label(s2, text="  Item IDs given in exchange. 0 = unused.", bg=BG, fg=FG_GREY,
                     font=("Consolas", 8)).pack(anchor="w", padx=8)
            res_vars = []
            for n in range(1, 4):
                v = tk.StringVar(value="0")
                row(s2, f"ResID{n}:", v)
                res_vars.append(v)

            s3 = sec("  Required Items (ReqID / ReqNum 1-5)  ")
            tk.Label(s3, text="  Items the player must hand in. 0 = unused.", bg=BG, fg=FG_GREY,
                     font=("Consolas", 8)).pack(anchor="w", padx=8)
            req_id_vars = []; req_num_vars = []
            for n in range(1, 6):
                r2 = tk.Frame(s3, bg=BG); r2.pack(fill="x", padx=8, pady=2)
                tk.Label(r2, text=f"ReqID{n}:", width=10, anchor="w", bg=BG, fg=FG,
                         font=("Consolas", 9)).pack(side="left")
                vid = tk.StringVar(value="0")
                tk.Entry(r2, textvariable=vid, width=12, bg=BG3, fg=FG,
                         insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
                tk.Label(r2, text=f"ReqNum{n}:", width=11, anchor="w", bg=BG, fg=FG_DIM,
                         font=("Consolas", 9)).pack(side="left", padx=(8, 0))
                vnum = tk.StringVar(value="0")
                tk.Entry(r2, textvariable=vnum, width=7, bg=BG3, fg=FG,
                         insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
                req_id_vars.append(vid); req_num_vars.append(vnum)

            s4 = sec("  Settings  ")
            v_fee = tk.StringVar(value="0")
            row(s4, "Fee:", v_fee, "Galder fee for this exchange.")

            def _add():
                cfg = {
                    "exchange_id": v_eid.get(), "name": v_name.get(), "comment": v_cmt.get(),
                    "res_lv": v_reslv.get(),
                    "res_id1": res_vars[0].get(), "res_id2": res_vars[1].get(),
                    "res_id3": res_vars[2].get(),
                    "fee": v_fee.get(),
                }
                for n in range(1, 6):
                    cfg[f"req_id{n}"]  = req_id_vars[n-1].get()
                    cfg[f"req_num{n}"] = req_num_vars[n-1].get()
                try: _set_last_id("exchange", int(v_eid.get()))
                except: pass
                self._exchange_rows.append((build_exchange_row(cfg),
                                            build_exchange_location_row(cfg["exchange_id"])))
                self.session.exchange_rows = list(self._exchange_rows)
                if messagebox.askyesno("Added",
                        f"Exchange ID {v_eid.get()} added.\nAdd another?"):
                    try: v_eid.set(str(int(v_eid.get()) + 1))
                    except: pass
                else:
                    self._build_output_screen()

        # ── SHOP LISTING FORM ──────────────────────────────────────────────
        elif mode == "shop":
            s1 = sec("  Shop Item Entry  ")
            tk.Label(s1, text="  Output format:  ItemID# Count.000000 Price.000000 Unk",
                     bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(anchor="w", padx=8, pady=(4, 0))

            # Batch text area
            tk.Label(s1, text="  Enter one Item ID per line (or paste many at once):",
                     bg=BG, fg=FG, font=("Consolas", 9)).pack(anchor="w", padx=8, pady=(6, 2))
            id_txt = scrolledtext.ScrolledText(s1, height=8, font=("Consolas", 9),
                                               bg=BG3, fg=FG, insertbackground=FG)
            id_txt.pack(fill="x", padx=8, pady=4)

            s2 = sec("  Defaults for all items  ")
            v_count = tk.StringVar(value="100")
            v_price = tk.StringVar(value="0")
            v_unk   = tk.StringVar(value="0")
            row(s2, "Count:", v_count, "How many the shop sells (default 100).")
            row(s2, "Price:", v_price, "Price (integer, will become X.000000).")
            row(s2, "Unk:", v_unk,    "Unknown trailing field — usually 0.")

            def _add():
                raw = id_txt.get("1.0", "end").strip()
                ids = [l.strip() for l in raw.splitlines() if l.strip()]
                if not ids:
                    messagebox.showwarning("No IDs", "Enter at least one Item ID."); return
                for iid in ids:
                    self._shop_rows.append(
                        build_shop_row(iid, v_count.get(), v_price.get(), v_unk.get()))
                if messagebox.askyesno("Added",
                        f"{len(ids)} shop row(s) added.\nAdd more?"):
                    id_txt.delete("1.0", "end")
                else:
                    self._build_output_screen()

        # ── Footer nav ─────────────────────────────────────────────────────
        nav = tk.Frame(self, bg=BG2); nav.pack(fill="x", side="bottom")
        mk_btn(nav, "◀  Back", self._build_start_screen,
               color=BG4).pack(side="left", padx=12, pady=6)
        mk_btn(nav, "✓  Add Entry" if mode != "shop" else "✓  Add Items",
               _add, color=GREEN, fg=BG2,
               font=("Consolas", 11, "bold")).pack(side="right", padx=14, pady=6)

        n_c, n_e, n_s = len(self._compound_rows), len(self._exchange_rows), len(self._shop_rows)
        if n_c or n_e or n_s:
            mk_btn(nav, f"📋  View ({n_c}C·{n_e}E·{n_s}S)", self._build_output_screen,
                   color=BG3).pack(side="right", padx=4, pady=6)

    # ─────────────────────────────────────────────────────────────────────────
    # IMPORT FROM FILE
    # ─────────────────────────────────────────────────────────────────────────
    def _import_from_file(self):
        path = filedialog.askopenfilename(
            title="Import spreadsheet",
            filetypes=[("Spreadsheet","*.csv *.xlsx *.xlsm *.xls"),
                       ("CSV","*.csv"),("Excel","*.xlsx *.xlsm *.xls"),("All","*.*")],
            parent=self.root)
        if not path: return
        ext = os.path.splitext(path)[1].lower()

        win = tk.Toplevel(self.root)
        win.title("Import Mode"); win.configure(bg=BG)
        win.resizable(False, False); win.grab_set()
        tk.Label(win, text="What data does this file contain?",
                 bg=BG, fg=FG, font=("Consolas", 11, "bold"), pady=14).pack(padx=28)
        bf = tk.Frame(win, bg=BG); bf.pack(padx=28, pady=8)

        def _do_import(mode):
            win.destroy()
            try:
                if ext in (".xlsx", ".xlsm", ".xls"):
                    if not _HAVE_OPENPYXL:
                        messagebox.showerror("Missing library",
                            "openpyxl required for Excel.\npip install openpyxl"); return
                    wb = openpyxl.load_workbook(path, data_only=True)
                    all_groups = []
                    for sh_name in wb.sheetnames:
                        csv_text = _sheet_to_csv(wb[sh_name])
                        if mode == "shop":
                            all_groups.extend(self._parse_shop_csv(csv_text, sh_name))
                        else:
                            all_groups.extend(_parse_ce_grouped_csv(csv_text, mode))
                else:
                    with open(path, encoding="utf-8-sig") as f:
                        csv_text = f.read()
                    if mode == "shop":
                        all_groups = self._parse_shop_csv(csv_text, os.path.basename(path))
                    else:
                        all_groups = _parse_ce_grouped_csv(csv_text, mode)
            except Exception as e:
                messagebox.showerror("Import Error", str(e)); return
            if not all_groups:
                messagebox.showwarning("No data", "No valid rows found."); return
            self._preview_and_confirm(all_groups, path, mode)

        def _do_both():
            win.destroy()
            self._import_mixed_file(path, ext)

        mk_btn(bf, "⚗  Compound", lambda: _do_import("compound"),
               color=ACC1, fg=BG2, font=("Consolas",10,"bold"),
               padx=14, pady=8).grid(row=0, column=0, padx=6, pady=4)
        mk_btn(bf, "🔄  Exchange", lambda: _do_import("exchange"),
               color=ACC2, fg=BG2, font=("Consolas",10,"bold"),
               padx=14, pady=8).grid(row=0, column=1, padx=6, pady=4)
        mk_btn(bf, "🛒  Shop List", lambda: _do_import("shop"),
               color=ACC5, fg=BG2, font=("Consolas",10,"bold"),
               padx=14, pady=8).grid(row=0, column=2, padx=6, pady=4)
        mk_btn(bf, "Auto-detect (compound+exchange)", _do_both,
               color=BG3, font=("Consolas",9),
               padx=10, pady=8).grid(row=1, column=0, columnspan=3, padx=6, pady=4)
        mk_btn(win, "Cancel", win.destroy, color=BG4).pack(pady=(0, 12))
        win.wait_window()

    def _parse_shop_csv(self, csv_text, group_name):
        """Parse a shop CSV. Columns: ItemID, Count, Price, Unk (only ItemID required)."""
        reader = csv.DictReader(io.StringIO(csv_text))
        rows = []
        id_key = next((k for k in (reader.fieldnames or [])
                       if k.strip().lower() in ("itemid","id","item_id","item")), None)
        if not id_key: return []
        for raw in reader:
            iid = raw.get(id_key, "").strip()
            if not iid: continue
            count = raw.get("Count", raw.get("count", "100")).strip() or "100"
            price = raw.get("Price", raw.get("price", "0")).strip() or "0"
            unk   = raw.get("Unk",   raw.get("unk",   "0")).strip() or "0"
            rows.append({"item_id": iid, "count": count, "price": price, "unk": unk})
        if not rows: return []
        return [{"mode": "shop", "group_name": group_name, "rows": rows}]

    def _import_mixed_file(self, path, ext):
        try:
            if ext in (".xlsx", ".xlsm", ".xls"):
                if not _HAVE_OPENPYXL:
                    messagebox.showerror("Missing library","openpyxl required."); return
                wb = openpyxl.load_workbook(path, data_only=True)
                all_groups = []
                for sh_name in wb.sheetnames:
                    csv_text = _sheet_to_csv(wb[sh_name])
                    grps = _parse_ce_grouped_csv(csv_text, "compound")
                    if not grps: grps = _parse_ce_grouped_csv(csv_text, "exchange")
                    all_groups.extend(grps)
            else:
                with open(path, encoding="utf-8-sig") as f: csv_text = f.read()
                all_groups = _parse_ce_grouped_csv(csv_text, "compound")
                if not all_groups: all_groups = _parse_ce_grouped_csv(csv_text, "exchange")
        except Exception as e:
            messagebox.showerror("Import Error", str(e)); return
        if not all_groups:
            messagebox.showwarning("No data","No valid rows found."); return
        self._preview_and_confirm(all_groups, path, "mixed")

    def _preview_and_confirm(self, groups, filepath, mode):
        win = tk.Toplevel(self.root)
        win.title(f"Import Preview — {len(groups)} group(s)")
        win.configure(bg=BG); win.geometry("920x560"); win.grab_set()

        hdr = tk.Frame(win, bg=BG2); hdr.pack(fill="x")
        tk.Label(hdr, text=f"📥  {os.path.basename(filepath)}",
                 bg=BG2, fg=self.ACC, font=("Consolas",12,"bold"), pady=8).pack(side="left", padx=14)
        total_rows = sum(len(g["rows"]) for g in groups)
        tk.Label(hdr, text=f"{total_rows} rows in {len(groups)} group(s)",
                 bg=BG2, fg=FG_DIM, font=("Consolas",9)).pack(side="left")

        sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True, padx=8, pady=6)
        canv, C = mk_scroll_canvas(sh)

        mode_colors = {"compound": ACC1, "exchange": ACC2, "shop": ACC5}
        for g in groups:
            gmode = g["mode"]
            gcolor = mode_colors.get(gmode, FG)
            gf = tk.LabelFrame(C,
                               text=f"  {g['group_name']}  [{gmode}]  ({len(g['rows'])} rows)",
                               bg=BG, fg=gcolor, font=("Consolas",9,"bold"), bd=1, relief="groove")
            gf.pack(fill="x", padx=8, pady=4)

            if gmode == "shop":
                cols = ["item_id", "count", "price", "unk"]
            elif gmode == "compound":
                cols = ["compound_id","name","res_id1","req_id1","req_num1","probability"]
            else:
                cols = ["exchange_id","name","res_id1","req_id1","req_num1","fee"]

            hrow = tk.Frame(gf, bg=BG2); hrow.pack(fill="x")
            for col in cols:
                tk.Label(hrow, text=col, bg=BG2, fg=BLUE, font=("Consolas",7,"bold"),
                         width=14, anchor="w", padx=2).pack(side="left")
            for cfg in g["rows"]:
                dr = tk.Frame(gf, bg=BG); dr.pack(fill="x")
                for col in cols:
                    tk.Label(dr, text=str(cfg.get(col,""))[:14], bg=BG, fg=FG,
                             font=("Consolas",7), width=14, anchor="w", padx=2).pack(side="left")

        nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")
        tk.Label(nav, text="Review above then confirm to add rows.",
                 bg=BG2, fg=FG_DIM, font=("Consolas",8)).pack(side="left", padx=14, pady=6)

        def _confirm():
            n_c = n_e = n_s = 0
            for g in groups:
                gmode = g["mode"]
                for cfg in g["rows"]:
                    if gmode == "compound":
                        self._compound_rows.append((
                            build_compound_row(cfg),
                            build_compound_location_row(cfg["compound_id"])))
                        n_c += 1
                    elif gmode == "exchange":
                        self._exchange_rows.append((
                            build_exchange_row(cfg),
                            build_exchange_location_row(cfg["exchange_id"])))
                        n_e += 1
                    elif gmode == "shop":
                        self._shop_rows.append(
                            build_shop_row(cfg["item_id"], cfg.get("count","100"),
                                           cfg.get("price","0"), cfg.get("unk","0")))
                        n_s += 1
            self.session.compound_rows = list(self._compound_rows)
            self.session.exchange_rows = list(self._exchange_rows)
            win.destroy()
            messagebox.showinfo("Imported",
                f"Added {n_c} compound · {n_e} exchange · {n_s} shop rows.")
            self._build_output_screen()

        mk_btn(nav, "✓  Confirm Import", _confirm,
               color=GREEN, fg=BG2, font=("Consolas",10,"bold")).pack(side="right", padx=14, pady=6)
        mk_btn(nav, "Cancel", win.destroy, color=BG4).pack(side="right", padx=4, pady=6)
        win.wait_window()

    def _import_session(self):
        n_c = len(self.session.compound_rows)
        n_e = len(self.session.exchange_rows)
        self._compound_rows = list(self.session.compound_rows)
        self._exchange_rows = list(self.session.exchange_rows)
        messagebox.showinfo("Session Imported",
            f"Imported: {n_c} compound  ·  {n_e} exchange")
        self._build_output_screen()

    def _save_template(self, mode):
        if mode == "compound":
            cols = list(_COMPOUND_DEFAULTS.keys())
        elif mode == "exchange":
            cols = list(_EXCHANGE_DEFAULTS.keys())
        else:
            cols = ["ItemID","Count","Price","Unk"]
        path = filedialog.asksaveasfilename(
            title=f"Save {mode} template", defaultextension=".csv",
            filetypes=[("CSV","*.csv")], initialfile=f"{mode}_template.csv",
            parent=self.root)
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            csv.writer(f).writerow(cols)
        messagebox.showinfo("Saved", f"Template saved:\n{path}")

    # ─────────────────────────────────────────────────────────────────────────
    # OUTPUT SCREEN
    # ─────────────────────────────────────────────────────────────────────────
    def _build_output_screen(self):
        self._clear()
        wrap = tk.Frame(self, bg=BG); wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(1, weight=1); wrap.grid_columnconfigure(0, weight=1)

        n_c, n_e, n_s = len(self._compound_rows), len(self._exchange_rows), len(self._shop_rows)
        hdr = tk.Frame(wrap, bg=BG2); hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="Compound / Exchange / Shop — Output",
                 font=("Consolas",13,"bold"), bg=BG2, fg=self.ACC, pady=8).pack(side="left", padx=15)
        tk.Label(hdr, text=f"  {n_c} compound  ·  {n_e} exchange  ·  {n_s} shop rows",
                 bg=BG2, fg=FG_DIM, font=("Consolas",9)).pack(side="left", padx=4)

        nb = ttk.Notebook(wrap); nb.grid(row=1, column=0, sticky="nsew", padx=6, pady=4)

        def _add_tab(title, content, fname):
            make_output_tab(nb, title, content, fname, self.root)

        if self._compound_rows:
            _add_tab("Compound_Potion rows",
                     "\n".join(r[0] for r in self._compound_rows),
                     "Compound_Potion_rows.xml")
            _add_tab("Compounder_Location rows",
                     "\n".join(r[1] for r in self._compound_rows),
                     "Compounder_Spot_rows.xml")
        if self._exchange_rows:
            _add_tab("ExchangeShopContents rows",
                     "\n".join(r[0] for r in self._exchange_rows),
                     "ExchangeShopContents_rows.xml")
            _add_tab("Exchange_Location rows",
                     "\n".join(r[1] for r in self._exchange_rows),
                     "Exchange_Location_rows.xml")
        if self._shop_rows:
            _add_tab("R_ShopItem rows",
                     "\n".join(self._shop_rows),
                     "R_ShopItem_rows.txt")

        if not n_c and not n_e and not n_s:
            tk.Label(wrap, text="No rows yet.",
                     bg=BG, fg=FG_GREY, font=("Consolas",10)).grid(row=1, column=0)

        nav = tk.Frame(wrap, bg=BG2); nav.grid(row=2, column=0, sticky="ew")
        mk_btn(nav, "◀  Back", self._build_start_screen, color=BG4).pack(side="left", padx=12, pady=6)
        mk_btn(nav, "⚗ +Compound", lambda: self._build_entry_form("compound"),
               color=BG4).pack(side="left", padx=2, pady=6)
        mk_btn(nav, "🔄 +Exchange", lambda: self._build_entry_form("exchange"),
               color=BG4).pack(side="left", padx=2, pady=6)
        mk_btn(nav, "🛒 +Shop",     lambda: self._build_entry_form("shop"),
               color=BG4).pack(side="left", padx=2, pady=6)
        mk_btn(nav, "📂 Import",    self._import_from_file, color=BG4).pack(side="left", padx=2, pady=6)
        mk_btn(nav, "🗑  Clear All",  self._clear_all,       color=BG4).pack(side="left", padx=2, pady=6)

        mk_btn(nav, "💾  Export All", self._export_all,
               color=GREEN, fg=BG2, font=("Consolas",10,"bold")).pack(side="right", padx=14, pady=6)

    def _clear_all(self):
        if messagebox.askyesno("Clear All", "Clear all compound, exchange, and shop rows?"):
            self._compound_rows.clear()
            self._exchange_rows.clear()
            self._shop_rows.clear()
            self.session.compound_rows = []
            self.session.exchange_rows = []
            self._build_start_screen()

    def _export_all(self):
        default_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(), "libconfig"))
        folder = filedialog.askdirectory(
            title="Export folder (libconfig)", initialdir=default_dir, parent=self.root)
        if not folder: folder = default_dir
        os.makedirs(folder, exist_ok=True)
        saved = []

        def _ts_name(key, fallback):
            fname = _APP_SETTINGS.get("filenames", {}).get(key, fallback)
            if _APP_SETTINGS.get("timestamp_files", False):
                import time as _t
                ts = _t.strftime("%d%m%y-%S%M%H")
                n, e = os.path.splitext(fname)
                fname = f"{n}_{ts}{e}"
            return fname

        if self._compound_rows:
            for key, fallback, content in [
                ("compound_potion",     "Compound_Potion.xml",
                 "\n".join(r[0] for r in self._compound_rows)),
                ("compounder_location", "Compounder_Location.xml",
                 "\n".join(r[1] for r in self._compound_rows)),
            ]:
                fn = _ts_name(key, fallback)
                with open(os.path.join(folder, fn), "w", encoding="utf-8") as f: f.write(content)
                saved.append(fn)

        if self._exchange_rows:
            for key, fallback, content in [
                ("exchange_contents",  "ExchangeShopContents.xml",
                 "\n".join(r[0] for r in self._exchange_rows)),
                ("exchange_location",  "Exchange_Location.xml",
                 "\n".join(r[1] for r in self._exchange_rows)),
            ]:
                fn = _ts_name(key, fallback)
                with open(os.path.join(folder, fn), "w", encoding="utf-8") as f: f.write(content)
                saved.append(fn)

        if self._shop_rows:
            fn = _ts_name("shop_item", "R_ShopItem.xml")
            with open(os.path.join(folder, fn), "w", encoding="utf-8") as f:
                f.write("\n".join(self._shop_rows))
            saved.append(fn)

        messagebox.showinfo("Export Complete",
            f"Saved to:\n{folder}\n\n" + "\n".join(saved))


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 18 — Fashion Creation
# ══════════════════════════════════════════════════════════════════════════════

ACC18 = "#f2cdcd"   # rosewater — fashion creation (defined again below with full data)

# Fox is special: inherits Goods_category0=15 as its category1 effectively
_FOX_SPECIAL = True  # Fox has no unique libcmgds goods_category1, uses default Goods_category0=15

# ── CSV column normaliser for fashion imports ─────────────────────────────────
def _norm_fashion_hdr(h):
    return re.sub(r"[\s_\-]+", "", h).lower()

_FASH_BOX_KEYS    = {"boxid","boxitemid","boxparamid"}
_FASH_SET_KEYS    = {"setid","setitemid","cmsetid"}
_FASH_ITEM_KEYS   = {"itemid","fashionid","itemparamid"}
_FASH_SHOPROW_KEYS= {"shoprowid","shopid","showr_id","shoprow"}


def _detect_fashion_section(headers):
    """Return 'box'|'set'|'item'|'shoprow'|'unknown' per header list."""
    nh = [_norm_fashion_hdr(h) for h in headers]
    # Heuristic: look for distinguishing columns
    has_box   = any(k in _FASH_BOX_KEYS   for k in nh)
    has_set   = any(k in _FASH_SET_KEYS   for k in nh)
    has_item  = any(k in _FASH_ITEM_KEYS  for k in nh)
    has_shop  = any(k in _FASH_SHOPROW_KEYS for k in nh)
    # Also check for column names unique to each section
    has_dropid = any("drop" in k for k in nh)
    has_ncash  = any("ncash" in k or "cash" in k for k in nh)
    has_count  = any(k in ("count","bundlenum","invbundlenum") for k in nh)

    if has_box or has_dropid: return "box"
    if has_set: return "set"
    if has_shop: return "shoprow"
    if has_item or has_ncash: return "item"
    return "unknown"


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 18 — Fashion Creation
# Single-screen design: character, pricing, filenames, set/box IDs, options,
# and fashion piece table all on one scrollable canvas.
# Import via button on start screen (like Tool 1/6).
# Session data persists to disk via _load_settings / _save_settings.
# ══════════════════════════════════════════════════════════════════════════════

ACC18 = "#f2cdcd"   # rosewater

# ── Character catalogue ────────────────────────────────────────────────────
# name → (goods_cat1, goods_char_sex, goods_char_type_for_sql, libcmgds_cat1)
_CHAR_DATA = {
    "Bunny 1st":   (1,  2, 8,  128),
    "Buffalo 1st": (2,  1, 8,  129),
    "Sheep 1st":   (3,  2, 4,  130),
    "Dragon 1st":  (4,  1, 4,  131),
    "Fox 1st":     (5,  2, 2,  None),
    "Lion 1st":    (6,  1, 2,  133),
    "Cat 1st":     (7,  2, 1,  134),
    "Raccoon 1st": (8,  1, 1,  135),
    "Bunny 2nd":   (1,  2, 8,  128),
    "Buffalo 2nd": (2,  1, 8,  129),
    "Sheep 2nd":   (3,  2, 4,  130),
    "Dragon 2nd":  (4,  1, 4,  131),
    "Fox 2nd":     (5,  2, 2,  None),
    "Lion 2nd":    (6,  1, 2,  133),
    "Cat 2nd":     (7,  2, 1,  134),
    "Raccoon 2nd": (8,  1, 1,  135),
    "Bunny 3rd":   (1,  2, 8,  128),
    "Buffalo 3rd": (2,  1, 8,  129),
    "Sheep 3rd":   (3,  2, 4,  130),
    "Dragon 3rd":  (4,  1, 4,  131),
    "Fox 3rd":     (5,  2, 2,  None),
    "Lion 3rd":    (6,  1, 2,  133),
    "Cat 3rd":     (7,  2, 1,  134),
    "Raccoon 3rd": (8,  1, 1,  135),
    "Paula 1st":   (1,  2, 9,  128),
    "Paula 2nd":   (1,  2, 9,  128),
    "Paula 3rd":   (1,  2, 9,  128),
}

# chr key  → fashion piece name list  (from constants.py FASHION_NAMES, mapped by char+job)
# Key format matching _CHAR_DATA: "Name Job"
_FASHION_PIECES = {
    "Bunny 1st":   ["Hoodie","Gloves","Skort","Backpack","Shoes"],
    "Buffalo 1st": ["Airshoes","Turtle Vest","Sash Belt","Warmups","Wraps","Hood Tie"],
    "Sheep 1st":   ["Blouse","Bow","Frill Dress","Flats","Socks","Spellbook"],
    "Dragon 1st":  ["Robe","Shirt","Jeans","Monk Shoes","Cane"],
    "Fox 1st":     ["Coat","Heels","Slit Skirt","Tank","Tote"],
    "Lion 1st":    ["Jacket","Shorts","Trainers","Open Glove","T-neck"],
    "Cat 1st":     ["Ribbon","Belt","Halter","Heels","Paws","Skirt"],
    "Raccoon 1st": ["Blazer","Slacks","Dress Shoes"],
    "Bunny 2nd":   ["Robe","Boxing Glove","Shorts","Gloves","Boxing Shoes","Stocking"],
    "Buffalo 2nd": ["Fur Collar","Tunic","Bolero","Gauntlet","Leather Shoes"],
    "Sheep 2nd":   ["Checkered Dress","Ribbon","Minisack","Gloves","Ribbon Boots"],
    "Dragon 2nd":  ["Shawl","Beads Necklace","Robe","Wrap Skirt","Ankle Boots"],
    "Fox 2nd":     ["Sports Suit","Tube Top","Elbow Wrap","Mittens","Walkers"],
    "Lion 2nd":    ["Turtleneck","Coil Coat","Utility Belt","Glove","Boots"],
    "Cat 2nd":     ["Hippie Shirt","Studded Belt","Checkered Skirt","Checkered Stockings","Heel Boots"],
    "Raccoon 2nd": ["Dress Shirt","Checkered Suit","Dress Shoes"],
    "Bunny 3rd":   ["Tube Top","Bolero Jacket","Gauntlets","Chord Skirt","Steel Boots"],
    "Buffalo 3rd": ["Asymmetrical Tee","Protector","Kilt","Steel Armlets","Ankle Shoes"],
    "Sheep 3rd":   ["Flower Ribbon","Puffy Blouse","Flower Brooch","Layered Dress","Flower Shoes"],
    "Dragon 3rd":  ["Wrap","Hooded Robe","Overcoat","Robe","Leather Boots"],
    "Fox 3rd":     ["Zip-up Coat","Leather Shorts","Leather Wristlets","Buckle Boots","Unknown"],
    "Lion 3rd":    ["Zip-up Jacket","Long Jacket","Shorts","Long Boots","Unknown"],
    "Cat 3rd":     ["Double Coat","Shirring Skirt","Buckle Shoes","Blouse"],
    "Raccoon 3rd": ["Dress Shirt","Opera Cape","Frock Coat","Dress Pants","Formal Shoes","Unknown"],
    "Paula 1st":   ["Stadium Jacket","Sleeveless Dress","Knee Socks","School Loafers",
                    "Ribbon Chou","Cutie Satchel","Extra"],
    "Paula 2nd":   ["Pocket One-piece","Animal Pocket Belt","Knee-high Boots",
                    "Ribbons","Arm Cover","Whip"],
    "Paula 3rd":   ["Blouse","Trench Dress","Frilly Socks","Cutie Buckle Boots",
                    "Ribbon Rubber","Ribbon Brooch","Mini Pocket Belt","Leather Buckle Gloves"],
}

MAX_FASHION_PIECES = 7

# ── Helpers for fashion section detection (CSV import) ─────────────────────
def _norm_fashion_hdr(h):
    return re.sub(r"[\s_\-]+", "", h).lower()

_FASH_BOX_KEYS     = {"boxid","boxitemid","dropid"}
_FASH_SET_KEYS     = {"setid","setitemid","cmsetid"}
_FASH_ITEM_KEYS    = {"itemid","fashionid","itemparamid"}
_FASH_SHOPROW_KEYS = {"shoprowid","shopid","shoprow"}

def _detect_fashion_section(headers):
    nh = [_norm_fashion_hdr(h) for h in headers]
    if any(k in _FASH_BOX_KEYS   for k in nh): return "box"
    if any(k in _FASH_SET_KEYS   for k in nh): return "set"
    if any(k in _FASH_SHOPROW_KEYS for k in nh): return "shoprow"
    if any(k in _FASH_ITEM_KEYS  for k in nh): return "item"
    if any("drop" in k for k in nh): return "box"
    return "item"



# ── Fashion ItemParam slot-type detection ─────────────────────────────────────
# Keywords (lowercase) in piece name → (Class, Type, SubType)
# Longer / more specific phrases checked first.
_FASH_SLOT_KEYWORDS = [
    # Footwear
    ("boxing shoes", 2, 5, 5), ("dress shoes", 2, 5, 5), ("ankle boots", 2, 5, 5),
    ("ankle shoes", 2, 5, 5), ("buckle boots", 2, 5, 5), ("buckle shoes", 2, 5, 5),
    ("heel boots", 2, 5, 5), ("long boots", 2, 5, 5), ("leather boots", 2, 5, 5),
    ("leather shoes", 2, 5, 5), ("ribbon shoes", 2, 5, 5), ("formal shoes", 2, 5, 5),
    ("steel boots", 2, 5, 5), ("utility boots", 2, 5, 5), ("school loafers", 2, 5, 5),
    ("knee-high boots", 2, 5, 5), ("ribbon boots", 2, 5, 5), ("cutie buckle boots", 2, 5, 5),
    ("monk shoes", 2, 5, 5), ("air shoes", 2, 5, 5), ("airshoes", 2, 5, 5),
    ("trainers", 2, 5, 5), ("walkers", 2, 5, 5), ("high heels", 2, 5, 5),
    ("heels", 2, 5, 5), ("flats", 2, 5, 5), ("loafers", 2, 5, 5),
    ("boots", 2, 5, 5), ("shoes", 2, 5, 5),
    # Stockings/Socks
    ("checkered stockings", 2, 27, 51), ("checkered stocking", 2, 27, 51),
    ("knee socks", 2, 27, 51), ("frilly socks", 2, 27, 51),
    ("stocking", 2, 27, 51), ("stockings", 2, 27, 51),
    ("socks", 2, 27, 51), ("sock", 2, 27, 51),
    # Gloves / Hands
    ("boxing glove", 2, 26, 50), ("boxing gloves", 2, 26, 50),
    ("open glove", 2, 26, 50), ("heavy glove", 2, 26, 50),
    ("leather buckle gloves", 2, 26, 50), ("leather wristlets", 2, 26, 50),
    ("elbow wrap", 2, 26, 50), ("steel armlets", 2, 26, 50),
    ("arm cover", 2, 26, 50), ("gauntlets", 2, 26, 50), ("gauntlet", 2, 26, 50),
    ("mittens", 2, 26, 50), ("armlets", 2, 26, 50), ("wristlets", 2, 26, 50),
    ("bandage wraps", 2, 26, 50), ("gloves", 2, 26, 50), ("glove", 2, 26, 50),
    ("paws", 2, 26, 50), ("paw", 2, 26, 50), ("wraps", 2, 26, 50),
    # Belt
    ("sash belt", 2, 22, 46), ("strap belt", 2, 22, 46), ("studded belt", 2, 22, 46),
    ("mini pocket belt", 2, 22, 46), ("animal pocket belt", 2, 22, 46),
    ("utility belt", 2, 22, 46), ("belt", 2, 22, 46), ("sash", 2, 22, 46),
    # Accessories (bags, wands, ribbons, necklaces, etc.)
    ("beads necklace", 2, 23, 47), ("fur collar", 2, 23, 47),
    ("opera cape", 2, 23, 47), ("robe wrap", 2, 23, 47), ("robe warp", 2, 23, 47),
    ("ribbon chou", 2, 23, 47), ("cutie satchel", 2, 23, 47),
    ("mini-bag", 2, 23, 47), ("mini bag", 2, 23, 47), ("mini sack", 2, 23, 47),
    ("tote bag", 2, 23, 47), ("tote", 2, 23, 47), ("backpack", 2, 23, 47),
    ("satchel", 2, 23, 47), ("spell book", 2, 23, 47), ("spellbook", 2, 23, 47),
    ("pearl wand", 2, 23, 47), ("wand", 2, 23, 47), ("cane", 2, 23, 47),
    ("whip", 2, 23, 47), ("brooch", 2, 23, 47), ("necklace", 2, 23, 47),
    ("ribbon brooch", 2, 23, 47), ("ribbon rubber", 2, 23, 47),
    ("hood tie", 2, 23, 47), ("bandana hood tie", 2, 23, 47), ("bandana", 2, 23, 47),
    ("collar", 2, 23, 47), ("bow", 2, 23, 47), ("ribbons", 2, 23, 47),
    ("ribbon", 2, 23, 47),  # catch-all ribbon = accessory
    # Bottoms
    ("wrap skirt", 2, 20, 44), ("slit skirt", 2, 20, 44),
    ("checkered skirt", 2, 20, 44), ("shirring skirt", 2, 20, 44),
    ("frill dress", 2, 20, 44), ("ruffle dress", 2, 20, 44),
    ("layered dress", 2, 20, 44), ("cord skirt", 2, 20, 44),
    ("hooded robe", 2, 20, 44), ("dress pants", 2, 20, 44),
    ("warm-up pants", 2, 20, 44), ("warm-ups", 2, 20, 44), ("warmups", 2, 20, 44),
    ("leather shorts", 2, 20, 44), ("skort", 2, 20, 44), ("skirt", 2, 20, 44),
    ("shorts", 2, 20, 44), ("jeans", 2, 20, 44), ("slacks", 2, 20, 44),
    ("kilt", 2, 20, 44), ("dress", 2, 20, 44),
    ("pants", 2, 20, 44),
    # Outerwear/Jacket (21/45)
    ("bolero jacket", 2, 21, 45), ("zip-up jacket", 2, 21, 45),
    ("zip-up coat", 2, 21, 45), ("double coat", 2, 21, 45),
    ("frock coat", 2, 21, 45), ("coil coat", 2, 21, 45),
    ("checkered suit", 2, 21, 45), ("sport suit", 2, 21, 45),
    ("sports suit", 2, 21, 45), ("stadium jacket", 2, 21, 45),
    ("pocket one-piece", 2, 21, 45),
    ("turtle vest", 2, 21, 45), ("turtleneck", 2, 21, 45), ("t-neck", 2, 21, 45),
    ("bolero", 2, 21, 45), ("blazer", 2, 21, 45), ("jacket", 2, 21, 45),
    ("coat", 2, 21, 45), ("protector", 2, 21, 45), ("tunic", 2, 21, 45),
    ("robe", 2, 21, 45),  # catch-all robe = outerwear
    # Tops (19/43) — fallback
    ("asymmetrical tee", 2, 19, 43), ("hippie shirt", 2, 19, 43),
    ("dress shirt", 2, 19, 43), ("long shirt", 2, 19, 43),
    ("puffy blouse", 2, 19, 43), ("tube top", 2, 19, 43),
    ("color tube top", 2, 19, 43),
    ("tank top", 2, 19, 43), ("tank", 2, 19, 43),
    ("hoodie", 2, 19, 43), ("blouse", 2, 19, 43),
    ("vest", 2, 19, 43), ("shirt", 2, 19, 43), ("shawl", 2, 19, 43),
    ("overcoat", 2, 19, 43), ("top", 2, 19, 43), ("halter", 2, 19, 43),
    ("trench dress", 2, 19, 43),  # trench dress = top slot for paula
]

def _fashion_detect_slot(piece_name):
    """Given a fashion piece name, return (Class, Type, SubType).
    Falls back to (2, 19, 43) = top if no keyword matches."""
    nl = piece_name.lower()
    for entry in _FASH_SLOT_KEYWORDS:
        kw = entry[0]
        if kw in nl:
            return entry[1], entry[2], entry[3]
    return 2, 19, 43   # default: top

# Color name substitution placeholders (case-insensitive)
_COLOR_PLACEHOLDERS = [
    "colorname", "color name", "colornameinlowercase", "colornamelc",
    "colorlowercase", "colorinlowercase", "colorlc", "colour",
    "colourname", "colornamelowerc", "colornamelc",
]

def _apply_color(text, color_name):
    """Replace all color placeholder tokens in text with color_name.
    Preserves case: ColorName->TitleCase, colorlc->lowercase.
    Removes 'colored ' after the word 'colored ' if color is empty."""
    if not text:
        return text
    result = text

    # Title-case placeholders (ColorName, Colorname, etc.)
    for ph in ["ColorName", "Colorname", "Color Name", "Color"]:
        if ph in result:
            result = result.replace(ph, color_name if color_name else "")

    # Lowercase placeholders
    for ph in ["colorname", "color name", "colornameinlowercase", "colornamelc",
               "colorlowercase", "colorinlowercase", "colorlc", "colornamelowerc"]:
        if ph in result.lower():
            # find actual casing in string
            lo = result.lower()
            idx = 0
            while True:
                idx = lo.find(ph, idx)
                if idx == -1: break
                result = result[:idx] + (color_name.lower() if color_name else "") + result[idx+len(ph):]
                lo = result.lower()
                idx += len(color_name) if color_name else 0

    # Clean up leading/trailing spaces left by empty color substitution
    if not color_name:
        result = re.sub(r'  +', ' ', result).strip()

    return result


class Tool18(tk.Frame):
    """Fashion Creation — single-screen workflow with persistent session."""

    ACC = ACC18
    _PERSIST_KEY = "t18_fashion"

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root    = root
        self.session = session
        self._item_row_widgets = []   # list of (frame, var_id, var_name)
        self._import_status    = tk.StringVar(value="")
        self._nb               = None  # reference to main notebook for tab switching
        # Persistent vars — loaded from disk
        self._set_id   = tk.StringVar()
        self._set_name = tk.StringVar()
        self._box_id   = tk.StringVar()
        self._box_name = tk.StringVar()
        self._box_comment = tk.StringVar()
        self._build_start_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()
        self._item_row_widgets = []

    # ─────────────────────────────────────────────────────────────────────────
    # START SCREEN — import or new
    # ─────────────────────────────────────────────────────────────────────────
    def _build_start_screen(self):
        self._clear()
        frm = tk.Frame(self, bg=BG); frm.pack(fill="both", expand=True)
        center = tk.Frame(frm, bg=BG); center.pack(expand=True)

        tk.Label(center, text="✨  FASHION CREATION",
                 font=("Consolas", 20, "bold"), bg=BG, fg=self.ACC).pack(pady=(32, 6))
        tk.Label(center,
                 text="Generate ItemParam rows · CMSetItemParam · Box row · libcmgds_e · SQL\n"
                      "R_ShopItem · Exchange · Compound",
                 bg=BG, fg=FG_DIM, font=("Consolas", 9), justify="center").pack(pady=(0, 16))

        # Info box
        info = tk.Frame(center, bg=BG2); info.pack(pady=6, padx=24, fill="x")
        tk.Label(info,
            text=(
                "  NEW:  Enter everything on one screen and generate.\n"
                "  IMPORT CSV/XLSX:  Auto-fills the form from your spreadsheet.\n"
                "  PASTE TXT:  Extracts IDs from a plain-text fashion note.\n"
                "  RESUME SESSION:  Reload last saved fashion set from disk.\n\n"
                "  Column guide (CSV):\n"
                "    Items:   ItemID, Name\n"
                "    Box:     BoxID, BoxName\n"
                "    Set:     SetID, SetName\n"
                "    Shop:    ShopRowID / ShopID\n"
            ),
            bg=BG2, fg=FG, font=("Consolas", 8), justify="left",
            padx=12, pady=8).pack(anchor="w")

        bf = tk.Frame(center, bg=BG); bf.pack(pady=16)
        mk_btn(bf, "✏️  New Fashion Set", self._build_editor,
               color=self.ACC, fg=BG2, font=("Consolas", 11, "bold")).pack(side="left", padx=8)
        mk_btn(bf, "📂  Import CSV / Excel", self._import_fashion_file,
               color=BG3).pack(side="left", padx=8)
        mk_btn(bf, "📋  Paste TXT", self._import_txt_paste,
               color=BG4).pack(side="left", padx=8)

        # Resume session button — always available if saved data exists
        saved = _load_settings(self._PERSIST_KEY)
        if saved.get("items") or saved.get("set_id"):
            n = len(saved.get("items", []))
            mk_btn(center,
                   f"⬇  Resume Last Session  ({n} item(s) — {saved.get('set_name','') or 'unnamed'})",
                   self._resume_session, color=GREEN, fg=BG2).pack(pady=4)

        if self._import_status.get():
            tk.Label(center, textvariable=self._import_status,
                     bg=BG, fg=GREEN, font=("Consolas", 9)).pack(pady=4)

    # ─────────────────────────────────────────────────────────────────────────
    # RESUME SESSION from disk
    # ─────────────────────────────────────────────────────────────────────────
    def _resume_session(self):
        saved = _load_settings(self._PERSIST_KEY)
        if not saved:
            messagebox.showinfo("No Session", "No saved fashion session found."); return
        self._build_editor(prefill=saved)

    # ─────────────────────────────────────────────────────────────────────────
    # MAIN EDITOR — all on one scrollable screen
    # ─────────────────────────────────────────────────────────────────────────
    def _build_editor(self, prefill=None):
        self._clear()
        pf = prefill or {}

        # ── Header ──────────────────────────────────────────────────────────
        hdr = tk.Frame(self, bg=BG2); hdr.pack(fill="x")
        tk.Label(hdr, text="✨  Fashion Creation",
                 font=("Consolas", 13, "bold"), bg=BG2, fg=self.ACC, pady=8).pack(side="left", padx=14)
        mk_btn(hdr, "◀  Back", self._build_start_screen,
               color=BG4, font=("Consolas", 8)).pack(side="left", padx=4)

        # Import from session button — always in editor
        def _import_session_btn():
            saved = _load_settings(self._PERSIST_KEY)
            if not saved.get("items"):
                messagebox.showinfo("No Session", "No saved fashion session on disk."); return
            if messagebox.askyesno("Load Session",
                    f"Load saved session?\n({len(saved.get('items',[]))} items — "
                    f"{saved.get('set_name','unnamed')})"):
                self._build_editor(prefill=saved)

        mk_btn(hdr, "💾  Load Session", _import_session_btn,
               color=BG3, font=("Consolas", 8)).pack(side="left", padx=4)

        # ── Scrollable body ──────────────────────────────────────────────────
        sh = tk.Frame(self, bg=BG); sh.pack(fill="both", expand=True)
        canv, C = mk_scroll_canvas(sh)

        def sec(title, color=BLUE):
            f = tk.LabelFrame(C, text=title, bg=BG, fg=color,
                              font=("Consolas", 9, "bold"), bd=1, relief="groove")
            f.pack(fill="x", padx=12, pady=5)
            return f

        def field(parent, label, var, tip="", w=28):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=10, pady=2)
            lw = tk.Label(r, text=label, width=22, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9))
            lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=w, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip and _APP_SETTINGS.get("tooltips_enabled", True):
                _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent

        # ── ① Character & Job ──────────────────────────────────────────────
        s_char = sec("  ① Character & Job  ")
        char_names = list(_CHAR_DATA.keys())
        self._char_var = tk.StringVar(value=pf.get("char", char_names[0]))
        cr = tk.Frame(s_char, bg=BG); cr.pack(fill="x", padx=10, pady=8)
        tk.Label(cr, text="Character:", bg=BG, fg=FG, font=("Consolas", 10),
                 width=14, anchor="w").pack(side="left")
        char_menu = ttk.Combobox(cr, textvariable=self._char_var,
                                 values=char_names, width=22, state="readonly")
        char_menu.pack(side="left", padx=6)

        char_info_lbl = tk.Label(s_char, text="", bg=BG, fg=BLUE,
                                 font=("Consolas", 8), justify="left")
        char_info_lbl.pack(anchor="w", padx=16, pady=(0, 4))

        # Color Name field (optional — applies to piece names / comments)
        color_row = tk.Frame(s_char, bg=BG); color_row.pack(fill="x", padx=10, pady=4)
        self._color_name_var = tk.StringVar(value=pf.get("color_name", ""))
        tk.Label(color_row, text="Set Color Name:", bg=BG, fg=FG,
                 font=("Consolas", 9), width=18, anchor="w").pack(side="left")
        tk.Entry(color_row, textvariable=self._color_name_var, width=22,
                 bg=BG3, fg=FG, insertbackground=FG, font=("Consolas", 9),
                 relief="flat").pack(side="left", padx=4)
        tk.Label(color_row,
                 text="  Replaces ColorName/colorlc placeholders in piece names & comments",
                 bg=BG, fg=FG_GREY, font=("Consolas", 7)).pack(side="left")

        def _update_char_display(*_):
            name = self._char_var.get()
            cat1, sex, ct, lib_cat1 = _CHAR_DATA.get(name, (0, 0, 0, 0))
            fox_note = "  (Fox: no unique libcmgds_category1)" if "Fox" in name else ""
            char_info_lbl.config(
                text=f"  goods_cat1={cat1}  goods_char_type={ct}"
                     f"  libcmgds_category1={lib_cat1 or 'none'}{fox_note}")
            # Auto-fill fashion piece names (guarded — defined later in _build_editor)
            if hasattr(self, '_auto_fill_pieces'):
                self._auto_fill_pieces()

        char_menu.bind("<<ComboboxSelected>>", _update_char_display)

                # ── ② Pricing ────────────────────────────────────────────────────
        s_price = sec("  ② Pricing  (check all that apply — default: none)  ")
        tk.Label(s_price, text="  Each pricing field is independent. Enable only what you need.",
                 bg=BG, fg=FG_GREY, font=("Consolas", 7)).pack(anchor="w", padx=10, pady=(2,4))

        # Galder (Value field in ItemParam) — entry only shown when checked
        self._use_galder = tk.BooleanVar(value=pf.get("use_galder", False))
        self._galder_val = tk.StringVar( value=pf.get("galder_val", "0"))
        pg_row = tk.Frame(s_price, bg=BG); pg_row.pack(fill="x", padx=10, pady=2)
        pg_ent_frame = tk.Frame(s_price, bg=BG)
        def _toggle_galder(*_):
            if self._use_galder.get(): pg_ent_frame.pack(fill="x", padx=28, pady=(0,2))
            else: pg_ent_frame.pack_forget()
        tk.Checkbutton(pg_row, text="Galder price  (→ Value field in ItemParam)",
                       variable=self._use_galder, bg=BG, fg=FG, selectcolor=BG3,
                       activebackground=BG, font=("Consolas", 9),
                       command=_toggle_galder).pack(side="left")
        tk.Label(pg_ent_frame, text="Value:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 8), width=8, anchor="w").pack(side="left")
        tk.Entry(pg_ent_frame, textvariable=self._galder_val, width=12, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
        if pf.get("use_galder"): _toggle_galder()

        # NCash / Recycling (Ncash field in ItemParam)
        self._use_ncash = tk.BooleanVar(value=pf.get("use_ncash", False))
        self._ncash_val = tk.StringVar( value=pf.get("ncash_val", "0"))
        pn_row = tk.Frame(s_price, bg=BG); pn_row.pack(fill="x", padx=10, pady=2)
        pn_ent_frame = tk.Frame(s_price, bg=BG)
        def _toggle_ncash(*_):
            if self._use_ncash.get(): pn_ent_frame.pack(fill="x", padx=28, pady=(0,2))
            else: pn_ent_frame.pack_forget()
        tk.Checkbutton(pn_row, text="NCash / Recycling value  (→ Ncash field in ItemParam)",
                       variable=self._use_ncash, bg=BG, fg=FG, selectcolor=BG3,
                       activebackground=BG, font=("Consolas", 9),
                       command=_toggle_ncash).pack(side="left")
        tk.Label(pn_ent_frame, text="NCash:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 8), width=8, anchor="w").pack(side="left")
        tk.Entry(pn_ent_frame, textvariable=self._ncash_val, width=12, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
        tk.Label(pn_ent_frame, text="(tickets × 133 = NCash)",
                 bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(side="left")
        if pf.get("use_ncash"): _toggle_ncash()

        # MyShop (goods_cash_price in libcmgds_e only)
        self._use_myshop   = tk.BooleanVar(value=pf.get("use_myshop",   False))
        self._myshop_price = tk.StringVar( value=pf.get("myshop_price", "0"))
        pm_row = tk.Frame(s_price, bg=BG); pm_row.pack(fill="x", padx=10, pady=2)
        pm_ent_frame = tk.Frame(s_price, bg=BG)
        def _toggle_myshop(*_):
            if self._use_myshop.get(): pm_ent_frame.pack(fill="x", padx=28, pady=(0,2))
            else: pm_ent_frame.pack_forget()
        tk.Checkbutton(pm_row, text="MyShop price  (→ goods_cash_price in libcmgds_e)",
                       variable=self._use_myshop, bg=BG, fg=FG, selectcolor=BG3,
                       activebackground=BG, font=("Consolas", 9),
                       command=_toggle_myshop).pack(side="left")
        tk.Label(pm_ent_frame, text="goods_cash_price:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 8), width=18, anchor="w").pack(side="left")
        tk.Entry(pm_ent_frame, textvariable=self._myshop_price, width=12, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
        # goods_list_code start — user can override, persists between sessions
        _glc_row = tk.Frame(s_price, bg=BG); _glc_row.pack(fill="x", padx=28, pady=(0,2))
        tk.Label(_glc_row, text="goods_list_code start:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 8), width=22, anchor="w").pack(side="left")
        tk.Entry(_glc_row, textvariable=self._goods_list_code_start, width=10, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
        tk.Label(_glc_row, text="(auto-increments per item; persisted)",
                 bg=BG, fg=FG_GREY, font=("Consolas", 7)).pack(side="left", padx=4)
        # Individual item price (separate from set price — items sold standalone cost more)
        self._myshop_price_item = tk.StringVar(value=pf.get("myshop_price_item", "12000"))
        _ipr_row = tk.Frame(s_price, bg=BG); _ipr_row.pack(fill="x", padx=28, pady=(0,4))
        tk.Label(_ipr_row, text="item price (standalone):", bg=BG, fg=FG_DIM,
                 font=("Consolas", 8), width=22, anchor="w").pack(side="left")
        tk.Entry(_ipr_row, textvariable=self._myshop_price_item, width=10, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
        tk.Label(_ipr_row, text="(each piece sold individually, e.g. 12000)",
                 bg=BG, fg=FG_GREY, font=("Consolas", 7)).pack(side="left", padx=4)
        if pf.get("use_myshop"): _toggle_myshop()

        # Keep legacy vars for backwards compat with _collect_state / _generate
        self._price_mode = tk.StringVar(value="none")  # kept for session compat
        self._price_val  = self._galder_val             # alias
        self._cash_price = self._myshop_price           # alias

        # ── ③ File Names ─────────────────────────────────────────────────
        s_fn = sec("  ③ File Names  ")

        self._filename_var     = tk.StringVar(value=pf.get("filename",    r"data\item\your.nri"))
        self._bundlenum_var    = tk.StringVar(value=pf.get("bundlenum",   "0"))
        self._cmtfilename_var  = tk.StringVar(value=pf.get("cmtfilename", r"data\item\yourportrait.nri"))
        self._cmtbundlenum_var = tk.StringVar(value=pf.get("cmtbundlenum","0"))

        field(s_fn, "FileName:", self._filename_var,
              "Path to the item icon .nri file. Duplicated into InvFileName automatically.")
        field(s_fn, "BundleNum:", self._bundlenum_var,
              "Bundle index (usually 0). Duplicated into InvBundleNum.", w=10)
        field(s_fn, "CmtFileName:", self._cmtfilename_var,
              "Path to the portrait .nri file.")
        field(s_fn, "CmtBundleNum:", self._cmtbundlenum_var, w=10)
        # Prefix/suffix stubs kept for _generate compat
        self._name_prefix_var = tk.StringVar(value="")
        self._name_suffix_var = tk.StringVar(value="")
        self._name_pos_var    = tk.StringVar(value="before")

        # %placeholder% variable editor — user defines tokens used in paths
        ph_hdr = tk.Frame(s_fn, bg=BG); ph_hdr.pack(fill="x", padx=10, pady=(6,2))
        tk.Label(ph_hdr, text="Path variables  (%placeholder% → value, used in FileName/CmtFileName):",
                 bg=BG, fg=FG, font=("Consolas", 8)).pack(side="left")
        self._ph_vars = {}   # name → StringVar
        ph_saved = pf.get("path_vars", {})
        ph_rows_frame = tk.Frame(s_fn, bg=BG); ph_rows_frame.pack(fill="x", padx=12)
        self._ph_rows_frame = ph_rows_frame

        def _add_ph_row(name="", val=""):
            r = tk.Frame(ph_rows_frame, bg=BG); r.pack(fill="x", pady=1)
            vn = tk.StringVar(value=name)
            vv = tk.StringVar(value=val)
            tk.Label(r, text="%", bg=BG, fg=FG_GREY, font=("Consolas", 9)).pack(side="left")
            tk.Entry(r, textvariable=vn, width=12, bg=BG3, fg=FG, insertbackground=FG,
                     font=("Consolas", 9), relief="flat").pack(side="left")
            tk.Label(r, text="%  →", bg=BG, fg=FG_GREY, font=("Consolas", 9)).pack(side="left", padx=2)
            tk.Entry(r, textvariable=vv, width=20, bg=BG3, fg=FG, insertbackground=FG,
                     font=("Consolas", 9), relief="flat").pack(side="left", padx=2)
            def _del(r_=r, n_=vn):
                self._ph_vars.pop(n_.get(), None)
                r_.destroy()
            tk.Button(r, text="✕", command=_del, bg=BG4, fg=ACC3,
                      font=("Consolas", 8), relief="flat", width=2).pack(side="left", padx=2)
            def _track(*_): self._ph_vars[vn.get()] = vv
            vn.trace_add("write", _track); _track()
            self._ph_vars[vn.get()] = vv

        for ph_name, ph_val in ph_saved.items():
            _add_ph_row(ph_name, ph_val)

        ph_btn_row = tk.Frame(s_fn, bg=BG); ph_btn_row.pack(anchor="w", padx=12, pady=2)
        mk_btn(ph_btn_row, "+ Add Variable", _add_ph_row, color=BG4,
               font=("Consolas", 8)).pack(side="left")
        tk.Label(ph_btn_row, text="  e.g.  %color% → White   %char% → BN2",
                 bg=BG, fg=FG_GREY, font=("Consolas", 7)).pack(side="left", padx=6)


        # ── ④ Set & Box ──────────────────────────────────────────────────
        s_set = sec("  ④ Set & Box  ")
        _last_set = _get_last_id("t18_set_id", 0)
        self._set_id.set(pf.get("set_id", "") or (str(_last_set + 1) if _last_set else ""))
        self._set_name.set(pf.get("set_name", ""))
        _last_box = _get_last_id("t18_box_id", 0)
        self._box_id.set(pf.get("box_id", "") or (str(_last_box + 1) if _last_box else ""))
        self._box_name.set(pf.get("box_name", ""))
        self._box_comment.set(pf.get("box_comment", ""))
        self._gen_box = tk.BooleanVar(value=pf.get("gen_box", True))

        self._set_id_keep = tk.BooleanVar(value=pf.get("set_id_keep", True))
        _id_row = tk.Frame(s_set, bg=BG); _id_row.pack(fill="x", padx=8, pady=2)
        tk.Label(_id_row, text="Set ID:", width=22, anchor="w", bg=BG, fg=FG,
                 font=("Consolas", 9)).pack(side="left")
        tk.Entry(_id_row, textvariable=self._set_id, width=14, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
        tk.Checkbutton(_id_row, text="Keep Counting", variable=self._set_id_keep,
                       bg=BG, fg=FG_DIM, selectcolor=BG3, activebackground=BG,
                       font=("Consolas", 8)).pack(side="left", padx=6)
        field(s_set, "Set Name:", self._set_name)
        sep_b = tk.Frame(s_set, bg=BG4, height=1); sep_b.pack(fill="x", padx=10, pady=4)

        # Box — checkbox + inline config panel
        box_cfg_lf = tk.LabelFrame(s_set, text="  Box <ROW> config  ", bg=BG, fg=ACC1,
                                   font=("Consolas", 8, "bold"), bd=1, relief="groove")
        box_cfg_inner = tk.Frame(box_cfg_lf, bg=BG)

        def _toggle_box_cfg(*_):
            if self._gen_box.get():
                box_cfg_lf.pack(fill="x", padx=10, pady=4)
                box_cfg_inner.pack(fill="x")
            else:
                box_cfg_lf.pack_forget()

        br = tk.Frame(s_set, bg=BG); br.pack(fill="x", padx=10, pady=2)
        tk.Checkbutton(br, text="Generate Box <ROW>",
                       variable=self._gen_box, bg=BG, fg=FG, selectcolor=BG3,
                       activebackground=BG, font=("Consolas", 9),
                       command=_toggle_box_cfg).pack(side="left")

        def _bfield(lbl, var, w=28):
            r = tk.Frame(box_cfg_inner, bg=BG); r.pack(fill="x", padx=10, pady=2)
            tk.Label(r, text=lbl, width=22, anchor="w", bg=BG, fg=FG,
                     font=("Consolas", 9)).pack(side="left")
            tk.Entry(r, textvariable=var, width=w, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)

        _bfield("Box Item ID:", self._box_id)
        _box_id_keep_row = tk.Frame(box_cfg_inner, bg=BG); _box_id_keep_row.pack(fill="x", padx=8, pady=1)
        self._box_id_keep = tk.BooleanVar(value=pf.get("box_id_keep", True))
        tk.Checkbutton(_box_id_keep_row, text="Keep Counting  (auto-increment Box ID each session)",
                       variable=self._box_id_keep, bg=BG, fg=FG_DIM, selectcolor=BG3,
                       activebackground=BG, font=("Consolas", 8)).pack(side="left", padx=24)
        _bfield("Box Name:", self._box_name)
        _bfield("Box Comment:", self._box_comment)
        tk.Label(box_cfg_inner, text="  Set item IDs are auto-filled from the pieces table below.",
                 bg=BG, fg=FG_GREY, font=("Consolas", 7)).pack(anchor="w", padx=12, pady=(0, 4))

        _toggle_box_cfg()

        # ── ⑤ Generator Steps ─────────────────────────────────────────────
        # Hidden config vars (pre-fill data for each generator)
        self._partfile_shared_var   = tk.StringVar(value=pf.get("partfile_shared", ""))
        self._shopfile_var          = tk.StringVar(value=pf.get("shopfile", ""))
        self._shopbundle_var        = tk.StringVar(value=pf.get("shopbundle", "0"))
        self._fash_min_level        = tk.StringVar(value=pf.get("fash_min_level", "1"))
        self._fash_recycle_var      = tk.IntVar(value=int(pf.get("fash_recycle", 0)))
        self._fash_ticket_var       = tk.StringVar(value=pf.get("fash_ticket", "0"))
        self._fash_ncash_input_mode = tk.StringVar(value=pf.get("fash_ncash_input_mode", "tickets"))
        self._use_equip_stats       = tk.BooleanVar(value=pf.get("use_equip_stats", False))
        self._fash_equip_stats      = {}
        self._fash_opt_checks       = [tk.BooleanVar(value=False) for _ in OPTIONS_CHECKS]
        for i, (_, fv) in enumerate(OPTIONS_CHECKS):
            saved_opts = set(pf.get("fash_options_flags", []))
            self._fash_opt_checks[i].set(fv in saved_opts)
        self._exchange_cfg = pf.get("exchange_cfg", {})
        self._compound_cfg = pf.get("compound_cfg", {})
        self._shop_count          = tk.StringVar(value=pf.get("shop_count", "100"))
        self._shop_price_override = tk.StringVar(value=pf.get("shop_price_override", ""))
        self._exch_fields = {}
        self._comp_fields = {}
        self._gen_itemparam = tk.BooleanVar(value=pf.get("gen_itemparam", True))
        self._gen_cmset     = tk.BooleanVar(value=pf.get("gen_cmset", True))
        self._gen_box_step  = tk.BooleanVar(value=pf.get("gen_box_step", True))
        self._gen_set_step  = tk.BooleanVar(value=pf.get("gen_set_step", True))
        self._gen_shop      = tk.BooleanVar(value=pf.get("gen_shop", False))
        self._gen_exchange  = tk.BooleanVar(value=pf.get("gen_exchange", False))
        self._gen_compound  = tk.BooleanVar(value=pf.get("gen_compound", False))
        self._gen_box       = self._gen_box_step  # alias used by _collect_state

        # Shop/CE import source
        self._shop_src_pieces = tk.BooleanVar(value=pf.get("shop_src_pieces", True))
        self._shop_src_box    = tk.BooleanVar(value=pf.get("shop_src_box",    False))

        s_opt = sec("  ⑤ Generator Steps  (choose which generators to step through)  ")
        tk.Label(s_opt,
                 text=("  Press [Next Step] to step through each selected"
                       " generator in order, pre-filled with your session data."),
                 bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(anchor="w", padx=10, pady=(2,6))

        def _step_row(parent, text, var, sub_frame=None, sub_toggle=None):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=12, pady=2)
            cb = tk.Checkbutton(r, text=text, variable=var, bg=BG, fg=FG,
                                selectcolor=BG3, activebackground=BG, font=("Consolas", 9))
            cb.pack(side="left")
            if sub_frame is not None and sub_toggle is not None:
                def _tog(*_):
                    sub_frame.pack(fill="x", padx=32, pady=1) if var.get() else sub_frame.pack_forget()
                    if sub_toggle: sub_toggle()
                cb.config(command=_tog)
                if var.get(): sub_frame.pack(fill="x", padx=32, pady=1)
                else: sub_frame.pack_forget()
            return r

        _step_row(s_opt, "① ItemParam Generator  (one screen per fashion piece)", self._gen_itemparam)
        _step_row(s_opt, "② Set Item Generator   (CMSetItemParam.xml row)",        self._gen_set_step)
        _step_row(s_opt, "③ Box Generator        (box ItemParam + box XML rows)",  self._gen_box_step)

        # Shop/Exchange/Compound — with import source sub-option
        ce_sub = tk.Frame(s_opt, bg=BG)
        def _tog_ce():
            ce_sub.pack(fill="x", padx=32, pady=1) if (self._gen_shop.get() or self._gen_exchange.get() or self._gen_compound.get()) else ce_sub.pack_forget()

        shop_r = tk.Frame(s_opt, bg=BG); shop_r.pack(fill="x", padx=12, pady=2)
        tk.Label(shop_r, text="④ Shop / Exchange / Compound:", bg=BG, fg=FG,
                 font=("Consolas", 9), width=36, anchor="w").pack(side="left")
        for lbl2, var2 in [("R_ShopItem", self._gen_shop),
                            ("Exchange",  self._gen_exchange),
                            ("Compound",  self._gen_compound)]:
            tk.Checkbutton(shop_r, text=lbl2, variable=var2, bg=BG, fg=FG,
                           selectcolor=BG3, activebackground=BG, font=("Consolas", 9),
                           command=_tog_ce).pack(side="left", padx=6)

        tk.Label(ce_sub, text="Pre-fill IDs from:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 8)).pack(side="left")
        tk.Checkbutton(ce_sub, text="Fashion pieces (ItemParam IDs)",
                       variable=self._shop_src_pieces, bg=BG, fg=FG,
                       selectcolor=BG3, activebackground=BG, font=("Consolas", 8)).pack(side="left", padx=8)
        tk.Checkbutton(ce_sub, text="Box ID",
                       variable=self._shop_src_box, bg=BG, fg=FG,
                       selectcolor=BG3, activebackground=BG, font=("Consolas", 8)).pack(side="left", padx=4)
        _tog_ce()

        # Step state tracker — which steps remain
        self._step_queue = []

        def _build_step_queue():
            q = []
            if self._gen_itemparam.get(): q.append("itemparam")
            if self._gen_set_step.get():  q.append("set")
            if self._gen_box_step.get():  q.append("box")
            if self._gen_shop.get() or self._gen_exchange.get() or self._gen_compound.get():
                q.append("shop_ce")
            return q

        def _next_step():
            _save_settings(self._PERSIST_KEY, self._collect_state())
            q = _build_step_queue()
            if not q:
                messagebox.showinfo("Nothing selected",
                    "Select at least one generator step above.", parent=self.root)
                return
            self._do_step(q[0])

        self._next_step_fn = _next_step

        # ── ⑥ Fashion Pieces ────────────────────────────────────────────
        s_pieces = sec("  ⑥ Fashion Pieces  (max 7)  ")
        tk.Label(s_pieces,
                 text="  When a character is selected, piece names fill automatically.\n"
                      "  ChrTypeFlags and ExistType are set automatically by the program.",
                 bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(anchor="w", padx=10, pady=4)

        # Header row — character-width labels that match the entry fields below
        hdr_f = tk.Frame(s_pieces, bg=BG2)
        hdr_f.pack(fill="x", padx=10, pady=(0, 2))
        # Widths match the Entry widget widths: #=3, ItemID=13, Name=26, Comment=22
        for lbl, w in [("#", 4), ("Item ID", 14), ("Name", 27), ("Comment", 23)]:
            tk.Label(hdr_f, text=lbl, width=w, anchor="w", bg=BG2, fg=BLUE,
                     font=("Consolas", 8, "bold")).pack(side="left", padx=1)

        # Rows host — solid BG so no weird shading between rows
        self._pieces_host = tk.Frame(s_pieces, bg=BG)
        self._pieces_host.pack(fill="x", padx=10)
        self._item_row_widgets = []

        def _add_piece_row(item_id="", name="", comment=""):
            if len(self._item_row_widgets) >= MAX_FASHION_PIECES:
                messagebox.showwarning("Limit", f"Maximum {MAX_FASHION_PIECES} fashion pieces per set.")
                return
            n = len(self._item_row_widgets) + 1
            rf = tk.Frame(self._pieces_host, bg=BG)
            rf.pack(fill="x", pady=1)
            tk.Label(rf, text=str(n), width=3, bg=BG, fg=FG_GREY,
                     font=("Consolas", 8), anchor="center").pack(side="left", padx=1)
            vid  = tk.StringVar(value=item_id)
            vn   = tk.StringVar(value=name)
            vcmt = tk.StringVar(value=comment)
            tk.Entry(rf, textvariable=vid, width=13, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat",
                     highlightthickness=1, highlightbackground=BG4).pack(side="left", padx=1)
            tk.Entry(rf, textvariable=vn, width=26, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat",
                     highlightthickness=1, highlightbackground=BG4).pack(side="left", padx=1)
            tk.Entry(rf, textvariable=vcmt, width=22, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat",
                     highlightthickness=1, highlightbackground=BG4).pack(side="left", padx=1)

            def _remove(f=rf):
                f.destroy()
                self._item_row_widgets[:] = [
                    x for x in self._item_row_widgets if x[0].winfo_exists()]
                # Renumber
                for i, (fr2, *_) in enumerate(self._item_row_widgets, 1):
                    for w2 in fr2.winfo_children():
                        if isinstance(w2, tk.Label) and w2.cget("width") == 3:
                            w2.config(text=str(i)); break

            mk_btn(rf, "✗", _remove, color=BG4, fg=FG_DIM,
                   font=("Consolas", 7), padx=3, pady=0).pack(side="left", padx=2)
            self._item_row_widgets.append((rf, vid, vn, vcmt))

        self._add_piece_row_fn = _add_piece_row

        # Restore prefill items (may include _partfile per row)
        for it in pf.get("items", []):
            _add_piece_row(it.get("item_id", ""), it.get("name", ""), it.get("comment", ""))
            # Tag the last added row's ID var with _partfile if present
            if it.get("_partfile") and self._item_row_widgets:
                rf_, vid_, vn_, vcmt_ = self._item_row_widgets[-1]
                vid_._part_fn = it["_partfile"]

        # If no prefill, add empty rows matching expected piece count for the character
        if not pf.get("items"):
            pieces = _FASHION_PIECES.get(self._char_var.get(), [])
            for name in pieces[:MAX_FASHION_PIECES]:
                _add_piece_row(name=name)
            while len(self._item_row_widgets) < 3:
                _add_piece_row()

        add_row = tk.Frame(s_pieces, bg=BG); add_row.pack(anchor="w", padx=10, pady=6)
        mk_btn(add_row, "➕  Add piece", _add_piece_row, color=BG4,
               font=("Consolas", 8)).pack(side="left", padx=4)
        tk.Label(add_row, text=f"(max {MAX_FASHION_PIECES})",
                 bg=BG, fg=FG_GREY, font=("Consolas", 7)).pack(side="left")

        # Auto-fill piece names when character changes
        def _auto_fill_pieces(*_):
            char = self._char_var.get()
            pieces = _FASHION_PIECES.get(char, [])
            existing = self._item_row_widgets
            # Fill names into existing rows, add rows for extra pieces
            for i, piece_name in enumerate(pieces[:MAX_FASHION_PIECES]):
                if i < len(existing):
                    _, _, vn, _ = existing[i]
                    if not vn.get().strip():  # only fill blank names
                        vn.set(piece_name)
                else:
                    _add_piece_row(name=piece_name)

        self._auto_fill_pieces = _auto_fill_pieces
        char_menu.bind("<<ComboboxSelected>>", lambda e: (_update_char_display(), _auto_fill_pieces()))

        # Initial display update
        _update_char_display()

        # ── Bottom nav ──────────────────────────────────────────────────
        bot = tk.Frame(self, bg=BG2); bot.pack(fill="x", side="bottom")
        mk_btn(bot, "⚡  Generate All Output", self._generate_all,
               color=self.ACC, fg=BG2,
               font=("Consolas", 12, "bold")).pack(side="right", padx=16, pady=8)
        mk_btn(bot, "💾  Save Session", self._save_session,
               color=GREEN, fg=BG2,
               font=("Consolas", 9, "bold")).pack(side="right", padx=4, pady=8)

        def _import_session_bottom():
            saved = _load_settings(self._PERSIST_KEY)
            if not saved.get("items"):
                messagebox.showinfo("No Session", "No saved fashion session on disk."); return
            if messagebox.askyesno("Load Session",
                    f"Load saved session?\n({len(saved.get('items',[]))} items — "
                    f"{saved.get('set_name','unnamed')})"):
                self._build_editor(prefill=saved)

        mk_btn(bot, "⬇  Import Session", _import_session_bottom,
               color=BG3, font=("Consolas", 8)).pack(side="right", padx=4, pady=8)
        mk_btn(bot, "🔄  Reset", self._build_start_screen,
               color=BG4).pack(side="left", padx=10, pady=8)

    # ─────────────────────────────────────────────────────────────────────────
    # IMPORT
    # ─────────────────────────────────────────────────────────────────────────
    def _do_step(self, step_name):
        """Execute one wizard step: pre-fill the target generator and switch to it."""
        # Collect common data
        char     = self._char_var.get()
        chr_flag = CHR_FLAG_MAP.get(char, 0)
        color    = getattr(self, "_color_name_var", tk.StringVar()).get().strip()
        fn       = _sanitise_xml_path(self._filename_var.get().strip())
        bn       = self._bundlenum_var.get().strip() or "0"
        cmt      = _sanitise_xml_path(self._cmtfilename_var.get().strip())
        cbn      = self._cmtbundlenum_var.get().strip() or "0"
        # Resolve %placeholders% in filenames
        ph_vars  = {k: v.get() for k, v in self._ph_vars.items()}
        def _ph(s):
            for k, v in ph_vars.items():
                s = s.replace(f"%{k}%", v)
            return s
        fn  = _sanitise_xml_path(_ph(fn))
        cmt = _sanitise_xml_path(_ph(cmt))

        # Pieces
        pieces = []
        for rf, vid, vn, vcmt in self._item_row_widgets:
            if rf.winfo_exists() and vn.get().strip():
                pfn = getattr(vid, "_part_fn",
                              self._partfile_shared_var.get().strip())
                pieces.append({"item_id":  vid.get().strip(),
                                "name":    vn.get().strip(),
                                "comment": vcmt.get().strip(),
                                "partfn":  _sanitise_xml_path(_ph(pfn))})

        # NCash / recycling
        recycle = self._fash_recycle_var.get()
        ncash_val = "0"
        if recycle == 262144:
            try:
                tv = float(self._fash_ticket_var.get() or "0")
                ncash_val = str(round(tv*133) if self._fash_ncash_input_mode.get()=="tickets" else int(tv))
            except: pass
        elif getattr(self, "_use_ncash", tk.BooleanVar()).get():
            try:
                tv = float(getattr(self, "_ncash_val", tk.StringVar(value="0")).get() or "0")
                ncash_val = str(round(tv*133) if getattr(self, "_ncash_mode", tk.StringVar(value="tickets")).get()=="tickets" else int(tv))
            except: pass
        galder = (getattr(self, "_galder_val", tk.StringVar(value="0")).get()
                  if getattr(self, "_use_galder", tk.BooleanVar()).get() else "0")

        # Options string
        opt_flags = [str(fv) for (_, fv), v in zip(OPTIONS_CHECKS, self._fash_opt_checks) if v.get()]
        base_opts = ["2", "16"]
        opt_str = "/".join(base_opts + opt_flags)

        if step_name == "itemparam":
            if not pieces:
                messagebox.showwarning("No Pieces",
                    "Add fashion pieces in ⑥ Fashion Pieces first.", parent=self.root)
                return
            p0 = pieces[0]
            cls2, typ2, sub2 = _fashion_detect_slot(p0["name"])
            last_id = int(_get_last_id("t18_fashion_item", 0) or 0)
            cfg = {
                "id":                  str(last_id + 1) if last_id else "",
                "class_val":           str(cls2),
                "type_val":            str(typ2),
                "subtype_val":         str(sub2),
                "itemftype_val":       "0",
                "name":                _apply_color(p0["name"], color),
                "comment":             _apply_color(p0.get("comment",""), color),
                "file_name":           fn,
                "bundle_num":          bn,
                "cmt_file_name":       cmt,
                "cmt_bundle_num":      cbn,
                "equip_file_name":     "",
                "chr_type_flags":      chr_flag,
                "exist_type":          "0",
                "part_file_name":      p0.get("partfn",""),
                "shop_file_name":      _sanitise_xml_path(self._shopfile_var.get().strip()),
                "shop_bundle_num":     self._shopbundle_var.get().strip() or "0",
                "min_level":           self._fash_min_level.get().strip() or "1",
                "options_raw_manual":  opt_str,
                "value":               galder,
                "ncash":               ncash_val,
                "_fashion_all_pieces": pieces,
                "_fashion_color":      color,
                "_fashion_char":       char,
                "_fashion_chr_flag":   chr_flag,
                "_fashion_fn":         fn,
                "_fashion_bn":         bn,
                "_fashion_cmt":        cmt,
                "_fashion_cbn":        cbn,
                "_fashion_recycle":    recycle,
                "_fashion_ncash":      ncash_val,
                "_fashion_galder":     galder,
                "_fashion_shopfile":   _sanitise_xml_path(self._shopfile_var.get().strip()),
                "_fashion_shopbn":     self._shopbundle_var.get().strip() or "0",
                "_fashion_minlevel":   self._fash_min_level.get().strip() or "1",
                "_fashion_opt_str":    opt_str,
            }
            _save_t6_settings(cfg)
            try:
                for i, (_, _, _, cls_) in enumerate(TOOLS):
                    if cls_ is Tool6:
                        self.root._switch_tool(i)
                        tool = self.root._tool_instances.get(i)
                        if tool:
                            tool._settings   = cfg
                            tool._first_run  = False
                            tool._build_editor()
                        return
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=self.root)

        elif step_name == "set":
            set_id   = self._set_id.get().strip()
            set_name = self._set_name.get().strip()
            if not set_id:
                messagebox.showwarning("No Set ID",
                    "Enter a Set ID in ④ Set & Box first.", parent=self.root)
                return
            pre = {"set_id": set_id, "set_name": set_name,
                   "last_id": set_id}
            for i, p in enumerate(pieces[:8]):
                pre[f"item{i}"] = p.get("item_id","")
            _save_settings("t8_set", pre)
            try:
                for i, (_, _, _, cls_) in enumerate(TOOLS):
                    if cls_ is Tool8:
                        self.root._switch_tool(i)
                        tool = self.root._tool_instances.get(i)
                        if tool:
                            tool._build_manual_editor(prefill=pre)
                        return
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=self.root)

        elif step_name == "box":
            box_id = self._box_id.get().strip()
            if not box_id:
                messagebox.showwarning("No Box ID",
                    "Enter a Box ID in ④ Set & Box first.", parent=self.root)
                return
            pre = {
                "id":         box_id,
                "box_name":   self._box_name.get().strip(),
                "comment":    self._box_comment.get().strip(),
                "file_name":  fn,
                "bundle_num": bn,
                "items":      [{"id": p.get("item_id",""), "name": p["name"]} for p in pieces],
            }
            _save_settings("t1_box", pre)
            try:
                for i, (_, _, _, cls_) in enumerate(TOOLS):
                    if cls_ is Tool1:
                        self.root._switch_tool(i)
                        tool = self.root._tool_instances.get(i)
                        if tool:
                            tool.box_configs = [{"id": box_id,
                                                  "box_name": pre["box_name"],
                                                  "comment":  pre["comment"],
                                                  "file_name": fn,
                                                  "bundle_num": bn,
                                                  "drop_items": [p.get("item_id","")
                                                                 for p in pieces]}]
                            tool._build_config_screen()
                        return
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=self.root)

        elif step_name == "shop_ce":
            # Build the ID list from selected sources
            ids = []
            if self._shop_src_pieces.get():
                ids += [p.get("item_id","") for p in pieces if p.get("item_id","")]
            if self._shop_src_box.get():
                box_id = self._box_id.get().strip()
                if box_id: ids.append(box_id)
            if not ids:
                messagebox.showwarning("No IDs",
                    "Enable Fashion pieces or Box ID under Shop/Exchange/Compound source,"
                    " and make sure those IDs are filled in.", parent=self.root)
                return
            try:
                for i, (_, _, _, cls_) in enumerate(TOOLS):
                    if cls_ is Tool7:
                        self.root._switch_tool(i)
                        tool = self.root._tool_instances.get(i)
                        if tool and hasattr(tool, "_prefill_ids"):
                            tool._prefill_ids(ids,
                                              do_shop=self._gen_shop.get(),
                                              do_exchange=self._gen_exchange.get(),
                                              do_compound=self._gen_compound.get())
                        return
            except Exception as e:
                messagebox.showerror("Error", str(e), parent=self.root)

    def _import_fashion_file(self):
        path = filedialog.askopenfilename(
            filetypes=[("Spreadsheet", "*.csv *.xlsx *.xlsm *.xls"), ("All", "*.*")],
            parent=self.root)
        if not path: return
        ext = os.path.splitext(path)[1].lower()
        sections = {}
        try:
            if ext in (".xlsx", ".xlsm", ".xls"):
                if not _HAVE_OPENPYXL:
                    messagebox.showerror("Missing library", "pip install openpyxl"); return
                wb = openpyxl.load_workbook(path, data_only=True)
                for sh_name in wb.sheetnames:
                    for k, v in self._parse_fashion_csv(_sheet_to_csv(wb[sh_name])).items():
                        sections.setdefault(k, []).extend(v)
            else:
                with open(path, encoding="utf-8-sig") as f:
                    sections = self._parse_fashion_csv(f.read())
        except Exception as e:
            messagebox.showerror("Import Error", str(e)); return
        pf = self._sections_to_prefill(sections)
        self._import_status.set(
            f"✓  Imported from {os.path.basename(path)}: "
            f"{len(pf.get('items',[]))} items")
        self._build_editor(prefill=pf)

    def _parse_fashion_csv(self, csv_text):
        reader = csv.DictReader(io.StringIO(csv_text))
        headers = reader.fieldnames or []
        section = _detect_fashion_section(headers)
        nh_map  = {h: _norm_fashion_hdr(h) for h in headers}
        result  = {section: []}
        for raw in reader:
            cl = {nh_map.get(k, k): (v or "").strip() for k, v in raw.items()}
            if section == "item":
                iid = next((cl[h] for h in cl if h in _FASH_ITEM_KEYS), "")
                result["item"].append({"item_id": iid,
                                       "name":    cl.get("name", ""),
                                       "comment": cl.get("comment", "")})
            elif section == "box":
                bid = next((cl[h] for h in cl if h in _FASH_BOX_KEYS), "")
                result["box"].append({"box_id": bid,
                                      "box_name": cl.get("boxname", cl.get("nameofbox", ""))})
            elif section == "set":
                sid = next((cl[h] for h in cl if h in _FASH_SET_KEYS), "")
                result["set"].append({"set_id": sid,
                                      "set_name": cl.get("setname", cl.get("name", ""))})
        return result

    def _sections_to_prefill(self, sections):
        pf = {}
        pf["items"] = sections.get("item", [])
        if sections.get("set"):
            pf["set_id"]   = sections["set"][0].get("set_id", "")
            pf["set_name"] = sections["set"][0].get("set_name", "")
        if sections.get("box"):
            pf["box_id"]   = sections["box"][0].get("box_id", "")
            pf["box_name"] = sections["box"][0].get("box_name", "")
        return pf

    def _parse_labels_txt(self, raw):
        """Parse a fashion labels .txt file (e.g. Bunny_2nd_Job_Labels.txt).

        Understands:
          BN2 = Char label          → prefix used in file paths
          Bunny = Character Type    → race
          Boxer = Job               → job tier (determines ChrTypeFlag)
          %color%                   → variable block: CODE = Full Name
          %part%                    → variable block: Code = Piece Name
          chr009_w05  Robe (White)  → PartFileName Code  Piece Name (Color)
          720119 = Test Box ID      → box_id
          740076 = Starting Set ID  → set_id
        """
        import re as _re

        result = {
            "char_label": "",   # e.g. BN2
            "char_type":  "",   # e.g. Bunny
            "job":        "",   # e.g. Boxer / 2nd
            "colors":     {},   # code → full name  e.g. WHT→White
            "parts":      {},   # code → piece name e.g. Robe→Robe
            "partfiles":  [],   # list of {partfile, piece_name, color_name}
            "box_id":     "",
            "set_id":     "",
            "variables":  {},   # all %var% → {code: name}
        }
        current_var_block = None
        lines = raw.splitlines()
        for raw_line in lines:
            line = raw_line.strip()
            if not line: continue

            # %varname% block header
            m_var = _re.match(r'^%(\w+)%\s*$', line)
            if m_var:
                current_var_block = m_var.group(1).lower()
                result["variables"].setdefault(current_var_block, {})
                continue

            # key = value lines
            if "=" in line:
                k, _, v = line.partition("=")
                k = k.strip(); v = v.strip()
                kl = k.lower().replace(" ", "").replace("_", "")

                # Inside a variable block
                if current_var_block:
                    result["variables"][current_var_block][k] = v
                    if current_var_block in ("color", "colours", "colors"):
                        result["colors"][k] = v
                    elif current_var_block == "part":
                        result["parts"][k] = v
                    continue

                # Box ID
                if "box" in kl or v.lower().replace(" ","").startswith("boxid") or "boxid" in kl or "testbox" in v.lower().replace(" ",""):
                    try: int(k.strip()); result["box_id"] = k.strip(); continue
                    except: pass
                if "box" in v.lower() and k.strip().isdigit():
                    result["box_id"] = k.strip(); continue
                if "box" in kl and k.strip().isdigit():
                    result["box_id"] = k.strip(); continue

                # Set ID
                if "set" in kl or "setid" in kl or "set" in v.lower():
                    if k.strip().isdigit():
                        result["set_id"] = k.strip(); continue

                # Detect numeric key with descriptive value
                if k.strip().isdigit():
                    vl = v.lower()
                    if "box" in vl: result["box_id"] = k.strip()
                    elif "set" in vl: result["set_id"] = k.strip()
                    continue

                # Char label line: BN2 = Char label
                if "charlabel" in v.lower().replace(" ","") or "prefix" in v.lower() or "charlabel" in kl:
                    result["char_label"] = k.strip(); continue
                if "charlabel" in kl:
                    result["char_label"] = v.strip(); continue

                # Character type: Bunny = Character Type
                if "charactertype" in v.lower().replace(" ","") or "chartype" in v.lower().replace(" ",""):
                    result["char_type"] = k.strip(); continue

                # Job: Boxer = Job
                if v.strip().lower() == "job" or "job" in v.lower().replace(" ",""):
                    result["job"] = k.strip(); continue

                # Fallback: treat as variable definition if inside block
                continue

            # PartFileName lines — any filename token, separated by tab OR " - " (space-dash-space)
            # Accepted: "chr009_w05\tRobe (White)"  "myfile_v2\tRobe (White)"  "myfile_v2 - Robe (White)"
            # Split on TAB first (unambiguous), then on " - " (space-dash-space, avoids splitting hyphenated names)
            m_pf = None
            if "\t" in line:
                parts_split = line.split("\t", 1)
                if len(parts_split) == 2 and parts_split[0].strip() and parts_split[1].strip():
                    m_pf = (parts_split[0].strip(), parts_split[1].strip())
            if not m_pf and " - " in line:
                idx_dash = line.index(" - ")
                lhs = line[:idx_dash].strip()
                rhs = line[idx_dash+3:].strip()
                if lhs and rhs and " " not in lhs:  # LHS must be a single token (no spaces)
                    m_pf = (lhs, rhs)
            if m_pf:
                partfile = m_pf[0]
                desc     = m_pf[1]
                # Extract piece name and color from "Robe (White)" or "Robe"
                m_pc = _re.match(r'^(.+?)\s*\((.+?)\)\s*$', desc)
                if m_pc:
                    piece_name = m_pc.group(1).strip()
                    color_name = m_pc.group(2).strip()
                else:
                    piece_name = desc
                    color_name = ""
                result["partfiles"].append({
                    "partfile":   partfile,
                    "piece_name": piece_name,
                    "color_name": color_name,
                })
                continue

            # Bare lines: if it resolves as a job via resolve_chr_flag, treat it as the job
            if not result["job"]:
                if resolve_chr_flag(line.strip()) is not None:
                    result["job"] = line.strip()

        return result

    def _import_txt_paste(self):
        win = tk.Toplevel(self.root)
        win.title("Paste Fashion TXT"); win.configure(bg=BG); win.geometry("920x620")
        win.grab_set()

        # ── Reference guide (left panel) ─────────────────────────────────
        panes = tk.PanedWindow(win, orient="horizontal", bg=BG, sashrelief="flat",
                               sashwidth=4)
        panes.pack(fill="both", expand=True, padx=0, pady=0)

        ref_frame = tk.Frame(panes, bg=BG2, width=340); ref_frame.pack_propagate(False)
        panes.add(ref_frame, minsize=280)

        tk.Label(ref_frame, text="📖  Format Reference",
                 bg=BG2, fg=BLUE, font=("Consolas", 10, "bold"), pady=6).pack(anchor="w", padx=10)
        ref_txt = scrolledtext.ScrolledText(ref_frame, font=("Consolas", 8), bg=BG2, fg=FG,
                                            wrap="word", state="normal")
        ref_txt.pack(fill="both", expand=True, padx=4, pady=4)
        REF = (
            "ACCEPTED FORMAT GUIDE\n"
            "═══════════════════════════════════════\n\n"
            "HEADER (any order, optional labels):\n"
            "  BN2 = Char label        ← prefix / char label\n"
            "  Bunny = Character Type  ← race name\n"
            "  Boxer = Job             ← job (sets ChrTypeFlag)\n\n"
            "IDS:\n"
            "  720119 = Test Box ID    ← box_id\n"
            "  740076 = Starting Set ID ← set_id\n\n"
            "VARIABLE BLOCKS  (used in path templates):\n"
            "  %color%\n"
            "  WHT = White             ← code = full name\n"
            "  BLK = Black\n"
            "  ...\n\n"
            "  %part%\n"
            "  Robe = Robe\n"
            "  BoxGloves = Boxing Gloves\n"
            "  ...\n\n"
            "PART FILE ENTRIES  (tab OR space-dash-space):\n"
            "  chr009_w05\tRobe (White)\n"
            "  chr009_w05 - Robe (White)\n"
            "  myitem_v2\tHoodie (Red)\n"
            "  any_filename - Piece Name (Color)\n"
            "  ^ PartFileName  ^ Piece (Color)\n\n"
            "COLOR GROUPS:\n"
            "  One blank line between each color group.\n"
            "  The program detects each group as one\n"
            "  complete fashion set per color.\n\n"
            "PATH TEMPLATE EXAMPLE:\n"
            "  data\\wear_parts\\BN2%color%%part%.pal\n"
            "  → BN2 = char label, %color% = WHT/BLK…\n"
            "    %part% = Robe/BoxGloves…\n\n"
            "NOTES:\n"
            "  • Job name determines ChrTypeFlag\n"
            "  • Character Type is optional\n"
            "  • Char label = prefix for file names\n"
            "  • Part entries group by blank lines\n"
        )
        ref_txt.insert("1.0", REF)
        ref_txt.config(state="disabled")

        # ── Input (right panel) ───────────────────────────────────────────
        input_frame = tk.Frame(panes, bg=BG)
        panes.add(input_frame, minsize=380)

        tk.Label(input_frame, text="Paste or load fashion definition text:",
                 bg=BG, fg=FG, font=("Consolas", 9)).pack(anchor="w", padx=10, pady=(8,2))

        txt = scrolledtext.ScrolledText(input_frame, font=("Consolas", 9), bg=BG3, fg=FG)
        txt.pack(fill="both", expand=True, padx=8, pady=4)

        status_lbl = tk.Label(input_frame, text="", bg=BG, fg=GREEN, font=("Consolas", 8))
        status_lbl.pack(anchor="w", padx=10)

        def _load_file():
            path = filedialog.askopenfilename(
                filetypes=[("Text files", "*.txt"), ("All", "*.*")], parent=win)
            if not path: return
            try:
                with open(path, encoding="utf-8-sig") as f:
                    txt.delete("1.0", "end")
                    txt.insert("1.0", f.read())
                status_lbl.config(text=f"✓ Loaded: {os.path.basename(path)}")
            except Exception as e:
                messagebox.showerror("Load Error", str(e), parent=win)

        def _parse():
            raw = txt.get("1.0", "end")
            parsed = self._parse_labels_txt(raw)

            # Build prefill from parsed data
            pf = {}

            # Determine character from job using the existing resolve_chr_flag + _JOB_NAME_MAP
            # This covers all in-game job names: Boxer, Librarian, Shaman, Fighter, etc.
            job_str   = parsed.get("job", "").strip()
            char_type = parsed.get("char_type", "").strip()

            char_found = None

            # 1. Try job name directly via resolve_chr_flag (handles "Boxer", "Librarian", etc.)
            if job_str:
                flag = resolve_chr_flag(job_str)
                if flag is not None:
                    # Reverse-look up the _CHAR_DATA key from the flag
                    for key in _CHAR_DATA:
                        if CHR_FLAG_MAP.get(key) == flag:
                            char_found = key; break

            # 2. If job alone didn't resolve, try combining char_type + job tier
            if not char_found and char_type and job_str:
                import re as _re2
                tier = None
                for t in ("1st","2nd","3rd","first","second","third"):
                    if t in job_str.lower():
                        tier = {"first":"1st","second":"2nd","third":"3rd"}.get(t, t)
                        break
                if tier:
                    combined = f"{char_type} {tier}"
                    flag = resolve_chr_flag(combined)
                    if flag is not None:
                        for key in _CHAR_DATA:
                            if CHR_FLAG_MAP.get(key) == flag:
                                char_found = key; break

            # 3. Last resort: fuzzy match char_type against _CHAR_DATA keys
            if not char_found and char_type:
                ct_norm = char_type.lower()
                for key in _CHAR_DATA:
                    if ct_norm in key.lower():
                        char_found = key; break

            if char_found:
                pf["char"] = char_found

            # Box and set IDs
            if parsed["box_id"]: pf["box_id"] = parsed["box_id"]
            if parsed["set_id"]: pf["set_id"] = parsed["set_id"]

            # Variables for path templates
            pf["_parsed_vars"]    = parsed["variables"]
            pf["_parsed_colors"]  = parsed["colors"]
            pf["_parsed_parts"]   = parsed["parts"]
            pf["_parsed_prefix"]  = parsed["char_label"]
            pf["_partfiles"]      = parsed["partfiles"]

            # Group partfiles by color → one set per color
            # Each group: same piece types, different color
            colors = parsed["colors"]  # code→name e.g. WHT→White
            partfiles = parsed["partfiles"]

            if colors and partfiles:
                # Group by color_name
                from collections import OrderedDict
                groups_by_color = OrderedDict()
                for pfe in partfiles:
                    cn = pfe["color_name"]
                    if cn not in groups_by_color:
                        groups_by_color[cn] = []
                    groups_by_color[cn].append(pfe)

                # Use first color group as the item list for initial prefill
                first_color = next(iter(groups_by_color)) if groups_by_color else ""
                first_group = groups_by_color.get(first_color, [])
                pf["items"] = [
                    {"item_id": "", "name": f"{p['color_name']} {p['piece_name']}".strip(),
                     "comment": "", "_partfile": f"data\\wear_parts\\{p['partfile']}.pal"}
                    for p in first_group
                ]
                pf["color_name"] = first_color
                pf["_color_groups"] = {k: [
                    {"item_id": "", "name": f"{p['color_name']} {p['piece_name']}".strip(),
                     "comment": "", "_partfile": f"data\\wear_parts\\{p['partfile']}.pal"}
                    for p in v
                ] for k, v in groups_by_color.items()}
                status_lbl.config(text=f"✓  Parsed: {len(partfiles)} pieces · "
                                       f"{len(groups_by_color)} color group(s) · "
                                       f"char={pf.get('char','?')}")
            else:
                # Plain item list (no color grouping)
                pf["items"] = [
                    {"item_id": "", "name": p["piece_name"], "comment": "",
                     "_partfile": f"data\\wear_parts\\{p['partfile']}.pal"}
                    for p in partfiles
                ]
                status_lbl.config(text=f"✓  Parsed TXT: {len(pf['items'])} pieces")

            win.destroy()
            self._build_editor(prefill=pf)

        btn_row = tk.Frame(input_frame, bg=BG); btn_row.pack(anchor="w", padx=8, pady=4)
        mk_btn(btn_row, "📂  Load File", _load_file, color=BG3,
               font=("Consolas", 9)).pack(side="left", padx=4)

        nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")
        mk_btn(nav, "✓  Parse & Apply", _parse, color=GREEN, fg=BG2,
               font=("Consolas", 10, "bold")).pack(side="right", padx=12, pady=6)
        mk_btn(nav, "Cancel", win.destroy, color=BG4).pack(side="left", padx=8, pady=6)

    # ─────────────────────────────────────────────────────────────────────────
    # SESSION PERSISTENCE
    # ─────────────────────────────────────────────────────────────────────────
    def _generate_all(self):
        """Step through generators in order via the step queue built in _build_editor."""
        _save_settings(self._PERSIST_KEY, self._collect_state())
        fn = getattr(self, "_next_step_fn", None)
        if fn:
            fn()
        else:
            # Fallback if called before editor is built
            self._do_step("itemparam")


    def _collect_state(self):
        """Collect current editor state into a serialisable dict."""
        items = []
        for rf, vid, vn, vcmt in self._item_row_widgets:
            if rf.winfo_exists() and vid.get().strip():
                items.append({"item_id": vid.get().strip(),
                              "name":    vn.get().strip(),
                              "comment": vcmt.get().strip()})
        exch_cfg = {k: v.get() for k, v in self._exch_fields.items()} if self._exch_fields else {}
        comp_cfg = {k: v.get() for k, v in self._comp_fields.items()} if self._comp_fields else {}
        return {
            "char":          self._char_var.get(),
            "price_mode":    self._price_mode.get(),   # legacy
            "price_val":     self._price_val.get(),      # legacy alias for galder
            "cash_price":    self._cash_price.get(),     # legacy alias for myshop
            "use_galder":    getattr(self, "_use_galder",   tk.BooleanVar()).get(),
            "galder_val":    getattr(self, "_galder_val",   tk.StringVar()).get(),
            "use_ncash":     getattr(self, "_use_ncash",    tk.BooleanVar()).get(),
            "ncash_val":     getattr(self, "_ncash_val",    tk.StringVar()).get(),
            "use_myshop":    getattr(self, "_use_myshop",   tk.BooleanVar()).get(),
            "myshop_price":  getattr(self, "_myshop_price", tk.StringVar()).get(),
            "goods_list_code_start": getattr(self, "_goods_list_code_start", tk.StringVar(value="21000")).get(),
            "myshop_price_item":     getattr(self, "_myshop_price_item", tk.StringVar(value="12000")).get(),
            "filename":      self._filename_var.get(),
            "bundlenum":     self._bundlenum_var.get(),
            "cmtfilename":   self._cmtfilename_var.get(),
            "cmtbundlenum":  self._cmtbundlenum_var.get(),
            "name_prefix":   getattr(self, '_name_prefix_var', tk.StringVar()).get(),
            "name_suffix":   getattr(self, '_name_suffix_var', tk.StringVar()).get(),
            "name_pos":      getattr(self, '_name_pos_var',    tk.StringVar(value="before")).get(),
            "set_id":        self._set_id.get(),
            "set_id_keep":   getattr(self, "_set_id_keep",  tk.BooleanVar(value=True)).get(),
            "box_id_keep":   getattr(self, "_box_id_keep",  tk.BooleanVar(value=True)).get(),
            "set_name":      self._set_name.get(),
            "box_id":        self._box_id.get(),
            "box_name":      self._box_name.get(),
            "box_comment":   self._box_comment.get(),
            "gen_box":       self._gen_box.get(),
            "gen_shop":         self._gen_shop.get(),
            "gen_exchange":     self._gen_exchange.get(),
            "gen_compound":     self._gen_compound.get(),
            "gen_box_step":     getattr(self,"_gen_box_step",   tk.BooleanVar(value=True)).get(),
            "gen_set_step":     getattr(self,"_gen_set_step",   tk.BooleanVar(value=True)).get(),
            "shop_src_pieces":  getattr(self,"_shop_src_pieces",tk.BooleanVar(value=True)).get(),
            "shop_src_box":     getattr(self,"_shop_src_box",   tk.BooleanVar(value=False)).get(),
            "shop_count":    self._shop_count.get(),
            "shop_price_override": self._shop_price_override.get(),
            "exchange_cfg":    exch_cfg,
            "compound_cfg":   comp_cfg,
            "items":          items,
            "color_name":     getattr(self, "_color_name_var",     tk.StringVar()).get(),
            "path_vars":      {k: v.get() for k,v in getattr(self, "_ph_vars", {}).items() if k},

            "gen_itemparam":  getattr(self, "_gen_itemparam",      tk.BooleanVar()).get(),
            "gen_cmset":      getattr(self, "_gen_cmset",          tk.BooleanVar(value=True)).get(),
            "partfile_shared":getattr(self, "_partfile_shared_var",tk.StringVar()).get(),
            "shopfile":       getattr(self, "_shopfile_var",       tk.StringVar()).get(),
            "shopbundle":     getattr(self, "_shopbundle_var",     tk.StringVar()).get(),
            "fash_min_level": getattr(self, "_fash_min_level",     tk.StringVar(value="1")).get(),
            "use_equip_stats":getattr(self, "_use_equip_stats",    tk.BooleanVar()).get(),
            "fash_recycle":   getattr(self, "_fash_recycle_var",   tk.IntVar()).get(),
            "do_itemparam":   getattr(self, "_do_itemparam", tk.BooleanVar(value=True)).get(),
            "do_set":         getattr(self, "_do_set",       tk.BooleanVar(value=True)).get(),
            "do_box":         getattr(self, "_do_box",       tk.BooleanVar(value=True)).get(),
            "do_shop_ce":     getattr(self, "_do_shop_ce",   tk.BooleanVar(value=False)).get(),

        }

    def _save_session(self):
        state = self._collect_state()
        _save_settings(self._PERSIST_KEY, state)
        if getattr(self, "_set_id_keep", tk.BooleanVar(value=True)).get():
            try: _set_last_id("t18_set_id", int(self._set_id.get()))
            except: pass
        if getattr(self, "_box_id_keep", tk.BooleanVar(value=True)).get():
            try: _set_last_id("t18_box_id", int(self._box_id.get()))
            except: pass
        messagebox.showinfo("Session Saved",
            f"Fashion session saved to disk.\n{len(state['items'])} items — {state['set_name'] or 'unnamed'}")

    # ─────────────────────────────────────────────────────────────────────────
    # GENERATION
    # ─────────────────────────────────────────────────────────────────────────
    def _generate(self):
        items_raw = []
        for rf, vid, vn, vcmt in self._item_row_widgets:
            if rf.winfo_exists() and vid.get().strip():
                items_raw.append({
                    "item_id": vid.get().strip(),
                    "name":    vn.get().strip(),
                    "comment": vcmt.get().strip(),
                })

        if not items_raw:
            messagebox.showwarning("No Items",
                "Enter at least one Item ID in the Fashion Pieces table."); return

        # ── Color name substitution ───────────────────────────────────────
        color_name = getattr(self, "_color_name_var", tk.StringVar()).get().strip()

        # Apply name prefix / suffix
        prefix = getattr(self, "_name_prefix_var", tk.StringVar()).get()
        suffix = getattr(self, "_name_suffix_var", tk.StringVar()).get()
        pos    = getattr(self, "_name_pos_var",    tk.StringVar(value="before")).get()
        items  = []
        for it in items_raw:
            n   = _apply_color(it["name"],    color_name)
            cmt = _apply_color(it["comment"], color_name)
            if pos == "before" and prefix:   n = prefix + n
            elif pos == "after" and suffix:  n = n + suffix
            elif pos == "both":              n = prefix + n + suffix
            items.append({**it, "name": n, "comment": cmt})

        char_name = self._char_var.get()
        cat1, sex, char_sql_type, lib_cat1 = _CHAR_DATA.get(char_name, (1, 2, 8, 128))
        chr_flag = CHR_FLAG_MAP.get(char_name, 0)

        # Pricing (new independent checkboxes)
        use_galder   = getattr(self, "_use_galder",   tk.BooleanVar()).get()
        galder_val   = getattr(self, "_galder_val",   tk.StringVar(value="0")).get().strip() or "0"
        use_ncash    = getattr(self, "_use_ncash",    tk.BooleanVar()).get()
        ncash_val    = getattr(self, "_ncash_val",    tk.StringVar(value="0")).get().strip() or "0"
        # Recycling NCash from ticket calc if Recyclable is selected
        _fash_recycle = getattr(self, "_fash_recycle_var", tk.IntVar(value=0)).get()
        if _fash_recycle == 262144:
            _fash_tick_var = getattr(self, "_fash_ticket_var", tk.StringVar(value="0"))
            _fash_mode = getattr(self, "_fash_ncash_input_mode", tk.StringVar(value="tickets"))
            try:
                _tv = float(_fash_tick_var.get() or "0")
                ncash_val = str(round(_tv * 133) if _fash_mode.get() == "tickets" else int(_tv))
            except: pass
            opt_recycle = _fash_recycle
        use_myshop   = getattr(self, "_use_myshop",   tk.BooleanVar()).get()
        myshop_price = getattr(self, "_myshop_price", tk.StringVar(value="0")).get().strip() or "0"

        set_id   = self._set_id.get().strip()
        set_name = self._set_name.get().strip()

        # Paths — sanitise to single backslashes, never \\
        fn     = _sanitise_xml_path(self._filename_var.get().strip())
        bn     = self._bundlenum_var.get().strip() or "0"
        cmt_fn = _sanitise_xml_path(self._cmtfilename_var.get().strip())
        cmt_bn = self._cmtbundlenum_var.get().strip() or "0"
        inv_fn = fn   # always mirrors FileName
        inv_bn = bn   # always mirrors BundleNum

        box_id = self._box_id.get().strip() if hasattr(self, "_box_id") else ""

        # Recycling / Options
        opt_recycle = getattr(self, "_fash_recycle_var", tk.IntVar(value=0)).get()
        opt_checks  = getattr(self, "_fash_opt_checks", [tk.BooleanVar() for _ in OPTIONS_CHECKS])
        gen_ip      = getattr(self, "_gen_itemparam",  tk.BooleanVar(value=False)).get()
        gen_set     = getattr(self, "_gen_cmset",      tk.BooleanVar(value=True)).get()

        # ── Error if nothing selected ────────────────────────────────────
        gen_box      = self._gen_box.get() if hasattr(self, "_gen_box") else False
        gen_shop     = self._gen_shop.get()
        gen_exchange = self._gen_exchange.get()
        gen_compound = self._gen_compound.get()
        if not any([gen_ip, gen_set, gen_box, gen_shop, gen_exchange, gen_compound]):
            messagebox.showerror("Nothing Selected",
                "Please enable at least one output:\n"
                "ItemParam Rows, Set, Box, R_ShopItem, Exchange, or Compound.")
            return

        # ── 1. Fashion ItemParam rows (full 65-field XML) ─────────────────
        itemparam_rows = []
        recycle_rows   = []
        if gen_ip:
            options_str = build_options_str(
                [v.get() for v in opt_checks], opt_recycle)
            # Last ID tracking — starts from last generated fashion item +1
            last_fash_id = _get_last_id("t18_fashion_item", 0)
            for i, it in enumerate(items):
                iid = it["item_id"].strip()
                # Auto-increment if ID is blank
                if not iid:
                    last_fash_id += 1
                    iid = str(last_fash_id)
                else:
                    try: last_fash_id = max(last_fash_id, int(iid))
                    except: pass

                # Detect Class/Type/SubType from piece name (override from row)
                per_slot = getattr(self, "_piece_slot_overrides", {})
                slot_override = per_slot.get(i)
                if slot_override:
                    cls, typ, sub = slot_override
                else:
                    cls, typ, sub = _fashion_detect_slot(it["name"])

                # PartFileName from piece row if given, else from shared field
                part_fn_raw = ""
                pw = self._item_row_widgets
                if i < len(pw):
                    rf_, vid_, vn_, vcmt_ = pw[i]
                    # check if there are per-row part fields
                    part_fn_raw = getattr(vid_, "_part_fn", "")
                if not part_fn_raw:
                    part_fn_raw = getattr(self, "_partfile_shared_var", tk.StringVar()).get().strip()
                part_fn = _sanitise_xml_path(part_fn_raw)

                # Shop file per-piece
                shop_fn_raw = getattr(self, "_shopfile_var", tk.StringVar()).get().strip()
                shop_fn     = _sanitise_xml_path(shop_fn_raw)
                shop_bn_raw = getattr(self, "_shopbundle_var", tk.StringVar(value="0")).get().strip() or "0"

                value_field  = galder_val  if use_galder else "0"
                ncash_field  = ncash_val   if use_ncash  else "0"
                min_level    = "1"
                if hasattr(self, "_fash_min_level"):
                    min_level = self._fash_min_level.get().strip() or "1"

                # Equip stats
                es = getattr(self, "_fash_equip_stats", {})
                def _es(k): return es.get(k, tk.StringVar(value="0")).get().strip() or "0"
                def _esf(k): v = _es(k); return v if "." in v else v+".000000"
                use_equip = getattr(self, "_use_equip_stats", tk.BooleanVar()).get()
                if not use_equip:
                    ap=hp=hpc=mp=mpc=ap2=ac=dx=mmp=ma=md=mwt=da=lk=mhp=dp=hv="0"
                    hpr=mpr="0.000000"
                    min_stat_type=min_stat_lv="0"
                else:
                    ap=_es("ap");  hp=_es("hp"); hpc=_es("hpcon"); mp=_es("mp")
                    mpc=_es("mpcon"); ap2=_es("applus"); ac=_es("acplus")
                    dx=_es("dxplus"); mmp=_es("maxmpplus"); ma=_es("maplus")
                    md=_es("mdplus"); mwt=_es("maxwtplus"); da=_es("daplus")
                    lk=_es("lkplus"); mhp=_es("maxhpplus"); dp=_es("dpplus")
                    hv=_es("hvplus"); hpr=_esf("hprecoveryrate")
                    mpr=_esf("mprecoveryrate")
                    min_stat_type=_es("minstattype"); min_stat_lv=_es("minstatlv")

                row_xml = (
                    f"<ROW>\n"
                    f"<ID>{iid}</ID>\n"
                    f"<Class>{cls}</Class>\n"
                    f"<Type>{typ}</Type>\n"
                    f"<SubType>{sub}</SubType>\n"
                    f"<ItemFType>0</ItemFType>\n"
                    f"<Name><![CDATA[{it['name']}]]></Name>\n"
                    f"<Comment><![CDATA[{it['comment']}]]></Comment>\n"
                    f"<Use><![CDATA[{it.get('use', ' ')}]]></Use>\n"
                    f"<Name_Eng><![CDATA[ ]]></Name_Eng>\n"
                    f"<Comment_Eng><![CDATA[ ]]></Comment_Eng>\n"
                    f"<FileName><![CDATA[{fn}]]></FileName>\n"
                    f"<BundleNum>{bn}</BundleNum>\n"
                    f"<InvFileName><![CDATA[{inv_fn}]]></InvFileName>\n"
                    f"<InvBundleNum>{inv_bn}</InvBundleNum>\n"
                    f"<CmtFileName><![CDATA[{cmt_fn}]]></CmtFileName>\n"
                    f"<CmtBundleNum>{cmt_bn}</CmtBundleNum>\n"
                    f"<EquipFileName><![CDATA[ ]]></EquipFileName>\n"
                    f"<PivotID>0</PivotID>\n"
                    f"<PaletteId>0</PaletteId>\n"
                    f"<Options>{options_str}</Options>\n"
                    f"<HideHat>0</HideHat>\n"
                    f"<ChrTypeFlags>{chr_flag}</ChrTypeFlags>\n"
                    f"<GroundFlags>0</GroundFlags>\n"
                    f"<SystemFlags>0</SystemFlags>\n"
                    f"<OptionsEx>0</OptionsEx>\n"
                    f"<Weight>0</Weight>\n"
                    f"<Value>{value_field}</Value>\n"
                    f"<MinLevel>{min_level}</MinLevel>\n"
                    f"<Effect>0</Effect>\n"
                    f"<EffectFlags2>0</EffectFlags2>\n"
                    f"<SelRange>0</SelRange>\n"
                    f"<Life>0</Life>\n"
                    f"<Depth>0</Depth>\n"
                    f"<Delay>0.000000</Delay>\n"
                    f"<AP>{ap}</AP>\n"
                    f"<HP>{hp}</HP>\n"
                    f"<HPCon>{hpc}</HPCon>\n"
                    f"<MP>{mp}</MP>\n"
                    f"<MPCon>{mpc}</MPCon>\n"
                    f"<Money>0</Money>\n"
                    f"<APPlus>{ap2}</APPlus>\n"
                    f"<ACPlus>{ac}</ACPlus>\n"
                    f"<DXPlus>{dx}</DXPlus>\n"
                    f"<MaxMPPlus>{mmp}</MaxMPPlus>\n"
                    f"<MAPlus>{ma}</MAPlus>\n"
                    f"<MDPlus>{md}</MDPlus>\n"
                    f"<MaxWTPlus>{mwt}</MaxWTPlus>\n"
                    f"<DAPlus>{da}</DAPlus>\n"
                    f"<LKPlus>{lk}</LKPlus>\n"
                    f"<MaxHPPlus>{mhp}</MaxHPPlus>\n"
                    f"<DPPlus>{dp}</DPPlus>\n"
                    f"<HVPlus>{hv}</HVPlus>\n"
                    f"<HPRecoveryRate>{hpr}</HPRecoveryRate>\n"
                    f"<MPRecoveryRate>{mpr}</MPRecoveryRate>\n"
                    f"<CardNum>0</CardNum>\n"
                    f"<CardGenGrade>0</CardGenGrade>\n"
                    f"<CardGenParam>0.000000</CardGenParam>\n"
                    f"<DailyGenCnt>0</DailyGenCnt>\n"
                    f"<PartFileName><![CDATA[{part_fn}]]></PartFileName>\n"
                    f"<ChrFTypeFlag>0</ChrFTypeFlag>\n"
                    f"<ChrGender>0</ChrGender>\n"
                    f"<ExistType>0</ExistType>\n"
                    f"<Ncash>{ncash_field}</Ncash>\n"
                    f"<NewCM>0</NewCM>\n"
                    f"<FamCM>0</FamCM>\n"
                    f"<Summary><![CDATA[ ]]></Summary>\n"
                    f"<ShopFileName><![CDATA[{shop_fn}]]></ShopFileName>\n"
                    f"<ShopBundleNum>{shop_bn_raw}</ShopBundleNum>\n"
                    f"<MinStatType>{min_stat_type}</MinStatType>\n"
                    f"<MinStatLv>{min_stat_lv}</MinStatLv>\n"
                    f"<RefineIndex>0</RefineIndex>\n"
                    f"<RefineType>0</RefineType>\n"
                    f"<CompoundSlot>0</CompoundSlot>\n"
                    f"<SetItemID>0</SetItemID>\n"
                    f"<ReformCount>0</ReformCount>\n"
                    f"<GroupId>0</GroupId>\n"
                    f"</ROW>"
                )
                itemparam_rows.append(row_xml)

                # RecycleExcept row if Non-Recyclable
                if opt_recycle == 8388608:
                    recycle_rows.append(build_recycle_except_row(iid, it["name"]))

            # Persist last fashion item ID
            if last_fash_id > 0:
                _set_last_id("t18_fashion_item", last_fash_id)

        # ── 2. CMSetItemParam row ─────────────────────────────────────────
        cmset_row = ""
        if gen_set and set_id:
            slot_lines = []
            for i in range(MAX_FASHION_PIECES):
                if i < len(items):
                    slot_lines.append(
                        f"<Item{i}>{items[i]['item_id']}</Item{i}> <!-- {items[i]['name']} -->")
                else:
                    slot_lines.append(f"<Item{i}>0</Item{i}>")
            slots = "\n".join(slot_lines) + "\n"
            cmset_row = (f"<ROW>\n<ID>{set_id}</ID>\n"
                         f"<n><![CDATA[{set_name}]]></n>\n{slots}</ROW>")

        # ── 3. Box ItemParam row ──────────────────────────────────────────
        box_row = ""
        if gen_box and box_id:
            box_row = (
                f"<ROW>\n<ID>{box_id}</ID>\n"
                f"<Class>4</Class>\n<Type>20</Type>\n<SubType>0</SubType>\n<ItemFType>0</ItemFType>\n"
                f"<Name><![CDATA[{self._box_name.get().strip()}]]></Name>\n"
                f"<Comment><![CDATA[{self._box_comment.get().strip()}]]></Comment>\n"
                f"<Use><![CDATA[ ]]></Use>\n"
                f"<Name_Eng><![CDATA[ ]]></Name_Eng>\n"
                f"<Comment_Eng><![CDATA[ ]]></Comment_Eng>\n"
                f"<FileName><![CDATA[{fn}]]></FileName>\n"
                f"<BundleNum>{bn}</BundleNum>\n"
                f"<InvFileName><![CDATA[{inv_fn}]]></InvFileName>\n"
                f"<InvBundleNum>{inv_bn}</InvBundleNum>\n"
                f"<CmtFileName><![CDATA[{cmt_fn}]]></CmtFileName>\n"
                f"<CmtBundleNum>{cmt_bn}</CmtBundleNum>\n"
                f"<EquipFileName><![CDATA[ ]]></EquipFileName>\n"
                f"<PivotID>0</PivotID>\n<PaletteId>0</PaletteId>\n"
                f"<Options>1/16</Options>\n<HideHat>0</HideHat>\n"
                f"<ChrTypeFlags>0</ChrTypeFlags>\n"
                f"<GroundFlags>0</GroundFlags>\n<SystemFlags>0</SystemFlags>\n<OptionsEx>0</OptionsEx>\n"
                f"<Weight>0</Weight>\n<Value>0</Value>\n<MinLevel>1</MinLevel>\n"
                f"<Effect>0</Effect>\n<EffectFlags2>0</EffectFlags2>\n<SelRange>0</SelRange>\n"
                f"<Life>0</Life>\n<Depth>0</Depth>\n<Delay>0.000000</Delay>\n"
                f"<AP>0</AP>\n<HP>0</HP>\n<HPCon>0</HPCon>\n<MP>0</MP>\n<MPCon>0</MPCon>\n"
                f"<Money>0</Money>\n<APPlus>0</APPlus>\n<ACPlus>0</ACPlus>\n<DXPlus>0</DXPlus>\n"
                f"<MaxMPPlus>0</MaxMPPlus>\n<MAPlus>0</MAPlus>\n<MDPlus>0</MDPlus>\n"
                f"<MaxWTPlus>0</MaxWTPlus>\n<DAPlus>0</DAPlus>\n<LKPlus>0</LKPlus>\n"
                f"<MaxHPPlus>0</MaxHPPlus>\n<DPPlus>0</DPPlus>\n<HVPlus>0</HVPlus>\n"
                f"<HPRecoveryRate>0.000000</HPRecoveryRate>\n<MPRecoveryRate>0.000000</MPRecoveryRate>\n"
                f"<CardNum>0</CardNum>\n<CardGenGrade>0</CardGenGrade>\n<CardGenParam>0.000000</CardGenParam>\n"
                f"<DailyGenCnt>0</DailyGenCnt>\n<PartFileName><![CDATA[ ]]></PartFileName>\n"
                f"<ChrFTypeFlag>0</ChrFTypeFlag>\n<ChrGender>0</ChrGender>\n"
                f"<ExistType>0</ExistType>\n<Ncash>0</Ncash>\n<NewCM>0</NewCM>\n<FamCM>0</FamCM>\n"
                f"<Summary><![CDATA[ ]]></Summary>\n<ShopFileName><![CDATA[ ]]></ShopFileName>\n"
                f"<ShopBundleNum>0</ShopBundleNum>\n<MinStatType>0</MinStatType>\n"
                f"<MinStatLv>0</MinStatLv>\n<RefineIndex>0</RefineIndex>\n<RefineType>0</RefineType>\n"
                f"<CompoundSlot>0</CompoundSlot>\n<SetItemID>0</SetItemID>\n"
                f"<ReformCount>0</ReformCount>\n<GroupId>0</GroupId>\n</ROW>"
            )

        # ── 4. R_ShopItem rows ────────────────────────────────────────────
        shop_rows = ""
        if gen_shop:
            sp = self._shop_price_override.get().strip() or (myshop_price if use_myshop else "0")
            sc = self._shop_count.get().strip() or "100"
            for it in items:
                shop_rows += build_shop_row(it["item_id"], sc, sp) + "\n"
            if box_id:
                shop_rows += build_shop_row(box_id, sc, sp) + "\n"

        # ── 5. Exchange / Compound rows ───────────────────────────────────
        exchange_row = exchange_loc = ""
        if gen_exchange and self._exch_fields:
            cfg_e = {k: v.get() for k, v in self._exch_fields.items()}
            exchange_row = build_exchange_row(cfg_e)
            exchange_loc = build_exchange_location_row(cfg_e.get("exchange_id", "0"))
            try: _set_last_id("exchange", int(cfg_e.get("exchange_id", 0)))
            except: pass

        compound_row = compound_loc = ""
        if gen_compound and self._comp_fields:
            cfg_c = {k: v.get() for k, v in self._comp_fields.items()}
            compound_row = build_compound_row(cfg_c)
            compound_loc = build_compound_location_row(cfg_c.get("compound_id", "0"))
            try: _set_last_id("compound", int(cfg_c.get("compound_id", 0)))
            except: pass

        # ── 6. libcmgds_e XML GOODS blocks ──────────────────────────────────
        libcmgds = ""
        if use_myshop:
            today_lib  = datetime.date.today().strftime("%Y%m%d")
            price_item = getattr(self, "_myshop_price_item",
                                 tk.StringVar(value="12000")).get().strip() or "12000"
            goods_cat1 = str(lib_cat1) if lib_cat1 else "128"
            goods_limit_desc_lib = CFG_LIMIT_DESC_MAP.get(char_name, "All Characters")

            # glc counter — starts from user-set field (or last persisted)
            try:
                _lib_glc = int(getattr(self, "_goods_list_code_start",
                                       tk.StringVar(value="21000")).get() or "21000")
            except:
                _lib_glc = _get_last_id("t18_goods_list_code", 20999) + 1

            def _goods_attrs(gcode, gname, gset_count, gprice, issell):
                return (
                    f'goods_code="{gcode}" goods_name="{gname}" goods_desc="" '
                    f'goods_set_count="{gset_count}" goods_limit_use="2" goods_limit_time="0" '
                    f'goods_cash_price="{gprice}" goods_shop_new="1" goods_shop_popular="0" '
                    f'goods_category="1" goods_category0="15" goods_category1="{goods_cat1}" '
                    f'goods_category2="0" goods_limit_desc="{goods_limit_desc_lib}" '
                    f'goods_char_level="0" goods_char_sex="{sex}" goods_char_type="15" '
                    f'goods_issell="{issell}" goods_created="{today_lib}" '
                    f'goods_filtercode1="0" goods_filtercode2="0" goods_filtercode3="0" '
                    f'goods_filtercode4="0" goods_filterlevel="0" '
                    f'goods_discount_price="{gprice}" '
                    f'discount_start_date="1900-01-01 00:00:00" '
                    f'discount_end_date="1900-01-01 00:00:00" discount_display_date=""'
                )

            lib_blocks = []

            # ── SET block: one GOODS entry with all pieces as GOODS_LIST ─────
            if set_id and items:
                set_lines = [f'\t\t<GOODS {_goods_attrs(set_id, set_name, len(items), myshop_price, "0")}>']
                for it in items:
                    iid  = it["item_id"]
                    inam = it["name"]
                    set_lines.append(
                        f'\t\t\t<GOODS_LIST item_index="{iid}" goods_name="{inam}" '
                        f'item_count="1" item_class="2" preview_x="" preview_y="" '
                        f'preview_z="" preview_d="" '
                        f'goods_list_code="{_lib_glc}" parents_list_code="{_lib_glc}" />'
                    )
                    _lib_glc += 1
                set_lines.append('\t\t</GOODS>')
                lib_blocks.append("\n".join(set_lines))

            # ── Individual item blocks: one GOODS per piece ──────────────────
            for it in items:
                iid  = it["item_id"]
                inam = it["name"]
                item_block = (
                    f'\t\t<GOODS {_goods_attrs(iid, inam, 1, price_item, "0")}>\n'
                    f'\t\t\t<GOODS_LIST item_index="{iid}" goods_name="{inam}" '
                    f'item_count="1" item_class="2" preview_x="" preview_y="" '
                    f'preview_z="" preview_d="" '
                    f'goods_list_code="{_lib_glc}" parents_list_code="{_lib_glc}" />\n'
                    f'\t\t</GOODS>'
                )
                _lib_glc += 1
                lib_blocks.append(item_block)

            # Persist the last code used so next session continues from here
            _set_last_id("t18_goods_list_code", _lib_glc - 1)

            _n_lib = len(lib_blocks)
            libcmgds = "\n".join(lib_blocks)

        # ── 7. SQL block — tbl_goods + tbl_goods_list + tbl_goods_limit ────
        sql_block = ""
        if use_myshop:
            today_sql = datetime.date.today().strftime("%Y-%m-%d")
            limit_use = "2"
            price_item_sql = getattr(self, "_myshop_price_item",
                                     tk.StringVar(value="12000")).get().strip() or "12000"
            _sql_goods  = []
            _sql_list   = []
            _sql_limit  = []

            # goods_list_code starts from user-set field (pre-filled from last used +1)
            try:
                _glc = int(getattr(self, "_goods_list_code_start",
                                   tk.StringVar(value="21000")).get() or "21000")
            except:
                _glc = _get_last_id("t18_goods_list_code", 20999) + 1

            # ── SET entry in tbl_goods ────────────────────────────────────
            if set_id:
                _sql_goods.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods ("
                    f"goods_code, goods_name, goods_desc, goods_capacity, goods_category, "
                    f"goods_set_count, goods_item_index, goods_item_count, "
                    f"goods_limit_use, goods_limit_time, goods_cash_price, goods_created, "
                    f"goods_shop_new, goods_shop_popular, goods_sellcount, "
                    f"goods_category0, goods_category1, goods_category2, "
                    f"goods_limit_desc, goods_char_level, goods_char_sex, goods_char_type, "
                    f"version_code, goods_issell, goods_image"
                    f") VALUES ("
                    f"{set_id}, N'{set_name}', NULL, NULL, 0, "
                    f"{len(items)}, NULL, NULL, "
                    f"{limit_use}, NULL, {myshop_price}, '{today_sql} 00:00:00', "
                    f"0, 0, 0, "
                    f"15, {cat1}, 0, "
                    f"N'{CFG_LIMIT_DESC_MAP.get(char_name, 'All Characters')}', 0, {sex}, {char_sql_type}, "
                    f"10001, 1, ''"
                    f");"
                )
                # Each fashion piece gets its own GOODS_LIST row under the set
                for it in items:
                    iid  = it["item_id"]
                    inam = it["name"]
                    _sql_list.append(
                        f"INSERT INTO gmg_account.dbo.tbl_goods_list ("
                        f"goods_code, item_index, item_count, goods_scode, item_class, "
                        f"preview_x, preview_y, preview_z, preview_d, "
                        f"goods_list_code, parents_list_code, goods_list_limit"
                        f") VALUES ("
                        f"{set_id}, {iid}, 1, {iid}, 2, "
                        f"NULL, NULL, NULL, NULL, "
                        f"{_glc}, {_glc}, 0"
                        f");"
                    )
                    _glc += 1
                # tbl_goods_limit for the set
                _sql_limit.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods_limit ("
                    f"goods_code, limit_code, goods_limit_price, default_display"
                    f") VALUES ("
                    f"{set_id}, {limit_use}, {myshop_price}, True"
                    f");"
                )

            # ── Individual item entries in tbl_goods + tbl_goods_list ─────
            # Each fashion piece also gets its own standalone listing
            for it in items:
                iid  = it["item_id"]
                inam = it["name"]
                _sql_goods.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods ("
                    f"goods_code, goods_name, goods_desc, goods_capacity, goods_category, "
                    f"goods_set_count, goods_item_index, goods_item_count, "
                    f"goods_limit_use, goods_limit_time, goods_cash_price, goods_created, "
                    f"goods_shop_new, goods_shop_popular, goods_sellcount, "
                    f"goods_category0, goods_category1, goods_category2, "
                    f"goods_limit_desc, goods_char_level, goods_char_sex, goods_char_type, "
                    f"version_code, goods_issell, goods_image"
                    f") VALUES ("
                    f"{iid}, N'{inam}', NULL, NULL, 0, "
                    f"1, NULL, NULL, "
                    f"{limit_use}, NULL, {price_item_sql}, '{today_sql} 00:00:00', "
                    f"0, 0, 0, "
                    f"15, {cat1}, 0, "
                    f"N'{CFG_LIMIT_DESC_MAP.get(char_name, "All Characters")}', 0, {sex}, {char_sql_type}, "
                    f"10001, 0, ''"
                    f");"
                )
                _sql_list.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods_list ("
                    f"goods_code, item_index, item_count, goods_scode, item_class, "
                    f"preview_x, preview_y, preview_z, preview_d, "
                    f"goods_list_code, parents_list_code, goods_list_limit"
                    f") VALUES ("
                    f"{iid}, {iid}, 1, {iid}, 2, "
                    f"NULL, NULL, NULL, NULL, "
                    f"{_glc}, {_glc}, 0"
                    f");"
                )
                _glc += 1
                _sql_limit.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods_limit ("
                    f"goods_code, limit_code, goods_limit_price, default_display"
                    f") VALUES ("
                    f"{iid}, {limit_use}, {price_item_sql}, True"
                    f");"
                )

            # ── BOX entry if present ──────────────────────────────────────
            if box_id:
                _box_name = self._box_name.get().strip() if hasattr(self, "_box_name") else ""
                _sql_goods.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods ("
                    f"goods_code, goods_name, goods_desc, goods_capacity, goods_category, "
                    f"goods_set_count, goods_item_index, goods_item_count, "
                    f"goods_limit_use, goods_limit_time, goods_cash_price, goods_created, "
                    f"goods_shop_new, goods_shop_popular, goods_sellcount, "
                    f"goods_category0, goods_category1, goods_category2, "
                    f"goods_limit_desc, goods_char_level, goods_char_sex, goods_char_type, "
                    f"version_code, goods_issell, goods_image"
                    f") VALUES ("
                    f"{box_id}, N'{_box_name}', NULL, NULL, 0, "
                    f"1, NULL, NULL, "
                    f"{limit_use}, NULL, {myshop_price}, '{today_sql} 00:00:00', "
                    f"0, 0, 0, "
                    f"15, {cat1}, 0, "
                    f"N'All Characters', 0, {sex}, {char_sql_type}, "
                    f"10001, 0, ''"
                    f");"
                )
                _sql_list.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods_list ("
                    f"goods_code, item_index, item_count, goods_scode, item_class, "
                    f"preview_x, preview_y, preview_z, preview_d, "
                    f"goods_list_code, parents_list_code, goods_list_limit"
                    f") VALUES ("
                    f"{box_id}, {box_id}, 1, {box_id}, 1, "
                    f"NULL, NULL, NULL, NULL, "
                    f"{_glc}, {_glc}, 0"
                    f");"
                )
                _glc += 1
                _sql_limit.append(
                    f"INSERT INTO gmg_account.dbo.tbl_goods_limit ("
                    f"goods_code, limit_code, goods_limit_price, default_display"
                    f") VALUES ("
                    f"{box_id}, {limit_use}, {myshop_price}, True"
                    f");"
                )

            # NOTE: goods_list_code persistence is handled by the libcmgds block above.
            # If only SQL is generated (no libcmgds), persist here.
            if not use_myshop or not libcmgds:
                _set_last_id("t18_goods_list_code", _glc - 1)

            sql_block = (
                "BEGIN TRANSACTION;\n\n"
                "-- tbl_goods\n" + "\n\n".join(_sql_goods) + "\n\n"
                "-- tbl_goods_list\n" + "\n\n".join(_sql_list) + "\n\n"
                "-- tbl_goods_limit\n" + "\n\n".join(_sql_limit) + "\n\n"
                "COMMIT TRANSACTION;"
            ) if (_sql_goods or _sql_list or _sql_limit) else ""

        # Auto-save session
        _save_settings(self._PERSIST_KEY, self._collect_state())

        self._show_output(itemparam_rows, cmset_row, box_row, shop_rows,
                          exchange_row, exchange_loc, compound_row, compound_loc,
                          libcmgds, sql_block, use_myshop, recycle_rows)

    def _show_output(self, itemparam_rows, cmset_row, box_row, shop_rows,
                     exchange_row, exchange_loc, compound_row, compound_loc,
                     libcmgds, sql_block, use_myshop, recycle_rows=None):
        price_mode = "myshop" if use_myshop else "none"  # compat
        self._clear()
        wrap = tk.Frame(self, bg=BG); wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(1, weight=1); wrap.grid_columnconfigure(0, weight=1)

        hdr = tk.Frame(wrap, bg=BG2); hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="✨  Fashion Creation — Output",
                 font=("Consolas", 13, "bold"), bg=BG2, fg=self.ACC, pady=8
                 ).pack(side="left", padx=14)

        nb = ttk.Notebook(wrap); nb.grid(row=1, column=0, sticky="nsew", padx=6, pady=4)

        ip_text = "\n\n".join(itemparam_rows)  # kept for export compat
        if box_row:
            make_output_tab(nb, "Box ItemParam",   box_row,     "fashion_box_itemparam.xml",self.root)
        if cmset_row:
            make_output_tab(nb, "CMSetItemParam",  cmset_row,   "fashion_cmset.xml",        self.root)
        if itemparam_rows:
            make_output_tab(nb, "ItemParam rows",  "\n\n".join(itemparam_rows), "fashion_itemparam.xml", self.root)
        if recycle_rows:
            make_output_tab(nb, "RecycleExceptItem", "\n".join(recycle_rows), "fashion_recycle_except.xml", self.root)
        if shop_rows:
            make_output_tab(nb, "R_ShopItem rows", shop_rows,   "fashion_shopitem.txt",     self.root)
        if exchange_row:
            make_output_tab(nb, "ExchangeContents",exchange_row,"fashion_exchange.xml",     self.root)
            make_output_tab(nb, "Exchange_Location",exchange_loc,"fashion_exchange_loc.xml",self.root)
        if compound_row:
            make_output_tab(nb, "Compound_Potion", compound_row,"fashion_compound.xml",     self.root)
            make_output_tab(nb, "Compounder_Spot",compound_loc,"fashion_compounder_spot.xml",self.root)
        if price_mode == "myshop":
            if libcmgds:
                make_output_tab(nb, "libcmgds_e",  libcmgds,    "fashion_libcmgds_e.xml",  self.root)
            if sql_block:
                make_output_tab(nb, "SQL",         sql_block,   "fashion_goods_limit.sql",  self.root)

        nav = tk.Frame(wrap, bg=BG2); nav.grid(row=2, column=0, sticky="ew")
        mk_btn(nav, "◀  Back / Edit", self._build_editor,
               color=BG4).pack(side="left", padx=12, pady=6)

        def _next_tab():
            cur = nb.index("current")
            total = nb.index("end")
            if cur + 1 < total:
                nb.select(cur + 1)

        def _prev_tab():
            cur = nb.index("current")
            if cur > 0:
                nb.select(cur - 1)

        mk_btn(nav, "◀ Prev Tab", _prev_tab,
               color=BG3, font=("Consolas", 8)).pack(side="left", padx=2, pady=6)
        mk_btn(nav, "Next Tab ▶", _next_tab,
               color=BG3, font=("Consolas", 8)).pack(side="left", padx=2, pady=6)
        mk_btn(nav, "💾  Export All to Folders", lambda: self._export_fashion(
               ip_text, box_row, cmset_row, shop_rows,
               exchange_row, exchange_loc, compound_row, compound_loc,
               libcmgds, sql_block, price_mode),
               color=self.ACC, fg=BG2,
               font=("Consolas", 10, "bold")).pack(side="right", padx=14, pady=6)

    def _export_fashion(self, ip_text, box_row, cmset_row, shop_rows,
                        exchange_row, exchange_loc, compound_row, compound_loc,
                        libcmgds, sql_block, price_mode):
        lib_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(), "libconfig"))
        ms_dir  = _APP_SETTINGS.get("myshop_dir",    os.path.join(os.getcwd(), "MyShop"))
        os.makedirs(lib_dir, exist_ok=True)
        saved = []

        def _w(d, fname, text):
            if not text.strip(): return
            p = os.path.join(d, fname)
            with open(p, "w", encoding="utf-8") as f: f.write(text)
            saved.append(os.path.basename(p))

        _w(lib_dir, "fashion_itemparam.xml",    ip_text)
        _w(lib_dir, "fashion_box_itemparam.xml",box_row)
        _w(lib_dir, "fashion_cmset.xml",        cmset_row)
        _w(lib_dir, "fashion_shopitem.txt",     shop_rows)
        _w(lib_dir, "fashion_exchange.xml",     exchange_row)
        _w(lib_dir, "fashion_exchange_loc.xml", exchange_loc)
        _w(lib_dir, "fashion_compound.xml",     compound_row)
        _w(lib_dir, "fashion_compounder_spot.xml", compound_loc)
        if price_mode == "myshop":
            os.makedirs(ms_dir, exist_ok=True)
            _w(ms_dir, "fashion_libcmgds_e.xml",  libcmgds)
            _w(ms_dir, "fashion_goods_limit.sql",  sql_block)

        messagebox.showinfo("Export Complete",
            f"libconfig  →  {lib_dir}\n"
            f"MyShop     →  {ms_dir if price_mode=='myshop' else '(not MyShop mode)'}\n\n"
            + "\n".join(saved))


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 8 — CMSetItemParam.xml Set Generator
# ══════════════════════════════════════════════════════════════════════════════

ACC8 = "#b4befe"   # lavender — set generator

_SET_COL_GUIDE = (
    "CMSetItemParam CSV / Excel column guide:\n\n"
    "Layout A — one row per set  (SetID column present):\n"
    "  SetID / Set ID     → the set's own <ID> tag\n"
    "  Name / Set Name    → the set's <n> name (CDATA)\n"
    "  ID  (repeated)     → item IDs for Item0..Item7\n"
    "  Item0..Item7  or  0..7   → item IDs by slot\n"
    "  Item Name / Name of Item / Box Name / Name of Box\n"
    "                     → inline <!-- comment --> next to each item\n\n"
    "Layout B — multiple rows per set  (group-header style):\n"
    "  Any column whose header is not a recognised keyword\n"
    "    becomes the set's group name (its <n>)\n"
    "  ID column          → item IDs (one per row)\n"
    "  Item Name / Name of Item → inline <!-- comment -->\n"
    "  SetID column       → the set's <ID> (optional)\n\n"
    "Rules:\n"
    "  • Up to 8 items per set (Item0..Item7). Extras are ignored.\n"
    "  • Slots filled with 0 if fewer than 8 items are provided.\n"
    "  • <!-- comments --> omitted when item name is blank.\n"
    "  • Excel: each sheet is treated as an independent batch."
)


class Tool8(tk.Frame):
    """Set Generator — builds CMSetItemParam.xml <ROW> entries."""

    ACC = ACC8

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root    = root
        self.session = session
        self._set_rows = []   # list of (set_id, set_name, items[8])
        self._build_start_screen()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    # ── Start screen ──────────────────────────────────────────────────────
    def _build_start_screen(self):
        self._clear()
        frm = tk.Frame(self, bg=BG); frm.pack(expand=True, fill="both")
        center = tk.Frame(frm, bg=BG); center.pack(expand=True)

        tk.Label(center, text="SET ITEM PARAM GENERATOR",
                 font=("Consolas", 18, "bold"), bg=BG, fg=self.ACC).pack(pady=(36, 4))
        tk.Label(center, text="Generates <ROW> entries for CMSetItemParam.xml",
                 bg=BG, fg=FG_DIM, font=("Consolas", 10)).pack(pady=(0, 10))

        # Column guide box
        info = tk.Frame(center, bg=BG2); info.pack(pady=6, padx=20, fill="x")
        tk.Label(info, text=_SET_COL_GUIDE, bg=BG2, fg=FG, font=("Consolas", 8),
                 justify="left", padx=12, pady=10).pack(anchor="w")

        bf = tk.Frame(center, bg=BG); bf.pack(pady=16)
        mk_btn(bf, "✏️  Manual Entry", self._build_manual_editor,
               color=self.ACC, fg=BG2, font=("Consolas", 11, "bold")).pack(side="left", padx=8)
        mk_btn(bf, "📂  Import Spreadsheet", self._import_spreadsheet,
               color=BG3).pack(side="left", padx=8)
        mk_btn(bf, "📄  Download Template", self._save_template,
               color=BG4).pack(side="left", padx=8)

        if self._set_rows:
            n = len(self._set_rows)
            tk.Label(center, text=f"{n} set row(s) pending",
                     bg=BG, fg=GREEN, font=("Consolas", 9, "italic")).pack(pady=(4, 0))
            mk_btn(center, "📋  View / Export current rows", self._build_output_screen,
                   color=BG3).pack(pady=4)

    # ── Manual editor ──────────────────────────────────────────────────────
    def _build_manual_editor(self, prefill=None):
        self._clear()
        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0, weight=0)
        wrap.grid_rowconfigure(1, weight=1)
        wrap.grid_rowconfigure(2, weight=0)
        wrap.grid_columnconfigure(0, weight=1)

        hdr = tk.Frame(wrap, bg=BG2); hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="Set Item Param — Manual Entry",
                 font=("Consolas", 13, "bold"), bg=BG2, fg=self.ACC, pady=8
                 ).pack(side="left", padx=15)
        tk.Label(hdr,
                 text="SetID = the set's own ID.  ID column = IDs of items inside the set.",
                 font=("Consolas", 8), bg=BG2, fg=FG_DIM).pack(side="left", padx=4)

        scroll_host = tk.Frame(wrap, bg=BG)
        scroll_host.grid(row=1, column=0, sticky="nsew")
        canv, C = mk_scroll_canvas(scroll_host)

        pf = prefill or {}
        last_set_id = int(_get_last_id("t8_set", 0) or 0)
        v_set_id   = tk.StringVar(value=str(pf.get("set_id",  int(last_set_id or 0) + 1)))
        v_set_name = tk.StringVar(value=str(pf.get("set_name", "")))

        # Identity section
        s1 = mk_section(C, "  Set Identity  ⚠ SetID is the set's ID, not the item IDs")
        def _lr(parent, lbl, var, w=20, tip=""):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=3)
            lw = tk.Label(r, text=lbl, width=24, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9)); lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=w, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip: _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent
        _lr(s1, "SetID: ⚠ REQUIRED", v_set_id, 14,
            "The Set's own unique ID.  This is the <ID> tag in CMSetItemParam.xml.")
        _lr(s1, "Set Name:", v_set_name, 40,
            "The display name for this set.  Stored as <n><![CDATA[...]]></n>.")

        # Items section
        s2 = mk_section(C, "  Items (Item0 – Item7)  —  ID = item inside the set")
        tk.Label(s2,
                 text="  Item ID = the ID of each item that belongs to this set.\n"
                      "  Item Name = optional inline comment (appears as <!-- name --> in XML).",
                 bg=BG, fg=FG_GREY, font=("Consolas", 8)).pack(anchor="w", padx=10, pady=(4,2))

        pf_items = pf.get("items", [])
        item_id_vars   = []
        item_name_vars = []
        for i in range(8):
            pfi = pf_items[i] if i < len(pf_items) else {}
            r = tk.Frame(s2, bg=BG); r.pack(fill="x", padx=8, pady=2)
            tk.Label(r, text=f"Item{i}  ID:", bg=BG, fg=FG,
                     font=("Consolas", 9), width=10, anchor="w").pack(side="left")
            vid = tk.StringVar(value=str(pfi.get("item_id", "0") or "0"))
            tk.Entry(r, textvariable=vid, width=12, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
            tk.Label(r, text="Name:", bg=BG, fg=FG_DIM,
                     font=("Consolas", 9)).pack(side="left", padx=(10, 0))
            vnm = tk.StringVar(value=str(pfi.get("item_name", "")))
            tk.Entry(r, textvariable=vnm, width=28, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
            tk.Label(r, text="← optional <!-- -->", bg=BG, fg=FG_GREY,
                     font=("Consolas", 7)).pack(side="left")
            item_id_vars.append(vid)
            item_name_vars.append(vnm)

        # Footer
        nav = tk.Frame(wrap, bg=BG2); nav.grid(row=2, column=0, sticky="ew")

        def _gather_entry():
            return {
                "set_id":   v_set_id.get().strip(),
                "set_name": v_set_name.get(),
                "items": [
                    {"item_id": item_id_vars[i].get().strip(),
                     "item_name": item_name_vars[i].get().strip()}
                    for i in range(8)
                ]
            }

        def _generate():
            entry = _gather_entry()
            if not entry["set_id"]:
                messagebox.showerror("Missing SetID", "Please enter a SetID."); return
            try: _set_last_id("t8_set", int(entry["set_id"]))
            except: pass
            self._set_rows.append(entry)
            self._build_output_screen()

        def _generate_next():
            entry = _gather_entry()
            if not entry["set_id"]:
                messagebox.showerror("Missing SetID", "Please enter a SetID."); return
            self._set_rows.append(entry)
            try:
                next_id = str(int(entry["set_id"]) + 1)
                _set_last_id("t8_set", int(entry["set_id"]))
            except: next_id = ""
            # Pre-fill next entry with incremented ID
            next_pf = {"set_id": next_id, "set_name": "", "items": []}
            self._build_manual_editor(prefill=next_pf)

        mk_btn(nav, "◀  Back", self._build_start_screen,
               color=BG4).pack(side="left", padx=14, pady=6)
        mk_btn(nav, "🗑  Clear", lambda: self._build_manual_editor(),
               color=BG4).pack(side="left", padx=4, pady=6)
        mk_btn(nav, "➕  Generate & Next", _generate_next,
               color=GREEN, fg=BG2).pack(side="right", padx=4, pady=6)
        mk_btn(nav, "⚡  Generate", _generate,
               color=self.ACC, fg=BG2,
               font=("Consolas", 11, "bold")).pack(side="right", padx=14, pady=6)

    # ── Spreadsheet import ────────────────────────────────────────────────
    def _import_spreadsheet(self):
        path = filedialog.askopenfilename(
            title="Import Set data from spreadsheet",
            filetypes=[
                ("Spreadsheet", "*.csv *.xlsx *.xlsm *.xls"),
                ("CSV", "*.csv"), ("Excel", "*.xlsx *.xlsm *.xls"),
                ("All", "*.*"),
            ], parent=self.root)
        if not path: return

        ext = os.path.splitext(path)[1].lower()
        try:
            if ext in (".xlsx", ".xlsm", ".xls"):
                if not _HAVE_OPENPYXL:
                    messagebox.showerror("Missing library",
                        "openpyxl is required for Excel files.\nInstall: pip install openpyxl")
                    return
                wb = openpyxl.load_workbook(path, data_only=True)
                all_sets = []
                for sh_name in wb.sheetnames:
                    csv_text = _sheet_to_csv(wb[sh_name])
                    sets = parse_set_csv(csv_text)
                    all_sets.extend(sets)
            else:
                with open(path, encoding="utf-8-sig") as f:
                    csv_text = f.read()
                all_sets = parse_set_csv(csv_text)
        except Exception as e:
            messagebox.showerror("Import Error", str(e)); return

        if not all_sets:
            messagebox.showwarning("No data",
                "No valid set rows found.\n\n"
                "Check that your spreadsheet has a SetID column (Layout A)\n"
                "or a non-keyword group column header (Layout B).\n\n"
                "Click '📄 Download Template' on the start screen for examples.")
            return

        self._show_import_preview(all_sets, path)

    def _show_import_preview(self, sets, filepath):
        win = tk.Toplevel(self.root)
        win.title(f"Set Import Preview — {len(sets)} set(s)")
        win.configure(bg=BG); win.geometry("860x540"); win.grab_set()

        hdr = tk.Frame(win, bg=BG2); hdr.pack(fill="x")
        tk.Label(hdr, text=f"📥  Set Import Preview — {os.path.basename(filepath)}",
                 bg=BG2, fg=self.ACC, font=("Consolas", 12, "bold"), pady=8).pack(side="left", padx=14)
        tk.Label(hdr, text=f"{len(sets)} set(s) found",
                 bg=BG2, fg=FG_DIM, font=("Consolas", 9)).pack(side="left")

        sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True, padx=8, pady=6)
        canv, C = mk_scroll_canvas(sh)

        # Editable preview table — SetID | Name | Item0..Item7 (ID + Name pairs)
        cols = (["SetID", "Set Name"] +
                [f"Item{i} ID" for i in range(8)] +
                [f"Item{i} Name" for i in range(8)])
        col_w = 11

        # Header row
        hrow = tk.Frame(C, bg=BG2); hrow.pack(fill="x")
        for col in cols:
            tk.Label(hrow, text=col, bg=BG2, fg=BLUE, font=("Consolas", 7, "bold"),
                     width=col_w, anchor="w", padx=2).pack(side="left")

        # Data rows — each set is one row
        row_vars = []
        for ri, s in enumerate(sets):
            bg_row = BG if ri % 2 == 0 else BG2
            rv = {}
            r = tk.Frame(C, bg=bg_row); r.pack(fill="x")

            sid_v  = tk.StringVar(value=s.get("set_id", ""))
            snm_v  = tk.StringVar(value=s.get("set_name", ""))
            for v, col in [(sid_v, "SetID"), (snm_v, "Set Name")]:
                tk.Entry(r, textvariable=v, width=col_w, bg=bg_row, fg=FG,
                         insertbackground=FG, font=("Consolas", 8), relief="flat", bd=0
                         ).pack(side="left", padx=1)
                rv[col] = v

            items = s.get("items", [])
            for i in range(8):
                itm = items[i] if i < len(items) else {}
                vid  = tk.StringVar(value=str(itm.get("item_id", "0")))
                vnm  = tk.StringVar(value=str(itm.get("item_name", "")))
                for v, col in [(vid, f"Item{i} ID"), (vnm, f"Item{i} Name")]:
                    tk.Entry(r, textvariable=v, width=col_w, bg=bg_row, fg=FG,
                             insertbackground=FG, font=("Consolas", 8), relief="flat", bd=0
                             ).pack(side="left", padx=1)
                    rv[col] = v

            row_vars.append(rv)

        nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")
        tk.Label(nav, text="You can edit cells above before confirming.",
                 bg=BG2, fg=FG_DIM, font=("Consolas", 8)).pack(side="left", padx=14, pady=6)

        def _confirm():
            for rv in row_vars:
                set_cfg = {
                    "set_id":   rv["SetID"].get().strip(),
                    "set_name": rv["Set Name"].get(),
                    "items": [
                        {"item_id":   rv[f"Item{i} ID"].get().strip(),
                         "item_name": rv[f"Item{i} Name"].get().strip()}
                        for i in range(8)
                    ]
                }
                if set_cfg["set_id"]:
                    self._set_rows.append(set_cfg)
            win.destroy()
            messagebox.showinfo("Imported",
                f"Added {len(self._set_rows)} set row(s) total.")
            self._build_output_screen()

        mk_btn(nav, f"✓  Import {len(sets)} set(s)", _confirm,
               color=GREEN, fg=BG2, font=("Consolas", 10, "bold")).pack(side="right", padx=14, pady=6)
        mk_btn(nav, "Cancel", win.destroy, color=BG4).pack(side="right", padx=4, pady=6)
        win.wait_window()

    # ── Template download ─────────────────────────────────────────────────
    def _save_template(self):
        path = filedialog.asksaveasfilename(
            title="Save Set template", defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            initialfile="set_template.csv", parent=self.root)
        if not path: return
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(["SetID", "Set Name",
                         "Item0 ID", "Item0 Name",
                         "Item1 ID", "Item1 Name",
                         "Item2 ID", "Item2 Name",
                         "Item3 ID", "Item3 Name",
                         "Item4 ID", "Item4 Name",
                         "Item5 ID", "Item5 Name",
                         "Item6 ID", "Item6 Name",
                         "Item7 ID", "Item7 Name"])
            w.writerow(["1001", "Example Set",
                         "12001", "Sword of Fire",
                         "12002", "Shield of Fire",
                         "12003", "Helm of Fire",
                         "0", "", "0", "", "0", "", "0", "", "0", ""])
        messagebox.showinfo("Template saved", f"Saved to:\n{path}")

    # ── Output screen ─────────────────────────────────────────────────────
    def _build_output_screen(self):
        self._clear()
        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(1, weight=1); wrap.grid_columnconfigure(0, weight=1)

        hdr = tk.Frame(wrap, bg=BG2); hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="CMSetItemParam.xml Output",
                 font=("Consolas", 14, "bold"), bg=BG2, fg=self.ACC, pady=8
                 ).pack(side="left", padx=15)
        tk.Label(hdr, text=f"  {len(self._set_rows)} set row(s)",
                 bg=BG2, fg=FG_DIM, font=("Consolas", 9)).pack(side="left")

        if self._set_rows:
            xml_rows = [build_set_row(s) for s in self._set_rows]
            xml_all  = "\n".join(xml_rows)

            nb = ttk.Notebook(wrap)
            nb.grid(row=1, column=0, sticky="nsew", padx=6, pady=4)
            fname_key = _APP_SETTINGS["filenames"].get("set_item_param", "CMSetItemParam.xml")
            make_output_tab(nb, "CMSetItemParam.xml rows", xml_all, fname_key, self.root)
        else:
            tk.Label(wrap, text="No set rows yet.",
                     bg=BG, fg=FG_GREY, font=("Consolas", 10)).grid(row=1, column=0)

        nav = tk.Frame(wrap, bg=BG2); nav.grid(row=2, column=0, sticky="ew")

        mk_btn(nav, "◀  Back", self._build_start_screen, color=BG4).pack(side="left", padx=14, pady=6)
        mk_btn(nav, "➕  Add Manual Entry", self._build_manual_editor, color=BG4).pack(side="left", padx=4, pady=6)
        mk_btn(nav, "📂  Import Spreadsheet", self._import_spreadsheet, color=BG4).pack(side="left", padx=4, pady=6)
        mk_btn(nav, "🗑  Clear All",
               lambda: (self._set_rows.clear(), self._build_start_screen())
                       if messagebox.askyesno("Clear", "Clear all set rows?") else None,
               color=BG4).pack(side="left", padx=4, pady=6)

        def _export():
            if not self._set_rows:
                messagebox.showwarning("Nothing to export", "Generate some set rows first.")
                return
            default_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(), "libconfig"))
            folder = filedialog.askdirectory(
                title="Choose export folder (default: libconfig)", initialdir=default_dir)
            if not folder: folder = default_dir
            os.makedirs(folder, exist_ok=True)
            xml_rows = [build_set_row(s) for s in self._set_rows]
            xml_all  = "\n".join(xml_rows)
            fname = _APP_SETTINGS["filenames"].get("set_item_param", "CMSetItemParam.xml")
            if _APP_SETTINGS.get("timestamp_files", False):
                import time as _t
                ts = _t.strftime("%d%m%y-%S%M%H")
                n, e = os.path.splitext(fname)
                fname = f"{n}_{ts}{e}"
            out_path = os.path.join(folder, fname)
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(xml_all)
            messagebox.showinfo("Export Complete", f"Saved:\n{out_path}")

        mk_btn(nav, "💾  Export", _export,
               color=GREEN, fg=BG2, font=("Consolas", 10, "bold")).pack(side="right", padx=14, pady=6)


class Tool6(tk.Frame):
    """ItemParam Generator — builds full <ROW> entries for any item type."""

    ACC = "#94e2d5"   # teal — tool 6

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root    = root
        self.session = session
        self._settings = _load_t6_settings()
        self._first_run = not bool(self._settings.get("_first_run_done", False))
        self._rows = []           # list of generated row strings
        self._build_start_screen()

    # ─────────────────────────────────────────────────────────────────────────
    def _clear(self):
        for w in self.winfo_children(): w.destroy()

    def _build_start_screen(self):
        """Welcome / entry screen for ItemParam Generator — mirrors Tool1's load screen."""
        self._clear()
        frm = tk.Frame(self, bg=BG); frm.pack(expand=True, fill="both")
        center = tk.Frame(frm, bg=BG); center.pack(expand=True)

        tk.Label(center, text="ITEMPARAM GENERATOR",
                 font=("Consolas", 20, "bold"), bg=BG, fg=self.ACC).pack(pady=(40, 4))
        tk.Label(center,
                 text="Mewsie's ItemParam Toolbox — build full <ROW> entries for any item type.",
                 bg=BG, fg=FG_DIM, font=("Consolas", 10)).pack(pady=(0, 6))

        info = tk.Frame(center, bg=BG2); info.pack(pady=8, padx=20, fill="x")
        tk.Label(info,
            text=(
                "Three layout modes:\n"
                "  Normal  — follows XML field order with functional group labels\n"
                "  Simple  — grouped by purpose (Identity, Equip Stats, Drilling Params, etc.)\n"
                "  Custom  — same as Normal but drag section headers to reorder\n\n"
                "Hover any label for a full description.  "
                "All values are saved between sessions automatically."
            ),
            bg=BG2, fg=FG, font=("Consolas", 9), justify="left",
            padx=12, pady=10).pack(anchor="w")

        saved_name = self._settings.get("name", "")
        saved_id   = self._settings.get("id", "")
        if saved_name or saved_id:
            resume_lbl = (f"Last session:  ID {saved_id}  —  {saved_name}"
                          if saved_name else f"Last session:  ID {saved_id}")
            tk.Label(center, text=resume_lbl, bg=BG, fg=GREEN,
                     font=("Consolas", 9, "italic")).pack(pady=(4, 0))

        bf = tk.Frame(center, bg=BG); bf.pack(pady=20)
        mk_btn(bf, "✏️  New / Continue Entry", self._build_editor,
               color=self.ACC, fg=BG2,
               font=("Consolas", 11, "bold")).pack(side="left", padx=8)

        # Session import — pull CE rows from session
        if self.session.compound_rows or self.session.exchange_rows:
            def _import_ce_session():
                n_c = len(self.session.compound_rows)
                n_e = len(self.session.exchange_rows)
                messagebox.showinfo("Session CE Rows",
                    f"Session has:\n  {n_c} compound row(s)\n  {n_e} exchange row(s)\n"
                    "These will be included when you Export from the output screen.")
            mk_btn(bf, "⬇  CE from Session", _import_ce_session,
                   color=BG4).pack(side="left", padx=8)

        def _import_from_start():
            path = filedialog.askopenfilename(
                title="Import ItemParam from spreadsheet",
                filetypes=[
                    ("Spreadsheet", "*.csv *.xlsx *.xlsm *.xls"),
                    ("CSV", "*.csv"), ("Excel", "*.xlsx *.xlsm *.xls"),
                    ("All", "*.*"),
                ],
                parent=self.root,
            )
            if not path: return
            ext = os.path.splitext(path)[1].lower()
            try:
                if ext in (".xlsx", ".xlsm", ".xls"):
                    _, raw_rows = _read_xlsx_rows(path)
                else:
                    _, raw_rows = _read_csv_rows(path)
            except Exception as e:
                messagebox.showerror("Import Error", str(e)); return
            if not raw_rows:
                messagebox.showwarning("Empty", "No data rows found."); return
            _alias = {
                "id":"id","itemid":"id","class":"class_val","type":"type_val",
                "subtype":"subtype_val","itemftype":"itemftype_val",
                "name":"name","comment":"comment","use":"use",
                "nameeng":"name_eng","commenteng":"comment_eng",
                "filename":"file_name","fn":"file_name",
                "bundlenum":"bundle_num","bn":"bundle_num",
                "cmtfilename":"cmt_file_name","cmtfn":"cmt_file_name",
                "cmtbundlenum":"cmt_bundle_num","cmtbn":"cmt_bundle_num",
                "equipfilename":"equip_file_name",
                "shopfilename":"shop_file_name","shopbundlenum":"shop_bundle_num",
                "partfilename":"part_file_name",
                "pivotid":"pivot_id","paletteid":"palette_id","groupid":"group_id",
                "options":"options_raw_manual","optionsex":"options_ex",
                "effect":"effect","existtype":"exist_type",
                "weight":"weight","value":"value","minlevel":"min_level",
                "money":"money","ncash":"ncash",
                "chrtypeflags":"chr_type_flags","hidehat":"hide_hat",
            }
            import re as _re
            new_s = dict(self._settings)
            for col, val in raw_rows[0].items():
                key = _alias.get(_re.sub(r"[^a-z0-9]", "", col.lower()))
                # Values: plain strip only — never strip . or \ from paths
                if key and str(val).strip(): new_s[key] = str(val).strip()
            self._settings = new_s; _save_t6_settings(new_s)
            if len(raw_rows) > 1:
                messagebox.showinfo("Import",
                    f"Loaded row 1 of {len(raw_rows)}.\n"
                    "Use Generate & Continue to step through all rows.")
            self._build_editor()

        mk_btn(bf, "📥  Import Spreadsheet", _import_from_start,
               color=BG4).pack(side="left", padx=8)
        mk_btn(bf, "📖  Field Reference", self._show_reference,
               color=BG4).pack(side="left", padx=8)
        if self._settings:
            mk_btn(bf, "🗑  Reset All Settings", self._reset,
                   color=BG4).pack(side="left", padx=8)

    # ─────────────────────────────────────────────────────────────────────────
    def _build_editor(self):
        self._clear()
        s = self._settings
        ACC = self.ACC

        # ── Layout preference (Normal / Simple / Custom) ──────────────────
        layout_var = tk.StringVar(value=s.get("layout_mode", "normal"))

        # ── Outer wrapper ─────────────────────────────────────────────────
        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0, weight=0)
        wrap.grid_rowconfigure(1, weight=1)
        wrap.grid_rowconfigure(2, weight=0)
        wrap.grid_columnconfigure(0, weight=1)

        # ── Header ────────────────────────────────────────────────────────
        hdr = tk.Frame(wrap, bg=BG2)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text="ItemParam Generator",
                 font=("Consolas", 14, "bold"), bg=BG2, fg=ACC, pady=8
                 ).pack(side="left", padx=15)
        tk.Label(hdr, text="Generates full <ROW> entries for any item type.",
                 font=("Consolas", 7), bg=BG2, fg=FG_DIM).pack(side="left", padx=4)
        mk_btn(hdr, "📖  Field Reference", self._show_reference,
               color=BG4, font=("Consolas", 8)).pack(side="right", padx=10, pady=4)

        # ── Layout selector (always visible in header area) ───────────────
        lay_bar = tk.Frame(wrap, bg=BG3)
        lay_bar.grid(row=0, column=0, sticky="e")
        # Re-position: pack inside hdr on right
        lay_bar.grid_forget()
        lay_row = tk.Frame(hdr, bg=BG2); lay_row.pack(side="right", padx=16)
        tk.Label(lay_row, text="Layout:", bg=BG2, fg=FG_DIM,
                 font=("Consolas", 8)).pack(side="left")
        for ltxt, lval in [("Normal", "normal"), ("Simple", "simple"), ("Custom", "custom")]:
            tk.Radiobutton(lay_row, text=ltxt, variable=layout_var, value=lval,
                           bg=BG2, fg=FG, selectcolor=BG3, activebackground=BG2,
                           font=("Consolas", 8),
                           command=lambda: (s.update({"layout_mode": layout_var.get()}),
                                            _save_t6_settings(s),
                                            self._build_editor())
                           ).pack(side="left", padx=4)

        # ── Scrollable body ───────────────────────────────────────────────
        scroll_host = tk.Frame(wrap, bg=BG)
        scroll_host.grid(row=1, column=0, sticky="nsew")
        canvas, C = mk_scroll_canvas(scroll_host)

        def sec(title):
            return mk_section(C, title)

        def lbl_entry(parent, label, var, width=30, tip=None):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            lw = tk.Label(r, text=label, width=24, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9)); lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=width, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip: _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent

        def lbl_note(parent, text, fg=FG_GREY):
            tk.Label(parent, text=text, bg=BG, fg=fg,
                     font=("Consolas", 8)).pack(anchor="w", padx=10, pady=(2, 0))

        # ── Welcome banner ────────────────────────────────────────────────
        if self._first_run:
            wb = tk.Frame(C, bg=BG4, padx=12, pady=8)
            wb.pack(fill="x", padx=12, pady=(10, 4))
            tk.Label(wb, text="👋  Welcome to the ItemParam Generator",
                     bg=BG4, fg=ACC, font=("Consolas", 11, "bold")).pack(anchor="w")
            tk.Label(wb, text=(
                "All fields pre-filled with safe defaults.  Hover any label for its description.\n"
                "Class and Type default to Unselected — set them before generating.\n"
                "Choose a layout with the Normal / Simple / Custom selector in the header."),
                bg=BG4, fg=FG, font=("Consolas", 9), justify="left").pack(anchor="w", pady=(4, 0))
            mk_btn(wb, "Got it — dismiss", lambda: wb.destroy(),
                   color=BG3, font=("Consolas", 9)).pack(anchor="w", pady=(6, 0))

        # ── Mode toggle (Dropdown / Manual) ──────────────────────────────
        mode_var = tk.StringVar(value=s.get("input_mode", "dropdown"))
        mode_frm = tk.Frame(C, bg=BG); mode_frm.pack(fill="x", padx=14, pady=(8, 2))
        tk.Label(mode_frm, text="Input mode:", bg=BG, fg=FG_DIM,
                 font=("Consolas", 9)).pack(side="left")
        for lbl_m, val_m in [("Dropdown", "dropdown"), ("Manual", "manual")]:
            tk.Radiobutton(mode_frm, text=lbl_m, variable=mode_var, value=val_m,
                           bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                           font=("Consolas", 9)).pack(side="left", padx=8)
        lbl_note(mode_frm.master,
                 "  Manual mode: Class/Type/SubType/ItemFType, Options & OptionsEx become free-text fields.")

        # ══════════════════════════════════════════════════════════════════
        # ALL FIELD VARIABLES
        # ══════════════════════════════════════════════════════════════════
        _default_id = s.get("id", "") or str(_get_last_id("t6_item", 0) + 1 or "")
        v_id        = tk.StringVar(value=_default_id)
        v_class     = tk.StringVar(value=s.get("class_val", "0"))
        v_type      = tk.StringVar(value=s.get("type_val", "0"))
        v_sub       = tk.StringVar(value=s.get("subtype_val", "0"))
        v_ift       = tk.StringVar(value=s.get("itemftype_val", "0"))
        v_name      = tk.StringVar(value=s.get("name", ""))
        v_comment   = tk.StringVar(value=s.get("comment", ""))
        v_use       = tk.StringVar(value=s.get("use", ""))
        v_name_eng  = tk.StringVar(value=s.get("name_eng", " "))
        v_cmt_eng   = tk.StringVar(value=s.get("comment_eng", " "))
        v_fn        = tk.StringVar(value=s.get("file_name",     r"data\item\itm000.nri"))
        v_bn        = tk.StringVar(value=s.get("bundle_num",    "0"))
        v_cmtfn     = tk.StringVar(value=s.get("cmt_file_name", r"data\item\itm_illu000.nri"))
        v_cmtbn     = tk.StringVar(value=s.get("cmt_bundle_num","0"))
        v_equipfn   = tk.StringVar(value=s.get("equip_file_name"," "))
        v_pivot     = tk.StringVar(value=str(s.get("pivot_id",  0)))
        v_palette   = tk.StringVar(value=str(s.get("palette_id", 0)))
        v_exist_type= tk.StringVar(value=str(s.get("exist_type", 0)))
        v_ground    = tk.StringVar(value=str(s.get("ground_flags", 0)))
        v_system    = tk.StringVar(value=str(s.get("system_flags", 0)))
        v_hide_hat  = tk.StringVar(value=str(s.get("hide_hat", 0)))
        v_chr_raw   = tk.StringVar(value=str(s.get("chr_type_flags", 0)))
        v_eff2      = tk.StringVar(value=str(s.get("effect_flags2", 0)))
        v_sel_range = tk.StringVar(value=str(s.get("sel_range", 0)))
        v_life      = tk.StringVar(value=str(s.get("life", 0)))
        v_depth     = tk.StringVar(value=str(s.get("depth", 0)))
        v_cardnum   = tk.StringVar(value=str(s.get("cardnum",     0)))
        v_cardgrade = tk.StringVar(value=str(s.get("cardgengrade",0)))
        v_daily     = tk.StringVar(value=str(s.get("dailygencnt", 0)))
        v_refine_idx  = tk.StringVar(value=str(s.get("refine_index",  0)))
        v_refine_type = tk.StringVar(value=str(s.get("refine_type",   0)))
        v_minstattype = tk.StringVar(value=str(s.get("min_stat_type", 0)))
        v_minstatLv   = tk.StringVar(value=str(s.get("min_stat_lv",   0)))
        v_compound    = tk.StringVar(value=str(s.get("compound_slot", 0)))
        v_setitem     = tk.StringVar(value=str(s.get("set_item_id",   0)))
        v_reform      = tk.StringVar(value=str(s.get("reform_count",  0)))
        v_group       = tk.StringVar(value=str(s.get("group_id",    0)))
        v_shopfn    = tk.StringVar(value=s.get("shop_file_name",  " "))
        v_shopbn    = tk.StringVar(value=str(s.get("shop_bundle_num", 0)))
        v_partfn    = tk.StringVar(value=s.get("part_file_name",  ""))

        # Options
        saved_opts_raw = s.get("options_raw_manual", "0")
        _saved_active = set()
        for _p in str(saved_opts_raw).replace(",", "/").split("/"):
            try: _saved_active.add(int(_p.strip()))
            except: pass
        _saved_flags_list = s.get("options_flags", [])
        while len(_saved_flags_list) < len(_OPTIONS_FULL): _saved_flags_list.append(False)
        if not _saved_active and any(_saved_flags_list):
            _saved_active = {f for (f, _), on in zip(_OPTIONS_FULL, _saved_flags_list) if on}
        opts_vars = [tk.BooleanVar(value=(fval in _saved_active)) for fval, _ in _OPTIONS_FULL]
        v_opts_manual = tk.StringVar(value=saved_opts_raw)

        # OptionsEx
        saved_optex = s.get("options_ex", 0)
        v_optex_raw    = tk.StringVar(value=str(saved_optex))
        v_optex_manual = tk.StringVar(value=str(saved_optex))
        optex_vars = {}

        # Effect
        saved_effect_raw = str(s.get("effect", 0))
        _eff_active = set()
        for _ep in str(saved_effect_raw).replace("/", ",").split(","):
            try: _eff_active.add(int(_ep.strip()))
            except: pass
        eff_vars = {}
        v_effect_raw   = tk.StringVar(value=saved_effect_raw)

        # Numeric stats dict
        num_vars = {}
        _num_defaults = {
            "weight":"1","value":"0","min_level":"1","money":"0","ncash":"0",
            "ap":"0","hp":"0","hpcon":"0","mp":"0","mpcon":"0",
            "applus":"0","acplus":"0","dxplus":"0","maxmpplus":"0","maplus":"0",
            "mdplus":"0","maxwtplus":"0","daplus":"0","lkplus":"0","maxhpplus":"0",
            "dpplus":"0","hvplus":"0",
            "hprecoveryrate":"0","mprecoveryrate":"0","delay":"0","cardgenparam":"0",
        }
        for k, dflt in _num_defaults.items():
            num_vars[k] = tk.StringVar(value=str(s.get(k, dflt)))

        # ── Widget factories ──────────────────────────────────────────────
        def _build_dd_row(parent, label, mapping, var, tip, show_raw=False):
            """Dropdown row. show_raw=False hides the manual textbox (Class/Type/SubType/ItemFType)."""
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            lw = tk.Label(r, text=label, width=24, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9)); lw.pack(side="left")
            vals_dd = [f"{v} — {d}" for v, d in mapping]
            combo = ttk.Combobox(r, values=vals_dd, state="readonly",
                                 width=38, font=("Consolas", 9))
            combo.pack(side="left", padx=4)
            # Raw entry only shown when show_raw=True (e.g. Options/Effect)
            raw_ent = tk.Entry(r, textvariable=var, width=7, bg=BG3, fg=FG,
                               insertbackground=FG, font=("Consolas", 9), relief="flat")
            if show_raw:
                raw_ent.pack(side="left", padx=(0, 4))
            def _dd_sel(e):
                sel = combo.current()
                if sel >= 0: var.set(str(mapping[sel][0]))
            combo.bind("<<ComboboxSelected>>", _dd_sel)
            def _raw_ch(*_):
                val = var.get().strip()
                for i, (v, _) in enumerate(mapping):
                    if str(v) == val: combo.set(vals_dd[i]); return
                combo.set("")
            var.trace_add("write", _raw_ch); _raw_ch()
            if tip:
                _attach_tooltip(lw, tip); _attach_tooltip(combo, tip)
                if show_raw: _attach_tooltip(raw_ent, tip)
            return combo, raw_ent

        def _build_man_row(parent, label, var, tip):
            r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
            lw = tk.Label(r, text=label, width=24, anchor="w", bg=BG, fg=FG,
                          font=("Consolas", 9)); lw.pack(side="left")
            ent = tk.Entry(r, textvariable=var, width=14, bg=BG3, fg=FG,
                           insertbackground=FG, font=("Consolas", 9), relief="flat")
            ent.pack(side="left", padx=4)
            if tip: _attach_tooltip(lw, tip); _attach_tooltip(ent, tip)
            return ent

        # ── Options widgets (built once, parented to placeholder frames) ──
        def _build_options_section(parent_dd, parent_man):
            """Build Options checkboxes + manual entry into given parent frames."""
            opt_chk_frm = tk.Frame(parent_dd, bg=BG)
            opt_chk_frm.pack(fill="x", padx=8, pady=4)
            for i, (fval, flbl) in enumerate(_OPTIONS_FULL):
                cb = tk.Checkbutton(opt_chk_frm, text=f"{flbl}  ({fval})",
                                    variable=opts_vars[i], bg=BG, fg=FG,
                                    selectcolor=BG3, activebackground=BG,
                                    font=("Consolas", 8))
                cb.grid(row=i // 3, column=i % 3, sticky="w", padx=6, pady=1)
                _attach_tooltip(cb, f"Flag: {fval}  —  {flbl}")
            opt_preview = tk.Label(parent_dd, text="Value: ", bg=BG, fg=GREEN,
                                   font=("Consolas", 9))
            opt_preview.pack(anchor="w", padx=10, pady=(2, 4))

            def _update_opts(*_):
                selected = [f for (f, _), v in zip(_OPTIONS_FULL, opts_vars) if v.get()]
                raw = "/".join(str(f) for f in selected) if selected else "0"
                opt_preview.config(text=f"Value: {raw}")
                v_opts_manual.set(raw)
            for v in opts_vars: v.trace_add("write", _update_opts)
            _update_opts()

            # Manual
            r_mo = tk.Frame(parent_man, bg=BG); r_mo.pack(fill="x", padx=8, pady=4)
            tk.Label(r_mo, text="Options (slash-sep):", width=24, anchor="w",
                     bg=BG, fg=FG, font=("Consolas", 9)).pack(side="left")
            man_opts_ent = tk.Entry(r_mo, textvariable=v_opts_manual, width=14,
                                    bg=BG3, fg=FG, insertbackground=FG,
                                    font=("Consolas", 9), relief="flat")
            man_opts_ent.pack(side="left", padx=4)
            _attach_tooltip(man_opts_ent, _TOOLTIPS["Options"])

            def _manual_opts_to_chk(*_):
                raw = v_opts_manual.get().strip()
                active = set()
                for p in raw.replace(",", "/").split("/"):
                    p = p.strip()
                    try: active.add(int(p))
                    except: pass
                for (fval, _), v in zip(_OPTIONS_FULL, opts_vars):
                    v.set(fval in active)
            v_opts_manual.trace_add("write", _manual_opts_to_chk)

        def _build_optionsex_section(parent_dd, parent_man):
            optex_chk_frm = tk.Frame(parent_dd, bg=BG)
            optex_chk_frm.pack(fill="x", padx=8, pady=4)
            for i, (fval, fdesc) in enumerate(_OPTIONSEX_MAP):
                if fval == 0: continue
                v = tk.BooleanVar()
                cb = tk.Checkbutton(optex_chk_frm, text=f"{fval} — {fdesc}", variable=v,
                                    bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                                    font=("Consolas", 8))
                cb.grid(row=(i-1)//2, column=(i-1)%2, sticky="w", padx=6, pady=1)
                _attach_tooltip(cb, f"Flag: {fval}  —  {fdesc}")
                optex_vars[fval] = v
            r_optex = tk.Frame(parent_dd, bg=BG); r_optex.pack(fill="x", padx=8, pady=(2,4))
            tk.Label(r_optex, text="Value:", bg=BG, fg=FG_DIM,
                     font=("Consolas", 9)).pack(side="left")
            optex_preview = tk.Label(r_optex, text="0", bg=BG, fg=GREEN,
                                     font=("Consolas", 9))
            optex_preview.pack(side="left", padx=6)

            def _optex_chk_to_raw(*_):
                total = sum(f for f, v in optex_vars.items() if v.get())
                v_optex_raw.set(str(total)); v_optex_manual.set(str(total))
                optex_preview.config(text=str(total))
            def _optex_raw_to_chk(*_):
                try: total = int(v_optex_raw.get())
                except: total = 0
                for f, v in optex_vars.items(): v.set(bool(total & f))
                optex_preview.config(text=str(total))
            for v in optex_vars.values(): v.trace_add("write", _optex_chk_to_raw)
            v_optex_raw.trace_add("write", _optex_raw_to_chk)
            _optex_raw_to_chk()

            r_moe = tk.Frame(parent_man, bg=BG); r_moe.pack(fill="x", padx=8, pady=4)
            tk.Label(r_moe, text="OptionsEx (raw int):", width=24, anchor="w",
                     bg=BG, fg=FG, font=("Consolas", 9)).pack(side="left")
            man_optex_ent = tk.Entry(r_moe, textvariable=v_optex_manual, width=14,
                                     bg=BG3, fg=FG, insertbackground=FG,
                                     font=("Consolas", 9), relief="flat")
            man_optex_ent.pack(side="left", padx=4)
            _attach_tooltip(man_optex_ent, _TOOLTIPS["OptionsEx"])
            def _man_optex_to_raw(*_):
                v_optex_raw.set(v_optex_manual.get())
            v_optex_manual.trace_add("write", _man_optex_to_raw)

        def _build_effect_section(parent_dd, parent_man, parent_preview):
            eff_chk_frm = tk.Frame(parent_dd, bg=BG)
            eff_chk_frm.pack(fill="x", padx=8, pady=4)
            for i, (fval, fdesc) in enumerate(_EFFECT_MAP):
                v = tk.BooleanVar(value=(fval in _eff_active))
                cb = tk.Checkbutton(eff_chk_frm, text=f"{fval} — {fdesc}", variable=v,
                                    bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                                    font=("Consolas", 8))
                cb.grid(row=i//3, column=i%3, sticky="w", padx=6, pady=1)
                _attach_tooltip(cb, f"Effect {fval}: {fdesc}")
                eff_vars[fval] = v

            eff_preview = tk.Label(parent_preview, text="Value: 0", bg=BG, fg=GREEN,
                                   font=("Consolas", 9))
            eff_preview.pack(anchor="w", padx=10)

            r_eff = tk.Frame(parent_preview, bg=BG)
            r_eff.pack(fill="x", padx=8, pady=(0, 4))
            tk.Label(r_eff, text="Raw value:", bg=BG, fg=FG_DIM,
                     font=("Consolas", 9)).pack(side="left")
            eff_raw_ent = tk.Entry(r_eff, textvariable=v_effect_raw, width=14,
                                   bg=BG3, fg=FG, insertbackground=FG,
                                   font=("Consolas", 9), relief="flat")
            eff_raw_ent.pack(side="left", padx=6)
            _attach_tooltip(eff_raw_ent, _TOOLTIPS["Effect"])

            def _eff_chk_to_raw(*_):
                vals = sorted(f for f, v in eff_vars.items() if v.get())
                raw = "/".join(str(x) for x in vals if x != 0) or "0"
                v_effect_raw.set(raw)
                eff_preview.config(text="Value: " + raw)
            def _eff_raw_to_chk(*_):
                raw = v_effect_raw.get()
                active = set()
                for p in raw.replace("/", ",").split(","):
                    try: active.add(int(p.strip()))
                    except: pass
                for fval, v in eff_vars.items():
                    v.set(fval in active)
                eff_preview.config(text="Value: " + raw)
            for v in eff_vars.values(): v.trace_add("write", _eff_chk_to_raw)
            v_effect_raw.trace_add("write", _eff_raw_to_chk)
            _eff_raw_to_chk()

            # Manual
            r_me = tk.Frame(parent_man, bg=BG); r_me.pack(fill="x", padx=8, pady=4)
            tk.Label(r_me, text="Effect (slash-sep):", width=24, anchor="w",
                     bg=BG, fg=FG, font=("Consolas", 9)).pack(side="left")
            tk.Entry(r_me, textvariable=v_effect_raw, width=14, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)

        def _build_hidehat_section(parent):
            lbl_note(parent,
                     "  Same flag values as ChrTypeFlags. Default: 0 unless item hides ears.")
            hh_sel_frm = tk.Frame(parent, bg=BG); hh_sel_frm.pack(fill="x", padx=8, pady=4)
            tk.Label(hh_sel_frm, text="Character:", bg=BG, fg=FG,
                     font=("Consolas", 9)).pack(side="left")
            hh_name_dd = ttk.Combobox(hh_sel_frm, values=CHR_NAMES, state="readonly",
                                       width=12, font=("Consolas", 9))
            hh_name_dd.pack(side="left", padx=6)
            tk.Label(hh_sel_frm, text="Job:", bg=BG, fg=FG,
                     font=("Consolas", 9)).pack(side="left")
            hh_job_dd = ttk.Combobox(hh_sel_frm, values=["1st","2nd","3rd"],
                                      state="readonly", width=6, font=("Consolas", 9))
            hh_job_dd.pack(side="left", padx=6)
            hh_add_btn = mk_btn(hh_sel_frm, "+", None, color=GREEN, fg=BG2,
                                font=("Consolas","11","bold"), width=3)
            hh_add_btn.pack(side="left", padx=2)
            hh_rem_btn = mk_btn(hh_sel_frm, "−", None, color=ACC3, fg=BG2,
                                font=("Consolas","11","bold"), width=3)
            hh_rem_btn.pack(side="left", padx=2)
            hh_lb_frm = tk.Frame(parent, bg=BG); hh_lb_frm.pack(fill="x", padx=8)
            hh_lb = tk.Listbox(hh_lb_frm, height=3, width=36, bg=BG3, fg=FG,
                               font=("Consolas", 9), selectbackground=BG4, activestyle="none")
            hh_lb.pack(anchor="w")
            r_hh = tk.Frame(parent, bg=BG); r_hh.pack(fill="x", padx=8, pady=(0,4))
            tk.Label(r_hh, text="Raw value:", bg=BG, fg=FG_DIM,
                     font=("Consolas", 9)).pack(side="left")
            hh_raw_ent = tk.Entry(r_hh, textvariable=v_hide_hat, width=12,
                                   bg=BG3, fg=FG, insertbackground=FG,
                                   font=("Consolas", 9), relief="flat")
            hh_raw_ent.pack(side="left", padx=6)
            _attach_tooltip(hh_raw_ent, _TOOLTIPS["HideHat"])
            hh_selected = []
            def _refresh_hh():
                hh_lb.delete(0, "end")
                for val in hh_selected: hh_lb.insert("end", CHR_FLAG_REVERSE.get(val, str(val)))
                v_hide_hat.set(str(sum(hh_selected)))
            def _add_hh():
                name = hh_name_dd.get(); job = hh_job_dd.get()
                if not name or not job: return
                val = CHR_FLAG_MAP.get(f"{name} {job}")
                if val and val not in hh_selected: hh_selected.append(val); _refresh_hh()
            def _rem_hh():
                sel = hh_lb.curselection()
                if sel: hh_selected.pop(sel[0]); _refresh_hh()
            hh_add_btn.config(command=_add_hh); hh_rem_btn.config(command=_rem_hh)
            try: saved_hh_int = int(s.get("hide_hat", 0))
            except: saved_hh_int = 0
            for v_hh in sorted(CHR_FLAG_MAP.values()):
                if saved_hh_int & v_hh: hh_selected.append(v_hh)
            _refresh_hh()

        def _build_chrtype_section(parent):
            lbl_note(parent, "  0 = no restriction (all characters allowed).")
            chr_sel_frm = tk.Frame(parent, bg=BG); chr_sel_frm.pack(fill="x", padx=8, pady=4)
            tk.Label(chr_sel_frm, text="Character:", bg=BG, fg=FG,
                     font=("Consolas", 9)).pack(side="left")
            chr_name_dd = ttk.Combobox(chr_sel_frm, values=CHR_NAMES, state="readonly",
                                        width=12, font=("Consolas", 9))
            chr_name_dd.pack(side="left", padx=6)
            tk.Label(chr_sel_frm, text="Job:", bg=BG, fg=FG,
                     font=("Consolas", 9)).pack(side="left")
            chr_job_dd = ttk.Combobox(chr_sel_frm, values=["1st","2nd","3rd"],
                                       state="readonly", width=6, font=("Consolas", 9))
            chr_job_dd.pack(side="left", padx=6)
            chr_add_btn = mk_btn(chr_sel_frm, "+", None, color=GREEN, fg=BG2,
                                  font=("Consolas","11","bold"), width=3)
            chr_add_btn.pack(side="left", padx=2)
            chr_rem_btn = mk_btn(chr_sel_frm, "−", None, color=ACC3, fg=BG2,
                                  font=("Consolas","11","bold"), width=3)
            chr_rem_btn.pack(side="left", padx=2)
            chr_lb_frm = tk.Frame(parent, bg=BG); chr_lb_frm.pack(fill="x", padx=8)
            chr_lb = tk.Listbox(chr_lb_frm, height=4, width=36, bg=BG3, fg=FG,
                                font=("Consolas", 9), selectbackground=BG4, activestyle="none")
            chr_lb.pack(anchor="w")
            r_chr = tk.Frame(parent, bg=BG); r_chr.pack(fill="x", padx=8, pady=(0,4))
            tk.Label(r_chr, text="Raw value:", bg=BG, fg=FG_DIM,
                     font=("Consolas", 9)).pack(side="left")
            chr_raw_ent = tk.Entry(r_chr, textvariable=v_chr_raw, width=12,
                                    bg=BG3, fg=FG, insertbackground=FG,
                                    font=("Consolas", 9), relief="flat")
            chr_raw_ent.pack(side="left", padx=6)
            _attach_tooltip(chr_raw_ent, _TOOLTIPS["ChrTypeFlags"])
            chr_selected6 = []
            def _refresh_chr6():
                chr_lb.delete(0, "end")
                for val in chr_selected6: chr_lb.insert("end", CHR_FLAG_REVERSE.get(val, str(val)))
                v_chr_raw.set(str(sum(chr_selected6)))
            def _add_chr6():
                name = chr_name_dd.get(); job = chr_job_dd.get()
                if not name or not job: return
                val = CHR_FLAG_MAP.get(f"{name} {job}")
                if val and val not in chr_selected6: chr_selected6.append(val); _refresh_chr6()
            def _rem_chr6():
                sel = chr_lb.curselection()
                if sel: chr_selected6.pop(sel[0]); _refresh_chr6()
            chr_add_btn.config(command=_add_chr6); chr_rem_btn.config(command=_rem_chr6)
            try: saved_chr_int = int(s.get("chr_type_flags", 0))
            except: saved_chr_int = 0
            for v_c in sorted(CHR_FLAG_MAP.values()):
                if saved_chr_int & v_c: chr_selected6.append(v_c)
            _refresh_chr6()

        def _build_numeric_stats(parent):
            nr = tk.Frame(parent, bg=BG); nr.pack(fill="x", padx=8, pady=4)
            for ci, (lbl_t, key, tip) in enumerate([
                ("Weight:",   "weight",    _TOOLTIPS["Weight"]),
                ("Value:",    "value",     _TOOLTIPS["Value"]),
                ("MinLevel:", "min_level", _TOOLTIPS["MinLevel"]),
                ("Money:",    "money",     _TOOLTIPS["Money"]),
                ("Ncash:",    "ncash",     _TOOLTIPS["Ncash"]),
            ]):
                lw2 = tk.Label(nr, text=lbl_t, bg=BG, fg=FG, font=("Consolas", 9), width=10, anchor="w")
                lw2.grid(row=0, column=ci*2, padx=3)
                ent2 = tk.Entry(nr, textvariable=num_vars[key], width=10, bg=BG3, fg=FG,
                                insertbackground=FG, font=("Consolas", 9), relief="flat")
                ent2.grid(row=0, column=ci*2+1, padx=3)
                if tip: _attach_tooltip(lw2, tip); _attach_tooltip(ent2, tip)

        def _build_equip_stats(parent):
            sf = tk.Frame(parent, bg=BG); sf.pack(fill="x", padx=8, pady=4)
            stat_list = [
                ("AP:",       "ap",        _TOOLTIPS["AP"]),
                ("HP:",       "hp",        _TOOLTIPS["HP"]),
                ("HPCon:",    "hpcon",     _TOOLTIPS["HPCon"]),
                ("MP:",       "mp",        _TOOLTIPS["MP"]),
                ("MPCon:",    "mpcon",     _TOOLTIPS["MPCon"]),
                ("APPlus:",   "applus",    _TOOLTIPS["APPlus"]),
                ("ACPlus:",   "acplus",    _TOOLTIPS["ACPlus"]),
                ("DXPlus:",   "dxplus",    _TOOLTIPS["DXPlus"]),
                ("MaxMPPlus:","maxmpplus", _TOOLTIPS["MaxMPPlus"]),
                ("MAPlus:",   "maplus",    _TOOLTIPS["MAPlus"]),
                ("MDPlus:",   "mdplus",    _TOOLTIPS["MDPlus"]),
                ("MaxWTPlus:","maxwtplus", _TOOLTIPS["MaxWTPlus"]),
                ("DAPlus:",   "daplus",    _TOOLTIPS["DAPlus"]),
                ("LKPlus:",   "lkplus",    _TOOLTIPS["LKPlus"]),
                ("MaxHPPlus:","maxhpplus", _TOOLTIPS["MaxHPPlus"]),
                ("DPPlus:",   "dpplus",    _TOOLTIPS["DPPlus"]),
                ("HVPlus:",   "hvplus",    _TOOLTIPS["HVPlus"]),
            ]
            for i, (lbl_t, key, tip) in enumerate(stat_list):
                lw2 = tk.Label(sf, text=lbl_t, bg=BG, fg=FG, font=("Consolas", 9), width=10, anchor="w")
                lw2.grid(row=i//4, column=(i%4)*2, padx=3, pady=1)
                ent2 = tk.Entry(sf, textvariable=num_vars[key], width=8, bg=BG3, fg=FG,
                                insertbackground=FG, font=("Consolas", 9), relief="flat")
                ent2.grid(row=i//4, column=(i%4)*2+1, padx=3, pady=1)
                if tip: _attach_tooltip(lw2, tip); _attach_tooltip(ent2, tip)
            ff = tk.Frame(parent, bg=BG); ff.pack(fill="x", padx=8, pady=4)
            for ci, (lbl_t, key, tip) in enumerate([
                ("HPRecovery:", "hprecoveryrate", _TOOLTIPS["HPRecoveryRate"]),
                ("MPRecovery:", "mprecoveryrate", _TOOLTIPS["MPRecoveryRate"]),
            ]):
                lw2 = tk.Label(ff, text=lbl_t, bg=BG, fg=FG_DIM, font=("Consolas", 9), width=12, anchor="w")
                lw2.grid(row=0, column=ci*2, padx=3)
                ent2 = tk.Entry(ff, textvariable=num_vars[key], width=12, bg=BG3, fg=FG,
                                insertbackground=FG, font=("Consolas", 9), relief="flat")
                ent2.grid(row=0, column=ci*2+1, padx=3)
                if tip: _attach_tooltip(lw2, tip); _attach_tooltip(ent2, tip)

        def _build_drilling_params(parent):
            lbl_entry(parent, "Life:",  v_life,  10, _TOOLTIPS["Life"])
            lbl_entry(parent, "Depth:", v_depth, 10, _TOOLTIPS["Depth"])
            lbl_entry(parent, "Delay:", num_vars["delay"], 12, _TOOLTIPS["Delay"])
            lbl_note(parent, "  Life: timed item duration or drill life. Depth: test drills only.")

        def _build_refine_section(parent):
            def _dd_row_ref(label, mapping, raw_var, tip_key=""):
                r = tk.Frame(parent, bg=BG); r.pack(fill="x", padx=8, pady=2)
                tip = _TOOLTIPS.get(tip_key, "")
                lw2 = tk.Label(r, text=label, width=24, anchor="w", bg=BG, fg=FG, font=("Consolas", 9))
                lw2.pack(side="left")
                vals_dd = [f"{v} — {d}" for v, d in mapping]
                combo2 = ttk.Combobox(r, values=vals_dd, state="readonly", width=34, font=("Consolas", 9))
                combo2.pack(side="left", padx=4)
                raw_ent2 = tk.Entry(r, textvariable=raw_var, width=7, bg=BG3, fg=FG,
                                    insertbackground=FG, font=("Consolas", 9), relief="flat")
                raw_ent2.pack(side="left", padx=(0,4))
                def _sel2(e):
                    sel2 = combo2.current()
                    if sel2 >= 0: raw_var.set(str(mapping[sel2][0]))
                combo2.bind("<<ComboboxSelected>>", _sel2)
                def _raw2(*_):
                    val2 = raw_var.get().strip()
                    for i2, (v2, _) in enumerate(mapping):
                        if str(v2) == val2: combo2.set(vals_dd[i2]); return
                    combo2.set("")
                raw_var.trace_add("write", _raw2); _raw2()
                if tip: _attach_tooltip(lw2, tip); _attach_tooltip(combo2, tip); _attach_tooltip(raw_ent2, tip)
            _dd_row_ref("RefineIndex:", _REFINEINDEX_MAP, v_refine_idx, "RefineIndex")
            _dd_row_ref("RefineType:",  _REFINETYPE_MAP,  v_refine_type,"RefineType")
            _dd_row_ref("MinStatType:", _MINSTATTYPE_MAP, v_minstattype,"MinStatType")
            lbl_entry(parent, "MinStatLv:",     v_minstatLv, 8, _TOOLTIPS["MinStatLv"])
            lbl_entry(parent, "CompoundSlot:",  v_compound,  8, _TOOLTIPS["CompoundSlot"])
            lbl_entry(parent, "SetItemID: (Default: 0)", v_setitem, 8, _TOOLTIPS["SetItemID"])
            lbl_entry(parent, "ReformCount: (Default: 0)", v_reform, 8, _TOOLTIPS["ReformCount"])
            lbl_note(parent, "  CompoundSlot: range 0-5. Higher values cause UI errors on inspect.")

        def _build_ground_system(parent):
            def _attach_warn_nonzero(ew, fname):
                def _cb(e):
                    val = ew.get().strip()
                    if val not in ("", "0"):
                        messagebox.showwarning("Warning",
                            f"{fname} is not 0.\nThis may have unintended consequences.")
                ew.bind("<FocusOut>", _cb)
            gnd_ent = lbl_entry(parent, "GroundFlags: (Default: 0)", v_ground, 10, _TOOLTIPS["GroundFlags"])
            sys_ent = lbl_entry(parent, "SystemFlags: (Default: 0)", v_system, 10, _TOOLTIPS["SystemFlags"])
            lbl_note(parent, "  Always 0 in standard items. Warning appears if changed.")
            _attach_warn_nonzero(gnd_ent, "GroundFlags"); _attach_warn_nonzero(sys_ent, "SystemFlags")

        # ══════════════════════════════════════════════════════════════════
        # LAYOUT: NORMAL — follows XML field order with functional group labels
        # ══════════════════════════════════════════════════════════════════
        def _build_normal_layout():
            # ── Identity ─────────────────────────────────────────────────
            s1 = sec("  Identity  ⚠ Class and Type MUST be set")
            lbl_entry(s1, "ID:", v_id, width=14,
                      tip="Item ID — must be unique in the XML table. Required.")
            dd_class_frm = tk.Frame(s1, bg=BG)
            dd_type_frm  = tk.Frame(s1, bg=BG)
            dd_sub_frm   = tk.Frame(s1, bg=BG)
            dd_ift_frm   = tk.Frame(s1, bg=BG)
            man_class_frm = tk.Frame(s1, bg=BG)
            man_type_frm  = tk.Frame(s1, bg=BG)
            man_sub_frm   = tk.Frame(s1, bg=BG)
            man_ift_frm   = tk.Frame(s1, bg=BG)
            _build_dd_row(dd_class_frm,  "Class: *",    _CLASS_MAP,    v_class,
                          _TOOLTIPS["Class"]+"\n⚠ REQUIRED — must not be 0.")
            _build_dd_row(dd_type_frm,   "Type: *",     _TYPE_MAP,     v_type,
                          _TOOLTIPS["Type"]+"\n⚠ REQUIRED — must not be 0.")
            _build_dd_row(dd_sub_frm,    "SubType:",     _SUBTYPE_MAP,  v_sub,   _TOOLTIPS["SubType"])
            _build_dd_row(dd_ift_frm,    "ItemFType:",   _ITEMFTYPE_MAP,v_ift,   _TOOLTIPS["ItemFType"])
            _build_man_row(man_class_frm,"Class: *",     v_class, _TOOLTIPS["Class"]+"\n⚠ REQUIRED.")
            _build_man_row(man_type_frm, "Type: *",      v_type,  _TOOLTIPS["Type"]+"\n⚠ REQUIRED.")
            _build_man_row(man_sub_frm,  "SubType:",     v_sub,   _TOOLTIPS["SubType"])
            _build_man_row(man_ift_frm,  "ItemFType:",   v_ift,   _TOOLTIPS["ItemFType"])

            # ── Names ────────────────────────────────────────────────────
            s2 = sec("  Names & Text  ")
            lbl_entry(s2, "Name:",        v_name,    40, _TOOLTIPS["Name"])
            lbl_entry(s2, "Comment:",     v_comment, 60, _TOOLTIPS["Comment"])
            lbl_entry(s2, "Use:",         v_use,     60, _TOOLTIPS["Use"])
            lbl_entry(s2, "Name_Eng:",    v_name_eng, 40, _TOOLTIPS["Name_Eng"])
            lbl_entry(s2, "Comment_Eng:", v_cmt_eng,  40, _TOOLTIPS["Comment_Eng"])

            # ── Files ────────────────────────────────────────────────────
            s3 = sec("  Filepaths & Bundle Numbers  ")
            lbl_entry(s3, "FileName:",     v_fn,   50, _TOOLTIPS["FileName"])
            lbl_entry(s3, "BundleNum:",    v_bn,    8, _TOOLTIPS["BundleNum"])
            lbl_note(s3, "  InvFileName / InvBundleNum are auto-copied from FileName / BundleNum.")
            lbl_entry(s3, "CmtFileName:",  v_cmtfn, 50, _TOOLTIPS["CmtFileName"])
            lbl_entry(s3, "CmtBundleNum:", v_cmtbn,  8, _TOOLTIPS["CmtBundleNum"])
            lbl_entry(s3, "EquipFileName:", v_equipfn, 50, _TOOLTIPS["EquipFileName"])
            lbl_note(s3, "  EquipFileName: leave blank if not equipment/drill.")

            # ── PivotID / PaletteId ──────────────────────────────────────
            s_piv = sec("  PivotID & PaletteId  (Default: 0)")
            lbl_entry(s_piv, "PivotID:   (Default: 0)", v_pivot,   10, _TOOLTIPS["PivotID"])
            lbl_entry(s_piv, "PaletteId: (Default: 0)", v_palette, 10, _TOOLTIPS["PaletteId"])

            # ── Options ──────────────────────────────────────────────────
            s4 = sec("  Options  (eItemOption — slash-separated flags e.g. 1/2/256)")
            dd_opts_frm = tk.Frame(s4, bg=BG)
            man_opts_frm = tk.Frame(s4, bg=BG)
            _build_options_section(dd_opts_frm, man_opts_frm)

            # ── HideHat ──────────────────────────────────────────────────
            s_hh = sec("  HideHat  (Default: 0 — per-character ear hide)")
            _build_hidehat_section(s_hh)

            # ── ChrTypeFlags ─────────────────────────────────────────────
            s6 = sec("  ChrTypeFlags  (Default: 0 = all characters allowed)")
            _build_chrtype_section(s6)

            # ── GroundFlags / SystemFlags ─────────────────────────────────
            s7 = sec("  GroundFlags & SystemFlags  (Default: 0 — warning if changed)")
            _build_ground_system(s7)
            lbl_entry(s7, "OptionsEx: (Default: 0)", v_optex_raw, 10, _TOOLTIPS["OptionsEx"])
            dd_optex_frm  = tk.Frame(s7, bg=BG)
            man_optex_frm = tk.Frame(s7, bg=BG)
            _build_optionsex_section(dd_optex_frm, man_optex_frm)

            # ── Numeric Stats ─────────────────────────────────────────────
            s8 = sec("  Numeric Stats  (Weight, Value, MinLevel, Money, Ncash)")
            _build_numeric_stats(s8)

            # ── Equip Stats ───────────────────────────────────────────────
            s8b = sec("  Equip Stats  (AP, HP, MP, stat bonuses, recovery rates)")
            _build_equip_stats(s8b)

            # ── Effect ────────────────────────────────────────────────────
            s_eff = sec("  Effect  (slash-separated, e.g. 22 for boxes)  ⚠ Type 15 needs Effect 22")
            lbl_note(s_eff, "  Default: 0 for non-use items.")
            dd_eff_frm  = tk.Frame(s_eff, bg=BG)
            man_eff_frm = tk.Frame(s_eff, bg=BG)
            _build_effect_section(dd_eff_frm, man_eff_frm, s_eff)

            # ── Misc ─────────────────────────────────────────────────────
            s9 = sec("  Misc Fields  (EffectFlags2, SelRange, Life — Default: 0 unless needed)")
            lbl_entry(s9, "EffectFlags2: (Default: 0)", v_eff2,      10, _TOOLTIPS["EffectFlags2"])
            lbl_entry(s9, "SelRange: (Default: 0)",     v_sel_range, 10, _TOOLTIPS["SelRange"])

            # ── Drilling Parameters ───────────────────────────────────────
            s_drill = sec("  Drilling Parameters  (Life, Depth, Delay)")
            _build_drilling_params(s_drill)

            # ── Card Parameters ───────────────────────────────────────────
            s10 = sec("  Card Parameters  (Default: 0 for non-card items)")
            lbl_entry(s10, "CardNum:      (Default: 0)", v_cardnum,   8, _TOOLTIPS["CardNum"])
            lbl_entry(s10, "CardGenGrade: (Default: 0)", v_cardgrade, 8, _TOOLTIPS["CardGenGrade"])
            lbl_entry(s10, "CardGenParam: (Default: 0)", num_vars["cardgenparam"], 12, _TOOLTIPS["CardGenParam"])
            lbl_entry(s10, "DailyGenCnt: (Default: 0)", v_daily, 8, _TOOLTIPS["DailyGenCnt"])
            lbl_note(s10, "  Skill cards: CardNum=7, CardGenGrade=0, CardGenParam=0.000000")

            # ── Refine & Equipment ────────────────────────────────────────
            s11 = sec("  Refine & Equipment Params  ")
            _build_refine_section(s11)

            # ── CM / Shop / Part ─────────────────────────────────────────
            s12 = sec("  CM / Shop / PartFileName  (Default: 0 or blank)")
            lbl_entry(s12, "ShopFileName:  (Default: 0)", v_shopfn, 50, _TOOLTIPS["ShopFileName"])
            lbl_entry(s12, "ShopBundleNum: (Default: 0)", v_shopbn,  8, _TOOLTIPS["ShopBundleNum"])
            lbl_entry(s12, "PartFileName:  (Default: 0)", v_partfn, 50, _TOOLTIPS["PartFileName"])
            lbl_note(s12, "  ChrFTypeFlag and ChrGender are always 0 — hidden.")
            lbl_entry(s12, "ExistType:     (Default: 0)", v_exist_type, 10, _TOOLTIPS["ExistType"])
            lbl_note(s12, "  ExistType: 0=disabled, 1=timer/cannot stack simultaneously (sprints, boosters).")
            lbl_entry(s12, "GroupId:       (Default: 0)", v_group,  10, _TOOLTIPS["GroupId"])
            lbl_note(s12, "  NewCM, FamCM, Summary always 0/blank — hidden.")

            return (dd_class_frm, dd_type_frm, dd_sub_frm, dd_ift_frm,
                    man_class_frm, man_type_frm, man_sub_frm, man_ift_frm,
                    dd_opts_frm, man_opts_frm, dd_optex_frm, man_optex_frm,
                    dd_eff_frm, man_eff_frm)

        # ══════════════════════════════════════════════════════════════════
        # LAYOUT: SIMPLE — grouped by function
        # ══════════════════════════════════════════════════════════════════
        def _build_simple_layout():
            # ── Identity (includes ExistType) ─────────────────────────────
            s1 = sec("  Identity  ⚠ Class and Type MUST be set")
            lbl_entry(s1, "ID:", v_id, 14, "Item ID — must be unique. Required.")
            lbl_entry(s1, "ExistType:", v_exist_type, 10, _TOOLTIPS["ExistType"])
            lbl_note(s1, "  ExistType: 0=disabled, 1=timer/cannot stack (sprints, boosters).")
            dd_class_frm = tk.Frame(s1, bg=BG); dd_type_frm  = tk.Frame(s1, bg=BG)
            dd_sub_frm   = tk.Frame(s1, bg=BG); dd_ift_frm   = tk.Frame(s1, bg=BG)
            man_class_frm = tk.Frame(s1, bg=BG); man_type_frm = tk.Frame(s1, bg=BG)
            man_sub_frm   = tk.Frame(s1, bg=BG); man_ift_frm  = tk.Frame(s1, bg=BG)
            _build_dd_row(dd_class_frm,  "Class: *",    _CLASS_MAP,    v_class,
                          _TOOLTIPS["Class"]+"\n⚠ REQUIRED.")
            _build_dd_row(dd_type_frm,   "Type: *",     _TYPE_MAP,     v_type,
                          _TOOLTIPS["Type"]+"\n⚠ REQUIRED.")
            _build_dd_row(dd_sub_frm,    "SubType:",     _SUBTYPE_MAP,  v_sub,  _TOOLTIPS["SubType"])
            _build_dd_row(dd_ift_frm,    "ItemFType:",   _ITEMFTYPE_MAP,v_ift,  _TOOLTIPS["ItemFType"])
            _build_man_row(man_class_frm,"Class: *",     v_class, _TOOLTIPS["Class"])
            _build_man_row(man_type_frm, "Type: *",      v_type,  _TOOLTIPS["Type"])
            _build_man_row(man_sub_frm,  "SubType:",     v_sub,   _TOOLTIPS["SubType"])
            _build_man_row(man_ift_frm,  "ItemFType:",   v_ift,   _TOOLTIPS["ItemFType"])

            # ── Names ─────────────────────────────────────────────────────
            s2 = sec("  Names & Text  ")
            lbl_entry(s2, "Name:",        v_name,    40, _TOOLTIPS["Name"])
            lbl_entry(s2, "Comment:",     v_comment, 60, _TOOLTIPS["Comment"])
            lbl_entry(s2, "Use:",         v_use,     60, _TOOLTIPS["Use"])
            lbl_entry(s2, "Name_Eng:",    v_name_eng, 40, _TOOLTIPS["Name_Eng"])
            lbl_entry(s2, "Comment_Eng:", v_cmt_eng,  40, _TOOLTIPS["Comment_Eng"])

            # ── Files ─────────────────────────────────────────────────────
            s3 = sec("  Filepaths & Bundle Numbers  ")
            lbl_entry(s3, "FileName:",     v_fn,   50, _TOOLTIPS["FileName"])
            lbl_entry(s3, "BundleNum:",    v_bn,    8, _TOOLTIPS["BundleNum"])
            lbl_note(s3, "  InvFileName / InvBundleNum auto-copied.")
            lbl_entry(s3, "CmtFileName:",  v_cmtfn, 50, _TOOLTIPS["CmtFileName"])
            lbl_entry(s3, "CmtBundleNum:", v_cmtbn,  8, _TOOLTIPS["CmtBundleNum"])
            lbl_entry(s3, "EquipFileName:", v_equipfn, 50, _TOOLTIPS["EquipFileName"])

            # ── Options ───────────────────────────────────────────────────
            s4 = sec("  Options  (eItemOption — slash-separated flags e.g. 1/2/256)")
            dd_opts_frm  = tk.Frame(s4, bg=BG)
            man_opts_frm = tk.Frame(s4, bg=BG)
            _build_options_section(dd_opts_frm, man_opts_frm)

            # ── ChrTypeFlags + HideHat ────────────────────────────────────
            s6 = sec("  ChrTypeFlags  (Default: 0 = all characters allowed)")
            _build_chrtype_section(s6)
            s_hh = sec("  HideHat  (Default: 0 — per-character ear hide)")
            _build_hidehat_section(s_hh)

            # ── OptionsEx ─────────────────────────────────────────────────
            dd_optex_frm  = tk.Frame(C, bg=BG)
            man_optex_frm = tk.Frame(C, bg=BG)
            s5_host = sec("  OptionsEx  (Default: 0)")
            dd_optex_frm2  = tk.Frame(s5_host, bg=BG)
            man_optex_frm2 = tk.Frame(s5_host, bg=BG)
            _build_optionsex_section(dd_optex_frm2, man_optex_frm2)
            dd_optex_frm = dd_optex_frm2; man_optex_frm = man_optex_frm2

            # ── Numeric Stats ─────────────────────────────────────────────
            s8 = sec("  Numeric Stats  (Weight, Value, MinLevel, Money, Ncash)")
            _build_numeric_stats(s8)

            # ── Equip Stats ───────────────────────────────────────────────
            s8b = sec("  Equip Stats  (AP, HP, MP, stat bonuses, recovery rates)")
            _build_equip_stats(s8b)

            # ── Effect ────────────────────────────────────────────────────
            s_eff = sec("  Effect  (slash-separated)  ⚠ Type 15 boxes MUST include 22")
            lbl_note(s_eff, "  Default: 0 for non-use items.")
            dd_eff_frm  = tk.Frame(s_eff, bg=BG)
            man_eff_frm = tk.Frame(s_eff, bg=BG)
            _build_effect_section(dd_eff_frm, man_eff_frm, s_eff)

            # ── Drilling Parameters ───────────────────────────────────────
            s_drill = sec("  Drilling Parameters  (Life, Depth, Delay)")
            _build_drilling_params(s_drill)

            # ── Card Parameters ───────────────────────────────────────────
            s10 = sec("  Card Parameters  (Default: 0 for non-card items)")
            lbl_entry(s10, "CardNum:      (Default: 0)", v_cardnum,   8, _TOOLTIPS["CardNum"])
            lbl_entry(s10, "CardGenGrade: (Default: 0)", v_cardgrade, 8, _TOOLTIPS["CardGenGrade"])
            lbl_entry(s10, "CardGenParam: (Default: 0)", num_vars["cardgenparam"], 12, _TOOLTIPS["CardGenParam"])
            lbl_entry(s10, "DailyGenCnt: (Default: 0)", v_daily, 8, _TOOLTIPS["DailyGenCnt"])

            # ── Refine & Equipment ────────────────────────────────────────
            s11 = sec("  Refine & Equipment Params  ")
            _build_refine_section(s11)

            # ── Misc Fields (PaletteId, GroundFlags, SystemFlags) ─────────
            s_misc = sec("  Misc Fields  (Default: 0)")
            lbl_entry(s_misc, "PaletteId:   (Default: 0)", v_palette, 10, _TOOLTIPS["PaletteId"])
            lbl_entry(s_misc, "PivotID:     (Default: 0)", v_pivot,   10, _TOOLTIPS["PivotID"])
            _build_ground_system(s_misc)
            lbl_entry(s_misc, "EffectFlags2: (Default: 0)", v_eff2,      10, _TOOLTIPS["EffectFlags2"])
            lbl_entry(s_misc, "SelRange:     (Default: 0)", v_sel_range, 10, _TOOLTIPS["SelRange"])

            # ── CM / Shop ─────────────────────────────────────────────────
            s12 = sec("  CM / Shop / PartFileName  (Default: 0 or blank)")
            lbl_entry(s12, "ShopFileName:  (Default: 0)", v_shopfn, 50, _TOOLTIPS["ShopFileName"])
            lbl_entry(s12, "ShopBundleNum: (Default: 0)", v_shopbn,  8, _TOOLTIPS["ShopBundleNum"])
            lbl_entry(s12, "PartFileName:  (Default: 0)", v_partfn, 50, _TOOLTIPS["PartFileName"])
            lbl_entry(s12, "GroupId:       (Default: 0)", v_group,  10, _TOOLTIPS["GroupId"])

            return (dd_class_frm, dd_type_frm, dd_sub_frm, dd_ift_frm,
                    man_class_frm, man_type_frm, man_sub_frm, man_ift_frm,
                    dd_opts_frm, man_opts_frm, dd_optex_frm, man_optex_frm,
                    dd_eff_frm, man_eff_frm)

        # ══════════════════════════════════════════════════════════════════
        # LAYOUT: CUSTOM — drag-to-reorder sections
        # ══════════════════════════════════════════════════════════════════
        def _build_custom_layout():
            # Use normal layout as base, but wrap each section in a draggable frame
            # For now build same as normal — user can reorder by dragging section headers
            tk.Label(C, text="Custom Layout — drag section headers to reorder",
                     bg=BG, fg=ACC4, font=("Consolas", 9, "italic")).pack(
                     anchor="w", padx=14, pady=(4, 0))

            result = _build_normal_layout()

            # Make all mk_section frames draggable by their header labels
            # (basic drag-reorder: click header to grab, release to drop)
            _sections = [w for w in C.winfo_children()
                         if isinstance(w, tk.LabelFrame)]
            _drag_state = {"source": None, "y_start": 0}

            for sf in _sections:
                for child in sf.winfo_children():
                    if isinstance(child, tk.Label):
                        child.config(cursor="fleur")
                        def _on_press(e, frame=sf):
                            _drag_state["source"] = frame
                            _drag_state["y_start"] = e.y_root
                        def _on_release(e, frame=sf):
                            src = _drag_state["source"]
                            if src is None or src is frame: return
                            # Swap pack order
                            src_info = src.pack_info()
                            frame_info = frame.pack_info()
                            src.pack_forget(); frame.pack_forget()
                            # Re-insert in swapped order
                            frame.pack(fill="x", padx=8, pady=4)
                            src.pack(fill="x", padx=8, pady=4)
                            _drag_state["source"] = None
                        child.bind("<ButtonPress-1>", _on_press)
                        child.bind("<ButtonRelease-1>", _on_release)
                        break
            return result

        # ── Build chosen layout ───────────────────────────────────────────
        mode = layout_var.get()
        if mode == "simple":
            (dd_class_frm, dd_type_frm, dd_sub_frm, dd_ift_frm,
             man_class_frm, man_type_frm, man_sub_frm, man_ift_frm,
             dd_opts_frm, man_opts_frm, dd_optex_frm, man_optex_frm,
             dd_eff_frm, man_eff_frm) = _build_simple_layout()
        elif mode == "custom":
            (dd_class_frm, dd_type_frm, dd_sub_frm, dd_ift_frm,
             man_class_frm, man_type_frm, man_sub_frm, man_ift_frm,
             dd_opts_frm, man_opts_frm, dd_optex_frm, man_optex_frm,
             dd_eff_frm, man_eff_frm) = _build_custom_layout()
        else:
            (dd_class_frm, dd_type_frm, dd_sub_frm, dd_ift_frm,
             man_class_frm, man_type_frm, man_sub_frm, man_ift_frm,
             dd_opts_frm, man_opts_frm, dd_optex_frm, man_optex_frm,
             dd_eff_frm, man_eff_frm) = _build_normal_layout()

        # ── Mode show/hide ────────────────────────────────────────────────
        def _apply_mode(*_):
            dd = mode_var.get() == "dropdown"
            for f in (dd_class_frm, dd_type_frm, dd_sub_frm, dd_ift_frm):
                if dd:  f.pack(fill="x")
                else:   f.pack_forget()
            for f in (man_class_frm, man_type_frm, man_sub_frm, man_ift_frm):
                if not dd: f.pack(fill="x")
                else:      f.pack_forget()
            if dd:
                dd_opts_frm.pack(fill="x"); man_opts_frm.pack_forget()
                dd_optex_frm.pack(fill="x"); man_optex_frm.pack_forget()
                dd_eff_frm.pack(fill="x"); man_eff_frm.pack_forget()
            else:
                man_opts_frm.pack(fill="x"); dd_opts_frm.pack_forget()
                man_optex_frm.pack(fill="x"); dd_optex_frm.pack_forget()
                man_eff_frm.pack(fill="x"); dd_eff_frm.pack_forget()

        mode_var.trace_add("write", _apply_mode)
        _apply_mode()

        # ══════════════════════════════════════════════════════════════════
        # GATHER
        # ══════════════════════════════════════════════════════════════════
        def _gather():
            return {
                "id":            v_id.get().strip(),
                "class_val":     v_class.get().strip(),
                "type_val":      v_type.get().strip(),
                "subtype_val":   v_sub.get().strip(),
                "itemftype_val": v_ift.get().strip(),
                "name":          v_name.get(),
                "comment":       v_comment.get(),
                "use":           v_use.get(),
                "name_eng":      v_name_eng.get() or " ",
                "comment_eng":   v_cmt_eng.get() or " ",
                "file_name":     v_fn.get(),
                "bundle_num":    v_bn.get(),
                "cmt_file_name": v_cmtfn.get(),
                "cmt_bundle_num":v_cmtbn.get(),
                "equip_file_name": v_equipfn.get() or "",
                "pivot_id":      v_pivot.get(),
                "palette_id":    v_palette.get(),
                "options_flags": [v.get() for v in opts_vars],
                "options_raw_manual": v_opts_manual.get(),
                "options_ex":    "/".join(str(f) for f,v in sorted(optex_vars.items()) if v.get()) or "0",
                "hide_hat":      v_hide_hat.get(),
                "chr_type_flags":int(v_chr_raw.get() or 0),
                "ground_flags":  v_ground.get(),
                "system_flags":  v_system.get(),
                "exist_type":    v_exist_type.get(),
                "weight":        num_vars["weight"].get(),
                "value":         num_vars["value"].get(),
                "min_level":     num_vars["min_level"].get(),
                "money":         num_vars["money"].get(),
                "ncash":         num_vars["ncash"].get(),
                "ap":            num_vars["ap"].get(),
                "hp":            num_vars["hp"].get(),
                "hpcon":         num_vars["hpcon"].get(),
                "mp":            num_vars["mp"].get(),
                "mpcon":         num_vars["mpcon"].get(),
                "applus":        num_vars["applus"].get(),
                "acplus":        num_vars["acplus"].get(),
                "dxplus":        num_vars["dxplus"].get(),
                "maxmpplus":     num_vars["maxmpplus"].get(),
                "maplus":        num_vars["maplus"].get(),
                "mdplus":        num_vars["mdplus"].get(),
                "maxwtplus":     num_vars["maxwtplus"].get(),
                "daplus":        num_vars["daplus"].get(),
                "lkplus":        num_vars["lkplus"].get(),
                "maxhpplus":     num_vars["maxhpplus"].get(),
                "dpplus":        num_vars["dpplus"].get(),
                "hvplus":        num_vars["hvplus"].get(),
                "hprecoveryrate": num_vars["hprecoveryrate"].get(),
                "mprecoveryrate": num_vars["mprecoveryrate"].get(),
                "delay":          num_vars["delay"].get(),
                "cardgenparam":   num_vars["cardgenparam"].get(),
                "effect":         v_effect_raw.get(),
                "effect_flags2":  v_eff2.get(),
                "sel_range":      v_sel_range.get(),
                "life":           v_life.get(),
                "depth":          v_depth.get(),
                "cardnum":        v_cardnum.get(),
                "cardgengrade":   v_cardgrade.get(),
                "dailygencnt":    v_daily.get(),
                "refine_index":   v_refine_idx.get(),
                "refine_type":    v_refine_type.get(),
                "min_stat_type":  v_minstattype.get(),
                "min_stat_lv":    v_minstatLv.get(),
                "compound_slot":  v_compound.get(),
                "set_item_id":    v_setitem.get(),
                "reform_count":   v_reform.get(),
                "shop_file_name": v_shopfn.get() or " ",
                "shop_bundle_num":v_shopbn.get(),
                "part_file_name": v_partfn.get() or " ",
                "group_id":       v_group.get(),
                "input_mode":     mode_var.get(),
                "layout_mode":    layout_var.get(),
            }

        # ── Generate actions ──────────────────────────────────────────────
        def _generate():
            cfg = _gather()
            if not cfg["id"].strip():
                messagebox.showerror("Missing ID", "Please enter an Item ID."); return
            if cfg["class_val"] in ("", "0", "Unselected"):
                if not messagebox.askyesno("Confirm", "Class is Unselected / 0. Continue?"):
                    return
            if cfg["type_val"] in ("", "0", "Unselected"):
                if not messagebox.askyesno("Confirm", "Type is Unselected / 0. Continue?"):
                    return
            xml = build_generic_itemparam_row(cfg)
            cfg["_first_run_done"] = True
            _save_t6_settings(cfg)
            try: _set_last_id("t6_item", int(cfg["id"]))
            except: pass
            self._settings = cfg; self._first_run = False
            try:  type_val_i = int(cfg.get("type_val", 0))
            except: type_val_i = 0
            try:
                eff_str = str(cfg.get("effect", "0"))
                eff_set = {int(x.strip()) for x in eff_str.replace("/",",").split(",") if x.strip()}
            except: eff_set = set()
            present_xml = None
            if type_val_i == 15:
                if 22 not in eff_set:
                    messagebox.showwarning("Box Effect Warning",
                        "Type 15 (Useables/Boxes) but Effect 22 (Open Box) is NOT set.\n"
                        "Add Effect 22, or continue if intentional.")
                if messagebox.askyesno("PresentItemParam2",
                        "This is a box (Type 15).\nGenerate a PresentItemParam2 row too?"):
                    present_xml = self._ask_present_contents(cfg)
            self._show_output(xml, cfg, present_xml=present_xml)

        def _generate_and_next():
            cfg = _gather()
            if not cfg["id"].strip():
                messagebox.showerror("Missing ID", "Please enter an Item ID."); return
            xml = build_generic_itemparam_row(cfg)
            try:    cfg["id"] = str(int(cfg["id"]) + 1)
            except: pass
            cfg["_first_run_done"] = True
            _save_t6_settings(cfg)
            try: _set_last_id("t6_item", int(cfg["id"]) - 1)
            except: pass
            self._settings = cfg; self._first_run = False
            self._show_output(xml, cfg, auto_next=True, present_xml=None)

        def _import_spreadsheet():
            """Import a single ItemParam row (or batch) from CSV/Excel."""
            path = filedialog.askopenfilename(
                title="Import ItemParam from spreadsheet",
                filetypes=[
                    ("Spreadsheet", "*.csv *.xlsx *.xlsm *.xls"),
                    ("CSV", "*.csv"),
                    ("Excel", "*.xlsx *.xlsm *.xls"),
                    ("All", "*.*"),
                ],
                parent=self.root,
            )
            if not path: return
            ext = os.path.splitext(path)[1].lower()
            try:
                if ext in (".xlsx", ".xlsm", ".xls"):
                    _, raw_rows = _read_xlsx_rows(path)
                else:
                    _, raw_rows = _read_csv_rows(path)
            except Exception as e:
                messagebox.showerror("Import Error", f"Could not read file:\n{e}"); return
            if not raw_rows:
                messagebox.showwarning("Empty", "No data rows found."); return
            # Map first row to settings using a broad alias map
            _alias = {
                "id":"id","itemid":"id","class":"class_val","type":"type_val",
                "subtype":"subtype_val","itemftype":"itemftype_val",
                "name":"name","comment":"comment","use":"use",
                "nameeng":"name_eng","commenteng":"comment_eng",
                "filename":"file_name","bundlenum":"bundle_num",
                "cmtfilename":"cmt_file_name","cmtbundlenum":"cmt_bundle_num",
                "equipfilename":"equip_file_name",
                "pivotid":"pivot_id","paletteid":"palette_id",
                "options":"options_raw_manual","optionsex":"options_ex",
                "hidehat":"hide_hat","chrtypeflags":"chr_type_flags",
                "groundflags":"ground_flags","systemflags":"system_flags",
                "existtype":"exist_type","weight":"weight","value":"value",
                "minlevel":"min_level","money":"money","ncash":"ncash",
                "ap":"ap","hp":"hp","hpcon":"hpcon","mp":"mp","mpcon":"mpcon",
                "applus":"applus","acplus":"acplus","dxplus":"dxplus",
                "maxmpplus":"maxmpplus","maplus":"maplus","mdplus":"mdplus",
                "maxwtplus":"maxwtplus","daplus":"daplus","lkplus":"lkplus",
                "maxhpplus":"maxhpplus","dpplus":"dpplus","hvplus":"hvplus",
                "hprecoveryrate":"hprecoveryrate","mprecoveryrate":"mprecoveryrate",
                "effect":"effect","effectflags2":"effect_flags2","selrange":"sel_range",
                "life":"life","depth":"depth","delay":"delay",
                "cardnum":"cardnum","cardgengrade":"cardgengrade",
                "cardgenparam":"cardgenparam","dailygencnt":"dailygencnt",
                "refineindex":"refine_index","refinetype":"refine_type",
                "minstattype":"min_stat_type","minstatlv":"min_stat_lv",
                "compoundslot":"compound_slot","setitemid":"set_item_id",
                "reformcount":"reform_count","groupid":"group_id",
                "shopfilename":"shop_file_name","shopbundlenum":"shop_bundle_num",
                "partfilename":"part_file_name",
            }
            new_s = dict(self._settings)
            row = raw_rows[0]
            for col, val in row.items():
                key = _alias.get(re.sub(r"[^a-z0-9]", "", col.lower()))  # normalise HEADER only
                if key and str(val).strip():
                    new_s[key] = str(val).strip()  # value kept as-is — \\ and . preserved
            self._settings = new_s
            _save_t6_settings(new_s)
            if len(raw_rows) > 1:
                messagebox.showinfo("Import",
                    f"Loaded first row. File has {len(raw_rows)} rows total.\n"
                    "Generate row-by-row using Generate & Continue to process all.")
            self._build_editor()

        # ── Footer nav ────────────────────────────────────────────────────
        nav = tk.Frame(wrap, bg=BG2)
        nav.grid(row=2, column=0, sticky="ew")

        mk_btn(nav, "◀  Back", self._build_start_screen,
               color=BG4).pack(side="left", padx=14, pady=6)
        mk_btn(nav, "🗑  Clear / Reset", self._reset,
               color=BG4).pack(side="left", padx=4, pady=6)
        mk_btn(nav, "📥  Import Spreadsheet", _import_spreadsheet,
               color=BG4).pack(side="left", padx=4, pady=6)
        mk_btn(nav, "➕  Generate & Continue", _generate_and_next,
               color=GREEN, fg=BG2).pack(side="right", padx=4, pady=6)
        mk_btn(nav, "⚡  Generate", _generate,
               color=self.ACC, fg=BG2,
               font=("Consolas", 11, "bold")).pack(side="right", padx=14, pady=6)

    # ─────────────────────────────────────────────────────────────────────────
    def _ask_present_contents(self, cfg):
        """Simple dialog to collect item IDs for a PresentItemParam2 row."""
        win = tk.Toplevel(self.root)
        win.title("PresentItemParam2 — Box Contents")
        win.geometry("500x480")
        win.configure(bg=BG)
        win.grab_set()

        tk.Label(win, text="Box Drop Contents",
                 bg=BG, fg=ACC6, font=("Consolas",12,"bold")).pack(pady=(12,2), padx=12, anchor="w")
        tk.Label(win, text="Enter the item IDs that drop from this box (up to 20).",
                 bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(padx=12, anchor="w")

        sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True, padx=8, pady=4)
        canv, C = mk_scroll_canvas(sh, init_width=460)

        rows_frame = tk.Frame(C, bg=BG); rows_frame.pack(fill="x", padx=4, pady=4)
        item_id_vars = []; item_rate_vars = []; item_cnt_vars = []
        for n in range(1, 21):
            r = tk.Frame(rows_frame, bg=BG); r.pack(fill="x", pady=1)
            tk.Label(r, text=f"Drop {n:2d}:", bg=BG, fg=FG,
                     font=("Consolas",9), width=8).pack(side="left")
            vid = tk.StringVar(value="0")
            tk.Entry(r, textvariable=vid, width=10, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=2)
            tk.Label(r, text="Rate:", bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(side="left", padx=(6,0))
            vrate = tk.StringVar(value="50")
            tk.Entry(r, textvariable=vrate, width=6, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=2)
            tk.Label(r, text="Cnt:", bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(side="left", padx=(6,0))
            vcnt = tk.StringVar(value="1")
            tk.Entry(r, textvariable=vcnt, width=4, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=2)
            item_id_vars.append(vid); item_rate_vars.append(vrate); item_cnt_vars.append(vcnt)

        ptype_var = tk.IntVar(value=1)
        pt_frm = tk.Frame(C, bg=BG); pt_frm.pack(fill="x", padx=8, pady=(4,0))
        tk.Label(pt_frm, text="Present Type:", bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        for lbl, val in [("Normal (1)","1"),("Distributive (2)","2")]:
            tk.Radiobutton(pt_frm, text=lbl, variable=ptype_var, value=int(val),
                           bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                           font=("Consolas",9)).pack(side="left", padx=8)

        drop_cnt_var = tk.StringVar(value="1")
        dc_frm = tk.Frame(C, bg=BG); dc_frm.pack(fill="x", padx=8, pady=2)
        tk.Label(dc_frm, text="Drop Cnt:", bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        tk.Entry(dc_frm, textvariable=drop_cnt_var, width=6, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left", padx=6)

        result = [None]
        nav = tk.Frame(win, bg=BG2); nav.pack(fill="x", side="bottom")

        def _build():
            items = []
            for vid, vr, vc in zip(item_id_vars, item_rate_vars, item_cnt_vars):
                iid = vid.get().strip()
                if iid and iid != "0":
                    try: rate = int(vr.get())
                    except: rate = 50
                    try: cnt = int(vc.get())
                    except: cnt = 1
                    items.append({"id": iid, "rate": rate, "item_cnt": cnt})
            if not items:
                messagebox.showwarning("No Items", "Add at least one drop item."); return
            try: dc = int(drop_cnt_var.get())
            except: dc = 1
            result[0] = build_presentparam_row(
                cfg["id"], items, ptype_var.get(), dc, 50,
                item_cnts=[it["item_cnt"] for it in items],
                box_name=cfg.get("name",""))
            win.destroy()

        mk_btn(nav, "✓  Build Present Row", _build, color=GREEN, fg=BG2,
               font=("Consolas",10,"bold")).pack(side="left", padx=12, pady=6)
        mk_btn(nav, "Skip", win.destroy).pack(side="left", padx=4, pady=6)
        win.wait_window()
        return result[0]

    def _show_output(self, xml, cfg, auto_next=False,
                     _compound_rows=None, _exchange_rows=None, present_xml=None):
        self._clear()
        if _compound_rows is None: _compound_rows = []
        if _exchange_rows is None: _exchange_rows = []
        ACC = self.ACC

        wrap = tk.Frame(self, bg=BG)
        wrap.pack(fill="both", expand=True)
        wrap.grid_rowconfigure(0, weight=0)
        wrap.grid_rowconfigure(1, weight=1)
        wrap.grid_rowconfigure(2, weight=0)
        wrap.grid_columnconfigure(0, weight=1)

        hdr = tk.Frame(wrap, bg=BG2)
        hdr.grid(row=0, column=0, sticky="ew")
        tk.Label(hdr, text=f"✓  Generated: ID {cfg['id']}  —  {cfg.get('name','')}",
                 font=("Consolas", 13, "bold"), bg=BG2, fg=ACC, pady=8
                 ).pack(side="left", padx=15)

        # Warn if Effect != 22 but PresentItemParam is expected (box use)
        try:
            effect_val = int(cfg.get("effect", 0))
        except: effect_val = 0
        if effect_val != 0 and effect_val != 22:
            pass  # non-box item, no warning needed
        # If type is 15 (Useables/Boxes) and effect != 22 warn
        try: type_val = int(cfg.get("type_val", 0))
        except: type_val = 0
        if type_val == 15 and effect_val != 22:
            tk.Label(hdr, text="⚠ Effect should be 22 (Open Box) for Type 15 items!",
                     bg=BG2, fg=ACC3, font=("Consolas",9,"bold")).pack(side="right", padx=10)

        nb = ttk.Notebook(wrap)
        nb.grid(row=1, column=0, sticky="nsew", padx=8, pady=4)
        make_output_tab(nb, "itemparam row", xml, "itemparam_row.xml", self.root)
        if present_xml:
            make_output_tab(nb, "PresentItemParam2 row", present_xml,
                            "presentparam_row.xml", self.root)

        # Compound/exchange extra tabs
        if _compound_rows:
            cp_xml = "\n".join(r[0] for r in _compound_rows)
            cl_xml = "\n".join(r[1] for r in _compound_rows)
            make_output_tab(nb,"Compound_Potion rows",cp_xml,"Compound_Potion_rows.xml",self.root)
            make_output_tab(nb,"Compounder_Location rows",cl_xml,"Compounder_Spot_rows.xml",self.root)
        if _exchange_rows:
            es_xml = "\n".join(r[0] for r in _exchange_rows)
            el_xml = "\n".join(r[1] for r in _exchange_rows)
            make_output_tab(nb,"ExchangeShopContents rows",es_xml,"ExchangeShopContents_rows.xml",self.root)
            make_output_tab(nb,"Exchange_Location rows",el_xml,"Exchange_Location_rows.xml",self.root)

        nav = tk.Frame(wrap, bg=BG2)
        nav.grid(row=2, column=0, sticky="ew")

        def _add_ce():
            # Build item list: current item + any batch rows if present
            items = [{"id": cfg.get("id",""), "name": cfg.get("name",""),
                      "comment": cfg.get("comment","")}]
            # Also include any rows generated in this batch session
            for prev in getattr(self, "_batch_rows", []):
                iid = prev.get("id","")
                nm  = prev.get("name","")
                if iid and not any(it["id"]==iid for it in items):
                    items.append({"id": iid, "name": nm, "comment": prev.get("comment","")})
            _shop_rows_t6 = getattr(self, "_extra_shop_rows_t6", [])
            def _on_done(comp_cfgs, exch_cfgs, shop_cfgs):
                for c in comp_cfgs:
                    _compound_rows.append((build_compound_row(c),
                                           build_compound_location_row(c["compound_id"])))
                    try: _set_last_id("compound", int(c["compound_id"]))
                    except: pass
                for c in exch_cfgs:
                    _exchange_rows.append((build_exchange_row(c),
                                           build_exchange_location_row(c["exchange_id"])))
                    try: _set_last_id("exchange", int(c["exchange_id"]))
                    except: pass
                for c in shop_cfgs:
                    _shop_rows_t6.append(
                        build_shop_row(c["id"], c.get("count","100"), c.get("price","0")))
                self._extra_shop_rows_t6 = _shop_rows_t6
                self._show_output(xml, cfg, _compound_rows=_compound_rows,
                                  _exchange_rows=_exchange_rows)
            _show_multi_ce_picker(self.root, items, _on_done)

        def _export_all():
            default_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(), "libconfig"))
            folder = filedialog.askdirectory(title="Choose export folder (default: libconfig)",
                                             initialdir=default_dir)
            if not folder: folder = default_dir
            os.makedirs(folder, exist_ok=True)
            exports = [("itemparam_row.xml", xml)]
            if present_xml:
                exports.append(("presentparam_row.xml", present_xml))
            if _compound_rows:
                exports += [("Compound_Potion_rows.xml", "\n".join(r[0] for r in _compound_rows)),
                            ("Compounder_Spot_rows.xml", "\n".join(r[1] for r in _compound_rows))]
            if _exchange_rows:
                exports += [("ExchangeShopContents_rows.xml", "\n".join(r[0] for r in _exchange_rows)),
                            ("Exchange_Location_rows.xml", "\n".join(r[1] for r in _exchange_rows))]
            _sh_t6 = getattr(self, "_extra_shop_rows_t6", [])
            if _sh_t6:
                exports.append(("R_ShopItem_rows.xml", "\n".join(_sh_t6)))
            saved = []
            for fname, content in exports:
                if _APP_SETTINGS.get("timestamp_files", False):
                    import time as _time
                    ts = _time.strftime("%d%m%y-%S%M%H")
                    name, ext = os.path.splitext(fname)
                    fname = f"{name}_{ts}{ext}"
                with open(os.path.join(folder, fname), "w", encoding="utf-8") as f:
                    f.write(content)
                saved.append(fname)
            messagebox.showinfo("Export Complete", f"Saved to:\n{folder}\n\n" + "\n".join(saved))

        mk_btn(nav, "➕  New Item (next ID)", self._next_from_output(cfg),
               color=self.ACC, fg=BG2, font=("Consolas",10,"bold")).pack(
               side="left", padx=14, pady=6)
        mk_btn(nav, "⚗  Add Compound/Exchange", _add_ce, color=BG4).pack(
               side="left", padx=4, pady=6)
        def _import_ce_t6():
            def _on_compound_cfgs(cfgs):
                for ce_cfg in cfgs:
                    cpr = build_compound_row(ce_cfg)
                    clr = build_compound_location_row(ce_cfg["compound_id"])
                    _compound_rows.append((cpr, clr))
                self._show_output(xml, cfg,
                                  _compound_rows=_compound_rows,
                                  _exchange_rows=_exchange_rows)
            def _on_exchange_cfgs(cfgs):
                for ce_cfg in cfgs:
                    esr = build_exchange_row(ce_cfg)
                    elr = build_exchange_location_row(ce_cfg["exchange_id"])
                    _exchange_rows.append((esr, elr))
                self._show_output(xml, cfg,
                                  _compound_rows=_compound_rows,
                                  _exchange_rows=_exchange_rows)
            _ask_import_mode_then_file(self.root, _on_compound_cfgs, _on_exchange_cfgs)
        mk_btn(nav, "📥  Import CSV/Excel", _import_ce_t6, color=BG4).pack(
               side="left", padx=4, pady=6)
        mk_btn(nav, "💾  Export All", _export_all, color=GREEN, fg=BG2).pack(
               side="left", padx=4, pady=6)
        mk_btn(nav, "◀  Back to Edit", self._build_editor).pack(
               side="left", padx=4, pady=6)
        mk_btn(nav, "🗑  Reset All", self._reset).pack(side="right", padx=14, pady=6)
        if auto_next:
            self.after(10, self._build_editor)

    def _next_from_output(self, cfg):
        def _go():
            try: cfg["id"] = str(int(cfg["id"]) + 1)
            except: pass
            try: _set_last_id("t6_item", int(cfg["id"]) - 1)
            except: pass
            _save_t6_settings(cfg)
            self._settings = cfg
            self._build_editor()
        return _go

    def _show_reference(self):
        """Scrollable reference window — all lookup tables in XML field order."""
        win = tk.Toplevel(self.root)
        win.title("ItemParam Field Reference")
        win.geometry("920x700")
        win.configure(bg=BG)
        sh = tk.Frame(win, bg=BG); sh.pack(fill="both", expand=True)
        canv, cont = mk_scroll_canvas(sh, init_width=880)

        def rsec(title, required=False):
            f = tk.Frame(cont, bg=BG2, pady=4)
            f.pack(fill="x", padx=8, pady=(8, 2))
            color = ACC3 if required else BLUE
            tk.Label(f, text=title, bg=BG2, fg=color,
                     font=("Consolas", 10, "bold")).pack(anchor="w", padx=8)
            body = tk.Frame(cont, bg=BG); body.pack(fill="x", padx=12, pady=2)
            return body

        def rtable(parent, rows, cols=None):
            frm = tk.Frame(parent, bg=BG); frm.pack(anchor="w", pady=2, padx=4)
            if cols:
                for ci, c in enumerate(cols):
                    w = max(12, len(str(c)) + 2)
                    tk.Label(frm, text=c, bg=BG2, fg=BLUE,
                             font=("Consolas", 8, "bold"), width=w,
                             anchor="w").grid(row=0, column=ci, padx=2, pady=1, sticky="w")
            for ri, row in enumerate(rows):
                bg = BG if ri % 2 == 0 else BG2
                for ci, cell in enumerate(row):
                    w = max(12, len(str(cell)) + 2)
                    tk.Label(frm, text=str(cell), bg=bg, fg=FG,
                             font=("Consolas", 8), width=w,
                             anchor="w").grid(row=ri + (1 if cols else 0),
                                              column=ci, padx=2, pady=1, sticky="w")

        def rnote(parent, text):
            tk.Label(parent, text=text, bg=BG, fg=FG_GREY,
                     font=("Consolas", 8), justify="left", wraplength=840).pack(
                     anchor="w", padx=4, pady=(1, 4))

        # ── In XML order ─────────────────────────────────────────────────
        s = rsec("ID", required=False)
        rnote(s, "Unique item ID number. Must not duplicate any existing ID in the table.")

        s = rsec("Class  ⚠ REQUIRED — must not be 0", required=True)
        rtable(s, _CLASS_MAP, ["Value","Description"])

        s = rsec("Type  ⚠ REQUIRED — must not be 0", required=True)
        rtable(s, _TYPE_MAP, ["Value","Description"])

        s = rsec("SubType  (default 0)")
        rtable(s, _SUBTYPE_MAP, ["Value","Description"])
        rnote(s, "Values marked > in original docs are from ItemParamCM2.")

        s = rsec("ItemFType  (default 0)")
        rtable(s, _ITEMFTYPE_MAP, ["Value","Description"])

        s = rsec("Name / Comment / Use / Name_Eng / Comment_Eng")
        rnote(s, "All are CDATA text fields.  Name_Eng and Comment_Eng are usually a single space.")

        s = rsec("FileName / BundleNum / InvFileName / InvBundleNum")
        rnote(s,
            "FileName: path to sprite NRI (e.g. data\\item\\itm000.nri).\n"
            "BundleNum: sprite index starting at 0.  InvFileName and InvBundleNum are "
            "always identical to FileName and BundleNum — auto-copied by this tool.\n"
            "To find a sprite: open NRI in viewer → Animations tab → slot number = BundleNum + 1.")

        s = rsec("CmtFileName / CmtBundleNum")
        rnote(s, "Points to the item illustration shown in the tooltip window. "
                 "Different file from FileName. Same BundleNum +1 offset rule.")

        s = rsec("EquipFileName")
        rnote(s, "Path to equipment or drill model. Leave blank if not equipment/drill.")

        s = rsec("PivotID  (Default: 0)")
        rnote(s, "Source item ID reference — mainly used for equipment with multiple level/option variants.")

        s = rsec("PaletteId  (Default: 0)")
        rnote(s, "Palette ID. 0 for almost all items. Rare exceptions can be ignored.")

        s = rsec("Options  (eItemOption — flags OR together)")
        rtable(s, _OPTIONS_FULL, ["Flag Value", "Description"])
        rnote(s, "Common defaults to always include: 2 (Usable) and 32 (UsableToSelf) for use-items.")

        s = rsec("HideHat  (Default: 0)")
        rnote(s,
            "Defines per character-type which model hides its ear when this item is equipped.\n"
            "Uses the SAME flag values as ChrTypeFlags (see below).  Usually 0.")
        rtable(s, [
            ("Bunny",   1, 512,    262144),
            ("Buffalo", 2, 1024,   524288),
            ("Sheep",   4, 2048,   1048576),
            ("Dragon",  8, 4096,   2097152),
            ("Fox",    16, 8192,   4194304),
            ("Lion",   32, 16384,  8388608),
            ("Cat",    64, 32768,  16777216),
            ("Raccoon",124,65536,  33554432),
            ("Paula",  256,131072, 67108864),
        ], ["Character", "1st Job Flag", "2nd Job Flag", "3rd Job Flag"])

        s = rsec("ChrTypeFlags  (Default: 0)")
        rnote(s,
            "Character-type restrictions/events. Sum the flags for every character+job "
            "that should be allowed.  0 = no restriction (all characters can use).")
        rtable(s, [
            ("Bunny",   1, 512,    262144),
            ("Buffalo", 2, 1024,   524288),
            ("Sheep",   4, 2048,   1048576),
            ("Dragon",  8, 4096,   2097152),
            ("Fox",    16, 8192,   4194304),
            ("Lion",   32, 16384,  8388608),
            ("Cat",    64, 32768,  16777216),
            ("Raccoon",124,65536,  33554432),
            ("Paula",  256,131072, 67108864),
        ], ["Character", "1st Job Flag", "2nd Job Flag", "3rd Job Flag"])

        s = rsec("GroundFlags / SystemFlags  (ALWAYS 0 — Suggested 0)")
        rnote(s, "Never changed. Non-zero values may cause unintended behaviour.")

        s = rsec("OptionsEx  (flags OR together — Suggested 0)")
        rtable(s, _OPTIONSEX_MAP, ["Flag Value", "Description"])
        rnote(s, "Values can be combined. Example: 1/2 = random stat range + elemental property.")

        s = rsec("Weight / Value / MinLevel  (defaults: 1, 0, 1)")
        rnote(s, "Weight = WT stat cost.  Value = Galder NPC sell price.  MinLevel = level gate.")

        s = rsec("Effect  (action on use — Suggested 0 for non-use items)")
        rtable(s, _EFFECT_MAP, ["Value", "Description"])
        rnote(s, "⚠ Type 15 (Useables/Boxes) MUST have Effect 22 (Open Box) to be usable.")

        s = rsec("EffectFlags2 / SelRange / Depth / Delay  (ALWAYS 0 — Suggested 0)")
        rnote(s, "All left at 0 in standard items. SelRange has a non-zero use only for Beta Magic Cards.")

        s = rsec("Life")
        rnote(s, "Duration for timed items (EXP/TM Boosters) or drill life span.  0 = no limit.")

        s = rsec("AP / HP / HPCon / MP / MPCon / Money  (Suggested 0 unless item grants stats)")
        rnote(s,
            "AP = Gun/Throwing ATK (not the APPlus equip stat).\n"
            "HP/MP = recovery amount.  HPCon/MPCon = consumed amount.  "
            "Money = Galder from coupon.")

        s = rsec("APPlus / ACPlus / DXPlus / MAPlus / MDPlus / DPPlus / HVPlus / DAPlus / LKPlus / MaxHP/MP/WTPlus")
        rnote(s, "Static stat bonuses for equipment. All default 0.")

        s = rsec("HPRecoveryRate / MPRecoveryRate  (format: 0.000000)")
        rnote(s, "HP/MP regen rate — used on pets. Format: 0.000000.")

        s = rsec("CardNum / CardGenGrade / CardGenParam / DailyGenCnt  (Suggested 0 for non-cards)")
        rtable(s, [
            (0,"None/dummy/invalid","—","Lowest"),
            (1,"Boss Cards","5-6","—"),
            (2,"High Tier","4-5","—"),
            (3,"Low-mid","4-5","—"),
            (4,"Mid","4","—"),
            (5,"High-low Monster","4","—"),
            (6,"Mid-Low Monster","4","—"),
            (7,"Characters / Skill","3","—"),
            (8,"Low Monster","3","—"),
            (9,"NPCs","3","Highest"),
        ], ["CardNum","Target","Life","Fortune Rank"])
        rnote(s, "Skill cards: CardNum=7, CardGenGrade=0, CardGenParam=0.000000\n"
                 "Star cards: CardNum=1, CardGenGrade=0, span full range.")

        s = rsec("PartFileName  (Suggested blank/space)")
        rnote(s, "ItemParamCM2 only — path to fashion item model. Usually a single space.")

        s = rsec("ChrFTypeFlag / ChrGender / NewCM / FamCM / Summary  (ALWAYS 0 / blank)")
        rnote(s, "All left at 0 or blank. Not shown in editor.")

        s = rsec("ExistType")
        rnote(s, "0 = disabled.  1 = timer enabled / cannot stack (sprints, boosters, etc).")

        s = rsec("Ncash")
        rnote(s, "Cash shop (MyShop) price. 0 = not sold in cash shop.")

        s = rsec("ShopFileName / ShopBundleNum  (Suggested 0/blank)")
        rnote(s, "Points to promotional Cash Mall image. Same NRI +1 offset rule for BundleNum.")

        s = rsec("MinStatType / MinStatLv")
        rtable(s, _MINSTATTYPE_MAP, ["Value", "Description"])
        rnote(s, "MinStatLv = 0 shows no requirement.\n"
                 "⚠ Non-zero MinStatLv blocks equip-swapping while under the Rust debuff.")

        s = rsec("RefineIndex  (Default: 0)")
        rtable(s, _REFINEINDEX_MAP, ["Value", "Description"])

        s = rsec("RefineType  (Default: 0)")
        rtable(s, _REFINETYPE_MAP, ["Value", "Description"])

        s = rsec("CompoundSlot  (Default: 0)")
        rnote(s, "Number of compound slots.  Intended range: 0-5.  "
                 "Higher values cause UI errors (EE) when item is inspected.")

        s = rsec("SetItemID  (Default: 0)")
        rnote(s, "Equipment set ID reference for set bonuses.")

        s = rsec("ReformCount  (Default: 0)")
        rnote(s, "Reform count — possibly tied to Skins. Not fully researched.")

        s = rsec("GroupId  (ALWAYS 0)")
        rnote(s, "Always 0 in all standard items.")

        mk_btn(win, "Close", win.destroy, color=BG4).pack(pady=8)


    def _reset(self):
        if messagebox.askyesno("Reset", "Clear all fields and reset to defaults?"):
            _save_t6_settings({})
            self._settings = {}
            self._first_run = True
            self._build_start_screen()


# ══════════════════════════════════════════════════════════════════════════════
# COMBINED SHELL
# ══════════════════════════════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════════════════════════════
# XML UTILITY TOOLS — shared helpers
# ══════════════════════════════════════════════════════════════════════════════

import time as _time_mod

ACC9  = "#eba0ac"  # maroon-red  — Row Counter
ACC10 = "#a6e3a1"  # green       — Range Auditor
ACC11 = "#cba6f7"  # purple      — XML Compare
ACC12 = "#89b4fa"  # blue        — Data Extract
ACC13 = "#f9e2af"  # yellow      — Duplicator
ACC14 = "#fab387"  # peach       — Mass Variable Manipulation
ACC15 = "#94e2d5"  # teal        — Reorder XML
ACC16 = "#f38ba8"  # red         — ID Checker
ACC17 = "#b4befe"  # lavender    — Fix ItemParam


def _reports_dir():
    d = _APP_SETTINGS.get("reports_dir", os.path.join(os.getcwd(), "reports"))
    os.makedirs(d, exist_ok=True)
    return d


def _ts():
    return _time_mod.strftime("%Y%m%d_%H%M%S")


def _report_path(name):
    """Return a path in the reports folder, with timestamp appended."""
    base, ext = os.path.splitext(name)
    return os.path.join(_reports_dir(), f"{base}_{_ts()}{ext}")


def _write_report(name, text, parent=None):
    """Write text to a timestamped file in the reports dir and offer to show it."""
    path = _report_path(name)
    with open(path, "w", encoding="utf-8") as f:
        f.write(text)
    if messagebox.askyesno("Report saved",
            f"Saved:\n{path}\n\nOpen report in viewer?", parent=parent):
        _show_text_window(path, text, parent)
    return path


def _show_text_window(title, text, parent=None):
    """Display text in a scrollable popup."""
    win = tk.Toplevel(parent)
    win.title(title if len(title) < 80 else title[:77] + "…")
    win.configure(bg=BG); win.geometry("900x600")
    txt = scrolledtext.ScrolledText(win, font=("Consolas", 8), bg=BG2, fg=FG,
                                    wrap="none")
    txt.pack(fill="both", expand=True, padx=4, pady=4)
    txt.insert("1.0", text)
    txt.config(state="disabled")


# ── Shared XML row iterator ────────────────────────────────────────────────
def _iter_xml_rows(path):
    """Yield each <ROW>…</ROW> block as a string from the file at path."""
    buf = []; in_row = False
    with open(path, encoding="utf-8", errors="replace", newline="") as f:
        for line in f:
            if not in_row:
                if "<ROW>" in line:
                    in_row = True; buf = [line]
                # also yield non-row lines so caller knows structure
            else:
                buf.append(line)
                if "</ROW>" in line:
                    in_row = False
                    yield "".join(buf)


def _xml_tag_val(row_text, tag):
    m = re.search(rf"<{re.escape(tag)}>\s*(.*?)\s*</{re.escape(tag)}>",
                  row_text, re.DOTALL)
    return m.group(1).strip() if m else None


def _detect_row_tags(path):
    """Return sorted list of child tag names from first <ROW> in file."""
    for row in _iter_xml_rows(path):
        tags = re.findall(r"<([A-Za-z_][A-Za-z0-9_]*?)>", row)
        # deduplicate keeping order, skip ROW itself
        seen = set(); result = []
        for t in tags:
            if t != "ROW" and t not in seen:
                seen.add(t); result.append(t)
        return result
    return []


def _count_rows_in_file(path):
    """Count <ROW> occurrences in an XML file."""
    n = 0
    with open(path, encoding="utf-8", errors="replace") as f:
        for line in f:
            n += line.count("<ROW>")
    return n


def _update_rowcount_in_file(path):
    """Update RowCount='N' (ItemParam-style) or CHARACTER count="N" (libcmgds_e-style).
    Returns (old, new) or None if neither found."""
    text = open(path, encoding="utf-8", errors="replace").read()

    # ── libcmgds_e: <CHARACTER count="N"> — count <GOODS  occurrences ────
    char_pat = re.compile(r'(<CHARACTER\s+count=")(\d+)(")')
    mc = char_pat.search(text)
    if mc:
        goods_count = text.count("<GOODS ")
        old = int(mc.group(2))
        new_text = char_pat.sub(lambda x: x.group(1) + str(goods_count) + x.group(3), text, count=1)
        with open(path, "w", encoding="utf-8") as f:
            f.write(new_text)
        return old, goods_count

    # ── ItemParam-style: RowCount='N' ─────────────────────────────────────
    count = text.count("<ROW>")
    pat = re.compile(r'(RowCount\s*=\s*["\'])(\d+)(["\'])', re.IGNORECASE)
    m = pat.search(text)
    if not m:
        return None
    old = int(m.group(2))
    new_text = pat.sub(lambda x: x.group(1) + str(count) + x.group(3), text, count=1)
    with open(path, "w", encoding="utf-8") as f:
        f.write(new_text)
    return old, count


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 9 — Row Counter / RowCount Updater
# ══════════════════════════════════════════════════════════════════════════════

class Tool9(tk.Frame):
    ACC = ACC9

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._build()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build(self):
        self._clear()
        tk.Label(self, text="ROW COUNTER / ROWCOUNT UPDATER",
                 font=("Consolas", 16, "bold"), bg=BG, fg=self.ACC).pack(pady=(24,4))
        tk.Label(self,
                 text="Counts <ROW> elements in one XML or a whole folder.\n"
                      "If the root element has a RowCount='N' attribute, it is updated in-place.",
                 bg=BG, fg=FG_DIM, font=("Consolas", 9), justify="center").pack(pady=(0,12))

        bf = tk.Frame(self, bg=BG); bf.pack(pady=8)
        mk_btn(bf, "📄  Count Single File", self._count_file,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="left", padx=8)
        mk_btn(bf, "📁  Count Folder", self._count_folder,
               color=BG3).pack(side="left", padx=8)

        self._out_var = tk.StringVar(value="")
        out_frame = tk.Frame(self, bg=BG); out_frame.pack(fill="both", expand=True, padx=20, pady=8)
        self._txt = scrolledtext.ScrolledText(out_frame, font=("Consolas",9),
                                              bg=BG2, fg=FG, height=18)
        self._txt.pack(fill="both", expand=True)
        self._txt.config(state="disabled")

    def _log(self, s):
        self._txt.config(state="normal")
        self._txt.insert("end", s + "\n")
        self._txt.see("end"); self._txt.config(state="disabled")
        self.update_idletasks()

    def _count_file(self):
        path = filedialog.askopenfilename(
            title="Select XML file",
            filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not path: return
        self._txt.config(state="normal"); self._txt.delete("1.0","end"); self._txt.config(state="disabled")
        n = _count_rows_in_file(path)
        self._log(f"File: {os.path.basename(path)}")
        self._log(f"  <ROW> count: {n:,}")
        result = _update_rowcount_in_file(path)
        if result:
            old, new = result
            self._log(f"  RowCount updated: {old} → {new}")
        else:
            self._log("  (No RowCount attribute found — nothing to update)")

    def _count_folder(self):
        folder = filedialog.askdirectory(title="Select folder", parent=self.root)
        if not folder: return
        self._txt.config(state="normal"); self._txt.delete("1.0","end"); self._txt.config(state="disabled")
        xmls = [f for f in os.listdir(folder) if f.lower().endswith(".xml")]
        if not xmls:
            self._log("No XML files found in folder."); return
        total = 0
        for fname in sorted(xmls):
            p = os.path.join(folder, fname)
            n = _count_rows_in_file(p)
            total += n
            result = _update_rowcount_in_file(p)
            suffix = ""
            if result:
                old, new = result
                suffix = f"  [RowCount: {old} → {new}]"
            self._log(f"  {fname}: {n:,} rows{suffix}")
        self._log(f"\nTotal: {total:,} rows across {len(xmls)} files")


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 10 — Range Auditor + Used/Unused Values
# ══════════════════════════════════════════════════════════════════════════════

class Tool10(tk.Frame):
    ACC = ACC10

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._xml_path = None; self._tags = []
        self._build_load()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build_load(self):
        self._clear()
        tk.Label(self, text="RANGE AUDITOR & USED / UNUSED REPORT",
                 font=("Consolas", 15, "bold"), bg=BG, fg=self.ACC).pack(pady=(24,4))
        tk.Label(self,
                 text="Load an XML file, pick which fields to audit, and get a report\n"
                      "of value ranges and gaps saved to the Reports folder.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9), justify="center").pack(pady=(0,12))

        self._status = tk.StringVar(value="No file loaded")
        sf = mk_section(self, "  Step 1 — Load XML  ")
        tk.Label(sf, textvariable=self._status, bg=BG, fg=FG_GREY,
                 font=("Consolas",9)).pack(side="left", padx=10)
        mk_btn(sf, "📂  Load XML", self._load_xml,
               color=self.ACC, fg=BG2).pack(side="right", padx=8, pady=6)

    def _load_xml(self):
        path = filedialog.askopenfilename(
            title="Select XML", filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not path: return
        self._xml_path = path
        self._tags = _detect_row_tags(path)
        n = _count_rows_in_file(path)
        self._status.set(f"✓  {os.path.basename(path)}  ({n:,} rows, {len(self._tags)} fields)")
        self._build_picker()

    def _build_picker(self):
        # Keep load section, add picker below
        for w in list(self.winfo_children()):
            if hasattr(w, '_is_picker'): w.destroy()

        pf = tk.Frame(self, bg=BG); pf._is_picker = True; pf.pack(fill="both", expand=True, padx=12)

        tk.Label(pf, text="Step 2 — Select fields to audit:",
                 bg=BG, fg=FG, font=("Consolas",10,"bold")).pack(anchor="w", pady=(8,4))
        tk.Label(pf,
                 text="  Numeric fields → range analysis + gap list\n"
                      "  Text fields    → unique values list",
                 bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(anchor="w")

        # Scrollable checkboxes
        sh = tk.Frame(pf, bg=BG, height=220); sh.pack(fill="x", pady=4)
        sh.pack_propagate(False)
        canv = tk.Canvas(sh, bg=BG, bd=0, highlightthickness=0)
        sb = tk.Scrollbar(sh, orient="vertical", command=canv.yview)
        canv.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y"); canv.pack(side="left", fill="both", expand=True)
        inner = tk.Frame(canv, bg=BG); canv.create_window((0,0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canv.configure(scrollregion=canv.bbox("all")))

        self._chk_vars = {}
        cols = 4
        for i, tag in enumerate(self._tags):
            v = tk.BooleanVar(value=(tag.upper() in ("ID","LEVEL","LV")))
            self._chk_vars[tag] = v
            tk.Checkbutton(inner, text=tag, variable=v, bg=BG, fg=FG,
                           selectcolor=BG3, activebackground=BG,
                           font=("Consolas",8)).grid(row=i//cols, column=i%cols,
                                                     sticky="w", padx=6, pady=1)

        bf = tk.Frame(pf, bg=BG); bf.pack(pady=10)
        mk_btn(bf, "✓ All",   lambda: [v.set(True)  for v in self._chk_vars.values()], color=BG4).pack(side="left", padx=4)
        mk_btn(bf, "✗ None",  lambda: [v.set(False) for v in self._chk_vars.values()], color=BG4).pack(side="left", padx=4)

        nav = tk.Frame(pf, bg=BG); nav.pack(pady=8)
        mk_btn(nav, "📊  Generate Report", self._run_audit,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="left", padx=8)
        mk_btn(nav, "🔄  Load Different File", self._build_load,
               color=BG4).pack(side="left", padx=4)

    def _run_audit(self):
        selected = [t for t,v in self._chk_vars.items() if v.get()]
        if not selected:
            messagebox.showwarning("Nothing selected", "Select at least one field."); return

        # Collect values
        field_vals = {t: [] for t in selected}
        for row in _iter_xml_rows(self._xml_path):
            for t in selected:
                v = _xml_tag_val(row, t)
                if v is not None:
                    field_vals[t].append(v)

        lines = [f"Range Audit Report",
                 f"File: {os.path.basename(self._xml_path)}",
                 f"Generated: {_time_mod.strftime('%Y-%m-%d %H:%M:%S')}",
                 "=" * 60]

        for tag in selected:
            vals = field_vals[tag]
            lines.append(f"\n[{tag}]  ({len(vals)} values)")
            nums = []
            for v in vals:
                try: nums.append(int(v))
                except:
                    try: nums.append(float(v))
                    except: pass
            if nums and len(nums) == len(vals):
                # Numeric
                int_nums = [int(n) for n in nums if n == int(n)] if all(isinstance(n,float) for n in nums) else []
                sorted_n = sorted(set(int(n) for n in nums if float(n) == int(float(n))))
                if sorted_n:
                    lines.append(f"  Min: {sorted_n[0]}   Max: {sorted_n[-1]}   Distinct: {len(sorted_n)}")
                    # Used ranges
                    used_ranges = []; s = sorted_n[0]; p = sorted_n[0]
                    for n in sorted_n[1:]:
                        if n == p+1: p = n
                        else: used_ranges.append((s,p)); s = n; p = n
                    used_ranges.append((s,p))
                    lines.append("  Used ranges:")
                    for a,b in used_ranges:
                        lines.append(f"    {a}" if a==b else f"    {a} ~ {b}")
                    # Gaps
                    gaps = []
                    for a,b in zip(sorted_n, sorted_n[1:]):
                        if b - a > 1:
                            gaps.append((a+1, b-1))
                    if gaps:
                        lines.append("  Unused gaps:")
                        for a,b in gaps:
                            lines.append(f"    {a}" if a==b else f"    {a} ~ {b}")
                    else:
                        lines.append("  Unused gaps: none")
                else:
                    lines.append("  (float values — range summary only)")
                    lines.append(f"  Min: {min(nums):.4f}   Max: {max(nums):.4f}")
            else:
                # Text
                unique = sorted(set(vals))
                lines.append(f"  Unique values ({len(unique)}):")
                for uv in unique[:200]:
                    lines.append(f"    {uv}")
                if len(unique) > 200:
                    lines.append(f"    … and {len(unique)-200} more")

        report = "\n".join(lines)
        _write_report("range_audit.txt", report, self.root)


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 11 — XML Compare (two-file + CSV lookup)
# ══════════════════════════════════════════════════════════════════════════════

class Tool11(tk.Frame):
    ACC = ACC11

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._file1 = None; self._file2 = None; self._build()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build(self):
        self._clear()
        tk.Label(self, text="XML COMPARATOR",
                 font=("Consolas",16,"bold"), bg=BG, fg=self.ACC).pack(pady=(24,4))
        tk.Label(self,
                 text="Compare two XML files by field value.  Find rows in A not in B, and vice versa.\n"
                      "Optionally use a CSV lookup list to check which IDs from the CSV appear in the XML.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9), justify="center").pack(pady=(0,12))

        # File A
        self._a_var = tk.StringVar(value="No file loaded")
        sf_a = mk_section(self, "  File A  (base)  ")
        tk.Label(sf_a, textvariable=self._a_var, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf_a, "📂  Load", lambda: self._pick("a"), color=self.ACC, fg=BG2).pack(side="right",padx=8,pady=6)

        # File B
        self._b_var = tk.StringVar(value="No file loaded")
        sf_b = mk_section(self, "  File B  (compare against)  ")
        tk.Label(sf_b, textvariable=self._b_var, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf_b, "📂  Load", lambda: self._pick("b"), color=BG3).pack(side="right",padx=8,pady=6)

        # Key field
        kf = mk_section(self, "  Key Field (the tag to compare by, e.g. ID)  ")
        self._key_var = tk.StringVar(value="ID")
        tk.Entry(kf, textvariable=self._key_var, width=18, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",10), relief="flat").pack(side="left",padx=10,pady=6)
        tk.Label(kf, text="  Also compare full row content (slow):", bg=BG, fg=FG_DIM,
                 font=("Consolas",8)).pack(side="left",padx=4)
        self._deep_var = tk.BooleanVar(value=False)
        tk.Checkbutton(kf, variable=self._deep_var, bg=BG, fg=FG,
                       selectcolor=BG3, activebackground=BG).pack(side="left")

        # CSV lookup (optional)
        self._csv_var = tk.StringVar(value="None  (optional)")
        sf_csv = mk_section(self, "  CSV Lookup List  (optional — column 'ID' or first column)  ")
        tk.Label(sf_csv, textvariable=self._csv_var, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf_csv, "📂  Load CSV", self._pick_csv, color=BG3).pack(side="right",padx=8,pady=6)
        mk_btn(sf_csv, "✗  Clear", lambda: (setattr(self,"_csv_path",None), self._csv_var.set("None  (optional)")),
               color=BG4).pack(side="right",padx=4,pady=6)
        self._csv_path = None

        nav = tk.Frame(self, bg=BG); nav.pack(pady=14)
        mk_btn(nav, "🔍  Compare Files", self._run,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="left", padx=8)

    def _pick(self, which):
        p = filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not p: return
        if which == "a":
            self._file1 = p; self._a_var.set(f"✓  {os.path.basename(p)}")
            # auto-fill key from first row
            tags = _detect_row_tags(p)
            if tags: self._key_var.set(tags[0])
        else:
            self._file2 = p; self._b_var.set(f"✓  {os.path.basename(p)}")

    def _pick_csv(self):
        p = filedialog.askopenfilename(
            filetypes=[("CSV/Spreadsheet","*.csv *.xlsx *.xlsm *.xls"),("All","*.*")], parent=self.root)
        if not p: return
        self._csv_path = p; self._csv_var.set(f"✓  {os.path.basename(p)}")

    def _run(self):
        if not self._file1 or not self._file2:
            messagebox.showwarning("Missing files","Load both File A and File B."); return
        key = self._key_var.get().strip() or "ID"

        def _get_map(path):
            m = {}
            for row in _iter_xml_rows(path):
                v = _xml_tag_val(row, key)
                if v is not None:
                    m[v] = row
            return m

        map_a = _get_map(self._file1)
        map_b = _get_map(self._file2)

        only_a = sorted(set(map_a) - set(map_b))
        only_b = sorted(set(map_b) - set(map_a))
        both   = sorted(set(map_a) & set(map_b))

        diff_rows = []
        if self._deep_var.get():
            for k in both:
                if map_a[k].strip() != map_b[k].strip():
                    diff_rows.append(k)

        lines = [
            "XML Comparison Report",
            f"File A: {os.path.basename(self._file1)}",
            f"File B: {os.path.basename(self._file2)}",
            f"Key field: <{key}>",
            f"Generated: {_time_mod.strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 60,
            f"\nRows ONLY in A ({len(only_a)}):"]
        lines += [f"  {k}" for k in only_a] or ["  (none)"]
        lines += [f"\nRows ONLY in B ({len(only_b)}):"]
        lines += [f"  {k}" for k in only_b] or ["  (none)"]
        lines += [f"\nIn both: {len(both)}"]
        if self._deep_var.get():
            lines += [f"\nRows with DIFFERENT content ({len(diff_rows)}):"]
            lines += [f"  {k}" for k in diff_rows] or ["  (none)"]

        # CSV lookup
        if self._csv_path:
            csv_ids = set()
            ext = os.path.splitext(self._csv_path)[1].lower()
            if ext in (".xlsx",".xlsm",".xls"):
                if _HAVE_OPENPYXL:
                    wb = openpyxl.load_workbook(self._csv_path, data_only=True)
                    ws = wb.active
                    rows_raw = list(ws.iter_rows(values_only=True))
                    headers = [str(c).strip() if c else "" for c in rows_raw[0]]
                    id_ci = next((i for i,h in enumerate(headers) if h.upper()=="ID"), 0)
                    for row in rows_raw[1:]:
                        v = row[id_ci]
                        if v: csv_ids.add(str(v).strip())
            else:
                with open(self._csv_path, encoding="utf-8-sig") as f:
                    reader = csv.DictReader(f)
                    id_key = "ID" if "ID" in (reader.fieldnames or []) else (reader.fieldnames or ["ID"])[0]
                    for row in reader:
                        v = row.get(id_key,"").strip()
                        if v: csv_ids.add(v)

            in_xml   = csv_ids & set(map_a)
            not_in   = csv_ids - set(map_a)
            lines += [
                f"\nCSV Lookup against File A ({len(csv_ids)} IDs in CSV):",
                f"  Found in XML: {len(in_xml)}",
                f"  Missing from XML ({len(not_in)}):"]
            lines += [f"    {i}" for i in sorted(not_in, key=lambda x: (not x.isdigit(), int(x) if x.isdigit() else x))]

        report = "\n".join(lines)
        _write_report("xml_compare.txt", report, self.root)


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 12 — Data Extract / Field Reporter
# ══════════════════════════════════════════════════════════════════════════════

class Tool12(tk.Frame):
    ACC = ACC12

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._xml_path = None; self._tags = []; self._build_load()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build_load(self):
        self._clear()
        tk.Label(self, text="DATA EXTRACT / FIELD REPORTER",
                 font=("Consolas",16,"bold"), bg=BG, fg=self.ACC).pack(pady=(24,4))
        tk.Label(self,
                 text="Load an XML, pick which fields to export.\n"
                      "Optionally filter rows by a field value. Output as CSV to Reports folder.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9), justify="center").pack(pady=(0,12))
        sf = mk_section(self, "  Load XML  ")
        self._status = tk.StringVar(value="No file loaded")
        tk.Label(sf, textvariable=self._status, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf, "📂  Load XML", self._load, color=self.ACC, fg=BG2).pack(side="right",padx=8,pady=6)

    def _load(self):
        p = filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not p: return
        self._xml_path = p
        self._tags = _detect_row_tags(p)
        n = _count_rows_in_file(p)
        self._status.set(f"✓  {os.path.basename(p)}  ({n:,} rows)")
        self._build_picker()

    def _build_picker(self):
        for w in list(self.winfo_children()):
            if hasattr(w,'_picker'): w.destroy()

        pf = tk.Frame(self, bg=BG); pf._picker=True; pf.pack(fill="both",expand=True,padx=12)

        # Field checkboxes
        tk.Label(pf, text="Fields to include in output:",
                 bg=BG, fg=FG, font=("Consolas",10,"bold")).pack(anchor="w",pady=(8,4))
        sh = tk.Frame(pf, bg=BG, height=180); sh.pack(fill="x", pady=2); sh.pack_propagate(False)
        canv = tk.Canvas(sh, bg=BG, bd=0, highlightthickness=0)
        sb = tk.Scrollbar(sh, orient="vertical", command=canv.yview)
        canv.configure(yscrollcommand=sb.set)
        sb.pack(side="right",fill="y"); canv.pack(side="left",fill="both",expand=True)
        inner = tk.Frame(canv,bg=BG); canv.create_window((0,0),window=inner,anchor="nw")
        inner.bind("<Configure>", lambda e: canv.configure(scrollregion=canv.bbox("all")))
        self._field_vars = {}
        cols = 4
        for i,t in enumerate(self._tags):
            v = tk.BooleanVar(value=True)
            self._field_vars[t] = v
            tk.Checkbutton(inner, text=t, variable=v, bg=BG, fg=FG,
                           selectcolor=BG3, activebackground=BG,
                           font=("Consolas",8)).grid(row=i//cols,column=i%cols,sticky="w",padx=4,pady=1)

        brow = tk.Frame(pf,bg=BG); brow.pack(anchor="w",pady=2)
        mk_btn(brow,"✓ All",   lambda:[v.set(True)  for v in self._field_vars.values()],color=BG4).pack(side="left",padx=4)
        mk_btn(brow,"✗ None",  lambda:[v.set(False) for v in self._field_vars.values()],color=BG4).pack(side="left",padx=4)

        # Filter
        filt = mk_section(pf, "  Filter Rows  (optional)  ")
        fr = tk.Frame(filt,bg=BG); fr.pack(fill="x",padx=8,pady=6)
        tk.Label(fr,text="Where field:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        self._filt_tag = tk.StringVar(value=self._tags[0] if self._tags else "")
        tag_menu = ttk.Combobox(fr, textvariable=self._filt_tag,
                                values=self._tags, width=16, state="readonly")
        tag_menu.pack(side="left",padx=4)
        tk.Label(fr,text="contains:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left",padx=4)
        self._filt_val = tk.StringVar()
        tk.Entry(fr,textvariable=self._filt_val,width=20,bg=BG3,fg=FG,
                 insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=4)
        tk.Label(fr,text="(blank = no filter)",bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(side="left")

        nav = tk.Frame(pf,bg=BG); nav.pack(pady=10)
        mk_btn(nav,"📊  Extract to CSV",self._run,
               color=self.ACC,fg=BG2,font=("Consolas",11,"bold")).pack(side="left",padx=8)
        mk_btn(nav,"👁  Preview in-app",lambda:self._run(preview=True),color=BG3).pack(side="left",padx=4)

    def _run(self, preview=False):
        fields = [t for t,v in self._field_vars.items() if v.get()]
        if not fields:
            messagebox.showwarning("Nothing selected","Pick at least one field."); return
        filt_tag = self._filt_tag.get()
        filt_val = self._filt_val.get().strip().lower()

        rows_out = []
        for row in _iter_xml_rows(self._xml_path):
            if filt_val:
                v = (_xml_tag_val(row, filt_tag) or "").lower()
                if filt_val not in v: continue
            row_data = [_xml_tag_val(row, f) or "" for f in fields]
            rows_out.append(row_data)

        if not rows_out:
            messagebox.showinfo("No results","No rows matched the filter."); return

        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(fields)
        w.writerows(rows_out)
        csv_text = out.getvalue()

        if preview:
            _show_text_window(f"Data Extract — {os.path.basename(self._xml_path)}", csv_text, self.root)
        else:
            _write_report("data_extract.csv", csv_text, self.root)


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 13 — XML Row Duplicator
# ══════════════════════════════════════════════════════════════════════════════

def _apply_eq(text, expr):
    """Apply equation to the numeric part of text. Supports +N, -N, y+N, (y+N)^2, etc."""
    m = re.search(r'(\d+)', text)
    if m:
        prefix = text[:m.start()]; num = int(m.group(1)); suffix = text[m.end():]
        if num == 0 and 'y' in expr: num = 1
    else:
        prefix = text; num = 1; suffix = ""
    e = expr.strip()
    if not e: return text
    if e[0] in "+-": e = "y" + e
    e = e.replace("^", "**")
    try:
        result = eval(e, {"__builtins__": {}}, {"y": num})
        return prefix + str(int(round(result))) + suffix
    except Exception as ex:
        raise ValueError(f"Bad equation '{expr}': {ex}")


class Tool13(tk.Frame):
    ACC = ACC13

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._files = []; self._tags = []; self._build()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build(self):
        self._clear()
        tk.Label(self, text="XML ROW DUPLICATOR",
                 font=("Consolas",16,"bold"), bg=BG, fg=self.ACC).pack(pady=(20,4))
        tk.Label(self,
                 text="Load one or more XML files, configure per-field equations (+1, y*2, etc.),\n"
                      "and generate N clones of each existing row with transformed field values.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9), justify="center").pack(pady=(0,8))

        sf = mk_section(self, "  Step 1 — Load XML File(s)  ")
        self._file_status = tk.StringVar(value="No files loaded")
        tk.Label(sf, textvariable=self._file_status, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf, "📂  Load Files", self._load_files,
               color=self.ACC, fg=BG2).pack(side="right",padx=8,pady=6)

        sf2 = mk_section(self, "  Step 2 — Detect Fields  ")
        mk_btn(sf2, "🔍  Detect from first file", self._detect,
               color=BG3).pack(side="left",padx=8,pady=6)
        self._detect_status = tk.StringVar(value="")
        tk.Label(sf2, textvariable=self._detect_status, bg=BG, fg=FG_GREY, font=("Consolas",8)).pack(side="left",padx=4)

        # Repeat count
        rf = mk_section(self, "  Clones per original row  ")
        self._repeat = tk.IntVar(value=3)
        tk.Spinbox(rf, from_=1, to=9999, textvariable=self._repeat, width=8,
                   bg=BG3, fg=FG, insertbackground=FG, font=("Consolas",10),
                   relief="flat").pack(side="left",padx=10,pady=6)

        # Fields area (scrollable, built after detect)
        self._fields_host = tk.Frame(self, bg=BG)
        self._fields_host.pack(fill="both", expand=True, padx=8)

        nav = tk.Frame(self, bg=BG2); nav.pack(fill="x", side="bottom")
        mk_btn(nav, "⚡  Generate", self._run,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="right",padx=14,pady=6)

    def _load_files(self):
        paths = filedialog.askopenfilenames(
            filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not paths: return
        self._files = list(paths)
        self._file_status.set(f"✓  {len(self._files)} file(s) loaded")

    def _detect(self):
        if not self._files:
            messagebox.showwarning("No files","Load files first."); return
        self._tags = _detect_row_tags(self._files[0])
        self._detect_status.set(f"{len(self._tags)} fields detected")
        self._build_fields()

    def _build_fields(self):
        for w in self._fields_host.winfo_children(): w.destroy()
        tk.Label(self._fields_host,
                 text="  Field       │  Equation (blank=unchanged, +1, y+1, (y+1)^2, etc.)",
                 bg=BG, fg=BLUE, font=("Consolas",8,"bold")).pack(anchor="w",padx=4,pady=(4,2))
        tk.Label(self._fields_host,
                 text="  Equations use 'y' for the current numeric value. +N and -N are shorthand for y+N.",
                 bg=BG, fg=FG_GREY, font=("Consolas",7)).pack(anchor="w",padx=4)

        sh = tk.Frame(self._fields_host, bg=BG, height=200)
        sh.pack(fill="x", pady=4); sh.pack_propagate(False)
        canv = tk.Canvas(sh, bg=BG, bd=0, highlightthickness=0)
        sb = tk.Scrollbar(sh, orient="vertical", command=canv.yview)
        canv.configure(yscrollcommand=sb.set)
        sb.pack(side="right",fill="y"); canv.pack(side="left",fill="both",expand=True)
        inner = tk.Frame(canv,bg=BG); canv.create_window((0,0),window=inner,anchor="nw")
        inner.bind("<Configure>", lambda e: canv.configure(scrollregion=canv.bbox("all")))

        self._eq_vars = {}
        for tag in self._tags:
            r = tk.Frame(inner,bg=BG); r.pack(fill="x",padx=4,pady=1)
            tk.Label(r, text=tag, width=22, anchor="w", bg=BG, fg=FG,
                     font=("Consolas",9)).pack(side="left")
            v = tk.StringVar()
            tk.Entry(r, textvariable=v, width=28, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left",padx=4)
            self._eq_vars[tag] = v

    def _run(self):
        if not self._files: messagebox.showwarning("No files","Load files first."); return
        if not self._tags: messagebox.showwarning("No fields","Detect fields first."); return
        repeat = max(1, self._repeat.get())
        eqs = {t: v.get().strip() for t,v in self._eq_vars.items()}

        default_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(),"libconfig"))
        out_folder = filedialog.askdirectory(title="Save duplicated files to…",
                                             initialdir=default_dir, parent=self.root)
        if not out_folder: return

        results = []
        for path in self._files:
            out_lines = []
            in_row = False; buf = []
            with open(path, encoding="utf-8", errors="replace", newline="") as f:
                content = f.read()

            # Pass through non-ROW lines unchanged; duplicate each ROW
            out_parts = []
            in_row = False; buf = []
            for line in content.splitlines(keepends=True):
                if not in_row:
                    if "<ROW>" in line:
                        in_row = True; buf = [line]
                    else:
                        out_parts.append(line)
                else:
                    buf.append(line)
                    if "</ROW>" in line:
                        in_row = False
                        row_text = "".join(buf)
                        out_parts.append(row_text)
                        # last known values per tag
                        last = {}
                        for t in self._tags:
                            v = _xml_tag_val(row_text, t)
                            if v is not None: last[t] = v
                        for _ in range(repeat):
                            clone = row_text
                            for t,eq in eqs.items():
                                if not eq: continue
                                cur = last.get(t,"")
                                try:
                                    new_val = _apply_eq(cur, eq)
                                    # Replace tag value in clone
                                    clone = re.sub(
                                        rf"(<{re.escape(t)}>)(.*?)(</{re.escape(t)}>)",
                                        lambda m, nv=new_val: m.group(1)+nv+m.group(3),
                                        clone, count=1, flags=re.DOTALL)
                                    last[t] = new_val
                                except ValueError as e:
                                    messagebox.showerror("Equation error", str(e)); return
                            out_parts.append(clone)

            out_text = "".join(out_parts)
            stem = os.path.splitext(os.path.basename(path))[0]
            out_name = stem + "_duplicated.xml"
            out_path = os.path.join(out_folder, out_name)
            with open(out_path, "w", encoding="utf-8") as f:
                f.write(out_text)
            results.append(out_name)

        messagebox.showinfo("Done",
            f"Generated {len(results)} file(s):\n" + "\n".join(results[:10]))


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 14 — Mass Variable Manipulation
# ══════════════════════════════════════════════════════════════════════════════

class Tool14(tk.Frame):
    ACC = ACC14
    _WARNED_KEY = "t14_warned"

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._xml_path = None; self._tags = []
        # Show first-open warning once per install
        if not _load_settings(self._WARNED_KEY).get("warned"):
            self.after(200, self._first_open_warning)
        self._build()

    def _first_open_warning(self):
        messagebox.showwarning(
            "⚠  Mass Variable Manipulation — Read First",
            "This tool modifies XML field values across all (or filtered) rows.\n\n"
            "IMPORTANT:\n"
            "• Always keep a backup of your XML before applying changes.\n"
            "• Regex and text replacement operate on raw field content.\n"
            "• CDATA tags and decimal precision are preserved where possible,\n"
            "  but complex patterns can still corrupt your output.\n"
            "• Use Preview before applying to verify results.\n\n"
            "Advanced Manual Renaming (regex patterns) can be enabled in ⚙ Settings.",
            parent=self.root)
        _save_settings(self._WARNED_KEY, {"warned": True})

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build(self):
        self._clear()
        outer = tk.Frame(self,bg=BG); outer.pack(fill="both",expand=True)
        sh = tk.Frame(outer,bg=BG); sh.pack(fill="both",expand=True)
        canv, C = mk_scroll_canvas(sh)

        tk.Label(C, text="MASS VARIABLE MANIPULATION",
                 font=("Consolas",15,"bold"), bg=BG, fg=self.ACC).pack(pady=(20,4))
        tk.Label(C,
                 text="Alter any field across all (or filtered) rows.  Math, text replace, conditional.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(pady=(0,8))

        # Load XML
        sf = mk_section(container, "  Load XML  ")
        self._status = tk.StringVar(value="No file loaded")
        tk.Label(sf, textvariable=self._status, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf, "📂  Load XML", self._load, color=self.ACC, fg=BG2).pack(side="right",padx=8,pady=6)

        # Target field
        tf = mk_section(container, "  Target Field  (field to modify)")
        tf_row = tk.Frame(tf,bg=BG); tf_row.pack(fill="x",padx=8,pady=6)
        tk.Label(tf_row, text="Field:", bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        self._tgt_var = tk.StringVar()
        self._tgt_menu = ttk.Combobox(tf_row, textvariable=self._tgt_var, width=20, state="readonly")
        self._tgt_menu.pack(side="left",padx=6)

        # Operation
        op_f = mk_section(container, "  Operation")
        self._op_var = tk.StringVar(value="math")
        ops = [("Math expression  (e.g. +1, y*2, round(y/100)*100)", "math"),
               ("Text replace  (literal)", "replace"),
               ("Regex replace", "regex"),
               ("Friendly regex  (contains / starts-with / ends-with)", "friendly"),
               ("Set to fixed value", "set")]
        for lbl, val in ops:
            tk.Radiobutton(C, text=lbl, variable=self._op_var, value=val,
                           bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                           font=("Consolas",9)).pack(anchor="w", padx=16, pady=1)

        # Expression/value
        expr_f = mk_section(container, "  Expression / Value / Pattern → Replacement")
        er = tk.Frame(expr_f,bg=BG); er.pack(fill="x",padx=8,pady=6)
        tk.Label(er, text="From / Expr:", bg=BG, fg=FG, font=("Consolas",9), width=14, anchor="w").pack(side="left")
        self._from_var = tk.StringVar()
        tk.Entry(er, textvariable=self._from_var, width=36, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left",padx=4)
        tk.Label(er, text="To:", bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(side="left",padx=(8,2))
        self._to_var = tk.StringVar()
        tk.Entry(er, textvariable=self._to_var, width=20, bg=BG3, fg=FG,
                 insertbackground=FG, font=("Consolas",9), relief="flat").pack(side="left",padx=2)
        tk.Label(C, text="  For Math: use 'y' for current value.  Preserve decimal places: checked below.",
                 bg=BG, fg=FG_GREY, font=("Consolas",7)).pack(anchor="w",padx=16)

        dec_f = tk.Frame(C,bg=BG); dec_f.pack(anchor="w",padx=16,pady=2)
        self._preserve_dec = tk.BooleanVar(value=True)
        tk.Checkbutton(dec_f, text="Preserve decimal place count  (e.g. 1.530000 stays 6dp)",
                       variable=self._preserve_dec, bg=BG, fg=FG,
                       selectcolor=BG3, activebackground=BG, font=("Consolas",8)).pack(side="left")

        # Regex reference guide — always visible
        ref_bar = tk.Frame(C, bg=BG); ref_bar.pack(anchor="w", padx=16, pady=(4, 2))

        def _open_regex_ref():
            win = tk.Toplevel(self.root)
            win.title("📖  Regex & Pattern Reference Guide")
            win.geometry("720x560"); win.configure(bg=BG)
            txt = scrolledtext.ScrolledText(win, font=("Consolas", 8), bg=BG2, fg=FG, wrap="word")
            txt.pack(fill="both", expand=True, padx=8, pady=8)
            ref = (
                "REGEX & PATTERN REFERENCE  —  Mass Variable Manipulation\n"
                "══════════════════════════════════════════════════════════════\n\n"
                "OPERATION MODES\n"
                "───────────────\n"
                "  Math expression   Use 'y' for the current numeric value.\n"
                "    Examples:  y+1   y*2   round(y/100)*100   y-500\n"
                "    Preserve decimal:  1.530000 → stays 6dp if option checked.\n\n"
                "  Text replace (literal)\n"
                "    From: exact text to find.   To: replacement text.\n"
                "    Case-sensitive, no wildcards.\n\n"
                "  Regex replace\n"
                "    From: Python regex pattern.  To: replacement (supports \\1, \\2 groups).\n"
                "    Examples:\n"
                "      From: ^Red    To: Blue       → replaces prefix 'Red' with 'Blue'\n"
                "      From: (\\d+)   To: [\\1]        → wraps numbers in brackets\n"
                "      From: \\s+     To: _           → replaces whitespace with underscore\n\n"
                "  Friendly regex  (easier syntax)\n"
                "    'contains X'   → matches anything containing X\n"
                "    'starts X'     → matches strings starting with X\n"
                "    'ends X'       → matches strings ending with X\n\n"
                "  Set to fixed value\n"
                "    Replaces the entire field with the From value.\n\n"
                "CDATA SAFETY\n"
                "───────────────\n"
                "  Fields wrapped in <![CDATA[...]]> are handled transparently.\n"
                "  The CDATA wrapper is preserved — only the inner value is changed.\n"
                "  Do NOT include <![CDATA[ or ]]> in your pattern or replacement.\n\n"
                "DECIMAL PRESERVATION\n"
                "───────────────\n"
                "  When checked: if old value was '1.530000' (6dp), result keeps 6dp.\n"
                "  Example: y*2 on '0.750000' → '1.500000' (not '1.5')\n\n"
                "COMMON REGEX PATTERNS\n"
                "───────────────\n"
                "  \\d+        one or more digits\n"
                "  \\w+        word characters (letters, digits, _)\n"
                "  ^          start of string\n"
                "  $          end of string\n"
                "  .          any single character\n"
                "  .*         any sequence of characters\n"
                "  [abc]      character class (a, b, or c)\n"
                "  (group)    capture group — reference with \\1 in replacement\n"
                "  (?i)       case-insensitive flag at start of pattern\n\n"
                "SAFETY RULES\n"
                "───────────────\n"
                "  • Always use Preview before Apply & Save.\n"
                "  • Keep a backup of your XML file.\n"
                "  • Do not include < > or & in replacement values for non-CDATA fields.\n"
                "  • For CDATA fields (n, Comment, FileName, etc.) special chars are fine.\n"
            )
            txt.insert("1.0", ref)
            txt.config(state="disabled")
            mk_btn(win, "Close", win.destroy, color=BG4).pack(pady=4)

        mk_btn(ref_bar, "📖  Regex Reference Guide", _open_regex_ref,
               color=BG3, font=("Consolas", 8)).pack(side="left", padx=0)

        # Advanced Manual Renaming section (only if enabled in settings)
        if _APP_SETTINGS.get("advanced_renaming_enabled", False):
            adv_f = mk_section(container, "  Advanced Name Replacement  (⚠ Advanced Renaming enabled in Settings)")
            tk.Label(adv_f,
                     text="  Apply a rename pattern to the <n> (Name) field using prefix/suffix or regex.\n"
                          "  This is additive to the main operation above — use conditions to target specific rows.",
                     bg=BG, fg=FG_GREY, font=("Consolas", 8), justify="left").pack(anchor="w", padx=10, pady=4)
            adv_row = tk.Frame(adv_f, bg=BG); adv_row.pack(fill="x", padx=10, pady=4)
            tk.Label(adv_row, text="Name prefix:", bg=BG, fg=FG, font=("Consolas", 9),
                     width=14, anchor="w").pack(side="left")
            self._adv_prefix = tk.StringVar()
            tk.Entry(adv_row, textvariable=self._adv_prefix, width=16, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
            tk.Label(adv_row, text="suffix:", bg=BG, fg=FG_DIM, font=("Consolas", 9)).pack(side="left", padx=(8,2))
            self._adv_suffix = tk.StringVar()
            tk.Entry(adv_row, textvariable=self._adv_suffix, width=16, bg=BG3, fg=FG,
                     insertbackground=FG, font=("Consolas", 9), relief="flat").pack(side="left", padx=4)
            adv_pos_row = tk.Frame(adv_f, bg=BG); adv_pos_row.pack(fill="x", padx=10, pady=2)
            self._adv_pos = tk.StringVar(value="before")
            for lbl, val in [("Before", "before"), ("After", "after"), ("Both", "both")]:
                tk.Radiobutton(adv_pos_row, text=lbl, variable=self._adv_pos, value=val,
                               bg=BG, fg=FG, selectcolor=BG3, activebackground=BG,
                               font=("Consolas", 8)).pack(side="left", padx=6)
            mk_btn(adv_f, "📖  Regex Reference", _open_regex_ref,
                   color=BG3, font=("Consolas", 8)).pack(anchor="w", padx=10, pady=(2, 6))

        # Conditions
        cond_f = mk_section(container, "  Conditions  (only apply to rows where…)  all optional")
        cond_rows = []
        for i in range(3):
            r = tk.Frame(cond_f,bg=BG); r.pack(fill="x",padx=8,pady=2)
            tk.Label(r, text=f"Cond {i+1}:", bg=BG, fg=FG, font=("Consolas",8), width=8).pack(side="left")
            ct = ttk.Combobox(r, values=[], width=16, state="readonly")
            ct.pack(side="left",padx=2)
            tk.Label(r,text="value",bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(side="left",padx=2)
            op = ttk.Combobox(r,values=["==","!=","contains","not contains",">","<",">=","<="],
                              width=12,state="readonly"); op.set("=="); op.pack(side="left",padx=2)
            cv = tk.StringVar()
            tk.Entry(r,textvariable=cv,width=18,bg=BG3,fg=FG,
                     insertbackground=FG,font=("Consolas",8),relief="flat").pack(side="left",padx=2)
            cond_rows.append((ct,op,cv))
        self._cond_rows = cond_rows

        def _refresh_menus(*_):
            for ct,_,_ in cond_rows:
                ct["values"] = self._tags
                if self._tags and not ct.get(): ct.set(self._tags[0])
            if self._tags and not self._tgt_var.get(): self._tgt_menu["values"] = self._tags

        self._tgt_menu.bind("<FocusIn>", lambda e: (setattr(self._tgt_menu,"values",self._tags),))

        nav = tk.Frame(C,bg=BG); nav.pack(pady=12)
        mk_btn(nav, "🔍  Preview (first 20)", lambda: self._run(preview=True),
               color=BG3).pack(side="left",padx=6)
        mk_btn(nav, "⚡  Apply & Save", self._run,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="left",padx=8)

        self._refresh_menus = _refresh_menus

    def _load(self):
        p = filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not p: return
        self._xml_path = p
        self._tags = _detect_row_tags(p)
        n = _count_rows_in_file(p)
        self._status.set(f"✓  {os.path.basename(p)}  ({n:,} rows)")
        self._tgt_menu["values"] = self._tags
        if self._tags: self._tgt_menu.set(self._tags[0])
        self._refresh_menus()

    def _matches_conds(self, row_text):
        for ct, op_box, cv in self._cond_rows:
            tag = ct.get().strip()
            cval = cv.get().strip()
            if not tag or not cval: continue
            row_val = (_xml_tag_val(row_text, tag) or "").strip()
            op = op_box.get()
            if op == "==":
                if row_val != cval: return False
            elif op == "!=":
                if row_val == cval: return False
            elif op == "contains":
                if cval.lower() not in row_val.lower(): return False
            elif op == "not contains":
                if cval.lower() in row_val.lower(): return False
            elif op in (">","<",">=","<="):
                try:
                    a,b = float(row_val), float(cval)
                    if not eval(f"{a}{op}{b}"): return False
                except: return False
        return True

    def _apply_op(self, old_val):
        op = self._op_var.get()
        frm = self._from_var.get()
        to  = self._to_var.get()
        if op == "math":
            # Detect decimal places
            dp = 0
            if "." in old_val:
                dp = len(old_val.split(".",1)[1])
            m = re.search(r'[\d.]+', old_val)
            if not m: return old_val
            num = float(m.group())
            e = frm.strip() or "y"
            if e[0] in "+-": e = "y" + e
            e = e.replace("^","**")
            try:
                result = eval(e, {"__builtins__":{},"round":round,"abs":abs,"int":int,"float":float}, {"y":num})
            except Exception as ex:
                raise ValueError(f"Math error: {ex}")
            if self._preserve_dec.get() and dp > 0:
                new_val = f"{result:.{dp}f}"
            else:
                new_val = str(int(round(result))) if result == int(result) else str(result)
            return old_val[:m.start()] + new_val + old_val[m.end():]
        elif op == "replace":
            return old_val.replace(frm, to)
        elif op == "regex":
            return re.sub(frm, to, old_val)
        elif op == "friendly":
            # Pattern helpers: contains X → .*X.*, starts → ^X, ends → X$
            pat = frm
            if pat.startswith("contains "): pat = ".*" + re.escape(pat[9:]) + ".*"
            elif pat.startswith("starts "): pat = "^" + re.escape(pat[7:])
            elif pat.startswith("ends "):   pat = re.escape(pat[5:]) + "$"
            return re.sub(pat, to, old_val, flags=re.IGNORECASE)
        elif op == "set":
            return frm
        return old_val

    def _run(self, preview=False):
        if not self._xml_path:
            messagebox.showwarning("No file","Load an XML first."); return
        tgt = self._tgt_var.get().strip()
        if not tgt:
            messagebox.showwarning("No target","Select a target field."); return

        changed = []; preview_lines = []
        in_row = False; buf = []; out_parts = []
        rows_changed = 0

        with open(self._xml_path, encoding="utf-8", errors="replace", newline="") as f:
            content = f.read()

        for line in content.splitlines(keepends=True):
            if not in_row:
                if "<ROW>" in line: in_row=True; buf=[line]
                else: out_parts.append(line)
            else:
                buf.append(line)
                if "</ROW>" in line:
                    in_row = False
                    row_text = "".join(buf)
                    if self._matches_conds(row_text):
                        old_val = _xml_tag_val(row_text, tgt) or ""
                        try:
                            new_val = self._apply_op(old_val)
                        except ValueError as e:
                            messagebox.showerror("Error", str(e)); return
                        if new_val != old_val:
                            row_text = re.sub(
                                rf"(<{re.escape(tgt)}>)(.*?)(</{re.escape(tgt)}>)",
                                lambda m, nv=new_val: m.group(1)+nv+m.group(3),
                                row_text, count=1, flags=re.DOTALL)
                            rows_changed += 1
                            if preview and rows_changed <= 20:
                                preview_lines.append(f"  {tgt}: {old_val!r} → {new_val!r}")
                    out_parts.append(row_text)

        if preview:
            msg = f"Would change {rows_changed} row(s).\n\nFirst examples:\n" + "\n".join(preview_lines)
            messagebox.showinfo("Preview", msg); return

        if rows_changed == 0:
            messagebox.showinfo("No changes","No rows matched the conditions."); return

        out_text = "".join(out_parts)
        # Save to libconfig
        default_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(),"libconfig"))
        stem = os.path.splitext(os.path.basename(self._xml_path))[0]
        out_path = filedialog.asksaveasfilename(
            title="Save modified XML",
            initialdir=default_dir,
            initialfile=stem+"_modified.xml",
            defaultextension=".xml",
            filetypes=[("XML","*.xml"),("All","*.*")],
            parent=self.root)
        if not out_path: return
        with open(out_path,"w",encoding="utf-8") as f: f.write(out_text)
        messagebox.showinfo("Done", f"Changed {rows_changed} row(s).\nSaved to:\n{out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 15 — Reorder XML
# ══════════════════════════════════════════════════════════════════════════════

class Tool15(tk.Frame):
    ACC = ACC15

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._xml_path = None; self._tags = []; self._build()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build(self):
        self._clear()
        tk.Label(self, text="REORDER XML",
                 font=("Consolas",16,"bold"), bg=BG, fg=self.ACC).pack(pady=(24,4))
        tk.Label(self,
                 text="Sort <ROW> blocks by any field, alphabetically or numerically.\n"
                      "Optionally renumber a sequence field after sorting.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9), justify="center").pack(pady=(0,12))

        sf = mk_section(self, "  Load XML  ")
        self._status = tk.StringVar(value="No file loaded")
        tk.Label(sf, textvariable=self._status, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf, "📂  Load XML", self._load, color=self.ACC, fg=BG2).pack(side="right",padx=8,pady=6)

        sf2 = mk_section(self, "  Sort Settings  ")
        r = tk.Frame(sf2,bg=BG); r.pack(fill="x",padx=8,pady=6)
        tk.Label(r, text="Sort by field:", bg=BG, fg=FG, font=("Consolas",9)).pack(side="left")
        self._sort_field = tk.StringVar()
        self._sort_menu = ttk.Combobox(r, textvariable=self._sort_field, width=20)
        self._sort_menu.pack(side="left",padx=6)
        tk.Label(r, text="Order:", bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(side="left",padx=(10,2))
        self._sort_order = tk.StringVar(value="asc")
        tk.Radiobutton(r, text="Asc", variable=self._sort_order, value="asc",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9)).pack(side="left",padx=2)
        tk.Radiobutton(r, text="Desc", variable=self._sort_order, value="desc",
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9)).pack(side="left",padx=2)
        tk.Label(r, text="  Numeric:", bg=BG, fg=FG_DIM, font=("Consolas",9)).pack(side="left",padx=(10,2))
        self._sort_num = tk.BooleanVar(value=True)
        tk.Checkbutton(r, variable=self._sort_num, bg=BG, selectcolor=BG3, activebackground=BG).pack(side="left")

        sf3 = mk_section(self, "  Renumber Field  (optional — renumber after sort)  ")
        r3 = tk.Frame(sf3,bg=BG); r3.pack(fill="x",padx=8,pady=6)
        self._renum_var = tk.BooleanVar(value=False)
        tk.Checkbutton(r3, text="Renumber field:", variable=self._renum_var,
                       bg=BG,fg=FG,selectcolor=BG3,activebackground=BG,font=("Consolas",9)).pack(side="left")
        self._renum_field = tk.StringVar()
        self._renum_menu = ttk.Combobox(r3, textvariable=self._renum_field, width=16)
        self._renum_menu.pack(side="left",padx=6)
        tk.Label(r3, text="starting at:", bg=BG, fg=FG_DIM, font=("Consolas",8)).pack(side="left")
        self._renum_start = tk.IntVar(value=1)
        tk.Spinbox(r3, from_=0, to=999999, textvariable=self._renum_start,
                   width=8, bg=BG3, fg=FG, font=("Consolas",9), relief="flat").pack(side="left",padx=4)

        nav = tk.Frame(self,bg=BG); nav.pack(pady=12)
        mk_btn(nav, "⚡  Sort & Save", self._run,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="left",padx=8)

    def _load(self):
        p = filedialog.askopenfilename(filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not p: return
        self._xml_path = p
        self._tags = _detect_row_tags(p)
        n = _count_rows_in_file(p)
        self._status.set(f"✓  {os.path.basename(p)}  ({n:,} rows)")
        self._sort_menu["values"] = self._tags
        self._renum_menu["values"] = self._tags
        if self._tags:
            self._sort_menu.set(self._tags[0])
            self._renum_menu.set(self._tags[0])

    def _run(self):
        if not self._xml_path:
            messagebox.showwarning("No file","Load an XML first."); return
        sort_field = self._sort_field.get().strip()
        if not sort_field:
            messagebox.showwarning("No field","Select a sort field."); return

        # Read all rows plus prefix/suffix
        with open(self._xml_path, encoding="utf-8", errors="replace", newline="") as f:
            content = f.read()

        row_blocks = list(re.finditer(r"<ROW>.*?</ROW>", content, re.DOTALL))
        if not row_blocks:
            messagebox.showwarning("No rows","No <ROW> blocks found."); return

        prefix = content[:row_blocks[0].start()]
        suffix = content[row_blocks[-1].end():]
        if len(row_blocks) > 1:
            sep = content[row_blocks[0].end():row_blocks[1].start()]
        else:
            sep = "\n"

        rows = [(i, m.group(0)) for i,m in enumerate(row_blocks)]

        do_num = self._sort_num.get()
        desc   = self._sort_order.get() == "desc"

        def sort_key(item):
            _, row_text = item
            v = _xml_tag_val(row_text, sort_field) or ""
            if do_num:
                try: return (0, float(v), v)
                except: pass
            return (1, 0.0, v)

        rows.sort(key=sort_key, reverse=desc)

        # Renumber
        if self._renum_var.get():
            renum_f = self._renum_field.get().strip()
            start   = self._renum_start.get()
            new_rows = []
            for i,(_,rt) in enumerate(rows):
                nid = str(start + i)
                rt = re.sub(rf"(<{re.escape(renum_f)}>).*?(</{re.escape(renum_f)}>)",
                            lambda m, nv=nid: m.group(1)+nv+m.group(2), rt, count=1, flags=re.DOTALL)
                new_rows.append(rt)
        else:
            new_rows = [rt for _,rt in rows]

        out_text = prefix + sep.join(new_rows) + suffix

        default_dir = _APP_SETTINGS.get("libconfig_dir", os.path.join(os.getcwd(),"libconfig"))
        stem = os.path.splitext(os.path.basename(self._xml_path))[0]
        out_path = filedialog.asksaveasfilename(
            title="Save reordered XML",
            initialdir=default_dir, initialfile=stem+"_reordered.xml",
            defaultextension=".xml", filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if not out_path: return
        with open(out_path,"w",encoding="utf-8") as f: f.write(out_text)
        messagebox.showinfo("Done", f"Sorted {len(new_rows)} rows.\nSaved to:\n{out_path}")


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 16 — Duplicate / Invalid ID Checker
# ══════════════════════════════════════════════════════════════════════════════

class Tool16(tk.Frame):
    ACC = ACC16

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._files = []; self._build()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build(self):
        self._clear()
        tk.Label(self, text="DUPLICATE / INVALID ID CHECKER",
                 font=("Consolas",15,"bold"), bg=BG, fg=self.ACC).pack(pady=(24,4))
        tk.Label(self,
                 text="Check one or multiple XML files for duplicate or conflicting IDs.\n"
                      "Cross-file conflicts (same ID in different files) are also detected.",
                 bg=BG, fg=FG_DIM, font=("Consolas",9), justify="center").pack(pady=(0,10))

        sf = mk_section(self, "  Load XML Files  ")
        self._file_status = tk.StringVar(value="No files loaded")
        tk.Label(sf, textvariable=self._file_status, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf, "📂  Add Files", self._add_files,
               color=self.ACC, fg=BG2).pack(side="right",padx=4,pady=6)
        mk_btn(sf, "✗  Clear", self._clear_files, color=BG4).pack(side="right",padx=4,pady=6)

        id_f = mk_section(self, "  ID Field Name  ")
        r = tk.Frame(id_f,bg=BG); r.pack(fill="x",padx=8,pady=6)
        tk.Label(r,text="Tag:",bg=BG,fg=FG,font=("Consolas",9)).pack(side="left")
        self._id_field = tk.StringVar(value="ID")
        tk.Entry(r,textvariable=self._id_field,width=16,bg=BG3,fg=FG,
                 insertbackground=FG,font=("Consolas",9),relief="flat").pack(side="left",padx=6)
        tk.Label(r,text="(usually ID, may also be CompoundID, ExchangeID, etc.)",
                 bg=BG,fg=FG_GREY,font=("Consolas",8)).pack(side="left",padx=4)

        nav = tk.Frame(self,bg=BG); nav.pack(pady=10)
        mk_btn(nav, "🔍  Run Check", self._run,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="left",padx=8)

        self._out_txt = scrolledtext.ScrolledText(self, font=("Consolas",8), bg=BG2, fg=FG, height=14)
        self._out_txt.pack(fill="both", expand=True, padx=10, pady=6)
        self._out_txt.config(state="disabled")

    def _add_files(self):
        paths = filedialog.askopenfilenames(filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if paths: self._files.extend(paths)
        self._file_status.set(f"✓  {len(self._files)} file(s) loaded")

    def _clear_files(self):
        self._files.clear(); self._file_status.set("No files loaded")

    def _log(self, s):
        self._out_txt.config(state="normal")
        self._out_txt.insert("end", s+"\n")
        self._out_txt.see("end"); self._out_txt.config(state="disabled")
        self.update_idletasks()

    def _run(self):
        if not self._files:
            messagebox.showwarning("No files","Load files first."); return
        id_field = self._id_field.get().strip() or "ID"
        self._out_txt.config(state="normal"); self._out_txt.delete("1.0","end"); self._out_txt.config(state="disabled")

        all_ids = {}   # id_val -> [(filename, name_val)]
        internal = {}  # filename -> {id_val -> [name_vals]}
        has_issues = False

        for path in self._files:
            fname = os.path.basename(path)
            file_map = {}
            for row in _iter_xml_rows(path):
                id_val  = _xml_tag_val(row, id_field) or ""
                name_val = _xml_tag_val(row, "Name") or _xml_tag_val(row, "n") or ""
                # Strip CDATA
                name_val = re.sub(r"<!\[CDATA\[(.*?)\]\]>","\\1",name_val).strip()
                if not id_val: continue
                file_map.setdefault(id_val,[]).append(name_val)
                all_ids.setdefault(id_val,[]).append((fname,name_val))

            dupes = {i:n for i,n in file_map.items() if len(n)>1}
            if dupes:
                has_issues = True
                self._log(f"\n⚠ Internal duplicates in {fname}:")
                for k,names in dupes.items():
                    self._log(f"  ID {k}: {', '.join(names)}")
            else:
                self._log(f"✓  {fname}: no internal duplicates")
            internal[fname] = {i:n for i,n in file_map.items() if len(n)>1}

        # Cross-file
        cross = {i:data for i,data in all_ids.items()
                 if len({f for f,_ in data})>1}
        if cross:
            has_issues = True
            self._log(f"\n⚠ Cross-file conflicts ({len(cross)}):")
            for k,entries in cross.items():
                desc = ", ".join(f"{f}({n})" for f,n in entries)
                self._log(f"  ID {k}: {desc}")
        else:
            self._log("\n✓  No cross-file conflicts")

        # Report
        lines = [
            "Duplicate / Invalid ID Report",
            f"Files: {', '.join(os.path.basename(p) for p in self._files)}",
            f"ID field: <{id_field}>",
            f"Generated: {_time_mod.strftime('%Y-%m-%d %H:%M:%S')}",
            "=" * 60,
        ]
        for fname, dupes in internal.items():
            if dupes:
                lines.append(f"\nInternal duplicates in {fname}:")
                for k,names in dupes.items():
                    lines.append(f"  ID {k}: {', '.join(names)}")
        if cross:
            lines.append(f"\nCross-file conflicts:")
            for k,entries in cross.items():
                desc = ", ".join(f"{f}({n})" for f,n in entries)
                lines.append(f"  ID {k}: {desc}")
        if not has_issues:
            lines.append("\nNo issues found.")
        _write_report("id_check.txt", "\n".join(lines), self.root)


# ══════════════════════════════════════════════════════════════════════════════
# TOOL 17 — Fix ItemParam (CDATA + decimal places)
# ══════════════════════════════════════════════════════════════════════════════

# Fields that must have CDATA wrappers
_CDATA_TAGS = [
    "Name","Comment","Use","Name_Eng","Comment_Eng",
    "FileName","InvFileName","CmtFileName","EquipFileName",
    "PartFileName","Summary","ShopFileName",
]
# Fields that require exactly 6 decimal places
_SIX_DP_TAGS = [
    "Delay","HPRecoveryRate","MPRecoveryRate","CardGenParam",
]

def _fix_itemparam_text(text):
    # 1. CDATA fix
    pair_re = re.compile(
        rf"<({'|'.join(_CDATA_TAGS)})>(.*?)</\1>", re.DOTALL | re.IGNORECASE)
    lone_re = re.compile(
        rf"<({'|'.join(_CDATA_TAGS)})>\s*(?=\r?\n|$)", re.IGNORECASE)

    def _wrap_pair(m):
        tag = m.group(1); content = m.group(2).strip()
        if "<![CDATA[" in content: return m.group(0)
        return f"<{tag}><![CDATA[{content or ' '}]]></{tag}>"

    def _wrap_lone(m):
        return f"<{m.group(1)}><![CDATA[ ]]></{m.group(1)}>"

    text = pair_re.sub(_wrap_pair, text)
    text = lone_re.sub(_wrap_lone, text)

    # 2. Decimal places fix for 6dp fields
    for tag in _SIX_DP_TAGS:
        def _fix_dp(m, t=tag):
            inner = m.group(2).strip()
            try:
                val = float(inner)
                fixed = f"{val:.6f}"
                return m.group(1) + fixed + m.group(3)
            except: return m.group(0)
        text = re.sub(
            rf"(<{re.escape(tag)}>)([\d.eE+\-]+)(</{re.escape(tag)}>)",
            _fix_dp, text)

    return text


class Tool17(tk.Frame):
    ACC = ACC17

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root = root; self.session = session
        self._build()

    def _clear(self): [w.destroy() for w in self.winfo_children()]

    def _build(self):
        self._clear()
        tk.Label(self, text="FIX ITEMPARAM XML",
                 font=("Consolas",16,"bold"), bg=BG, fg=self.ACC).pack(pady=(24,4))
        tk.Label(self,
                 text="Restores CDATA wrappers to string fields and normalises decimal places.\n"
                      f"CDATA tags: {', '.join(_CDATA_TAGS[:6])}…\n"
                      f"6-decimal tags: {', '.join(_SIX_DP_TAGS)}",
                 bg=BG, fg=FG_DIM, font=("Consolas",8), justify="center").pack(pady=(0,12))

        sf = mk_section(self, "  Load ItemParam XML  ")
        self._status = tk.StringVar(value="No file loaded")
        tk.Label(sf, textvariable=self._status, bg=BG, fg=FG_GREY, font=("Consolas",9)).pack(side="left",padx=10)
        mk_btn(sf, "📂  Load XML", self._load, color=self.ACC, fg=BG2).pack(side="right",padx=8,pady=6)
        mk_btn(sf, "📁  Load Folder", self._load_folder, color=BG3).pack(side="right",padx=4,pady=6)

        opt_f = mk_section(self, "  Options  ")
        self._backup = tk.BooleanVar(value=True)
        tk.Checkbutton(opt_f, text="Create .bak backup before overwriting",
                       variable=self._backup, bg=BG, fg=FG,
                       selectcolor=BG3, activebackground=BG, font=("Consolas",9)).pack(anchor="w",padx=10,pady=4)

        nav = tk.Frame(self,bg=BG); nav.pack(pady=10)
        mk_btn(nav, "⚡  Fix Files", self._run,
               color=self.ACC, fg=BG2, font=("Consolas",11,"bold")).pack(side="left",padx=8)

        self._out_txt = scrolledtext.ScrolledText(self, font=("Consolas",8), bg=BG2, fg=FG, height=14)
        self._out_txt.pack(fill="both", expand=True, padx=10, pady=6)
        self._out_txt.config(state="disabled")
        self._paths = []

    def _log(self, s):
        self._out_txt.config(state="normal")
        self._out_txt.insert("end",s+"\n")
        self._out_txt.see("end"); self._out_txt.config(state="disabled")
        self.update_idletasks()

    def _load(self):
        paths = filedialog.askopenfilenames(filetypes=[("XML","*.xml"),("All","*.*")], parent=self.root)
        if paths: self._paths = list(paths); self._status.set(f"✓  {len(self._paths)} file(s)")

    def _load_folder(self):
        folder = filedialog.askdirectory(parent=self.root)
        if not folder: return
        self._paths = [os.path.join(folder,f) for f in os.listdir(folder) if f.lower().endswith(".xml")]
        self._status.set(f"✓  {len(self._paths)} XML files in folder")

    def _run(self):
        if not self._paths:
            messagebox.showwarning("No files","Load files first."); return
        self._out_txt.config(state="normal"); self._out_txt.delete("1.0","end"); self._out_txt.config(state="disabled")
        changed = 0
        for path in self._paths:
            orig = open(path, encoding="utf-8", errors="replace").read()
            fixed = _fix_itemparam_text(orig)
            if fixed != orig:
                if self._backup.get():
                    bak = path+".bak"
                    with open(bak,"w",encoding="utf-8") as f: f.write(orig)
                with open(path,"w",encoding="utf-8") as f: f.write(fixed)
                self._log(f"✓  Fixed: {os.path.basename(path)}")
                changed += 1
            else:
                self._log(f"—  No changes: {os.path.basename(path)}")
        self._log(f"\nDone: {changed}/{len(self._paths)} files updated.")


# ══════════════════════════════════════════════════════════════════════════════
# NCASH UPDATER (Combined) — landing screen that routes to Tool3 or Tool4,
# with a pop-out Ticket Calculator widget always available
# ══════════════════════════════════════════════════════════════════════════════
class ToolNCashCombined(tk.Frame):
    """Combined NCash Updater — landing screen + Ticket Calc popout."""

    ACC = ACC3

    def __init__(self, parent, root, session):
        super().__init__(parent, bg=BG)
        self.root    = root
        self.session = session
        self._sub    = None   # currently shown sub-tool frame
        self._build_landing()

    def _clear(self):
        for w in self.winfo_children(): w.destroy()
        self._sub = None

    def _popout_calc(self):
        win = tk.Toplevel(self.root)
        win.title("NCash ↔ Ticket Calculator")
        win.geometry("480x300")
        win.configure(bg=BG)
        win.resizable(False, False)
        # embed a Tool5 instance directly
        frm = Tool5(win, self.root, self.session)
        frm.pack(fill="both", expand=True)

    def _build_landing(self):
        self._clear()
        center = tk.Frame(self, bg=BG); center.pack(expand=True)

        tk.Label(center, text="NCASH UPDATER",
                 font=("Consolas", 20, "bold"), bg=BG, fg=self.ACC).pack(pady=(32, 6))
        tk.Label(center, text="What would you like to update?",
                 bg=BG, fg=FG_DIM, font=("Consolas", 11)).pack(pady=(0, 24))

        cards = tk.Frame(center, bg=BG); cards.pack(pady=8)

        # Individual items card
        card_a = tk.Frame(cards, bg=BG2, padx=20, pady=16, cursor="hand2")
        card_a.pack(side="left", padx=12, ipadx=8, ipady=4)
        tk.Label(card_a, text="Individual Items",
                 font=("Consolas", 13, "bold"), bg=BG2, fg=ACC3).pack()
        tk.Label(card_a, text="Change NCash of specific items\nvia a simple CSV or session import.",
                 bg=BG2, fg=FG, font=("Consolas", 8), justify="center").pack(pady=(4, 10))
        mk_btn(card_a, "Open  →  Simple Updater",
               self._open_simple, color=ACC3, fg=BG2,
               font=("Consolas", 9, "bold")).pack()

        # Boxes card
        card_b = tk.Frame(cards, bg=BG2, padx=20, pady=16, cursor="hand2")
        card_b.pack(side="left", padx=12, ipadx=8, ipady=4)
        tk.Label(card_b, text="Boxes",
                 font=("Consolas", 13, "bold"), bg=BG2, fg=ACC4).pack()
        tk.Label(card_b, text="Update NCash for parent boxes\nand their sub-box contents.",
                 bg=BG2, fg=FG, font=("Consolas", 8), justify="center").pack(pady=(4, 10))
        mk_btn(card_b, "Open  →  Parent Updater",
               self._open_parent, color=ACC4, fg=BG2,
               font=("Consolas", 9, "bold")).pack()

        # Ticket calculator popout widget (always visible between landing and menus)
        sep = tk.Frame(center, bg=BG4, height=1); sep.pack(fill="x", padx=40, pady=(20, 0))
        calc_bar = tk.Frame(center, bg=BG); calc_bar.pack(pady=8)
        tk.Label(calc_bar, text="🧮  NCash ↔ Ticket Calculator",
                 bg=BG, fg=FG_GREY, font=("Consolas", 9)).pack(side="left", padx=6)
        mk_btn(calc_bar, "Open Calculator ↗", self._popout_calc,
               color=BG3, font=("Consolas", 8)).pack(side="left", padx=6)

    def _open_simple(self):
        self._clear()
        hdr = tk.Frame(self, bg=BG2); hdr.pack(fill="x")
        mk_btn(hdr, "◀  Back", self._build_landing, color=BG4,
               font=("Consolas", 8)).pack(side="left", padx=8, pady=4)
        tk.Label(hdr, text="NCash Updater — Simple",
                 font=("Consolas", 11, "bold"), bg=BG2, fg=ACC3, pady=6).pack(side="left", padx=4)
        mk_btn(hdr, "🧮  Ticket Calc", self._popout_calc,
               color=BG3, font=("Consolas", 8)).pack(side="right", padx=8, pady=4)
        sub = Tool3(self, self.root, self.session)
        sub.pack(fill="both", expand=True)
        self._sub = sub

    def _open_parent(self):
        self._clear()
        hdr = tk.Frame(self, bg=BG2); hdr.pack(fill="x")
        mk_btn(hdr, "◀  Back", self._build_landing, color=BG4,
               font=("Consolas", 8)).pack(side="left", padx=8, pady=4)
        tk.Label(hdr, text="NCash Updater — Parent Boxes",
                 font=("Consolas", 11, "bold"), bg=BG2, fg=ACC4, pady=6).pack(side="left", padx=4)
        mk_btn(hdr, "🧮  Ticket Calc", self._popout_calc,
               color=BG3, font=("Consolas", 8)).pack(side="right", padx=8, pady=4)
        sub = Tool4(self, self.root, self.session)
        sub.pack(fill="both", expand=True)
        self._sub = sub


TOOLS = [
    ("1",  "ItemParam\nGenerator",        ACC6,  Tool6),
    ("2",  "Box XML\nGenerator",          ACC1,  Tool1),
    ("3",  "Fashion\nCreation",           ACC18, Tool18),
    ("4",  "Set Item\nGenerator",         ACC8,  Tool8),
    ("5",  "Compound /\nExchange / Shop", ACC7,  Tool7),
    ("6",  "Box Rate / Count\nAdjuster",  ACC2,  Tool2),
    ("7",  "NCash\nUpdater",              ACC3,  ToolNCashCombined),
    ("8",  "Reorder\nXML",               ACC15, Tool15),
    ("9",  "Row Counter /\nUpdater",      ACC9,  Tool9),
    ("10", "ID\nChecker",                ACC16, Tool16),
    ("11", "Fix\nItemParam",             ACC17, Tool17),
    ("12", "Range\nAuditor",             ACC10, Tool10),
    ("13", "XML\nComparator",            ACC11, Tool11),
    ("14", "Row\nDuplicator",            ACC13, Tool13),
    ("15", "Data\nExtract",              ACC12, Tool12),
    ("16", "Mass Variable\nManip.",      ACC14, Tool14),
]

class CombinedApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Mewsie's ItemParam Toolbox")
        # Geometry + minsize. update() forces the WM to apply before widgets render.
        self.geometry("1280x900")
        self.minsize(1100, 860)
        self.resizable(True, True)
        self.configure(bg=BG2)
        self.update_idletasks()
        self._current_tool = None
        self._tool_instances = {}
        self._nav_buttons = {}
        self.session = AppSession()
        self._build_layout()
        self._switch_tool(0)

    def _build_layout(self):
        # ── Left sidebar — scrollable so all tools are always reachable ───
        sidebar_outer = tk.Frame(self, bg=BG2, width=158)
        sidebar_outer.pack(side="left", fill="y")
        sidebar_outer.pack_propagate(False)

        # Static header (never scrolls)
        tk.Label(sidebar_outer, text="MEWSIE'S\nITEMPARAM\nTOOLBOX",
                 font=("Consolas",10,"bold"), bg=BG2, fg=FG,
                 justify="center").pack(pady=(14,8))
        tk.Frame(sidebar_outer, bg=BG4, height=1).pack(fill="x", padx=10, pady=2)

        # Scrollable nav area
        nav_canvas = tk.Canvas(sidebar_outer, bg=BG2, bd=0, highlightthickness=0,
                               width=156)
        nav_sb = tk.Scrollbar(sidebar_outer, orient="vertical", command=nav_canvas.yview)
        nav_canvas.configure(yscrollcommand=nav_sb.set)
        # scrollbar on far right, canvas fills rest
        nav_sb.pack(side="right", fill="y")
        nav_canvas.pack(side="top", fill="both", expand=True)

        nav_inner = tk.Frame(nav_canvas, bg=BG2)
        nav_win = nav_canvas.create_window((0, 0), window=nav_inner, anchor="nw")

        def _on_nav_configure(e):
            nav_canvas.configure(scrollregion=nav_canvas.bbox("all"))
            nav_canvas.itemconfig(nav_win, width=nav_canvas.winfo_width())
        nav_inner.bind("<Configure>", _on_nav_configure)
        nav_canvas.bind("<Configure>", lambda e: nav_canvas.itemconfig(nav_win, width=e.width))

        # Mouse-wheel scrolling on the nav
        def _nav_scroll(e):
            nav_canvas.yview_scroll(int(-1*(e.delta/120)) if e.delta else (-1 if e.num==4 else 1), "units")
        nav_canvas.bind("<MouseWheel>", _nav_scroll)
        nav_inner.bind("<MouseWheel>", _nav_scroll)

        for i,(num,label,color,_) in enumerate(TOOLS):
            frm = tk.Frame(nav_inner, bg=BG2, cursor="hand2")
            frm.pack(fill="x", padx=4, pady=2)
            dot = tk.Label(frm, text="●", font=("Consolas",8), bg=BG2, fg=color, width=2)
            dot.pack(side="left")
            btn = tk.Button(frm, text=f" {label}", font=("Consolas",8),
                            bg=BG2, fg=FG_DIM, relief="flat", anchor="w",
                            justify="left", padx=2, pady=4,
                            activebackground=BG3, activeforeground=FG,
                            command=lambda idx=i: self._switch_tool(idx))
            btn.pack(side="left", fill="x", expand=True)
            self._nav_buttons[i] = (frm, btn, dot, color)
            frm.bind("<Button-1>", lambda e, idx=i: self._switch_tool(idx))
            # propagate scroll to canvas
            for w in (frm, btn, dot):
                w.bind("<MouseWheel>", _nav_scroll)

        # Static footer (gear + info)
        footer = tk.Frame(sidebar_outer, bg=BG2)
        footer.pack(side="bottom", fill="x")
        tk.Frame(footer, bg=BG4, height=1).pack(fill="x", padx=10, pady=4)
        mk_btn(footer, "⚙  Settings", lambda: _open_settings_window(self),
               color=BG2, fg=FG_GREY, font=("Consolas",8)).pack(fill="x", padx=6, pady=2)
        tk.Label(footer, text="Scroll list ↕ for more tools",
                 font=("Consolas",6), bg=BG2, fg=FG_GREY).pack(pady=(0,4))

        # ── Content area ──────────────────────────────────────────────────
        self._content = tk.Frame(self, bg=BG)
        self._content.pack(side="left", fill="both", expand=True)

    def _switch_tool(self, idx):
        if self._current_tool == idx: return

        # Update nav highlight
        for i,(frm,btn,dot,color) in self._nav_buttons.items():
            if i == idx:
                frm.config(bg=BG3); btn.config(bg=BG3, fg=FG); dot.config(bg=BG3, fg=color)
            else:
                frm.config(bg=BG2); btn.config(bg=BG2, fg=FG_DIM); dot.config(bg=BG2, fg=self._nav_buttons[i][3])

        # Hide current
        if self._current_tool is not None:
            if self._current_tool in self._tool_instances:
                self._tool_instances[self._current_tool].pack_forget()

        # Instantiate or show
        if idx not in self._tool_instances:
            _,_,_,ToolClass = TOOLS[idx]
            instance = ToolClass(self._content, self, self.session)
            self._tool_instances[idx] = instance

        self._tool_instances[idx].pack(fill="both", expand=True)
        self._current_tool = idx
        num,label,_,_ = TOOLS[idx]
        self.title(f"Mewsie's ItemParam Toolbox  —  Tool {num}: {label.replace(chr(10),' ')}")


if __name__ == "__main__":
    CombinedApp().mainloop()